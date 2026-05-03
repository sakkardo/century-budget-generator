"""
Parse Yardi YSL Annual Budget exports.

Reads YSL .xlsx files, extracts GL codes and values, and returns structured data.
"""

import logging
import re
from pathlib import Path
from typing import Dict, Tuple, Optional, Any
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# Regex patterns for period extraction from YSL header metadata (rows 1-15).
# YSL files commonly include strings like:
#   "Period: 01/2026 - 04/2026"
#   "Through: 04/30/2026"
#   "From 1/1/2026 To 4/30/2026"
# The "through" month is what we want for budget_period.
_MONTH_NAMES = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
# Match MM/YYYY or MM-YYYY (with optional leading zero)
_MM_YYYY_RE = re.compile(r'(\d{1,2})[\-/](20\d{2})')
# Match MM/DD/YYYY or M/D/YYYY (date with day)
_MM_DD_YYYY_RE = re.compile(r'(\d{1,2})[\-/]\d{1,2}[\-/](20\d{2})')
# Match "Apr 2026" or "April 2026"
_MONTH_NAME_RE = re.compile(r'\b(' + '|'.join(_MONTH_NAMES) + r')[a-z]*\s+(20\d{2})\b', re.IGNORECASE)


class YSLParser:
    """Parser for Yardi YSL Annual Budget export files."""

    # Metadata rows that should be skipped (rows 1-15)
    METADATA_SKIP_ROWS = 15

    # Column indices (0-based)
    COL_METADATA = 0      # Column A - contains row type metadata
    COL_GL_CODE = 1       # Column B - GL code
    COL_PERIOD1 = 3       # Column D
    COL_PERIOD2 = 4       # Column E
    COL_PERIOD3 = 5       # Column F
    COL_PERIOD4 = 6       # Column G
    COL_PERIOD5 = 7       # Column H

    # Row type indicators in column A
    ROW_TYPE_HEADER = "H"
    ROW_TYPE_DETAIL = "R"
    ROW_TYPE_TOTAL = "T"

    def __init__(self, file_path: Path):
        """
        Initialize parser with file path.

        Args:
            file_path: Path to YSL .xlsx file
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"YSL file not found: {self.file_path}")
        self.wb = None
        self.ws = None

    def parse(self) -> Tuple[Dict[str, Dict[str, float]], Dict[str, str]]:
        """
        Parse YSL file and extract GL codes with values and property info.

        Returns:
            Tuple of:
                - gl_data: Dict keyed by GL code with period values
                - property_info: Dict with property_code and property_name

        Example:
            gl_data = {
                "4010-0000": {
                    "period_1": 1000.0,
                    "period_2": 1050.0,
                    "period_3": 2100.0,
                    "period_4": 1200.0,
                    "period_5": 1300.0,
                }
            }
            property_info = {
                "property_code": "204",
                "property_name": "123 Main Street"
            }
        """
        self.wb = load_workbook(self.file_path, data_only=True)
        self.ws = self.wb.active

        gl_data = {}
        property_info = self._extract_property_info()

        # Start after metadata rows
        for row_idx in range(self.METADATA_SKIP_ROWS + 1, self.ws.max_row + 1):
            row = self.ws[row_idx]

            # Skip blank rows
            if not row[self.COL_GL_CODE].value:
                continue

            # Get row type from metadata (column A)
            metadata = row[self.COL_METADATA].value
            if not metadata:
                continue

            # Skip header and total rows
            if self._get_row_type(metadata) in (self.ROW_TYPE_HEADER, self.ROW_TYPE_TOTAL):
                continue

            # Extract GL code and values
            gl_code = str(row[self.COL_GL_CODE].value).strip()

            # Validate GL code format (should be XXXX-XXXX)
            if not self._is_valid_gl_code(gl_code):
                logger.warning(f"Invalid GL code format at row {row_idx}: {gl_code}")
                continue

            # Skip duplicates — Adjustments section re-lists GL codes with
            # negated values for cash-basis reporting. Keep first occurrence only.
            if gl_code in gl_data:
                logger.debug(f"Skipping duplicate GL {gl_code} at row {row_idx}")
                continue

            # Extract period values
            values = self._extract_period_values(row)

            gl_data[gl_code] = values
            logger.debug(f"Parsed GL {gl_code}: {values}")

        self.wb.close()
        return gl_data, property_info

    def _extract_period_month(self) -> Optional[int]:
        """
        Scan header metadata rows (1-15) for the report period and return
        the "actuals through" month as an integer 1-12.

        Looks for patterns like:
          "Period: 01/2026 - 04/2026"          -> 4 (last month of range)
          "Through: 04/30/2026"                -> 4
          "Apr 2026" / "April 2026"            -> 4
          "From 01/01/2026 To 04/30/2026"      -> 4

        Strategy: scan all metadata cells, collect every (month, year) pair
        we can parse, return the LATEST month (highest month number that
        appears in the most recent year present). The "through" month is
        always the latest one.

        Returns None if no recognizable period is found — caller falls
        back to manual selection.
        """
        if not self.ws or self.ws.max_row < 1:
            return None
        scan_rows = min(self.ws.max_row, self.METADATA_SKIP_ROWS)
        candidates = []  # list of (year, month) tuples
        for row_idx in range(1, scan_rows + 1):
            for cell in self.ws[row_idx]:
                v = cell.value
                if v is None:
                    continue
                text = str(v)
                # MM/DD/YYYY first (more specific) so we don't read "04/30" as month=4 year=30
                for m in _MM_DD_YYYY_RE.finditer(text):
                    month = int(m.group(1))
                    year = int(m.group(2))
                    if 1 <= month <= 12 and 2000 <= year <= 2100:
                        candidates.append((year, month))
                # MM/YYYY (only if MM_DD didn't already swallow it)
                # Use a simple heuristic: scan substrings that don't have an
                # extra "/dd/" pattern around them.
                for m in _MM_YYYY_RE.finditer(text):
                    pos_start = m.start()
                    pos_end = m.end()
                    surrounding = text[max(0, pos_start - 3):min(len(text), pos_end + 3)]
                    # Skip if it looks like part of MM/DD/YYYY
                    if re.search(r'\d/\d{1,2}/' + m.group(0), surrounding):
                        continue
                    month = int(m.group(1))
                    year = int(m.group(2))
                    if 1 <= month <= 12 and 2000 <= year <= 2100:
                        candidates.append((year, month))
                # Month name + year
                for m in _MONTH_NAME_RE.finditer(text):
                    name = m.group(1).lower()[:3]
                    if name in _MONTH_NAMES:
                        month = _MONTH_NAMES.index(name) + 1
                        year = int(m.group(2))
                        candidates.append((year, month))
        if not candidates:
            return None
        # The "through" month is the highest (year, month) tuple
        candidates.sort()
        latest_year, latest_month = candidates[-1]
        logger.info(f"YSL period auto-detected: month={latest_month}, year={latest_year} (from {len(candidates)} candidate matches)")
        return latest_month

    def _extract_property_info(self) -> Dict[str, Any]:
        """
        Extract property information from metadata rows (rows 1-15).

        Returns:
            Dict with property_code, property_name, and (when detectable)
            budget_period_month (int 1-12).
        """
        property_info = {
            "property_code": None,
            "property_name": None,
            "budget_period_month": self._extract_period_month(),
        }

        # YSL format: property code is in row 3, column B; property name is in row 11, column C
        if self.ws.max_row >= 3:
            property_code = self.ws["B3"].value
            if property_code:
                property_info["property_code"] = str(property_code).strip()

        if self.ws.max_row >= 11:
            property_name = self.ws["C11"].value
            if property_name:
                property_info["property_name"] = str(property_name).strip()

        # Fallback: search through metadata rows if direct method fails
        if not property_info["property_code"] or not property_info["property_name"]:
            for row_idx in range(1, min(self.METADATA_SKIP_ROWS + 1, self.ws.max_row + 1)):
                row = self.ws[row_idx]
                cell_a = row[0].value
                cell_b = row[1].value
                cell_c = row[2].value

                if not cell_a:
                    continue

                cell_a_str = str(cell_a).lower()

                # Common YSL metadata indicators
                if "property" in cell_a_str or "building" in cell_a_str:
                    if not property_info["property_name"]:
                        property_info["property_name"] = str(cell_b) if cell_b else str(cell_c) if cell_c else None

                if "code" in cell_a_str and "entity" in cell_a_str:
                    if not property_info["property_code"]:
                        property_info["property_code"] = str(cell_b) if cell_b else None

                if "entity" in cell_a_str and "code" not in cell_a_str:
                    if not property_info["property_code"]:
                        property_info["property_code"] = str(cell_b) if cell_b else None

        logger.info(f"Extracted property info: {property_info}")
        return property_info

    def _extract_period_values(self, row: Any) -> Dict[str, Optional[float]]:
        """
        Extract numeric values from period columns (D through H).

        Args:
            row: openpyxl row object

        Returns:
            Dict with period keys and numeric values (or None)
        """
        values = {}
        col_indices = [
            (self.COL_PERIOD1, "period_1"),
            (self.COL_PERIOD2, "period_2"),
            (self.COL_PERIOD3, "period_3"),
            (self.COL_PERIOD4, "period_4"),
            (self.COL_PERIOD5, "period_5"),
        ]

        for col_idx, period_name in col_indices:
            try:
                val = row[col_idx].value
                if val is None:
                    values[period_name] = None
                else:
                    values[period_name] = float(val)
            except (ValueError, TypeError):
                logger.warning(f"Could not convert value to float: {val}")
                values[period_name] = None

        return values

    @staticmethod
    def _get_row_type(metadata: str) -> Optional[str]:
        """
        Extract row type from metadata string.

        Args:
            metadata: Metadata string from column A (e.g., "$t=H")

        Returns:
            Row type character (H, R, T) or None
        """
        if not metadata:
            return None

        metadata_str = str(metadata).lower()
        if "$t=" in metadata_str:
            parts = metadata_str.split("$t=")
            if len(parts) > 1:
                return parts[1][0].upper()

        return None

    @staticmethod
    def _is_valid_gl_code(code: str) -> bool:
        """
        Validate GL code format (should be XXXX-XXXX).

        Args:
            code: GL code string

        Returns:
            True if valid format, False otherwise
        """
        parts = code.split("-")
        if len(parts) != 2:
            return False

        try:
            int(parts[0])
            int(parts[1])
            return True
        except ValueError:
            return False


def parse_ysl_file(file_path: Path) -> Tuple[Dict[str, Dict[str, float]], Dict[str, str]]:
    """
    Convenience function to parse a YSL file.

    Args:
        file_path: Path to YSL .xlsx file

    Returns:
        Tuple of (gl_data, property_info)
    """
    parser = YSLParser(file_path)
    return parser.parse()
