"""
Fill the budget template with YSL data.

Loads template, maps YSL data to template locations, preserves formulas.
"""

import logging
from pathlib import Path
from typing import Dict, Optional, Tuple, Any
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import YSL_COLUMN_MAPPING, SETUP_SHEET_FIELDS
from gl_mapper import GLMapper

logger = logging.getLogger(__name__)


class TemplatePopulator:
    """Populates template workbook with YSL data."""

    def __init__(self, template_path: Path, output_path: Path):
        """
        Initialize populator.

        Args:
            template_path: Path to template.xlsx
            output_path: Path where filled template will be saved
        """
        self.template_path = Path(template_path)
        self.output_path = Path(output_path)

        # Load template with formulas preserved (data_only=False)
        self.wb = load_workbook(self.template_path, data_only=False)
        self.gl_mapper = GLMapper(self.template_path)
        self.gl_mapping = self.gl_mapper.build_mapping()

        self.stats = {
            "gl_codes_matched": 0,
            "gl_codes_unmatched": 0,
            "cells_filled": 0,
            "errors": [],
        }

    # Budget Summary row → detail sheet/row ranges for current year budget
    # Maps Budget Summary row number to (sheet_name, start_row, end_row) for summing period_5
    BUDGET_SUMMARY_MAPPING = {
        8: [("Income", None, None)],           # Total Operating Income (all Income GL codes)
        11: [("Payroll", None, None)],          # Payroll & Related
        12: [("Energy", None, None)],           # Energy
        13: [("Water & Sewer", None, None)],    # Water & Sewer
        14: [("Repairs & Supplies", None, None)], # Repairs & Supplies
        15: [("Gen & Admin", 8, 16)],           # Professional Fees: SUM(D8:D16)
        16: [("Gen & Admin", 20, 49)],          # Administrative & Other: SUM(D20:D49)
        17: [("Gen & Admin", 53, 64)],          # Insurance: SUM(D53:D64)
        18: [("Gen & Admin", 68, 78)],          # Taxes (RE + Corp): SUM(D68:D78)
        19: [("Gen & Admin", 82, 90)],          # Financial Expenses: SUM(D82:D90)
    }

    def populate(
        self,
        gl_data: Dict[str, Dict[str, float]],
        property_info: Dict[str, str],
        ytd_months: int = 0,
        remaining_months: int = 0,
    ) -> bool:
        """
        Populate template with YSL data.

        Args:
            gl_data: Dict of GL codes with period values
            property_info: Dict with property_code and property_name
            ytd_months: Number of months into year (for Setup sheet)
            remaining_months: Remaining months in year (for Setup sheet)

        Returns:
            True if successful, False otherwise
        """
        try:
            # Update Setup sheet
            self._populate_setup_sheet(property_info, ytd_months, remaining_months)

            # Populate GL codes
            for gl_code, values in gl_data.items():
                if gl_code not in self.gl_mapping:
                    self.stats["gl_codes_unmatched"] += 1
                    logger.debug(f"GL code {gl_code} not found in template mapping")
                    continue

                sheet_name, row_num = self.gl_mapping[gl_code]
                self._populate_gl_row(sheet_name, row_num, values)
                self.stats["gl_codes_matched"] += 1

            # Populate Insurance Schedule Col C with prior year actual (period_2)
            self._populate_insurance_schedule(gl_data)

            # Populate Budget Summary Col D with current year budget (period_5)
            self._populate_budget_summary_current_year(gl_data)

            logger.info(f"Population complete. Matched: {self.stats['gl_codes_matched']}, "
                       f"Unmatched: {self.stats['gl_codes_unmatched']}")
            return True

        except Exception as e:
            logger.error(f"Error during population: {e}")
            self.stats["errors"].append(str(e))
            return False

    # Insurance Schedule GL code → row mapping
    INSURANCE_SCHEDULE_GL_ROWS = {
        "6105-0000": 12, "6110-0000": 13, "6115-0000": 14,
        "6125-0000": 15, "6126-0000": 16, "6135-0000": 17,
        "6195-0000": 18, "6120-0000": 19, "6180-0000": 20,
    }

    def _populate_insurance_schedule(
        self, gl_data: Dict[str, Dict[str, float]]
    ) -> None:
        """
        Populate Insurance Schedule:
          Col C (Current Annual Premium) with period_2 (prior year actual)
          Col D (Current Year Budget) with period_5 (2026 approved budget)
        """
        if "Insurance Schedule" not in self.wb.sheetnames:
            logger.warning("Insurance Schedule sheet not found")
            return

        ws = self.wb["Insurance Schedule"]

        for gl_code, ins_row in self.INSURANCE_SCHEDULE_GL_ROWS.items():
            if gl_code in gl_data:
                # Col C = prior year actual (period_2)
                value_c = gl_data[gl_code].get("period_2", 0) or 0
                ws.cell(row=ins_row, column=3, value=value_c)  # Col C
                self.stats["cells_filled"] += 1
                logger.debug(f"Insurance Schedule C{ins_row} ({gl_code}) = {value_c:,.2f}")

                # Col D = current year budget (period_5)
                value_d = gl_data[gl_code].get("period_5", 0) or 0
                ws.cell(row=ins_row, column=4, value=value_d)  # Col D
                self.stats["cells_filled"] += 1
                logger.debug(f"Insurance Schedule D{ins_row} ({gl_code}) = {value_d:,.2f}")

    def _populate_budget_summary_current_year(
        self, gl_data: Dict[str, Dict[str, float]]
    ) -> None:
        """
        Populate Budget Summary Col D (Current Year Budget) with period_5 totals.

        Sums YSL period_5 values by category based on which detail sheet and
        row range each GL code maps to.
        """
        if "Budget Summary" not in self.wb.sheetnames:
            logger.warning("Budget Summary sheet not found")
            return

        ws = self.wb["Budget Summary"]

        for summary_row, sources in self.BUDGET_SUMMARY_MAPPING.items():
            total = 0.0
            for sheet_name, start_row, end_row in sources:
                for gl_code, values in gl_data.items():
                    if gl_code not in self.gl_mapping:
                        continue
                    mapped_sheet, mapped_row = self.gl_mapping[gl_code]
                    if mapped_sheet != sheet_name:
                        continue
                    if start_row is not None and not (start_row <= mapped_row <= end_row):
                        continue
                    p5 = values.get("period_5", 0) or 0
                    total += p5

            ws.cell(row=summary_row, column=4, value=total)  # Col D
            self.stats["cells_filled"] += 1
            logger.debug(f"Budget Summary D{summary_row} = {total:,.2f}")

    def _populate_setup_sheet(
        self,
        property_info: Dict[str, str],
        ytd_months: int,
        remaining_months: int,
    ) -> None:
        """
        Update Setup sheet with property info and time periods.

        Args:
            property_info: Dict with property_code and property_name
            ytd_months: Number of months YTD
            remaining_months: Remaining months in year
        """
        if "Setup" not in self.wb.sheetnames:
            logger.warning("Setup sheet not found in template")
            return

        ws = self.wb["Setup"]

        # Update property code (B4)
        if property_info.get("property_code"):
            ws["B4"] = property_info["property_code"]
            self.stats["cells_filled"] += 1

        # Update property name (B5)
        if property_info.get("property_name"):
            ws["B5"] = property_info["property_name"]
            self.stats["cells_filled"] += 1

        # Update YTD months (B10)
        if ytd_months > 0:
            ws["B10"] = ytd_months
            self.stats["cells_filled"] += 1

        # Update remaining months (B11)
        if remaining_months > 0:
            ws["B11"] = remaining_months
            self.stats["cells_filled"] += 1

    def _populate_gl_row(
        self, sheet_name: str, row_num: int, values: Dict[str, Optional[float]]
    ) -> None:
        """
        Populate a GL code row with YSL values.

        Maps YSL period columns to template columns D, E, H based on configuration.

        Args:
            sheet_name: Name of sheet containing this GL code
            row_num: Row number (1-based)
            values: Dict with period_1 through period_5 values
        """
        ws = self.wb[sheet_name]

        # Map YSL column indices to template columns
        # YSL structure: period_1 (D), period_2 (E), period_3 (F), period_4 (G), period_5 (H)
        column_mapping = [
            ("period_2", "D"),  # YSL col E (2025 Actual) -> Template col D (Prior Year Actual)
            ("period_3", "E"),  # YSL col F (YTD 02/2026 Actual) -> Template col E (YTD Actual)
            ("period_4", "H"),  # YSL col G (YTD 02/2026 Budget) -> Template col H (YTD Budget)
            ("period_5", "K"),  # YSL col H (2026 Approved Budget) -> Template col K (Current Year Budget)
        ]

        for ysl_key, template_col in column_mapping:
            if ysl_key not in values:
                continue

            value = values[ysl_key]
            if value is None:
                continue

            try:
                cell_addr = f"{template_col}{row_num}"
                ws[cell_addr] = value
                self.stats["cells_filled"] += 1
                logger.debug(f"Filled {sheet_name}!{cell_addr} with {value}")
            except Exception as e:
                logger.error(f"Error populating {sheet_name}!{template_col}{row_num}: {e}")
                self.stats["errors"].append(
                    f"Error populating {sheet_name}!{template_col}{row_num}: {e}"
                )

    def save(self) -> bool:
        """
        Save the populated workbook to output path.

        Returns:
            True if successful, False otherwise
        """
        try:
            # Ensure output directory exists
            self.output_path.parent.mkdir(parents=True, exist_ok=True)

            self.wb.save(self.output_path)
            logger.info(f"Saved populated template to {self.output_path}")
            return True

        except Exception as e:
            logger.error(f"Error saving workbook: {e}")
            self.stats["errors"].append(f"Save error: {e}")
            return False

    def close(self) -> None:
        """Close workbooks and cleanup."""
        self.wb.close()
        self.gl_mapper.close()

    def get_stats(self) -> Dict[str, Any]:
        """
        Get population statistics.

        Returns:
            Dict with stats on matches, fills, and errors
        """
        return self.stats.copy()


def apply_assumptions(workbook_path: Path, assumptions: Dict) -> bool:
    """
    Apply budget assumptions to yellow input cells in a generated workbook.

    Args:
        workbook_path: Path to the generated budget .xlsx
        assumptions: Merged assumptions dict (portfolio defaults + building overrides)

    Returns:
        True if successful
    """
    if not assumptions:
        logger.info("No assumptions provided, skipping")
        return True

    try:
        wb = load_workbook(workbook_path, data_only=False)

        # === PAYROLL SHEET ===
        if "Payroll" in wb.sheetnames:
            ws = wb["Payroll"]

            # Tax rates: B7-B12
            tax = assumptions.get("payroll_tax", {})
            tax_mapping = [
                (7, "FICA"), (8, "SUI"), (9, "FUI"),
                (10, "MTA"), (11, "NYS_Disability"), (12, "PFL"),
            ]
            for row, key in tax_mapping:
                val = tax.get(key)
                if val:
                    ws.cell(row=row, column=2, value=val)

            # Union benefit rates: H7-H12
            union = assumptions.get("union_benefits", {})
            union_mapping = [
                (7, "welfare_monthly"), (8, "pension_weekly"),
                (9, "supp_retirement_weekly"), (10, "legal_monthly"),
                (11, "training_monthly"), (12, "profit_sharing_quarterly"),
            ]
            for row, key in union_mapping:
                val = union.get(key)
                if val:
                    ws.cell(row=row, column=8, value=val)  # Col H

            # Workers comp %: L7
            wc = assumptions.get("workers_comp", {})
            if wc.get("percent"):
                ws.cell(row=7, column=12, value=wc["percent"])  # L7

            # Wage increase: L8, L9, L10, L11
            wage = assumptions.get("wage_increase", {})
            if wage.get("percent"):
                ws.cell(row=8, column=12, value=wage["percent"])  # L8
            if wage.get("effective_week"):
                ws.cell(row=9, column=12, value=wage["effective_week"])  # L9
            if wage.get("pre_increase_weeks"):
                ws.cell(row=10, column=12, value=wage["pre_increase_weeks"])  # L10
            if wage.get("post_increase_weeks"):
                ws.cell(row=11, column=12, value=wage["post_increase_weeks"])  # L11

            # Employee positions: A17-C24 (name, count, rate)
            payroll = assumptions.get("payroll", {})
            positions = payroll.get("positions", [])
            for i, pos in enumerate(positions[:8]):
                row = 17 + i
                if pos.get("name"):
                    ws.cell(row=row, column=1, value=pos["name"])
                if pos.get("employee_count"):
                    ws.cell(row=row, column=2, value=pos["employee_count"])
                if pos.get("hourly_rate"):
                    ws.cell(row=row, column=3, value=pos["hourly_rate"])

        # === INSURANCE SCHEDULE ===
        if "Insurance Schedule" in wb.sheetnames:
            ws = wb["Insurance Schedule"]

            ins_renewal = assumptions.get("insurance_renewal", {})
            if ins_renewal.get("increase_percent"):
                ws.cell(row=6, column=3, value=ins_renewal["increase_percent"])  # C6
            if ins_renewal.get("effective_date"):
                ws.cell(row=7, column=3, value=ins_renewal["effective_date"])  # C7
            if ins_renewal.get("pre_renewal_months"):
                ws.cell(row=8, column=3, value=ins_renewal["pre_renewal_months"])  # C8
            if ins_renewal.get("post_renewal_months"):
                ws.cell(row=9, column=3, value=ins_renewal["post_renewal_months"])  # C9

            # Per-policy data: rows 12-20
            ins_gl_rows = {
                "6105-0000": 12, "6110-0000": 13, "6115-0000": 14,
                "6125-0000": 15, "6126-0000": 16, "6135-0000": 17,
                "6195-0000": 18, "6120-0000": 19, "6180-0000": 20,
            }
            for policy in assumptions.get("insurance", []):
                gl = policy.get("gl_code", "")
                if gl in ins_gl_rows:
                    row = ins_gl_rows[gl]
                    if policy.get("current_premium"):
                        ws.cell(row=row, column=3, value=policy["current_premium"])  # C
                    if policy.get("current_budget"):
                        ws.cell(row=row, column=4, value=policy["current_budget"])  # D
                    if policy.get("expiration_date"):
                        ws.cell(row=row, column=5, value=policy["expiration_date"])  # E
                    if policy.get("override_increase"):
                        ws.cell(row=row, column=11, value=policy["override_increase"])  # K

        # === INCOME SHEET ===
        if "Income" in wb.sheetnames:
            ws = wb["Income"]

            income = assumptions.get("income", {})
            maint = income.get("maintenance", {})
            if maint.get("total_shares"):
                ws.cell(row=6, column=2, value=maint["total_shares"])  # B6
            if maint.get("per_share_monthly"):
                ws.cell(row=6, column=4, value=maint["per_share_monthly"])  # D6
            if maint.get("increase_percent"):
                ws.cell(row=7, column=2, value=maint["increase_percent"])  # B7

            # Storage: rows 12-15 (B=count, C=occupied, D=rate)
            for i, unit in enumerate(income.get("storage", [])[:4]):
                row = 12 + i
                if unit.get("size_label"):
                    ws.cell(row=row, column=1, value=unit["size_label"])  # A
                if unit.get("units"):
                    ws.cell(row=row, column=2, value=unit["units"])  # B
                if unit.get("occupied"):
                    ws.cell(row=row, column=3, value=unit["occupied"])  # C
                if unit.get("monthly"):
                    ws.cell(row=row, column=4, value=unit["monthly"])  # D

            # Bike storage: B19, C19, D19
            bike = income.get("bike_storage", {})
            if bike.get("racks"):
                ws.cell(row=19, column=2, value=bike["racks"])
            if bike.get("occupied"):
                ws.cell(row=19, column=3, value=bike["occupied"])
            if bike.get("monthly"):
                ws.cell(row=19, column=4, value=bike["monthly"])

            # Laundry: C23, D23
            laundry = income.get("laundry_vending", {})
            if laundry.get("description"):
                ws.cell(row=23, column=3, value=laundry["description"])
            if laundry.get("monthly"):
                ws.cell(row=23, column=4, value=laundry["monthly"])

        # === ENERGY SHEET ===
        if "Energy" in wb.sheetnames:
            ws = wb["Energy"]

            energy = assumptions.get("energy", {})
            if energy.get("gas_esco_rate"):
                ws.cell(row=6, column=2, value=energy["gas_esco_rate"])  # B6
            if energy.get("electric_esco_rate"):
                ws.cell(row=7, column=2, value=energy["electric_esco_rate"])  # B7
            if energy.get("gas_rate_increase"):
                ws.cell(row=8, column=2, value=energy["gas_rate_increase"])  # B8
            if energy.get("electric_rate_increase"):
                ws.cell(row=9, column=2, value=energy["electric_rate_increase"])  # B9
            if energy.get("consumption_basis"):
                ws.cell(row=10, column=2, value=energy["consumption_basis"])  # B10
            if energy.get("oil_price_per_gallon"):
                ws.cell(row=11, column=2, value=energy["oil_price_per_gallon"])  # B11 (repurposed)
            if energy.get("oil_rate_increase"):
                # Oil rate increase goes to the GL adjustment row L59
                ws.cell(row=59, column=12, value=energy["oil_rate_increase"])

            # Energy GL adjustments
            energy_gl_rows = {
                "5252-0000": 56, "5252-0001": 57, "5252-0010": 58,
                "5253-0000": 59, "5250-0000": 60,
            }
            for gl, row in energy_gl_rows.items():
                adj = energy.get("gl_adjustments", {}).get(gl, {})
                if adj.get("accrual"):
                    ws.cell(row=row, column=6, value=adj["accrual"])  # F
                if adj.get("unpaid"):
                    ws.cell(row=row, column=7, value=adj["unpaid"])  # G
                if adj.get("rate_increase"):
                    ws.cell(row=row, column=12, value=adj["rate_increase"])  # L

        # === WATER & SEWER SHEET ===
        if "Water & Sewer" in wb.sheetnames:
            ws = wb["Water & Sewer"]

            water = assumptions.get("water_sewer", {})
            if water.get("rate_increase"):
                ws.cell(row=6, column=2, value=water["rate_increase"])  # B6

            # Water GL adjustments
            water_gl_rows = {
                "6305-0000": 30, "6305-0010": 31, "6305-0020": 32,
            }
            for gl, row in water_gl_rows.items():
                adj = water.get("gl_adjustments", {}).get(gl, {})
                if adj.get("accrual"):
                    ws.cell(row=row, column=6, value=adj["accrual"])  # F
                if adj.get("unpaid"):
                    ws.cell(row=row, column=7, value=adj["unpaid"])  # G
                if adj.get("rate_increase"):
                    ws.cell(row=row, column=12, value=adj["rate_increase"])  # L

        wb.save(workbook_path)
        wb.close()
        logger.info(f"Assumptions applied to {workbook_path.name}")
        return True

    except Exception as e:
        logger.error(f"Error applying assumptions: {e}")
        return False


def apply_pm_projections(workbook_path: Path, projections: Dict) -> bool:
    """
    Apply PM R&M projections to the Repairs & Supplies sheet in a generated workbook.

    Writes to:
        - Col C (notes)
        - Col F (accrual adjustment)
        - Col G (unpaid bills)
        - Col L (increase %)

    The template formulas then compute:
        - Col I: Sep-Dec Estimate
        - Col J: 12-Month Forecast = E + F + G + I
        - Col M: Proposed Budget = J * (1 + L)

    Args:
        workbook_path: Path to the generated budget .xlsx
        projections: Dict of {gl_code: {accrual_adj, unpaid_bills, increase_pct, notes}}
                     Typically from workflow_helpers['get_pm_projections'](entity)

    Returns:
        True if successful
    """
    if not projections:
        logger.info("No PM projections to apply")
        return True

    # Import the GL→row mapping from workflow module
    try:
        from workflow import RM_GL_MAP
    except ImportError:
        # Fallback: build mapping dynamically from template
        logger.warning("Could not import RM_GL_MAP, skipping PM projections")
        return False

    try:
        wb = load_workbook(workbook_path, data_only=False)

        if "Repairs & Supplies" not in wb.sheetnames:
            logger.warning("No 'Repairs & Supplies' sheet found")
            wb.close()
            return False

        ws = wb["Repairs & Supplies"]
        cells_written = 0

        for gl_code, pm_data in projections.items():
            if gl_code not in RM_GL_MAP:
                continue

            _, row_num, _ = RM_GL_MAP[gl_code]

            # Col C (3) = Notes
            notes = pm_data.get("notes", "")
            if notes:
                ws.cell(row=row_num, column=3, value=notes)
                cells_written += 1

            # Col F (6) = Accrual Adjustment
            accrual = float(pm_data.get("accrual_adj", 0) or 0)
            if accrual:
                ws.cell(row=row_num, column=6, value=accrual)
                cells_written += 1

            # Col G (7) = Unpaid Bills
            unpaid = float(pm_data.get("unpaid_bills", 0) or 0)
            if unpaid:
                ws.cell(row=row_num, column=7, value=unpaid)
                cells_written += 1

            # Col L (12) = Increase %
            increase = float(pm_data.get("increase_pct", 0) or 0)
            if increase:
                ws.cell(row=row_num, column=12, value=increase)
                cells_written += 1

        wb.save(workbook_path)
        wb.close()
        logger.info(f"PM projections applied: {cells_written} cells written to {workbook_path.name}")
        return True

    except Exception as e:
        logger.error(f"Error applying PM projections: {e}")
        return False


def populate_template(
    template_path: Path,
    gl_data: Dict[str, Dict[str, float]],
    property_info: Dict[str, str],
    output_path: Path,
    ytd_months: int = 0,
    remaining_months: int = 0,
) -> bool:
    """
    Convenience function to populate template.

    Args:
        template_path: Path to template.xlsx
        gl_data: Dict of GL codes with values
        property_info: Dict with property info
        output_path: Path to save filled template
        ytd_months: Number of months YTD
        remaining_months: Remaining months

    Returns:
        True if successful
    """
    populator = TemplatePopulator(template_path, output_path)
    success = populator.populate(gl_data, property_info, ytd_months, remaining_months)
    if success:
        populator.save()
    populator.close()
    return success
