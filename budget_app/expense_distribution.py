"""
Expense Distribution Report Parser & PM Review Blueprint for Century Management.

Parses Yardi "Expense Distribution (Paid Only)" .xlsx exports and provides:
- Invoice-level drill-down per GL code in PM review
- Reclassification of invoices between GL codes with adjusted totals
- Budget vs actual variance analysis per GL
"""

from flask import Blueprint, render_template_string, request, jsonify
from datetime import datetime
import logging
import re

logger = logging.getLogger(__name__)

# G&A GL codes that appear in the Expense Distribution report
# but are not in workflow.py's RM_GL_MAP
GA_GL_MAP = {
    "6505-0000": ("Engineering & Architectural Fees", "professional"),
    "6510-0000": ("Management Fees", "professional"),
    "6515-0000": ("Legal Fees", "professional"),
    "6525-0000": ("Accounting Fees", "professional"),
    "6590-0000": ("Other Professional Fees", "professional"),
    "6710-0000": ("Postage Expense", "admin"),
    "6710-0015": ("Messenger Expense", "admin"),
    "6714-0000": ("Telephone Expense", "admin"),
    "6716-0000": ("Cable TV Expense", "admin"),
    "6718-0000": ("Photocopy Expense", "admin"),
    "6720-0000": ("Dues & Subscriptions", "admin"),
    "6726-0000": ("Inspection Fees & Permits", "admin"),
    "6734-0000": ("Gratuities", "admin"),
    "6746-0000": ("Lobby & Hallway Decorations", "admin"),
    "6754-0000": ("Credit Check", "admin"),
    "6762-0000": ("1098 Fee's", "admin"),
    "6763-0000": ("Safety Notices", "admin"),
    "6795-0000": ("Other Administrative", "admin"),
}


def parse_expense_distribution(file_path):
    """
    Parse an Expense Distribution (Paid Only) .xlsx export from Yardi.

    Report structure:
      Row 1: Title ("Expense Distribution (Paid Only)")
      Row 2: Entity code
      Row 3: Period range ("Period: From MM/YYYY to MM/YYYY")
      Row 4: Column headers (A-P)
      Row 5+: GL groups:
        - GL header: col A = "XXXX-XXXX", col B = account name
        - Invoice rows: col A empty, cols C-P have detail
        - Total row: col A = "Total XXXX-XXXX", col L = sum
      Last row: "Grand Total" with overall sum

    Returns:
        tuple: (entity_code, period_from, period_to, invoices_list)
        Each invoice is a dict with keys matching ExpenseInvoice model fields.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Row 2: Entity code
    entity_code = str(ws.cell(row=2, column=1).value or "").strip()

    # Row 3: Period range
    period_str = str(ws.cell(row=3, column=1).value or "")
    period_from = None
    period_to = None
    period_match = re.search(r'From\s+(\d{2}/\d{4})\s+to\s+(\d{2}/\d{4})', period_str)
    if period_match:
        period_from = period_match.group(1)
        period_to = period_match.group(2)

    invoices = []
    current_gl = None
    current_gl_name = None

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        a_val = row[0].value  # Column A
        b_val = row[1].value  # Column B

        if a_val and isinstance(a_val, str):
            a_val = a_val.strip()

            # Skip total and grand total rows
            if a_val.startswith("Total") or a_val.startswith("Grand Total"):
                current_gl = None
                current_gl_name = None
                continue

            # GL code header row (format: XXXX-XXXX)
            if re.match(r'^\d{4}-\d{4}$', a_val):
                current_gl = a_val
                current_gl_name = str(b_val or "").strip()
                continue

        # Invoice detail row: col A is empty, col C (payee code) has a value
        if current_gl and row[2].value is not None:
            amount = row[11].value  # Column L
            if amount is None:
                continue

            invoice_date = row[8].value  # Column I
            period = row[9].value       # Column J
            check_date = row[14].value  # Column O

            invoices.append({
                "gl_code": current_gl,
                "gl_name": current_gl_name,
                "payee_code": str(row[2].value or "").strip(),
                "payee_name": str(row[3].value or "").strip(),
                "payable_control": str(row[4].value or "").strip(),
                "batch": str(row[5].value or "").strip(),
                "property_code": str(row[6].value or "").strip(),
                "invoice_num": str(row[7].value or "").strip(),
                "invoice_date": invoice_date.isoformat() if hasattr(invoice_date, 'isoformat') else str(invoice_date or ""),
                "period": period.isoformat() if hasattr(period, 'isoformat') else str(period or ""),
                "payment_method": str(row[10].value or "").strip(),
                "amount": float(amount),
                "check_control": str(row[12].value or "").strip(),
                "check_num": str(row[13].value or "").strip(),
                "check_date": check_date.isoformat() if hasattr(check_date, 'isoformat') else str(check_date or ""),
                "notes": str(row[15].value or "").strip(),
            })

    wb.close()
    return (entity_code, period_from, period_to, invoices)


def create_expense_distribution_blueprint(db, workflow_models):
    """
    Create the expense distribution blueprint.

    Args:
        db: SQLAlchemy instance
        workflow_models: dict from workflow blueprint (Budget, BudgetLine, etc.)

    Returns:
        tuple: (blueprint, models_dict, helpers_dict)
    """
    bp = Blueprint("expense_dist", __name__)

    Budget = workflow_models["Budget"]
    BudgetLine = workflow_models["BudgetLine"]
    BudgetRevision = workflow_models.get("BudgetRevision")  # For audit trail

    # ─── Models ───────────────────────────────────────────────────────────────

    class ExpenseReport(db.Model):
        """Uploaded Expense Distribution report metadata."""
        __tablename__ = "expense_reports"

        id = db.Column(db.Integer, primary_key=True)
        entity_code = db.Column(db.String(50), nullable=False, index=True)
        period_from = db.Column(db.String(10))   # MM/YYYY
        period_to = db.Column(db.String(10))     # MM/YYYY
        file_name = db.Column(db.String(255))
        total_amount = db.Column(db.Float, default=0.0)
        uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

        invoices = db.relationship("ExpenseInvoice", back_populates="report",
                                   cascade="all, delete-orphan")

        def to_dict(self):
            return {
                "id": self.id,
                "entity_code": self.entity_code,
                "period_from": self.period_from,
                "period_to": self.period_to,
                "file_name": self.file_name,
                "total_amount": float(self.total_amount or 0),
                "uploaded_at": self.uploaded_at.isoformat() if self.uploaded_at else None,
                "invoice_count": len(self.invoices) if self.invoices else 0,
            }

    class ExpenseInvoice(db.Model):
        """Individual invoice line from Expense Distribution report."""
        __tablename__ = "expense_invoices"

        id = db.Column(db.Integer, primary_key=True)
        report_id = db.Column(db.Integer, db.ForeignKey("expense_reports.id"), nullable=False)
        gl_code = db.Column(db.String(50), nullable=False, index=True)
        gl_name = db.Column(db.String(255))
        payee_code = db.Column(db.String(100))
        payee_name = db.Column(db.String(255))
        payable_control = db.Column(db.String(50))
        batch = db.Column(db.String(50))
        property_code = db.Column(db.String(50))
        invoice_num = db.Column(db.String(100))
        invoice_date = db.Column(db.String(30))
        period = db.Column(db.String(30))
        payment_method = db.Column(db.String(50))
        amount = db.Column(db.Float, nullable=False)
        check_control = db.Column(db.String(50))
        check_num = db.Column(db.String(50))
        check_date = db.Column(db.String(30))
        notes = db.Column(db.Text, default="")

        # Reclass fields
        reclass_to_gl = db.Column(db.String(50))
        reclass_notes = db.Column(db.Text)
        reclassed_by = db.Column(db.String(100))
        reclassed_at = db.Column(db.DateTime)

        report = db.relationship("ExpenseReport", back_populates="invoices")

        def to_dict(self):
            return {
                "id": self.id,
                "report_id": self.report_id,
                "gl_code": self.gl_code,
                "gl_name": self.gl_name,
                "payee_code": self.payee_code,
                "payee_name": self.payee_name,
                "invoice_num": self.invoice_num,
                "invoice_date": self.invoice_date,
                "period": self.period,
                "payment_method": self.payment_method,
                "amount": float(self.amount or 0),
                "check_num": self.check_num,
                "check_date": self.check_date,
                "notes": self.notes,
                "reclass_to_gl": self.reclass_to_gl,
                "reclass_notes": self.reclass_notes,
                "reclassed_by": self.reclassed_by,
                "reclassed_at": self.reclassed_at.isoformat() if self.reclassed_at else None,
            }

    # ─── Helper: Store parsed invoices ────────────────────────────────────────

    def store_expense_report(entity_code, period_from, period_to, invoices, file_name=""):
        """Store parsed expense distribution data. Replaces existing report for same entity/period."""
        # Remove existing report for same entity and period range
        existing = ExpenseReport.query.filter_by(
            entity_code=entity_code,
            period_from=period_from,
            period_to=period_to
        ).first()
        if existing:
            db.session.delete(existing)
            db.session.flush()

        total = sum(inv["amount"] for inv in invoices)
        report = ExpenseReport(
            entity_code=entity_code,
            period_from=period_from,
            period_to=period_to,
            file_name=file_name,
            total_amount=total,
        )
        db.session.add(report)
        db.session.flush()

        for inv in invoices:
            db.session.add(ExpenseInvoice(
                report_id=report.id,
                gl_code=inv["gl_code"],
                gl_name=inv["gl_name"],
                payee_code=inv.get("payee_code", ""),
                payee_name=inv.get("payee_name", ""),
                payable_control=inv.get("payable_control", ""),
                batch=inv.get("batch", ""),
                property_code=inv.get("property_code", ""),
                invoice_num=inv.get("invoice_num", ""),
                invoice_date=inv.get("invoice_date", ""),
                period=inv.get("period", ""),
                payment_method=inv.get("payment_method", ""),
                amount=inv["amount"],
                check_control=inv.get("check_control", ""),
                check_num=inv.get("check_num", ""),
                check_date=inv.get("check_date", ""),
                notes=inv.get("notes", ""),
            ))

        db.session.commit()
        logger.info(f"Stored {len(invoices)} invoices for entity {entity_code} ({period_from}-{period_to})")
        return report

    # ─── Helper: Get adjusted GL totals (after reclass) ───────────────────────

    def get_adjusted_gl_totals(entity_code):
        """
        Calculate GL totals adjusted for reclassifications.

        Returns dict: {gl_code: {"original": float, "reclass_out": float, "reclass_in": float, "adjusted": float}}
        """
        report = ExpenseReport.query.filter_by(entity_code=entity_code)\
            .order_by(ExpenseReport.uploaded_at.desc()).first()
        if not report:
            return {}

        invoices = ExpenseInvoice.query.filter_by(report_id=report.id).all()

        totals = {}

        for inv in invoices:
            # Original GL totals
            if inv.gl_code not in totals:
                totals[inv.gl_code] = {"original": 0, "reclass_out": 0, "reclass_in": 0, "gl_name": inv.gl_name}
            totals[inv.gl_code]["original"] += inv.amount

            # Handle reclassifications
            if inv.reclass_to_gl:
                totals[inv.gl_code]["reclass_out"] += inv.amount
                if inv.reclass_to_gl not in totals:
                    totals[inv.reclass_to_gl] = {"original": 0, "reclass_out": 0, "reclass_in": 0, "gl_name": ""}
                totals[inv.reclass_to_gl]["reclass_in"] += inv.amount

        # Calculate adjusted totals
        for gl, t in totals.items():
            t["adjusted"] = t["original"] - t["reclass_out"] + t["reclass_in"]

        return totals

    # ─── API Routes ───────────────────────────────────────────────────────────

    @bp.route("/api/expense-dist/upload", methods=["POST"])
    def upload_expense_dist():
        """Upload and parse an Expense Distribution .xlsx file."""
        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files["file"]
        if not file.filename.endswith(".xlsx"):
            return jsonify({"error": "File must be .xlsx"}), 400

        import tempfile, os
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        file.save(tmp.name)
        tmp.close()

        try:
            entity_code, period_from, period_to, invoices = parse_expense_distribution(tmp.name)

            if not entity_code:
                return jsonify({"error": "Could not detect entity code from report"}), 400

            report = store_expense_report(
                entity_code, period_from, period_to, invoices, file.filename
            )

            return jsonify({
                "status": "ok",
                "report": report.to_dict(),
                "gl_codes_found": len(set(inv["gl_code"] for inv in invoices)),
                "invoices_parsed": len(invoices),
            })
        except Exception as e:
            logger.error(f"Failed to parse expense distribution: {e}")
            return jsonify({"error": str(e)}), 500
        finally:
            os.unlink(tmp.name)

    @bp.route("/api/expense-dist/<entity_code>", methods=["GET"])
    def get_expense_dist(entity_code):
        """Get the latest expense distribution report for a building."""
        report = ExpenseReport.query.filter_by(entity_code=entity_code)\
            .order_by(ExpenseReport.uploaded_at.desc()).first()
        if not report:
            return jsonify({"error": "No expense distribution report found"}), 404

        invoices = ExpenseInvoice.query.filter_by(report_id=report.id)\
            .order_by(ExpenseInvoice.gl_code, ExpenseInvoice.invoice_date).all()

        # Group by GL code
        gl_groups = {}
        for inv in invoices:
            gl = inv.gl_code
            if gl not in gl_groups:
                gl_groups[gl] = {"gl_code": gl, "gl_name": inv.gl_name, "invoices": [], "total": 0}
            gl_groups[gl]["invoices"].append(inv.to_dict())
            gl_groups[gl]["total"] += inv.amount

        return jsonify({
            "report": report.to_dict(),
            "gl_groups": list(gl_groups.values()),
        })

    @bp.route("/api/expense-dist/<entity_code>/<gl_code>", methods=["GET"])
    def get_gl_invoices(entity_code, gl_code):
        """Get invoices for a specific GL code."""
        report = ExpenseReport.query.filter_by(entity_code=entity_code)\
            .order_by(ExpenseReport.uploaded_at.desc()).first()
        if not report:
            return jsonify({"error": "No expense distribution report found"}), 404

        invoices = ExpenseInvoice.query.filter_by(
            report_id=report.id, gl_code=gl_code
        ).order_by(ExpenseInvoice.invoice_date).all()

        return jsonify({
            "gl_code": gl_code,
            "invoices": [inv.to_dict() for inv in invoices],
            "total": sum(inv.amount for inv in invoices),
            "reclass_count": sum(1 for inv in invoices if inv.reclass_to_gl),
        })

    @bp.route("/api/expense-dist/reclass/<int:invoice_id>", methods=["POST"])
    def reclass_invoice(invoice_id):
        """Reclassify an invoice to a different GL code."""
        inv = ExpenseInvoice.query.get(invoice_id)
        if not inv:
            return jsonify({"error": "Invoice not found"}), 404

        data = request.get_json()
        target_gl = data.get("reclass_to_gl", "").strip()
        notes = data.get("reclass_notes", "").strip()
        user = data.get("user", "PM")

        if not target_gl:
            # Clear reclass
            inv.reclass_to_gl = None
            inv.reclass_notes = None
            inv.reclassed_by = None
            inv.reclassed_at = None
        else:
            if target_gl == inv.gl_code:
                return jsonify({"error": "Cannot reclass to the same GL code"}), 400
            inv.reclass_to_gl = target_gl
            inv.reclass_notes = notes
            inv.reclassed_by = user
            inv.reclassed_at = datetime.utcnow()

        db.session.commit()

        return jsonify({"status": "ok", "invoice": inv.to_dict()})

    @bp.route("/api/expense-dist/<entity_code>/summary", methods=["GET"])
    def gl_summary(entity_code):
        """
        GL-level summary with budget variance.
        Combines expense distribution actuals with YSL budget data.
        """
        adjusted = get_adjusted_gl_totals(entity_code)

        # Get YSL budget data from BudgetLine
        budget = Budget.query.filter_by(entity_code=entity_code, year=2027).first()
        budget_lines = {}
        if budget:
            for line in BudgetLine.query.filter_by(budget_id=budget.id).all():
                budget_lines[line.gl_code] = {
                    "current_budget": float(line.current_budget or 0),
                    "ytd_actual": float(line.ytd_actual or 0),
                    "prior_year": float(line.prior_year or 0),
                }

        summary = []
        for gl_code, totals in sorted(adjusted.items()):
            budget_data = budget_lines.get(gl_code, {})
            current_budget = budget_data.get("current_budget", 0)
            adjusted_ytd = totals["adjusted"]

            over_under = adjusted_ytd - current_budget if current_budget else 0
            pct_over = (over_under / current_budget * 100) if current_budget else 0

            summary.append({
                "gl_code": gl_code,
                "gl_name": totals.get("gl_name", ""),
                "original_ytd": round(totals["original"], 2),
                "reclass_out": round(totals["reclass_out"], 2),
                "reclass_in": round(totals["reclass_in"], 2),
                "adjusted_ytd": round(adjusted_ytd, 2),
                "current_budget": round(current_budget, 2),
                "over_under": round(over_under, 2),
                "pct_over": round(pct_over, 1),
            })

        return jsonify({"entity_code": entity_code, "summary": summary})

    # ─── PM Review Page with Invoice Drill-Down ───────────────────────────────

    @bp.route("/pm/<entity_code>/expenses", methods=["GET"])
    def pm_expense_review(entity_code):
        """PM Expense Review page with invoice drill-down and reclass."""
        import json as json_mod
        try:
            from workflow import RM_GL_MAP
        except ImportError:
            from budget_app.workflow import RM_GL_MAP

        # Get all GL codes available for reclass dropdown
        all_gl_codes = {}
        all_gl_codes.update({k: v[0] for k, v in RM_GL_MAP.items()})
        all_gl_codes.update({k: v[0] for k, v in GA_GL_MAP.items()})

        # Get budget data for variance display
        budget = Budget.query.filter_by(entity_code=entity_code, year=2027).first()
        building_name = budget.building_name if budget else f"Entity {entity_code}"
        budget_lines = {}
        if budget:
            for line in BudgetLine.query.filter_by(budget_id=budget.id).all():
                budget_lines[line.gl_code] = {
                    "current_budget": float(line.current_budget or 0),
                    "ytd_actual": float(line.ytd_actual or 0),
                    "prior_year": float(line.prior_year or 0),
                }

        return render_template_string(
            PM_EXPENSE_TEMPLATE,
            entity_code=entity_code,
            building_name=building_name,
            all_gl_codes_json=json_mod.dumps(all_gl_codes),
            budget_lines_json=json_mod.dumps(budget_lines),
        )

    # ─── Helper: Apply accrual adjustments ─────────────────────────────────────

    def apply_accrual_adjustments(entity_code, report_id, period_from):
        """
        Identify prior-year invoices (invoice_date <= 12/31 of prior year)
        and sum them by GL code into BudgetLine.accrual_adj.

        Logic: The Expense Distribution report covers a current-year period
        (e.g., 01/2026–03/2026). Any invoice with an invoice_date in a prior
        year (i.e., <= 12/31/2025) represents a prior-year expense that was
        paid in the current year. These amounts are accrual adjustments that
        need to be backed out of YTD actuals for budgeting purposes.

        Args:
            entity_code: Building entity code (e.g., "204")
            report_id: ExpenseReport.id for the uploaded report
            period_from: Report start period as "MM/YYYY" string

        Returns:
            dict: {"applied": int, "accruals": {gl_code: amount}}
        """
        from datetime import datetime as _dt

        # Determine cutoff: 12/31 of the year BEFORE the report period
        try:
            parts = period_from.split("/")
            report_year = int(parts[1]) if len(parts) == 2 else int(parts[0])
        except (ValueError, IndexError):
            logger.warning(f"Could not parse period_from '{period_from}' for accrual cutoff")
            return {"applied": 0, "accruals": {}}

        prior_year = report_year - 1
        cutoff_str = f"{prior_year}-12-31"  # ISO format for comparison
        logger.info(f"Accrual cutoff for entity {entity_code}: invoice_date <= {cutoff_str}")

        # Get the report's invoices
        report = ExpenseReport.query.get(report_id)
        if not report or not report.invoices:
            logger.warning(f"No invoices found for report {report_id}")
            return {"applied": 0, "accruals": {}}

        # Sum prior-year invoices by GL code
        accruals = {}
        prior_count = 0
        for inv in report.invoices:
            inv_date_str = (inv.invoice_date or "").strip()
            if not inv_date_str:
                continue

            # Parse invoice date — handle multiple formats
            inv_date = None
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S"):
                try:
                    inv_date = _dt.strptime(inv_date_str[:10], fmt)
                    break
                except ValueError:
                    continue

            if inv_date is None:
                logger.warning(f"Could not parse invoice_date '{inv_date_str}' for invoice {inv.id}")
                continue

            # Check if this invoice is from prior year or earlier
            cutoff_date = _dt(prior_year, 12, 31)
            if inv_date <= cutoff_date:
                gl = inv.reclass_to_gl or inv.gl_code  # Use reclassed GL if applicable
                if gl not in accruals:
                    accruals[gl] = 0
                accruals[gl] += float(inv.amount or 0)
                prior_count += 1

        # Round totals
        accruals = {gl: round(amt, 2) for gl, amt in accruals.items()}

        if not accruals:
            logger.info(f"No prior-year invoices found for entity {entity_code} (checked {len(report.invoices)} invoices, cutoff={cutoff_str})")
            return {"applied": 0, "accruals": {}}

        logger.info(f"Found {prior_count} prior-year invoices across {len(accruals)} GL codes for entity {entity_code}")

        # Get the budget and update BudgetLine.accrual_adj
        budget = Budget.query.filter_by(entity_code=entity_code).order_by(Budget.year.desc(), Budget.version.desc()).first()
        if not budget:
            logger.warning(f"No budget found for entity {entity_code}, cannot apply accrual adjustments")
            return {"applied": 0, "accruals": accruals}

        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
        gl_to_line = {line.gl_code: line for line in lines}

        applied = 0
        for gl, total in accruals.items():
            if gl in gl_to_line:
                line = gl_to_line[gl]
                old_val = float(line.accrual_adj or 0)
                neg_total = -abs(total)  # Accruals are backed out of YTD → store as negative
                if abs(old_val - neg_total) > 0.01:
                    line.accrual_adj = neg_total

                    if BudgetRevision:
                        db.session.add(BudgetRevision(
                            budget_id=budget.id,
                            budget_line_id=line.id,
                            action="update",
                            field_name="accrual_adj",
                            old_value=str(old_val),
                            new_value=str(neg_total),
                            notes=f"GL {gl}: auto-calculated from Expense Distribution (invoices dated <= {cutoff_str})",
                            source="expense_dist_auto"
                        ))
                applied += 1
            else:
                logger.info(f"Accrual GL {gl} (${total:,.2f}) has no matching BudgetLine — skipped")

        db.session.commit()
        logger.info(f"Applied accrual adjustments to {applied} GL lines for entity {entity_code} (${sum(accruals.values()):,.2f} total)")
        return {"applied": applied, "accruals": accruals}

    # ─── Return blueprint ─────────────────────────────────────────────────────

    models = {
        "ExpenseReport": ExpenseReport,
        "ExpenseInvoice": ExpenseInvoice,
    }
    helpers = {
        "parse_expense_distribution": parse_expense_distribution,
        "store_expense_report": store_expense_report,
        "get_adjusted_gl_totals": get_adjusted_gl_totals,
        "apply_accrual_adjustments": apply_accrual_adjustments,
    }

    return (bp, models, helpers)


# ─── HTML Template ────────────────────────────────────────────────────────────

PM_EXPENSE_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Expense Review — {{ building_name }} — Century Management</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap');
  :root {
    --blue: #5a4a3f; --blue-light: #f5efe7;
    --green: #057a55; --green-light: #def7ec;
    --orange: #d97706; --orange-light: #fef3c7;
    --red: #e02424; --red-light: #fde8e8;
    --gray-50: #f4f1eb; --gray-100: #ede9e1; --gray-200: #e5e0d5;
    --gray-300: #d5cfc5; --gray-500: #8a7e72; --gray-700: #4a4039; --gray-900: #1a1714;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: var(--gray-50); color: var(--gray-900); line-height: 1.5; }
  header { background: linear-gradient(135deg, #2c2825 0%, #3d322a 100%); color: white; padding: 24px 20px; }
  header h1 { font-size: 24px; margin-bottom: 4px; }
  header p { opacity: 0.9; font-size: 14px; }
  .back-link { color: rgba(255,255,255,0.8); text-decoration: none; font-size: 14px; }
  .back-link:hover { color: white; }
  .container { max-width: 1500px; margin: 0 auto; padding: 24px 20px; }

  .upload-bar {
    display: flex; align-items: center; gap: 12px;
    background: white; border-radius: 12px; padding: 16px 24px;
    margin-bottom: 20px; border: 1px solid var(--gray-200);
  }
  .upload-bar input[type="file"] { flex: 1; }
  .upload-bar .status { font-size: 13px; color: var(--gray-500); }

  .summary-cards {
    display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 12px; margin-bottom: 20px;
  }
  .card {
    background: white; border-radius: 12px; padding: 16px 20px;
    border: 1px solid var(--gray-200);
  }
  .card .label { font-size: 12px; color: var(--gray-500); text-transform: uppercase; font-weight: 600; }
  .card .value { font-size: 24px; font-weight: 700; margin-top: 4px; }
  .card .value.negative { color: var(--red); }
  .card .value.positive { color: var(--green); }

  .grid-wrapper { background: white; border-radius: 12px; border: 1px solid var(--gray-200); overflow: hidden; }
  .grid-container { overflow-x: auto; max-height: 80vh; overflow-y: auto; }

  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  thead { background: var(--gray-100); position: sticky; top: 0; z-index: 10; }
  th { padding: 10px 12px; text-align: left; font-weight: 600; border-bottom: 2px solid var(--gray-300); white-space: nowrap; }
  th.number { text-align: right; }
  td { padding: 8px 12px; border-bottom: 1px solid var(--gray-200); }
  td.number { text-align: right; font-variant-numeric: tabular-nums; }

  .gl-row { cursor: pointer; }
  .gl-row:hover { background: var(--blue-light); }
  .gl-row td:first-child::before { content: '▶ '; font-size: 10px; color: var(--gray-500); }
  .gl-row.expanded td:first-child::before { content: '▼ '; }

  .invoice-row { background: var(--gray-50); }
  .invoice-row td { padding: 6px 12px 6px 36px; font-size: 12px; }
  .invoice-row.reclassed { background: #fef3c7; opacity: 0.7; text-decoration: line-through; }

  .over { color: var(--red); font-weight: 600; }
  .under { color: var(--green); }

  .btn {
    border: none; padding: 4px 10px; border-radius: 4px; font-size: 12px;
    font-weight: 600; cursor: pointer;
  }
  .btn-reclass { background: var(--orange-light); color: var(--orange); }
  .btn-reclass:hover { background: #fde68a; }
  .btn-undo { background: var(--gray-100); color: var(--gray-700); }
  .btn-upload { background: var(--blue); color: white; padding: 8px 20px; font-size: 14px; border-radius: 6px; }
  .btn-upload:hover { background: #3d322a; }

  /* Reclass modal */
  .modal-overlay {
    display: none; position: fixed; top: 0; left: 0; right: 0; bottom: 0;
    background: rgba(0,0,0,0.5); z-index: 100; align-items: center; justify-content: center;
  }
  .modal-overlay.active { display: flex; }
  .modal {
    background: white; border-radius: 12px; padding: 24px; width: 440px; max-width: 95vw;
    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
  }
  .modal h3 { margin-bottom: 16px; }
  .modal label { display: block; font-size: 13px; font-weight: 600; margin-bottom: 4px; margin-top: 12px; }
  .modal select, .modal textarea {
    width: 100%; padding: 8px 10px; border: 1px solid var(--gray-300);
    border-radius: 6px; font-size: 13px;
  }
  .modal textarea { height: 80px; resize: vertical; }
  .modal-actions { display: flex; gap: 8px; margin-top: 20px; justify-content: flex-end; }
  .modal-actions .btn { padding: 8px 20px; font-size: 14px; border-radius: 6px; }
  .btn-cancel { background: var(--gray-100); color: var(--gray-700); }
  .btn-save { background: var(--blue); color: white; }
  .btn-save:hover { background: #3d322a; }

  .no-data { text-align: center; padding: 60px 20px; color: var(--gray-500); }
  .no-data p { margin-bottom: 16px; }
</style>
</head>
<body>

<header>
  <a href="/" class="back-link">← Home</a>
  <a href="/pm" class="back-link">← Back to buildings</a>
  <h1>{{ building_name }}</h1>
  <p>Entity {{ entity_code }} — Expense Distribution Review</p>
</header>

<div class="container">
  <div class="upload-bar">
    <input type="file" id="fileInput" accept=".xlsx">
    <button class="btn btn-upload" onclick="uploadFile()">Upload Expense Report</button>
    <span class="status" id="uploadStatus"></span>
  </div>

  <div class="summary-cards" id="summaryCards" style="display:none;">
    <div class="card"><div class="label">Report Period</div><div class="value" id="cardPeriod" style="font-size:18px;">—</div></div>
    <div class="card"><div class="label">Total Invoices</div><div class="value" id="cardInvoices">0</div></div>
    <div class="card"><div class="label">GL Codes</div><div class="value" id="cardGLs">0</div></div>
    <div class="card"><div class="label">Total Paid</div><div class="value" id="cardTotal">$0</div></div>
  </div>

  <div class="grid-wrapper">
    <div class="grid-container">
      <div style="display:flex; justify-content:flex-end; margin-bottom:8px;">
        <button id="expZeroToggle" onclick="expToggleZeroRows()" style="font-size:12px; padding:5px 14px; background:#f5efe7; color:#5a4a3f; border:1px solid #5a4a3f; border-radius:6px; cursor:pointer;"></button>
      </div>
      <table id="expenseTable">
        <thead>
          <tr>
            <th>GL Code</th>
            <th>Description</th>
            <th class="number">YTD Paid</th>
            <th class="number">Reclass Adj</th>
            <th class="number">Adjusted YTD</th>
            <th class="number">Current Budget</th>
            <th class="number">Over/Under</th>
            <th class="number">% Variance</th>
            <th>Action</th>
          </tr>
        </thead>
        <tbody id="expenseBody">
          <tr><td colspan="9" class="no-data"><p>No expense distribution report loaded.</p><p>Upload a report above to get started.</p></td></tr>
        </tbody>
      </table>
    </div>
  </div>
</div>

<!-- Reclass Modal -->
<div class="modal-overlay" id="reclassModal">
  <div class="modal">
    <h3>Reclassify Invoice</h3>
    <div id="reclassInvoiceInfo" style="font-size:13px; color:var(--gray-500); margin-bottom:8px;"></div>
    <label>Move to GL Code</label>
    <select id="reclassTarget"></select>
    <label>Notes</label>
    <textarea id="reclassNotes" placeholder="Reason for reclassification..."></textarea>
    <div class="modal-actions">
      <button class="btn btn-cancel" onclick="closeReclassModal()">Cancel</button>
      <button class="btn btn-save" onclick="submitReclass()">Reclassify</button>
    </div>
  </div>
</div>

<script>
const ENTITY = "{{ entity_code }}";
const ALL_GL_CODES = {{ all_gl_codes_json | safe }};
const BUDGET_LINES = {{ budget_lines_json | safe }};

let reportData = null;
let expandedGLs = new Set();
let reclassInvoiceId = null;

function fmt(n) {
  if (n == null || isNaN(n)) return '$0';
  const neg = n < 0;
  const abs = Math.abs(Math.round(n));
  return (neg ? '-$' : '$') + abs.toLocaleString();
}

// ── Upload ──────────────────────────────────────────────────────────────────

async function uploadFile() {
  const fileInput = document.getElementById('fileInput');
  const status = document.getElementById('uploadStatus');
  if (!fileInput.files.length) { status.textContent = 'Select a file first'; return; }

  status.textContent = 'Uploading...';
  const form = new FormData();
  form.append('file', fileInput.files[0]);

  try {
    const resp = await fetch('/api/expense-dist/upload', { method: 'POST', body: form });
    const result = await resp.json();
    if (resp.ok) {
      status.textContent = `Parsed ${result.invoices_parsed} invoices across ${result.gl_codes_found} GL codes`;
      loadReport();
    } else {
      status.textContent = 'Error: ' + (result.error || 'Upload failed');
    }
  } catch(e) {
    status.textContent = 'Upload error: ' + e.message;
  }
}

// ── Load Report Data ────────────────────────────────────────────────────────

async function loadReport() {
  try {
    const resp = await fetch('/api/expense-dist/' + ENTITY);
    if (!resp.ok) return;
    reportData = await resp.json();
    renderSummaryCards();
    renderTable();
    expUpdateZeroToggle();
  } catch(e) {
    console.error('Failed to load report:', e);
  }
}

function renderSummaryCards() {
  if (!reportData || !reportData.report) return;
  const r = reportData.report;
  document.getElementById('summaryCards').style.display = 'grid';
  document.getElementById('cardPeriod').textContent = (r.period_from || '?') + ' – ' + (r.period_to || '?');
  document.getElementById('cardInvoices').textContent = r.invoice_count;
  document.getElementById('cardGLs').textContent = reportData.gl_groups.length;
  document.getElementById('cardTotal').textContent = fmt(r.total_amount);
}

// ── Zero-row toggle ──────────────────────────────────────────────────────────

let _expShowZeroRows = false;

function expCountZeroRows() {
  return document.querySelectorAll('#expenseBody .zero-row').length;
}

function expUpdateZeroToggle() {
  const btn = document.getElementById('expZeroToggle');
  if (!btn) return;
  const count = expCountZeroRows();
  if (count === 0) { btn.style.display = 'none'; return; }
  btn.style.display = '';
  btn.textContent = _expShowZeroRows ? 'Hide ' + count + ' Zero Rows' : 'Show ' + count + ' Hidden Zero Rows';
  btn.style.background = _expShowZeroRows ? '#ede9e1' : '#f5efe7';
  btn.style.color = _expShowZeroRows ? '#8a7e72' : '#5a4a3f';
  btn.style.borderColor = _expShowZeroRows ? '#d5cfc5' : '#5a4a3f';
}

function expToggleZeroRows() {
  _expShowZeroRows = !_expShowZeroRows;
  document.querySelectorAll('#expenseBody .zero-row').forEach(row => {
    row.style.display = _expShowZeroRows ? '' : 'none';
  });
  expUpdateZeroToggle();
}

// ── Render Table ────────────────────────────────────────────────────────────

function renderTable() {
  const tbody = document.getElementById('expenseBody');
  tbody.innerHTML = '';
  if (!reportData || !reportData.gl_groups.length) {
    tbody.innerHTML = '<tr><td colspan="9" class="no-data">No data</td></tr>';
    return;
  }

  // Calculate adjusted totals per GL
  const adjustments = {};
  reportData.gl_groups.forEach(g => {
    adjustments[g.gl_code] = { original: 0, reclass_out: 0, reclass_in: 0 };
    g.invoices.forEach(inv => {
      adjustments[g.gl_code].original += inv.amount;
      if (inv.reclass_to_gl) {
        adjustments[g.gl_code].reclass_out += inv.amount;
        if (!adjustments[inv.reclass_to_gl]) {
          adjustments[inv.reclass_to_gl] = { original: 0, reclass_out: 0, reclass_in: 0 };
        }
        adjustments[inv.reclass_to_gl].reclass_in += inv.amount;
      }
    });
  });

  reportData.gl_groups.forEach(group => {
    const gl = group.gl_code;
    const adj = adjustments[gl] || { original: 0, reclass_out: 0, reclass_in: 0 };
    const reclassAdj = adj.reclass_in - adj.reclass_out;
    const adjustedYtd = adj.original + reclassAdj;
    const budgetData = BUDGET_LINES[gl] || {};
    const currentBudget = budgetData.current_budget || 0;
    const overUnder = currentBudget ? adjustedYtd - currentBudget : 0;
    const pctVariance = currentBudget ? (overUnder / currentBudget * 100) : 0;

    // GL summary row
    const isZero = !adj.original && !adj.reclass_in && !adj.reclass_out && !currentBudget;
    const tr = document.createElement('tr');
    tr.className = 'gl-row' + (expandedGLs.has(gl) ? ' expanded' : '') + (isZero ? ' zero-row' : '');
    tr.dataset.gl = gl;
    if (isZero && !_expShowZeroRows) tr.style.display = 'none';
    tr.onclick = () => toggleGL(gl);
    tr.innerHTML = `
      <td><strong>${gl}</strong></td>
      <td>${group.gl_name}</td>
      <td class="number">${fmt(adj.original)}</td>
      <td class="number" style="color:${reclassAdj !== 0 ? 'var(--orange)' : 'inherit'}">${reclassAdj !== 0 ? fmt(reclassAdj) : '—'}</td>
      <td class="number"><strong>${fmt(adjustedYtd)}</strong></td>
      <td class="number">${currentBudget ? fmt(currentBudget) : '—'}</td>
      <td class="number ${overUnder > 0 ? 'over' : 'under'}">${currentBudget ? fmt(overUnder) : '—'}</td>
      <td class="number ${pctVariance > 0 ? 'over' : 'under'}">${currentBudget ? pctVariance.toFixed(1) + '%' : '—'}</td>
      <td>${group.invoices.length} invoices</td>
    `;
    tbody.appendChild(tr);

    // Invoice detail rows (if expanded)
    if (expandedGLs.has(gl)) {
      group.invoices.forEach(inv => {
        const itr = document.createElement('tr');
        itr.className = 'invoice-row' + (inv.reclass_to_gl ? ' reclassed' : '');
        itr.innerHTML = `
          <td></td>
          <td>${inv.payee_name}</td>
          <td class="number">${fmt(inv.amount)}</td>
          <td colspan="2" style="font-size:12px;">
            ${inv.invoice_num ? 'Inv# ' + inv.invoice_num : ''}
            ${inv.invoice_date ? ' — ' + inv.invoice_date.substring(0,10) : ''}
          </td>
          <td style="font-size:12px;">${inv.check_num ? 'Check# ' + inv.check_num : ''}</td>
          <td colspan="2" style="font-size:12px;">${inv.notes || ''}</td>
          <td>
            ${inv.reclass_to_gl
              ? '<span style="font-size:11px;color:var(--orange);">→ ' + inv.reclass_to_gl + '</span> <button class="btn btn-undo" onclick="event.stopPropagation();undoReclass(' + inv.id + ')">Undo</button>'
              : '<button class="btn btn-reclass" onclick="event.stopPropagation();openReclassModal(' + inv.id + ',\\''+gl+'\\',\\''+inv.payee_name+'\\','+inv.amount+')">Reclass</button>'
            }
          </td>
        `;
        tbody.appendChild(itr);
      });
    }
  });
}

function toggleGL(gl) {
  if (expandedGLs.has(gl)) expandedGLs.delete(gl);
  else expandedGLs.add(gl);
  renderTable();
  expUpdateZeroToggle();
}

// ── Reclass Modal ───────────────────────────────────────────────────────────

function openReclassModal(invoiceId, fromGL, payeeName, amount) {
  reclassInvoiceId = invoiceId;
  document.getElementById('reclassInvoiceInfo').textContent =
    `${payeeName} — ${fmt(amount)} (from ${fromGL})`;

  // Populate GL dropdown
  const select = document.getElementById('reclassTarget');
  select.innerHTML = '<option value="">Select target GL...</option>';
  Object.entries(ALL_GL_CODES).sort().forEach(([code, name]) => {
    if (code !== fromGL) {
      select.innerHTML += `<option value="${code}">${code} — ${name}</option>`;
    }
  });

  document.getElementById('reclassNotes').value = '';
  document.getElementById('reclassModal').classList.add('active');
}

function closeReclassModal() {
  document.getElementById('reclassModal').classList.remove('active');
  reclassInvoiceId = null;
}

async function submitReclass() {
  const target = document.getElementById('reclassTarget').value;
  const notes = document.getElementById('reclassNotes').value;
  if (!target) { alert('Select a target GL code'); return; }

  try {
    const resp = await fetch('/api/expense-dist/reclass/' + reclassInvoiceId, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ reclass_to_gl: target, reclass_notes: notes })
    });
    if (resp.ok) {
      closeReclassModal();
      loadReport();
    } else {
      const err = await resp.json();
      alert(err.error || 'Reclass failed');
    }
  } catch(e) {
    alert('Error: ' + e.message);
  }
}

async function undoReclass(invoiceId) {
  try {
    const resp = await fetch('/api/expense-dist/reclass/' + invoiceId, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({ reclass_to_gl: '' })
    });
    if (resp.ok) loadReport();
  } catch(e) {
    alert('Error: ' + e.message);
  }
}

// ── Init ────────────────────────────────────────────────────────────────────

loadReport();
</script>
</body>
</html>
"""
