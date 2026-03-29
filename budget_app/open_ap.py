"""
Open AP (Aging) Report Parser & Unpaid Bills Integration for Century Management.

Parses Yardi "Payable Analytics — Aging" .xls/.xlsx exports and provides:
- Invoice-level drill-down per GL code for unpaid bills
- Auto-population of BudgetLine.unpaid_bills from summed AP data
- GET endpoint for FA/PM drill-down on unpaid_bills column
"""

from flask import Blueprint, request, jsonify
from datetime import datetime
import logging
import re

logger = logging.getLogger(__name__)


def _open_workbook(file_path):
    """Open an Excel file with openpyxl, handling .xls extension that's actually .xlsx."""
    import openpyxl
    import shutil, tempfile, os
    try:
        return openpyxl.load_workbook(str(file_path), data_only=True)
    except Exception:
        # Yardi names .xlsx files as .xls — try copying with .xlsx extension
        if str(file_path).lower().endswith('.xls'):
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            tmp.close()
            shutil.copy2(str(file_path), tmp.name)
            try:
                wb = openpyxl.load_workbook(tmp.name, data_only=True)
                wb._tmp_path = tmp.name  # stash for cleanup
                return wb
            except Exception:
                os.unlink(tmp.name)
                raise
        raise


def detect_open_ap_file(file_path):
    """
    Check if a file is an Open AP (Aging) report from Yardi.

    Yardi AP Analytics exports as .xls that's actually .xlsx format.
    Detection: scan first ~10 rows for telltale header patterns.
    Handles multi-row headers where "Payee" and "Code" are split.

    Returns True if this looks like an AP Aging report.
    """
    # Try openpyxl first (real .xlsx, or .xls that's actually .xlsx)
    try:
        wb = _open_workbook(file_path)
        ws = wb.active
        for row_num in range(1, min(15, ws.max_row + 1)):
            vals = [str(ws.cell(row=row_num, column=c).value or "").strip().lower()
                    for c in range(1, min(20, ws.max_column + 1))]
            row_text = " ".join(vals)
            # Look for AP aging indicators
            if "payee code" in row_text and "current owed" in row_text:
                wb.close()
                return True
            if "payable analytics" in row_text or "ap aging" in row_text:
                wb.close()
                return True
            if "payables aging" in row_text:
                wb.close()
                return True
        # Also check for split headers: Row N has "Payee" + "Current", Row N+1 has "Code" + "Owed"
        for row_num in range(1, min(10, ws.max_row + 1)):
            row_vals = [str(ws.cell(row=row_num, column=c).value or "").strip().lower()
                        for c in range(1, min(20, ws.max_column + 1))]
            if "payee" in row_vals and "current" in row_vals:
                # Check next row for "code" and "owed"
                if row_num + 1 <= ws.max_row:
                    next_vals = [str(ws.cell(row=row_num + 1, column=c).value or "").strip().lower()
                                 for c in range(1, min(20, ws.max_column + 1))]
                    if "code" in next_vals and "owed" in next_vals:
                        wb.close()
                        return True
        wb.close()
        return False
    except Exception:
        pass

    # Try pandas HTML reader (Yardi .xls is sometimes HTML)
    try:
        import pandas as pd
        dfs = pd.read_html(str(file_path))
        if dfs:
            for df in dfs[:3]:
                cols = " ".join(str(c).lower() for c in df.columns)
                if "payee" in cols and ("current owed" in cols or "owed" in cols):
                    return True
                for _, row in df.head(10).iterrows():
                    row_text = " ".join(str(v).lower() for v in row.values if v is not None)
                    if "payee code" in row_text and "current owed" in row_text:
                        return True
    except Exception:
        pass

    # Try by filename pattern
    fname = str(file_path).lower()
    if "aging" in fname or "openap" in fname or "open_ap" in fname or "apanalytics" in fname:
        return True

    return False


def parse_open_ap_report(file_path):
    """
    Parse a Yardi AP Aging .xls/.xlsx export.

    Report structure (from AP Analytics — Aging with Detail + Grid):
      Grouped by Payee (vendor), each row has:
        Payee Code | Payee Name | Invoice Notes | Control | Batch Id |
        Property | Invoice Date | Account (GL code + description) |
        Invoice # | Current Owed | 0-30 Owed | 31-60 Owed |
        61-90 Owed | Over 90 Owed | Future Invoice | Notes

      Vendor subtotals: "Total {payee_code}" rows
      Report total: "Grand Total" row

    Returns:
        tuple: (entity_code, invoices_list)
        Each invoice dict has: gl_code, gl_name, payee_code, payee_name,
        invoice_notes, control, batch_id, property_code, invoice_date,
        invoice_num, current_owed, aging_0_30, aging_31_60, aging_61_90,
        aging_over_90, future_invoice, notes
    """
    invoices = []
    entity_code = None

    # Try openpyxl first
    parsed = _parse_xlsx(file_path)
    if parsed is None:
        # Fall back to HTML/pandas parsing
        parsed = _parse_html_xls(file_path)

    if parsed:
        entity_code, invoices = parsed

    return (entity_code, invoices)


def _parse_xlsx(file_path):
    """Parse as a real .xlsx file via openpyxl.

    Handles Yardi's multi-row header format where column headers are split:
      Row 5: Payee | Payee Name | Invoice | ... | Current | 0-30  | ...
      Row 6: Code  |            | Notes   | ... | Owed    | Owed  | ...
      Row 7:       |            |         | ... |         |       | ... | Owed
    We merge all header rows vertically (joining non-empty values with space).
    """
    try:
        wb = _open_workbook(file_path)
        ws = wb.active

        # Extract entity code from early rows (Row 2 typically has entity code)
        entity_code = None
        for row_num in range(1, min(6, ws.max_row + 1)):
            val = ws.cell(row=row_num, column=1).value
            if val is not None:
                s = str(val).strip()
                # Entity code is a short numeric string (e.g., "204")
                if s.isdigit() and len(s) <= 5:
                    entity_code = s
                    break

        # Find header start row: look for row where column 1 has "Payee"
        header_start = None
        for row_num in range(1, min(20, ws.max_row + 1)):
            val = str(ws.cell(row=row_num, column=1).value or "").strip().lower()
            if val == "payee":
                header_start = row_num
                break
            # Also check for single-row "Payee Code" header
            if val in ("payee code", "payeecode"):
                header_start = row_num
                break

        if not header_start:
            wb.close()
            return None

        # Merge up to 3 header rows vertically to handle split headers
        # E.g., Row 5: "Payee" + Row 6: "Code" → "payee code"
        merged_headers = {}
        header_end = header_start
        for col_num in range(1, ws.max_column + 1):
            parts = []
            for r_offset in range(3):  # up to 3 header rows
                rn = header_start + r_offset
                if rn > ws.max_row:
                    break
                v = str(ws.cell(row=rn, column=col_num).value or "").strip()
                if v:
                    parts.append(v)
                    header_end = max(header_end, rn)
            combined = " ".join(parts).lower().strip()
            if combined:
                merged_headers[combined] = col_num

        logger.info(f"Open AP headers (merged): {merged_headers}")

        # Map expected columns
        col_map = _build_column_map(merged_headers)
        if not col_map:
            wb.close()
            return None

        # Data starts after the last header row
        data_start = header_end + 1

        # Parse data rows
        invoices = []
        current_payee_code = None
        current_payee_name = None

        for row_num in range(data_start, ws.max_row + 1):
            def cell(col_key):
                idx = col_map.get(col_key)
                if idx is None:
                    return None
                return ws.cell(row=row_num, column=idx).value

            payee_code = cell("payee_code")
            payee_name = cell("payee_name")

            # Payee header row (has payee_code + payee_name, no invoice data)
            if payee_code and str(payee_code).strip():
                pc = str(payee_code).strip()
                # Skip Total/Grand Total rows
                if pc.lower().startswith("total") or pc.lower().startswith("grand total"):
                    current_payee_code = None
                    current_payee_name = None
                    continue
                current_payee_code = pc
                current_payee_name = str(payee_name or "").strip()
                continue

            # Check for Total rows that might be in a different column
            a_val = str(ws.cell(row=row_num, column=1).value or "").strip().lower()
            if a_val.startswith("total") or a_val.startswith("grand total"):
                continue

            # Invoice detail row — needs at least Current Owed and Account
            account_raw = cell("account")
            current_owed = cell("current_owed")

            if account_raw is None and current_owed is None:
                continue

            # Parse the Account field: "5630-0000 Plumbing Repairs"
            account_str = str(account_raw or "").strip()
            gl_code, gl_name = _parse_account_field(account_str)

            if not gl_code:
                continue

            # Extract entity from Property column (first time)
            prop = cell("property")
            if prop and not entity_code:
                entity_code = str(prop).strip()

            # Parse amounts
            amt_current = _safe_float(current_owed)
            amt_0_30 = _safe_float(cell("0-30 owed"))
            amt_31_60 = _safe_float(cell("31-60 owed"))
            amt_61_90 = _safe_float(cell("61-90 owed"))
            amt_over_90 = _safe_float(cell("over 90 owed"))
            amt_future = _safe_float(cell("future invoice"))

            # Skip rows where all amounts are zero (usually empty rows)
            if amt_current == 0 and amt_future == 0:
                continue

            invoice_date = cell("invoice date")
            inv_date_str = ""
            if invoice_date:
                if hasattr(invoice_date, 'isoformat'):
                    inv_date_str = invoice_date.isoformat()[:10]
                else:
                    inv_date_str = str(invoice_date).strip()

            invoices.append({
                "gl_code": gl_code,
                "gl_name": gl_name,
                "payee_code": current_payee_code or "",
                "payee_name": current_payee_name or "",
                "invoice_notes": str(cell("invoice notes") or "").strip(),
                "control": str(cell("control") or "").strip(),
                "batch_id": str(cell("batch id") or "").strip(),
                "property_code": str(cell("property") or "").strip(),
                "invoice_date": inv_date_str,
                "invoice_num": str(cell("invoice #") or cell("invoice_num") or "").strip(),
                "current_owed": amt_current,
                "aging_0_30": amt_0_30,
                "aging_31_60": amt_31_60,
                "aging_61_90": amt_61_90,
                "aging_over_90": amt_over_90,
                "future_invoice": amt_future,
                "notes": str(cell("notes") or "").strip(),
            })

        wb.close()
        return (entity_code, invoices) if invoices else None

    except Exception as e:
        logger.warning(f"openpyxl parse failed for open AP: {e}")
        return None


def _parse_html_xls(file_path):
    """Parse as HTML-table .xls (common Yardi export format)."""
    try:
        import pandas as pd

        dfs = pd.read_html(str(file_path))
        if not dfs:
            return None

        # Find the table with AP data (has "Payee Code" and "Current Owed" columns)
        target_df = None
        for df in dfs:
            cols_lower = [str(c).lower().strip() for c in df.columns]
            if any("payee" in c for c in cols_lower) and any("owed" in c for c in cols_lower):
                target_df = df
                break

        if target_df is None:
            # Try finding headers in data rows (Yardi sometimes has merged header cells)
            for df in dfs:
                for idx, row in df.head(10).iterrows():
                    row_vals = [str(v).lower().strip() for v in row.values if v is not None]
                    if any("payee code" in v for v in row_vals) and any("current owed" in v for v in row_vals):
                        # Re-index with this row as headers
                        new_headers = [str(v).strip() for v in row.values]
                        target_df = df.iloc[idx + 1:].copy()
                        target_df.columns = new_headers
                        break
                if target_df is not None:
                    break

        if target_df is None:
            return None

        # Build column map from DataFrame columns
        headers = {str(c).lower().strip(): i for i, c in enumerate(target_df.columns)}
        col_map = _build_column_map(headers)
        if not col_map:
            return None

        invoices = []
        entity_code = None
        current_payee_code = None
        current_payee_name = None

        for _, row in target_df.iterrows():
            def cell(col_key):
                idx = col_map.get(col_key)
                if idx is None:
                    return None
                return row.iloc[idx] if idx < len(row) else None

            payee_code = cell("payee_code")
            payee_name = cell("payee_name")

            if payee_code and str(payee_code).strip() and str(payee_code).strip().lower() not in ("nan",):
                pc = str(payee_code).strip()
                if pc.lower().startswith("total") or pc.lower().startswith("grand total"):
                    current_payee_code = None
                    current_payee_name = None
                    continue
                current_payee_code = pc
                current_payee_name = str(payee_name or "").strip()
                if current_payee_name.lower() == "nan":
                    current_payee_name = ""
                continue

            account_raw = cell("account")
            current_owed = cell("current_owed")

            if account_raw is None or str(account_raw).strip().lower() in ("nan", "none", ""):
                continue

            account_str = str(account_raw).strip()
            gl_code, gl_name = _parse_account_field(account_str)
            if not gl_code:
                continue

            prop = cell("property")
            if prop and not entity_code:
                p = str(prop).strip()
                if p.lower() not in ("nan", "none", ""):
                    entity_code = p

            amt_current = _safe_float(current_owed)
            amt_future = _safe_float(cell("future invoice"))

            if amt_current == 0 and amt_future == 0:
                continue

            inv_date = str(cell("invoice date") or "").strip()
            if inv_date.lower() == "nan":
                inv_date = ""

            invoices.append({
                "gl_code": gl_code,
                "gl_name": gl_name,
                "payee_code": current_payee_code or "",
                "payee_name": current_payee_name or "",
                "invoice_notes": _clean_nan(cell("invoice notes")),
                "control": _clean_nan(cell("control")),
                "batch_id": _clean_nan(cell("batch id")),
                "property_code": _clean_nan(cell("property")),
                "invoice_date": inv_date,
                "invoice_num": _clean_nan(cell("invoice #") or cell("invoice_num")),
                "current_owed": amt_current,
                "aging_0_30": _safe_float(cell("0-30 owed")),
                "aging_31_60": _safe_float(cell("31-60 owed")),
                "aging_61_90": _safe_float(cell("61-90 owed")),
                "aging_over_90": _safe_float(cell("over 90 owed")),
                "future_invoice": amt_future,
                "notes": _clean_nan(cell("notes")),
            })

        return (entity_code, invoices) if invoices else None

    except Exception as e:
        logger.warning(f"HTML/pandas parse failed for open AP: {e}")
        return None


# ─── Helpers ────────────────────────────────────────────────────────────────

def _build_column_map(headers):
    """
    Map standard column names to actual column indices.
    headers: dict of {lowercase_name: column_index}
    """
    mapping = {}
    aliases = {
        "payee_code": ["payee code", "payeecode", "payee\ncode"],
        "payee_name": ["payee name", "payeename", "payee\nname"],
        "invoice notes": ["invoice notes", "invoice\nnotes", "invoicenotes", "description"],
        "control": ["control"],
        "batch id": ["batch id", "batchid", "batch\nid"],
        "property": ["property", "prop"],
        "invoice date": ["invoice date", "invoicedate", "invoice\ndate", "inv date"],
        "account": ["account"],
        "invoice #": ["invoice #", "invoice#", "invoice no", "invoice\n#", "invoiceno"],
        "current_owed": ["current owed", "currentowed", "current\nowed", "amount owed"],
        "0-30 owed": ["0-30 owed", "0-30\nowed", "0-30owed"],
        "31-60 owed": ["31-60 owed", "31-60\nowed", "31-60owed"],
        "61-90 owed": ["61-90 owed", "61-90\nowed", "61-90owed"],
        "over 90 owed": ["over 90 owed", "over\n90\nowed", "over90owed", "over 90\nowed", "over 90 owed"],
        "future invoice": ["future invoice", "futureinvoice", "future\ninvoice"],
        "notes": ["notes"],
    }

    for key, possible_names in aliases.items():
        for name in possible_names:
            if name in headers:
                mapping[key] = headers[name]
                break

    # Must have at least account and current_owed
    if "account" not in mapping and "current_owed" not in mapping:
        return None

    return mapping


def _parse_account_field(account_str):
    """
    Parse "5630-0000 Plumbing Repairs" into ("5630-0000", "Plumbing Repairs").
    """
    if not account_str:
        return (None, None)
    m = re.match(r'^(\d{4}-\d{4})\s+(.*)', account_str)
    if m:
        return (m.group(1), m.group(2).strip())
    # Try without space
    m = re.match(r'^(\d{4}-\d{4})(.*)', account_str)
    if m:
        return (m.group(1), m.group(2).strip())
    return (None, None)


def _safe_float(val):
    """Safely convert to float, handling commas, None, NaN."""
    if val is None:
        return 0.0
    s = str(val).strip().replace(",", "").replace("$", "")
    if s.lower() in ("", "nan", "none", "-"):
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _clean_nan(val):
    """Clean NaN/None to empty string."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ("nan", "none"):
        return ""
    return s


# ─── Blueprint factory ──────────────────────────────────────────────────────

def create_open_ap_blueprint(db, workflow_models):
    """
    Create the Open AP blueprint with models and helpers.

    Args:
        db: SQLAlchemy instance
        workflow_models: dict with Budget, BudgetLine, BudgetRevision

    Returns:
        tuple: (blueprint, models_dict, helpers_dict)
    """
    bp = Blueprint("open_ap", __name__)

    Budget = workflow_models["Budget"]
    BudgetLine = workflow_models["BudgetLine"]
    BudgetRevision = workflow_models["BudgetRevision"]

    # ─── Models ────────────────────────────────────────────────────────────

    class OpenAPReport(db.Model):
        """Uploaded Open AP (Aging) report metadata."""
        __tablename__ = "open_ap_reports"

        id = db.Column(db.Integer, primary_key=True)
        entity_code = db.Column(db.String(50), nullable=False, index=True)
        report_date = db.Column(db.String(30))  # When the report was run
        file_name = db.Column(db.String(255))
        total_amount = db.Column(db.Float, default=0.0)
        invoice_count = db.Column(db.Integer, default=0)
        uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

        invoices = db.relationship("OpenAPInvoice", back_populates="report",
                                   cascade="all, delete-orphan")

        def to_dict(self):
            return {
                "id": self.id,
                "entity_code": self.entity_code,
                "report_date": self.report_date,
                "file_name": self.file_name,
                "total_amount": float(self.total_amount or 0),
                "invoice_count": self.invoice_count,
                "uploaded_at": self.uploaded_at.isoformat() if self.uploaded_at else None,
            }

    class OpenAPInvoice(db.Model):
        """Individual unpaid invoice from AP Aging report."""
        __tablename__ = "open_ap_invoices"

        id = db.Column(db.Integer, primary_key=True)
        report_id = db.Column(db.Integer, db.ForeignKey("open_ap_reports.id"), nullable=False)
        gl_code = db.Column(db.String(50), nullable=False, index=True)
        gl_name = db.Column(db.String(255))
        payee_code = db.Column(db.String(100))
        payee_name = db.Column(db.String(255))
        invoice_notes = db.Column(db.Text, default="")
        control = db.Column(db.String(50))
        batch_id = db.Column(db.String(50))
        property_code = db.Column(db.String(50))
        invoice_date = db.Column(db.String(30))
        invoice_num = db.Column(db.String(100))
        current_owed = db.Column(db.Float, default=0.0)
        aging_0_30 = db.Column(db.Float, default=0.0)
        aging_31_60 = db.Column(db.Float, default=0.0)
        aging_61_90 = db.Column(db.Float, default=0.0)
        aging_over_90 = db.Column(db.Float, default=0.0)
        future_invoice = db.Column(db.Float, default=0.0)
        notes = db.Column(db.Text, default="")

        report = db.relationship("OpenAPReport", back_populates="invoices")

        def to_dict(self):
            return {
                "id": self.id,
                "gl_code": self.gl_code,
                "gl_name": self.gl_name,
                "payee_code": self.payee_code,
                "payee_name": self.payee_name,
                "invoice_notes": self.invoice_notes,
                "control": self.control,
                "invoice_date": self.invoice_date,
                "invoice_num": self.invoice_num,
                "current_owed": float(self.current_owed or 0),
                "aging_0_30": float(self.aging_0_30 or 0),
                "aging_31_60": float(self.aging_31_60 or 0),
                "aging_61_90": float(self.aging_61_90 or 0),
                "aging_over_90": float(self.aging_over_90 or 0),
                "future_invoice": float(self.future_invoice or 0),
                "notes": self.notes,
            }

    # ─── Helpers ──────────────────────────────────────────────────────────

    def store_open_ap_report(entity_code, invoices, file_name="", report_date=None):
        """Store parsed Open AP data. Replaces existing report for same entity."""
        # Remove existing report for same entity
        existing = OpenAPReport.query.filter_by(entity_code=entity_code).first()
        if existing:
            db.session.delete(existing)
            db.session.flush()

        total = sum(inv["current_owed"] for inv in invoices)
        report = OpenAPReport(
            entity_code=entity_code,
            report_date=report_date or datetime.utcnow().strftime("%m/%d/%Y"),
            file_name=file_name,
            total_amount=total,
            invoice_count=len(invoices),
        )
        db.session.add(report)
        db.session.flush()

        for inv in invoices:
            db.session.add(OpenAPInvoice(
                report_id=report.id,
                gl_code=inv["gl_code"],
                gl_name=inv.get("gl_name", ""),
                payee_code=inv.get("payee_code", ""),
                payee_name=inv.get("payee_name", ""),
                invoice_notes=inv.get("invoice_notes", ""),
                control=inv.get("control", ""),
                batch_id=inv.get("batch_id", ""),
                property_code=inv.get("property_code", ""),
                invoice_date=inv.get("invoice_date", ""),
                invoice_num=inv.get("invoice_num", ""),
                current_owed=inv.get("current_owed", 0),
                aging_0_30=inv.get("aging_0_30", 0),
                aging_31_60=inv.get("aging_31_60", 0),
                aging_61_90=inv.get("aging_61_90", 0),
                aging_over_90=inv.get("aging_over_90", 0),
                future_invoice=inv.get("future_invoice", 0),
                notes=inv.get("notes", ""),
            ))

        db.session.commit()
        logger.info(f"Stored {len(invoices)} open AP invoices for entity {entity_code} (total=${total:,.2f})")
        return report

    def apply_unpaid_bills(entity_code):
        """
        Sum Open AP invoices by GL code and update BudgetLine.unpaid_bills.

        Uses current_owed (not future_invoice) as the unpaid amount.
        Future invoices are bills not yet due and should not be in the budget.

        Returns:
            dict: {"applied": int, "gl_totals": {gl_code: amount}}
        """
        budget = Budget.query.filter_by(entity_code=entity_code).order_by(Budget.year.desc(), Budget.version.desc()).first()
        if not budget:
            logger.warning(f"No budget found for entity {entity_code}, cannot apply unpaid bills")
            return {"applied": 0, "gl_totals": {}}

        report = OpenAPReport.query.filter_by(entity_code=entity_code).first()
        if not report:
            logger.warning(f"No Open AP report found for entity {entity_code}")
            return {"applied": 0, "gl_totals": {}}

        # Sum current_owed by GL code
        gl_totals = {}
        for inv in report.invoices:
            gl = inv.gl_code
            if gl not in gl_totals:
                gl_totals[gl] = 0
            gl_totals[gl] += float(inv.current_owed or 0)

        # Round totals
        gl_totals = {gl: round(amt, 2) for gl, amt in gl_totals.items()}

        # Update BudgetLine.unpaid_bills for matching GL codes
        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
        gl_to_line = {line.gl_code: line for line in lines}

        applied = 0
        for gl, total in gl_totals.items():
            if gl in gl_to_line:
                line = gl_to_line[gl]
                old_val = float(line.unpaid_bills or 0)
                if abs(old_val - total) > 0.01:  # Only update if changed
                    line.unpaid_bills = total

                    db.session.add(BudgetRevision(
                        budget_id=budget.id,
                        action="update",
                        field_name="unpaid_bills",
                        old_value=str(old_val),
                        new_value=str(total),
                        notes=f"GL {gl}: auto-populated from Open AP ({report.invoice_count} total invoices)",
                        source="open_ap_auto"
                    ))
                applied += 1
            else:
                logger.info(f"Open AP GL {gl} ({gl_totals[gl]:,.2f}) has no matching BudgetLine — skipped")

        # Zero out unpaid_bills for GL codes NOT in Open AP (clear stale data)
        for line in lines:
            if line.gl_code not in gl_totals and float(line.unpaid_bills or 0) != 0:
                old_val = float(line.unpaid_bills or 0)
                line.unpaid_bills = 0
                db.session.add(BudgetRevision(
                    budget_id=budget.id,
                    action="update",
                    field_name="unpaid_bills",
                    old_value=str(old_val),
                    new_value="0",
                    notes=f"GL {line.gl_code}: cleared — no open AP invoices",
                    source="open_ap_auto"
                ))

        db.session.commit()
        logger.info(f"Applied unpaid bills to {applied} GL lines for entity {entity_code}")
        return {"applied": applied, "gl_totals": gl_totals}

    # ─── API Routes ───────────────────────────────────────────────────────

    @bp.route("/api/open-ap/<entity_code>", methods=["GET"])
    def get_open_ap(entity_code):
        """Get Open AP invoices grouped by GL code for drill-down."""
        report = OpenAPReport.query.filter_by(entity_code=entity_code).first()
        if not report:
            return jsonify({"report": None, "gl_groups": []}), 200

        # Group invoices by GL code
        gl_groups = {}
        for inv in report.invoices:
            gl = inv.gl_code
            if gl not in gl_groups:
                gl_groups[gl] = {
                    "gl_code": gl,
                    "gl_name": inv.gl_name,
                    "total": 0,
                    "invoices": [],
                }
            gl_groups[gl]["total"] += float(inv.current_owed or 0)
            gl_groups[gl]["invoices"].append(inv.to_dict())

        # Round totals and sort by GL code
        for gl in gl_groups:
            gl_groups[gl]["total"] = round(gl_groups[gl]["total"], 2)

        sorted_groups = sorted(gl_groups.values(), key=lambda g: g["gl_code"])

        return jsonify({
            "report": report.to_dict(),
            "gl_groups": sorted_groups,
        }), 200

    @bp.route("/api/open-ap/<entity_code>/<gl_code>", methods=["GET"])
    def get_open_ap_gl(entity_code, gl_code):
        """Get Open AP invoices for a single GL code."""
        report = OpenAPReport.query.filter_by(entity_code=entity_code).first()
        if not report:
            return jsonify({"gl_code": gl_code, "invoices": [], "total": 0}), 200

        invoices = OpenAPInvoice.query.filter_by(
            report_id=report.id, gl_code=gl_code
        ).all()

        total = sum(float(inv.current_owed or 0) for inv in invoices)

        return jsonify({
            "gl_code": gl_code,
            "invoices": [inv.to_dict() for inv in invoices],
            "total": round(total, 2),
        }), 200

    # ─── Return blueprint, models, helpers ────────────────────────────────

    models = {
        "OpenAPReport": OpenAPReport,
        "OpenAPInvoice": OpenAPInvoice,
    }

    helpers = {
        "store_open_ap_report": store_open_ap_report,
        "apply_unpaid_bills": apply_unpaid_bills,
    }

    return bp, models, helpers
