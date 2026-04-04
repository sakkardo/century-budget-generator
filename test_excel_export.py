"""
Standalone test: pull real budget data from production API for entity 204,
generate test_budget_export.xlsx with Excel formulas for all calculated fields.

No existing files are modified. Creates only:
  - test_budget_export.xlsx
"""

import requests
import json

# Use production API — no DB credentials needed
API_BASE = "https://century-budget-generator-production.up.railway.app"
ENTITY = "204"


def fetch_data():
    """Pull dashboard data for entity 204 from production."""
    resp = requests.get(f"{API_BASE}/api/dashboard/{ENTITY}")
    resp.raise_for_status()
    return resp.json()


def build_workbook(data):
    """Build Excel workbook with formulas for all calculated fields."""
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    lines = data["lines"]
    sheets_map = data.get("sheets", {})
    sheet_order = data.get("sheet_order", list(sheets_map.keys()))
    ytd_months = data.get("ytd_months", 2)
    remaining_months = data.get("remaining_months", 10)
    assumptions = data.get("assumptions", {})

    # Build a lookup: gl_code -> line data
    line_lookup = {l["gl_code"]: l for l in lines}

    for sheet_name in sheet_order:
        sheet_lines = sheets_map.get(sheet_name, [])
        if not sheet_lines:
            continue

        # Sort by row_num
        sheet_lines.sort(key=lambda x: x.get("row_num", 0))

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit

        # --- Header row ---
        headers = [
            "GL Code",           # A
            "Description",       # B
            "Notes",             # C
            "Prior Year",        # D
            "YTD Actual",        # E
            "Accrual Adj",       # F
            "Unpaid Bills",      # G
            "YTD Budget",        # H
            "Mar-Dec Est",       # I  (FORMULA)
            "12 Mo Forecast",    # J  (FORMULA)
            "Curr Budget",       # K
            "Inc %",             # L
            "Proposed Budget",   # M  (FORMULA)
            "$ Variance",        # N  (FORMULA)
            "% Change",          # O  (FORMULA)
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        # --- Data rows ---
        for i, line in enumerate(sheet_lines):
            r = i + 2  # row 2, 3, 4, ...

            # A-H: raw data from DB
            ws.cell(row=r, column=1, value=line["gl_code"])
            ws.cell(row=r, column=2, value=line["description"])
            ws.cell(row=r, column=3, value=line.get("notes", "") or "")
            ws.cell(row=r, column=4, value=line.get("prior_year", 0) or 0)
            ws.cell(row=r, column=5, value=line.get("ytd_actual", 0) or 0)
            ws.cell(row=r, column=6, value=line.get("accrual_adj", 0) or 0)
            ws.cell(row=r, column=7, value=line.get("unpaid_bills", 0) or 0)
            ws.cell(row=r, column=8, value=line.get("ytd_budget", 0) or 0)

            # I: Mar-Dec Estimate (FORMULA)
            # base = E+F+G, prior = D
            # =IFERROR(IF(AND((E+F+G)>=D, D>0), (E+F+G)/ytd_months*remaining, MAX(D-(E+F+G),0)), 0)
            ws.cell(row=r, column=9).value = (
                f'=IFERROR(IF(AND((E{r}+F{r}+G{r})>=D{r},D{r}>0),'
                f'(E{r}+F{r}+G{r})/{ytd_months}*{remaining_months},'
                f'MAX(D{r}-(E{r}+F{r}+G{r}),0)),0)'
            )

            # J: 12 Mo Forecast = YTD + Accrual + Unpaid + Estimate
            # =E+F+G+I
            ws.cell(row=r, column=10).value = f'=E{r}+F{r}+G{r}+I{r}'

            # K: Current Budget (raw)
            ws.cell(row=r, column=11, value=line.get("current_budget", 0) or 0)

            # L: Increase % (raw, stored as decimal e.g. 0.05 = 5%)
            ws.cell(row=r, column=12, value=line.get("increase_pct", 0) or 0)

            # M: Proposed Budget = Forecast * (1 + Inc%)
            # =ROUND(J*(1+L), 0)
            ws.cell(row=r, column=13).value = f'=ROUND(J{r}*(1+L{r}),0)'

            # N: $ Variance = Curr Budget - 12 Mo Forecast
            # =K-J
            ws.cell(row=r, column=14).value = f'=K{r}-J{r}'

            # O: % Change = (Curr Budget - Forecast) / Forecast
            # =IFERROR((K-J)/J, 0)
            ws.cell(row=r, column=15).value = f'=IFERROR((K{r}-J{r})/J{r},0)'

        # --- Sheet Total row ---
        last_data_row = len(sheet_lines) + 1
        total_row = last_data_row + 1
        ws.cell(row=total_row, column=1, value="Sheet Total")

        # SUM formulas for numeric columns D through O
        for col_letter in ["D", "E", "F", "G", "H", "I", "J", "K", "M", "N"]:
            col_idx = ord(col_letter) - ord("A") + 1
            ws.cell(row=total_row, column=col_idx).value = (
                f'=SUM({col_letter}2:{col_letter}{last_data_row})'
            )

        # L (Inc %): average, not sum
        ws.cell(row=total_row, column=12).value = (
            f'=IFERROR(AVERAGE(L2:L{last_data_row}),0)'
        )

        # O (% Change): recalculate from totals, not sum of percentages
        ws.cell(row=total_row, column=15).value = (
            f'=IFERROR((K{total_row}-J{total_row})/J{total_row},0)'
        )

    # --- Summary sheet ---
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.cell(row=1, column=1, value="Sheet")
    ws_sum.cell(row=1, column=2, value="Prior Year")
    ws_sum.cell(row=1, column=3, value="12 Mo Forecast")
    ws_sum.cell(row=1, column=4, value="Curr Budget")
    ws_sum.cell(row=1, column=5, value="Proposed Budget")
    ws_sum.cell(row=1, column=6, value="$ Variance")
    ws_sum.cell(row=1, column=7, value="% Change")

    # Reference each sheet's total row
    for i, sheet_name in enumerate(sheet_order):
        sheet_lines = sheets_map.get(sheet_name, [])
        if not sheet_lines:
            continue
        r = i + 2
        safe_name = sheet_name[:31]
        total_row_in_sheet = len(sheet_lines) + 2

        ws_sum.cell(row=r, column=1, value=sheet_name)
        # Cross-sheet references to the total row
        ws_sum.cell(row=r, column=2).value = f"='{safe_name}'!D{total_row_in_sheet}"
        ws_sum.cell(row=r, column=3).value = f"='{safe_name}'!J{total_row_in_sheet}"
        ws_sum.cell(row=r, column=4).value = f"='{safe_name}'!K{total_row_in_sheet}"
        ws_sum.cell(row=r, column=5).value = f"='{safe_name}'!M{total_row_in_sheet}"
        ws_sum.cell(row=r, column=6).value = f"='{safe_name}'!N{total_row_in_sheet}"
        ws_sum.cell(row=r, column=7).value = f"=IFERROR('{safe_name}'!N{total_row_in_sheet}/'{safe_name}'!J{total_row_in_sheet},0)"

    # Grand total
    active_sheets = [sn for sn in sheet_order if sheets_map.get(sn)]
    gt_row = len(active_sheets) + 2
    ws_sum.cell(row=gt_row, column=1, value="GRAND TOTAL")
    for col in range(2, 7):  # B through F
        col_letter = chr(ord("A") + col - 1)
        ws_sum.cell(row=gt_row, column=col).value = (
            f'=SUM({col_letter}2:{col_letter}{gt_row - 1})'
        )
    # % Change for grand total
    ws_sum.cell(row=gt_row, column=7).value = (
        f'=IFERROR(F{gt_row}/C{gt_row},0)'
    )

    # --- Metadata sheet ---
    ws_meta = wb.create_sheet(title="Metadata")
    ws_meta.cell(row=1, column=1, value="Entity Code")
    ws_meta.cell(row=1, column=2, value=ENTITY)
    ws_meta.cell(row=2, column=1, value="Building")
    ws_meta.cell(row=2, column=2, value=data.get("budget", {}).get("building_name", ""))
    ws_meta.cell(row=3, column=1, value="Year")
    ws_meta.cell(row=3, column=2, value=data.get("budget", {}).get("year", 2027))
    ws_meta.cell(row=4, column=1, value="YTD Months")
    ws_meta.cell(row=4, column=2, value=ytd_months)
    ws_meta.cell(row=5, column=1, value="Remaining Months")
    ws_meta.cell(row=5, column=2, value=remaining_months)
    ws_meta.cell(row=6, column=1, value="Status")
    ws_meta.cell(row=6, column=2, value=data.get("budget", {}).get("status", ""))
    ws_meta.cell(row=7, column=1, value="Total Lines")
    ws_meta.cell(row=7, column=2, value=len(lines))

    return wb


def main():
    print("Fetching data from production API...")
    data = fetch_data()
    budget = data.get("budget", {})
    print(f"  Building: {budget.get('building_name', '?')}")
    print(f"  Entity: {budget.get('entity_code', '?')}")
    print(f"  Lines: {len(data.get('lines', []))}")
    print(f"  Sheets: {data.get('sheet_order', [])}")
    print(f"  YTD Months: {data.get('ytd_months', '?')}")

    print("\nBuilding workbook with Excel formulas...")
    wb = build_workbook(data)

    out_path = "test_budget_export.xlsx"
    wb.save(out_path)
    print(f"\nSaved: {out_path}")

    # Verification: count sheets, rows, formula cells
    for ws in wb.worksheets:
        total_cells = 0
        formula_cells = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    total_cells += 1
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells += 1
        print(f"  Sheet '{ws.title}': {total_cells} cells, {formula_cells} formulas")

    print("\nDone. Open test_budget_export.xlsx in Excel to verify formulas calculate.")


if __name__ == "__main__":
    main()
