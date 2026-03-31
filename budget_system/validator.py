"""
Validation and reporting for budget automation.

Checks for data quality issues and generates reports.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class BudgetValidator:
    """Validates populated budget templates."""

    def __init__(self, template_path: Path):
        """
        Initialize validator.

        Args:
            template_path: Path to populated template.xlsx
        """
        self.template_path = Path(template_path)
        self.wb = load_workbook(self.template_path, data_only=True)
        self.validation_results = {
            "file": str(self.template_path),
            "timestamp": datetime.now().isoformat(),
            "gl_codes_matched": 0,
            "gl_codes_unmatched": 0,
            "zero_value_alerts": [],
            "missing_data_alerts": [],
            "summary_stats": {},
        }

    def validate(self, gl_codes_expected: List[str],
                gl_codes_found: List[str]) -> Dict[str, Any]:
        """
        Run validation checks.

        Args:
            gl_codes_expected: List of GL codes that should be in template
            gl_codes_found: List of GL codes found in YSL data

        Returns:
            Dict with validation results
        """
        # Check which GL codes were matched
        matched = set(gl_codes_found) & set(gl_codes_expected)
        unmatched = set(gl_codes_expected) - set(gl_codes_found)

        self.validation_results["gl_codes_matched"] = len(matched)
        self.validation_results["gl_codes_unmatched"] = len(unmatched)

        # Check for zero-value and missing data issues
        self._check_for_data_issues(matched)

        self.wb.close()
        return self.validation_results

    def _check_for_data_issues(self, matched_gl_codes: set) -> None:
        """
        Check for zero values where prior year had data and other issues.

        Args:
            matched_gl_codes: Set of GL codes that were matched
        """
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            for row_idx in range(1, ws.max_row + 1):
                gl_code_cell = ws[f"A{row_idx}"]
                gl_code = gl_code_cell.value

                if not gl_code or gl_code not in matched_gl_codes:
                    continue

                # Check columns: D (Prior Year), E (YTD Actual), H (YTD Budget)
                prior_year = ws[f"D{row_idx}"].value
                ytd_actual = ws[f"E{row_idx}"].value
                ytd_budget = ws[f"H{row_idx}"].value

                # Alert if prior year had data but current is zero
                if prior_year and isinstance(prior_year, (int, float)) and prior_year != 0:
                    if ytd_actual is None or (isinstance(ytd_actual, (int, float)) and ytd_actual == 0):
                        alert = {
                            "sheet": sheet_name,
                            "gl_code": gl_code,
                            "issue": "YTD Actual is zero but prior year had data",
                            "prior_year": prior_year,
                            "ytd_actual": ytd_actual,
                        }
                        self.validation_results["zero_value_alerts"].append(alert)

                # Check for missing data
                if ytd_actual is None or ytd_budget is None:
                    if prior_year:  # Only alert if there was prior year data
                        alert = {
                            "sheet": sheet_name,
                            "gl_code": gl_code,
                            "issue": "Missing YTD Actual or Budget",
                            "ytd_actual": ytd_actual,
                            "ytd_budget": ytd_budget,
                        }
                        self.validation_results["missing_data_alerts"].append(alert)

        # Generate summary stats
        self._generate_summary_stats()

    def _generate_summary_stats(self) -> None:
        """Generate summary statistics per sheet."""
        stats = {}

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            sheet_stats = {
                "total_rows": ws.max_row,
                "filled_cells": 0,
                "empty_cells": 0,
                "numeric_values": 0,
                "total_value": 0,
            }

            for row_idx in range(1, ws.max_row + 1):
                # Check key columns
                for col in ["D", "E", "H"]:
                    cell = ws[f"{col}{row_idx}"]
                    if cell.value is not None:
                        sheet_stats["filled_cells"] += 1
                        if isinstance(cell.value, (int, float)):
                            sheet_stats["numeric_values"] += 1
                            sheet_stats["total_value"] += cell.value
                    else:
                        sheet_stats["empty_cells"] += 1

            stats[sheet_name] = sheet_stats

        self.validation_results["summary_stats"] = stats

    def generate_report(self, output_path: Path) -> bool:
        """
        Generate text validation report.

        Args:
            output_path: Path to save report

        Returns:
            True if successful
        """
        try:
            with open(output_path, "w") as f:
                f.write("BUDGET VALIDATION REPORT\n")
                f.write("=" * 80 + "\n\n")

                f.write(f"File: {self.validation_results['file']}\n")
                f.write(f"Timestamp: {self.validation_results['timestamp']}\n\n")

                # GL code summary
                f.write("GL CODE SUMMARY\n")
                f.write("-" * 80 + "\n")
                f.write(f"GL Codes Matched: {self.validation_results['gl_codes_matched']}\n")
                f.write(f"GL Codes Unmatched: {self.validation_results['gl_codes_unmatched']}\n\n")

                # Zero value alerts
                if self.validation_results["zero_value_alerts"]:
                    f.write("ZERO VALUE ALERTS\n")
                    f.write("-" * 80 + "\n")
                    f.write("These accounts had data in prior year but are zero now:\n\n")

                    for alert in self.validation_results["zero_value_alerts"]:
                        f.write(f"  GL Code: {alert['gl_code']} ({alert['sheet']})\n")
                        f.write(f"    Prior Year: {alert['prior_year']}\n")
                        f.write(f"    YTD Actual: {alert['ytd_actual']}\n")
                        f.write(f"    Issue: {alert['issue']}\n\n")

                # Missing data alerts
                if self.validation_results["missing_data_alerts"]:
                    f.write("\nMISSING DATA ALERTS\n")
                    f.write("-" * 80 + "\n")
                    f.write("These accounts have missing YTD data:\n\n")

                    for alert in self.validation_results["missing_data_alerts"]:
                        f.write(f"  GL Code: {alert['gl_code']} ({alert['sheet']})\n")
                        f.write(f"    YTD Actual: {alert['ytd_actual']}\n")
                        f.write(f"    YTD Budget: {alert['ytd_budget']}\n")
                        f.write(f"    Issue: {alert['issue']}\n\n")

                # Summary stats
                if self.validation_results["summary_stats"]:
                    f.write("\nSUMMARY STATISTICS BY SHEET\n")
                    f.write("-" * 80 + "\n\n")

                    for sheet, stats in self.validation_results["summary_stats"].items():
                        f.write(f"{sheet}\n")
                        f.write(f"  Total Rows: {stats['total_rows']}\n")
                        f.write(f"  Filled Cells: {stats['filled_cells']}\n")
                        f.write(f"  Empty Cells: {stats['empty_cells']}\n")
                        f.write(f"  Numeric Values: {stats['numeric_values']}\n")
                        f.write(f"  Total Value: ${stats['total_value']:,.2f}\n\n")

                logger.info(f"Saved validation report to {output_path}")
                return True

        except Exception as e:
            logger.error(f"Error generating report: {e}")
            return False

    def generate_html_report(self, output_path: Path) -> bool:
        """
        Generate HTML validation report.

        Args:
            output_path: Path to save HTML report

        Returns:
            True if successful
        """
        try:
            html = self._build_html()

            with open(output_path, "w") as f:
                f.write(html)

            logger.info(f"Saved HTML report to {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error generating HTML report: {e}")
            return False

    def _build_html(self) -> str:
        """
        Build HTML report content.

        Returns:
            HTML string
        """
        html = """<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Budget Validation Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1 { color: #333; }
        h2 { color: #666; border-bottom: 2px solid #ccc; padding-bottom: 5px; }
        table { border-collapse: collapse; width: 100%; margin: 20px 0; }
        th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }
        th { background-color: #4CAF50; color: white; }
        tr:nth-child(even) { background-color: #f9f9f9; }
        .alert { background-color: #fff3cd; padding: 10px; margin: 10px 0; }
        .error { background-color: #f8d7da; color: #721c24; }
        .success { background-color: #d4edda; color: #155724; }
        .summary { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
        .stat-box { background: #f0f0f0; padding: 15px; border-radius: 5px; }
    </style>
</head>
<body>
    <h1>Budget Validation Report</h1>
    <p><strong>File:</strong> {file}</p>
    <p><strong>Generated:</strong> {timestamp}</p>

    <h2>GL Code Summary</h2>
    <div class="summary">
        <div class="stat-box success">
            <p><strong>Matched GL Codes</strong></p>
            <p style="font-size: 24px; margin: 10px 0;">{gl_matched}</p>
        </div>
        <div class="stat-box error">
            <p><strong>Unmatched GL Codes</strong></p>
            <p style="font-size: 24px; margin: 10px 0;">{gl_unmatched}</p>
        </div>
    </div>

    {alerts_section}
    {stats_section}
</body>
</html>
"""

        # Build alerts section
        alerts_html = ""
        if self.validation_results["zero_value_alerts"]:
            alerts_html += "<h2>Zero Value Alerts</h2>\n"
            alerts_html += "<p>Accounts with prior year data but zero YTD actual:</p>\n"
            for alert in self.validation_results["zero_value_alerts"]:
                alerts_html += f"""<div class="alert error">
                    <strong>{alert['gl_code']}</strong> ({alert['sheet']})<br>
                    Prior Year: ${alert['prior_year']:,.2f} | YTD: {alert['ytd_actual']}
                </div>\n"""

        if self.validation_results["missing_data_alerts"]:
            alerts_html += "<h2>Missing Data Alerts</h2>\n"
            for alert in self.validation_results["missing_data_alerts"][:10]:  # Limit to 10
                alerts_html += f"""<div class="alert">
                    <strong>{alert['gl_code']}</strong> ({alert['sheet']})<br>
                    {alert['issue']}
                </div>\n"""

        # Build stats section
        stats_html = ""
        if self.validation_results["summary_stats"]:
            stats_html += "<h2>Summary Statistics by Sheet</h2>\n"
            stats_html += "<table><tr><th>Sheet</th><th>Rows</th><th>Filled</th>"
            stats_html += "<th>Empty</th><th>Numeric</th><th>Total Value</th></tr>\n"

            for sheet, stats in self.validation_results["summary_stats"].items():
                stats_html += f"""<tr>
                    <td>{sheet}</td>
                    <td>{stats['total_rows']}</td>
                    <td>{stats['filled_cells']}</td>
                    <td>{stats['empty_cells']}</td>
                    <td>{stats['numeric_values']}</td>
                    <td>${stats['total_value']:,.2f}</td>
                </tr>\n"""

            stats_html += "</table>\n"

        return html.format(
            file=self.validation_results["file"],
            timestamp=self.validation_results["timestamp"],
            gl_matched=self.validation_results["gl_codes_matched"],
            gl_unmatched=self.validation_results["gl_codes_unmatched"],
            alerts_section=alerts_html,
            stats_section=stats_html,
        )
