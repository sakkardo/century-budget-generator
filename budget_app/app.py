"""
Century Management Unified Budget & Assumptions System
Run: python app.py  |  Open: http://localhost:5000
"""
import os
import sys
import csv
import json
import shutil
import logging
import tempfile
import zipfile
from pathlib import Path
from io import BytesIO
from flask import Flask, render_template_string, request, jsonify, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from functools import wraps
from werkzeug.exceptions import InternalServerError

# Detect cloud deployment (Railway sets PORT env var)
IS_CLOUD = "PORT" in os.environ or "RAILWAY_ENVIRONMENT" in os.environ

# Add budget_system to path for pipeline imports
BUDGET_SYSTEM = Path(__file__).parent.parent / "budget_system"
sys.path.insert(0, str(BUDGET_SYSTEM))

# Add budget_app to path so workflow.py can import sibling modules (e.g. dof_taxes)
BUDGET_APP = Path(__file__).parent
sys.path.insert(0, str(BUDGET_APP))

from ysl_parser import parse_ysl_file
from template_populator import populate_template, apply_assumptions, apply_pm_projections

app = Flask(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# ─── Database Configuration ──────────────────────────────────────────────────
DATABASE_URL = os.environ.get("DATABASE_URL", "")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
if not DATABASE_URL:
    DATABASE_URL = "sqlite:///" + str(Path(__file__).parent / "data" / "budget.db")

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "century-budget-dev-key")

db = SQLAlchemy(app)

# ─── Admin Authentication ────────────────────────────────────────────────────
# Shared-secret gate for /api/admin/* endpoints. Until proper SSO is in place,
# this prevents anyone-with-the-URL from calling destructive admin endpoints.
#
# Accepts the secret via either:
#   - X-Admin-Key request header  (use for curl / scripts / Postman)
#   - admin_key cookie            (set by /admin/login form for browser use)
#
# To configure: set ADMIN_KEY in Railway env vars. If unset, admin endpoints
# remain open (with a startup warning) so we don't accidentally lock ourselves
# out of an existing deployment. Once ADMIN_KEY is set, every admin call
# without the right value gets a 401.
ADMIN_KEY = os.environ.get("ADMIN_KEY", "").strip()
ADMIN_COOKIE_NAME = "century_admin_key"
ADMIN_COOKIE_MAX_AGE = 60 * 60 * 24 * 30  # 30 days

if not ADMIN_KEY:
    logger.warning(
        "ADMIN_KEY env var is not set — /api/admin/* endpoints are OPEN. "
        "Set ADMIN_KEY in Railway env to enable authentication."
    )


def require_admin(f):
    """Decorator: require ADMIN_KEY via X-Admin-Key header or admin_key cookie.

    Returns 401 if the value doesn't match. If ADMIN_KEY env var is unset,
    requests pass through unchanged (failsafe — won't lock you out by default).
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not ADMIN_KEY:
            return f(*args, **kwargs)
        provided = (
            (request.headers.get("X-Admin-Key") or "").strip()
            or (request.cookies.get(ADMIN_COOKIE_NAME) or "").strip()
        )
        if provided != ADMIN_KEY:
            # Generic 401 — don't leak whether the key was wrong vs missing
            if request.path.startswith("/api/"):
                return jsonify({"error": "unauthorized — admin key required"}), 401
            # For non-API requests, redirect to login page
            from flask import redirect
            return redirect(f"/admin/login?next={request.path}")
        return f(*args, **kwargs)
    return wrapper


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    """Tiny login page that accepts the ADMIN_KEY and sets a cookie.
    Browser users hit this once; the cookie remembers them for 30 days."""
    next_url = request.args.get("next") or request.form.get("next") or "/"
    error = None
    if request.method == "POST":
        provided = (request.form.get("key") or "").strip()
        if not ADMIN_KEY:
            error = "Server has no ADMIN_KEY set. Configure it in Railway env."
        elif provided == ADMIN_KEY:
            from flask import redirect
            resp = make_response(redirect(next_url))
            resp.set_cookie(
                ADMIN_COOKIE_NAME, provided,
                max_age=ADMIN_COOKIE_MAX_AGE,
                httponly=True,
                secure=True,
                samesite="Lax",
            )
            return resp
        else:
            error = "Invalid key."
    # Minimal inline HTML — no template needed
    html = """
    <!DOCTYPE html>
    <html><head><title>Admin Login</title>
    <style>
      body{font-family:-apple-system,sans-serif;max-width:400px;margin:80px auto;padding:0 20px;color:#374151}
      h1{font-size:18px;margin-bottom:8px}
      p{color:#6b7280;font-size:13px;margin-bottom:20px}
      input{width:100%;padding:10px 12px;border:1px solid #d1d5db;border-radius:6px;font-size:14px;font-family:monospace}
      button{margin-top:12px;padding:10px 20px;background:#1a56db;color:white;border:none;border-radius:6px;cursor:pointer;font-weight:600}
      .err{color:#dc2626;font-size:13px;margin-top:8px}
    </style></head>
    <body>
      <h1>Admin login</h1>
      <p>Paste the ADMIN_KEY value to access admin endpoints. Cookie remembers for 30 days.</p>
      <form method="POST">
        <input type="password" name="key" autofocus placeholder="ADMIN_KEY" required>
        <input type="hidden" name="next" value="%NEXT%">
        <button type="submit">Sign in</button>
        %ERR%
      </form>
    </body></html>
    """.replace("%NEXT%", next_url).replace(
        "%ERR%", f'<div class="err">{error}</div>' if error else ""
    )
    return html


@app.route("/admin/logout", methods=["GET"])
def admin_logout():
    """Clear the admin cookie."""
    from flask import redirect
    resp = make_response(redirect("/admin/login"))
    resp.delete_cookie(ADMIN_COOKIE_NAME)
    return resp

# Register workflow blueprint (PM review, admin, dashboard)
try:
    from workflow import create_workflow_blueprint
except ImportError:
    from budget_app.workflow import create_workflow_blueprint
workflow_bp, workflow_models, workflow_helpers = create_workflow_blueprint(db)

# ─── One-time idempotent migrations ──────────────────────────────────────
# PostgreSQL supports ADD COLUMN IF NOT EXISTS — safe to run on every boot.
# Add new columns here when the model gains a field; remove the migration
# once the column is verified across all environments.
def _run_idempotent_migrations():
    """Run additive schema migrations that are safe to retry."""
    statements = [
        "ALTER TABLE budgets ADD COLUMN IF NOT EXISTS wizard_selections_json TEXT",
        # Phase E — Foundation gate
        "ALTER TABLE budgets ADD COLUMN IF NOT EXISTS foundation_confirmed_at TIMESTAMP",
        "ALTER TABLE budgets ADD COLUMN IF NOT EXISTS foundation_confirmed_by INTEGER",
        "ALTER TABLE budgets ADD COLUMN IF NOT EXISTS foundation_no_prior_budget BOOLEAN DEFAULT FALSE NOT NULL",
        "CREATE INDEX IF NOT EXISTS ix_budgets_foundation_confirmed_at ON budgets (foundation_confirmed_at)",
        # FA directive 2026-05-05: editable green-tab cells on summary row.
        # Override fields parallel BudgetLine.estimate_override / forecast_override.
        # When set, they take precedence over the live-computed value in
        # /api/summary; when NULL, the row falls back to GL aggregation.
        "ALTER TABLE budget_summary_rows ADD COLUMN IF NOT EXISTS col3_override DOUBLE PRECISION",
        "ALTER TABLE budget_summary_rows ADD COLUMN IF NOT EXISTS col4_override DOUBLE PRECISION",
        "ALTER TABLE budget_summary_rows ADD COLUMN IF NOT EXISTS col5_override DOUBLE PRECISION",
        # FA directive 2026-05-05: per-position benefit adjustments on payroll tab.
        # Lets the FA flag "N of M employees in this position have an extra
        # rate × periods adjustment on welfare/pension/etc". Math is additive
        # to the building default. JSON shape:
        #   {"adjusted_count": int, "label": str?, "benefits": {
        #       "welfare":{"rate":float,"periods":float,"label":str?}, ...}}
        "ALTER TABLE payroll_positions ADD COLUMN IF NOT EXISTS benefit_adjustments_json TEXT",
        # FA directive 2026-05-10: per-FA visit tracking for the diff-strip
        # feature. Each row = one (user, building) visit with a small JSON
        # snapshot of the building's state. Diff endpoint compares the
        # latest row to current state to surface "what changed since you
        # were last here". Lazy-GC: rows older than 90 days are deleted on
        # next diff call for the same (user, entity).
        """
        CREATE TABLE IF NOT EXISTS building_visits (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL,
            entity_code VARCHAR(50) NOT NULL,
            visited_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            snapshot_json TEXT NOT NULL,
            diff_dismissed_at TIMESTAMP NULL
        )
        """,
        "CREATE INDEX IF NOT EXISTS ix_building_visits_user_entity_visited "
        "ON building_visits (user_id, entity_code, visited_at DESC)",
    ]
    with app.app_context():
        for stmt in statements:
            try:
                db.session.execute(db.text(stmt))
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                logger.warning(f"Migration skipped or failed (non-fatal): {stmt} :: {e}")

try:
    _run_idempotent_migrations()
except Exception as e:
    logger.warning(f"Idempotent migration runner errored (non-fatal): {e}")

app.register_blueprint(workflow_bp)

# Register audited financials blueprint
try:
    from audited_financials import create_audited_financials_blueprint
except ImportError:
    from budget_app.audited_financials import create_audited_financials_blueprint
af_bp, af_models, af_helpers = create_audited_financials_blueprint(db)
app.register_blueprint(af_bp)

# Register expense distribution blueprint
try:
    from expense_distribution import create_expense_distribution_blueprint
except ImportError:
    from budget_app.expense_distribution import create_expense_distribution_blueprint
ed_bp, ed_models, ed_helpers = create_expense_distribution_blueprint(db, workflow_models)
app.register_blueprint(ed_bp)

# Register maintenance proof blueprint
try:
    from maintenance_proof import create_maintenance_proof_blueprint
except ImportError:
    from budget_app.maintenance_proof import create_maintenance_proof_blueprint
mp_bp, mp_models, mp_helpers = create_maintenance_proof_blueprint(db, workflow_models)
app.register_blueprint(mp_bp)

# Register open AP (aging) blueprint
try:
    from open_ap import create_open_ap_blueprint, detect_open_ap_file, parse_open_ap_report
except ImportError:
    from budget_app.open_ap import create_open_ap_blueprint, detect_open_ap_file, parse_open_ap_report
oa_bp, oa_models, oa_helpers = create_open_ap_blueprint(db, workflow_models)
app.register_blueprint(oa_bp)

# Register file repository blueprint
try:
    from file_repository import create_file_repository_blueprint
except ImportError:
    from budget_app.file_repository import create_file_repository_blueprint
fr_bp, fr_models, fr_helpers = create_file_repository_blueprint(db, workflow_models)
app.register_blueprint(fr_bp)

# Ensure every request starts with a clean DB session.
# Prevents poisoned PostgreSQL transactions from leaking via connection pool.
@app.before_request
def _ensure_clean_db_session():
    try:
        db.session.rollback()
    except Exception:
        pass

# ─── Global error handlers ────────────────────────────────────────────────
# Catches uncaught 500s anywhere in the app, logs the full traceback to
# Railway stdout, and shows a branded error page (HTML) or JSON error
# (for /api/* routes so frontend AJAX still parses cleanly).
# Set SHOW_TRACEBACKS=1 in env to show tracebacks inline in the browser.
@app.errorhandler(500)
@app.errorhandler(InternalServerError)
def _handle_500(err):
    import traceback
    tb = traceback.format_exc()
    path = request.path if request else '?'
    logger.error(f"[500] {path}: {err}\n{tb}")

    # API requests get JSON so frontend code can parse it
    wants_json = (
        path.startswith('/api/')
        or request.is_json
        or 'application/json' in (request.headers.get('Accept') or '')
    )
    if wants_json:
        return jsonify({
            "success": False,
            "error": "Internal server error",
            "path": path
        }), 500

    show_tb = os.environ.get("SHOW_TRACEBACKS", "").lower() in ("1", "true", "yes")
    safe_tb = (tb or "").replace("<", "&lt;").replace(">", "&gt;") if show_tb else ""
    tb_block = (
        f"<pre style='background:#111;color:#0f0;padding:12px;white-space:pre-wrap;font-size:12px;border-radius:6px;margin-top:20px;'>{safe_tb}</pre>"
        if show_tb else ""
    )
    return (
        "<!DOCTYPE html><html><head><title>Error - Century Budget</title>"
        "<style>body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:720px;margin:60px auto;padding:20px;color:#1a1714;background:#f4f1eb;}"
        "h1{color:#e02424;margin-bottom:12px;font-size:24px;}p{color:#4a4039;line-height:1.6;}"
        "a{color:#5a4a3f;font-weight:600;text-decoration:none;}a:hover{text-decoration:underline;}"
        "code{background:#ede9e1;padding:2px 8px;border-radius:4px;font-size:13px;}"
        ".card{background:white;border-radius:12px;padding:32px;border:1px solid #e5e0d5;}</style></head><body>"
        "<div class='card'>"
        "<h1>Something went wrong</h1>"
        f"<p>We hit an unexpected error on <code>{path}</code>. The issue has been logged and we'll take a look.</p>"
        "<p style='margin-top:20px;'><a href='/'>← Home</a> &nbsp;·&nbsp; <a href='/dashboard'>Dashboard</a></p>"
        f"{tb_block}"
        "</div></body></html>"
    ), 500

@app.errorhandler(404)
def _handle_404(err):
    if request.path.startswith('/api/'):
        return jsonify({"success": False, "error": "Not found", "path": request.path}), 404
    return (
        "<!DOCTYPE html><html><head><title>Not Found - Century Budget</title>"
        "<style>body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;max-width:720px;margin:60px auto;padding:20px;color:#1a1714;background:#f4f1eb;}"
        "h1{color:#5a4a3f;font-size:24px;margin-bottom:12px;}p{color:#4a4039;line-height:1.6;}"
        "a{color:#5a4a3f;font-weight:600;text-decoration:none;}a:hover{text-decoration:underline;}"
        ".card{background:white;border-radius:12px;padding:32px;border:1px solid #e5e0d5;}</style></head><body>"
        "<div class='card'>"
        "<h1>Page not found</h1><p>That page doesn't exist.</p>"
        "<p style='margin-top:20px;'><a href='/'>← Home</a> &nbsp;·&nbsp; <a href='/dashboard'>Dashboard</a></p>"
        "</div></body></html>"
    ), 404

# Resolve all model relationships after ALL blueprints are registered
try:
    db.configure_mappers()
    logger.info("Model mappers configured successfully")
except Exception as e:
    logger.warning(f"configure_mappers() failed: {e} — will retry on first request")

# Create all tables on startup + migrate missing columns
with app.app_context():
    db.create_all()
    logger.info("Database tables initialized")

    # Auto-migrate: add columns that exist in models but not in the DB
    if IS_CLOUD:
        _migrations = [
            ("budgets", "initiated_by", "INTEGER"),
            ("budgets", "initiated_at", "TIMESTAMP"),
            ("budgets", "return_to_status", "VARCHAR(20)"),
            ("budgets", "presentation_token", "VARCHAR(64)"),
            ("budgets", "approved_by", "INTEGER"),
            ("budgets", "approved_at", "TIMESTAMP"),
            ("budgets", "increase_pct", "FLOAT"),
            ("budgets", "effective_date", "VARCHAR(20)"),
            ("budgets", "ar_notes", "TEXT DEFAULT ''"),
            ("budget_lines", "sheet_name", "VARCHAR(50) DEFAULT ''"),
            ("budget_lines", "pm_editable", "BOOLEAN DEFAULT FALSE"),
            ("budget_lines", "reclass_to_gl", "VARCHAR(50)"),
            ("budget_lines", "reclass_amount", "FLOAT DEFAULT 0"),
            ("budget_lines", "reclass_notes", "TEXT DEFAULT ''"),
            ("budget_lines", "proposed_budget", "FLOAT DEFAULT 0"),
            ("budgets", "assumptions_json", "TEXT DEFAULT '{}'"),
            ("budget_lines", "estimate_override", "FLOAT"),
            ("budget_lines", "forecast_override", "FLOAT"),
            ("budget_lines", "accrual_adj", "FLOAT DEFAULT 0"),
            ("budgets", "building_type", "VARCHAR(50) DEFAULT ''"),
            ("budget_lines", "proposed_formula", "TEXT"),
            ("budgets", "version", "INTEGER DEFAULT 1"),
            ("budget_lines", "fa_proposed_status", "VARCHAR(20)"),
            ("budget_lines", "fa_proposed_note", "TEXT DEFAULT ''"),
            ("budget_lines", "fa_override_value", "FLOAT"),
            ("budget_lines", "backup_json", "TEXT"),
            ("budgets", "pre_merge_snapshot", "TEXT"),
            ("budgets", "pre_merge_snapshot_at", "TIMESTAMP"),
            ("payroll_positions", "bonus_per_employee", "FLOAT DEFAULT 0"),
            ("payroll_positions", "effective_week_override", "FLOAT"),
            ("payroll_positions", "wage_increase_mode", "VARCHAR(10)"),
            ("payroll_positions", "wage_increase_value", "FLOAT"),
            ("payroll_positions", "extra_bonuses_json", "TEXT"),
            ("budgets", "assumptions_history_json", "TEXT"),
            ("budgets", "wizard_completed_at", "TIMESTAMP"),
            ("budgets", "wizard_step", "INTEGER DEFAULT 0"),
            ("budgets", "pm_sent_at", "TIMESTAMP"),
            ("audit_uploads", "summary_overrides", "TEXT"),
            ("audit_uploads", "sharepoint_web_url", "TEXT"),
            ("building_info", "common_charges_history_json", "TEXT"),
        ]
        # Create payroll tables if they don't exist
        _payroll_tables = [
            """CREATE TABLE IF NOT EXISTS payroll_positions (
                id SERIAL PRIMARY KEY,
                entity_code VARCHAR(50) NOT NULL,
                budget_year INTEGER NOT NULL,
                position_name VARCHAR(100) NOT NULL,
                employee_count INTEGER DEFAULT 0,
                hourly_rate FLOAT DEFAULT 0,
                sort_order INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS payroll_assumptions (
                id SERIAL PRIMARY KEY,
                entity_code VARCHAR(50) NOT NULL,
                budget_year INTEGER NOT NULL,
                assumptions_json TEXT DEFAULT '{}',
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
        ]
        for ddl in _payroll_tables:
            try:
                db.session.execute(db.text(ddl))
                db.session.commit()
            except Exception:
                db.session.rollback()
        for table, col, col_type in _migrations:
            try:
                db.session.execute(db.text(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}"))
                db.session.commit()
                logger.info(f"Added column {table}.{col}")
            except Exception:
                db.session.rollback()  # Column already exists, skip

        # Migrate unique constraint: drop old, add new with version
        try:
            db.session.execute(db.text("ALTER TABLE budgets DROP CONSTRAINT IF EXISTS uq_entity_year"))
            db.session.commit()
            logger.info("Dropped old uq_entity_year constraint")
        except Exception:
            db.session.rollback()
        try:
            db.session.execute(db.text("ALTER TABLE budgets ADD CONSTRAINT uq_entity_year_ver UNIQUE (entity_code, year, version)"))
            db.session.commit()
            logger.info("Added new uq_entity_year_ver constraint")
        except Exception:
            db.session.rollback()

        # Backfill building_type on existing budgets from buildings.csv
        try:
            import csv as _csv
            _csv_path = BUDGET_SYSTEM / "buildings.csv"
            _type_map = {}
            if _csv_path.exists():
                with open(_csv_path, newline="", encoding="utf-8") as _f:
                    for _r in _csv.DictReader(_f):
                        ec = str(_r.get("entity_code", "")).strip()
                        bt = (_r.get("type", "") or "").strip()
                        if ec and bt:
                            _type_map[ec] = bt
            empty_bt = db.session.execute(
                db.text("SELECT id, entity_code FROM budgets WHERE building_type IS NULL OR building_type = ''")
            ).fetchall()
            if empty_bt and _type_map:
                for row in empty_bt:
                    bt = _type_map.get(str(row[1]).strip(), "")
                    if bt:
                        db.session.execute(
                            db.text("UPDATE budgets SET building_type = :bt WHERE id = :id"),
                            {"bt": bt, "id": row[0]}
                        )
                db.session.commit()
                logger.info(f"Backfilled building_type on {len(empty_bt)} budgets")
        except Exception as e:
            db.session.rollback()
            logger.warning(f"Building type backfill skipped: {e}")

        # Backfill: move 7xxx GL codes from 'Unmapped' sheet to 'Capital' sheet
        try:
            from workflow import CAPITAL_GL_PREFIX
            cap_lines = db.session.execute(
                db.text("SELECT id, gl_code FROM budget_lines WHERE sheet_name = 'Unmapped' AND gl_code LIKE '7%'")
            ).fetchall()
            moved = 0
            for row in cap_lines:
                line_id, gl_code = row[0], row[1]
                prefix = (gl_code or "")[:4]
                desc = CAPITAL_GL_PREFIX.get(prefix, f"Cap - {prefix}")
                db.session.execute(
                    db.text("UPDATE budget_lines SET sheet_name = 'Capital', category = 'capital', pm_editable = TRUE, description = :desc WHERE id = :id"),
                    {"desc": desc, "id": line_id}
                )
                moved += 1
            if moved:
                db.session.commit()
                logger.info(f"Backfilled {moved} 7xxx lines from Unmapped to Capital")
        except Exception as e:
            db.session.rollback()
            logger.warning(f"Capital backfill skipped: {e}")

        # Backfill: route unmapped budget_lines to their proper tab using GL_Mapping.csv.
        # Only moves lines whose 4-digit prefix has an explicit routing rule in the CSV.
        # Balance sheet codes (1xxx/2xxx/3xxx) and codes missing from the mapping stay Unmapped.
        try:
            from workflow import GL_MAPPING_CSV
            if GL_MAPPING_CSV:
                rows = db.session.execute(
                    db.text("SELECT id, gl_code FROM budget_lines WHERE sheet_name = 'Unmapped'")
                ).fetchall()
                moved = 0
                for r in rows:
                    line_id, gl_code = r[0], r[1] or ""
                    hit = GL_MAPPING_CSV.get(gl_code[:4])
                    if hit:
                        desc, sheet_name, category = hit
                        db.session.execute(
                            db.text("UPDATE budget_lines SET sheet_name = :s, category = :c, pm_editable = TRUE, description = :d WHERE id = :id"),
                            {"s": sheet_name, "c": category, "d": desc, "id": line_id}
                        )
                        moved += 1
                if moved:
                    db.session.commit()
                    logger.info(f"Routed {moved} unmapped budget_lines to proper tabs via GL_Mapping.csv")
        except Exception as e:
            db.session.rollback()
            logger.warning(f"GL_Mapping routing backfill skipped: {e}")

        # Backfill: split Repairs & Supplies lines from the lumped 'rm' category
        # into proper 'supplies' / 'repairs' / 'maintenance' sub-buckets.
        # The FA dashboard, PM portal, and Board Presentation all group the R&S
        # tab by matching BudgetLine.category against those three strings, so
        # any line left as 'rm' silently disappeared from its sub-group.
        # Re-look-up each 'rm' row by 4-digit prefix in GL_Mapping.csv; if the
        # CSV has no hit, leave the row untouched (we do not guess).
        try:
            from workflow import GL_MAPPING_CSV
            if GL_MAPPING_CSV:
                rm_rows = db.session.execute(
                    db.text("SELECT id, gl_code FROM budget_lines WHERE category = 'rm'")
                ).fetchall()
                fixed = 0
                for r in rm_rows:
                    line_id, gl_code = r[0], r[1] or ""
                    hit = GL_MAPPING_CSV.get(gl_code[:4])
                    if not hit:
                        continue
                    new_cat = hit[2]
                    if new_cat in ("supplies", "repairs", "maintenance"):
                        db.session.execute(
                            db.text("UPDATE budget_lines SET category = :c WHERE id = :id"),
                            {"c": new_cat, "id": line_id}
                        )
                        fixed += 1
                if fixed:
                    db.session.commit()
                    logger.info(f"Split {fixed} budget_lines from 'rm' into supplies/repairs/maintenance")
        except Exception as e:
            db.session.rollback()
            logger.warning(f"R&S sub-category backfill skipped: {e}")

        # Backfill: correct stale gl_prefixes_json on budget_summary_rows.
        # Legacy push files carried pre-Yardi chart-of-accounts prefixes for
        # Electric/Steam/Gas/Water & Sewer/Supplies, which prevented YTD from
        # flowing on the Budget Summary tab. SUMMARY_PREFIX_OVERRIDES is the
        # canonical Yardi prefix list keyed by label; this backfill rewrites
        # gl_prefixes_json for any row whose label matches.
        try:
            import json as _json
            from workflow import SUMMARY_PREFIX_OVERRIDES
            fixed = 0
            for label, correct_prefixes in SUMMARY_PREFIX_OVERRIDES.items():
                new_json = _json.dumps(list(correct_prefixes))
                result = db.session.execute(
                    db.text(
                        "UPDATE budget_summary_rows "
                        "SET gl_prefixes_json = :p "
                        "WHERE label = :l AND (gl_prefixes_json IS NULL OR gl_prefixes_json != :p)"
                    ),
                    {"p": new_json, "l": label}
                )
                fixed += result.rowcount or 0
            if fixed:
                db.session.commit()
                logger.info(f"Corrected gl_prefixes_json on {fixed} budget_summary_rows via SUMMARY_PREFIX_OVERRIDES")
        except Exception as e:
            db.session.rollback()
            logger.warning(f"Summary prefix override backfill skipped: {e}")

# Health check for Railway
@app.route("/healthz")
def healthz():
    return "ok", 200

# Paths
TEMPLATE_PATH = BUDGET_SYSTEM / "Budget_Final_Template_v2.xlsx"
BUILDINGS_CSV = BUDGET_SYSTEM / "buildings.csv"
SETTINGS_FILE = Path(__file__).parent / "settings.json"
DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)

# Data files
PORTFOLIO_DEFAULTS_FILE = DATA_DIR / "portfolio_defaults.json"
BUILDING_ASSUMPTIONS_FILE = DATA_DIR / "building_assumptions.json"

# Default save location
DEFAULT_SAVE_DIR = str(BUDGET_SYSTEM / "budgets")

# Console JS script templates — entities/email/period get injected
# Normalize CRLF→LF so string replace patches work on all platforms
CONSOLE_SCRIPT = (BUDGET_SYSTEM / "YSL Budget Script.js").read_text(encoding="utf-8").replace("\r\n", "\n")
MAINT_PROOF_SCRIPT = (BUDGET_SYSTEM / "Maintenance Proof Script.js").read_text(encoding="utf-8").replace("\r\n", "\n")
try:
    AP_AGING_SCRIPT = (BUDGET_SYSTEM / "AP Aging Script.js").read_text(encoding="utf-8").replace("\r\n", "\n")
except FileNotFoundError:
    AP_AGING_SCRIPT = None

# Default portfolio values
DEFAULT_PORTFOLIO = {
    "payroll_tax": {
        "FICA": 0.0765,
        "SUI": 0.0525,
        "FUI": 0.006,
        "MTA": 0.0034,
        "NYS_Disability": 0.005,
        "PFL": 0.00455
    },
    "union_benefits": {
        "welfare_monthly": 1072.55,
        "pension_weekly": 82.75,
        "supp_retirement_weekly": 10.00,
        "legal_monthly": 16.63,
        "training_monthly": 14.13,
        "profit_sharing_quarterly": 130.00
    },
    "workers_comp": {
        "percent": 0.15
    },
    "wage_increase": {
        "percent": 0.03,
        "effective_week": "Wk 16",
        "pre_increase_weeks": 15,
        "post_increase_weeks": 37
    },
    "insurance_renewal": {
        "increase_percent": 0.15,
        "effective_date": "Mar 2027",
        "pre_renewal_months": 3,
        "post_renewal_months": 9
    },
    "energy": {
        "gas_esco_rate": 0,
        "electric_esco_rate": 0,
        "oil_price_per_gallon": 0,
        "gas_rate_increase": 0.05,
        "electric_rate_increase": 0.05,
        "oil_rate_increase": 0.05,
        "consumption_basis": "2-Year Average"
    },
    "water_sewer": {
        "rate_increase": 0.05
    }
}

INSURANCE_POLICIES = [
    {"gl_code": "6105-0000", "name": "Package"},
    {"gl_code": "6110-0000", "name": "Gen Liability"},
    {"gl_code": "6115-0000", "name": "Umbrella Excess"},
    {"gl_code": "6120-0000", "name": "Umbrella Primary"},
    {"gl_code": "6125-0000", "name": "D&O"},
    {"gl_code": "6126-0000", "name": "Cyber"},
    {"gl_code": "6135-0000", "name": "Crime"},
    {"gl_code": "6180-0000", "name": "D&O Excess"},
    {"gl_code": "6195-0000", "name": "Other"}
]

ENERGY_GL_ACCOUNTS = [
    {"gl": "5252-0000", "desc": "Gas - Heating"},
    {"gl": "5252-0001", "desc": "Gas - Cooking"},
    {"gl": "5252-0010", "desc": "Gas - Common Area"},
    {"gl": "5253-0000", "desc": "Oil / Fuel"},
    {"gl": "5250-0000", "desc": "Electric"}
]

WATER_GL_ACCOUNTS = [
    {"gl": "6305-0000", "desc": "Water/Sewer"},
    {"gl": "6305-0010", "desc": "Water - Common Area"},
    {"gl": "6305-0020", "desc": "Sewer Charges"}
]


# ─── Settings Functions ───────────────────────────────────────────────────────

def load_settings():
    """Load app settings from disk."""
    if SETTINGS_FILE.exists():
        return json.loads(SETTINGS_FILE.read_text())
    return {"save_dir": DEFAULT_SAVE_DIR}


def save_settings(settings):
    """Persist app settings to disk."""
    SETTINGS_FILE.write_text(json.dumps(settings, indent=2))


# ─── Buildings Functions ──────────────────────────────────────────────────────

def load_buildings():
    """Load building list from CSV."""
    buildings = []
    if BUILDINGS_CSV.exists():
        with open(BUILDINGS_CSV) as f:
            reader = csv.DictReader(f)
            for row in reader:
                buildings.append(row)
    return buildings


def save_buildings(buildings):
    """Save building list back to CSV."""
    if not buildings:
        return
    fieldnames = ["entity_code", "building_name", "address", "city", "zip", "type", "units"]
    with open(BUILDINGS_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(buildings)


# ─── Assumptions Functions ─────────────────────────────────────────────────────

def load_portfolio_defaults():
    """Load portfolio-wide defaults, or return defaults."""
    if PORTFOLIO_DEFAULTS_FILE.exists():
        return json.loads(PORTFOLIO_DEFAULTS_FILE.read_text())
    return DEFAULT_PORTFOLIO


def save_portfolio_defaults(data):
    """Save portfolio defaults to disk."""
    PORTFOLIO_DEFAULTS_FILE.write_text(json.dumps(data, indent=2))


def load_building_assumptions():
    """Load all building assumptions, or return empty dict."""
    if BUILDING_ASSUMPTIONS_FILE.exists():
        return json.loads(BUILDING_ASSUMPTIONS_FILE.read_text())
    return {}


def save_building_assumptions(data):
    """Save building assumptions to disk."""
    BUILDING_ASSUMPTIONS_FILE.write_text(json.dumps(data, indent=2))


def merge_assumptions(entity_code):
    """Merge portfolio defaults with building-specific overrides."""
    defaults = load_portfolio_defaults()
    building_data = load_building_assumptions().get(entity_code, {})

    # Start with defaults
    merged = json.loads(json.dumps(defaults))

    # Override with building-specific data
    if "payroll_tax" in building_data:
        merged["payroll_tax"].update(building_data["payroll_tax"])
    if "union_benefits" in building_data:
        merged["union_benefits"].update(building_data["union_benefits"])
    if "workers_comp" in building_data:
        merged["workers_comp"].update(building_data["workers_comp"])
    if "wage_increase" in building_data:
        merged["wage_increase"].update(building_data["wage_increase"])
    if "insurance_renewal" in building_data:
        merged["insurance_renewal"].update(building_data["insurance_renewal"])
    if "energy" in building_data:
        merged["energy"].update(building_data["energy"])
    if "water_sewer" in building_data:
        merged["water_sewer"].update(building_data["water_sewer"])

    merged["payroll"] = building_data.get("payroll", {})
    merged["income"] = building_data.get("income", {})
    merged["insurance"] = building_data.get("insurance", [])

    return merged


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    """Unified home page."""
    return render_template_string(HOME_TEMPLATE)


@app.route("/generate")
def generate():
    """Budget generator page."""
    settings = load_settings()
    return render_template_string(
        GENERATE_TEMPLATE,
        buildings=load_buildings(),
        save_dir=settings.get("save_dir", DEFAULT_SAVE_DIR),
    )


@app.route("/assumptions")
def assumptions():
    """Portfolio defaults page."""
    data = load_portfolio_defaults()
    return render_template_string(ASSUMPTIONS_TEMPLATE, data=json.dumps(data))


@app.route("/assumptions/buildings")
def assumptions_buildings():
    """Building assumptions grid."""
    bldgs = load_buildings()
    policies = INSURANCE_POLICIES
    return render_template_string(
        ASSUMPTIONS_BUILDINGS_TEMPLATE,
        buildings=json.dumps(bldgs),
        policies=json.dumps(policies)
    )


@app.route("/assumptions/workbench")
def assumptions_workbench():
    """Unified assumptions workbench — portfolio defaults + per-building overrides in one view."""
    defaults = load_portfolio_defaults()
    bldgs = load_buildings()
    all_bldg = load_building_assumptions()
    policies = INSURANCE_POLICIES

    # Compute override counts per building for sidebar badges.
    # Count non-empty/non-zero overrides in energy + water_sewer sections
    # (those are the only overridable fields surfaced in the current UI).
    override_counts = {}
    for b in bldgs:
        code = b["entity_code"]
        bd = all_bldg.get(code, {})
        count = 0
        for section_key in ("energy", "water_sewer"):
            section = bd.get(section_key, {}) or {}
            for k, v in section.items():
                if k == "consumption_basis":
                    if v:
                        count += 1
                else:
                    try:
                        if v not in (None, "", 0, 0.0) and float(v) != 0:
                            count += 1
                    except (TypeError, ValueError):
                        if v:
                            count += 1
        override_counts[code] = count

    return render_template_string(
        ASSUMPTIONS_WORKBENCH_TEMPLATE,
        defaults=json.dumps(defaults),
        buildings=json.dumps(bldgs),
        policies=json.dumps(policies),
        override_counts=json.dumps(override_counts),
    )


# ─── API Routes ───────────────────────────────────────────────────────────────

@app.route("/api/settings", methods=["GET", "POST"])
def settings_api():
    """Get or update app settings."""
    if request.method == "GET":
        return jsonify(load_settings())

    data = request.json
    settings = load_settings()
    if "save_dir" in data:
        save_dir = data["save_dir"].strip()
        if save_dir:
            # Validate path exists or can be created
            try:
                Path(save_dir).mkdir(parents=True, exist_ok=True)
                settings["save_dir"] = save_dir
            except Exception as e:
                return jsonify({"error": f"Invalid path: {e}"}), 400
        else:
            settings["save_dir"] = DEFAULT_SAVE_DIR

    save_settings(settings)
    return jsonify(settings)


@app.route("/api/buildings", methods=["GET", "POST"])
def buildings_api():
    """Get all buildings or add a new building."""
    if request.method == "GET":
        return jsonify(load_buildings())

    data = request.json
    required_fields = ["entity_code", "building_name", "address", "city", "zip", "type", "units"]
    if not all(f in data for f in required_fields):
        return jsonify({"error": "Missing required fields"}), 400

    buildings = load_buildings()
    # Check if entity_code already exists
    if any(b["entity_code"] == data["entity_code"] for b in buildings):
        return jsonify({"error": "Entity code already exists"}), 400

    new_building = {f: data.get(f, "") for f in required_fields}
    buildings.append(new_building)
    save_buildings(buildings)
    return jsonify(new_building), 201


@app.route("/api/buildings/<entity_code>", methods=["PUT"])
def manage_building(entity_code):
    """Edit a building."""
    buildings = load_buildings()
    building_idx = next((i for i, b in enumerate(buildings) if b["entity_code"] == entity_code), None)

    if building_idx is None:
        return jsonify({"error": "Building not found"}), 404

    data = request.json
    building = buildings[building_idx]
    # Update allowed fields
    for field in ["building_name", "address", "city", "zip", "type", "units"]:
        if field in data:
            building[field] = data[field]
    buildings[building_idx] = building
    save_buildings(buildings)
    return jsonify(building), 200


@app.route("/api/buildings/<entity_code>/delete", methods=["POST"])
def delete_building(entity_code):
    """Delete a building."""
    buildings = load_buildings()
    building_idx = next((i for i, b in enumerate(buildings) if b["entity_code"] == entity_code), None)

    if building_idx is None:
        return jsonify({"error": "Building not found"}), 404

    buildings.pop(building_idx)
    save_buildings(buildings)
    return jsonify({"message": "Building deleted"}), 200


# ─── CORS Helper for Yardi Auto-Upload ────────────────────────────────────
def cors_headers(f):
    """Decorator to add CORS headers for cross-origin Yardi→Railway requests."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.method == "OPTIONS":
            resp = make_response()
            resp.headers["Access-Control-Allow-Origin"] = "*"
            resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
            resp.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Upload-Session, X-Entity-Code, X-File-Type, X-Filename"
            resp.headers["Access-Control-Max-Age"] = "3600"
            return resp
        result = f(*args, **kwargs)
        if isinstance(result, tuple):
            resp = make_response(result[0], result[1])
        else:
            resp = make_response(result)
        resp.headers["Access-Control-Allow-Origin"] = "*"
        return resp
    return decorated


# ─── Auto-Upload: Yardi → Budget App (no manual file upload needed) ───────
_upload_sessions = {}  # session_id → {files: [{name, data, type}], entities: [], period: str}

@app.route("/api/auto-upload", methods=["POST", "OPTIONS"])
@cors_headers
def auto_upload_file():
    """Accept a single file blob from the Yardi console script.

    The Yardi script POSTs each downloaded report here as it completes.
    Files accumulate in a server-side session until /api/auto-process is called.
    """
    if request.method == "OPTIONS":
        return make_response(""), 200

    session_id = request.headers.get("X-Upload-Session", "default")
    entity_code = request.headers.get("X-Entity-Code", "")
    file_type = request.headers.get("X-File-Type", "unknown")  # ysl, expense, maint, ap

    if session_id not in _upload_sessions:
        _upload_sessions[session_id] = {"files": [], "entities": set(), "period": ""}

    sess = _upload_sessions[session_id]

    # Accept the file from the request body
    if request.content_type and "multipart" in request.content_type:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No file in request"}), 400
        filename = f.filename or f"upload_{entity_code}_{file_type}.xls"
        file_data = f.read()
    else:
        # Raw binary body
        file_data = request.get_data()
        filename = request.headers.get("X-Filename", f"upload_{entity_code}_{file_type}.xls")

    if not file_data:
        return jsonify({"error": "Empty file"}), 400

    sess["files"].append({
        "name": filename,
        "data": file_data,
        "type": file_type,
        "entity": entity_code,
    })
    if entity_code:
        sess["entities"].add(entity_code)

    # Persist to disk: data/yardi_archives/YYYY-MM-DD/<entity>/<filename>
    try:
        from datetime import datetime as _dt
        archive_root = os.path.join(os.path.dirname(__file__), "data", "yardi_archives")
        today = _dt.now().strftime("%Y-%m-%d")
        entity_dir = os.path.join(archive_root, today, entity_code or "unknown")
        os.makedirs(entity_dir, exist_ok=True)
        archive_path = os.path.join(entity_dir, filename)
        with open(archive_path, "wb") as af:
            af.write(file_data)
        logger.info(f"Archived Yardi file to {archive_path}")
    except Exception as arch_err:
        logger.warning(f"Failed to archive {filename}: {arch_err}")

    logger.info(f"Auto-upload: received {filename} ({len(file_data)} bytes) for entity {entity_code}, session {session_id}")

    return jsonify({
        "status": "received",
        "filename": filename,
        "size": len(file_data),
        "total_files": len(sess["files"]),
    })


@app.route("/api/qa/verify", methods=["GET", "OPTIONS"])
@cors_headers
def qa_verify():
    """Read-only QA endpoint: returns archive listing + DB state for an entity.

    Query params:
      entity  (required) - entity code to inspect
      date    (optional) - archive date YYYY-MM-DD, defaults to today
    """
    if request.method == "OPTIONS":
        return make_response(""), 200

    from datetime import datetime as _dt
    entity = request.args.get("entity", "").strip()
    date_str = request.args.get("date", _dt.now().strftime("%Y-%m-%d"))
    if not entity:
        return jsonify({"error": "entity param required"}), 400

    result = {"entity": entity, "date": date_str}

    # Archive listing
    archive_root = os.path.join(os.path.dirname(__file__), "data", "yardi_archives")
    entity_dir = os.path.join(archive_root, date_str, entity)
    archived = []
    if os.path.isdir(entity_dir):
        for fn in sorted(os.listdir(entity_dir)):
            fp = os.path.join(entity_dir, fn)
            try:
                st = os.stat(fp)
                archived.append({
                    "name": fn,
                    "size": st.st_size,
                    "mtime": _dt.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
                })
            except Exception:
                pass
    result["archive"] = {
        "dir": entity_dir,
        "exists": os.path.isdir(entity_dir),
        "files": archived,
        "count": len(archived),
    }

    # DB state — latest budget for this entity
    try:
        from sqlalchemy import text as _sql
        row = db.session.execute(_sql(
            "SELECT id, year, status, updated_at FROM budgets "
            "WHERE entity_code = :e ORDER BY updated_at DESC LIMIT 1"
        ), {"e": entity}).fetchone()
        if row:
            bid = row[0]
            stats = db.session.execute(_sql(
                "SELECT COUNT(*) lines, "
                "COALESCE(SUM(unpaid_bills),0) unpaid_total, "
                "COALESCE(SUM(CASE WHEN unpaid_bills<>0 THEN 1 ELSE 0 END),0) lines_with_unpaid, "
                "COALESCE(SUM(ytd_actual),0) ytd_actual_total, "
                "COALESCE(SUM(accrual_adj),0) accrual_total "
                "FROM budget_lines WHERE budget_id = :bid"
            ), {"bid": bid}).fetchone()
            result["db"] = {
                "budget_id": bid,
                "year": row[1],
                "status": row[2],
                "updated_at": row[3].isoformat(timespec="seconds") if row[3] else None,
                "lines": stats[0],
                "unpaid_bills_total": float(stats[1] or 0),
                "lines_with_unpaid": int(stats[2] or 0),
                "ytd_actual_total": float(stats[3] or 0),
                "accrual_total": float(stats[4] or 0),
            }
        else:
            result["db"] = {"budget_id": None, "note": "no budget row for this entity"}
    except Exception as db_err:
        result["db"] = {"error": str(db_err)}

    return jsonify(result)


@app.route("/api/auto-process", methods=["POST", "OPTIONS"])
@cors_headers
def auto_process():
    """Trigger processing of all files accumulated via auto-upload.

    Called by the Yardi script after all downloads complete.
    Reuses the same processing logic as /api/process.
    """
    if request.method == "OPTIONS":
        return make_response(""), 200

    data = request.get_json(silent=True) or {}
    session_id = data.get("session_id", request.headers.get("X-Upload-Session", "default"))
    period = data.get("period", "02/2026")

    if session_id not in _upload_sessions or not _upload_sessions[session_id]["files"]:
        return jsonify({"error": "No files in session. Upload files first via /api/auto-upload"}), 400

    sess = _upload_sessions.pop(session_id)
    file_list = sess["files"]

    logger.info(f"Auto-process: session {session_id}, {len(file_list)} files, entities: {sess['entities']}")

    # Save files to a temp dir and reuse the existing /api/process logic
    # by writing them as real files and processing them
    save_dir_str = load_settings().get("save_dir", DEFAULT_SAVE_DIR)
    save_dir = Path(save_dir_str)
    ysl_archive = BUDGET_SYSTEM / "ysl_downloads"
    ysl_archive.mkdir(exist_ok=True)

    results = {"success": [], "failed": [], "warnings": [], "save_dir": str(save_dir), "auto_upload": True}

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        output_files = []
        saved_files = []
        ysl_entities = set()
        expense_files = []
        open_ap_files = []
        maint_proof_files = []

        # Pass 1: Save all files, classify, process YSL first
        for finfo in file_list:
            filename = finfo["name"]
            if not filename.endswith((".xlsx", ".xls")):
                results["warnings"].append(f"Skipped non-Excel: {filename}")
                continue

            file_path = tmp / filename
            file_path.write_bytes(finfo["data"])

            # Use upload-time file type hint as primary classifier
            upload_type = finfo.get("type", "unknown")

            # If upload type is 'ap', route directly to Open AP pipeline
            if upload_type == "ap":
                logger.info(f"Auto-upload routed by type hint: {filename} → Open AP")
                open_ap_files.append((file_path, filename))
                continue

            # If upload type is 'maint', route to Maintenance Proof pipeline
            if upload_type == "maint":
                logger.info(f"Auto-upload routed by type hint: {filename} → Maintenance Proof")
                maint_proof_files.append((file_path, filename))
                continue

            # Detect Open AP from content (fallback for manual uploads)
            try:
                if detect_open_ap_file(str(file_path)):
                    logger.info(f"Auto-upload detection for {filename}: Open AP (Aging) report")
                    open_ap_files.append((file_path, filename))
                    continue
            except Exception:
                pass

            # Handle .xls files that are actually .xlsx (Yardi naming quirk)
            # Copy to .xlsx temp file if needed for openpyxl
            check_path = file_path
            _tmp_xlsx = None
            if str(file_path).lower().endswith('.xls') and not str(file_path).lower().endswith('.xlsx'):
                import shutil as _shutil
                _tmp_xlsx = tmp / (filename + 'x')  # .xls → .xlsx
                _shutil.copy2(str(file_path), str(_tmp_xlsx))
                check_path = _tmp_xlsx

            # Detect Expense Distribution vs YSL via Row 1
            try:
                from openpyxl import load_workbook as _lwb
                _wb_check = _lwb(str(check_path), data_only=True)
                _ws_check = _wb_check.active
                _row1 = str(_ws_check.cell(row=1, column=1).value or "")
                _row2 = str(_ws_check.cell(row=2, column=1).value or "")
                _wb_check.close()

                is_expense = "expense distribution" in _row1.lower() or "expense dist" in _row1.lower()
                logger.info(f"Auto-upload detection for {filename}: Row1='{_row1}', is_expense={is_expense}")

                if is_expense:
                    expense_files.append((file_path, filename))
                    continue

                is_maint = "maintenance proof" in _row1.lower() or "adhoc_amp" in _row1.lower()
                if is_maint:
                    logger.info(f"Auto-upload detection for {filename}: Maintenance Proof report")
                    maint_proof_files.append((file_path, filename))
                    continue
            except Exception as detect_err:
                logger.error(f"Auto-upload detection error for {filename}: {detect_err}")
                results["warnings"].append(f"Could not read {filename}: {str(detect_err)}")
                continue

            # Process YSL file
            try:
                gl_data, property_info = parse_ysl_file(file_path)
                entity = property_info.get("property_code", "unknown")
                name = property_info.get("property_name", f"Entity_{entity}")
                ysl_entities.add(str(entity))

                building_folder_name = f"{entity} - {name}"
                building_dir = save_dir / building_folder_name
                building_dir.mkdir(parents=True, exist_ok=True)

                output_name = f"{entity}_{name}_2027_Budget.xlsx"
                output_path = building_dir / output_name

                _gen_ytd = 2
                try:
                    _gen_ytd = int(period.split("/")[0])
                except Exception:
                    pass

                success = populate_template(
                    template_path=TEMPLATE_PATH,
                    gl_data=gl_data,
                    property_info=property_info,
                    output_path=output_path,
                    ytd_months=_gen_ytd,
                    remaining_months=12 - _gen_ytd,
                )

                if success and output_path.exists():
                    merged = None
                    try:
                        merged = merge_assumptions(entity)
                    except Exception:
                        pass

                    try:
                        fresh = data.get("fresh_start", False)
                        workflow_helpers["store_all_lines"](entity, name, gl_data, TEMPLATE_PATH, assumptions=merged, fresh_start=fresh)
                    except Exception as wfe:
                        logger.warning(f"Could not store GL data for {entity}: {wfe}")

                    try:
                        if merged:
                            apply_assumptions(output_path, merged)
                    except Exception as ae:
                        logger.warning(f"Could not apply assumptions for {entity}: {ae}")

                    try:
                        pm_proj = workflow_helpers["get_pm_projections"](entity)
                        if pm_proj:
                            apply_pm_projections(output_path, pm_proj)
                    except Exception as pe:
                        logger.warning(f"Could not apply PM projections for {entity}: {pe}")

                    output_files.append((output_name, output_path))
                    shutil.copy2(file_path, ysl_archive / filename)
                    yardi_drops = building_dir / "yardi_drops"
                    yardi_drops.mkdir(exist_ok=True)
                    shutil.copy2(file_path, yardi_drops / filename)

                    results["success"].append({
                        "entity": entity, "name": name, "file": output_name,
                        "size_kb": round(output_path.stat().st_size / 1024),
                        "saved_to": str(building_dir),
                    })
                else:
                    results["failed"].append({"entity": entity, "file": filename, "reason": "Pipeline returned failure"})
            except Exception as e:
                logger.exception(f"Auto-upload error processing {filename}")
                results["failed"].append({"file": filename, "reason": str(e)})

        # Pass 2: Expense Distribution
        for exp_path, exp_filename in expense_files:
            try:
                from expense_distribution import parse_expense_distribution
            except ImportError:
                from budget_app.expense_distribution import parse_expense_distribution
            try:
                exp_entity, exp_from, exp_to, exp_invoices = parse_expense_distribution(str(exp_path))
                if not exp_invoices:
                    results["warnings"].append(f"Expense file {exp_filename}: no invoices found")
                    continue

                import re as _re
                fname_match = _re.search(r'ExpenseDistribution[_-]?(\d+)', exp_filename)
                fname_entity = fname_match.group(1) if fname_match else None

                target_entity = exp_entity
                if fname_entity and fname_entity != exp_entity:
                    if fname_entity in ysl_entities:
                        target_entity = fname_entity
                    elif not ysl_entities:
                        target_entity = fname_entity

                report = ed_helpers["store_expense_report"](target_entity, exp_from, exp_to, exp_invoices, exp_filename)
                accrual_result = {"applied": 0, "accruals": {}}
                if exp_from:
                    try:
                        accrual_result = ed_helpers["apply_accrual_adjustments"](target_entity, report.id, exp_from)
                    except Exception as accrual_err:
                        results["warnings"].append(f"Accrual adjustments failed for {target_entity}: {str(accrual_err)}")

                accrual_msg = f", {accrual_result['applied']} accrual adj" if accrual_result["applied"] > 0 else ""
                results["success"].append(f"Expense Distribution: {target_entity} ({len(exp_invoices)} invoices{accrual_msg})")
            except Exception as exp_err:
                results["warnings"].append(f"Expense file {exp_filename} error: {str(exp_err)}")

        # Pass 3: Open AP
        for ap_path, ap_filename in open_ap_files:
            try:
                ap_entity, ap_invoices = parse_open_ap_report(str(ap_path))
                if not ap_invoices:
                    results["warnings"].append(f"Open AP file {ap_filename}: no invoices found")
                    continue

                import re as _re
                fname_match = _re.search(r'(?:Aging|OpenAP|AP)[_-]?(\d+)', ap_filename, _re.IGNORECASE)
                fname_entity = fname_match.group(1) if fname_match else None

                target_entity = ap_entity
                if fname_entity and fname_entity != ap_entity:
                    if fname_entity in ysl_entities:
                        target_entity = fname_entity
                    elif not ysl_entities:
                        target_entity = fname_entity
                elif not target_entity and fname_entity:
                    target_entity = fname_entity

                if not target_entity:
                    results["warnings"].append(f"Open AP file {ap_filename}: could not determine entity code")
                    continue

                report = oa_helpers["store_open_ap_report"](target_entity, ap_invoices, ap_filename)
                unpaid_result = {"applied": 0, "gl_totals": {}}
                try:
                    unpaid_result = oa_helpers["apply_unpaid_bills"](target_entity)
                    logger.info(f"Applied unpaid bills: {unpaid_result['applied']} GL lines for entity {target_entity}")
                except Exception as unpaid_err:
                    logger.error(f"CRITICAL: apply_unpaid_bills failed for {target_entity}: {unpaid_err}")
                    results["warnings"].append(f"Unpaid bills FAILED for {target_entity}: {str(unpaid_err)}")

                unpaid_msg = f", {unpaid_result['applied']} GL lines updated" if unpaid_result["applied"] > 0 else ""
                results["success"].append(f"Open AP: {target_entity} ({len(ap_invoices)} invoices, ${report.total_amount:,.2f}{unpaid_msg})")
            except Exception as ap_err:
                results["warnings"].append(f"Open AP file {ap_filename} error: {str(ap_err)}")

        # Pass 4: Maintenance Proof
        for mp_path, mp_filename in maint_proof_files:
            try:
                report_title, units, total_shares = mp_helpers["parse_maintenance_proof"](str(mp_path))
                if not units:
                    results["warnings"].append(f"Maintenance Proof file {mp_filename}: no units found")
                    continue

                import re as _re
                fname_match = _re.search(r'(?:Adhoc_AMP|MaintProof|MP)[_-]?(\d+)', mp_filename, _re.IGNORECASE)
                fname_entity = fname_match.group(1) if fname_match else None

                target_entity = fname_entity
                if not target_entity and ysl_entities:
                    target_entity = next(iter(ysl_entities))

                if not target_entity:
                    results["warnings"].append(f"Maintenance Proof file {mp_filename}: could not determine entity code")
                    continue

                report = mp_helpers["store_maintenance_proof"](target_entity, report_title, units, total_shares, mp_filename)
                results["success"].append(f"Maintenance Proof: {target_entity} ({len(units)} units, {report_title})")
            except Exception as mp_err:
                results["warnings"].append(f"Maintenance Proof file {mp_filename} error: {str(mp_err)}")

        return jsonify({"message": f"Auto-processed {len(file_list)} files", **results}), 200


@app.route("/api/generate-script", methods=["POST", "OPTIONS"])
@cors_headers
def generate_script():
    if request.method == "OPTIONS":
        return make_response(""), 200
    """Generate a customized Console script for selected buildings.

    Combines both the YSL Annual Budget and Expense Distribution scripts
    into a single script that downloads both reports for each entity.
    """
    data = request.json
    entities = data.get("entities", [])
    email = data.get("email", "")
    period = data.get("period", "02/2026")

    if not entities:
        return jsonify({"error": "No buildings selected"}), 400

    entities_js = ', '.join(str(e) for e in entities)

    # Build YSL script with user settings
    ysl_script = CONSOLE_SCRIPT
    ysl_script = ysl_script.replace(
        "const AS_OF_PERIOD = '02/2026';",
        f"const AS_OF_PERIOD = '{period}';"
    )
    ysl_script = ysl_script.replace(
        "const ENTITIES = [148, 204, 206, 805];",
        f"const ENTITIES = [{entities_js}];"
    )
    ysl_script = ysl_script.replace(
        "const EMAIL = 'JSirotkin@Centuryny.com';",
        f"const EMAIL = '{email}';"
    )

    # Build AP Aging script with user settings
    ap_script = ""
    if AP_AGING_SCRIPT:
        ap_script = AP_AGING_SCRIPT
        ap_script = ap_script.replace(
            "const ENTITIES = [148, 204, 206, 805];",
            f"const ENTITIES = [{entities_js}];"
        )
        ap_script = ap_script.replace(
            "const PERIOD_TO = '03/2026';",
            f"const PERIOD_TO = '{period}';"
        )
        # Inject auto-upload into AP Aging triggerDownload
        ap_script = ap_script.replace(
            "    URL.revokeObjectURL(a.href);\n  }",
            "    URL.revokeObjectURL(a.href);\n    if (typeof _autoUpload === 'function') _uploadPromises.push(_autoUpload(blob, a.download, entity, 'ap'));\n  }"
        )

    # Build Maintenance Proof script with user settings
    # Map entity → charge code based on building type from CSV
    building_charges = {}
    try:
        with open(BUILDINGS_CSV, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                ec = str(row.get("entity_code", "")).strip()
                btype = (row.get("type", "") or "").strip().lower()
                if btype in ("coop", "cond-op"):
                    building_charges[ec] = "maint"
                elif btype == "condo":
                    building_charges[ec] = "common"
                # Rental, Mitchell Lama, Comm/Retail → skip (no maint proof)
    except Exception:
        pass

    # Only include entities that need a maintenance proof
    mp_entities = [e for e in entities if str(e) in building_charges]
    mp_mapping_js = ", ".join(f"{e}: '{building_charges[str(e)]}'" for e in mp_entities)

    mp_script = MAINT_PROOF_SCRIPT
    mp_script = mp_script.replace(
        "const ENTITY_CHARGES = {204: 'maint', 206: 'common', 148: 'maint', 805: 'maint'};",
        f"const ENTITY_CHARGES = {{{mp_mapping_js}}};"
    )

    # AP Aging is back in the combined script now that Expense Distribution
    # runs separately. No RT contamination risk.
    if ap_script:
        ap_script_block = f"""
  // ── Part 3: AP Aging ──
  console.log('\\n>>> Starting Part 3: AP Aging <<<\\n');
  try {{
    _partResults.ap = await {ap_script}
    console.log('\\n>>> Part 3 (AP Aging) completed successfully <<<\\n');
  }} catch (_e3) {{
    console.error('>>> Part 3 (AP Aging) FAILED with error:', _e3.message);
    console.error(_e3.stack);
    _partResults.ap = 'ERROR: ' + _e3.message;
  }}"""
    else:
        ap_script_block = ""

    # ── Inject auto-upload into each script's triggerDownload function ──
    # The 3 scripts (Expense, MaintProof, AP Aging) all have a triggerDownload(blob, entity)
    # function. We inject a call to the global _autoUpload helper after the local download.
    # YSL has inline download code — we handle that separately.

    # Expense Distribution patching removed — runs as separate standalone script

    # Patch Maintenance Proof triggerDownload
    mp_script = mp_script.replace(
        "    URL.revokeObjectURL(a.href);\n  }",
        "    URL.revokeObjectURL(a.href);\n    if (typeof _autoUpload === 'function') _uploadPromises.push(_autoUpload(blob, a.download, entity, 'maint'));\n  }"
    )

    # AP Aging patching removed — runs as separate standalone script

    # Patch YSL inline download (it doesn't use triggerDownload function)
    ysl_script = ysl_script.replace(
        "URL.revokeObjectURL(a.href);",
        "URL.revokeObjectURL(a.href);\n          if (typeof _autoUpload === 'function') _uploadPromises.push(_autoUpload(blob, a.download, entity, 'ysl'));"
    )

    # Get the Railway app URL for auto-upload target
    railway_url = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
    if railway_url and not railway_url.startswith("http"):
        railway_url = f"https://{railway_url}"
    if not railway_url:
        # Fallback to known production URL
        railway_url = "https://century-budget-generator-production.up.railway.app"

    # Combine into one script that runs all four sequentially
    # Each part is wrapped in try/catch so errors don't kill subsequent parts
    combined = f"""/**
 * Century Budget — Combined Yardi Download Script
 * Downloads YSL Annual Budget + Maintenance Proof + AP Aging for all selected entities.
 * Expense Distribution runs separately (own button).
 * Generated for: {email}
 * Entities: {entities_js}
 * Period: {period}
 *
 * AUTO-UPLOAD: Files are automatically sent to the budget app for processing.
 * Local downloads are also saved as backup.
 */
(async function() {{
  'use strict';
  const _partResults = {{ysl: null, mp: null, ap: null}};
  const _uploadedFiles = [];
  const _uploadPromises = [];
  const _SESSION_ID = 'yardi_' + Date.now();
  const _BUDGET_APP = '{railway_url}';
  const _PERIOD = '{period}';
  const _FRESH_START = {str(data.get('fresh_start', False)).lower()};

  // ── Auto-Upload Helper ──
  // Sends each downloaded blob to the budget app in the background
  async function _autoUpload(blob, filename, entity, fileType) {{
    try {{
      const formData = new FormData();
      formData.append('file', blob, filename);
      const resp = await fetch(_BUDGET_APP + '/api/auto-upload', {{
        method: 'POST',
        headers: {{
          'X-Upload-Session': _SESSION_ID,
          'X-Entity-Code': String(entity),
          'X-File-Type': fileType,
          'X-Filename': filename,
        }},
        body: formData,
      }});
      const data = await resp.json();
      if (resp.ok) {{
        _uploadedFiles.push(filename);
        console.log('  ↑ Auto-uploaded: ' + filename + ' (' + data.total_files + ' files in session)');
      }} else {{
        console.warn('  ↑ Upload failed for ' + filename + ': ' + (data.error || resp.status));
      }}
    }} catch (err) {{
      console.warn('  ↑ Upload error for ' + filename + ': ' + err.message);
    }}
  }}

  // ── Auto-Process: triggers server-side processing after all downloads ──
  async function _autoProcess() {{
    // Wait for all in-flight uploads to finish before processing
    if (_uploadPromises.length > 0) {{
      console.log('\\n>>> Waiting for ' + _uploadPromises.length + ' uploads to complete... <<<');
      await Promise.allSettled(_uploadPromises);
    }}
    if (_uploadedFiles.length === 0) {{
      console.log('No files uploaded — skipping auto-process.');
      return;
    }}
    try {{
      console.log('\\n>>> Auto-processing ' + _uploadedFiles.length + ' files on server... <<<');
      const resp = await fetch(_BUDGET_APP + '/api/auto-process', {{
        method: 'POST',
        headers: {{ 'Content-Type': 'application/json' }},
        body: JSON.stringify({{ session_id: _SESSION_ID, period: _PERIOD, fresh_start: _FRESH_START }}),
      }});
      const data = await resp.json();
      if (resp.ok) {{
        console.log('✓ Server processed ' + _uploadedFiles.length + ' files successfully!');
        if (data.success) console.log('  Successes:', data.success);
        if (data.warnings && data.warnings.length) console.log('  Warnings:', data.warnings);
        if (data.failed && data.failed.length) console.log('  Failures:', data.failed);
        // Show a Done banner with link to FA Dashboard
        const banner = document.createElement('div');
        banner.style.cssText = 'position:fixed;top:20px;right:20px;z-index:99999;background:#065f46;color:white;padding:16px 24px;border-radius:12px;font-family:system-ui;font-size:14px;box-shadow:0 8px 24px rgba(0,0,0,0.3);display:flex;flex-direction:column;gap:10px;max-width:360px;';
        banner.innerHTML = '<div style="font-weight:700;font-size:16px;">✓ Budget data updated</div>'
          + '<div style="font-size:13px;opacity:0.9;">' + _uploadedFiles.length + ' files processed'
          + (data.warnings && data.warnings.length ? ' · ' + data.warnings.length + ' warning(s)' : '')
          + '</div>'
          + '<a href="' + _BUDGET_APP + '/dashboard" target="_blank" style="display:inline-block;background:white;color:#065f46;padding:8px 20px;border-radius:8px;text-decoration:none;font-weight:600;font-size:14px;text-align:center;margin-top:4px;">Open FA Dashboard →</a>';
        document.body.appendChild(banner);
        // Auto-dismiss after 30s
        setTimeout(() => banner.remove(), 30000);
      }} else {{
        console.error('✗ Server processing failed:', data.error || resp.status);
      }}
    }} catch (err) {{
      console.error('✗ Auto-process error:', err.message);
      console.log('Files were downloaded locally — you can upload them manually via the Generator page.');
    }}
  }}

  console.log('='.repeat(60));
  console.log('Century Budget — Combined Download + Auto-Upload');
  console.log('Target: ' + _BUDGET_APP);
  console.log('Session: ' + _SESSION_ID);
  console.log('Part 1: YSL Annual Budget reports');
  console.log('Part 2: Maintenance Proof reports');
  console.log('Part 3: AP Aging reports');
  console.log('(Expense Distribution runs separately — use its button)');
  console.log('='.repeat(60));

  // ── Part 1: YSL Annual Budget ──
  console.log('\\n>>> Starting Part 1: YSL Annual Budget <<<\\n');
  try {{
    _partResults.ysl = await {ysl_script}
    console.log('\\n>>> Part 1 (YSL) completed successfully <<<\\n');
  }} catch (_e1) {{
    console.error('>>> Part 1 (YSL) FAILED with error:', _e1.message);
    console.error(_e1.stack);
    _partResults.ysl = 'ERROR: ' + _e1.message;
  }}

  // ── Part 2: Maintenance Proof ({len(mp_entities)} of {len(entities)} buildings — coops/condos only) ──
  console.log('\\n>>> Starting Part 2: Maintenance Proof ({len(mp_entities)} of {len(entities)} buildings — coops/condos only) <<<\\n');
  try {{
    _partResults.mp = await {mp_script}
    console.log('\\n>>> Part 2 (Maintenance Proof) completed successfully <<<\\n');
  }} catch (_e2) {{
    console.error('>>> Part 2 (Maintenance Proof) FAILED with error:', _e2.message);
    console.error(_e2.stack);
    _partResults.mp = 'ERROR: ' + _e2.message;
  }}
{ap_script_block}

  // ── Auto-Process all uploaded files ──
  await _autoProcess();

  console.log('\\n' + '='.repeat(60));
  console.log('ALL DONE — Summary:');
  console.log('  Part 1 (YSL):', _partResults.ysl === null ? 'SKIPPED' : (typeof _partResults.ysl === 'string' && _partResults.ysl.startsWith('ERROR') ? _partResults.ysl : 'OK'));
  console.log('  Part 2 (MP): ', _partResults.mp === null ? 'SKIPPED' : (typeof _partResults.mp === 'string' && _partResults.mp.startsWith('ERROR') ? _partResults.mp : 'OK'));
  console.log('  Part 3 (AP): ', _partResults.ap === null ? 'SKIPPED' : (typeof _partResults.ap === 'string' && _partResults.ap.startsWith('ERROR') ? _partResults.ap : 'OK'));
  console.log('  Auto-Upload:', _uploadedFiles.length + ' files sent to server');
  console.log(_uploadedFiles.length > 0 ? 'Files were auto-processed — check the dashboard!' : 'Files were downloaded locally — upload them via the Generator page.');
  console.log('');
  console.log('Now run the Expense Distribution script separately (use its button on the Generate page).');
  console.log('='.repeat(60));
}})();"""

    return jsonify({"script": combined})



@app.route("/api/generate-ap-aging-script", methods=["POST"])
def generate_ap_aging_script():
    """Generate a standalone AP Aging script that runs independently.

    Separated from the combined script because Yardi's ASP.NET session
    remembers ReportType=2 (Expense Distribution) and contaminates the
    AP Aging download when both run in the same browser session.
    Running AP Aging standalone avoids this session state issue.
    """
    data = request.json
    entities = data.get("entities", [])
    email = data.get("email", "")
    period = data.get("period", "02/2026")

    if not entities:
        return jsonify({"error": "No buildings selected"}), 400

    if not AP_AGING_SCRIPT:
        return jsonify({"error": "AP Aging script not available on server"}), 500

    entities_js = ', '.join(str(e) for e in entities)

    ap_aging_script = AP_AGING_SCRIPT
    ap_aging_script = ap_aging_script.replace(
        "const ENTITIES = [148, 204, 206, 805];",
        f"const ENTITIES = [{entities_js}];"
    )
    ap_aging_script = ap_aging_script.replace(
        "const PERIOD_TO = '03/2026';",
        f"const PERIOD_TO = '{period}';"
    )

    # Get the Railway app URL for auto-upload target
    railway_url = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
    if railway_url and not railway_url.startswith("http"):
        railway_url = f"https://{railway_url}"
    if not railway_url:
        railway_url = "https://century-budget-generator-production.up.railway.app"

    # Inject auto-upload into triggerDownload
    ap_aging_script = ap_aging_script.replace(
        "URL.revokeObjectURL(a.href);\n  }",
        "URL.revokeObjectURL(a.href);\n    if (typeof _autoUpload === 'function') _autoUpload(blob, a.download, entity, 'ap');\n  }"
    )

    # Wrap with auto-upload helper and auto-process
    standalone = f"""/**
 * Century Budget — AP Aging (Open AP) Standalone Script
 * Run this AFTER the main Yardi script (YSL + Expense + Maint Proof).
 * Entities: {entities_js}
 * Period: {period}
 */
(async function() {{
  'use strict';
  const _uploadedFiles = [];
  const _SESSION_ID = 'yardi_' + Date.now();
  const _BUDGET_APP = '{railway_url}';

  async function _autoUpload(blob, filename, entity, fileType) {{
    try {{
      const resp = await fetch(_BUDGET_APP + '/api/auto-upload', {{
        method: 'POST',
        headers: {{
          'X-Upload-Session': _SESSION_ID,
          'X-Entity-Code': String(entity),
          'X-File-Type': fileType,
          'X-Filename': filename,
          'Content-Type': 'application/octet-stream'
        }},
        body: blob
      }});
      const data = await resp.json();
      _uploadedFiles.push({{ filename, entity, fileType, ok: resp.ok }});
      console.log('  \\u2191 Auto-uploaded: ' + filename + ' (' + data.files_in_session + ' files in session)');
    }} catch (err) {{
      console.warn('  Auto-upload failed for ' + filename + ':', err.message);
    }}
  }}

  async function _autoProcess() {{
    if (!_uploadedFiles.length) return;
    console.log('\\n>>> Auto-processing ' + _uploadedFiles.length + ' files on server... <<<');
    try {{
      const resp = await fetch(_BUDGET_APP + '/api/auto-process', {{
        method: 'POST',
        headers: {{ 'Content-Type': 'application/json' }},
        body: JSON.stringify({{ session_id: _SESSION_ID }})
      }});
      const data = await resp.json();
      if (resp.ok) {{
        console.log('\\u2713 Server processed ' + (data.successes?.length || 0) + ' files successfully!');
        if (data.successes?.length) console.log('  Successes:', data.successes);
        if (data.warnings?.length) console.log('  Warnings:', data.warnings);
        if (data.failures?.length) console.log('  Failures:', data.failures);
      }}
    }} catch (err) {{
      console.error('Auto-process error:', err.message);
    }}
  }}

  console.log('='.repeat(60));
  console.log('AP Aging (Open AP) — Standalone Download + Auto-Upload');
  console.log('Target: ' + _BUDGET_APP);
  console.log('='.repeat(60));

  try {{
    await {ap_aging_script}
    console.log('\\n>>> AP Aging completed successfully <<<');
  }} catch (e) {{
    console.error('>>> AP Aging FAILED:', e.message);
  }}

  await _autoProcess();

  console.log('\\n' + '='.repeat(60));
  console.log('AP Aging DONE — ' + _uploadedFiles.length + ' files uploaded.');
  console.log(_uploadedFiles.length > 0 ? 'Check the FA Dashboard for results.' : 'No files uploaded — check console for errors.');
  console.log('='.repeat(60));
}})();"""

    return jsonify({"script": standalone})


@app.route("/api/process", methods=["POST"])
def process_files():
    """Accept YSL uploads, run pipeline, save + return budgets."""
    if "files" not in request.files:
        return jsonify({"error": "No files uploaded"}), 400

    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400

    # Get save location from form data or settings
    save_dir_str = request.form.get("save_dir", "").strip()
    if not save_dir_str:
        save_dir_str = load_settings().get("save_dir", DEFAULT_SAVE_DIR)
    save_dir = Path(save_dir_str)

    # Also archive YSLs to the central ysl_downloads folder
    ysl_archive = BUDGET_SYSTEM / "ysl_downloads"
    ysl_archive.mkdir(exist_ok=True)

    results = {"success": [], "failed": [], "warnings": [], "save_dir": str(save_dir)}

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        output_files = []

        # Multi-pass approach: separate file types so we process in correct order
        # Order: YSL first (creates BudgetLines), then Expense Dist, then Open AP
        saved_files = []  # (path, filename, file_type, row1, row2)
        ysl_entities = set()  # entity codes found in YSL files
        expense_files = []  # (path, filename) for expense files to process after YSL
        open_ap_files = []  # (path, filename) for AP Aging files to process after YSL
        maint_proof_files = []  # (path, filename) for Maintenance Proof files

        # Pass 1: Save all files, classify them, process YSL files first
        for f in files:
            if not f.filename or not f.filename.endswith((".xlsx", ".xls")):
                results["warnings"].append(f"Skipped non-Excel: {f.filename}")
                continue

            file_path = tmp / f.filename
            f.save(str(file_path))

            try:
                # Check if it's an Open AP (Aging) file first (uses robust detection)
                if detect_open_ap_file(str(file_path)):
                    logger.info(f"File detection for {f.filename}: Open AP (Aging) report")
                    open_ap_files.append((file_path, f.filename))
                    continue

                # Handle .xls files that are actually .xlsx (Yardi naming quirk)
                _check_path = file_path
                _tmp_xlsx_manual = None
                if str(file_path).lower().endswith('.xls') and not str(file_path).lower().endswith('.xlsx'):
                    import shutil as _shutil_m
                    _tmp_xlsx_manual = tmp / (f.filename + 'x')
                    _shutil_m.copy2(str(file_path), str(_tmp_xlsx_manual))
                    _check_path = _tmp_xlsx_manual

                from openpyxl import load_workbook as _lwb
                _wb_check = _lwb(str(_check_path), data_only=True)
                _ws_check = _wb_check.active
                _row1 = str(_ws_check.cell(row=1, column=1).value or "")
                _row2 = str(_ws_check.cell(row=2, column=1).value or "")
                _wb_check.close()

                is_expense = "expense distribution" in _row1.lower() or "expense dist" in _row1.lower()
                logger.info(f"File detection for {f.filename}: Row1='{_row1}', Row2='{_row2}', is_expense={is_expense}")

                if is_expense:
                    expense_files.append((file_path, f.filename))
                    continue

                is_maint = "maintenance proof" in _row1.lower() or "adhoc_amp" in _row1.lower()
                if is_maint:
                    logger.info(f"File detection for {f.filename}: Maintenance Proof report")
                    maint_proof_files.append((file_path, f.filename))
                    continue
            except Exception as detect_err:
                logger.error(f"File detection error for {f.filename}: {detect_err}")
                results["warnings"].append(f"Could not read {f.filename}: {str(detect_err)}")
                continue

            # Process YSL file immediately
            ysl_path = file_path
            try:
                # Parse YSL
                gl_data, property_info = parse_ysl_file(ysl_path)
                entity = property_info.get("property_code", "unknown")
                name = property_info.get("property_name", f"Entity_{entity}")
                ysl_entities.add(str(entity))

                # Build the building folder name (e.g., "148 - 130 E. 18 Owners Corp.")
                building_folder_name = f"{entity} - {name}"
                building_dir = save_dir / building_folder_name
                building_dir.mkdir(parents=True, exist_ok=True)

                # Generate budget into the building folder
                from datetime import datetime as _dt_now
                _ts = _dt_now.now().strftime("%Y%m%d_%H%M")
                output_name = f"{entity}_{name}_2027_Budget_{_ts}.xlsx"
                output_path = building_dir / output_name

                # Derive YTD months from period (e.g., "02/2026" → 2)
                _gen_ytd = 2
                try:
                    _gen_ytd = int(period.split("/")[0])
                except Exception:
                    pass

                success = populate_template(
                    template_path=TEMPLATE_PATH,
                    gl_data=gl_data,
                    property_info=property_info,
                    output_path=output_path,
                    ytd_months=_gen_ytd,
                    remaining_months=12 - _gen_ytd,
                )

                if success and output_path.exists():
                    # Merge assumptions for this building
                    merged = None
                    try:
                        merged = merge_assumptions(entity)
                    except Exception:
                        pass

                    # Store ALL GL data in database for budget review workflow
                    try:
                        fresh = request.form.get("fresh_start", "").lower() in ("true", "1", "yes")
                        workflow_helpers["store_all_lines"](entity, name, gl_data, TEMPLATE_PATH, assumptions=merged, fresh_start=fresh)
                        logger.info(f"All GL data stored for entity {entity} (fresh_start={fresh})")
                    except Exception as wfe:
                        logger.warning(f"Could not store GL data for {entity}: {wfe}")

                    # Apply assumptions to Excel file
                    try:
                        if merged:
                            apply_assumptions(output_path, merged)
                            logger.info(f"Assumptions applied for entity {entity}")
                    except Exception as ae:
                        logger.warning(f"Could not apply assumptions for {entity}: {ae}")

                    # Apply PM R&M projections if available
                    try:
                        pm_proj = workflow_helpers["get_pm_projections"](entity)
                        if pm_proj:
                            apply_pm_projections(output_path, pm_proj)
                            logger.info(f"PM projections applied for entity {entity}")
                    except Exception as pe:
                        logger.warning(f"Could not apply PM projections for {entity}: {pe}")

                    output_files.append((output_name, output_path))
                    size_kb = output_path.stat().st_size / 1024

                    # Archive YSL to central ysl_downloads/
                    shutil.copy2(ysl_path, ysl_archive / f.filename)

                    # Archive YSL to building's yardi_drops/
                    yardi_drops = building_dir / "yardi_drops"
                    yardi_drops.mkdir(exist_ok=True)
                    shutil.copy2(ysl_path, yardi_drops / f.filename)

                    results["success"].append({
                        "entity": entity,
                        "name": name,
                        "file": output_name,
                        "size_kb": round(size_kb),
                        "saved_to": str(building_dir),
                    })
                else:
                    results["failed"].append({
                        "entity": entity,
                        "file": f.filename,
                        "reason": "Pipeline returned failure",
                    })

            except Exception as e:
                logger.exception(f"Error processing {f.filename}")
                results["failed"].append({
                    "file": f.filename,
                    "reason": str(e),
                })

        # Pass 2: Process Expense Distribution files
        # Use entity code from filename (our script names them ExpenseDistribution_XXX.xlsx)
        # If that doesn't match the file contents, use the filename entity (Yardi bug workaround)
        for exp_path, exp_filename in expense_files:
            try:
                from expense_distribution import parse_expense_distribution
            except ImportError:
                from budget_app.expense_distribution import parse_expense_distribution
            try:
                exp_entity, exp_from, exp_to, exp_invoices = parse_expense_distribution(str(exp_path))
                logger.info(f"Expense parse: file={exp_filename}, file_entity={exp_entity}, invoices={len(exp_invoices) if exp_invoices else 0}")

                if not exp_invoices:
                    results["warnings"].append(f"Expense file {exp_filename}: no invoices found")
                    continue

                # Determine correct entity code:
                # 1. Try to extract from filename (ExpenseDistribution_204.xlsx → 204)
                # 2. If filename entity matches a YSL entity we just processed, use it
                # 3. Otherwise fall back to what's in the file
                import re as _re
                fname_match = _re.search(r'ExpenseDistribution[_-]?(\d+)', exp_filename)
                fname_entity = fname_match.group(1) if fname_match else None

                target_entity = exp_entity  # default: trust the file
                if fname_entity and fname_entity != exp_entity:
                    # Filename says one entity, file contents say another — Yardi bug
                    if fname_entity in ysl_entities:
                        # The filename matches a YSL we just processed, so trust the filename
                        target_entity = fname_entity
                        logger.warning(f"Yardi entity mismatch: file says {exp_entity}, filename says {fname_entity}. Using {fname_entity} (matches YSL upload)")
                        results["warnings"].append(f"Note: Yardi returned entity {exp_entity} data but file was requested for {fname_entity}. Stored under {fname_entity}.")
                    elif not ysl_entities:
                        # No YSL files uploaded — standalone expense upload, trust filename
                        target_entity = fname_entity
                        logger.warning(f"Yardi entity mismatch: file says {exp_entity}, filename says {fname_entity}. Using {fname_entity} (filename override)")
                    else:
                        # Filename doesn't match any YSL entity — log warning but store as-is
                        logger.warning(f"Yardi entity mismatch: file says {exp_entity}, filename says {fname_entity}. Neither matches YSL entities {ysl_entities}. Using file entity {exp_entity}.")

                report = ed_helpers["store_expense_report"](target_entity, exp_from, exp_to, exp_invoices, exp_filename)
                logger.info(f"Expense distribution stored for entity {target_entity}")

                # Apply accrual adjustments (prior-year invoices backed out of YTD)
                accrual_result = {"applied": 0, "accruals": {}}
                if exp_from:
                    try:
                        accrual_result = ed_helpers["apply_accrual_adjustments"](target_entity, report.id, exp_from)
                        if accrual_result["applied"] > 0:
                            logger.info(f"Auto-applied accrual adjustments to {accrual_result['applied']} GL lines for entity {target_entity}")
                    except Exception as accrual_err:
                        logger.error(f"Accrual adjustment failed for {target_entity}: {accrual_err}")
                        results["warnings"].append(f"Accrual adjustments failed for {target_entity}: {str(accrual_err)}")

                accrual_msg = f", {accrual_result['applied']} accrual adj" if accrual_result["applied"] > 0 else ""
                results["success"].append(f"Expense Distribution: {target_entity} ({len(exp_invoices)} invoices{accrual_msg})")
            except Exception as exp_err:
                logger.error(f"Expense parse error for {exp_filename}: {exp_err}")
                results["warnings"].append(f"Expense file {exp_filename} error: {str(exp_err)}")

        # Pass 3: Process Open AP (Aging) files
        # These provide unpaid invoice data that populates BudgetLine.unpaid_bills
        for ap_path, ap_filename in open_ap_files:
            try:
                ap_entity, ap_invoices = parse_open_ap_report(str(ap_path))
                logger.info(f"Open AP parse: file={ap_filename}, entity={ap_entity}, invoices={len(ap_invoices) if ap_invoices else 0}")

                if not ap_invoices:
                    results["warnings"].append(f"Open AP file {ap_filename}: no invoices found")
                    continue

                # Determine correct entity code
                # 1. Try to extract from filename (e.g., APAging_204.xlsx or OpenAP_204.xlsx)
                import re as _re
                fname_match = _re.search(r'(?:Aging|OpenAP|AP)[_-]?(\d+)', ap_filename, _re.IGNORECASE)
                fname_entity = fname_match.group(1) if fname_match else None

                target_entity = ap_entity  # default: from file data
                if fname_entity and fname_entity != ap_entity:
                    if fname_entity in ysl_entities:
                        target_entity = fname_entity
                        logger.warning(f"Open AP entity mismatch: data says {ap_entity}, filename says {fname_entity}. Using {fname_entity}")
                    elif not ysl_entities:
                        target_entity = fname_entity
                elif not target_entity and fname_entity:
                    target_entity = fname_entity

                if not target_entity:
                    results["warnings"].append(f"Open AP file {ap_filename}: could not determine entity code")
                    continue

                # Store parsed invoices
                report = oa_helpers["store_open_ap_report"](target_entity, ap_invoices, ap_filename)
                logger.info(f"Open AP stored for entity {target_entity}: {len(ap_invoices)} invoices, ${report.total_amount:,.2f}")

                # Auto-apply unpaid bills to BudgetLines
                unpaid_result = {"applied": 0, "gl_totals": {}}
                try:
                    unpaid_result = oa_helpers["apply_unpaid_bills"](target_entity)
                    if unpaid_result["applied"] > 0:
                        logger.info(f"Auto-applied unpaid bills to {unpaid_result['applied']} GL lines for entity {target_entity}")
                except Exception as unpaid_err:
                    logger.error(f"Unpaid bills application failed for {target_entity}: {unpaid_err}")
                    results["warnings"].append(f"Unpaid bills failed for {target_entity}: {str(unpaid_err)}")

                unpaid_msg = f", {unpaid_result['applied']} GL lines updated" if unpaid_result["applied"] > 0 else ""
                results["success"].append(f"Open AP: {target_entity} ({len(ap_invoices)} invoices, ${report.total_amount:,.2f}{unpaid_msg})")
            except Exception as ap_err:
                logger.error(f"Open AP parse error for {ap_filename}: {ap_err}")
                results["warnings"].append(f"Open AP file {ap_filename} error: {str(ap_err)}")

        # Pass 4: Process Maintenance Proof files
        for mp_path, mp_filename in maint_proof_files:
            try:
                report_title, units, total_shares = mp_helpers["parse_maintenance_proof"](str(mp_path))
                if not units:
                    results["warnings"].append(f"Maintenance Proof file {mp_filename}: no units found")
                    continue

                import re as _re
                fname_match = _re.search(r'(?:Adhoc_AMP|MaintProof|MP)[_-]?(\d+)', mp_filename, _re.IGNORECASE)
                fname_entity = fname_match.group(1) if fname_match else None

                target_entity = fname_entity
                if not target_entity and ysl_entities:
                    target_entity = next(iter(ysl_entities))

                if not target_entity:
                    results["warnings"].append(f"Maintenance Proof file {mp_filename}: could not determine entity code")
                    continue

                report = mp_helpers["store_maintenance_proof"](target_entity, report_title, units, total_shares, mp_filename)
                results["success"].append(f"Maintenance Proof: {target_entity} ({len(units)} units, {report_title})")
                logger.info(f"Maintenance proof stored for entity {target_entity}")
            except Exception as mp_err:
                logger.error(f"Maintenance proof parse error for {mp_filename}: {mp_err}")
                results["warnings"].append(f"Maintenance Proof file {mp_filename} error: {str(mp_err)}")

        if not output_files:
            # If expense/AP/maint files were processed successfully, return 200
            if results["success"]:
                return jsonify({"message": "Files processed", **results}), 200
            return jsonify({"error": "No budgets generated", **results}), 400

        # Return JSON with results — files are already saved to disk
        return jsonify({"message": f"{len(output_files)} budget(s) generated", **results}), 200


@app.route("/api/defaults", methods=["GET", "POST"])
def api_defaults():
    """Get or save portfolio defaults."""
    if request.method == "GET":
        return jsonify(load_portfolio_defaults())

    data = request.json
    save_portfolio_defaults(data)
    return jsonify({"status": "saved"})


@app.route("/api/building-assumptions/<entity_code>", methods=["GET", "POST"])
def api_building_assumptions(entity_code):
    """Get or save building-specific assumptions."""
    if request.method == "GET":
        assumptions = load_building_assumptions()
        return jsonify(assumptions.get(entity_code, {}))

    data = request.json
    assumptions = load_building_assumptions()
    assumptions[entity_code] = data
    save_building_assumptions(assumptions)
    return jsonify({"status": "saved"})


@app.route("/api/assumptions/<entity_code>")
def merged_assumptions(entity_code):
    """Return merged assumptions (portfolio + building overrides)."""
    return jsonify(merge_assumptions(entity_code))


@app.route("/api/export-assumptions")
def bulk_export():
    """Export all building assumptions."""
    buildings = {b["entity_code"]: merge_assumptions(b["entity_code"]) for b in load_buildings()}
    return jsonify(buildings)


# ─── Monday.com Sync ─────────────────────────────────────────────────────────

MONDAY_API_TOKEN = os.environ.get("MONDAY_API_TOKEN", "")
MONDAY_BOARD_ID = "3473581362"  # Building Master List


# ─── Monday.com sync helpers ─────────────────────────────────────────────
# In-memory staleness signal. Resets on every Railway redeploy — by design,
# so a deploy auto-refreshes Monday data on the first wizard hit.
_LAST_MONDAY_SYNC_AT = None     # datetime of last successful sync
_LAST_MONDAY_SYNC_ERROR = None  # str of last sync error, if any
MONDAY_STALE_MINUTES = 10       # auto-resync threshold


def _fetch_monday_buildings():
    """Fetch buildings + PM/FA assignments from Monday.com Building Master List.

    Returns: list[dict] with keys entity_code, building_name, address, city,
             zip, type, units, pm, fa.
    Raises: RuntimeError on missing token or API failure.
    """
    if not MONDAY_API_TOKEN:
        raise RuntimeError("MONDAY_API_TOKEN not configured")

    import urllib.request
    import ssl

    query = """{
      boards(ids: %s) {
        groups { id title }
        items_page(limit: 500) {
          items {
            name
            group { id title }
            column_values(ids: ["numeric", "numbers9", "pm8", "people", "text8", "text7", "text03", "status1"]) {
              id
              text
            }
          }
        }
      }
    }""" % MONDAY_BOARD_ID

    req = urllib.request.Request(
        "https://api.monday.com/v2",
        data=json.dumps({"query": query}).encode("utf-8"),
        headers={
            "Authorization": MONDAY_API_TOKEN,
            "Content-Type": "application/json",
            "API-Version": "2024-10",
        },
    )

    ctx = ssl.create_default_context()
    if not IS_CLOUD:
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

    try:
        with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
            result = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        logger.error(f"Monday.com API error: {e}")
        raise RuntimeError(f"Monday.com API error: {str(e)}")

    board0 = result.get("data", {}).get("boards", [{}])[0]
    items = board0.get("items_page", {}).get("items", [])

    # Only include items in the "Active Buildings (non-Lemle)" group.
    # Match is case-insensitive on the trimmed title for resilience to capitalization changes.
    ACTIVE_GROUP = "active buildings (non-lemle)"
    all_titles = [g.get("title") for g in (board0.get("groups") or [])]
    matched = [g for g in (board0.get("groups") or [])
               if (g.get("title") or "").strip().lower() == ACTIVE_GROUP]
    if not matched:
        logger.warning(
            f"Monday group '{ACTIVE_GROUP}' not found on board {MONDAY_BOARD_ID}. "
            f"Available groups: {all_titles}. Returning ALL items as fallback."
        )
    else:
        active_id = matched[0].get("id")
        items = [it for it in items if (it.get("group") or {}).get("id") == active_id]

    buildings = []
    for item in items:
        cols = {c["id"]: c["text"] for c in item.get("column_values", [])}
        entity = cols.get("numeric", "").replace(" ", "").split(".")[0]
        if not entity:
            continue
        units = cols.get("numbers9", "").replace(" ", "").split(".")[0]
        buildings.append({
            "entity_code": entity,
            "building_name": item["name"],
            "client": item["name"],
            "address": cols.get("text8") or "",
            "city": cols.get("text7") or "",
            "zip": cols.get("text03") or "",
            "type": cols.get("status1") or "",
            "units": units,
            "pm": cols.get("pm8") or None,
            "fa": cols.get("people") or None,
        })

    return buildings


def _apply_monday_sync(data):
    """Apply a buildings list to the DB — upsert Users + BuildingAssignments,
    and refresh buildings.csv. Commits the session.

    Args:
        data: list[dict] (same shape returned by _fetch_monday_buildings)
    Returns:
        dict of sync stats.
    """
    User = workflow_models["User"]
    BuildingAssignment = workflow_models["BuildingAssignment"]

    if not isinstance(data, list):
        raise ValueError("_apply_monday_sync expects a list of building dicts")

    stats = {"users_created": 0, "users_updated": 0, "assignments_created": 0,
             "assignments_removed": 0, "buildings_synced": 0}

    # Collect all unique people names from the data
    people = {}  # name -> set of roles
    for bldg in data:
        pm = (bldg.get("pm") or "").strip()
        fa = (bldg.get("fa") or "").strip()
        if pm:
            people.setdefault(pm, set()).add("pm")
        if fa:
            people.setdefault(fa, set()).add("fa")

    # Ensure all people exist as users
    for name, roles in people.items():
        primary_role = "pm" if "pm" in roles else "fa"

        user = User.query.filter_by(name=name).first()
        if not user:
            parts = name.lower().split()
            email_slug = parts[0][0] + parts[-1] if len(parts) > 1 else parts[0]
            email = f"{email_slug}@centuryny.com"

            existing = User.query.filter_by(email=email).first()
            if existing:
                if existing.name == name:
                    user = existing
                else:
                    email = f"{parts[0]}.{parts[-1]}@centuryny.com" if len(parts) > 1 else f"{parts[0]}2@centuryny.com"
                    user = User(name=name, email=email, role=primary_role)
                    db.session.add(user)
                    stats["users_created"] += 1
            else:
                user = User(name=name, email=email, role=primary_role)
                db.session.add(user)
                stats["users_created"] += 1

        db.session.flush()

    # Sync building assignments
    for bldg in data:
        entity_code = str(bldg.get("entity_code", "")).strip()
        if not entity_code:
            continue

        pm_name = (bldg.get("pm") or "").strip()
        fa_name = (bldg.get("fa") or "").strip()

        existing_assignments = BuildingAssignment.query.filter_by(entity_code=entity_code).all()
        for a in existing_assignments:
            should_keep = False
            if a.role == "pm" and a.user and a.user.name == pm_name:
                should_keep = True
            if a.role == "fa" and a.user and a.user.name == fa_name:
                should_keep = True
            if not should_keep:
                db.session.delete(a)
                stats["assignments_removed"] += 1

        db.session.flush()

        if pm_name:
            pm_user = User.query.filter_by(name=pm_name).first()
            if pm_user:
                existing = BuildingAssignment.query.filter_by(
                    entity_code=entity_code, user_id=pm_user.id, role="pm"
                ).first()
                if not existing:
                    db.session.add(BuildingAssignment(
                        entity_code=entity_code, user_id=pm_user.id, role="pm"
                    ))
                    stats["assignments_created"] += 1

        if fa_name:
            fa_user = User.query.filter_by(name=fa_name).first()
            if fa_user:
                existing = BuildingAssignment.query.filter_by(
                    entity_code=entity_code, user_id=fa_user.id, role="fa"
                ).first()
                if not existing:
                    db.session.add(BuildingAssignment(
                        entity_code=entity_code, user_id=fa_user.id, role="fa"
                    ))
                    stats["assignments_created"] += 1

        stats["buildings_synced"] += 1

    db.session.commit()

    # Also sync buildings.csv from the same data
    csv_buildings = []
    for bldg in data:
        entity_code = str(bldg.get("entity_code", "")).strip()
        if not entity_code:
            continue
        csv_buildings.append({
            "entity_code": entity_code,
            "building_name": bldg.get("building_name", ""),
            "address": bldg.get("address", ""),
            "city": bldg.get("city", ""),
            "zip": bldg.get("zip", ""),
            "type": bldg.get("type", ""),
            "units": bldg.get("units", ""),
        })
    if csv_buildings:
        save_buildings(csv_buildings)
        stats["buildings_csv_updated"] = len(csv_buildings)

    # Ensure a Budget row exists for every Monday-active building for the
    # current budget cycle. Existing rows are left untouched (no overwrite).
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY
    existing_codes = {
        row[0]
        for row in db.session.query(Budget.entity_code).filter_by(year=_BY).all()
    }
    stats["budgets_created"] = 0
    for bldg in data:
        entity_code = str(bldg.get("entity_code", "")).strip()
        if not entity_code or entity_code in existing_codes:
            continue
        bname = (bldg.get("building_name") or entity_code).strip()
        btype = (bldg.get("type") or "").strip().lower()
        db.session.add(Budget(
            entity_code=entity_code,
            building_name=bname,
            year=_BY,
            status="not_started",
            wizard_step=0,
            building_type=btype,
        ))
        existing_codes.add(entity_code)
        stats["budgets_created"] += 1
    if stats["budgets_created"]:
        db.session.commit()

    return stats


def _get_monday_status():
    """Read-only snapshot of last sync state. Does not touch the DB or network.
    Safe to call from any request handler regardless of blueprint/app context.
    """
    return {
        "last_synced_at": _LAST_MONDAY_SYNC_AT.isoformat() if _LAST_MONDAY_SYNC_AT else None,
        "error": _LAST_MONDAY_SYNC_ERROR,
        "stale_minutes": MONDAY_STALE_MINUTES,
    }


def _ensure_monday_fresh(stale_minutes=MONDAY_STALE_MINUTES, force=False):
    """Auto-sync from Monday.com if cached data is older than stale_minutes.

    Never raises — on failure, returns prior sync state with `error` set.
    Safe to call from page-load handlers.

    Args:
        stale_minutes: TTL threshold; if last sync was within this window, skip.
        force: if True, always sync regardless of TTL.
    Returns:
        dict: {synced, last_synced_at (ISO|None), error (str|None), stats|None}
    """
    from datetime import datetime as _dt, timedelta as _td
    global _LAST_MONDAY_SYNC_AT, _LAST_MONDAY_SYNC_ERROR
    now = _dt.utcnow()
    is_stale = (
        force
        or _LAST_MONDAY_SYNC_AT is None
        or (now - _LAST_MONDAY_SYNC_AT) > _td(minutes=stale_minutes)
    )
    if not is_stale:
        return {
            "synced": False,
            "last_synced_at": _LAST_MONDAY_SYNC_AT.isoformat() if _LAST_MONDAY_SYNC_AT else None,
            "error": _LAST_MONDAY_SYNC_ERROR,
            "stats": None,
        }
    try:
        buildings = _fetch_monday_buildings()
        stats = _apply_monday_sync(buildings)
        _LAST_MONDAY_SYNC_AT = _dt.utcnow()
        _LAST_MONDAY_SYNC_ERROR = None
        logger.info(f"Monday.com auto-sync complete: {stats}")
        return {
            "synced": True,
            "last_synced_at": _LAST_MONDAY_SYNC_AT.isoformat(),
            "error": None,
            "stats": stats,
        }
    except Exception as e:
        _LAST_MONDAY_SYNC_ERROR = str(e)
        logger.warning(f"Monday.com auto-sync failed (using stale data): {e}")
        try:
            db.session.rollback()
        except Exception:
            pass
        return {
            "synced": False,
            "last_synced_at": _LAST_MONDAY_SYNC_AT.isoformat() if _LAST_MONDAY_SYNC_AT else None,
            "error": _LAST_MONDAY_SYNC_ERROR,
            "stats": None,
        }


# ─── Monday.com sync routes ──────────────────────────────────────────────

@app.route("/api/sync-monday-fetch", methods=["GET"])
def sync_monday_fetch():
    """Fetch buildings + PM/FA assignments from Monday.com (read-only)."""
    try:
        buildings = _fetch_monday_buildings()
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"buildings": buildings, "count": len(buildings)})


@app.route("/api/sync-monday", methods=["POST"])
def sync_monday():
    """Apply a client-supplied buildings list to the DB.

    Expects JSON array of buildings (same shape as /api/sync-monday-fetch returns).
    """
    from datetime import datetime as _dt
    global _LAST_MONDAY_SYNC_AT, _LAST_MONDAY_SYNC_ERROR
    data = request.get_json()
    if not data or not isinstance(data, list):
        return jsonify({"error": "Expected JSON array of buildings"}), 400
    try:
        stats = _apply_monday_sync(data)
        _LAST_MONDAY_SYNC_AT = _dt.utcnow()
        _LAST_MONDAY_SYNC_ERROR = None
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        _LAST_MONDAY_SYNC_ERROR = str(e)
        logger.error(f"Monday.com sync failed: {e}")
        return jsonify({"error": str(e)}), 500
    logger.info(f"Monday.com sync complete: {stats}")
    return jsonify({
        "status": "ok",
        "stats": stats,
        "last_synced_at": _LAST_MONDAY_SYNC_AT.isoformat(),
    })


@app.route("/api/sync-monday-now", methods=["POST"])
def sync_monday_now():
    """One-call fetch + apply. Used by wizard's Refresh button.

    Always forces a sync regardless of staleness.
    """
    result = _ensure_monday_fresh(force=True)
    if result.get("error") and not result.get("synced"):
        return jsonify({"ok": False, **result}), 500
    return jsonify({"ok": True, **result})


# Cache the template GL list
_template_gls_cache = None
def _get_template_gls():
    global _template_gls_cache
    if _template_gls_cache is None:
        from gl_mapper import build_gl_mapping
        _template_gls_cache = set(build_gl_mapping(TEMPLATE_PATH).keys())
    return _template_gls_cache


# ─── HTML Templates ──────────────────────────────────────────────────────────

HOME_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<title>Century Management Budget System</title>
<style>
/* Force scrollbars always visible (fixes macOS auto-hide) */
::-webkit-scrollbar { width: 12px; height: 12px; -webkit-appearance: none; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 6px; border: 2px solid #f1f5f9; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }
::-webkit-scrollbar-corner { background: #f1f5f9; }
* { scrollbar-width: thin; scrollbar-color: #cbd5e1 #f1f5f9; }
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, sans-serif; background: #f5f3f1; color: #2d2520; min-height: 100vh; }

  /* Navbar */
  .navbar { position: sticky; top: 0; z-index: 100; background: #2d2520; padding: 0 40px; display: flex; align-items: center; justify-content: space-between; height: 54px; }
  .navbar .brand { font-size: 15px; font-weight: 700; color: #f5f3f1; letter-spacing: 0.5px; text-decoration: none; }
  .navbar .brand span { font-weight: 300; color: #c4b5a6; margin-left: 6px; }
  .nav-links { display: flex; gap: 4px; align-items: center; }
  .nav-links a { font-size: 12px; font-weight: 500; color: #c4b5a6; text-decoration: none; padding: 8px 14px; border-radius: 6px; transition: all 0.15s; }
  .nav-links a:hover { color: #f5f3f1; background: rgba(255,255,255,0.08); }
  .nav-links a.active { color: #f5f3f1; background: rgba(255,255,255,0.12); }

  /* Hero */
  .hero { background: linear-gradient(135deg, #3d342c 0%, #2d2520 50%, #1a1512 100%); padding: 48px 40px 40px; position: relative; overflow: hidden; }
  .hero::after { content: ''; position: absolute; top: -50%; right: -10%; width: 500px; height: 500px; background: radial-gradient(circle, rgba(196,181,166,0.06) 0%, transparent 70%); pointer-events: none; }
  .hero-inner { max-width: 1200px; margin: 0 auto; display: flex; justify-content: space-between; align-items: flex-end; }
  .hero h1 { font-size: 32px; font-weight: 300; color: #f5f3f1; line-height: 1.2; }
  .hero h1 strong { font-weight: 700; }
  .hero .subtitle { font-size: 13px; color: #a89888; margin-top: 8px; letter-spacing: 1px; text-transform: uppercase; font-weight: 500; }

  /* Role Toggle */
  .role-toggle { display: flex; background: rgba(255,255,255,0.08); border-radius: 8px; padding: 3px; gap: 2px; }
  .role-btn { padding: 8px 20px; font-size: 13px; font-weight: 600; border: none; border-radius: 6px; cursor: pointer; transition: all 0.2s; background: transparent; color: #a89888; font-family: inherit; }
  .role-btn.active { background: #f5f3f1; color: #2d2520; box-shadow: 0 1px 3px rgba(0,0,0,0.2); }
  .role-btn:not(.active):hover { color: #f5f3f1; }

  /* Main */
  .main { max-width: 1200px; margin: 0 auto; padding: 32px 40px 60px; }

  /* Section Labels */
  .section-label { font-size: 11px; font-weight: 700; color: #8b7b6b; text-transform: uppercase; letter-spacing: 1.5px; margin-bottom: 16px; display: flex; align-items: center; gap: 10px; }
  .section-label::after { content: ''; flex: 1; height: 1px; background: #ddd5cc; }

  /* Phase Stepper */
  .stepper { display: grid; grid-template-columns: repeat(3, 1fr); gap: 20px; margin-bottom: 40px; }
  .phase-card { background: #fff; border-radius: 12px; padding: 28px 24px; border: 1px solid #e8e0d8; position: relative; transition: all 0.2s; }
  .phase-card:hover { border-color: #c4b5a6; box-shadow: 0 4px 16px rgba(45,37,32,0.06); }
  .phase-num { display: inline-flex; align-items: center; justify-content: center; width: 28px; height: 28px; border-radius: 50%; background: #2d2520; color: #f5f3f1; font-size: 13px; font-weight: 700; margin-bottom: 14px; }
  .phase-card h3 { font-size: 17px; font-weight: 700; color: #2d2520; margin-bottom: 6px; }
  .phase-card .phase-desc { font-size: 13px; color: #8b7b6b; line-height: 1.5; margin-bottom: 18px; }
  .checklist { list-style: none; margin-bottom: 20px; }
  .checklist li { font-size: 13px; color: #4a3f35; padding: 7px 0; border-bottom: 1px solid #f5f0eb; display: flex; align-items: flex-start; gap: 10px; line-height: 1.4; }
  .checklist li:last-child { border-bottom: none; }
  .checklist li::before { content: '\25CB'; color: #c4b5a6; font-size: 14px; flex-shrink: 0; margin-top: 1px; }
  .phase-link { display: inline-flex; align-items: center; gap: 6px; font-size: 13px; font-weight: 600; color: #2d2520; text-decoration: none; padding: 8px 16px; border-radius: 8px; background: #f5f0eb; transition: all 0.15s; }
  .phase-link:hover { background: #2d2520; color: #f5f3f1; }
  .phase-link svg { width: 14px; height: 14px; }

  /* Quick Links */
  .quick-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 14px; margin-bottom: 40px; }
  .quick-card { background: #fff; border: 1px solid #e8e0d8; border-radius: 10px; padding: 20px; display: flex; align-items: flex-start; gap: 14px; text-decoration: none; color: inherit; transition: all 0.15s; }
  .quick-card:hover { border-color: #c4b5a6; box-shadow: 0 2px 8px rgba(45,37,32,0.05); transform: translateY(-1px); }
  .quick-icon { width: 40px; height: 40px; border-radius: 10px; display: flex; align-items: center; justify-content: center; font-size: 18px; flex-shrink: 0; }
  .quick-icon.blue { background: #eff6ff; }
  .quick-icon.green { background: #f0fdf4; }
  .quick-icon.amber { background: #fffbeb; }
  .quick-icon.purple { background: #faf5ff; }
  .quick-icon.slate { background: #f8fafc; }
  .quick-icon.rose { background: #fff1f2; }
  .quick-card h4 { font-size: 14px; font-weight: 600; color: #2d2520; margin-bottom: 3px; }
  .quick-card p { font-size: 12px; color: #8b7b6b; line-height: 1.4; }

  /* Footer */
  .footer { text-align: center; padding: 24px 40px; font-size: 11px; color: #a89888; letter-spacing: 0.5px; border-top: 1px solid #e8e0d8; }

  /* Responsive */
  @media (max-width: 900px) {
    .stepper { grid-template-columns: 1fr; }
    .quick-grid { grid-template-columns: repeat(2, 1fr); }
    .hero-inner { flex-direction: column; align-items: flex-start; gap: 16px; }
  }
  @media (max-width: 600px) {
    .quick-grid { grid-template-columns: 1fr; }
    .navbar { padding: 0 20px; }
    .hero { padding: 32px 20px; }
    .main { padding: 24px 20px 40px; }
  }

  /* Phase transition */
  .phase-card { opacity: 1; transition: opacity 0.25s ease, transform 0.25s ease; }
  .phase-card.fade-out { opacity: 0; transform: translateY(8px); }
</style>
</head>
<body>

<!-- Navbar -->
<nav class="navbar">
  <a href="/" class="brand">CENTURY MANAGEMENT <span>Budget System</span></a>
  <div class="nav-links">
    <a href="/" class="active">Home</a>
    <a href="/dashboard">FA Dashboard</a>
    <a href="/pm">PM Portal</a>
    <a href="/audited-financials">Audited Financials</a>
    <a href="/files">Files</a>
  </div>
</nav>

<!-- Hero -->
<div class="hero">
  <div class="hero-inner">
    <div>
      <h1>Budget & <strong>Assumptions</strong></h1>
      <div class="subtitle">2027 Budget Cycle</div>
    </div>
    <div class="role-toggle" id="roleToggle">
      <button class="role-btn active" data-role="fa" onclick="switchRole('fa')">Financial Analyst</button>
      <button class="role-btn" data-role="pm" onclick="switchRole('pm')">Property Manager</button>
    </div>
  </div>
</div>

<!-- Main -->
<div class="main">

  <!-- Process Checklist -->
  <div class="section-label">Your Budget Process</div>
  <div class="stepper" id="stepper"></div>

  <!-- Quick Links -->
  <div class="section-label">Quick Links</div>
  <div class="quick-grid">
    <a href="/generate" class="quick-card">
      <div class="quick-icon blue">&#x1F4CA;</div>
      <div><h4>Budget Generator</h4><p>Download Yardi reports and generate budgets for entities</p></div>
    </a>
    <a href="/dashboard" class="quick-card">
      <div class="quick-icon green">&#x1F4C8;</div>
      <div><h4>FA Dashboard</h4><p>Review status, manage workflow, approve submissions</p></div>
    </a>
    <a href="/pm" class="quick-card">
      <div class="quick-icon amber">&#x1F527;</div>
      <div><h4>PM Portal</h4><p>R&amp;M budget projections and building-level edits</p></div>
    </a>
    <a href="/assumptions/workbench" class="quick-card">
      <div class="quick-icon purple">&#x2699;&#xFE0F;</div>
      <div><h4>Assumptions</h4><p>Portfolio defaults and per-building overrides</p></div>
    </a>
    <a href="/audited-financials" class="quick-card">
      <div class="quick-icon slate">&#x1F4CB;</div>
      <div><h4>Audited Financials</h4><p>Extract and review prior-year financial data</p></div>
    </a>
    <a href="/files" class="quick-card">
      <div class="quick-icon rose">&#x1F4C1;</div>
      <div><h4>File Repository</h4><p>Upload and manage budget documents</p></div>
    </a>
  </div>

</div>

<!-- Footer -->
<div class="footer">Century Management &middot; Budget &amp; Assumptions System &middot; 2027 Cycle</div>

<script>
var phases = {
  fa: [
    { num: 1, title: 'Setup', desc: 'Load Yardi data into the system for each building.',
      items: ['Generate and run the Yardi script (YSL + AP Aging + Maint Proof)', 'Upload Expense Distribution manually for each entity', 'Verify all 4 data sources show a checkmark on the FA Dashboard', 'Set portfolio-wide assumptions (tax rates, escalations)'],
      link: { label: 'Go to Generator', href: '/generate' } },
    { num: 2, title: 'Review', desc: 'Edit budgets, review PM submissions, and check variances.',
      items: ['Review each building on the FA Dashboard across all tabs', 'Monitor PM submissions and send back for edits if needed', 'Adjust line items, formulas, and overrides as needed', 'Verify variance and percent-change columns for accuracy'],
      link: { label: 'Go to FA Dashboard', href: '/dashboard' } },
    { num: 3, title: 'Finalize', desc: 'Approve budgets and prepare for board presentation.',
      items: ['Run final review across all buildings', 'Approve each entity to lock edits', 'Open Board Presentation for each building', 'Export or print for board meeting materials'],
      link: { label: 'Go to FA Dashboard', href: '/dashboard' } }
  ],
  pm: [
    { num: 1, title: 'Setup', desc: 'Wait for your FA to load data. You will get access once it is ready.',
      items: ['Your FA uploads Yardi reports and creates the budget', 'Once loaded, your assigned buildings appear in the PM Portal', 'Review the assumptions set by your FA for your buildings'],
      link: { label: 'Check PM Portal', href: '/pm' } },
    { num: 2, title: 'Review', desc: 'Edit R&M projections and review budget lines for your buildings.',
      items: ['Open each assigned building in the PM Portal', 'Review and edit R&M line items and projections', 'Check income, utilities, and other expense tabs', 'Add notes or comments for your FA on any line'],
      link: { label: 'Go to PM Portal', href: '/pm' } },
    { num: 3, title: 'Finalize', desc: 'Submit your edits back to the FA for approval.',
      items: ['Do a final check on all your buildings', 'Click Submit to FA for each completed building', 'Your FA may send it back for revisions. Check status.', 'Once approved, the budget is locked for board review'],
      link: { label: 'Go to PM Portal', href: '/pm' } }
  ]
};

function renderStepper(role) {
  var el = document.getElementById('stepper');
  var cards = el.querySelectorAll('.phase-card');
  if (cards.length) {
    cards.forEach(function(c) { c.classList.add('fade-out'); });
    setTimeout(function() { buildCards(role); }, 250);
  } else {
    buildCards(role);
  }
}

function buildCards(role) {
  var el = document.getElementById('stepper');
  var data = phases[role];
  el.innerHTML = data.map(function(p) {
    return '<div class="phase-card">' +
      '<div class="phase-num">' + p.num + '</div>' +
      '<h3>' + p.title + '</h3>' +
      '<p class="phase-desc">' + p.desc + '</p>' +
      '<ul class="checklist">' + p.items.map(function(i) { return '<li>' + i + '</li>'; }).join('') + '</ul>' +
      '<a href="' + p.link.href + '" class="phase-link">' + p.link.label + ' <svg viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 8h10M9 4l4 4-4 4"/></svg></a>' +
      '</div>';
  }).join('');
}

function switchRole(role) {
  document.querySelectorAll('.role-btn').forEach(function(b) {
    b.classList.toggle('active', b.dataset.role === role);
  });
  renderStepper(role);
}

renderStepper('fa');
</script>

</body>
</html>
"""

GENERATE_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<title>Budget Generator — Century Management</title>
<style>
/* Force scrollbars always visible (fixes macOS auto-hide) */
::-webkit-scrollbar { width: 12px; height: 12px; -webkit-appearance: none; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 6px; border: 2px solid #f1f5f9; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }
::-webkit-scrollbar-corner { background: #f1f5f9; }
* { scrollbar-width: thin; scrollbar-color: #cbd5e1 #f1f5f9; }
  :root {
    --blue: #5a4a3f;
    --blue-light: #f5efe7;
    --green: #057a55;
    --green-light: #def7ec;
    --red: #e02424;
    --red-light: #fde8e8;
    --gray-50: #f4f1eb;
    --gray-100: #ede9e1;
    --gray-200: #e5e0d5;
    --gray-300: #d5cfc5;
    --gray-500: #8a7e72;
    --gray-700: #4a4039;
    --gray-900: #1a1714;
    --radius: 8px;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: var(--gray-50);
    color: var(--gray-900);
    line-height: 1.5;
  }
  .container { max-width: 900px; margin: 0 auto; padding: 40px 20px; }
  h1 { font-size: 28px; font-weight: 700; margin-bottom: 4px; }
  .subtitle { color: var(--gray-500); font-size: 15px; margin-bottom: 32px; }
  .back-link {
    color: var(--blue);
    text-decoration: none;
    font-weight: 600;
    padding: 8px 16px;
    border: 1px solid var(--blue);
    border-radius: 6px;
    transition: all 0.15s;
    display: inline-block;
    margin-bottom: 24px;
  }
  .back-link:hover { background: var(--blue-light); }

  /* Steps */
  .step {
    background: white;
    border: 1px solid var(--gray-200);
    border-radius: var(--radius);
    padding: 24px;
    margin-bottom: 20px;
  }
  .step-header {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 16px;
  }
  .step-num {
    background: var(--blue);
    color: white;
    width: 32px;
    height: 32px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 700;
    font-size: 14px;
    flex-shrink: 0;
  }
  .step-title { font-size: 18px; font-weight: 600; }
  .step-desc { color: var(--gray-500); font-size: 14px; margin-bottom: 16px; }

  /* Building grid */
  .search-box {
    width: 100%;
    padding: 10px 14px;
    border: 1px solid var(--gray-300);
    border-radius: var(--radius);
    font-size: 14px;
    margin-bottom: 12px;
    outline: none;
  }
  .search-box:focus { border-color: var(--blue); box-shadow: 0 0 0 3px rgba(90,74,63,0.1); }
  .building-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
    gap: 8px;
    max-height: 280px;
    overflow-y: auto;
    padding: 4px;
  }
  .building-item {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 8px 12px;
    border: 1px solid var(--gray-200);
    border-radius: 6px;
    cursor: pointer;
    transition: all 0.15s;
    font-size: 13px;
  }
  .building-item:hover { background: var(--blue-light); border-color: var(--blue); }
  .building-item.selected { background: var(--blue-light); border-color: var(--blue); }
  .building-item input { accent-color: var(--blue); }
  .select-actions { display: flex; gap: 8px; margin-bottom: 8px; }
  .select-actions button {
    background: none;
    border: none;
    color: var(--blue);
    cursor: pointer;
    font-size: 13px;
    padding: 4px 0;
  }
  .select-actions button:hover { text-decoration: underline; }

  /* Settings row */
  .settings-row {
    display: flex;
    gap: 16px;
    margin-bottom: 16px;
    flex-wrap: wrap;
  }
  .setting { flex: 1; min-width: 200px; }
  .setting label { display: block; font-size: 13px; font-weight: 600; margin-bottom: 4px; color: var(--gray-700); }
  .setting input {
    width: 100%;
    padding: 8px 12px;
    border: 1px solid var(--gray-300);
    border-radius: 6px;
    font-size: 14px;
  }

  /* Script output */
  .script-box {
    position: relative;
    background: var(--gray-900);
    color: #e5e0d5;
    border-radius: var(--radius);
    padding: 16px;
    font-family: 'SF Mono', Monaco, Consolas, monospace;
    font-size: 12px;
    max-height: 200px;
    overflow: auto;
    white-space: pre-wrap;
    display: none;
  }
  .copy-btn {
    position: absolute;
    top: 8px;
    right: 8px;
    background: var(--gray-700);
    color: white;
    border: none;
    border-radius: 4px;
    padding: 6px 14px;
    cursor: pointer;
    font-size: 12px;
    font-weight: 600;
  }
  .copy-btn:hover { background: var(--gray-500); }
  .copy-btn.copied { background: var(--green); }

  /* Upload zone */
  .upload-zone {
    border: 2px dashed var(--gray-300);
    border-radius: var(--radius);
    padding: 40px;
    text-align: center;
    cursor: pointer;
    transition: all 0.2s;
  }
  .upload-zone:hover, .upload-zone.dragover {
    border-color: var(--blue);
    background: var(--blue-light);
  }
  .upload-zone svg { margin-bottom: 8px; }
  .upload-text { color: var(--gray-500); font-size: 14px; }
  .upload-text strong { color: var(--blue); }
  .file-list { margin-top: 12px; }
  .file-item {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 8px 12px;
    background: var(--gray-100);
    border-radius: 6px;
    font-size: 13px;
    margin-bottom: 4px;
  }
  .file-item .remove { color: var(--red); cursor: pointer; border: none; background: none; font-size: 16px; }

  /* Buttons */
  .btn {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 10px 24px;
    border-radius: var(--radius);
    font-size: 14px;
    font-weight: 600;
    cursor: pointer;
    border: none;
    transition: all 0.15s;
  }
  .btn-primary { background: var(--blue); color: white; }
  .btn-primary:hover { background: #3d342c; }
  .btn-primary:disabled { background: var(--gray-300); cursor: not-allowed; }
  .btn-green { background: var(--green); color: white; }
  .btn-green:hover { background: #046c4e; }
  .btn-green:disabled { background: var(--gray-300); cursor: not-allowed; }

  /* Results */
  .results { display: none; }
  .result-card {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 12px 16px;
    border-radius: 6px;
    margin-bottom: 6px;
    font-size: 14px;
  }
  .result-card.success { background: var(--green-light); }
  .result-card.failed { background: var(--red-light); }
  .result-info { display: flex; align-items: center; gap: 8px; }

  /* Progress */
  .progress-bar {
    display: none;
    height: 4px;
    background: var(--gray-200);
    border-radius: 2px;
    overflow: hidden;
    margin: 16px 0;
  }
  .progress-bar .fill {
    height: 100%;
    background: var(--blue);
    border-radius: 2px;
    animation: progress 2s ease-in-out infinite;
    width: 40%;
  }
  @keyframes progress {
    0% { transform: translateX(-100%); }
    100% { transform: translateX(350%); }
  }

  .count-badge {
    background: var(--blue);
    color: white;
    font-size: 12px;
    padding: 2px 10px;
    border-radius: 12px;
    font-weight: 600;
  }
  .gen-pill { display:inline-block; font-size:10px; font-weight:600; padding:1px 7px; border-radius:10px; margin-left:6px; vertical-align:middle; }
  .gen-pill-done { background:#dcfce7; color:#166534; }
  .gen-pill-partial { background:#fef3c7; color:#92400e; }
  .gen-pill-none { background:#f1f5f9; color:#64748b; }
  .building-item.gen-done { opacity:0.55; }

  .page-header {
    background: linear-gradient(135deg, var(--blue) 0%, #3d342c 100%);
    color: white;
    padding: 30px 20px;
    margin-bottom: 0;
  }
  .page-header a { color: white; text-decoration: none; font-size: 14px; }
  .page-header a:hover { text-decoration: underline; }
  .page-header h1 { font-size: 28px; font-weight: 700; }
  .page-header p { font-size: 14px; opacity: 0.85; margin-top: 4px; }
</style>
</head>
<body>
<div class="page-header">
  <a href="/">← Home</a>
  <h1>Budget Generator</h1>
  <p>Download YSL reports from Yardi, then generate 2027 budgets in one click.</p>
</div>
<div class="container">

  <!-- STEP 1 -->
  <div class="step">
    <div class="step-header">
      <div class="step-num">1</div>
      <div class="step-title">Select Buildings & Run from Yardi</div>
    </div>
    <p class="step-desc">Pick buildings, copy the script into your Yardi Console (F12). Reports download locally and auto-upload to the budget system.</p>

    <input type="text" class="search-box" id="buildingSearch" placeholder="Search buildings..." oninput="filterBuildings()">
    <div class="select-actions">
      <button onclick="selectRemaining()" id="runRemainingBtn" style="background:var(--blue);color:white;font-weight:600;">▶ Run All Remaining</button>
      <button onclick="selectAll()">Select All</button>
      <button onclick="selectNone()">Select None</button>
      <span id="selectedCount" class="count-badge">0 selected</span>
    </div>
    <div class="building-grid" id="buildingGrid">
      {% for b in buildings %}
      <label class="building-item" data-code="{{ b.entity_code }}" data-name="{{ b.building_name|lower }}">
        <input type="checkbox" value="{{ b.entity_code }}" onchange="updateCount()">
        <span><strong>{{ b.entity_code }}</strong> — {{ b.building_name }}</span>
      </label>
      {% endfor %}
    </div>

    <div class="settings-row" style="margin-top: 16px;">
      <div class="setting">
        <label>Your Email</label>
        <input type="email" id="email" placeholder="you@centuryny.com">
      </div>
      <div class="setting">
        <label>Budget Period</label>
        <div style="display:flex; gap:6px;">
          <select id="periodMonth" style="padding:8px 10px; border:1px solid var(--gray-300); border-radius:6px; font-size:14px; background:white;">
            <option value="01">January</option><option value="02">February</option><option value="03">March</option>
            <option value="04">April</option><option value="05">May</option><option value="06">June</option>
            <option value="07">July</option><option value="08">August</option><option value="09">September</option>
            <option value="10">October</option><option value="11">November</option><option value="12">December</option>
          </select>
          <select id="periodYear" style="padding:8px 10px; border:1px solid var(--gray-300); border-radius:6px; font-size:14px; background:white;">
            <option value="2025">2025</option><option value="2026">2026</option><option value="2027">2027</option><option value="2028">2028</option>
          </select>
        </div>
      </div>
      <div class="setting">
        <label style="display:flex; align-items:center; gap:8px; cursor:pointer;">
          <input type="checkbox" id="freshStart" style="width:16px; height:16px;">
          <span>Fresh Start</span>
        </label>
        <span style="font-size:11px; color:var(--gray-500);">Creates a brand new budget version. Existing budget stays untouched.</span>
      </div>
    </div>

    <div style="display:flex; gap:10px; flex-wrap:wrap;">
      <button class="btn btn-primary" onclick="generateScript()" id="genBtn">
        <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M16 18l2-2-2-2M8 18l-2-2 2-2M14 4l-4 16"/></svg>
        Generate Yardi Script
      </button>
    </div>
    <p style="font-size:11px; color:var(--gray-500); margin-top:6px;">Runs YSL + Maint Proof + AP Aging. Expense Distribution is run manually in Yardi and uploaded via Manual Upload below.</p>

    <div class="script-box" id="scriptBox">
      <button class="copy-btn" id="copyBtn" onclick="copyScript()">Copy</button>
      <code id="scriptCode"></code>
    </div>

    <!-- AP Aging script box hidden — AP Aging is now in the combined script -->
    <div class="script-box" id="apScriptBox" style="display:none; border-color:var(--amber-200, #fde68a);">
      <button class="copy-btn" id="apCopyBtn" onclick="copyAPScript()">Copy AP Aging</button>
      <code id="apScriptCode"></code>
    </div>
  </div>

  <!-- Auto-upload status (populated by Yardi script) -->
  <div id="autoStatus" style="display:none; margin-top:16px; padding:16px; background:#f0fdf4; border:1px solid #86efac; border-radius:8px;">
    <p style="margin:0; font-weight:600; color:#166534;">Files auto-uploaded and processed. Check the FA Dashboard for results.</p>
  </div>

  <!-- Manual upload fallback -->
  <div style="margin-top:24px; padding:16px; background:#f8f5f0; border:1px solid #e8e0d4; border-radius:8px;">
    <p style="margin:0 0 8px; font-weight:600; color:#5a4a3f;">Manual Upload</p>
    <p style="margin:0 0 12px; font-size:12px; color:#8a7a6f;">Upload Yardi Excel files directly. The app auto-detects file types — select one or all four at once.</p>
    <!-- Visual stepper (decorative — all 4 file types can be uploaded together) -->
    <div style="display:flex; align-items:center; gap:4px; margin-bottom:14px; flex-wrap:wrap;">
      <div style="display:flex; align-items:center; gap:8px;">
        <div style="width:26px; height:26px; border-radius:50%; background:#5a4a3f; color:#fff; display:flex; align-items:center; justify-content:center; font-size:12px; font-weight:600;">1</div>
        <span style="font-size:12px; color:#5a4a3f; font-weight:500;">YSL</span>
      </div>
      <div style="flex:0 0 24px; height:2px; background:#d5cfc5; margin:0 4px;"></div>
      <div style="display:flex; align-items:center; gap:8px;">
        <div style="width:26px; height:26px; border-radius:50%; background:#5a4a3f; color:#fff; display:flex; align-items:center; justify-content:center; font-size:12px; font-weight:600;">2</div>
        <span style="font-size:12px; color:#5a4a3f; font-weight:500;">AP Aging</span>
      </div>
      <div style="flex:0 0 24px; height:2px; background:#d5cfc5; margin:0 4px;"></div>
      <div style="display:flex; align-items:center; gap:8px;">
        <div style="width:26px; height:26px; border-radius:50%; background:#5a4a3f; color:#fff; display:flex; align-items:center; justify-content:center; font-size:12px; font-weight:600;">3</div>
        <span style="font-size:12px; color:#5a4a3f; font-weight:500;">Maint Proof</span>
      </div>
      <div style="flex:0 0 24px; height:2px; background:#d5cfc5; margin:0 4px;"></div>
      <div style="display:flex; align-items:center; gap:8px;">
        <div style="width:26px; height:26px; border-radius:50%; background:#5a4a3f; color:#fff; display:flex; align-items:center; justify-content:center; font-size:12px; font-weight:600;">4</div>
        <span style="font-size:12px; color:#5a4a3f; font-weight:500;">Exp Dist</span>
      </div>
    </div>
    <div style="display:flex; gap:10px; align-items:center; flex-wrap:wrap;">
      <input type="file" id="manualFiles" multiple accept=".xlsx,.xls" style="font-size:13px;">
      <button class="btn btn-primary" id="uploadBtn" onclick="manualUpload()" style="background:#5a4a3f;">Upload &amp; Process</button>
    </div>
  </div>
</div>

<script>
// Auto-upload: no manual file management needed

// ── Batch: tag buildings with budget status on load ──
const _budgetStatus = {};  // entity_code → {status, done}
(async function tagBuildings() {
  try {
    const resp = await fetch('/api/budgets');
    const budgets = await resp.json();
    budgets.forEach(b => {
      const ts = b.timestamps || {};
      const hasYsl = !!ts.ysl;
      const hasAp = !!ts.open_ap;
      const hasBudget = !!ts.budget_summary;
      const hasExp = !!ts.expense_dist;
      const allDone = hasYsl && hasAp && hasBudget;
      const partial = hasYsl || hasAp || hasBudget || hasExp;
      _budgetStatus[String(b.entity_code)] = { done: allDone, partial: partial && !allDone };
    });
    // Tag each building checkbox
    let doneCount = 0, remainCount = 0;
    document.querySelectorAll('.building-item').forEach(el => {
      const code = el.dataset.code;
      const info = _budgetStatus[code];
      const pill = document.createElement('span');
      pill.className = 'gen-pill';
      if (info && info.done) {
        pill.className += ' gen-pill-done';
        pill.textContent = 'Done';
        el.classList.add('gen-done');
        doneCount++;
      } else if (info && info.partial) {
        pill.className += ' gen-pill-partial';
        pill.textContent = 'Partial';
        remainCount++;
      } else {
        pill.className += ' gen-pill-none';
        pill.textContent = 'Not started';
        remainCount++;
      }
      el.querySelector('span').appendChild(pill);
    });
    // Update button label with count
    const btn = document.getElementById('runRemainingBtn');
    if (btn) btn.textContent = `▶ Run All Remaining (${remainCount})`;
  } catch (e) {
    console.warn('Could not load budget status:', e);
  }
})();

function selectRemaining() {
  document.querySelectorAll('#buildingGrid .building-item').forEach(el => {
    const code = el.dataset.code;
    const info = _budgetStatus[code];
    const cb = el.querySelector('input');
    const visible = !el.style.display || el.style.display !== 'none';
    cb.checked = visible && !(info && info.done);
  });
  updateCount();
}

function getSelected() {
  return [...document.querySelectorAll('#buildingGrid input:checked')].map(c => parseInt(c.value));
}

function updateCount() {
  const n = getSelected().length;
  document.getElementById('selectedCount').textContent = n + ' selected';
  document.querySelectorAll('.building-item').forEach(el => {
    el.classList.toggle('selected', el.querySelector('input').checked);
  });
}

function selectAll() {
  document.querySelectorAll('#buildingGrid input').forEach(c => {
    if (!c.closest('.building-item').style.display || c.closest('.building-item').style.display !== 'none')
      c.checked = true;
  });
  updateCount();
}

function selectNone() {
  document.querySelectorAll('#buildingGrid input').forEach(c => c.checked = false);
  updateCount();
}

function filterBuildings() {
  const q = document.getElementById('buildingSearch').value.toLowerCase();
  document.querySelectorAll('.building-item').forEach(el => {
    const code = el.dataset.code;
    const name = el.dataset.name;
    el.style.display = (code.includes(q) || name.includes(q)) ? '' : 'none';
  });
}

// Default period dropdowns to current month/year
(function() {
  const now = new Date();
  const m = document.getElementById('periodMonth');
  const y = document.getElementById('periodYear');
  if (m) m.value = String(now.getMonth()+1).padStart(2,'0');
  if (y) y.value = String(now.getFullYear());
})();

async function generateScript() {
  const entities = getSelected();
  const email = document.getElementById('email').value;
  const period = document.getElementById('periodMonth').value + '/' + document.getElementById('periodYear').value;

  if (!entities.length) { alert('Select at least one building'); return; }
  if (!email) { alert('Enter your email'); return; }

  const freshStart = document.getElementById('freshStart').checked;
  const resp = await fetch('/api/generate-script', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ entities, email, period, fresh_start: freshStart }),
  });

  const data = await resp.json();
  if (data.error) { alert(data.error); return; }

  document.getElementById('scriptCode').textContent = data.script;
  document.getElementById('scriptBox').style.display = 'block';
}

function copyScript() {
  const code = document.getElementById('scriptCode').textContent;
  navigator.clipboard.writeText(code).then(() => {
    const btn = document.getElementById('copyBtn');
    btn.textContent = 'Copied!';
    btn.classList.add('copied');
    setTimeout(() => { btn.textContent = 'Copy'; btn.classList.remove('copied'); }, 2000);
  });
}

async function generateAPAgingScript() {
  const entities = getSelected();
  const email = document.getElementById('email').value;
  const period = document.getElementById('periodMonth').value + '/' + document.getElementById('periodYear').value;

  if (!entities.length) { alert('Select at least one building'); return; }

  const resp = await fetch('/api/generate-ap-aging-script', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ entities, email, period }),
  });

  const data = await resp.json();
  if (data.error) { alert(data.error); return; }

  document.getElementById('apScriptCode').textContent = data.script;
  document.getElementById('apScriptBox').style.display = 'block';
}

function copyAPScript() {
  const code = document.getElementById('apScriptCode').textContent;
  navigator.clipboard.writeText(code).then(() => {
    const btn = document.getElementById('apCopyBtn');
    btn.textContent = 'Copied!';
    btn.classList.add('copied');
    setTimeout(() => { btn.textContent = 'Copy AP Aging'; btn.classList.remove('copied'); }, 2000);
  });
}

// ── Manual Upload (fallback for files that don't auto-upload) ──
async function manualUpload() {
  const input = document.getElementById('manualFiles');
  if (!input.files.length) { alert('Select at least one file'); return; }

  const formData = new FormData();
  for (const f of input.files) formData.append('files', f);
  formData.append('fresh_start', 'true');

  const btn = document.getElementById('uploadBtn');
  btn.disabled = true;
  btn.textContent = 'Processing...';

  try {
    const resp = await fetch('/api/process', { method: 'POST', body: formData });
    const data = await resp.json();
    if (data.error) { alert('Error: ' + data.error); return; }

    let msg = '';
    if (data.success?.length) msg += 'Success: ' + data.success.join(', ') + '\\n';
    if (data.warnings?.length) msg += 'Warnings: ' + data.warnings.join(', ') + '\\n';
    if (data.failed?.length) msg += 'Failed: ' + data.failed.join(', ');
    alert(msg || 'Files processed.');
    input.value = '';
  } catch (err) {
    alert('Upload failed: ' + err.message);
  } finally {
    btn.disabled = false;
    btn.textContent = 'Upload & Process';
  }
}
</script>
</body>
</html>
"""

# MANAGE_TEMPLATE removed — buildings now sync from Monday.com

_MANAGE_REMOVED = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<title>Manage Buildings — Century Management</title>
<style>
/* Force scrollbars always visible (fixes macOS auto-hide) */
::-webkit-scrollbar { width: 12px; height: 12px; -webkit-appearance: none; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 6px; border: 2px solid #f1f5f9; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }
::-webkit-scrollbar-corner { background: #f1f5f9; }
* { scrollbar-width: thin; scrollbar-color: #cbd5e1 #f1f5f9; }
  :root {
    --blue: #5a4a3f;
    --blue-light: #f5efe7;
    --green: #057a55;
    --green-light: #def7ec;
    --red: #e02424;
    --red-light: #fde8e8;
    --gray-50: #f4f1eb;
    --gray-100: #ede9e1;
    --gray-200: #e5e0d5;
    --gray-300: #d5cfc5;
    --gray-500: #8a7e72;
    --gray-700: #4a4039;
    --gray-900: #1a1714;
    --radius: 8px;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: var(--gray-50);
    color: var(--gray-900);
    line-height: 1.5;
  }
  .container { max-width: 1200px; margin: 0 auto; padding: 40px 20px; }
  h1 { font-size: 28px; font-weight: 700; margin-bottom: 4px; }
  .subtitle { color: var(--gray-500); font-size: 15px; margin-bottom: 32px; }
  .header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 32px;
  }
  .back-link {
    color: var(--blue);
    text-decoration: none;
    font-weight: 600;
    padding: 8px 16px;
    border: 1px solid var(--blue);
    border-radius: 6px;
    transition: all 0.15s;
  }
  .back-link:hover { background: var(--blue-light); }

  /* Card */
  .card {
    background: white;
    border: 1px solid var(--gray-200);
    border-radius: var(--radius);
    padding: 24px;
    margin-bottom: 24px;
  }
  .card-title { font-size: 18px; font-weight: 600; margin-bottom: 16px; }

  /* Form */
  .form-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
    gap: 12px;
    margin-bottom: 16px;
  }
  .form-group {
    display: flex;
    flex-direction: column;
  }
  .form-group label {
    font-size: 13px;
    font-weight: 600;
    margin-bottom: 4px;
    color: var(--gray-700);
  }
  .form-group input {
    padding: 8px 12px;
    border: 1px solid var(--gray-300);
    border-radius: 6px;
    font-size: 14px;
    outline: none;
  }
  .form-group input:focus { border-color: var(--blue); box-shadow: 0 0 0 3px rgba(90,74,63,0.1); }

  /* Search */
  .search-box {
    width: 100%;
    padding: 10px 14px;
    border: 1px solid var(--gray-300);
    border-radius: var(--radius);
    font-size: 14px;
    margin-bottom: 16px;
    outline: none;
  }
  .search-box:focus { border-color: var(--blue); box-shadow: 0 0 0 3px rgba(90,74,63,0.1); }

  /* Table */
  .table-wrapper {
    overflow-x: auto;
    border: 1px solid var(--gray-200);
    border-radius: var(--radius);
  }
  table {
    width: 100%;
    border-collapse: collapse;
    font-size: 14px;
  }
  th {
    background: var(--gray-100);
    padding: 12px;
    text-align: left;
    font-weight: 600;
    border-bottom: 1px solid var(--gray-200);
  }
  td {
    padding: 12px;
    border-bottom: 1px solid var(--gray-200);
  }
  tr:hover { background: var(--gray-50); }
  .code-cell { font-weight: 600; color: var(--blue); }

  /* Buttons */
  .btn {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 10px 24px;
    border-radius: var(--radius);
    font-size: 14px;
    font-weight: 600;
    cursor: pointer;
    border: none;
    transition: all 0.15s;
  }
  .btn-primary { background: var(--blue); color: white; }
  .btn-primary:hover { background: #3d342c; }
  .btn-primary:disabled { background: var(--gray-300); cursor: not-allowed; }
  .btn-sm {
    padding: 6px 12px;
    font-size: 12px;
    gap: 4px;
  }
  .btn-edit {
    background: var(--blue);
    color: white;
    padding: 6px 12px;
    border: none;
    border-radius: 4px;
    font-size: 12px;
    cursor: pointer;
    transition: all 0.15s;
  }
  .btn-edit:hover { background: #3d342c; }
  .btn-delete {
    background: var(--red);
    color: white;
    padding: 6px 12px;
    border: none;
    border-radius: 4px;
    font-size: 12px;
    cursor: pointer;
    transition: all 0.15s;
    margin-left: 4px;
  }
  .btn-delete:hover { background: #c91c1c; }

  /* Modal */
  .modal {
    display: none;
    position: fixed;
    top: 0;
    left: 0;
    right: 0;
    bottom: 0;
    background: rgba(0,0,0,0.5);
    z-index: 1000;
    align-items: center;
    justify-content: center;
  }
  .modal.active { display: flex; }
  .modal-content {
    background: white;
    border-radius: var(--radius);
    padding: 24px;
    max-width: 500px;
    width: 90%;
    max-height: 90vh;
    overflow-y: auto;
  }
  .modal-title { font-size: 18px; font-weight: 600; margin-bottom: 16px; }
  .modal-actions {
    display: flex;
    gap: 8px;
    margin-top: 24px;
    justify-content: flex-end;
  }

  /* Alert */
  .alert {
    padding: 12px 16px;
    border-radius: 6px;
    margin-bottom: 16px;
    font-size: 14px;
    display: none;
  }
  .alert.active { display: block; }
  .alert.success { background: var(--green-light); color: var(--green); }
  .alert.error { background: var(--red-light); color: var(--red); }

  /* Empty state */
  .empty-state {
    text-align: center;
    padding: 40px;
    color: var(--gray-500);
  }
  .empty-state p { margin-bottom: 8px; }
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <div>
      <h1>Manage Buildings</h1>
      <p class="subtitle">Add, edit, or remove buildings from the budget generator.</p>
    </div>
    <a href="/" class="back-link">← Home</a>
  </div>

  <div id="alert" class="alert"></div>

  <!-- Add Building Card -->
  <div class="card">
    <div class="card-title">Add New Building</div>
    <div class="form-grid">
      <div class="form-group">
        <label>Entity Code</label>
        <input type="text" id="newEntityCode" placeholder="e.g., 148">
      </div>
      <div class="form-group">
        <label>Building Name</label>
        <input type="text" id="newBuildingName" placeholder="e.g., Main Office">
      </div>
      <div class="form-group">
        <label>Address</label>
        <input type="text" id="newAddress" placeholder="e.g., 130 E. 18th St">
      </div>
      <div class="form-group">
        <label>City</label>
        <input type="text" id="newCity" placeholder="e.g., New York">
      </div>
      <div class="form-group">
        <label>ZIP</label>
        <input type="text" id="newZIP" placeholder="e.g., 10003">
      </div>
      <div class="form-group">
        <label>Type</label>
        <input type="text" id="newType" placeholder="e.g., Residential">
      </div>
      <div class="form-group">
        <label>Units</label>
        <input type="text" id="newUnits" placeholder="e.g., 50">
      </div>
    </div>
    <button class="btn btn-primary" onclick="addBuilding()">
      + Add Building
    </button>
  </div>

  <!-- Buildings List Card -->
  <div class="card">
    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px;">
      <div class="card-title">Buildings (<span id="buildingCount">0</span>)</div>
    </div>
    <input type="text" class="search-box" id="searchBox" placeholder="Search buildings..." oninput="filterTable()">
    <div class="table-wrapper">
      <table id="buildingsTable">
        <thead>
          <tr>
            <th>Entity Code</th>
            <th>Building Name</th>
            <th>Address</th>
            <th>City</th>
            <th>ZIP</th>
            <th>Type</th>
            <th>Units</th>
            <th>Actions</th>
          </tr>
        </thead>
        <tbody id="buildingsBody">
        </tbody>
      </table>
    </div>
    <div id="emptyState" class="empty-state" style="display: none;">
      <p>No buildings found.</p>
    </div>
  </div>
</div>

<!-- Edit Modal -->
<div id="editModal" class="modal">
  <div class="modal-content">
    <div class="modal-title">Edit Building</div>
    <div class="form-grid" style="grid-template-columns: 1fr;">
      <div class="form-group">
        <label>Entity Code (read-only)</label>
        <input type="text" id="editEntityCode" readonly style="background: var(--gray-100); cursor: not-allowed;">
      </div>
      <div class="form-group">
        <label>Building Name</label>
        <input type="text" id="editBuildingName">
      </div>
      <div class="form-group">
        <label>Address</label>
        <input type="text" id="editAddress">
      </div>
      <div class="form-group">
        <label>City</label>
        <input type="text" id="editCity">
      </div>
      <div class="form-group">
        <label>ZIP</label>
        <input type="text" id="editZIP">
      </div>
      <div class="form-group">
        <label>Type</label>
        <input type="text" id="editType">
      </div>
      <div class="form-group">
        <label>Units</label>
        <input type="text" id="editUnits">
      </div>
    </div>
    <div class="modal-actions">
      <button class="btn" style="background: var(--gray-200); color: var(--gray-900);" onclick="closeEditModal()">Cancel</button>
      <button class="btn btn-primary" onclick="saveEdit()">Save Changes</button>
    </div>
  </div>
</div>

<script>
let buildings = [];

async function loadBuildings() {
  const resp = await fetch('/api/buildings');
  buildings = await resp.json();
  renderTable();
  updateCount();
}

function renderTable() {
  const tbody = document.getElementById('buildingsBody');
  const searchTerm = document.getElementById('searchBox').value.toLowerCase();

  const filtered = buildings.filter(b =>
    b.entity_code.toLowerCase().includes(searchTerm) ||
    b.building_name.toLowerCase().includes(searchTerm) ||
    b.address.toLowerCase().includes(searchTerm) ||
    b.city.toLowerCase().includes(searchTerm)
  );

  const empty = document.getElementById('emptyState');
  if (filtered.length === 0) {
    tbody.innerHTML = '';
    empty.style.display = 'block';
    return;
  }
  empty.style.display = 'none';

  tbody.innerHTML = filtered.map(b =>
    `<tr>
      <td class="code-cell">${escapeHtml(b.entity_code)}</td>
      <td>${escapeHtml(b.building_name)}</td>
      <td>${escapeHtml(b.address)}</td>
      <td>${escapeHtml(b.city)}</td>
      <td>${escapeHtml(b.zip)}</td>
      <td>${escapeHtml(b.type)}</td>
      <td>${escapeHtml(b.units)}</td>
      <td>
        <button class="btn-edit" onclick="openEditModal('${escapeAttr(b.entity_code)}')">Edit</button>
        <button class="btn-delete" onclick="deleteBuilding('${escapeAttr(b.entity_code)}')">Delete</button>
      </td>
    </tr>`
  ).join('');
}

function filterTable() {
  renderTable();
}

function updateCount() {
  document.getElementById('buildingCount').textContent = buildings.length;
}

function escapeHtml(text) {
  const div = document.createElement('div');
  div.textContent = text;
  return div.innerHTML;
}

function escapeAttr(text) {
  return text.replace(/'/g, "\\'");
}

async function addBuilding() {
  const code = document.getElementById('newEntityCode').value.trim();
  const name = document.getElementById('newBuildingName').value.trim();
  const address = document.getElementById('newAddress').value.trim();
  const city = document.getElementById('newCity').value.trim();
  const zip = document.getElementById('newZIP').value.trim();
  const type = document.getElementById('newType').value.trim();
  const units = document.getElementById('newUnits').value.trim();

  if (!code || !name) {
    showAlert('Entity code and building name are required', 'error');
    return;
  }

  const resp = await fetch('/api/buildings', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ entity_code: code, building_name: name, address, city, zip, type, units }),
  });

  if (resp.ok) {
    document.getElementById('newEntityCode').value = '';
    document.getElementById('newBuildingName').value = '';
    document.getElementById('newAddress').value = '';
    document.getElementById('newCity').value = '';
    document.getElementById('newZIP').value = '';
    document.getElementById('newType').value = '';
    document.getElementById('newUnits').value = '';
    await loadBuildings();
    showAlert(`Building "${name}" added successfully`, 'success');
  } else {
    const data = await resp.json();
    showAlert(data.error || 'Failed to add building', 'error');
  }
}

function openEditModal(code) {
  const building = buildings.find(b => b.entity_code === code);
  if (!building) return;
  document.getElementById('editEntityCode').value = building.entity_code;
  document.getElementById('editBuildingName').value = building.building_name;
  document.getElementById('editAddress').value = building.address;
  document.getElementById('editCity').value = building.city;
  document.getElementById('editZIP').value = building.zip;
  document.getElementById('editType').value = building.type;
  document.getElementById('editUnits').value = building.units;
  document.getElementById('editModal').classList.add('active');
}

function closeEditModal() {
  document.getElementById('editModal').classList.remove('active');
}

async function saveEdit() {
  const code = document.getElementById('editEntityCode').value;
  const name = document.getElementById('editBuildingName').value.trim();
  const address = document.getElementById('editAddress').value.trim();
  const city = document.getElementById('editCity').value.trim();
  const zip = document.getElementById('editZIP').value.trim();
  const type = document.getElementById('editType').value.trim();
  const units = document.getElementById('editUnits').value.trim();

  if (!name) {
    showAlert('Building name is required', 'error');
    return;
  }

  const resp = await fetch(`/api/buildings/${code}`, {
    method: 'PUT',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ building_name: name, address, city, zip, type, units }),
  });

  if (resp.ok) {
    closeEditModal();
    await loadBuildings();
    showAlert(`Building "${name}" updated successfully`, 'success');
  } else {
    const data = await resp.json();
    showAlert(data.error || 'Failed to update building', 'error');
  }
}

async function deleteBuilding(code) {
  const building = buildings.find(b => b.entity_code === code);
  if (!confirm(`Are you sure you want to delete "${building.building_name}"?`)) return;

  const resp = await fetch(`/api/buildings/${code}/delete`, { method: 'POST' });

  if (resp.ok) {
    await loadBuildings();
    showAlert(`Building deleted successfully`, 'success');
  } else {
    const data = await resp.json();
    showAlert(data.error || 'Failed to delete building', 'error');
  }
}

function showAlert(msg, type) {
  const el = document.getElementById('alert');
  el.textContent = msg;
  el.className = `alert active ${type}`;
  setTimeout(() => el.classList.remove('active'), 4000);
}

// Load on page load
document.addEventListener('DOMContentLoaded', loadBuildings);
</script>
</body>
</html>
"""

ASSUMPTIONS_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
    <title>Portfolio Defaults</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; background: #f5f5f5; }
        header { background: #5a4a3f; color: white; padding: 24px; }
        header h1 { font-size: 24px; font-weight: 600; }
        header .back-link { color: white; margin-bottom: 8px; }
        .container { max-width: 900px; margin: 0 auto; padding: 40px 20px; }
        .back-link { display: inline-block; margin-bottom: 24px; color: #5a4a3f; text-decoration: none; font-size: 14px; }
        .back-link:hover { text-decoration: underline; }
        .section { background: white; border-radius: 8px; padding: 24px; margin-bottom: 24px; border: 1px solid #e0e0e0; }
        .section h2 { font-size: 18px; color: #5a4a3f; margin-bottom: 16px; font-weight: 600; }
        .form-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 16px; }
        .form-group { display: flex; flex-direction: column; }
        .form-group label { font-size: 13px; font-weight: 500; color: #333; margin-bottom: 6px; }
        .form-group input, .form-group select { padding: 10px; border: 1px solid #ddd; border-radius: 4px; font-size: 14px; }
        .form-group input:focus, .form-group select:focus { outline: none; border-color: #5a4a3f; box-shadow: 0 0 0 2px rgba(90,74,63, 0.1); }
        .form-group input[type="number"] { font-family: 'Courier New', monospace; }
        .button-row { display: flex; gap: 12px; margin-top: 24px; }
        button { padding: 10px 20px; border: none; border-radius: 4px; font-size: 14px; font-weight: 500; cursor: pointer; }
        .btn-primary { background: #5a4a3f; color: white; }
        .btn-primary:hover { background: #0f1f38; }
        .toast { position: fixed; bottom: 20px; right: 20px; background: #4caf50; color: white; padding: 16px 20px; border-radius: 4px; display: none; z-index: 1000; }
        .toast.show { display: block; animation: slideIn 0.3s ease; }
        @keyframes slideIn { from { transform: translateX(400px); } to { transform: translateX(0); } }
    </style>
</head>
<body>
    <header>
        <a href="/" class="back-link">← Home</a>
        <h1>Portfolio Defaults</h1>
    </header>
    <div class="container">

        <div class="section">
            <h2>Payroll Tax Rates</h2>
            <div class="form-grid">
                <div class="form-group">
                    <label>FICA</label>
                    <input type="number" step="0.0001" id="fica" />
                </div>
                <div class="form-group">
                    <label>SUI</label>
                    <input type="number" step="0.0001" id="sui" />
                </div>
                <div class="form-group">
                    <label>FUI</label>
                    <input type="number" step="0.0001" id="fui" />
                </div>
                <div class="form-group">
                    <label>MTA</label>
                    <input type="number" step="0.0001" id="mta" />
                </div>
                <div class="form-group">
                    <label>NYS Disability</label>
                    <input type="number" step="0.0001" id="nys_disability" />
                </div>
                <div class="form-group">
                    <label>Paid Family Leave</label>
                    <input type="number" step="0.0001" id="pfl" />
                </div>
            </div>
        </div>

        <div class="section">
            <h2>32BJ Union Benefits</h2>
            <div class="form-grid">
                <div class="form-group">
                    <label>Welfare ($/mo)</label>
                    <input type="number" step="0.01" id="welfare_monthly" />
                </div>
                <div class="form-group">
                    <label>Pension ($/wk)</label>
                    <input type="number" step="0.01" id="pension_weekly" />
                </div>
                <div class="form-group">
                    <label>Supp Retirement ($/wk)</label>
                    <input type="number" step="0.01" id="supp_retirement_weekly" />
                </div>
                <div class="form-group">
                    <label>Legal ($/mo)</label>
                    <input type="number" step="0.01" id="legal_monthly" />
                </div>
                <div class="form-group">
                    <label>Training ($/mo)</label>
                    <input type="number" step="0.01" id="training_monthly" />
                </div>
                <div class="form-group">
                    <label>Profit Sharing ($/qtr)</label>
                    <input type="number" step="0.01" id="profit_sharing_quarterly" />
                </div>
            </div>
        </div>

        <div class="section">
            <h2>Workers Comp</h2>
            <div class="form-grid">
                <div class="form-group">
                    <label>Workers Comp %</label>
                    <input type="number" step="0.0001" id="workers_comp_percent" />
                </div>
            </div>
        </div>

        <div class="section">
            <h2>Mid-Year Wage Increase</h2>
            <div class="form-grid">
                <div class="form-group">
                    <label>Increase %</label>
                    <input type="number" step="0.0001" id="wage_increase_percent" />
                </div>
                <div class="form-group">
                    <label>Effective Week</label>
                    <input type="text" id="wage_increase_effective_week" />
                </div>
                <div class="form-group">
                    <label>Pre-increase Weeks</label>
                    <input type="number" id="wage_increase_pre_weeks" />
                </div>
                <div class="form-group">
                    <label>Post-increase Weeks</label>
                    <input type="number" id="wage_increase_post_weeks" />
                </div>
            </div>
        </div>

        <div class="section">
            <h2>Insurance Renewal</h2>
            <div class="form-grid">
                <div class="form-group">
                    <label>Expected Renewal Increase %</label>
                    <input type="number" step="0.0001" id="insurance_renewal_increase_percent" />
                </div>
                <div class="form-group">
                    <label>Renewal Effective Date</label>
                    <input type="text" id="insurance_renewal_effective_date" />
                </div>
                <div class="form-group">
                    <label>Pre-renewal Months</label>
                    <input type="number" id="insurance_renewal_pre_months" />
                </div>
                <div class="form-group">
                    <label>Post-renewal Months</label>
                    <input type="number" id="insurance_renewal_post_months" />
                </div>
            </div>
        </div>

        <div class="section">
            <h2>Energy Rates</h2>
            <div class="form-grid">
                <div class="form-group">
                    <label>Gas ESCO Rate ($/Therm)</label>
                    <input type="number" step="0.0001" id="energy_gas_esco_rate" />
                </div>
                <div class="form-group">
                    <label>Electric ESCO Rate ($/kWh)</label>
                    <input type="number" step="0.0001" id="energy_electric_esco_rate" />
                </div>
                <div class="form-group">
                    <label>Gas Rate Increase %</label>
                    <input type="number" step="0.0001" id="energy_gas_rate_increase" />
                </div>
                <div class="form-group">
                    <label>Electric Rate Increase %</label>
                    <input type="number" step="0.0001" id="energy_electric_rate_increase" />
                </div>
                <div class="form-group">
                    <label>Oil/Fuel Price ($/Gallon)</label>
                    <input type="number" step="0.01" id="energy_oil_price" />
                </div>
                <div class="form-group">
                    <label>Oil/Fuel Rate Increase %</label>
                    <input type="number" step="0.0001" id="energy_oil_rate_increase" />
                </div>
                <div class="form-group">
                    <label>Consumption Basis</label>
                    <select id="energy_consumption_basis">
                        <option value="2-Year Average">2-Year Average</option>
                        <option value="Prior Year">Prior Year</option>
                        <option value="Custom">Custom</option>
                    </select>
                </div>
            </div>
        </div>

        <div class="section">
            <h2>Water & Sewer</h2>
            <div class="form-grid">
                <div class="form-group">
                    <label>Rate Increase %</label>
                    <input type="number" step="0.0001" id="water_sewer_rate_increase" />
                </div>
            </div>
        </div>

        <div class="button-row">
            <button class="btn-primary" onclick="saveDefaults()">Save Defaults</button>
        </div>
    </div>

    <div class="toast" id="toast"></div>

    <script>
        const data = {{ data | safe }};

        function loadUI() {
            // Payroll tax
            document.getElementById('fica').value = data.payroll_tax.FICA;
            document.getElementById('sui').value = data.payroll_tax.SUI;
            document.getElementById('fui').value = data.payroll_tax.FUI;
            document.getElementById('mta').value = data.payroll_tax.MTA;
            document.getElementById('nys_disability').value = data.payroll_tax.NYS_Disability;
            document.getElementById('pfl').value = data.payroll_tax.PFL;

            // Union benefits
            document.getElementById('welfare_monthly').value = data.union_benefits.welfare_monthly;
            document.getElementById('pension_weekly').value = data.union_benefits.pension_weekly;
            document.getElementById('supp_retirement_weekly').value = data.union_benefits.supp_retirement_weekly;
            document.getElementById('legal_monthly').value = data.union_benefits.legal_monthly;
            document.getElementById('training_monthly').value = data.union_benefits.training_monthly;
            document.getElementById('profit_sharing_quarterly').value = data.union_benefits.profit_sharing_quarterly;

            // Workers comp
            document.getElementById('workers_comp_percent').value = data.workers_comp.percent;

            // Wage increase
            document.getElementById('wage_increase_percent').value = data.wage_increase.percent;
            document.getElementById('wage_increase_effective_week').value = data.wage_increase.effective_week;
            document.getElementById('wage_increase_pre_weeks').value = data.wage_increase.pre_increase_weeks;
            document.getElementById('wage_increase_post_weeks').value = data.wage_increase.post_increase_weeks;

            // Insurance renewal
            document.getElementById('insurance_renewal_increase_percent').value = data.insurance_renewal.increase_percent;
            document.getElementById('insurance_renewal_effective_date').value = data.insurance_renewal.effective_date;
            document.getElementById('insurance_renewal_pre_months').value = data.insurance_renewal.pre_renewal_months;
            document.getElementById('insurance_renewal_post_months').value = data.insurance_renewal.post_renewal_months;

            // Energy
            if (data.energy) {
                document.getElementById('energy_gas_esco_rate').value = data.energy.gas_esco_rate || 0;
                document.getElementById('energy_electric_esco_rate').value = data.energy.electric_esco_rate || 0;
                document.getElementById('energy_oil_price').value = data.energy.oil_price_per_gallon || 0;
                document.getElementById('energy_gas_rate_increase').value = data.energy.gas_rate_increase || 0;
                document.getElementById('energy_electric_rate_increase').value = data.energy.electric_rate_increase || 0;
                document.getElementById('energy_oil_rate_increase').value = data.energy.oil_rate_increase || 0;
                document.getElementById('energy_consumption_basis').value = data.energy.consumption_basis || '2-Year Average';
            }

            // Water & Sewer
            if (data.water_sewer) {
                document.getElementById('water_sewer_rate_increase').value = data.water_sewer.rate_increase || 0;
            }
        }

        function saveDefaults() {
            const payload = {
                payroll_tax: {
                    FICA: parseFloat(document.getElementById('fica').value),
                    SUI: parseFloat(document.getElementById('sui').value),
                    FUI: parseFloat(document.getElementById('fui').value),
                    MTA: parseFloat(document.getElementById('mta').value),
                    NYS_Disability: parseFloat(document.getElementById('nys_disability').value),
                    PFL: parseFloat(document.getElementById('pfl').value)
                },
                union_benefits: {
                    welfare_monthly: parseFloat(document.getElementById('welfare_monthly').value),
                    pension_weekly: parseFloat(document.getElementById('pension_weekly').value),
                    supp_retirement_weekly: parseFloat(document.getElementById('supp_retirement_weekly').value),
                    legal_monthly: parseFloat(document.getElementById('legal_monthly').value),
                    training_monthly: parseFloat(document.getElementById('training_monthly').value),
                    profit_sharing_quarterly: parseFloat(document.getElementById('profit_sharing_quarterly').value)
                },
                workers_comp: {
                    percent: parseFloat(document.getElementById('workers_comp_percent').value)
                },
                wage_increase: {
                    percent: parseFloat(document.getElementById('wage_increase_percent').value),
                    effective_week: document.getElementById('wage_increase_effective_week').value,
                    pre_increase_weeks: parseInt(document.getElementById('wage_increase_pre_weeks').value),
                    post_increase_weeks: parseInt(document.getElementById('wage_increase_post_weeks').value)
                },
                insurance_renewal: {
                    increase_percent: parseFloat(document.getElementById('insurance_renewal_increase_percent').value),
                    effective_date: document.getElementById('insurance_renewal_effective_date').value,
                    pre_renewal_months: parseInt(document.getElementById('insurance_renewal_pre_months').value),
                    post_renewal_months: parseInt(document.getElementById('insurance_renewal_post_months').value)
                },
                energy: {
                    gas_esco_rate: parseFloat(document.getElementById('energy_gas_esco_rate').value) || 0,
                    electric_esco_rate: parseFloat(document.getElementById('energy_electric_esco_rate').value) || 0,
                    oil_price_per_gallon: parseFloat(document.getElementById('energy_oil_price').value) || 0,
                    gas_rate_increase: parseFloat(document.getElementById('energy_gas_rate_increase').value) || 0,
                    electric_rate_increase: parseFloat(document.getElementById('energy_electric_rate_increase').value) || 0,
                    oil_rate_increase: parseFloat(document.getElementById('energy_oil_rate_increase').value) || 0,
                    consumption_basis: document.getElementById('energy_consumption_basis').value
                },
                water_sewer: {
                    rate_increase: parseFloat(document.getElementById('water_sewer_rate_increase').value) || 0
                }
            };

            fetch('/api/defaults', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(payload)
            })
            .then(() => {
                showToast('Defaults saved successfully');
            })
            .catch(e => console.error(e));
        }

        function showToast(msg) {
            const toast = document.getElementById('toast');
            toast.textContent = msg;
            toast.classList.add('show');
            setTimeout(() => toast.classList.remove('show'), 2000);
        }

        loadUI();
    </script>
</body>
</html>
"""

ASSUMPTIONS_BUILDINGS_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin><link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
    <title>Building Assumptions</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; background: #f5f5f5; }
        header { background: #5a4a3f; color: white; padding: 24px; }
        header h1 { font-size: 24px; font-weight: 600; }
        header .back-link { color: white; margin-bottom: 8px; }
        .main { display: flex; height: calc(100vh - 80px); }
        .sidebar {
            width: 280px;
            background: white;
            border-right: 1px solid #e0e0e0;
            overflow-y: auto;
            padding: 16px;
        }
        .sidebar-header { font-size: 13px; font-weight: 600; color: #666; text-transform: uppercase; margin-bottom: 12px; }
        .sidebar-search { display: flex; margin-bottom: 12px; }
        .sidebar-search input { flex: 1; padding: 8px; border: 1px solid #ddd; border-radius: 4px; font-size: 13px; }
        .sidebar-list { display: flex; flex-direction: column; gap: 4px; }
        .building-item {
            padding: 12px;
            border: 1px solid #e0e0e0;
            border-radius: 4px;
            cursor: pointer;
            font-size: 13px;
            background: white;
        }
        .building-item:hover { background: #f9f9f9; border-color: #5a4a3f; }
        .building-item.active { background: #5a4a3f; color: white; border-color: #5a4a3f; }
        .content {
            flex: 1;
            padding: 40px;
            overflow-y: auto;
        }
        .back-link { display: inline-block; margin-bottom: 24px; color: #5a4a3f; text-decoration: none; font-size: 14px; }
        .back-link:hover { text-decoration: underline; }
        .building-title { font-size: 24px; color: #5a4a3f; margin-bottom: 8px; font-weight: 600; }
        .building-code { font-size: 13px; color: #999; margin-bottom: 24px; }
        .tabs {
            display: flex;
            gap: 8px;
            border-bottom: 1px solid #e0e0e0;
            margin-bottom: 24px;
        }
        .tab {
            padding: 12px 16px;
            border: none;
            background: none;
            cursor: pointer;
            font-size: 14px;
            font-weight: 500;
            color: #666;
            border-bottom: 2px solid transparent;
        }
        .tab:hover { color: #5a4a3f; }
        .tab.active { color: #5a4a3f; border-bottom-color: #5a4a3f; }
        .tab-content { display: none; }
        .tab-content.active { display: block; }
        .section { background: white; border-radius: 8px; padding: 24px; border: 1px solid #e0e0e0; margin-bottom: 24px; }
        .section h3 { font-size: 16px; color: #5a4a3f; margin-bottom: 16px; font-weight: 600; }
        table { width: 100%; border-collapse: collapse; font-size: 14px; }
        th { background: #f5f5f5; padding: 12px; text-align: left; font-weight: 600; color: #333; border-bottom: 1px solid #e0e0e0; }
        td { padding: 12px; border-bottom: 1px solid #e0e0e0; }
        input { padding: 8px; border: 1px solid #ddd; border-radius: 4px; font-size: 13px; }
        input[type="number"] { font-family: 'Courier New', monospace; text-align: right; }
        input:focus { outline: none; border-color: #5a4a3f; box-shadow: 0 0 0 2px rgba(90,74,63, 0.1); }
        .form-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; }
        .form-group { display: flex; flex-direction: column; }
        .form-group label { font-size: 13px; font-weight: 500; color: #333; margin-bottom: 6px; }
        .button-row { display: flex; gap: 12px; margin-top: 24px; }
        button { padding: 10px 20px; border: none; border-radius: 4px; font-size: 14px; font-weight: 500; cursor: pointer; }
        .btn-primary { background: #5a4a3f; color: white; }
        .btn-primary:hover { background: #0f1f38; }
        .btn-secondary { background: #e0e0e0; color: #333; }
        .btn-secondary:hover { background: #d0d0d0; }
        .toast { position: fixed; bottom: 20px; right: 20px; background: #4caf50; color: white; padding: 16px 20px; border-radius: 4px; display: none; z-index: 1000; }
        .toast.show { display: block; animation: slideIn 0.3s ease; }
        @keyframes slideIn { from { transform: translateX(400px); } to { transform: translateX(0); } }
        .empty { padding: 40px; text-align: center; color: #999; }
    </style>
</head>
<body>
    <header>
        <a href="/" class="back-link">← Home</a>
        <h1>Building Assumptions</h1>
    </header>
    <div class="main">
        <div class="sidebar">
            <div class="sidebar-header">Buildings</div>
            <div class="sidebar-search">
                <input type="text" id="searchInput" placeholder="Search..." />
            </div>
            <div class="sidebar-list" id="buildingList"></div>
        </div>
        <div class="content">

            <div id="noBuildingSelected" class="empty">
                <p>Select a building from the list to view and edit assumptions</p>
            </div>

            <div id="buildingContent" style="display: none;">
                <div class="building-title" id="buildingName"></div>
                <div class="building-code" id="buildingCode"></div>

                <div class="tabs">
                    <button class="tab active" onclick="switchTab('payroll')">Payroll</button>
                    <button class="tab" onclick="switchTab('income')">Income</button>
                    <button class="tab" onclick="switchTab('insurance')">Insurance</button>
                    <button class="tab" onclick="switchTab('utilities')">Utilities</button>
                </div>

                <!-- Payroll Tab -->
                <div id="payroll" class="tab-content active">
                    <div class="section">
                        <h3>Position Details</h3>
                        <table>
                            <thead>
                                <tr>
                                    <th>Position</th>
                                    <th># Employees</th>
                                    <th>Hourly Rate</th>
                                </tr>
                            </thead>
                            <tbody id="payrollTable"></tbody>
                        </table>
                    </div>
                </div>

                <!-- Income Tab -->
                <div id="income" class="tab-content">
                    <div class="section">
                        <h3>Maintenance</h3>
                        <div class="form-grid">
                            <div class="form-group">
                                <label>Total Shares</label>
                                <input type="number" id="income_maint_shares" />
                            </div>
                            <div class="form-group">
                                <label>$/Share/Mo</label>
                                <input type="number" step="0.01" id="income_maint_per_share" />
                            </div>
                            <div class="form-group">
                                <label>Increase %</label>
                                <input type="number" step="0.0001" id="income_maint_increase" />
                            </div>
                        </div>
                    </div>

                    <div class="section">
                        <h3>Storage Units</h3>
                        <table>
                            <thead>
                                <tr>
                                    <th>Size Label</th>
                                    <th># Units</th>
                                    <th># Occupied</th>
                                    <th>$/Month</th>
                                </tr>
                            </thead>
                            <tbody id="storageTable"></tbody>
                        </table>
                    </div>

                    <div class="section">
                        <h3>Bike Storage</h3>
                        <div class="form-grid">
                            <div class="form-group">
                                <label># Racks</label>
                                <input type="number" id="income_bike_racks" />
                            </div>
                            <div class="form-group">
                                <label># Occupied</label>
                                <input type="number" id="income_bike_occupied" />
                            </div>
                            <div class="form-group">
                                <label>$/Month</label>
                                <input type="number" step="0.01" id="income_bike_monthly" />
                            </div>
                        </div>
                    </div>

                    <div class="section">
                        <h3>Laundry/Vending</h3>
                        <div class="form-grid">
                            <div class="form-group">
                                <label>Contract Description</label>
                                <input type="text" id="income_laundry_desc" />
                            </div>
                            <div class="form-group">
                                <label>Monthly Amount</label>
                                <input type="number" step="0.01" id="income_laundry_monthly" />
                            </div>
                        </div>
                    </div>
                </div>

                <!-- Insurance Tab -->
                <div id="insurance" class="tab-content">
                    <div class="section">
                        <h3>Insurance Policies</h3>
                        <table>
                            <thead>
                                <tr>
                                    <th>GL Code</th>
                                    <th>Policy Name</th>
                                    <th>Current Annual Premium</th>
                                    <th>Current Year Budget</th>
                                    <th>Expiration Date</th>
                                    <th>Override Increase %</th>
                                </tr>
                            </thead>
                            <tbody id="insuranceTable"></tbody>
                        </table>
                    </div>
                </div>

                <!-- Utilities Tab -->
                <div id="utilities" class="tab-content">
                    <div class="section">
                        <h3>Energy — Rate Overrides</h3>
                        <p style="font-size:13px;color:#666;margin-bottom:16px;">Leave blank or 0 to use portfolio defaults. Only enter values here to override for this building.</p>
                        <div class="form-grid">
                            <div class="form-group">
                                <label>Gas ESCO Rate ($/Therm)</label>
                                <input type="number" step="0.0001" id="util_gas_esco_rate" />
                            </div>
                            <div class="form-group">
                                <label>Electric ESCO Rate ($/kWh)</label>
                                <input type="number" step="0.0001" id="util_electric_esco_rate" />
                            </div>
                            <div class="form-group">
                                <label>Gas Rate Increase % Override</label>
                                <input type="number" step="0.0001" id="util_gas_rate_increase" />
                            </div>
                            <div class="form-group">
                                <label>Electric Rate Increase % Override</label>
                                <input type="number" step="0.0001" id="util_electric_rate_increase" />
                            </div>
                            <div class="form-group">
                                <label>Oil/Fuel Price ($/Gallon) Override</label>
                                <input type="number" step="0.01" id="util_oil_price" />
                            </div>
                            <div class="form-group">
                                <label>Oil/Fuel Rate Increase % Override</label>
                                <input type="number" step="0.0001" id="util_oil_rate_increase" />
                            </div>
                            <div class="form-group">
                                <label>Consumption Basis</label>
                                <select id="util_consumption_basis">
                                    <option value="">Use Default</option>
                                    <option value="2-Year Average">2-Year Average</option>
                                    <option value="Prior Year">Prior Year</option>
                                    <option value="Custom">Custom</option>
                                </select>
                            </div>
                        </div>
                    </div>

                    <div class="section">
                        <h3>Energy — GL Account Adjustments</h3>
                        <table>
                            <thead>
                                <tr>
                                    <th>GL Code</th>
                                    <th>Description</th>
                                    <th>Accrual Adjustment</th>
                                    <th>Unpaid Bills</th>
                                    <th>Rate Increase % Override</th>
                                </tr>
                            </thead>
                            <tbody id="energyGLTable"></tbody>
                        </table>
                    </div>

                    <div class="section">
                        <h3>Water & Sewer</h3>
                        <div class="form-grid">
                            <div class="form-group">
                                <label>Rate Increase % Override</label>
                                <input type="number" step="0.0001" id="util_water_rate_increase" />
                            </div>
                        </div>
                    </div>

                    <div class="section">
                        <h3>Water & Sewer — GL Account Adjustments</h3>
                        <table>
                            <thead>
                                <tr>
                                    <th>GL Code</th>
                                    <th>Description</th>
                                    <th>Accrual Adjustment</th>
                                    <th>Unpaid Bills</th>
                                    <th>Rate Increase % Override</th>
                                </tr>
                            </thead>
                            <tbody id="waterGLTable"></tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <div class="toast" id="toast"></div>

    <script>
        const buildings = {{ buildings | safe }};
        const policies = {{ policies | safe }};
        let currentBuilding = null;
        let buildingData = {};
        let debounceTimer = null;

        const ENERGY_GL_ACCOUNTS = [
            { gl: '5252-0000', desc: 'Gas - Heating' },
            { gl: '5252-0001', desc: 'Gas - Cooking' },
            { gl: '5252-0010', desc: 'Gas - Common Area' },
            { gl: '5253-0000', desc: 'Oil / Fuel' },
            { gl: '5250-0000', desc: 'Electric' }
        ];
        const WATER_GL_ACCOUNTS = [
            { gl: '6305-0000', desc: 'Water/Sewer' },
            { gl: '6305-0010', desc: 'Water - Common Area' },
            { gl: '6305-0020', desc: 'Sewer Charges' }
        ];

        function loadBuildingsList() {
            const list = document.getElementById('buildingList');
            list.innerHTML = buildings.map(b =>
                `<div class="building-item" onclick="selectBuilding('${b.entity_code}', '${b.building_name}')">
                    <strong>${b.entity_code}</strong><br/>
                    ${b.building_name.substring(0, 40)}
                </div>`
            ).join('');
        }

        function selectBuilding(code, name) {
            currentBuilding = code;
            document.querySelectorAll('.building-item').forEach(el => el.classList.remove('active'));
            event.currentTarget.classList.add('active');

            fetch(`/api/building-assumptions/${code}`)
                .then(r => r.json())
                .then(data => {
                    buildingData = data;
                    document.getElementById('noBuildingSelected').style.display = 'none';
                    document.getElementById('buildingContent').style.display = 'block';
                    document.getElementById('buildingName').textContent = name;
                    document.getElementById('buildingCode').textContent = `Entity Code: ${code}`;
                    loadPayrollTab();
                    loadIncomeTab();
                    loadInsuranceTab();
                    loadUtilitiesTab();
                });
        }

        function switchTab(tabName) {
            document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
            document.querySelectorAll('.tab').forEach(el => el.classList.remove('active'));
            document.getElementById(tabName).classList.add('active');
            event.currentTarget.classList.add('active');
        }

        function loadPayrollTab() {
            const payroll = buildingData.payroll || {};
            const positions = payroll.positions || [
                { name: 'Resident Manager', employee_count: 0, hourly_rate: 0 },
                { name: 'Handyman', employee_count: 0, hourly_rate: 0 },
                { name: 'Porter/Doorman', employee_count: 0, hourly_rate: 0 },
                { name: 'Porter 80%', employee_count: 0, hourly_rate: 0 },
                { name: 'Employee 5', employee_count: 0, hourly_rate: 0 },
                { name: 'Employee 6', employee_count: 0, hourly_rate: 0 },
                { name: 'Employee 7', employee_count: 0, hourly_rate: 0 },
                { name: 'Employee 8', employee_count: 0, hourly_rate: 0 }
            ];

            const table = document.getElementById('payrollTable');
            table.innerHTML = positions.map((p, i) => `
                <tr>
                    <td><input type="text" value="${p.name}" class="payroll-name" data-idx="${i}" /></td>
                    <td><input type="number" value="${p.employee_count || 0}" class="payroll-count" data-idx="${i}" /></td>
                    <td><input type="number" step="0.01" value="${p.hourly_rate || 0}" class="payroll-rate" data-idx="${i}" /></td>
                </tr>
            `).join('');

            document.querySelectorAll('.payroll-name, .payroll-count, .payroll-rate').forEach(el => {
                el.addEventListener('change', savePayroll);
            });
        }

        function savePayroll() {
            const positions = [];
            document.querySelectorAll('#payrollTable tr').forEach(tr => {
                const name = tr.querySelector('.payroll-name').value;
                const count = parseInt(tr.querySelector('.payroll-count').value) || 0;
                const rate = parseFloat(tr.querySelector('.payroll-rate').value) || 0;
                positions.push({ name, employee_count: count, hourly_rate: rate });
            });

            buildingData.payroll = { positions };
            debouncedSave();
        }

        function loadIncomeTab() {
            const income = buildingData.income || {};
            const maint = income.maintenance || { total_shares: 0, per_share_monthly: 0, increase_percent: 0 };
            const storage = income.storage || [];
            const bike = income.bike_storage || { racks: 0, occupied: 0, monthly: 0 };
            const laundry = income.laundry_vending || { description: '', monthly: 0 };

            document.getElementById('income_maint_shares').value = maint.total_shares || 0;
            document.getElementById('income_maint_per_share').value = maint.per_share_monthly || 0;
            document.getElementById('income_maint_increase').value = maint.increase_percent || 0;

            const storageTable = document.getElementById('storageTable');
            const storageRows = storage.length > 0 ? storage : [
                { size_label: 'Small', units: 0, occupied: 0, monthly: 0 },
                { size_label: 'Medium', units: 0, occupied: 0, monthly: 0 },
                { size_label: 'Large', units: 0, occupied: 0, monthly: 0 },
                { size_label: 'XL', units: 0, occupied: 0, monthly: 0 }
            ];
            storageTable.innerHTML = storageRows.map((s, i) => `
                <tr>
                    <td><input type="text" value="${s.size_label || ''}" class="storage-label" data-idx="${i}" /></td>
                    <td><input type="number" value="${s.units || 0}" class="storage-units" data-idx="${i}" /></td>
                    <td><input type="number" value="${s.occupied || 0}" class="storage-occupied" data-idx="${i}" /></td>
                    <td><input type="number" step="0.01" value="${s.monthly || 0}" class="storage-monthly" data-idx="${i}" /></td>
                </tr>
            `).join('');

            document.getElementById('income_bike_racks').value = bike.racks || 0;
            document.getElementById('income_bike_occupied').value = bike.occupied || 0;
            document.getElementById('income_bike_monthly').value = bike.monthly || 0;

            document.getElementById('income_laundry_desc').value = laundry.description || '';
            document.getElementById('income_laundry_monthly').value = laundry.monthly || 0;

            document.querySelectorAll('.storage-label, .storage-units, .storage-occupied, .storage-monthly').forEach(el => {
                el.addEventListener('change', saveIncome);
            });
            document.querySelectorAll('#income_maint_shares, #income_maint_per_share, #income_maint_increase, #income_bike_racks, #income_bike_occupied, #income_bike_monthly, #income_laundry_desc, #income_laundry_monthly').forEach(el => {
                el.addEventListener('change', saveIncome);
            });
        }

        function saveIncome() {
            const income = {
                maintenance: {
                    total_shares: parseFloat(document.getElementById('income_maint_shares').value) || 0,
                    per_share_monthly: parseFloat(document.getElementById('income_maint_per_share').value) || 0,
                    increase_percent: parseFloat(document.getElementById('income_maint_increase').value) || 0
                },
                storage: Array.from(document.querySelectorAll('#storageTable tr')).map(tr => ({
                    size_label: tr.querySelector('.storage-label').value,
                    units: parseInt(tr.querySelector('.storage-units').value) || 0,
                    occupied: parseInt(tr.querySelector('.storage-occupied').value) || 0,
                    monthly: parseFloat(tr.querySelector('.storage-monthly').value) || 0
                })),
                bike_storage: {
                    racks: parseInt(document.getElementById('income_bike_racks').value) || 0,
                    occupied: parseInt(document.getElementById('income_bike_occupied').value) || 0,
                    monthly: parseFloat(document.getElementById('income_bike_monthly').value) || 0
                },
                laundry_vending: {
                    description: document.getElementById('income_laundry_desc').value,
                    monthly: parseFloat(document.getElementById('income_laundry_monthly').value) || 0
                }
            };

            buildingData.income = income;
            debouncedSave();
        }

        function loadInsuranceTab() {
            const insurance = buildingData.insurance || [];
            const table = document.getElementById('insuranceTable');

            table.innerHTML = policies.map((policy, i) => {
                const existing = insurance.find(p => p.gl_code === policy.gl_code) || {};
                return `
                    <tr>
                        <td>${policy.gl_code}</td>
                        <td>${policy.name}</td>
                        <td><input type="number" step="0.01" value="${existing.current_premium || 0}" class="ins-current-premium" data-gl="${policy.gl_code}" /></td>
                        <td><input type="number" step="0.01" value="${existing.current_budget || 0}" class="ins-budget" data-gl="${policy.gl_code}" /></td>
                        <td><input type="text" value="${existing.expiration_date || ''}" class="ins-expiration" data-gl="${policy.gl_code}" /></td>
                        <td><input type="number" step="0.0001" value="${existing.override_increase || 0}" class="ins-override-increase" data-gl="${policy.gl_code}" /></td>
                    </tr>
                `;
            }).join('');

            document.querySelectorAll('.ins-current-premium, .ins-budget, .ins-expiration, .ins-override-increase').forEach(el => {
                el.addEventListener('change', saveInsurance);
            });
        }

        function saveInsurance() {
            const insurance = policies.map(policy => {
                const gl = policy.gl_code;
                return {
                    gl_code: gl,
                    name: policy.name,
                    current_premium: parseFloat(document.querySelector(`.ins-current-premium[data-gl="${gl}"]`).value) || 0,
                    current_budget: parseFloat(document.querySelector(`.ins-budget[data-gl="${gl}"]`).value) || 0,
                    expiration_date: document.querySelector(`.ins-expiration[data-gl="${gl}"]`).value,
                    override_increase: parseFloat(document.querySelector(`.ins-override-increase[data-gl="${gl}"]`).value) || 0
                };
            });

            buildingData.insurance = insurance;
            debouncedSave();
        }

        function loadUtilitiesTab() {
            const energy = buildingData.energy || {};
            const water = buildingData.water_sewer || {};

            document.getElementById('util_gas_esco_rate').value = energy.gas_esco_rate || 0;
            document.getElementById('util_electric_esco_rate').value = energy.electric_esco_rate || 0;
            document.getElementById('util_oil_price').value = energy.oil_price_per_gallon || 0;
            document.getElementById('util_gas_rate_increase').value = energy.gas_rate_increase || 0;
            document.getElementById('util_electric_rate_increase').value = energy.electric_rate_increase || 0;
            document.getElementById('util_oil_rate_increase').value = energy.oil_rate_increase || 0;
            document.getElementById('util_consumption_basis').value = energy.consumption_basis || '';
            document.getElementById('util_water_rate_increase').value = water.rate_increase || 0;

            const energyAdj = energy.gl_adjustments || {};
            const energyTable = document.getElementById('energyGLTable');
            energyTable.innerHTML = ENERGY_GL_ACCOUNTS.map(a => {
                const adj = energyAdj[a.gl] || {};
                return `<tr>
                    <td>${a.gl}</td>
                    <td>${a.desc}</td>
                    <td><input type="number" step="0.01" value="${adj.accrual || 0}" class="energy-accrual" data-gl="${a.gl}" /></td>
                    <td><input type="number" step="0.01" value="${adj.unpaid || 0}" class="energy-unpaid" data-gl="${a.gl}" /></td>
                    <td><input type="number" step="0.0001" value="${adj.rate_increase || 0}" class="energy-gl-rate" data-gl="${a.gl}" /></td>
                </tr>`;
            }).join('');

            const waterAdj = water.gl_adjustments || {};
            const waterTable = document.getElementById('waterGLTable');
            waterTable.innerHTML = WATER_GL_ACCOUNTS.map(a => {
                const adj = waterAdj[a.gl] || {};
                return `<tr>
                    <td>${a.gl}</td>
                    <td>${a.desc}</td>
                    <td><input type="number" step="0.01" value="${adj.accrual || 0}" class="water-accrual" data-gl="${a.gl}" /></td>
                    <td><input type="number" step="0.01" value="${adj.unpaid || 0}" class="water-unpaid" data-gl="${a.gl}" /></td>
                    <td><input type="number" step="0.0001" value="${adj.rate_increase || 0}" class="water-gl-rate" data-gl="${a.gl}" /></td>
                </tr>`;
            }).join('');

            // Bind change events
            document.querySelectorAll('#util_gas_esco_rate, #util_electric_esco_rate, #util_oil_price, #util_gas_rate_increase, #util_electric_rate_increase, #util_oil_rate_increase, #util_consumption_basis, #util_water_rate_increase').forEach(el => {
                el.addEventListener('change', saveUtilities);
            });
            document.querySelectorAll('.energy-accrual, .energy-unpaid, .energy-gl-rate, .water-accrual, .water-unpaid, .water-gl-rate').forEach(el => {
                el.addEventListener('change', saveUtilities);
            });
        }

        function saveUtilities() {
            const energyGLAdj = {};
            ENERGY_GL_ACCOUNTS.forEach(a => {
                energyGLAdj[a.gl] = {
                    accrual: parseFloat(document.querySelector(`.energy-accrual[data-gl="${a.gl}"]`).value) || 0,
                    unpaid: parseFloat(document.querySelector(`.energy-unpaid[data-gl="${a.gl}"]`).value) || 0,
                    rate_increase: parseFloat(document.querySelector(`.energy-gl-rate[data-gl="${a.gl}"]`).value) || 0
                };
            });

            const waterGLAdj = {};
            WATER_GL_ACCOUNTS.forEach(a => {
                waterGLAdj[a.gl] = {
                    accrual: parseFloat(document.querySelector(`.water-accrual[data-gl="${a.gl}"]`).value) || 0,
                    unpaid: parseFloat(document.querySelector(`.water-unpaid[data-gl="${a.gl}"]`).value) || 0,
                    rate_increase: parseFloat(document.querySelector(`.water-gl-rate[data-gl="${a.gl}"]`).value) || 0
                };
            });

            buildingData.energy = {
                gas_esco_rate: parseFloat(document.getElementById('util_gas_esco_rate').value) || 0,
                electric_esco_rate: parseFloat(document.getElementById('util_electric_esco_rate').value) || 0,
                oil_price_per_gallon: parseFloat(document.getElementById('util_oil_price').value) || 0,
                gas_rate_increase: parseFloat(document.getElementById('util_gas_rate_increase').value) || 0,
                electric_rate_increase: parseFloat(document.getElementById('util_electric_rate_increase').value) || 0,
                oil_rate_increase: parseFloat(document.getElementById('util_oil_rate_increase').value) || 0,
                consumption_basis: document.getElementById('util_consumption_basis').value,
                gl_adjustments: energyGLAdj
            };

            buildingData.water_sewer = {
                rate_increase: parseFloat(document.getElementById('util_water_rate_increase').value) || 0,
                gl_adjustments: waterGLAdj
            };

            debouncedSave();
        }

        function debouncedSave() {
            clearTimeout(debounceTimer);
            debounceTimer = setTimeout(() => {
                fetch(`/api/building-assumptions/${currentBuilding}`, {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify(buildingData)
                }).then(() => showToast('Saved'));
            }, 500);
        }

        function showToast(msg) {
            const toast = document.getElementById('toast');
            toast.textContent = msg;
            toast.classList.add('show');
            setTimeout(() => toast.classList.remove('show'), 2000);
        }

        document.getElementById('searchInput').addEventListener('input', (e) => {
            const search = e.target.value.toLowerCase();
            document.querySelectorAll('.building-item').forEach(item => {
                const text = item.textContent.toLowerCase();
                item.style.display = text.includes(search) ? '' : 'none';
            });
        });

        loadBuildingsList();
    </script>
</body>
</html>
"""


ASSUMPTIONS_WORKBENCH_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&display=swap" rel="stylesheet">
<title>Assumptions Workbench — Century Management</title>
<style>
/* Force scrollbars always visible (fixes macOS auto-hide) */
::-webkit-scrollbar { width: 12px; height: 12px; -webkit-appearance: none; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 6px; border: 2px solid #f1f5f9; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }
::-webkit-scrollbar-corner { background: #f1f5f9; }
* { scrollbar-width: thin; scrollbar-color: #cbd5e1 #f1f5f9; }
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; background: #f5f5f5; color: #222; font-size: 14px; }
  header { background: #5a4a3f; color: white; padding: 18px 24px; display: flex; align-items: center; gap: 20px; }
  header .back-link { color: white; text-decoration: none; font-size: 13px; opacity: 0.85; }
  header .back-link:hover { opacity: 1; }
  header h1 { font-size: 20px; font-weight: 600; }
  header .hint { margin-left: auto; font-size: 12px; opacity: 0.7; font-style: italic; }
  .main { display: flex; height: calc(100vh - 60px); }
  .sidebar { width: 300px; background: white; border-right: 1px solid #e0e0e0; overflow-y: auto; display: flex; flex-direction: column; }
  .defaults-pin { padding: 14px 16px; background: linear-gradient(135deg, #f8f5f1 0%, #ede5d9 100%); border-bottom: 2px solid #d4c5ae; cursor: pointer; display: flex; align-items: center; gap: 10px; }
  .defaults-pin:hover { background: linear-gradient(135deg, #f0ebe3 0%, #e5dbc9 100%); }
  .defaults-pin.active { background: #5a4a3f; color: white; border-bottom-color: #3d342c; }
  .defaults-pin .icon { font-size: 18px; }
  .defaults-pin .label { font-weight: 600; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px; }
  .defaults-pin .sub { font-size: 11px; opacity: 0.7; margin-top: 2px; }
  .sidebar-search { padding: 12px 16px 8px; }
  .sidebar-search input { width: 100%; padding: 8px 10px; border: 1px solid #ddd; border-radius: 4px; font-size: 13px; }
  .sidebar-search input:focus { outline: none; border-color: #5a4a3f; box-shadow: 0 0 0 2px rgba(90,74,63,0.1); }
  .sidebar-list { flex: 1; padding: 0 8px 12px; overflow-y: auto; }
  .building-item { padding: 10px 12px; border-radius: 6px; cursor: pointer; font-size: 13px; display: flex; justify-content: space-between; align-items: center; margin-bottom: 2px; }
  .building-item:hover { background: #f5f0e9; }
  .building-item.active { background: #5a4a3f; color: white; }
  .building-item .building-code { font-weight: 600; margin-right: 8px; }
  .override-badge { background: #e8935a; color: white; font-size: 10px; font-weight: 700; padding: 2px 7px; border-radius: 10px; min-width: 18px; text-align: center; }
  .building-item.active .override-badge { background: #ffb680; color: #3d342c; }
  .content { flex: 1; overflow-y: auto; padding: 28px 36px; max-width: 1100px; }
  .panel-title { font-size: 22px; font-weight: 700; color: #2d2520; margin-bottom: 4px; }
  .panel-sub { font-size: 13px; color: #666; margin-bottom: 24px; }
  .tabs { display: flex; gap: 4px; border-bottom: 1px solid #e0e0e0; margin-bottom: 24px; overflow-x: auto; }
  .tab { padding: 10px 16px; border: none; background: none; cursor: pointer; font-size: 13px; font-weight: 500; color: #666; border-bottom: 2px solid transparent; white-space: nowrap; font-family: inherit; }
  .tab:hover { color: #5a4a3f; }
  .tab.active { color: #5a4a3f; border-bottom-color: #5a4a3f; }
  .tab .badge-inline { background: #e8935a; color: white; font-size: 10px; font-weight: 700; padding: 1px 6px; border-radius: 8px; margin-left: 6px; }
  .section { background: white; border-radius: 8px; padding: 20px 24px; margin-bottom: 18px; border: 1px solid #e0e0e0; }
  .section h2 { font-size: 15px; color: #5a4a3f; margin-bottom: 14px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.3px; }
  .section h3 { font-size: 14px; color: #5a4a3f; margin-bottom: 12px; font-weight: 600; }
  .form-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 14px; }
  .form-group { display: flex; flex-direction: column; }
  .form-group label { font-size: 12px; font-weight: 500; color: #555; margin-bottom: 5px; }
  .form-group input, .form-group select { padding: 8px 10px; border: 1px solid #ddd; border-radius: 4px; font-size: 13px; font-family: 'Courier New', monospace; }
  .form-group input:focus, .form-group select:focus { outline: none; border-color: #5a4a3f; box-shadow: 0 0 0 2px rgba(90,74,63,0.1); }
  .override-field { padding: 12px; border: 1px solid #e0e0e0; border-radius: 6px; background: #fafafa; }
  .override-field.has-override { background: #fff7ef; border-left: 3px solid #e8935a; border-color: #f0d4b3 #f0d4b3 #f0d4b3 #e8935a; }
  .override-field label { font-size: 12px; font-weight: 600; color: #333; margin-bottom: 6px; display: block; }
  .override-field .default-chip { font-size: 11px; color: #666; margin-bottom: 6px; }
  .override-field .default-chip .val { font-family: 'Courier New', monospace; font-weight: 600; color: #2d2520; }
  .override-field input, .override-field select { width: 100%; padding: 7px 10px; border: 1px solid #ddd; border-radius: 4px; font-size: 13px; font-family: 'Courier New', monospace; background: white; }
  .override-field.has-override input, .override-field.has-override select { border-color: #e8935a; background: #fffaf3; }
  .override-field .status { font-size: 11px; margin-top: 5px; display: flex; align-items: center; gap: 5px; }
  .override-field .status.override { color: #c96a1e; font-weight: 600; }
  .override-field .status.using-default { color: #888; font-style: italic; }
  .override-field-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 12px; }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { background: #f5f0e9; padding: 10px; text-align: left; font-weight: 600; color: #3d342c; border-bottom: 1px solid #d4c5ae; font-size: 11px; text-transform: uppercase; letter-spacing: 0.3px; }
  td { padding: 10px; border-bottom: 1px solid #eee; }
  td input { width: 100%; padding: 6px 8px; border: 1px solid #ddd; border-radius: 3px; font-size: 13px; font-family: 'Courier New', monospace; }
  td input[type="text"] { font-family: inherit; }
  .button-row { display: flex; gap: 10px; margin-top: 20px; padding-top: 16px; border-top: 1px solid #e0e0e0; }
  button { padding: 9px 18px; border: none; border-radius: 5px; font-size: 13px; font-weight: 600; cursor: pointer; font-family: inherit; }
  .btn-primary { background: #5a4a3f; color: white; }
  .btn-primary:hover { background: #3d342c; }
  .btn-ghost { background: transparent; color: #c96a1e; border: 1px solid #e8935a; }
  .btn-ghost:hover { background: #fff7ef; }
  .summary-row { display: grid; grid-template-columns: 2fr 1fr 1fr 60px; gap: 16px; padding: 12px 0; border-bottom: 1px solid #eee; align-items: center; font-size: 13px; }
  .summary-row:last-child { border-bottom: none; }
  .summary-row .field-label { font-weight: 500; }
  .summary-row .field-path { font-size: 11px; color: #888; }
  .summary-row .val { font-family: 'Courier New', monospace; }
  .summary-row .default-val { color: #888; }
  .summary-row .override-val { color: #c96a1e; font-weight: 600; }
  .summary-row .arrow { color: #888; text-align: center; }
  .summary-header { display: grid; grid-template-columns: 2fr 1fr 1fr 60px; gap: 16px; padding: 10px 0; border-bottom: 2px solid #5a4a3f; font-size: 11px; text-transform: uppercase; letter-spacing: 0.3px; font-weight: 600; color: #5a4a3f; }
  .empty { text-align: center; padding: 60px 20px; color: #999; }
  .empty-icon { font-size: 42px; margin-bottom: 12px; opacity: 0.3; }
  .toast { position: fixed; bottom: 20px; right: 20px; background: #4caf50; color: white; padding: 14px 20px; border-radius: 5px; display: none; z-index: 1000; font-weight: 500; }
  .toast.show { display: block; animation: slideIn 0.3s ease; }
  @keyframes slideIn { from { transform: translateX(400px); } to { transform: translateX(0); } }
</style>
</head>
<body>
<header>
  <a href="/" class="back-link">← Home</a>
  <h1>Assumptions</h1>
  <span class="hint">Portfolio defaults + building overrides</span>
</header>
<div class="main">
  <div class="sidebar">
    <div class="defaults-pin active" id="pinDefaults" onclick="selectDefaults()">
      <div class="icon">📌</div>
      <div>
        <div class="label">Portfolio Defaults</div>
        <div class="sub">Applies to all buildings</div>
      </div>
    </div>
    <div class="sidebar-search">
      <input type="text" id="search" placeholder="Search buildings..." oninput="filterBuildings()" />
    </div>
    <div class="sidebar-list" id="buildingList"></div>
  </div>
  <div class="content" id="content"></div>
</div>
<div class="toast" id="toast">✓ Saved</div>

<script>
const defaults = {{ defaults | safe }};
const buildings = {{ buildings | safe }};
const policies = {{ policies | safe }};
let overrideCounts = {{ override_counts | safe }};
let currentView = 'defaults';
let currentTab = 'utilities';
let currentBldgData = {};
let debounceTimer = null;

const ENERGY_GL_ACCOUNTS = [
  { gl: '5252-0000', desc: 'Gas - Heating' },
  { gl: '5252-0001', desc: 'Gas - Cooking' },
  { gl: '5252-0010', desc: 'Gas - Common Area' },
  { gl: '5253-0000', desc: 'Oil / Fuel' },
  { gl: '5250-0000', desc: 'Electric' }
];
const WATER_GL_ACCOUNTS = [
  { gl: '6305-0000', desc: 'Water/Sewer' },
  { gl: '6305-0010', desc: 'Water - Common Area' },
  { gl: '6305-0020', desc: 'Sewer Charges' }
];

const OVERRIDE_FIELDS = [
  { id: 'util_gas_esco_rate', label: 'Gas ESCO Rate', unit: '$/Therm', defaultPath: 'energy.gas_esco_rate', saveKey: 'energy.gas_esco_rate', type: 'number', step: '0.0001' },
  { id: 'util_electric_esco_rate', label: 'Electric ESCO Rate', unit: '$/kWh', defaultPath: 'energy.electric_esco_rate', saveKey: 'energy.electric_esco_rate', type: 'number', step: '0.0001' },
  { id: 'util_oil_price', label: 'Oil/Fuel Price', unit: '$/Gallon', defaultPath: 'energy.oil_price', saveKey: 'energy.oil_price_per_gallon', type: 'number', step: '0.01' },
  { id: 'util_gas_rate_increase', label: 'Gas Rate Increase', unit: '%', defaultPath: 'energy.gas_rate_increase', saveKey: 'energy.gas_rate_increase', type: 'number', step: '0.0001' },
  { id: 'util_electric_rate_increase', label: 'Electric Rate Increase', unit: '%', defaultPath: 'energy.electric_rate_increase', saveKey: 'energy.electric_rate_increase', type: 'number', step: '0.0001' },
  { id: 'util_oil_rate_increase', label: 'Oil Rate Increase', unit: '%', defaultPath: 'energy.oil_rate_increase', saveKey: 'energy.oil_rate_increase', type: 'number', step: '0.0001' },
  { id: 'util_consumption_basis', label: 'Consumption Basis', unit: '', defaultPath: 'energy.consumption_basis', saveKey: 'energy.consumption_basis', type: 'select', options: ['2-Year Average','Prior Year','Custom'] },
  { id: 'util_water_rate_increase', label: 'Water/Sewer Rate Increase', unit: '%', defaultPath: 'water_sewer.rate_increase', saveKey: 'water_sewer.rate_increase', type: 'number', step: '0.0001' }
];

function getPath(obj, path) { return path.split('.').reduce((o,k) => o && o[k] !== undefined ? o[k] : undefined, obj); }
function setPath(obj, path, val) { const keys = path.split('.'); let o = obj; for (let i=0;i<keys.length-1;i++){ if(!o[keys[i]]) o[keys[i]]={}; o=o[keys[i]];} o[keys[keys.length-1]] = val; }

function computeOverrideCount(bd) {
  let count = 0;
  for (const f of OVERRIDE_FIELDS) {
    const v = getPath(bd, f.saveKey);
    if (f.type === 'select') { if (v) count++; }
    else { if (v !== undefined && v !== null && v !== '' && parseFloat(v) !== 0) count++; }
  }
  return count;
}

function renderSidebar(filter) {
  filter = (filter || '').toLowerCase();
  const list = document.getElementById('buildingList');
  list.innerHTML = buildings
    .filter(b => !filter || b.entity_code.includes(filter) || (b.building_name||'').toLowerCase().includes(filter))
    .map(b => {
      const c = overrideCounts[b.entity_code] || 0;
      const active = currentView === b.entity_code ? 'active' : '';
      const nameEsc = (b.building_name||'').replace(/"/g,'&quot;');
      return `<div class="building-item ${active}" onclick="selectBuilding('${b.entity_code}','${nameEsc}')">
        <div><span class="building-code">${b.entity_code}</span>${(b.building_name||'').substring(0,30)}</div>
        ${c>0 ? `<span class="override-badge" title="${c} override(s)">${c}</span>` : ''}
      </div>`;
    }).join('');
  document.getElementById('pinDefaults').classList.toggle('active', currentView === 'defaults');
}
function filterBuildings() { renderSidebar(document.getElementById('search').value); }

function selectDefaults() {
  currentView = 'defaults';
  renderSidebar(document.getElementById('search').value);
  renderDefaultsPanel();
}

function selectBuilding(code, name) {
  currentView = code;
  currentBldgData = {};
  renderSidebar(document.getElementById('search').value);
  fetch('/api/building-assumptions/' + code).then(r => r.json()).then(data => {
    currentBldgData = data || {};
    renderBuildingPanel(code, name);
  });
}

function renderDefaultsPanel() {
  const d = defaults;
  const pt = d.payroll_tax || {};
  const ub = d.union_benefits || {};
  const wc = d.workers_comp || {};
  const wi = d.wage_increase || {};
  const ir = d.insurance_renewal || {};
  const en = d.energy || {};
  const ws = d.water_sewer || {};
  const c = document.getElementById('content');
  c.innerHTML = `
    <div class="panel-title">Portfolio Defaults</div>
    <div class="panel-sub">These values apply to all buildings unless overridden. Edit a building to override energy/water fields per-property.</div>
    <div class="section">
      <h2>Payroll Tax Rates</h2>
      <div class="form-grid">
        ${fg('FICA','d_fica',pt.FICA,'number','0.0001')}
        ${fg('SUI','d_sui',pt.SUI,'number','0.0001')}
        ${fg('FUI','d_fui',pt.FUI,'number','0.0001')}
        ${fg('MTA','d_mta',pt.MTA,'number','0.0001')}
        ${fg('NYS Disability','d_nysd',pt.NYS_Disability,'number','0.0001')}
        ${fg('Paid Family Leave','d_pfl',pt.PFL,'number','0.0001')}
      </div>
    </div>
    <div class="section">
      <h2>32BJ Union Benefits</h2>
      <div class="form-grid">
        ${fg('Welfare ($/mo)','d_welf',ub.welfare_monthly,'number','0.01')}
        ${fg('Pension ($/wk)','d_pens',ub.pension_weekly,'number','0.01')}
        ${fg('Supp Retirement ($/wk)','d_supp',ub.supp_retirement_weekly,'number','0.01')}
        ${fg('Legal ($/mo)','d_legal',ub.legal_monthly,'number','0.01')}
        ${fg('Training ($/mo)','d_train',ub.training_monthly,'number','0.01')}
        ${fg('Profit Sharing ($/qtr)','d_prof',ub.profit_sharing_quarterly,'number','0.01')}
      </div>
    </div>
    <div class="section">
      <h2>Workers Comp</h2>
      <div class="form-grid">${fg('Workers Comp %','d_wc',wc.percent,'number','0.0001')}</div>
    </div>
    <div class="section">
      <h2>Mid-Year Wage Increase</h2>
      <div class="form-grid">
        ${fg('Increase %','d_winc',wi.percent,'number','0.0001')}
        ${fg('Effective Week','d_wwk',wi.effective_week,'text')}
        ${fg('Pre-increase Weeks','d_wpre',wi.pre_increase_weeks,'number')}
        ${fg('Post-increase Weeks','d_wpost',wi.post_increase_weeks,'number')}
      </div>
    </div>
    <div class="section">
      <h2>Insurance Renewal</h2>
      <div class="form-grid">
        ${fg('Renewal Increase %','d_iinc',ir.increase_percent,'number','0.0001')}
        ${fg('Renewal Effective Date','d_idate',ir.effective_date,'text')}
        ${fg('Pre-renewal Months','d_ipre',ir.pre_renewal_months,'number')}
        ${fg('Post-renewal Months','d_ipost',ir.post_renewal_months,'number')}
      </div>
    </div>
    <div class="section">
      <h2>Energy Rates <span style="font-size:10px;color:#e8935a;font-weight:700;margin-left:8px;">↓ OVERRIDABLE PER BUILDING</span></h2>
      <div class="form-grid">
        ${fg('Gas ESCO Rate ($/Therm)','d_gas',en.gas_esco_rate,'number','0.0001')}
        ${fg('Electric ESCO Rate ($/kWh)','d_elec',en.electric_esco_rate,'number','0.0001')}
        ${fg('Oil/Fuel Price ($/Gal)','d_oil',en.oil_price,'number','0.01')}
        ${fg('Gas Rate Increase %','d_ginc',en.gas_rate_increase,'number','0.0001')}
        ${fg('Electric Rate Increase %','d_einc',en.electric_rate_increase,'number','0.0001')}
        ${fg('Oil Rate Increase %','d_oinc',en.oil_rate_increase,'number','0.0001')}
        ${selectField('Consumption Basis','d_cbasis',en.consumption_basis||'2-Year Average',['2-Year Average','Prior Year','Custom'])}
      </div>
    </div>
    <div class="section">
      <h2>Water & Sewer <span style="font-size:10px;color:#e8935a;font-weight:700;margin-left:8px;">↓ OVERRIDABLE PER BUILDING</span></h2>
      <div class="form-grid">${fg('Rate Increase %','d_water',ws.rate_increase,'number','0.0001')}</div>
    </div>
    <div class="button-row">
      <button class="btn-primary" onclick="saveDefaults()">Save Defaults</button>
    </div>
  `;
}

function fg(label, id, val, type, step) {
  const v = (val===undefined||val===null)?'':val;
  const st = step?`step="${step}"`:'';
  return `<div class="form-group"><label>${label}</label><input type="${type}" id="${id}" value="${v}" ${st} /></div>`;
}
function selectField(label, id, val, options) {
  return `<div class="form-group"><label>${label}</label><select id="${id}">${options.map(o=>`<option ${o===val?'selected':''}>${o}</option>`).join('')}</select></div>`;
}

function saveDefaults() {
  const v = id => parseFloat(document.getElementById(id).value) || 0;
  const t = id => document.getElementById(id).value;
  const n = id => parseInt(document.getElementById(id).value) || 0;
  const payload = {
    payroll_tax: { FICA:v('d_fica'), SUI:v('d_sui'), FUI:v('d_fui'), MTA:v('d_mta'), NYS_Disability:v('d_nysd'), PFL:v('d_pfl') },
    union_benefits: { welfare_monthly:v('d_welf'), pension_weekly:v('d_pens'), supp_retirement_weekly:v('d_supp'), legal_monthly:v('d_legal'), training_monthly:v('d_train'), profit_sharing_quarterly:v('d_prof') },
    workers_comp: { percent:v('d_wc') },
    wage_increase: { percent:v('d_winc'), effective_week:t('d_wwk'), pre_increase_weeks:n('d_wpre'), post_increase_weeks:n('d_wpost') },
    insurance_renewal: { increase_percent:v('d_iinc'), effective_date:t('d_idate'), pre_renewal_months:n('d_ipre'), post_renewal_months:n('d_ipost') },
    energy: { gas_esco_rate:v('d_gas'), electric_esco_rate:v('d_elec'), oil_price:v('d_oil'), gas_rate_increase:v('d_ginc'), electric_rate_increase:v('d_einc'), oil_rate_increase:v('d_oinc'), consumption_basis:t('d_cbasis') },
    water_sewer: { rate_increase:v('d_water') }
  };
  fetch('/api/defaults', { method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(payload) })
    .then(r => r.json()).then(() => { Object.assign(defaults, payload); toast('Defaults saved'); });
}

function renderBuildingPanel(code, name) {
  const count = overrideCounts[code] || 0;
  const c = document.getElementById('content');
  c.innerHTML = `
    <div class="panel-title">${code} — ${name}</div>
    <div class="panel-sub">Entity Code: ${code} · <span id="overrideCountDisplay">${count}</span> override<span id="overrideCountPlural">${count===1?'':'s'}</span> active</div>
    <div class="tabs" id="tabs">
      <button class="tab" data-tab="payroll" onclick="setTab('payroll')">Payroll</button>
      <button class="tab" data-tab="income" onclick="setTab('income')">Income</button>
      <button class="tab" data-tab="insurance" onclick="setTab('insurance')">Insurance</button>
      <button class="tab" data-tab="utilities" onclick="setTab('utilities')">Utilities<span class="badge-inline" id="utilBadge" style="${count>0?'':'display:none'}">${count}</span></button>
      <button class="tab" data-tab="summary" onclick="setTab('summary')">Summary<span class="badge-inline" id="summBadge" style="${count>0?'':'display:none'}">${count}</span></button>
    </div>
    <div id="tabContent"></div>
  `;
  setTab(currentTab);
}

function setTab(name) {
  currentTab = name;
  document.querySelectorAll('.tab').forEach(t => t.classList.toggle('active', t.dataset.tab === name));
  const tc = document.getElementById('tabContent');
  if (name === 'payroll') { tc.innerHTML = payrollTabHTML(); initPayrollTab(); }
  else if (name === 'income') { tc.innerHTML = incomeTabHTML(); initIncomeTab(); }
  else if (name === 'insurance') { tc.innerHTML = insuranceTabHTML(); initInsuranceTab(); }
  else if (name === 'utilities') { tc.innerHTML = utilitiesTabHTML(); initUtilitiesTab(); }
  else if (name === 'summary') { tc.innerHTML = summaryTabHTML(); }
}

function payrollTabHTML() {
  return `<div class="section"><h3>Position Details</h3>
    <table><thead><tr><th>Position</th><th style="width:140px"># Employees</th><th style="width:160px">Hourly Rate</th></tr></thead>
    <tbody id="payrollTable"></tbody></table></div>`;
}
function initPayrollTab() {
  const positions = (currentBldgData.payroll && currentBldgData.payroll.positions) || [
    { name: 'Resident Manager', employee_count: 0, hourly_rate: 0 },
    { name: 'Handyman', employee_count: 0, hourly_rate: 0 },
    { name: 'Porter/Doorman', employee_count: 0, hourly_rate: 0 },
    { name: 'Porter 80%', employee_count: 0, hourly_rate: 0 },
    { name: 'Employee 5', employee_count: 0, hourly_rate: 0 },
    { name: 'Employee 6', employee_count: 0, hourly_rate: 0 },
    { name: 'Employee 7', employee_count: 0, hourly_rate: 0 },
    { name: 'Employee 8', employee_count: 0, hourly_rate: 0 }
  ];
  const table = document.getElementById('payrollTable');
  table.innerHTML = positions.map((p, i) => `<tr>
    <td><input type="text" value="${(p.name||'').replace(/"/g,'&quot;')}" class="payroll-name" data-idx="${i}" /></td>
    <td><input type="number" value="${p.employee_count||0}" class="payroll-count" data-idx="${i}" /></td>
    <td><input type="number" step="0.01" value="${p.hourly_rate||0}" class="payroll-rate" data-idx="${i}" /></td>
  </tr>`).join('');
  document.querySelectorAll('.payroll-name, .payroll-count, .payroll-rate').forEach(el => { el.addEventListener('change', savePayroll); });
}
function savePayroll() {
  const positions = [];
  document.querySelectorAll('#payrollTable tr').forEach(tr => {
    positions.push({
      name: tr.querySelector('.payroll-name').value,
      employee_count: parseInt(tr.querySelector('.payroll-count').value) || 0,
      hourly_rate: parseFloat(tr.querySelector('.payroll-rate').value) || 0
    });
  });
  currentBldgData.payroll = { positions };
  debouncedSave();
}

function incomeTabHTML() {
  return `
    <div class="section"><h3>Maintenance</h3><div class="form-grid">
      <div class="form-group"><label>Total Shares</label><input type="number" id="income_maint_shares" /></div>
      <div class="form-group"><label>$/Share/Mo</label><input type="number" step="0.01" id="income_maint_per_share" /></div>
      <div class="form-group"><label>Increase %</label><input type="number" step="0.0001" id="income_maint_increase" /></div>
    </div></div>
    <div class="section"><h3>Storage Units</h3>
    <table><thead><tr><th>Size Label</th><th># Units</th><th># Occupied</th><th>$/Month</th></tr></thead>
    <tbody id="storageTable"></tbody></table></div>
    <div class="section"><h3>Bike Storage</h3><div class="form-grid">
      <div class="form-group"><label># Racks</label><input type="number" id="income_bike_racks" /></div>
      <div class="form-group"><label># Occupied</label><input type="number" id="income_bike_occupied" /></div>
      <div class="form-group"><label>$/Month</label><input type="number" step="0.01" id="income_bike_monthly" /></div>
    </div></div>
    <div class="section"><h3>Laundry/Vending</h3><div class="form-grid">
      <div class="form-group"><label>Contract Description</label><input type="text" id="income_laundry_desc" /></div>
      <div class="form-group"><label>Monthly Amount</label><input type="number" step="0.01" id="income_laundry_monthly" /></div>
    </div></div>`;
}
function initIncomeTab() {
  const income = currentBldgData.income || {};
  const maint = income.maintenance || {};
  const storage = income.storage || [];
  const bike = income.bike_storage || {};
  const laundry = income.laundry_vending || {};
  document.getElementById('income_maint_shares').value = maint.total_shares || 0;
  document.getElementById('income_maint_per_share').value = maint.per_share_monthly || 0;
  document.getElementById('income_maint_increase').value = maint.increase_percent || 0;
  const rows = storage.length > 0 ? storage : [
    { size_label: 'Small', units: 0, occupied: 0, monthly: 0 },
    { size_label: 'Medium', units: 0, occupied: 0, monthly: 0 },
    { size_label: 'Large', units: 0, occupied: 0, monthly: 0 },
    { size_label: 'XL', units: 0, occupied: 0, monthly: 0 }
  ];
  document.getElementById('storageTable').innerHTML = rows.map((s, i) => `<tr>
    <td><input type="text" value="${s.size_label||''}" class="storage-label" data-idx="${i}" /></td>
    <td><input type="number" value="${s.units||0}" class="storage-units" data-idx="${i}" /></td>
    <td><input type="number" value="${s.occupied||0}" class="storage-occupied" data-idx="${i}" /></td>
    <td><input type="number" step="0.01" value="${s.monthly||0}" class="storage-monthly" data-idx="${i}" /></td>
  </tr>`).join('');
  document.getElementById('income_bike_racks').value = bike.racks || 0;
  document.getElementById('income_bike_occupied').value = bike.occupied || 0;
  document.getElementById('income_bike_monthly').value = bike.monthly || 0;
  document.getElementById('income_laundry_desc').value = laundry.description || '';
  document.getElementById('income_laundry_monthly').value = laundry.monthly || 0;
  document.querySelectorAll('.storage-label,.storage-units,.storage-occupied,.storage-monthly,#income_maint_shares,#income_maint_per_share,#income_maint_increase,#income_bike_racks,#income_bike_occupied,#income_bike_monthly,#income_laundry_desc,#income_laundry_monthly').forEach(el => { el.addEventListener('change', saveIncome); });
}
function saveIncome() {
  currentBldgData.income = {
    maintenance: {
      total_shares: parseFloat(document.getElementById('income_maint_shares').value) || 0,
      per_share_monthly: parseFloat(document.getElementById('income_maint_per_share').value) || 0,
      increase_percent: parseFloat(document.getElementById('income_maint_increase').value) || 0
    },
    storage: Array.from(document.querySelectorAll('#storageTable tr')).map(tr => ({
      size_label: tr.querySelector('.storage-label').value,
      units: parseInt(tr.querySelector('.storage-units').value) || 0,
      occupied: parseInt(tr.querySelector('.storage-occupied').value) || 0,
      monthly: parseFloat(tr.querySelector('.storage-monthly').value) || 0
    })),
    bike_storage: {
      racks: parseInt(document.getElementById('income_bike_racks').value) || 0,
      occupied: parseInt(document.getElementById('income_bike_occupied').value) || 0,
      monthly: parseFloat(document.getElementById('income_bike_monthly').value) || 0
    },
    laundry_vending: {
      description: document.getElementById('income_laundry_desc').value,
      monthly: parseFloat(document.getElementById('income_laundry_monthly').value) || 0
    }
  };
  debouncedSave();
}

function insuranceTabHTML() {
  return `<div class="section"><h3>Insurance Policies</h3>
    <table><thead><tr><th>GL Code</th><th>Policy Name</th><th>Current Annual Premium</th><th>Current Year Budget</th><th>Expiration Date</th><th>Override Increase %</th></tr></thead>
    <tbody id="insuranceTable"></tbody></table></div>`;
}
function initInsuranceTab() {
  const insurance = currentBldgData.insurance || [];
  const table = document.getElementById('insuranceTable');
  table.innerHTML = policies.map(p => {
    const ex = insurance.find(x => x.gl_code === p.gl_code) || {};
    return `<tr>
      <td>${p.gl_code}</td>
      <td>${p.name}</td>
      <td><input type="number" step="0.01" value="${ex.current_premium||0}" class="ins-cur" data-gl="${p.gl_code}" /></td>
      <td><input type="number" step="0.01" value="${ex.current_budget||0}" class="ins-bud" data-gl="${p.gl_code}" /></td>
      <td><input type="text" value="${ex.expiration_date||''}" class="ins-exp" data-gl="${p.gl_code}" /></td>
      <td><input type="number" step="0.0001" value="${ex.override_increase||0}" class="ins-inc" data-gl="${p.gl_code}" /></td>
    </tr>`;
  }).join('');
  document.querySelectorAll('.ins-cur,.ins-bud,.ins-exp,.ins-inc').forEach(el => { el.addEventListener('change', saveInsurance); });
}
function saveInsurance() {
  currentBldgData.insurance = policies.map(p => ({
    gl_code: p.gl_code,
    name: p.name,
    current_premium: parseFloat(document.querySelector(`.ins-cur[data-gl="${p.gl_code}"]`).value) || 0,
    current_budget: parseFloat(document.querySelector(`.ins-bud[data-gl="${p.gl_code}"]`).value) || 0,
    expiration_date: document.querySelector(`.ins-exp[data-gl="${p.gl_code}"]`).value,
    override_increase: parseFloat(document.querySelector(`.ins-inc[data-gl="${p.gl_code}"]`).value) || 0
  }));
  debouncedSave();
}

function utilitiesTabHTML() {
  const fields = OVERRIDE_FIELDS.map(f => overrideFieldHTML(f)).join('');
  return `<div class="section">
    <h3>Energy & Water — Rate Overrides</h3>
    <p style="font-size:12px;color:#666;margin-bottom:14px;">Defaults come from <strong>Portfolio Defaults</strong>. Override only where this building differs. Blank or 0 = use default.</p>
    <div class="override-field-grid" id="overrideGrid">${fields}</div>
  </div>
  <div class="section"><h3>Energy — GL Account Adjustments</h3>
    <table><thead><tr><th>GL Code</th><th>Description</th><th>Accrual Adjustment</th><th>Unpaid Bills</th><th>Rate Increase % Override</th></tr></thead>
    <tbody id="energyGLTable"></tbody></table></div>
  <div class="section"><h3>Water & Sewer — GL Account Adjustments</h3>
    <table><thead><tr><th>GL Code</th><th>Description</th><th>Accrual Adjustment</th><th>Unpaid Bills</th><th>Rate Increase % Override</th></tr></thead>
    <tbody id="waterGLTable"></tbody></table></div>
  <div class="button-row">
    <button class="btn-ghost" onclick="resetOverrides()">Reset All Overrides to Defaults</button>
  </div>`;
}
function overrideFieldHTML(f) {
  const dv = getPath(defaults, f.defaultPath);
  const ov = getPath(currentBldgData, f.saveKey);
  const hasOverride = f.type === 'select' ? !!ov : (ov !== undefined && ov !== null && ov !== '' && parseFloat(ov) !== 0);
  const dvDisplay = dv === undefined || dv === null ? '—' : dv;
  let input;
  if (f.type === 'select') {
    input = `<select id="${f.id}" onchange="onOverrideChange()"><option value="">— use default —</option>${f.options.map(o => `<option ${ov===o?'selected':''}>${o}</option>`).join('')}</select>`;
  } else {
    const val = hasOverride ? ov : '';
    input = `<input type="number" step="${f.step}" id="${f.id}" placeholder="(use default)" value="${val}" oninput="onOverrideChange()" />`;
  }
  const status = hasOverride
    ? `<div class="status override">◉ Overrides default (${dvDisplay})</div>`
    : `<div class="status using-default">○ Using portfolio default</div>`;
  return `<div class="override-field ${hasOverride?'has-override':''}" data-field="${f.id}">
    <label>${f.label}${f.unit?` <span style="color:#888;font-weight:400">(${f.unit})</span>`:''}</label>
    <div class="default-chip">Default: <span class="val">${dvDisplay}</span></div>
    ${input}${status}
  </div>`;
}
function initUtilitiesTab() {
  // Energy GL table
  const energyAdj = (currentBldgData.energy && currentBldgData.energy.gl_adjustments) || {};
  document.getElementById('energyGLTable').innerHTML = ENERGY_GL_ACCOUNTS.map(a => {
    const adj = energyAdj[a.gl] || {};
    return `<tr>
      <td>${a.gl}</td><td>${a.desc}</td>
      <td><input type="number" step="0.01" value="${adj.accrual||0}" class="energy-accrual" data-gl="${a.gl}" /></td>
      <td><input type="number" step="0.01" value="${adj.unpaid||0}" class="energy-unpaid" data-gl="${a.gl}" /></td>
      <td><input type="number" step="0.0001" value="${adj.rate_increase||0}" class="energy-gl-rate" data-gl="${a.gl}" /></td>
    </tr>`;
  }).join('');
  const waterAdj = (currentBldgData.water_sewer && currentBldgData.water_sewer.gl_adjustments) || {};
  document.getElementById('waterGLTable').innerHTML = WATER_GL_ACCOUNTS.map(a => {
    const adj = waterAdj[a.gl] || {};
    return `<tr>
      <td>${a.gl}</td><td>${a.desc}</td>
      <td><input type="number" step="0.01" value="${adj.accrual||0}" class="water-accrual" data-gl="${a.gl}" /></td>
      <td><input type="number" step="0.01" value="${adj.unpaid||0}" class="water-unpaid" data-gl="${a.gl}" /></td>
      <td><input type="number" step="0.0001" value="${adj.rate_increase||0}" class="water-gl-rate" data-gl="${a.gl}" /></td>
    </tr>`;
  }).join('');
  document.querySelectorAll('.energy-accrual,.energy-unpaid,.energy-gl-rate,.water-accrual,.water-unpaid,.water-gl-rate').forEach(el => { el.addEventListener('change', saveUtilities); });
}
function onOverrideChange() { saveUtilities(); refreshOverrideUI(); }

function saveUtilities() {
  // Collect override fields
  const energy = { gl_adjustments: {} };
  const water = { gl_adjustments: {} };
  for (const f of OVERRIDE_FIELDS) {
    const el = document.getElementById(f.id);
    if (!el) continue;
    let v = el.value;
    if (f.type === 'number') v = parseFloat(v) || 0;
    setPath({ energy, water_sewer: water }, f.saveKey, v);
  }
  // Collect GL adjustments
  ENERGY_GL_ACCOUNTS.forEach(a => {
    energy.gl_adjustments[a.gl] = {
      accrual: parseFloat(document.querySelector(`.energy-accrual[data-gl="${a.gl}"]`).value) || 0,
      unpaid: parseFloat(document.querySelector(`.energy-unpaid[data-gl="${a.gl}"]`).value) || 0,
      rate_increase: parseFloat(document.querySelector(`.energy-gl-rate[data-gl="${a.gl}"]`).value) || 0
    };
  });
  WATER_GL_ACCOUNTS.forEach(a => {
    water.gl_adjustments[a.gl] = {
      accrual: parseFloat(document.querySelector(`.water-accrual[data-gl="${a.gl}"]`).value) || 0,
      unpaid: parseFloat(document.querySelector(`.water-unpaid[data-gl="${a.gl}"]`).value) || 0,
      rate_increase: parseFloat(document.querySelector(`.water-gl-rate[data-gl="${a.gl}"]`).value) || 0
    };
  });
  currentBldgData.energy = energy;
  currentBldgData.water_sewer = water;
  debouncedSave();
}

function refreshOverrideUI() {
  // Re-render override fields to update status chips and amber highlights, without losing focus
  const grid = document.getElementById('overrideGrid');
  if (!grid) return;
  const focusId = document.activeElement && document.activeElement.id;
  const selStart = document.activeElement && document.activeElement.selectionStart;
  grid.innerHTML = OVERRIDE_FIELDS.map(f => overrideFieldHTML(f)).join('');
  if (focusId) {
    const el = document.getElementById(focusId);
    if (el) { el.focus(); try { el.setSelectionRange(selStart, selStart); } catch(e){} }
  }
  updateOverrideCount();
}

function updateOverrideCount() {
  const c = computeOverrideCount(currentBldgData);
  overrideCounts[currentView] = c;
  const disp = document.getElementById('overrideCountDisplay');
  const plural = document.getElementById('overrideCountPlural');
  const uBadge = document.getElementById('utilBadge');
  const sBadge = document.getElementById('summBadge');
  if (disp) disp.textContent = c;
  if (plural) plural.textContent = c === 1 ? '' : 's';
  if (uBadge) { uBadge.textContent = c; uBadge.style.display = c > 0 ? '' : 'none'; }
  if (sBadge) { sBadge.textContent = c; sBadge.style.display = c > 0 ? '' : 'none'; }
  renderSidebar(document.getElementById('search').value);
}

function summaryTabHTML() {
  const activeEntries = OVERRIDE_FIELDS
    .map(f => ({ f, ov: getPath(currentBldgData, f.saveKey), dv: getPath(defaults, f.defaultPath) }))
    .filter(e => e.f.type === 'select' ? !!e.ov : (e.ov !== undefined && e.ov !== null && e.ov !== '' && parseFloat(e.ov) !== 0));
  if (activeEntries.length === 0) {
    return `<div class="section"><div class="empty"><div class="empty-icon">✓</div><div><strong>No overrides</strong></div><div style="margin-top:6px;font-size:13px">This building uses all portfolio defaults.</div></div></div>`;
  }
  const rows = activeEntries.map(({f, ov, dv}) => `<div class="summary-row">
    <div><div class="field-label">${f.label}</div><div class="field-path">Utilities · ${f.unit||''}</div></div>
    <div class="val default-val">${dv===undefined||dv===null?'—':dv}</div>
    <div class="val override-val">${ov}</div>
    <div class="arrow">→</div>
  </div>`).join('');
  return `<div class="section"><h3>Overrides Summary — ${activeEntries.length} field${activeEntries.length===1?'':'s'}</h3>
    <div class="summary-header"><div>Field</div><div>Default</div><div>Override</div><div></div></div>
    ${rows}
    <div class="button-row"><button class="btn-ghost" onclick="resetOverrides()">Reset All Overrides to Defaults</button></div>
  </div>`;
}

function resetOverrides() {
  const count = overrideCounts[currentView] || 0;
  if (count === 0) { toast('No overrides to reset'); return; }
  if (!confirm(`Reset all ${count} override(s) for ${currentView}? Portfolio defaults will be used everywhere.`)) return;
  // Clear scalar overrides, keep GL adjustments
  if (currentBldgData.energy) {
    const keep = { gl_adjustments: currentBldgData.energy.gl_adjustments || {} };
    currentBldgData.energy = keep;
  }
  if (currentBldgData.water_sewer) {
    const keep = { gl_adjustments: currentBldgData.water_sewer.gl_adjustments || {} };
    currentBldgData.water_sewer = keep;
  }
  debouncedSave();
  setTab(currentTab);
  updateOverrideCount();
  toast('Overrides reset');
}

function debouncedSave() {
  clearTimeout(debounceTimer);
  debounceTimer = setTimeout(() => {
    fetch('/api/building-assumptions/' + currentView, {
      method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify(currentBldgData)
    }).then(() => { toast('Saved'); updateOverrideCount(); });
  }, 500);
}

function toast(msg) {
  const t = document.getElementById('toast');
  t.textContent = '✓ ' + msg;
  t.classList.add('show');
  setTimeout(() => t.classList.remove('show'), 2000);
}

renderSidebar();
renderDefaultsPanel();
</script>
</body>
</html>
"""



# ─── SharePoint / Microsoft Graph integration ────────────────────────────
# App-only auth (client credentials flow). Permissions: Sites.ReadWrite.All.
SHAREPOINT_SITE_ID = "centurynyc.sharepoint.com,beafa175-6707-4aa1-a349-85d1cf02eb39,f6ca63b2-9cdc-4a28-b23d-047ab7e715f4"
SHAREPOINT_2027_FOLDER_PATH = "01 - Accounting General/Budgets/Budgets - FAs only/2027 Budget"

_GRAPH_TOKEN_CACHE = {"token": None, "expires_at": 0}


def _get_graph_token():
    """Acquire (and cache) an app-only Graph access token via MSAL."""
    import time
    now = time.time()
    if _GRAPH_TOKEN_CACHE["token"] and _GRAPH_TOKEN_CACHE["expires_at"] - now > 60:
        return _GRAPH_TOKEN_CACHE["token"]

    tenant = os.environ.get("AZURE_TENANT_ID", "")
    client_id = os.environ.get("AZURE_CLIENT_ID", "")
    client_secret = os.environ.get("AZURE_CLIENT_SECRET", "")
    if not (tenant and client_id and client_secret):
        raise RuntimeError("Azure env vars missing (AZURE_TENANT_ID/CLIENT_ID/CLIENT_SECRET)")

    import msal
    authority = f"https://login.microsoftonline.com/{tenant}"
    app_msal = msal.ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=authority,
    )
    result = app_msal.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise RuntimeError(f"Graph token acquisition failed: {result.get('error_description') or result}")
    _GRAPH_TOKEN_CACHE["token"] = result["access_token"]
    _GRAPH_TOKEN_CACHE["expires_at"] = now + int(result.get("expires_in", 3600))
    return _GRAPH_TOKEN_CACHE["token"]


def _graph_get(path, params=None, _retry=True):
    """GET https://graph.microsoft.com/v1.0/<path>. Returns parsed JSON or raises.
    On 401, invalidates the token cache and retries once (handles consent-flip cases).
    """
    import urllib.request
    import urllib.parse
    token = _get_graph_token()
    url = "https://graph.microsoft.com/v1.0/" + path.lstrip("/")
    if params:
        url += ("&" if "?" in url else "?") + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.request.HTTPError as e:
        if e.code == 401 and _retry:
            _GRAPH_TOKEN_CACHE["token"] = None
            _GRAPH_TOKEN_CACHE["expires_at"] = 0
            return _graph_get(path, params=params, _retry=False)
        body = ""
        try:
            body = e.read().decode("utf-8")[:600]
        except Exception:
            pass
        raise RuntimeError(f"Graph {e.code} {e.reason} on {path}: {body}")


def _graph_get_drive_id():
    """Return the default document library drive id for the SharePoint site."""
    data = _graph_get(f"sites/{SHAREPOINT_SITE_ID}/drive")
    return data.get("id")




def _graph_post(path, body):
    """POST JSON to Graph. Returns parsed JSON or raises RuntimeError with body."""
    import urllib.request
    token = _get_graph_token()
    url = "https://graph.microsoft.com/v1.0/" + path.lstrip("/")
    req = urllib.request.Request(
        url,
        data=json.dumps(body).encode("utf-8"),
        method="POST",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.request.HTTPError as e:
        body_text = ""
        try:
            body_text = e.read().decode("utf-8")[:600]
        except Exception:
            pass
        raise RuntimeError(f"Graph {e.code} {e.reason} on POST {path}: {body_text}")


def _sharepoint_create_folder(parent_path, name):
    """Create a folder inside parent_path (relative to SP root). Fails on conflict.
    Returns the created item dict, or raises RuntimeError if it already exists.
    """
    import urllib.parse
    drive_id = _graph_get_drive_id()
    encoded = urllib.parse.quote(parent_path, safe="/")
    return _graph_post(
        f"drives/{drive_id}/root:/{encoded}:/children",
        {"name": name, "folder": {}, "@microsoft.graph.conflictBehavior": "fail"},
    )


def _sharepoint_entity_folder_exists(entity_code):
    """True if 2027 Budget/<entity>/ already exists."""
    import urllib.parse
    drive_id = _graph_get_drive_id()
    base = SHAREPOINT_2027_FOLDER_PATH + "/" + str(entity_code)
    encoded = urllib.parse.quote(base, safe="/")
    try:
        _graph_get(f"drives/{drive_id}/root:/{encoded}:")
        return True
    except RuntimeError as e:
        if "404" in str(e):
            return False
        raise


def _sharepoint_ensure_entity_folder(entity_code):
    """Ensure 2027 Budget/<entity>/Supporting Documents/ exists. Idempotent.
    Returns dict describing what was created (or what already existed).
    """
    result = {"entity_code": str(entity_code), "created": [], "existed": []}
    # 1. Entity folder
    if _sharepoint_entity_folder_exists(entity_code):
        result["existed"].append(str(entity_code))
    else:
        _sharepoint_create_folder(SHAREPOINT_2027_FOLDER_PATH, str(entity_code))
        result["created"].append(str(entity_code))
    # 2. Supporting Documents subfolder
    sub_parent = SHAREPOINT_2027_FOLDER_PATH + "/" + str(entity_code)
    try:
        _sharepoint_create_folder(sub_parent, "Supporting Documents")
        result["created"].append(f"{entity_code}/Supporting Documents")
    except RuntimeError as e:
        if "409" in str(e) or "nameAlreadyExists" in str(e):
            result["existed"].append(f"{entity_code}/Supporting Documents")
        else:
            raise
    return result


@app.route("/api/sharepoint/_create-folders", methods=["POST", "GET"])
def sharepoint_create_folders():
    """ADMIN: ensure entity folders + Supporting Documents subfolders exist on
    SharePoint for entities that have a Budget row for the current year.

    Query params:
      limit=N    — only create up to N new entity folders this run (default 10)
      dry_run=1  — list what would be created without touching SharePoint
    Skips entities whose top-level folder already exists.
    """
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY
    try:
        limit = int(request.args.get("limit", "10"))
    except ValueError:
        limit = 10
    dry_run = request.args.get("dry_run") in ("1", "true", "yes")

    rows = (db.session.query(Budget.entity_code)
            .filter_by(year=_BY)
            .order_by(Budget.entity_code)
            .all())
    all_entities = [r[0] for r in rows]

    summary = {
        "limit": limit,
        "dry_run": dry_run,
        "year": _BY,
        "total_budget_entities": len(all_entities),
        "skipped_existing": [],
        "created": [],
        "errors": [],
        "remaining_after_run": 0,
    }

    created_count = 0
    missing_entities = []
    try:
        # Single Graph call to list ALL existing folders under 2027 Budget,
        # then check membership in-memory. Avoids 143 sequential checks.
        import urllib.parse
        drive_id = _graph_get_drive_id()
        encoded = urllib.parse.quote(SHAREPOINT_2027_FOLDER_PATH, safe="/")
        listing = _graph_get(f"drives/{drive_id}/root:/{encoded}:/children")
        existing_names = {it.get("name") for it in listing.get("value", []) if "folder" in it}
        for ec in all_entities:
            if str(ec) in existing_names:
                summary["skipped_existing"].append(ec)
            else:
                missing_entities.append(ec)
        summary["total_missing"] = len(missing_entities)

        if dry_run:
            summary["would_create"] = missing_entities[:limit]
            return jsonify(summary)

        # Second pass: create up to `limit` of the missing ones.
        for ec in missing_entities:
            if created_count >= limit:
                break
            try:
                r = _sharepoint_ensure_entity_folder(ec)
                summary["created"].append(r)
                created_count += 1
            except Exception as e:
                summary["errors"].append({"entity_code": ec, "error": str(e)})
        summary["remaining_after_run"] = max(0, len(missing_entities) - created_count)
    except Exception as e:
        summary["fatal_error"] = str(e)
        return jsonify(summary), 500

    return jsonify(summary)



# Source-type filename patterns for SharePoint Supporting Documents.
# Case-insensitive substring match. First match wins (in order).
SHAREPOINT_SOURCE_PATTERNS = [
    # (source_type, [substrings to match — any one matches])
    # Order matters: first-match wins. Audit patterns checked first because audit
    # PDFs in Supporting Documents are now expected (placed there by audit-sync admin).
    ("audit_2025",            ["audit", "financial statement", "afs", " fs ", ".pdf"]),
    ("approved_2026",         ["approved"]),
    ("ysl",                   ["ysl"]),
    ("expense_distribution",  ["expensedistribution"]),
    ("ap_aging",              ["apaging", "payablesaging"]),
    ("maint_proof",           ["adhoc_amp", "amp_"]),
]


def _classify_sharepoint_filename(name):
    """Return source_type ('ysl', 'expense_dist', etc.) or None."""
    if not name:
        return None
    low = name.lower()
    for source_type, needles in SHAREPOINT_SOURCE_PATTERNS:
        for n in needles:
            if n in low:
                return source_type
    return None


def _sharepoint_list_entity_sources(entity_code):
    """Read-only: list files in <entity>/Supporting Documents/ and classify
    each by source type using SHAREPOINT_SOURCE_PATTERNS.

    Returns dict:
      {
        "entity_code": str,
        "folder_exists": bool,
        "folder_url": str|None,
        "by_source_type": {
            "ysl": [{name, web_url, size, last_modified, item_id}, ...],
            "expense_distribution": [...],
            "ap_aging": [...],
            "maint_proof": [...],
            "unmatched": [...]   # files in folder that don\'t match any pattern
        }
      }
    Never raises — returns folder_exists=False on any error.
    """
    import urllib.parse
    result = {
        "entity_code": str(entity_code),
        "folder_exists": False,
        "folder_url": None,
        "by_source_type": {
            "audit_2025": [], "approved_2026": [],
            "ysl": [], "expense_distribution": [], "ap_aging": [],
            "maint_proof": [], "unmatched": [],
        },
    }
    try:
        drive_id = _graph_get_drive_id()
        sub_path = SHAREPOINT_2027_FOLDER_PATH + "/" + str(entity_code) + "/Supporting Documents"
        encoded = urllib.parse.quote(sub_path, safe="/")
        listing = _graph_get(f"drives/{drive_id}/root:/{encoded}:/children")
    except RuntimeError as e:
        if "404" in str(e):
            return result
        result["error"] = str(e)
        return result

    result["folder_exists"] = True
    for it in listing.get("value", []):
        if "folder" in it:
            continue  # skip subfolders (e.g., Con Edison/, Water & Sewer/)
        entry = {
            "name": it.get("name"),
            "web_url": it.get("webUrl"),
            "size": it.get("size"),
            "last_modified": it.get("lastModifiedDateTime"),
            "item_id": it.get("id"),
        }
        st = _classify_sharepoint_filename(it.get("name", ""))
        if st:
            result["by_source_type"][st].append(entry)
        else:
            result["by_source_type"]["unmatched"].append(entry)

    # Second pass: scan the entity TOP folder. FAs commonly drop Yardi reports
    # AND the approved budget straight into 2027 Budget/<entity>/ rather than
    # Supporting Documents. Classify each file with the same patterns — but
    # skip audit_2025: the audit needles include ".pdf", which would grab any
    # stray PDF in the top folder. Audits live in Supporting Documents only.
    try:
        top_path = SHAREPOINT_2027_FOLDER_PATH + "/" + str(entity_code)
        top_enc = urllib.parse.quote(top_path, safe="/")
        top_listing = _graph_get(f"drives/{drive_id}/root:/{top_enc}:/children")
        for it in top_listing.get("value", []):
            if "folder" in it:
                continue  # skip Supporting Documents subfolder etc.
            name = (it.get("name") or "")
            st = _classify_sharepoint_filename(name)
            if not st or st == "audit_2025":
                continue
            entry = {
                "name": name,
                "web_url": it.get("webUrl"),
                "size": it.get("size"),
                "last_modified": it.get("lastModifiedDateTime"),
                "item_id": it.get("id"),
            }
            # Avoid duplicates if a file is also surfaced from Supporting Documents
            if not any(e.get("item_id") == entry["item_id"] for e in result["by_source_type"][st]):
                result["by_source_type"][st].append(entry)
    except RuntimeError as e:
        if "404" not in str(e):
            # Don't fail the whole request if top folder scan errors; just record it.
            result["top_folder_error"] = str(e)
    return result



def _sharepoint_download_item(item_id):
    """Download a SharePoint file by drive item id. Returns (filename, bytes)."""
    import urllib.request
    drive_id = _graph_get_drive_id()
    meta = _graph_get(f"drives/{drive_id}/items/{item_id}")
    filename = meta.get("name", f"sp_item_{item_id}")
    download_url = meta.get("@microsoft.graph.downloadUrl")
    if download_url:
        with urllib.request.urlopen(download_url, timeout=60) as resp:
            return filename, resp.read()
    # Fallback via authenticated /content endpoint
    token = _get_graph_token()
    req = urllib.request.Request(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": f"Bearer {token}"},
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return filename, resp.read()


@app.route("/api/wizard/<entity_code>/use-sp-source", methods=["POST"])
def wizard_use_sp_source(entity_code):
    """Stage a SharePoint file as the FA-selected source for a slot.
    Does NOT download or parse — just records the selection in
    Budget.wizard_selections_json. Build Budget at Step 5 is what actually
    downloads and parses.
    """
    return _wizard_record_selection(entity_code, source_label_default=None)


def _wizard_record_selection(entity_code, source_label_default=None):
    """Record an FA file-selection into Budget.wizard_selections_json.

    Body: {
        source_type: "ysl"|"expense_distribution"|"ap_aging"|"maint_proof"|"approved_2026",
        item_id: "<sharepoint drive item id>",
        filename: "<optional, for display only>"
    }
    Returns the updated selections dict.
    """
    from datetime import datetime as _dt
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY

    data = request.get_json() or {}
    source_type = (data.get("source_type") or source_label_default or "").strip()
    item_id = (data.get("item_id") or "").strip()
    filename = (data.get("filename") or "").strip()
    web_url = (data.get("web_url") or "").strip()
    if not source_type or not item_id:
        return jsonify({"error": "source_type and item_id required"}), 400

    valid_sources = {"ysl", "expense_distribution", "ap_aging", "maint_proof", "approved_2026", "audit_2025"}
    if source_type not in valid_sources:
        return jsonify({"error": f"source_type must be one of {sorted(valid_sources)}"}), 400

    budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {entity_code} year {_BY}"}), 404

    try:
        current = json.loads(budget.wizard_selections_json or "{}")
    except Exception:
        current = {}
    current[source_type] = {
        "item_id": item_id,
        "filename": filename,
        "web_url": web_url,
        "selected_at": _dt.utcnow().isoformat(),
        "source": "sharepoint",
    }
    budget.wizard_selections_json = json.dumps(current)

    # ── PARSE-ON-CLICK dispatch (Phase D) ────────────────────────────────────
    # Each source-type click runs its parser immediately so the FA gets feedback
    # and downstream sources (e.g. audit) have what they need (e.g. categories).
    # No auto-cascade — only the clicked source-type is parsed.
    parse_result = None
    parse_error = None
    try:
        if source_type == "approved_2026":
            BudgetSummaryRow = workflow_models["BudgetSummaryRow"]
            from workflow import BUDGET_YEAR, apply_summary_prefix_override
            # Clean replace: drop old rows so a different XLSX overwrites cleanly.
            BudgetSummaryRow.query.filter_by(
                entity_code=entity_code,
                budget_year=BUDGET_YEAR,
            ).delete()
            db.session.flush()
            written = _build_apply_approved_2026(
                entity_code,
                current[source_type],
                BudgetSummaryRow,
                BUDGET_YEAR,
                apply_summary_prefix_override,
            )
            parse_result = {"source_type": source_type, "rows_imported": written,
                            "filename": filename}
        elif source_type == "audit_2025":
            # Long-running: Claude extraction is 30-60s. Done synchronously here.
            parse_result = _build_apply_audit_2025(entity_code, current[source_type])
        elif source_type == "ysl":
            parse_result = _build_apply_ysl(entity_code, current[source_type])
        elif source_type == "expense_distribution":
            parse_result = _build_apply_expense_distribution(entity_code, current[source_type])
        elif source_type == "ap_aging":
            parse_result = _build_apply_ap_aging(entity_code, current[source_type])
        elif source_type == "maint_proof":
            parse_result = _build_apply_maint_proof(entity_code, current[source_type])
    except Exception as e:
        # Roll back parser changes; preserve the staged selection so the FA
        # can re-click after fixing the source file.
        db.session.rollback()
        # Re-stage the selection (rollback wiped the unflushed change above).
        budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
        try:
            current_after_rb = json.loads(budget.wizard_selections_json or "{}")
        except Exception:
            current_after_rb = {}
        current_after_rb[source_type] = current[source_type]
        budget.wizard_selections_json = json.dumps(current_after_rb)
        db.session.commit()
        parse_error = str(e)[:1000]
        return jsonify({
            "ok": False,
            "entity_code": entity_code,
            "selections": current_after_rb,
            "source_type": source_type,
            "parse_error": parse_error,
        }), 500

    db.session.commit()
    resp = {"ok": True, "entity_code": entity_code, "selections": current}
    if parse_result:
        resp["parse_result"] = parse_result
    return jsonify(resp)


@app.route("/api/wizard/<entity_code>/selections", methods=["GET"])
def wizard_get_selections(entity_code):
    """Return the current FA selections for this entity."""
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY
    budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {entity_code}"}), 404
    try:
        sels = json.loads(budget.wizard_selections_json or "{}")
    except Exception:
        sels = {}
    return jsonify({"entity_code": entity_code, "selections": sels})


@app.route("/api/wizard/<entity_code>/selections", methods=["DELETE"])
def wizard_clear_selections(entity_code):
    """Clear all FA selections for this entity (Discard build, before Build Budget click)."""
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY
    budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {entity_code}"}), 404
    budget.wizard_selections_json = None
    db.session.commit()
    return jsonify({"ok": True, "entity_code": entity_code, "selections": {}})


@app.route("/api/wizard/<entity_code>/selections/assumptions", methods=["POST"])
def wizard_save_assumptions(entity_code):
    """Stage FA-edited assumption values into wizard_selections_json["assumptions"].

    Body: a partial assumptions dict shaped like Budget.assumptions_json
      e.g. {"insurance_renewal": {"increase_percent": 0.18}, "energy": {...}}
    Deep-merges into the existing staged assumptions. Does NOT write to
    Budget.assumptions_json — that happens at Build Budget time.
    """
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY

    data = request.get_json(silent=True) or {}
    if not isinstance(data, dict):
        return jsonify({"error": "Body must be a dict"}), 400

    budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {entity_code} year {_BY}"}), 404

    try:
        current = json.loads(budget.wizard_selections_json or "{}")
    except Exception:
        current = {}

    staged = current.get("assumptions") or {}
    if not isinstance(staged, dict):
        staged = {}

    # Deep merge by section (one level deep — matches Budget.assumptions_json shape)
    for key, value in data.items():
        if isinstance(value, dict) and isinstance(staged.get(key), dict):
            staged[key].update(value)
        else:
            staged[key] = value

    current["assumptions"] = staged
    budget.wizard_selections_json = json.dumps(current)
    db.session.commit()
    return jsonify({"ok": True, "entity_code": entity_code, "assumptions": staged})





SHAREPOINT_APPROVED_BUDGETS_PATH = SHAREPOINT_2027_FOLDER_PATH + "/2026 budget approved budgets only"


def _sharepoint_list_approved_budgets(entity_code):
    """List 2026 Approved Budget Excel files for an entity. After the bulk
    move on 2026-04-28, files live at the top level of the entity folder
    (2027 Budget/<entity>/). Match: filename contains "approved" (case-
    insensitive) AND is .xlsx/.xls (not other doc types).

    Returns list of dicts with name, web_url, size, last_modified, item_id.
    Empty list if folder missing or no matches.
    """
    import urllib.parse
    ec = str(entity_code)
    entity_path = SHAREPOINT_2027_FOLDER_PATH + "/" + ec
    try:
        drive_id = _graph_get_drive_id()
        encoded = urllib.parse.quote(entity_path, safe="/")
        listing = _graph_get(f"drives/{drive_id}/root:/{encoded}:/children")
    except RuntimeError as e:
        if "404" in str(e):
            return []
        raise

    matches = []
    for it in listing.get("value", []):
        name = it.get("name", "")
        if "folder" in it:
            continue
        low = name.lower()
        if "approved" not in low:
            continue
        if not (low.endswith(".xlsx") or low.endswith(".xls") or low.endswith(".xlsm")):
            continue
        # Skip files that look like 2027 drafts (only want 2026 approved)
        if "2027" in low and "draft" in low:
            continue
        matches.append({
            "name": name,
            "web_url": it.get("webUrl"),
            "size": it.get("size"),
            "last_modified": it.get("lastModifiedDateTime"),
            "item_id": it.get("id"),
        })
    return matches


@app.route("/api/wizard/<entity_code>/approved-budget-files", methods=["GET"])
def wizard_approved_budget_files(entity_code):
    """Read-only: list 2026 Approved Budget Excel files in SharePoint matching
    the entity_code. Used by the wizard\'s Step 2 to surface what\'s available
    for FA confirmation.
    """
    try:
        files = _sharepoint_list_approved_budgets(entity_code)
        return jsonify({
            "entity_code": entity_code,
            "folder_path": SHAREPOINT_APPROVED_BUDGETS_PATH,
            "matches": files,
            "count": len(files),
        })
    except Exception as e:
        logger.error(f"approved-budget-files({entity_code}) failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/foundation-summary", methods=["GET"])
@require_admin
def admin_foundation_summary():
    """Return Foundation status for every entity in the current Budget year.

    Used by /admin/foundation page. One row per Budget. Pre-computes
    approved_budget state, audit state, and Foundation confirmed flag for
    fast rendering.
    """
    Budget = workflow_models["Budget"]
    BudgetSummaryRow = workflow_models["BudgetSummaryRow"]
    AuditUpload = af_models["AuditUpload"]
    BuildingAssignment = workflow_models["BuildingAssignment"]
    User = workflow_models["User"]
    from workflow import BUDGET_YEAR as _BY

    # All budgets for current year
    budgets = Budget.query.filter_by(year=_BY).order_by(Budget.entity_code).all()

    # Bulk fetch summary row counts per entity
    summary_counts = {}
    rows = db.session.execute(db.text(
        "SELECT entity_code, COUNT(*) FROM budget_summary_rows "
        "WHERE budget_year = :y GROUP BY entity_code"
    ), {"y": _BY}).fetchall()
    for ec, cnt in rows:
        summary_counts[ec] = cnt

    # Bulk fetch audit_uploads — latest per entity
    audits = {}
    rows = db.session.execute(db.text(
        "SELECT DISTINCT ON (entity_code) entity_code, id, status FROM audit_uploads "
        "ORDER BY entity_code, id DESC"
    )).fetchall()
    for ec, uid, status in rows:
        audits[ec] = {"id": uid, "status": status}

    # Bulk fetch FA assignment per entity (for display)
    fa_by_entity = {}
    rows = db.session.execute(db.text(
        "SELECT ba.entity_code, u.name FROM building_assignments ba "
        "JOIN users u ON u.id = ba.user_id "
        "WHERE ba.role = \'fa\'"
    )).fetchall()
    for ec, name in rows:
        fa_by_entity[ec] = name

    out = []
    for b in budgets:
        sc = summary_counts.get(b.entity_code, 0)
        a = audits.get(b.entity_code) or {}
        if sc > 0:
            approved_state = "imported"
        elif b.foundation_no_prior_budget:
            approved_state = "acknowledged_missing"
        else:
            approved_state = "missing"
        audit_status = a.get("status")
        if audit_status == "confirmed":
            audit_state = "confirmed"
        elif audit_status in ("mapped", "extracted"):
            audit_state = "extracted"
        elif audit_status == "uploaded":
            audit_state = "in_sp"
        else:
            audit_state = "missing"
        out.append({
            "entity_code": b.entity_code,
            "building_name": b.building_name,
            "fa_name": fa_by_entity.get(b.entity_code) or "",
            "approved_budget": approved_state,
            "approved_budget_summary_rows": sc,
            "audit": audit_state,
            "audit_upload_id": a.get("id"),
            "foundation_confirmed_at": (
                b.foundation_confirmed_at.isoformat()
                if b.foundation_confirmed_at else None
            ),
            "foundation_no_prior_budget": bool(b.foundation_no_prior_budget),
        })

    # Aggregate counts for filter chips
    counts = {
        "all": len(out),
        "confirmed": sum(1 for r in out if r["foundation_confirmed_at"]),
        "pending": sum(1 for r in out if not r["foundation_confirmed_at"]),
        "no_prior_budget": sum(1 for r in out if r["foundation_no_prior_budget"]),
        "audit_missing": sum(1 for r in out if r["audit"] == "missing"),
        "audit_extracted_unconfirmed": sum(1 for r in out if r["audit"] == "extracted"),
    }

    return jsonify({
        "year": _BY,
        "total": len(out),
        "counts": counts,
        "entities": out,
    })


@app.route("/admin/foundation", methods=["GET"])
def admin_foundation_page():
    """HTML dashboard listing every entity\'s Foundation status. Lead view to
    track progress across the portfolio between now and the August cycle start."""
    return r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Foundation Dashboard - Century Management</title>
<style>
  body { font-family: -apple-system,BlinkMacSystemFont,system-ui,sans-serif; margin: 0; background: #fafaf9; color: #111827; }
  header { background: #1f2937; color: white; padding: 14px 28px; display: flex; align-items: center; gap: 16px; }
  header a { color: #d1d5db; text-decoration: none; font-size: 13px; }
  header a:hover { color: white; }
  h1 { font-size: 20px; margin: 0; }
  main { max-width: 1280px; margin: 0 auto; padding: 24px; }
  .card { background: white; border: 1px solid #e5e7eb; border-radius: 10px; padding: 18px 22px; margin-bottom: 18px; box-shadow: 0 1px 3px rgba(0,0,0,0.04); }
  .chips { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 16px; }
  .chip { padding: 6px 14px; border: 1px solid #e5e7eb; border-radius: 20px; cursor: pointer; font-size: 12px; font-weight: 600; background: white; user-select: none; }
  .chip.active { background: #1f2937; color: white; border-color: #1f2937; }
  .chip-count { font-weight: 700; margin-left: 6px; opacity: 0.7; }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { text-align: left; padding: 10px 12px; background: #f9fafb; font-weight: 700; font-size: 11px; letter-spacing: 0.05em; text-transform: uppercase; color: #6b7280; border-bottom: 1px solid #e5e7eb; }
  td { padding: 10px 12px; border-bottom: 1px solid #f3f4f6; }
  tr:hover td { background: #fafafa; }
  .ec { font-family: ui-monospace,monospace; font-size: 12px; color: #4b5563; width: 70px; }
  .badge { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; }
  .badge-ok { background: #dcfce7; color: #15803d; }
  .badge-pending { background: #fef3c7; color: #92400e; }
  .badge-missing { background: #fee2e2; color: #b91c1c; }
  .badge-neutral { background: #f3f4f6; color: #4b5563; }
  .open-link { color: #2563eb; text-decoration: none; font-size: 12px; font-weight: 600; }
  .open-link:hover { text-decoration: underline; }
  .summary { font-size: 12px; color: #6b7280; margin-bottom: 12px; }
  input[type=text] { width: 280px; padding: 6px 10px; border: 1px solid #d1d5db; border-radius: 6px; font-size: 13px; margin-bottom: 12px; }
</style>
</head>
<body>
<header>
  <h1>Foundation Dashboard</h1>
  <span style="opacity:0.6">·</span>
  <a href="/wizard">Wizard</a>
  <a href="/dashboard">FA Dashboard</a>
  <a href="/audited-financials">Audited Financials</a>
</header>
<main>
  <div class="card">
    <div class="summary" id="summary">Loading…</div>
    <div class="chips" id="chips"></div>
    <input type="text" id="search" placeholder="Filter by entity code or building name…">
    <table>
      <thead>
        <tr>
          <th>Entity</th>
          <th>Building</th>
          <th>FA</th>
          <th>2026 Approved</th>
          <th>2025 Audit</th>
          <th>Foundation</th>
          <th></th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
</main>
<script>
let _data = null;
let _filter = "all";
let _search = "";

function badgeApproved(state, count) {
  if (state === "imported") return '<span class="badge badge-ok">\u2713 Imported ' + count + ' rows</span>';
  if (state === "acknowledged_missing") return '<span class="badge badge-neutral">\u2713 No prior budget ack\u2019d</span>';
  return '<span class="badge badge-missing">Missing</span>';
}
function badgeAudit(state) {
  if (state === "confirmed") return '<span class="badge badge-ok">\u2713 Confirmed</span>';
  if (state === "extracted") return '<span class="badge badge-pending">Awaiting confirm</span>';
  if (state === "in_sp") return '<span class="badge badge-pending">In SP, not extracted</span>';
  return '<span class="badge badge-missing">Missing</span>';
}
function badgeFoundation(at) {
  if (at) return '<span class="badge badge-ok">\u2713 Confirmed</span>';
  return '<span class="badge badge-pending">Pending</span>';
}

function renderChips() {
  if (!_data) return;
  const c = _data.counts;
  const chips = [
    {k: "all", label: "All", count: c.all},
    {k: "pending", label: "Foundation pending", count: c.pending},
    {k: "confirmed", label: "Foundation confirmed", count: c.confirmed},
    {k: "no_prior_budget", label: "No prior budget", count: c.no_prior_budget},
    {k: "audit_missing", label: "Audit missing", count: c.audit_missing},
    {k: "audit_extracted_unconfirmed", label: "Audit awaiting confirm", count: c.audit_extracted_unconfirmed},
  ];
  const wrap = document.getElementById("chips");
  wrap.innerHTML = chips.map(c => '<div class="chip ' + (_filter === c.k ? "active" : "") + '" data-k="' + c.k + '">' + c.label + '<span class="chip-count">' + c.count + '</span></div>').join("");
  wrap.querySelectorAll(".chip").forEach(el => {
    el.addEventListener("click", () => { _filter = el.dataset.k; renderChips(); renderRows(); });
  });
}

function matchesFilter(e) {
  if (_filter === "all") return true;
  if (_filter === "pending") return !e.foundation_confirmed_at;
  if (_filter === "confirmed") return !!e.foundation_confirmed_at;
  if (_filter === "no_prior_budget") return e.foundation_no_prior_budget;
  if (_filter === "audit_missing") return e.audit === "missing";
  if (_filter === "audit_extracted_unconfirmed") return e.audit === "extracted";
  return true;
}

function renderRows() {
  if (!_data) return;
  const search = _search.toLowerCase();
  const filtered = (_data.entities || []).filter(e => matchesFilter(e) && (
    !search || (e.entity_code + " " + e.building_name).toLowerCase().indexOf(search) !== -1
  ));
  const tbody = document.getElementById("tbody");
  tbody.innerHTML = filtered.map(e => (
    '<tr><td class="ec">' + e.entity_code + '</td>' +
    '<td>' + (e.building_name || "") + '</td>' +
    '<td>' + (e.fa_name || "\u2014") + '</td>' +
    '<td>' + badgeApproved(e.approved_budget, e.approved_budget_summary_rows) + '</td>' +
    '<td>' + badgeAudit(e.audit) + (e.audit_upload_id ? ' <a class="open-link" href="/audited-financials/review/' + e.audit_upload_id + '" target="_blank">review \u2197</a>' : "") + '</td>' +
    '<td>' + badgeFoundation(e.foundation_confirmed_at) + '</td>' +
    '<td><a class="open-link" href="/wizard" target="_blank">open \u2197</a></td></tr>'
  )).join("");
  document.getElementById("summary").textContent = filtered.length + " of " + _data.total + " entities";
}

document.getElementById("search").addEventListener("input", e => { _search = e.target.value; renderRows(); });

fetch("/api/admin/foundation-summary").then(r => r.json()).then(j => {
  _data = j;
  renderChips();
  renderRows();
});
</script>
</body>
</html>
"""


@app.route("/api/wizard/<entity_code>/use-approved-budget", methods=["POST"])
def wizard_use_approved_budget(entity_code):
    """Stage a 2026 Approved Budget file from SharePoint. Same staging-only
    semantics as use-sp-source — actual import happens at Build Budget time.
    """
    # Force source_type = approved_2026 regardless of body (this endpoint is dedicated)
    data = request.get_json() or {}
    data["source_type"] = "approved_2026"
    # Patch request to ensure source_type is set before _wizard_record_selection reads it
    from flask import g as _g
    _g.__forced_payload = data
    # Inject via a small monkey-patch: re-build the request body.
    # Simpler: just call _wizard_record_selection directly — it reads request.get_json().
    # We need to re-set the request payload — but Flask doesn\'t let us mutate it.
    # Workaround: pass via closure. _wizard_record_selection accepts a default.
    return _wizard_record_selection(entity_code, source_label_default="approved_2026")

@app.route("/api/admin/move-approved-budgets", methods=["POST", "GET"])
@require_admin
def admin_move_approved_budgets():
    """Move every file in the flat \"2026 budget approved budgets only/\" folder
    into its matching <entity>/ folder, based on the entity_code prefix in the
    filename.

    Query params:
      dry_run=1  — list what would move without touching anything (default 1)
      confirm=MOVE — required to actually move
    """
    import re
    import urllib.parse
    dry_run = request.args.get("dry_run") not in ("0", "false", "no")
    confirm = request.args.get("confirm")
    if not dry_run and confirm != "MOVE":
        return jsonify({"error": "confirm=MOVE required to actually move (or omit dry_run for default dry_run=1)"}), 400

    drive_id = _graph_get_drive_id()
    flat_folder = SHAREPOINT_APPROVED_BUDGETS_PATH
    flat_encoded = urllib.parse.quote(flat_folder, safe="/")

    # 1. List files in the flat folder.
    listing = _graph_get(f"drives/{drive_id}/root:/{flat_encoded}:/children")
    files = [it for it in listing.get("value", []) if "folder" not in it]

    # 2. Bulk-list all entity subfolders under 2027 Budget/ in one Graph call.
    parent_listing = _graph_get(f"drives/{drive_id}/root:/{urllib.parse.quote(SHAREPOINT_2027_FOLDER_PATH, safe='/')}:/children")
    entity_folder_cache = {}
    for it in parent_listing.get("value", []):
        if "folder" in it:
            entity_folder_cache[it.get("name")] = it.get("id")

    def get_entity_folder_id(ec):
        return entity_folder_cache.get(ec)

    # 3. For each file, find its target entity folder.
    plan = {"would_move": [], "no_match": [], "errors": []}
    entity_pattern = re.compile(r"^(\d{2,4})(?:\D|$)")
    for f in files:
        name = f.get("name", "")
        m = entity_pattern.match(name)
        if not m:
            plan["no_match"].append({"name": name, "reason": "filename does not start with 2-4 digit entity code"})
            continue
        ec = m.group(1)
        target_folder_id = get_entity_folder_id(ec)
        if not target_folder_id:
            plan["no_match"].append({"name": name, "entity_code": ec, "reason": f"entity folder {ec}/ does not exist"})
            continue
        plan["would_move"].append({
            "name": name,
            "entity_code": ec,
            "from": flat_folder,
            "to": SHAREPOINT_2027_FOLDER_PATH + "/" + ec,
            "item_id": f.get("id"),
            "target_folder_id": target_folder_id,
        })

    plan["count_would_move"] = len(plan["would_move"])
    plan["count_no_match"] = len(plan["no_match"])
    plan["dry_run"] = dry_run

    if dry_run:
        return jsonify(plan)

    # 4. Execute moves.
    moved = []
    failed = []
    for entry in plan["would_move"]:
        try:
            _graph_patch(
                f"drives/{drive_id}/items/{entry['item_id']}",
                {"parentReference": {"id": entry["target_folder_id"]}},
            )
            moved.append(entry)
        except Exception as e:
            entry["error"] = str(e)
            failed.append(entry)
    return jsonify({
        "dry_run": False,
        "moved_count": len(moved),
        "failed_count": len(failed),
        "no_match_count": len(plan["no_match"]),
        "moved_sample": moved[:5],
        "failed": failed,
        "no_match": plan["no_match"],
    })


# ─── 2025 Audit Sync ─────────────────────────────────────────────────────────
# Admin-triggered: copies new PDFs from the 2025 audit master folder into each
# entity's 2027 Budget/<entity>/Supporting Documents/. Idempotent. No scheduler
# in v1 — invoked by POST /api/admin/audit-sync/run. Run log queryable via
# GET /api/admin/audit-sync-log.

SHAREPOINT_AUDIT_MASTER_PATH = "01 - Accounting General/Audited Financials/2025 Audited Financial Statements"


def _graph_put_content(path, body_bytes, content_type="application/octet-stream"):
    """PUT raw bytes to a SharePoint path. Creates or replaces the file.
    For files <4MB this is a single PUT to /content; audit PDFs are all under that.
    Returns parsed JSON (item metadata) or raises RuntimeError.
    """
    import urllib.request
    import urllib.parse
    drive_id = _graph_get_drive_id()
    encoded = urllib.parse.quote(path, safe="/")
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded}:/content"
    token = _get_graph_token()
    req = urllib.request.Request(
        url,
        data=body_bytes,
        method="PUT",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": content_type,
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.request.HTTPError as e:
        body_text = ""
        try:
            body_text = e.read().decode("utf-8")[:600]
        except Exception:
            pass
        raise RuntimeError(f"Graph {e.code} {e.reason} on PUT {path}: {body_text}")


def _parse_audit_entity_code(filename):
    """Extract leading numeric token. Handles all variations seen in master folder.
    Examples that all return their entity code: "106 - 2025 audit.pdf", "140- 29-45.pdf",
    "140 29-45 ...", "933 - 132 E 35th ...". Returns string or None.
    """
    import re
    m = re.match(r"^\s*(\d+)", filename or "")
    return m.group(1) if m else None


def _parse_iso_dt(s):
    """Parse an ISO-8601 datetime string (e.g. from Graph) to naive UTC datetime."""
    if not s:
        return None
    from datetime import datetime as _dt
    try:
        if s.endswith("Z"):
            s = s[:-1]
        if "." in s:
            return _dt.strptime(s, "%Y-%m-%dT%H:%M:%S.%f")
        return _dt.strptime(s, "%Y-%m-%dT%H:%M:%S")
    except Exception:
        return None


def _sync_audit_folder_to_entities():
    """Copy 2025 audit PDFs from master folder to per-entity Supporting Documents.

    Behavior per file:
      - parse leading numeric token → entity_code
      - if entity_code not in active budgets → log "unmatched"
      - duplicate filenames for same entity → keep the most recently modified
      - dest folder has same filename + same size → log "skipped"
      - dest has same filename, different size, source mtime newer → REPLACE, log "replaced"
      - otherwise → COPY, log "copied"
      - on exception per-entity → log "error" with text, continue with next entity

    Returns {"run_id", "summary": {copied, skipped, replaced, unmatched, error}, "entries": [...]}.
    """
    import urllib.parse
    import uuid
    AuditSyncRun = workflow_models["AuditSyncRun"]
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY

    run_id = str(uuid.uuid4())
    summary = {"copied": 0, "skipped": 0, "replaced": 0, "unmatched": 0, "error": 0}
    entries = []

    active = {str(r[0]) for r in db.session.query(Budget.entity_code).filter_by(year=_BY).all()}

    drive_id = _graph_get_drive_id()
    audit_path_enc = urllib.parse.quote(SHAREPOINT_AUDIT_MASTER_PATH, safe="/")

    try:
        listing = _graph_get(f"drives/{drive_id}/root:/{audit_path_enc}:/children")
    except Exception as e:
        return {"run_id": run_id, "fatal_error": str(e), "summary": summary, "entries": []}

    pdfs = [it for it in listing.get("value", [])
            if "folder" not in it and (it.get("name") or "").lower().endswith(".pdf")]

    by_entity = {}
    unmatched_files = []
    for it in pdfs:
        ec = _parse_audit_entity_code(it.get("name", ""))
        if not ec or ec not in active:
            unmatched_files.append((ec, it))
            continue
        prev = by_entity.get(ec)
        if not prev or (it.get("lastModifiedDateTime", "") > prev.get("lastModifiedDateTime", "")):
            by_entity[ec] = it

    # Log unmatched
    for ec, it in unmatched_files:
        row = AuditSyncRun(
            run_id=run_id,
            entity_code=ec,
            source_filename=it.get("name") or "",
            source_size=it.get("size"),
            source_mtime=_parse_iso_dt(it.get("lastModifiedDateTime")),
            action="unmatched",
            error_text=("entity_code not active in current year"
                        if ec else "could not parse entity_code from filename"),
        )
        db.session.add(row)
        entries.append({"action": "unmatched", "entity_code": ec, "filename": it.get("name")})
        summary["unmatched"] += 1
    try:
        db.session.flush()
    except Exception:
        db.session.rollback()

    for ec, src in by_entity.items():
        src_name = src.get("name") or ""
        src_size = src.get("size") or 0
        src_mtime = _parse_iso_dt(src.get("lastModifiedDateTime"))
        dest_folder_path = f"{SHAREPOINT_2027_FOLDER_PATH}/{ec}/Supporting Documents"
        dest_full_path = f"{dest_folder_path}/{src_name}"
        dest_url = None
        action = None
        error_text = None

        try:
            dest_enc = urllib.parse.quote(dest_folder_path, safe="/")
            try:
                dest_listing = _graph_get(f"drives/{drive_id}/root:/{dest_enc}:/children")
                dest_files = {f.get("name"): f for f in dest_listing.get("value", []) if "folder" not in f}
            except RuntimeError as e:
                if "404" in str(e):
                    _sharepoint_ensure_entity_folder(ec)
                    dest_files = {}
                else:
                    raise

            existing = dest_files.get(src_name)
            if existing and (existing.get("size") or 0) == src_size:
                action = "skipped"
                dest_url = existing.get("webUrl")
            elif existing:
                exist_mtime = _parse_iso_dt(existing.get("lastModifiedDateTime"))
                if src_mtime and exist_mtime and src_mtime > exist_mtime:
                    _, body = _sharepoint_download_item(src.get("id"))
                    new_meta = _graph_put_content(dest_full_path, body, content_type="application/pdf")
                    action = "replaced"
                    dest_url = new_meta.get("webUrl")
                else:
                    action = "skipped"
                    dest_url = existing.get("webUrl")
            else:
                _, body = _sharepoint_download_item(src.get("id"))
                new_meta = _graph_put_content(dest_full_path, body, content_type="application/pdf")
                action = "copied"
                dest_url = new_meta.get("webUrl")
        except Exception as e:
            action = "error"
            error_text = str(e)[:1000]
            try:
                db.session.rollback()
            except Exception:
                pass

        row = AuditSyncRun(
            run_id=run_id,
            entity_code=ec,
            source_filename=src_name,
            source_size=src_size,
            source_mtime=src_mtime,
            dest_path=dest_full_path,
            dest_url=dest_url,
            action=action,
            error_text=error_text,
        )
        db.session.add(row)
        entries.append({"action": action, "entity_code": ec, "filename": src_name,
                        "dest_url": dest_url, "error": error_text})
        summary[action] = summary.get(action, 0) + 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return {"run_id": run_id, "fatal_error": f"commit failed: {e}",
                "summary": summary, "entries": entries}

    return {"run_id": run_id, "summary": summary, "entries": entries}


@app.route("/api/admin/audit-sync/run", methods=["POST"])
@require_admin
def admin_audit_sync_run():
    """ADMIN: trigger one run of _sync_audit_folder_to_entities. No body required.
    Returns {run_id, summary, entries}. Idempotent — re-running on a clean state
    is a no-op (everything skipped).
    """
    try:
        result = _sync_audit_folder_to_entities()
        return jsonify(result)
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/clear-source", methods=["POST"])
@require_admin
def admin_clear_source():
    """ADMIN: clear staged data for one or more source_types on an entity.

    Body: {"entity_code": "148", "source_types": ["approved_2026", "audit_2025"]}
      - source_types is a list; each is one of:
          approved_2026 | audit_2025 | ysl | expense_distribution | ap_aging | maint_proof
        Plus the special token "all" which clears every source_type and the
        staged wizard_selections_json.

    Per source_type, deletes the corresponding tables AND the staged selection
    in Budget.wizard_selections_json. Does NOT touch SharePoint files. Does NOT
    reset the Budget row's status / wizard_step (use /api/admin/wipe-entity-data
    for full reset).

    Returns {ok, entity_code, results: {<source_type>: {tables_cleared: {...}}}}.
    Idempotent — re-running on already-clear state is a no-op.
    """
    Budget = workflow_models["Budget"]
    BudgetSummaryRow = workflow_models["BudgetSummaryRow"]
    BudgetLine = workflow_models["BudgetLine"]
    from workflow import BUDGET_YEAR as _BY

    data = request.get_json() or {}
    ec = (data.get("entity_code") or "").strip()
    types = data.get("source_types") or []
    if not ec:
        return jsonify({"error": "entity_code required"}), 400
    if not isinstance(types, list) or not types:
        return jsonify({"error": "source_types (non-empty list) required"}), 400

    valid = {"approved_2026", "audit_2025", "ysl", "expense_distribution",
             "ap_aging", "maint_proof", "all"}
    bad = [t for t in types if t not in valid]
    if bad:
        return jsonify({"error": f"invalid source_types: {bad}; valid: {sorted(valid)}"}), 400
    if "all" in types:
        types = ["approved_2026", "audit_2025", "ysl", "expense_distribution",
                 "ap_aging", "maint_proof"]

    budget = Budget.query.filter_by(entity_code=ec, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {ec} year {_BY}"}), 404

    results = {}
    try:
        # Load staged selections to remove the cleared keys.
        try:
            staged = json.loads(budget.wizard_selections_json or "{}")
        except Exception:
            staged = {}

        for st in types:
            cleared = {}
            if st == "approved_2026":
                n = BudgetSummaryRow.query.filter_by(
                    entity_code=ec, budget_year=_BY).delete()
                cleared["budget_summary_rows"] = n
            elif st == "audit_2025":
                row = db.session.execute(db.text(
                    "DELETE FROM audit_uploads WHERE entity_code = :ec RETURNING id"
                ), {"ec": ec}).fetchall()
                cleared["audit_uploads"] = len(row)
                # Un-confirm Foundation: clearing the audit means mapping is gone.
                db.session.execute(db.text(
                    "UPDATE budgets SET foundation_confirmed_at = NULL, foundation_confirmed_by = NULL "
                    "WHERE entity_code = :ec"
                ), {"ec": ec})
                cleared["foundation_unconfirmed"] = 1
            elif st == "ysl":
                # YSL feeds budget_lines; identify by budget_id.
                n = BudgetLine.query.filter_by(budget_id=budget.id).delete()
                cleared["budget_lines"] = n
            elif st == "expense_distribution":
                db.session.execute(db.text(
                    "DELETE FROM expense_invoices WHERE report_id IN "
                    "(SELECT id FROM expense_reports WHERE entity_code = :ec)"
                ), {"ec": ec})
                row = db.session.execute(db.text(
                    "DELETE FROM expense_reports WHERE entity_code = :ec RETURNING id"
                ), {"ec": ec}).fetchall()
                cleared["expense_reports"] = len(row)
            elif st == "ap_aging":
                db.session.execute(db.text(
                    "DELETE FROM open_ap_invoices WHERE report_id IN "
                    "(SELECT id FROM open_ap_reports WHERE entity_code = :ec)"
                ), {"ec": ec})
                row = db.session.execute(db.text(
                    "DELETE FROM open_ap_reports WHERE entity_code = :ec RETURNING id"
                ), {"ec": ec}).fetchall()
                cleared["open_ap_reports"] = len(row)
            elif st == "maint_proof":
                db.session.execute(db.text(
                    "DELETE FROM maint_proof_units WHERE report_id IN "
                    "(SELECT id FROM maint_proof_reports WHERE entity_code = :ec)"
                ), {"ec": ec})
                row = db.session.execute(db.text(
                    "DELETE FROM maint_proof_reports WHERE entity_code = :ec RETURNING id"
                ), {"ec": ec}).fetchall()
                cleared["maint_proof_reports"] = len(row)
            staged.pop(st, None)
            results[st] = {"tables_cleared": cleared}
        budget.wizard_selections_json = json.dumps(staged) if staged else None
        db.session.commit()
        return jsonify({"ok": True, "entity_code": ec, "source_types": types, "results": results})
    except Exception as e:
        db.session.rollback()
        logger.exception(f"clear_source failed for entity {ec}")
        return jsonify({"error": str(e), "results_partial": results}), 500


@app.route("/api/admin/audit-sync-log", methods=["GET"])
@require_admin
def admin_audit_sync_log():
    """ADMIN: list recent AuditSyncRun rows. Query params:
       run_id (filter to a specific run), action (filter copied/skipped/etc),
       limit (default 100, max 500).
    """
    AuditSyncRun = workflow_models["AuditSyncRun"]
    try:
        limit = min(max(int(request.args.get("limit", "100")), 1), 500)
    except ValueError:
        limit = 100
    q = AuditSyncRun.query
    rid = request.args.get("run_id")
    if rid:
        q = q.filter(AuditSyncRun.run_id == rid)
    act = request.args.get("action")
    if act:
        q = q.filter(AuditSyncRun.action == act)
    rows = q.order_by(AuditSyncRun.id.desc()).limit(limit).all()
    return jsonify({"count": len(rows), "rows": [r.to_dict() for r in rows]})


@app.route("/api/wizard/<entity_code>/build-budget", methods=["POST"])
def wizard_build_budget(entity_code):
    """All-or-nothing build trigger.

    Reads Budget.wizard_selections_json, downloads every selected file from
    SharePoint, runs each through its parser, and writes the results to the DB.
    On ANY parser/import error, rolls back the entire transaction — DB stays
    exactly as it was before the click.

    On success: stamps Budget.wizard_completed_at = NOW.

    Phase E gate: aborts with 400 if Foundation is not confirmed.
    """
    from datetime import datetime as _dt
    Budget = workflow_models["Budget"]
    BudgetSummaryRow = workflow_models["BudgetSummaryRow"]
    from workflow import BUDGET_YEAR as _BY, apply_summary_prefix_override

    budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {entity_code}"}), 404

    # Phase E gate: Foundation must be confirmed before any build can run.
    if not budget.foundation_confirmed_at:
        return jsonify({
            "error": "Foundation not confirmed. Process the 2026 Approved Budget and confirm the 2025 Audit mapping before building.",
            "foundation_required": True,
        }), 400

    if budget.wizard_completed_at is not None:
        return jsonify({"error": "Budget already built. Discard via /selections DELETE if you want to re-run."}), 400

    try:
        selections = json.loads(budget.wizard_selections_json or "{}")
    except Exception:
        selections = {}

    if not selections:
        return jsonify({"error": "No file selections staged. Pick at least one file in Step 2."}), 400

    # Result containers — populated as each selection processes.
    summary = {
        "entity_code": entity_code,
        "selections_processed": [],
        "lines_written": 0,
        "summary_rows_written": 0,
        "errors": [],
    }

    # Snapshot for rollback awareness — Flask-SQLAlchemy will rollback on raise.
    try:
        # Process 2026 Approved Budget first (it sets up the row framework).
        if "approved_2026" in selections:
            sel = selections["approved_2026"]
            try:
                rows_count = _build_apply_approved_2026(
                    entity_code, sel, BudgetSummaryRow, _BY, apply_summary_prefix_override
                )
                summary["selections_processed"].append({"source_type": "approved_2026", "filename": sel.get("filename"), "rows": rows_count})
                summary["summary_rows_written"] += rows_count
            except Exception as e:
                logger.exception("approved_2026 parse failed")
                raise RuntimeError(f"2026 Approved Budget parse failed: {e}")

        # YSL / ExpDist / AP Aging / Maint Proof — TODO Phase B.2 wiring.
        # For now: acknowledge selection but flag as not yet wired so FA knows
        # they still need to handle these via the legacy upload flow.
        for st in ("ysl", "expense_distribution", "ap_aging", "maint_proof"):
            if st in selections:
                sel = selections[st]
                summary["selections_processed"].append({
                    "source_type": st,
                    "filename": sel.get("filename"),
                    "status": "selection_recorded_only",
                    "note": "Parser pipeline for this source type pending — file selection saved but not yet auto-imported.",
                })

        # Seed Budget.assumptions_json from staged wizard assumptions.
        # CFO defaults are file-based and read by the wizard UI for pre-population.
        # Anything the FA edited in Step 3 lands in selections["assumptions"];
        # we deep-merge into whatever may already be on the Budget row (typically
        # empty on a fresh build). Existing assumptions_json values win only for
        # keys the FA did NOT touch — FA edits override.
        staged_assumptions = selections.get("assumptions") or {}
        if isinstance(staged_assumptions, dict) and staged_assumptions:
            try:
                existing = json.loads(budget.assumptions_json or "{}")
            except Exception:
                existing = {}
            for key, value in staged_assumptions.items():
                if isinstance(value, dict) and isinstance(existing.get(key), dict):
                    existing[key].update(value)
                else:
                    existing[key] = value
            budget.assumptions_json = json.dumps(existing)
            summary["assumptions_seeded"] = True
        else:
            summary["assumptions_seeded"] = False

        # Stamp wizard_completed_at on success
        budget.wizard_completed_at = _dt.utcnow()
        # Clear selections — they\'ve been consumed
        budget.wizard_selections_json = None

        db.session.commit()
        summary["ok"] = True
        summary["wizard_completed_at"] = budget.wizard_completed_at.isoformat()
        return jsonify(summary)

    except Exception as e:
        db.session.rollback()
        logger.exception(f"build-budget failed for entity {entity_code}")
        summary["ok"] = False
        summary["errors"].append(str(e))
        summary["fatal_error"] = str(e)
        return jsonify(summary), 500


def _build_apply_audit_2025(entity_code, selection):
    """Download + extract a 2025 audit PDF from SharePoint into the AuditUpload
    pipeline. Sets status='extracted' (skips mapping — FA picks auditor profile
    and confirms at /audited-financials/review/<upload_id>).

    Returns dict with upload_id, raw_extraction summary, review_url.
    Raises on failure.
    """
    import sys
    from pathlib import Path
    from datetime import datetime as _dt
    AuditUpload = af_models["AuditUpload"]

    item_id = (selection.get("item_id") or "").strip()
    if not item_id:
        raise RuntimeError("audit_2025 selection missing item_id")

    filename, file_bytes = _sharepoint_download_item(item_id)

    # Save PDF to the audited_financials data dir so the existing review UI
    # can find it later.
    af_path = str(Path(__file__).resolve().parent)
    if af_path not in sys.path:
        sys.path.insert(0, af_path)
    from audited_financials import _category_section  # ensure module loaded
    # Use the same helper as af_helpers if exposed, else compute path
    data_dir = af_helpers["get_data_dir"]() if "get_data_dir" in af_helpers else None
    if data_dir is None:
        data_dir = Path(__file__).resolve().parent / "data" / "audit_pdfs"
        data_dir.mkdir(parents=True, exist_ok=True)
    safe_filename = f"wizard_{entity_code}_{filename}"
    pdf_path = Path(data_dir) / safe_filename
    with open(pdf_path, "wb") as f:
        f.write(file_bytes)

    # Building name lookup via af_helpers (canonical source)
    try:
        buildings = af_helpers["get_buildings_list"]()
        building_name = next(
            (b.get("building_name") for b in buildings
             if str(b.get("entity_code")) == str(entity_code)),
            None,
        ) or f"Entity {entity_code}"
    except Exception:
        building_name = f"Entity {entity_code}"

    # Upsert AuditUpload — clean replace any existing row for this entity so
    # re-clicks overwrite (idempotent). Also un-confirms Foundation since the
    # mapping for the new audit hasn\'t been done yet.
    #
    # PROTECTION: skip the wipe if there's already a confirmed audit with a
    # real mapped_data payload — re-clicking the wizard tile must not destroy
    # FA mapping work. Status='confirmed' + empty mapped_data is treated as a
    # broken state and is allowed to be wiped.
    from workflow import BUDGET_YEAR as _BY_AUDIT_PRECHECK
    expected_fy = str(_BY_AUDIT_PRECHECK - 2)
    existing_confirmed = AuditUpload.query.filter_by(
        entity_code=entity_code,
        fiscal_year_end=expected_fy,
        status="confirmed",
    ).first()
    if existing_confirmed:
        try:
            md = json.loads(existing_confirmed.mapped_data) if existing_confirmed.mapped_data else {}
        except Exception:
            md = {}
        if isinstance(md, dict) and len(md) > 0:
            logger.info(
                f"[wizard audit_2025] {entity_code}: existing confirmed upload "
                f"id={existing_confirmed.id} fy={expected_fy} has {len(md)} mapped categories — "
                f"skipping re-extract to preserve FA mapping work."
            )
            return {
                "source_type": "audit_2025",
                "upload_id": existing_confirmed.id,
                "filename": existing_confirmed.pdf_filename or filename,
                "revenue_lines": 0,
                "expense_lines": 0,
                "review_url": f"/audited-financials/review/{existing_confirmed.id}",
                "status": "confirmed",
                "note": "Existing confirmed mapping preserved — wizard re-click is a no-op.",
                "skipped": True,
            }
    AuditUpload.query.filter_by(entity_code=entity_code).delete()
    db.session.execute(db.text(
        "UPDATE budgets SET foundation_confirmed_at = NULL, foundation_confirmed_by = NULL "
        "WHERE entity_code = :ec"
    ), {"ec": entity_code})
    db.session.flush()

    # fiscal_year_end is the load-bearing key for the summary endpoint's Col 2
    # lookup ([workflow.py:4035-4039]: WHERE fiscal_year_end = :fy). Without it,
    # confirmed audit data never reaches the budget summary tab even though
    # status='confirmed'. The wizard's audit_2025 slot is BY-2 by definition.
    from workflow import BUDGET_YEAR as _BY_AUDIT
    # Capture the SharePoint web URL (Office viewer link) so the review
    # page can deep-link to the source document. Survives Railway
    # ephemeral-fs wipes — unlike the local PDF cache which can disappear.
    sp_web_url = (selection or {}).get("web_url") or None
    upload = AuditUpload(
        entity_code=entity_code,
        building_name=building_name,
        profile_id=None,  # FA will pick at review
        fiscal_year_end=str(_BY_AUDIT - 2),
        pdf_filename=safe_filename,
        sharepoint_web_url=sp_web_url,
        status="uploaded",
    )
    db.session.add(upload)
    db.session.flush()  # need upload.id

    # Run Claude extraction
    extract_fn = af_helpers["extract_from_pdf"]
    extracted = extract_fn(str(pdf_path), building_name, entity_code=entity_code)
    if not extracted:
        raise RuntimeError("Claude extraction returned no data")

    upload.raw_extraction = json.dumps(extracted)
    upload.status = "extracted"
    upload.updated_at = _dt.utcnow()

    # The extracted dict has nested structure: revenue.items (list) and
    # expenses.categories (list of dicts each containing items list).
    def _count_items(node):
        if not node:
            return 0
        if isinstance(node, list):
            return len(node)
        if isinstance(node, dict):
            return len(node.get("items") or [])
        return 0
    rev_count = _count_items(extracted.get("revenue"))
    exp_node = extracted.get("expenses") or {}
    exp_cats = exp_node.get("categories") if isinstance(exp_node, dict) else exp_node
    exp_count = 0
    if isinstance(exp_cats, list):
        for c in exp_cats:
            exp_count += _count_items(c)
    elif isinstance(exp_cats, dict):
        for c in exp_cats.values():
            exp_count += _count_items(c)

    return {
        "source_type": "audit_2025",
        "upload_id": upload.id,
        "filename": filename,
        "revenue_lines": rev_count,
        "expense_lines": exp_count,
        "review_url": f"/audited-financials/review/{upload.id}",
        "status": "extracted",
        "note": "Open the review URL to assign auditor profile and confirm mapping.",
    }


def _build_apply_approved_2026(entity_code, selection, BudgetSummaryRow, BUDGET_YEAR, apply_summary_prefix_override):
    """Download + parse + upsert the 2026 Approved Budget.

    Raises on any failure. Returns count of rows written.
    """
    import sys
    import tempfile
    from pathlib import Path
    from datetime import datetime as _dt
    import json as _json

    item_id = (selection.get("item_id") or "").strip()
    if not item_id:
        raise RuntimeError("approved_2026 selection missing item_id")

    filename, file_bytes = _sharepoint_download_item(item_id)

    # Add budget_summary parsers to path
    bs_path = str(Path(__file__).resolve().parent.parent / "budget_summary")
    if bs_path not in sys.path:
        sys.path.insert(0, bs_path)
    from budget_summary_parser import parse_yrlycomp
    from batch_import import extract_importable_data, enrich_with_gl_map

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    parsed = parse_yrlycomp(tmp_path)
    if "error" in parsed:
        raise RuntimeError(f"yrlycomp parse error: {parsed['error']}")
    # FA #27a: pass expected Col 1 year so the parser blanks the column
    # instead of silently using a wrong-year audited_actual column.
    from workflow import BUDGET_YEAR as _BY
    _expected_c1_year = _BY - 3   # for 2027 cycle, Col 1 = 2024 Actual
    imported = extract_importable_data(parsed, expected_col1_year=_expected_c1_year)
    if "error" in imported:
        raise RuntimeError(f"extract_importable_data error: {imported['error']}")
    if imported.get("col1_warning"):
        logger.warning(
            f"[batch-import] {entity_code}: {imported['col1_warning']} "
            f"Will attempt confirmed-audit fallback."
        )
        # FA #27a fallback: try to fill Col 1 from a confirmed audit for
        # the expected year before persisting blank rows.
        filled = 0
        try:
            audit_actuals = af_helpers["get_confirmed_actuals"](entity_code, _expected_c1_year)
            if audit_actuals:
                for row in imported.get("rows", []):
                    label = row.get("label")
                    if label and row.get("col1_prior_actual") is None:
                        # Try direct, then case-insensitive label match against
                        # whatever shape get_confirmed_actuals returns.
                        v = audit_actuals.get(label)
                        if v is None:
                            for k, candidate in audit_actuals.items():
                                if str(k).strip().lower() == str(label).strip().lower():
                                    v = candidate
                                    break
                        if v is not None:
                            try:
                                row["col1_prior_actual"] = round(float(v), 2)
                                filled += 1
                            except (TypeError, ValueError):
                                pass
                imported["col1_audit_fallback_used"] = True
                imported["col1_audit_rows_filled"] = filled
                logger.info(
                    f"[batch-import] {entity_code}: confirmed-audit fallback "
                    f"filled {filled} Col 1 rows for year {_expected_c1_year}"
                )
            else:
                imported["col1_audit_fallback_used"] = False
                imported["col1_audit_rows_filled"] = 0
        except Exception as _fallback_err:
            logger.warning(
                f"[batch-import] {entity_code}: confirmed-audit fallback failed: {_fallback_err}"
            )
            imported["col1_audit_fallback_used"] = False
        # Final fallback: if neither expected-year column nor confirmed-year
        # audit produced data, re-extract WITHOUT the year guard so we use
        # whatever year the yrlycomp actually has (e.g. 2023 in a 2026-cycle
        # file used during the 2027 run). Surfaced via col1_actual_year so
        # the FA can see the real source year. Better to show last-known
        # audited numbers than a wall of blanks.
        if filled == 0:
            relax = extract_importable_data(parsed, expected_col1_year=None)
            if "error" not in relax:
                relaxed_rows = {r.get("display_order"): r for r in relax.get("rows", [])}
                refilled = 0
                for row in imported.get("rows", []):
                    if row.get("col1_prior_actual") is None:
                        rr = relaxed_rows.get(row.get("display_order"))
                        if rr and rr.get("col1_prior_actual") is not None:
                            row["col1_prior_actual"] = rr["col1_prior_actual"]
                            refilled += 1
                imported["col1_actual_year"] = (
                    int(str(relax.get("col1_label", "")).split()[0])
                    if relax.get("col1_label") and str(relax["col1_label"]).split()[0].isdigit()
                    else None
                )
                imported["col1_yrlycomp_fallback_used"] = True
                imported["col1_yrlycomp_rows_filled"] = refilled
                logger.info(
                    f"[batch-import] {entity_code}: yrlycomp-year fallback "
                    f"filled {refilled} Col 1 rows from year "
                    f"{imported.get('col1_actual_year')!r}"
                )
    enriched = enrich_with_gl_map(imported)

    # Upsert budget_summary_rows
    written = 0
    for i, row in enumerate(enriched.get("rows", [])):
        display_order = row.get("display_order") or (i + 1)
        existing = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code,
            budget_year=BUDGET_YEAR,
            display_order=display_order,
        ).first()
        incoming_prefixes = row.get("gl_prefixes") or []
        corrected_prefixes = apply_summary_prefix_override(row.get("label"), incoming_prefixes)
        gl_pj = _json.dumps(corrected_prefixes) if corrected_prefixes else None
        if existing:
            existing.label = row["label"]
            existing.section = row.get("section")
            existing.row_type = row.get("row_type", "data")
            existing.footnote_marker = row.get("footnote_marker")
            existing.col1_prior_actual = row.get("col1_prior_actual")
            existing.col6_approved_budget = row.get("col6_approved_budget")
            existing.source_tab = row.get("source_tab") or existing.source_tab
            existing.gl_prefixes_json = gl_pj or existing.gl_prefixes_json
            existing.source_file = filename or existing.source_file
            existing.updated_at = _dt.utcnow()
        else:
            db.session.add(BudgetSummaryRow(
                entity_code=entity_code,
                budget_year=BUDGET_YEAR,
                display_order=display_order,
                label=row["label"],
                section=row.get("section"),
                row_type=row.get("row_type", "data"),
                footnote_marker=row.get("footnote_marker"),
                col1_prior_actual=row.get("col1_prior_actual"),
                col6_approved_budget=row.get("col6_approved_budget"),
                col7_proposed_budget=None,
                source_tab=row.get("source_tab"),
                gl_prefixes_json=gl_pj,
                source_file=filename,
            ))
        written += 1

    # Side-effect: also populate the Building Info → Maintenance History
    # (coop) or Common Charges History (condo) card from the same XLSX's
    # "Income" tab. Failure here is non-fatal; the main D1 import job has
    # already done its work on budget_summary_rows.
    try:
        _populate_building_info_from_income(entity_code, tmp_path)
    except Exception:
        logger.warning(f"populate_building_info_from_income skipped for {entity_code}", exc_info=True)

    return written


def _populate_building_info_from_income(entity_code, xlsx_path):
    """Parse the Income tab's year-by-year history block and persist it onto
    the entity's BuildingInfo row. Coop → maintenance_history_json (with
    shares + perShare). Condo → common_charges_history_json (no shares).

    Preserves FA edits — only writes if the existing record is null OR
    consists entirely of the default-zeros placeholder rows.

    Returns dict with status, or None on no-op.
    """
    import sys as _sys
    from pathlib import Path as _Path
    bs_path = str(_Path(__file__).resolve().parent.parent / "budget_summary")
    if bs_path not in _sys.path:
        _sys.path.insert(0, bs_path)
    from budget_summary_parser import parse_income_history
    result = parse_income_history(xlsx_path)
    if "error" in result or not result.get("history"):
        return None
    btype = result.get("building_type")
    history = result.get("history", [])
    if btype not in ("coop", "condo"):
        return {"skipped": "unknown_building_type"}

    BuildingInfo = workflow_models["BuildingInfo"]
    info = BuildingInfo.query.filter_by(entity_code=entity_code).first()
    if not info:
        info = BuildingInfo(entity_code=entity_code)
        db.session.add(info)

    def _has_real_data(rows):
        if not isinstance(rows, list):
            return False
        for r in rows:
            if not isinstance(r, dict):
                continue
            if any((r.get(k) or 0) != 0 for k in ("shares", "perShare", "monthly", "annual")):
                return True
        return False

    if btype == "coop":
        existing_raw = info.maintenance_history_json
        if existing_raw:
            try:
                existing_rows = json.loads(existing_raw)
                if _has_real_data(existing_rows):
                    return {"skipped": "FA-edited", "type": btype}
            except Exception:
                pass
        normalized = [
            {
                "year": (h.get("year") if isinstance(h.get("year"), int) else 0),
                "year_label": h.get("year_label"),
                "shares": h.get("shares") or 0,
                "perShare": h.get("perShare") or 0,
                "monthly": h.get("monthly") or 0,
                "annual": h.get("annual") or 0,
                # Parser emits raw fraction (0.015 = 1.5%); FE input expects percent value (1.5).
                "increase": round((h.get("increase") or 0) * 100, 4),
            }
            for h in history
        ]
        info.maintenance_history_json = json.dumps(normalized)
    else:  # condo
        existing_raw = info.common_charges_history_json
        if existing_raw:
            try:
                existing_rows = json.loads(existing_raw)
                if _has_real_data(existing_rows):
                    return {"skipped": "FA-edited", "type": btype}
            except Exception:
                pass
        normalized = [
            {
                "year": (h.get("year") if isinstance(h.get("year"), int) else 0),
                "year_label": h.get("year_label"),
                "monthly": h.get("monthly") or 0,
                "annual": h.get("annual") or 0,
                "increase": round((h.get("increase") or 0) * 100, 4),
            }
            for h in history
        ]
        info.common_charges_history_json = json.dumps(normalized)

    db.session.commit()
    return {"populated": len(history), "type": btype, "gl_code": result.get("gl_code")}


def _wizard_download_sp_to_tmp(selection, suffix):
    """Download a SharePoint file by item_id into a tmp .xlsx (or other suffix).
    Returns (Path, filename). Caller is responsible for deleting the tmp file.
    Raises RuntimeError on missing item_id or download failure.
    """
    import tempfile
    from pathlib import Path
    item_id = (selection.get("item_id") or "").strip()
    if not item_id:
        raise RuntimeError("selection missing item_id")
    filename, file_bytes = _sharepoint_download_item(item_id)
    fd, tmp_path = tempfile.mkstemp(suffix=suffix or ".xlsx")
    try:
        import os as _os
        with _os.fdopen(fd, "wb") as fh:
            fh.write(file_bytes)
    except Exception:
        try:
            import os as _os
            _os.unlink(tmp_path)
        except Exception:
            pass
        raise
    return Path(tmp_path), filename


def _stage_wizard_period_for_entity(entity_code, month_num):
    """Write budget_period to the wizard's staged assumptions for an entity.

    Used by YSL auto-detect: when we successfully parse a period from the
    YSL header, pre-fill the wizard's staged assumptions so the FA sees the
    detected month already selected in Step 3. FA can still change it.

    The Build Budget step already deep-merges staged assumptions into
    Budget.assumptions_json (app.py around line 6843), so no additional
    plumbing is needed.

    Idempotent: if the FA has already set a period manually, this overwrites
    it. The wizard's UI shows the value either way; FA gets visual confirmation
    on next render.
    """
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY
    budget = Budget.query.filter_by(entity_code=str(entity_code), year=_BY).first()
    if not budget:
        return  # No-op; build hasn't been started yet
    try:
        sels = json.loads(budget.wizard_selections_json or "{}")
    except Exception:
        sels = {}
    assumptions = sels.get("assumptions") or {}
    if not isinstance(assumptions, dict):
        assumptions = {}
    # YSL covers prior-year actuals; year portion is BUDGET_YEAR - 1.
    mm = str(int(month_num)).zfill(2)
    yyyy = _BY - 1
    assumptions["budget_period"] = f"{mm}/{yyyy}"
    sels["assumptions"] = assumptions
    budget.wizard_selections_json = json.dumps(sels)
    db.session.commit()


def _build_apply_ysl(entity_code, selection):
    """Download + parse a YSL file from SharePoint and write GL lines to budget_lines.
    Idempotent: re-running for the same entity overwrites prior YSL data via
    workflow_helpers["store_all_lines"] (fresh_start=False = upsert).
    """
    import os as _os
    from budget_system.ysl_parser import parse_ysl_file
    tmp_path, filename = _wizard_download_sp_to_tmp(selection, ".xlsx")
    try:
        gl_data, property_info = parse_ysl_file(tmp_path)
        # The YSL file's internal entity may or may not match the FA's chosen
        # entity (Yardi has a known bug returning wrong-entity data sometimes).
        # Trust the FA's wizard context: store under entity_code, not what
        # the file claims.
        building_name = property_info.get("property_name") or f"Entity {entity_code}"
        merged = None
        try:
            merged = merge_assumptions(entity_code)
        except Exception:
            pass
        workflow_helpers["store_all_lines"](
            str(entity_code), building_name, gl_data, TEMPLATE_PATH,
            assumptions=merged, fresh_start=False,
        )
        # FA #1 follow-up: auto-detect budget_period from YSL header.
        # Hybrid model — we PRE-FILL the wizard's staged assumptions so the
        # period dropdown shows the detected month; FA can still override.
        # If parsing failed or no month was detected, leave the dropdown blank.
        detected_month = property_info.get("budget_period_month")
        period_set = False
        if detected_month and 1 <= detected_month <= 12:
            try:
                _stage_wizard_period_for_entity(entity_code, detected_month)
                period_set = True
            except Exception as _e:
                # Don't fail the YSL apply if staging the period fails; just
                # log and let FA pick manually.
                import logging
                logging.getLogger(__name__).warning(
                    f"Failed to stage detected YSL period for {entity_code}: {_e}"
                )
        # Count GL rows we ingested (gl_data shape: {gl_code: {sheet, ...}})
        gl_count = len(gl_data) if hasattr(gl_data, "__len__") else 0
        return {
            "source_type": "ysl",
            "filename": filename,
            "gl_lines": gl_count,
            "file_entity": property_info.get("property_code"),
            "stored_under_entity": str(entity_code),
            "detected_period_month": detected_month,
            "period_auto_set": period_set,
        }
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


def _build_apply_expense_distribution(entity_code, selection):
    """Download + parse an Expense Distribution file from SharePoint, store
    the report, and apply accrual adjustments to budget_lines (mirrors the
    /generate flow at app.py:1176-1184). Skipping apply_accrual_adjustments
    leaves BudgetLine.accrual_adj at 0 across the entity, so the wizard's
    forecast/preview ignores expense-driven accruals.
    """
    import os as _os
    try:
        from expense_distribution import parse_expense_distribution
    except ImportError:
        from budget_app.expense_distribution import parse_expense_distribution
    tmp_path, filename = _wizard_download_sp_to_tmp(selection, ".xlsx")
    try:
        exp_entity, period_from, period_to, invoices = parse_expense_distribution(str(tmp_path))
        if not invoices:
            return {
                "source_type": "expense_distribution",
                "filename": filename,
                "invoices": 0,
                "warning": "Parser found no invoices in this file.",
            }
        report = ed_helpers["store_expense_report"](
            str(entity_code), period_from, period_to, invoices, filename,
        )
        accrual_applied = 0
        accrual_error = None
        if period_from and report and getattr(report, "id", None):
            try:
                accrual_result = ed_helpers["apply_accrual_adjustments"](
                    str(entity_code), report.id, period_from,
                )
                accrual_applied = (accrual_result or {}).get("applied", 0)
            except Exception as _e:
                accrual_error = str(_e)[:200]
        return {
            "source_type": "expense_distribution",
            "filename": filename,
            "invoices": len(invoices),
            "period_from": period_from,
            "period_to": period_to,
            "accrual_adjustments_applied": accrual_applied,
            "accrual_error": accrual_error,
            "file_entity": exp_entity,
            "stored_under_entity": str(entity_code),
        }
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


def _build_apply_ap_aging(entity_code, selection):
    """Download + parse an AP Aging file from SharePoint."""
    import os as _os
    from open_ap import parse_open_ap_report
    tmp_path, filename = _wizard_download_sp_to_tmp(selection, ".xlsx")
    try:
        ap_entity, invoices = parse_open_ap_report(str(tmp_path))
        if not invoices:
            return {
                "source_type": "ap_aging",
                "filename": filename,
                "invoices": 0,
                "warning": "Parser found no invoices in this file.",
            }
        oa_helpers["store_open_ap_report"](str(entity_code), invoices, filename)
        unpaid_applied = 0
        try:
            unpaid_result = oa_helpers["apply_unpaid_bills"](str(entity_code))
            unpaid_applied = (unpaid_result or {}).get("applied", 0)
        except Exception:
            pass
        return {
            "source_type": "ap_aging",
            "filename": filename,
            "invoices": len(invoices),
            "unpaid_bills_applied": unpaid_applied,
            "file_entity": ap_entity,
            "stored_under_entity": str(entity_code),
        }
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


def _build_apply_maint_proof(entity_code, selection):
    """Download + parse a Maintenance Proof file from SharePoint."""
    import os as _os
    tmp_path, filename = _wizard_download_sp_to_tmp(selection, ".xlsx")
    try:
        report_title, units, total_shares = mp_helpers["parse_maintenance_proof"](str(tmp_path))
        mp_helpers["store_maintenance_proof"](
            str(entity_code), report_title, units, total_shares, filename,
        )
        return {
            "source_type": "maint_proof",
            "filename": filename,
            "report_title": report_title,
            "units": len(units) if units else 0,
            "total_shares": total_shares,
            "stored_under_entity": str(entity_code),
        }
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass


@app.route("/api/wizard/<entity_code>/foundation-status", methods=["GET"])
def wizard_foundation_status(entity_code):
    """Return the Foundation gate state for an entity. Used by Step 2 to render
    the cards and decide whether Steps 3+ are unlocked.

    Response shape:
    {
      "entity_code": "148",
      "approved_budget": "imported"|"in_sp_not_imported"|"missing"|"acknowledged_missing",
      "approved_budget_summary_rows": <int>,
      "audit": "missing"|"in_sp"|"extracted"|"confirmed",
      "audit_upload_id": <int|null>,
      "foundation_confirmed_at": "...iso..."|null,
      "foundation_no_prior_budget": <bool>,
      "blocking_reason": <str|null>,
      "review_url": "/audited-financials/review/<id>"|null,
    }
    """
    Budget = workflow_models["Budget"]
    BudgetSummaryRow = workflow_models["BudgetSummaryRow"]
    AuditUpload = af_models["AuditUpload"]
    from workflow import BUDGET_YEAR as _BY

    budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {entity_code}"}), 404

    # Approved budget state
    summary_count = BudgetSummaryRow.query.filter_by(
        entity_code=entity_code, budget_year=_BY
    ).count()

    # SP scan tells us whether the file is even available
    try:
        sp = _sharepoint_list_entity_sources(entity_code)
        approved_in_sp = bool(sp.get("by_source_type", {}).get("approved_2026"))
    except Exception:
        approved_in_sp = False

    if summary_count > 0:
        approved_state = "imported"
    elif budget.foundation_no_prior_budget:
        approved_state = "acknowledged_missing"
    elif approved_in_sp:
        approved_state = "in_sp_not_imported"
    else:
        approved_state = "missing"

    # Audit state
    audit = AuditUpload.query.filter_by(entity_code=entity_code).order_by(
        AuditUpload.id.desc()
    ).first()
    audit_state = "missing"
    audit_upload_id = None
    review_url = None
    if audit:
        audit_upload_id = audit.id
        review_url = f"/audited-financials/review/{audit.id}"
        if audit.status == "confirmed":
            audit_state = "confirmed"
        elif audit.status in ("mapped", "extracted"):
            audit_state = "extracted"
        else:
            audit_state = "in_sp"
    else:
        try:
            sp_audit = sp.get("by_source_type", {}).get("audit_2025") if approved_in_sp or True else []
            if sp_audit:
                audit_state = "in_sp"
        except Exception:
            pass

    # Compute blocking_reason
    blocking_reason = None
    if not budget.foundation_confirmed_at:
        if approved_state == "missing":
            blocking_reason = "No 2026 approved budget in SharePoint. Acknowledge \"no prior budget\" or upload one."
        elif approved_state == "in_sp_not_imported":
            blocking_reason = "Click Process on the 2026 Approved Budget card to import categories."
        elif audit_state == "missing":
            blocking_reason = "No 2025 audit in SharePoint for this entity. Upload one to proceed."
        elif audit_state == "in_sp":
            blocking_reason = "Click Process on the 2025 Audited Financial card to extract."
        elif audit_state == "extracted":
            blocking_reason = "Click Review & Confirm Mapping to finalize the audit and complete the Foundation."

    return jsonify({
        "entity_code": entity_code,
        "approved_budget": approved_state,
        "approved_budget_summary_rows": summary_count,
        "audit": audit_state,
        "audit_upload_id": audit_upload_id,
        "foundation_confirmed_at": (
            budget.foundation_confirmed_at.isoformat()
            if budget.foundation_confirmed_at else None
        ),
        "foundation_no_prior_budget": bool(budget.foundation_no_prior_budget),
        "blocking_reason": blocking_reason,
        "review_url": review_url,
    })


@app.route("/api/wizard/<entity_code>/acknowledge-no-prior-budget", methods=["POST"])
def wizard_ack_no_prior_budget(entity_code):
    """Set Budget.foundation_no_prior_budget = True for entities without a 2026
    approved budget XLSX. Audit extraction will then use CENTURY_CATEGORIES.
    Body: {"acknowledged": true}.
    """
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY
    budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {entity_code}"}), 404
    body = request.get_json(silent=True) or {}
    budget.foundation_no_prior_budget = bool(body.get("acknowledged", True))
    db.session.commit()
    return jsonify({
        "ok": True,
        "entity_code": entity_code,
        "foundation_no_prior_budget": budget.foundation_no_prior_budget,
    })


@app.route("/api/wizard/<entity_code>/sharepoint-sources", methods=["GET"])
def wizard_sharepoint_sources(entity_code):
    """Read-only: return what's in this entity\'s SharePoint Supporting
    Documents folder, classified by source type. Used by Step 2 to show
    the FA which sources are pre-staged in SharePoint.
    """
    try:
        return jsonify(_sharepoint_list_entity_sources(entity_code))
    except Exception as e:
        logger.error(f"wizard_sharepoint_sources({entity_code}) failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/wizard/<entity_code>/upload-to-sp", methods=["POST"])
def wizard_upload_to_sp(entity_code):
    """Manual-upload escape hatch: FA drops a file from their computer; we
    classify it via SHAREPOINT_SOURCE_PATTERNS and PUT to the right SP path:
      - approved_2026 (filename contains "approved" + .xlsx) → 2027 Budget/<ec>/
      - everything else → 2027 Budget/<ec>/Supporting Documents/

    Then the FROM SHAREPOINT panel refresh picks it up. Replaces files at the
    same path (Graph default conflictBehavior=replace for PUT /content).

    Body: multipart, field "file".
    Returns: {ok, classified_as, dest_path, dest_url, note?}
    """
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "file required (multipart field 'file')"}), 400

    filename = file.filename
    body = file.read()
    if not body:
        return jsonify({"error": "file is empty"}), 400

    # Classify destination
    classified = _classify_sharepoint_filename(filename)
    base = SHAREPOINT_2027_FOLDER_PATH + "/" + str(entity_code)
    if classified == "approved_2026":
        dest_path = f"{base}/{filename}"
    else:
        # Ensure Supporting Documents folder exists
        try:
            _sharepoint_ensure_entity_folder(entity_code)
        except Exception as e:
            logger.warning(f"ensure_entity_folder for {entity_code} failed (continuing): {e}")
        dest_path = f"{base}/Supporting Documents/{filename}"

    # Sniff content-type for upload
    low = filename.lower()
    if low.endswith(".pdf"):
        ct = "application/pdf"
    elif low.endswith(".xlsx"):
        ct = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif low.endswith(".xls"):
        ct = "application/vnd.ms-excel"
    elif low.endswith(".csv"):
        ct = "text/csv"
    else:
        ct = "application/octet-stream"

    try:
        meta = _graph_put_content(dest_path, body, content_type=ct)
        return jsonify({
            "ok": True,
            "classified_as": classified or "unmatched",
            "dest_path": dest_path,
            "dest_url": meta.get("webUrl"),
            "filename": filename,
            "size": len(body),
            "note": ("Upload routed to entity top folder (approved budget pattern)."
                     if classified == "approved_2026"
                     else ("Upload routed to Supporting Documents (classified as "
                           + str(classified) + ")."
                           if classified else
                           "Upload routed to Supporting Documents (unmatched filename — "
                           "FROM SHAREPOINT panel will list it under Other files.)")),
        })
    except Exception as e:
        logger.exception(f"upload-to-sp failed for {entity_code}/{filename}")
        return jsonify({"error": str(e)}), 500




@app.route("/api/admin/entity-trace/<entity_code>", methods=["GET"])
@require_admin
def admin_entity_trace(entity_code):
    """Trace upload provenance for a single entity. Shows which columns of
    budget_lines have non-default values, distinct sources of revisions, and
    a sample line so we can infer how the source file was processed.
    """
    Budget = workflow_models["Budget"]
    BudgetLine = workflow_models["BudgetLine"]
    BudgetRevision = workflow_models["BudgetRevision"]

    budget = Budget.query.filter_by(entity_code=entity_code, year=2027).first()
    if not budget:
        return jsonify({"error": f"No 2027 budget for entity {entity_code}"}), 404

    # Column population: how many lines have non-zero/non-null in each numeric column.
    cols = ["prior_year", "ytd_actual", "ytd_budget", "current_budget",
            "accrual_adj", "unpaid_bills", "increase_pct", "proposed_budget",
            "estimate_override", "forecast_override", "fa_override_value"]
    col_filled = {}
    for c in cols:
        # count rows where col is non-null and non-zero (or just non-null for nullable cols)
        if c in ("estimate_override", "forecast_override", "fa_override_value"):
            res = db.session.execute(
                db.text(f"SELECT COUNT(*) FROM budget_lines WHERE budget_id = :bid AND {c} IS NOT NULL"),
                {"bid": budget.id},
            ).scalar()
        else:
            res = db.session.execute(
                db.text(f"SELECT COUNT(*) FROM budget_lines WHERE budget_id = :bid AND ABS(COALESCE({c}, 0)) > 0.005"),
                {"bid": budget.id},
            ).scalar()
        col_filled[c] = int(res or 0)

    # Distinct revision sources + actions for this budget
    rev_summary = db.session.execute(db.text("""
        SELECT action, source, COUNT(*) as cnt,
               MIN(created_at) as first_at, MAX(created_at) as last_at
          FROM budget_revisions
         WHERE budget_id = :bid
         GROUP BY action, source
         ORDER BY cnt DESC
    """), {"bid": budget.id}).fetchall()

    # 3 sample lines: pick income, expense, payroll
    sample_lines = db.session.execute(db.text("""
        SELECT gl_code, description, sheet_name, category,
               prior_year, ytd_actual, ytd_budget, current_budget,
               accrual_adj, unpaid_bills, increase_pct, proposed_budget,
               estimate_override, forecast_override, notes
          FROM budget_lines
         WHERE budget_id = :bid
         ORDER BY sheet_name, row_num
         LIMIT 5
    """), {"bid": budget.id}).fetchall()

    # Total line count + counts by sheet
    sheet_counts = db.session.execute(db.text("""
        SELECT sheet_name, COUNT(*) as cnt
          FROM budget_lines WHERE budget_id = :bid
         GROUP BY sheet_name ORDER BY cnt DESC
    """), {"bid": budget.id}).fetchall()

    return jsonify({
        "entity_code": entity_code,
        "budget_id": budget.id,
        "year": budget.year,
        "wizard_step": budget.wizard_step,
        "wizard_completed_at": budget.wizard_completed_at.isoformat() if budget.wizard_completed_at else None,
        "status": budget.status,
        "total_lines": sum(s[1] for s in sheet_counts),
        "lines_by_sheet": [{"sheet_name": s[0], "count": s[1]} for s in sheet_counts],
        "column_population": col_filled,
        "revisions_grouped": [
            {"action": r[0], "source": r[1], "count": r[2],
             "first_at": r[3].isoformat() if r[3] else None,
             "last_at": r[4].isoformat() if r[4] else None}
            for r in rev_summary
        ],
        "sample_lines": [
            {"gl_code": l[0], "description": l[1], "sheet_name": l[2], "category": l[3],
             "prior_year": l[4], "ytd_actual": l[5], "ytd_budget": l[6], "current_budget": l[7],
             "accrual_adj": l[8], "unpaid_bills": l[9], "increase_pct": l[10],
             "proposed_budget": l[11], "estimate_override": l[12], "forecast_override": l[13],
             "notes": l[14]}
            for l in sample_lines
        ],
    })



@app.route("/api/admin/sync-afs/<entity_code>", methods=["POST"])
@require_admin
def admin_sync_afs(entity_code):
    """ADMIN: Scan SharePoint for AFS PDFs for one entity and create
    AuditUpload rows for any not yet ingested.

    Resolves FA #27b — currently AFS PDFs sit in SharePoint visible to FAs
    but no automation stages them for the audit-extraction pipeline. After
    this runs, new AFS PDFs appear in /audited-financials with status
    "uploaded", waiting for the FA to click "Extract" (existing flow).

    Body (optional):
      {"dry_run": true}  - report what would be created without writing
      {"force": true}    - re-upload even if a matching row exists

    Returns: {ok, entity_code, scanned, created, skipped, results: [{filename, year, status, ...}]}
    """
    AuditUpload = af_models["AuditUpload"]
    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get("dry_run"))
    force = bool(body.get("force"))
    return _do_afs_sync(entity_code, dry_run=dry_run, force=force, AuditUpload=AuditUpload)


@app.route("/api/admin/portfolio-health", methods=["GET"])
@require_admin
def admin_portfolio_health():
    """ADMIN: Portfolio-wide health check.

    Runs a battery of invariants across every Budget and surfaces issues
    grouped by severity. Read-only. Safe to run anytime. Built to catch
    the bug patterns we discovered manually for entity 168 — silent
    fallbacks, magic constants, per-building variation, math edge cases,
    half-built features.

    Query params:
      severity=critical|high|medium|low   - filter to one severity
      check=<check_name>                  - filter to one check
      entity_code=<ec>                    - scope to one entity
      include_passing=true                - include entities with no issues

    Returns:
      {
        "scanned_at": "...",
        "entities_scanned": N,
        "issues_total": M,
        "by_severity": {"critical": A, "high": B, "medium": C, "low": D},
        "by_check": {"period_not_set": X, ...},
        "issues": [
          {entity, check, severity, message, fix, details}, ...
        ]
      }
    """
    import json as _json
    from datetime import datetime as _dt
    from statistics import median as _median, stdev as _stdev

    Budget = workflow_models["Budget"]
    BudgetLine = workflow_models["BudgetLine"]
    BudgetSummaryRow = workflow_models["BudgetSummaryRow"]
    AuditUpload = af_models["AuditUpload"]
    from workflow import BUDGET_YEAR as _BY

    severity_filter = (request.args.get("severity") or "").strip().lower() or None
    check_filter = (request.args.get("check") or "").strip() or None
    entity_filter = (request.args.get("entity_code") or "").strip() or None
    include_passing = (request.args.get("include_passing") or "").lower() == "true"

    # ─── 1. Pre-fetch in-bulk to avoid N+1 queries ───
    budgets = Budget.query.filter_by(year=_BY).all()
    if entity_filter:
        budgets = [b for b in budgets if b.entity_code == entity_filter]
    budget_ids = [b.id for b in budgets]
    bid_to_ec = {b.id: b.entity_code for b in budgets}

    summary_rows_all = (
        BudgetSummaryRow.query
        .filter(BudgetSummaryRow.entity_code.in_([b.entity_code for b in budgets]))
        .filter_by(budget_year=_BY)
        .all()
    )
    rows_by_ec = {}
    for r in summary_rows_all:
        rows_by_ec.setdefault(r.entity_code, []).append(r)

    lines_all = BudgetLine.query.filter(BudgetLine.budget_id.in_(budget_ids)).all() if budget_ids else []
    lines_by_bid = {}
    for l in lines_all:
        lines_by_bid.setdefault(l.budget_id, []).append(l)

    audits_all = (
        AuditUpload.query
        .filter(AuditUpload.entity_code.in_([b.entity_code for b in budgets]))
        .all()
    )
    audits_by_ec = {}
    for au in audits_all:
        audits_by_ec.setdefault(au.entity_code, []).append(au)

    # ─── 2. Build peer-statistics lookup for outlier detection ───
    # For each summary row label, collect col6 (current budget) values across
    # all entities. We'll flag values >2σ from peer median.
    peer_values = {}  # label → [value, ...]
    for r in summary_rows_all:
        if r.row_type != "data" or r.col6_approved_budget is None:
            continue
        if r.col6_approved_budget <= 0:
            continue
        peer_values.setdefault(r.label, []).append(r.col6_approved_budget)
    peer_stats = {}  # label → {median, stdev, n}
    for label, vals in peer_values.items():
        if len(vals) < 5:  # need a meaningful sample
            continue
        try:
            peer_stats[label] = {
                "median": _median(vals),
                "stdev": _stdev(vals) if len(vals) > 1 else 0,
                "n": len(vals),
            }
        except Exception:
            pass

    # ─── 3. Run checks per entity ───
    issues = []

    def add(severity, check, entity, message, fix=None, details=None):
        if severity_filter and severity != severity_filter:
            return
        if check_filter and check != check_filter:
            return
        issues.append({
            "entity": entity, "check": check, "severity": severity,
            "message": message,
            "fix": fix, "details": details or {},
        })

    for budget in budgets:
        ec = budget.entity_code
        bid = budget.id
        try:
            assum = _json.loads(budget.assumptions_json or "{}")
        except Exception:
            assum = {}
        rows = rows_by_ec.get(ec, [])
        lines = lines_by_bid.get(bid, [])
        audits = audits_by_ec.get(ec, [])
        is_coop = (budget.building_type or "").lower() in ("coop", "co-op")

        # ── Check: period_not_set (CRITICAL) ──
        bp = (assum.get("budget_period") or "").strip()
        if not bp or "/" not in bp:
            add("critical", "period_not_set", ec,
                "budget_period is unset; forecasts use default 2-month YTD which produces wrong numbers",
                fix="Set period via dashboard banner or wizard Step 3")

        # ── Check: building_type_unknown (HIGH) ──
        if not budget.building_type:
            add("high", "building_type_unknown", ec,
                "Budget.building_type is empty; is_coop() heuristic falls back through legacy paths",
                fix="Backfill from buildings.csv via app startup hook (auto), or set manually")

        # ── Check: afs_not_confirmed (HIGH) ──
        # Filter audit uploads by status; expected confirmed for BUDGET_YEAR-2 (e.g., 2025 for 2027 cycle)
        target_year = str(_BY - 2)
        target_audits = [au for au in audits if str(au.fiscal_year_end or "") == target_year]
        confirmed = [au for au in target_audits if au.status == "confirmed"]
        non_confirmed = [au for au in target_audits if au.status != "confirmed"]
        if target_audits and not confirmed and non_confirmed:
            au = non_confirmed[0]
            add("high", "afs_not_confirmed", ec,
                f"FY{target_year} audit upload (id={au.id}) at status='{au.status}' — needs FA to click Confirm",
                fix=f"Open /audited-financials, find upload {au.id}, click Review & Confirm",
                details={"upload_id": au.id, "status": au.status, "filename": au.pdf_filename})
        elif not target_audits and is_coop:
            add("medium", "afs_missing", ec,
                f"No FY{target_year} audit upload found for this coop — Col 2 (BY-2 Actual) will be blank",
                fix=f"Run POST /api/admin/sync-afs/{ec} to import from SharePoint")

        # ── Check: summary_rows_missing (MEDIUM) ──
        if not rows:
            add("medium", "summary_rows_missing", ec,
                "No BudgetSummaryRow records — approved 2026 budget hasn't been imported yet",
                fix="Run the approved budget import via the wizard")
            continue  # downstream checks need rows; skip for empty entity

        # ── Check: row_missing_prefix (HIGH for income/expenses; LOW for non-op) ──
        for r in rows:
            if r.row_type != "data":
                continue
            if not r.gl_prefixes_json or r.gl_prefixes_json in ("[]", "null"):
                sev = "high" if r.section in ("Income", "Expenses") else "low"
                add(sev, "row_missing_prefix", ec,
                    f"Summary row '{r.label}' has no GL prefix → can't auto-aggregate",
                    fix=f"Run POST /api/admin/resolve-summary-aliases/{ec}, or add to LABEL_ALIASES if needed",
                    details={"label": r.label, "section": r.section})

        # ── Check: capital_forecast_extrapolated (HIGH) ──
        # FA #18: capital lines should not extrapolate. Find Capital lines where
        # forecast > YTD by more than 1% (i.e., extrapolation happened).
        for l in lines:
            if not (l.sheet_name == "Capital" or (l.category or "").lower() == "capital"):
                continue
            ytd = float(l.ytd_actual or 0) + float(l.accrual_adj or 0) + float(l.unpaid_bills or 0)
            prop = float(l.proposed_budget or 0)
            # If proposed_budget > ytd_total * 2 on a capital line → extrapolation slipped through
            if abs(ytd) >= 100 and prop > abs(ytd) * 2:
                add("high", "capital_forecast_extrapolated", ec,
                    f"Capital line {l.gl_code} has proposed_budget=${prop:,.0f} >> ytd=${ytd:,.0f} — #18 cap may not have applied",
                    fix=f"Re-run PUT /api/budget-assumptions/{ec} with any non-empty body to trigger recompute",
                    details={"gl_code": l.gl_code, "ytd": ytd, "proposed": prop})

        # ── Check: negative_forecast_anomaly (HIGH) ──
        # FA #7 cap should prevent this. Find lines where proposed_budget is negative
        # but prior_year is positive — should be capped to 0 or YTD.
        for l in lines:
            prior = float(l.prior_year or 0)
            prop = float(l.proposed_budget or 0)
            if prop < -1000 and prior >= 0:
                add("high", "negative_forecast_anomaly", ec,
                    f"Line {l.gl_code} ({l.description}) has proposed=${prop:,.0f} despite prior_year=${prior:,.0f}",
                    fix=f"Re-run PUT /api/budget-assumptions/{ec} to trigger anomaly cap",
                    details={"gl_code": l.gl_code, "ytd": float(l.ytd_actual or 0),
                             "prior": prior, "proposed": prop})

        # ── Check: orphan_dollar_total (MEDIUM) ──
        # Sum of GLs with non-zero data not claimed by any summary row.
        matched_gls = set()
        for r in rows:
            if r.row_type != "data" or not r.gl_prefixes_json:
                continue
            try:
                pfx = _json.loads(r.gl_prefixes_json)
            except Exception:
                continue
            for l in lines:
                gl_str = str(l.gl_code or "").strip()
                gl_base = gl_str.split("-")[0]
                for p in pfx:
                    if "-" in str(p):
                        if gl_str.startswith(str(p)):
                            matched_gls.add(l.gl_code); break
                    else:
                        if gl_base.startswith(str(p)):
                            matched_gls.add(l.gl_code); break
        orphan_ytd = 0.0
        orphan_count = 0
        for l in lines:
            if l.gl_code in matched_gls:
                continue
            ytd = float(l.ytd_actual or 0)
            cb = float(l.current_budget or 0)
            if abs(ytd) >= 100 or abs(cb) >= 100:
                orphan_ytd += ytd
                orphan_count += 1
        if orphan_count >= 3 and abs(orphan_ytd) >= 1000:
            add("medium", "summary_orphans", ec,
                f"{orphan_count} GLs with data not claimed by any summary row (YTD total ${orphan_ytd:,.0f})",
                fix=f"Run GET /api/admin/summary-debug/{ec} to see the orphans, then add aliases or extend prefix lists",
                details={"orphan_count": orphan_count, "orphan_ytd_total": orphan_ytd})

        # ── Check: peer_outlier (MEDIUM) ──
        # For each row that has a value, compare to peer median. Flag when delta > 2σ.
        for r in rows:
            if r.row_type != "data" or r.col6_approved_budget is None or r.col6_approved_budget <= 0:
                continue
            stats = peer_stats.get(r.label)
            if not stats or stats["stdev"] == 0:
                continue
            z = (r.col6_approved_budget - stats["median"]) / stats["stdev"]
            if abs(z) >= 2.5:
                pct = (r.col6_approved_budget / stats["median"] - 1) * 100 if stats["median"] else 0
                add("low", "peer_outlier", ec,
                    f"'{r.label}' = ${r.col6_approved_budget:,.0f} vs peer median ${stats['median']:,.0f} (σ={z:+.1f}, {pct:+.0f}%)",
                    fix="Confirm value is real, not a missed import or wrong unit",
                    details={"label": r.label, "value": r.col6_approved_budget,
                             "peer_median": stats["median"], "z_score": round(z, 2),
                             "peer_n": stats["n"]})

        # ── Check: total_imbalance (MEDIUM) ──
        # Coop budgets should net near zero. Flag entities with Total Income vs
        # Total Expenses delta > 10% of income.
        total_inc = next((r.col6_approved_budget or 0 for r in rows
                          if r.row_type == "subtotal" and "total income" in (r.label or "").lower()), None)
        total_exp = next((r.col6_approved_budget or 0 for r in rows
                          if r.row_type == "subtotal" and "total expense" in (r.label or "").lower()
                          and "non" not in (r.label or "").lower()), None)
        if total_inc and total_exp:
            delta = total_inc - total_exp
            pct = abs(delta) / total_inc * 100 if total_inc else 0
            if pct > 10:
                add("medium", "total_imbalance", ec,
                    f"Total Income ${total_inc:,.0f} vs Total Expenses ${total_exp:,.0f} ({pct:.1f}% off — coops should balance)",
                    fix="Investigate which categories are over/under-counted",
                    details={"total_income": total_inc, "total_expenses": total_exp, "delta_pct": round(pct, 1)})

    # ─── 4. Aggregate response ───
    by_severity = {"critical": 0, "high": 0, "medium": 0, "low": 0}
    by_check = {}
    by_entity = {}
    for issue in issues:
        by_severity[issue["severity"]] = by_severity.get(issue["severity"], 0) + 1
        by_check[issue["check"]] = by_check.get(issue["check"], 0) + 1
        by_entity[issue["entity"]] = by_entity.get(issue["entity"], 0) + 1

    # Sort issues: critical > high > medium > low; then by entity
    sev_rank = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    issues.sort(key=lambda i: (sev_rank.get(i["severity"], 9), i["entity"], i["check"]))

    # If include_passing, list entities with zero issues
    passing = []
    if include_passing:
        all_ec = {b.entity_code for b in budgets}
        for ec in sorted(all_ec - set(by_entity.keys())):
            passing.append(ec)

    return jsonify({
        "scanned_at": _dt.utcnow().isoformat() + "Z",
        "budget_year": _BY,
        "entities_scanned": len(budgets),
        "issues_total": len(issues),
        "by_severity": by_severity,
        "by_check": by_check,
        "by_entity_count": dict(sorted(by_entity.items(), key=lambda x: -x[1])[:20]),
        "issues": issues,
        "passing_entities": passing if include_passing else None,
        "filters_applied": {
            "severity": severity_filter,
            "check": check_filter,
            "entity_code": entity_filter,
        },
    })


@app.route("/api/admin/sync-afs/all", methods=["POST"])
@require_admin
def admin_sync_afs_all():
    """ADMIN: Bulk variant of sync-afs — scan SP for every entity that has
    a Budget row this cycle. Per-entity errors are captured but don't
    abort the rest. Body: same as per-entity endpoint.
    """
    Budget = workflow_models["Budget"]
    AuditUpload = af_models["AuditUpload"]
    from workflow import BUDGET_YEAR as _BY
    body = request.get_json(silent=True) or {}
    dry_run = bool(body.get("dry_run"))
    force = bool(body.get("force"))

    entities = [b.entity_code for b in Budget.query.filter_by(year=_BY).all()]
    summary = {"ok": True, "dry_run": dry_run, "entities_scanned": 0, "total_created": 0, "total_skipped": 0, "results": []}
    for ec in entities:
        try:
            r = _do_afs_sync(ec, dry_run=dry_run, force=force, AuditUpload=AuditUpload, return_dict=True)
            summary["entities_scanned"] += 1
            summary["total_created"] += r.get("created", 0)
            summary["total_skipped"] += r.get("skipped", 0)
            summary["results"].append({"entity_code": ec, **r})
        except Exception as e:
            summary["results"].append({"entity_code": ec, "error": str(e)[:200]})
    return jsonify(summary)


def _do_afs_sync(entity_code, dry_run=False, force=False, AuditUpload=None, return_dict=False):
    """Shared implementation for sync-afs single + bulk. Scans SharePoint's
    audit_2025-classified PDFs for the given entity, creates AuditUpload
    rows for any not already in the DB.
    """
    import re as _re
    import base64 as _base64
    from pathlib import Path as _Path

    if AuditUpload is None:
        AuditUpload = af_models["AuditUpload"]

    # 1. Scan SharePoint for AFS PDFs
    try:
        sp = _sharepoint_list_entity_sources(entity_code)
        afs_files = sp.get("by_source_type", {}).get("audit_2025") or []
    except Exception as e:
        result = {"ok": False, "error": f"SP scan failed: {str(e)[:200]}", "scanned": 0, "created": 0, "skipped": 0, "results": []}
        return result if return_dict else jsonify(result)

    # 2. Building name lookup (af_helpers["get_buildings_list"]() shape)
    try:
        bldgs = af_helpers["get_buildings_list"]()
        building_name = next((b["building_name"] for b in bldgs if b["entity_code"] == entity_code), f"Entity {entity_code}")
    except Exception:
        building_name = f"Entity {entity_code}"

    # 3. For each AFS PDF, extract fiscal year from filename and process
    created = 0
    skipped = 0
    results = []
    data_dir = af_helpers["get_data_dir"]()

    for f in afs_files:
        name = f.get("name", "")
        # Year detection: prefer the largest 20xx that appears in the filename.
        years_in_name = [int(m.group(1)) for m in _re.finditer(r"(20\d{2})", name)]
        fiscal_year_end = str(max(years_in_name)) if years_in_name else ""

        # Idempotency check: AuditUpload keyed by entity + fiscal_year_end + filename
        existing = AuditUpload.query.filter_by(
            entity_code=entity_code,
            fiscal_year_end=fiscal_year_end,
            pdf_filename=f"{entity_code}_{fiscal_year_end}_{name}",
        ).first()
        if existing and not force:
            skipped += 1
            results.append({
                "filename": name, "year": fiscal_year_end,
                "status": "skipped_exists", "upload_id": existing.id,
                "current_status": existing.status,
            })
            continue

        if dry_run:
            results.append({"filename": name, "year": fiscal_year_end, "status": "would_create"})
            created += 1
            continue

        # 4. Download PDF from SharePoint via item_id
        try:
            item_id = f.get("item_id")
            if not item_id:
                results.append({"filename": name, "year": fiscal_year_end, "error": "no item_id"})
                continue
            _, pdf_bytes = _sharepoint_download_item(item_id)
        except Exception as e:
            results.append({"filename": name, "year": fiscal_year_end, "error": f"download_failed: {str(e)[:150]}"})
            continue

        # 5. Save bytes to local data_dir (mirrors bulk-upload flow)
        safe_filename = f"{entity_code}_{fiscal_year_end}_{name}"
        try:
            with open(str(data_dir / safe_filename), "wb") as fh:
                fh.write(pdf_bytes)
        except Exception as e:
            results.append({"filename": name, "year": fiscal_year_end, "error": f"local_save_failed: {str(e)[:150]}"})
            continue

        # 6. Create AuditUpload row at "uploaded" status; capture SP web URL
        try:
            upload = AuditUpload(
                entity_code=entity_code,
                building_name=building_name,
                fiscal_year_end=fiscal_year_end,
                pdf_filename=safe_filename,
                sharepoint_web_url=f.get("web_url"),
                status="uploaded",
            )
            db.session.add(upload)
            db.session.commit()
            created += 1
            results.append({
                "filename": name, "year": fiscal_year_end,
                "status": "created", "upload_id": upload.id, "size": f.get("size"),
            })
        except Exception as e:
            db.session.rollback()
            results.append({"filename": name, "year": fiscal_year_end, "error": f"db_failed: {str(e)[:150]}"})

    out = {
        "ok": True,
        "entity_code": entity_code,
        "scanned": len(afs_files),
        "created": created,
        "skipped": skipped,
        "results": results,
    }
    if dry_run:
        out["dry_run"] = True
    return out if return_dict else jsonify(out)


@app.route("/api/admin/set-summary-row-prefixes", methods=["POST"])
@require_admin
def admin_set_summary_row_prefixes():
    """ADMIN: Surgically set gl_prefixes_json for a single (entity, label) row.

    Use case: targeted prefix correction (e.g. removing a wrongly-included GL
    range from an Income row) without running the full resolve-summary-aliases
    flow which would touch every label-matched row on the entity. Lets us test
    a prefix change on one building before rolling out portfolio-wide.

    Body: {
        "entity_code": "168",
        "label": "Tax Benefit Credits (Abatement, Star,etc)",
        "prefixes": ["4105", "4106", ..., "4125"],
        "section": "income"  # optional — adds extra filter to disambiguate
                              # rows when same label appears in multiple sections
    }

    Returns the old & new prefix values per matched row. If multiple rows match
    (rare — usually means duplicate display_order or same label across
    Income+Expense sections), all matched rows are updated; the response shows
    each one so the caller can audit.
    """
    BudgetSummaryRow = workflow_models["BudgetSummaryRow"]
    from workflow import BUDGET_YEAR as _BY

    data = request.get_json() or {}
    ec = (data.get("entity_code") or "").strip()
    label = (data.get("label") or "").strip()
    prefixes = data.get("prefixes")
    section_filter = (data.get("section") or "").strip() or None

    if not ec or not label:
        return jsonify({"error": "entity_code and label required"}), 400
    if not isinstance(prefixes, list) or not all(isinstance(p, str) for p in prefixes):
        return jsonify({"error": "prefixes must be a list of strings"}), 400

    q = BudgetSummaryRow.query.filter_by(
        entity_code=ec, budget_year=_BY, label=label
    )
    if section_filter:
        q = q.filter_by(section=section_filter)
    rows = q.all()
    if not rows:
        return jsonify({
            "error": f"No row found for entity={ec} label={label!r} section={section_filter!r}"
        }), 404

    new_pj = json.dumps(prefixes)
    diffs = []
    for row in rows:
        try:
            old = json.loads(row.gl_prefixes_json) if row.gl_prefixes_json else []
        except Exception:
            old = []
        diffs.append({
            "id": row.id,
            "display_order": row.display_order,
            "section": row.section,
            "old_prefixes": old,
            "new_prefixes": prefixes,
        })
        row.gl_prefixes_json = new_pj
    db.session.commit()
    return jsonify({
        "ok": True,
        "rows_updated": len(rows),
        "diffs": diffs,
    })


@app.route("/api/admin/delete-summary-row", methods=["POST"])
@require_admin
def admin_delete_summary_row():
    """ADMIN: Remove a BudgetSummaryRow by entity + label.

    Sister to add-summary-row. Used to undo a mis-added row from the FA's
    "+ Add Row" flow on the dashboard. Only deletes rows the FA created
    interactively (row_type=data, no col6_approved_budget — i.e., not from
    the imported template). Refuses to delete rows that came from the
    approved 2026 budget yrlycomp import.

    Body: {
        "entity_code": "168",
        "label": "Commercial",
        "merge_into_label": "Other Income",  // optional: sum col1/col6/col7
                                              // into target row before delete
                                              // (used to consolidate duplicate
                                              // rows like Gas / Gas Heating
                                              // without losing imported budget
                                              // values)
    }
    """
    BudgetSummaryRow = workflow_models["BudgetSummaryRow"]
    from workflow import BUDGET_YEAR as _BY
    from datetime import datetime as _dt

    data = request.get_json(silent=True) or {}
    entity_code = (data.get("entity_code") or "").strip()
    label = (data.get("label") or "").strip()
    merge_into_label = (data.get("merge_into_label") or "").strip()
    if not entity_code or not label:
        return jsonify({"error": "entity_code and label required"}), 400

    row = BudgetSummaryRow.query.filter_by(
        entity_code=entity_code, budget_year=_BY, label=label
    ).first()
    if not row:
        return jsonify({"ok": True, "noop": "row not found"})

    if row.row_type != "data":
        return jsonify({"error": "Cannot delete subtotal/section_header rows"}), 403

    # Two paths depending on whether the caller wants to merge first:
    #   - merge_into_label set → sum col1/col6/col7 into the target row, then delete
    #     (target must exist; preserves budget data when consolidating duplicates)
    #   - merge_into_label unset → refuse to delete if col6 is set, to prevent
    #     accidentally losing imported budget values (legacy safety behavior)
    merge_summary = None
    if merge_into_label:
        target = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=_BY, label=merge_into_label
        ).first()
        if not target:
            return jsonify({
                "error": f"merge_into_label '{merge_into_label}' not found for entity {entity_code}"
            }), 404
        # Sum the imported/FA-entered numeric columns into the target.
        # NULL handling: treat NULL as 0 for the addition; keep target NULL
        # if both are NULL (don't introduce a 0 where there was no data).
        def _add(a, b):
            if a is None and b is None: return None
            return (a or 0) + (b or 0)
        old_target = {
            "col1": target.col1_prior_actual,
            "col6": target.col6_approved_budget,
            "col7": target.col7_proposed_budget,
        }
        target.col1_prior_actual = _add(target.col1_prior_actual, row.col1_prior_actual)
        target.col6_approved_budget = _add(target.col6_approved_budget, row.col6_approved_budget)
        target.col7_proposed_budget = _add(target.col7_proposed_budget, row.col7_proposed_budget)
        target.updated_at = _dt.utcnow()
        merge_summary = {
            "target_label": merge_into_label,
            "target_id": target.id,
            "before": old_target,
            "after": {
                "col1": target.col1_prior_actual,
                "col6": target.col6_approved_budget,
                "col7": target.col7_proposed_budget,
            },
            "added_from_source": {
                "col1": row.col1_prior_actual,
                "col6": row.col6_approved_budget,
                "col7": row.col7_proposed_budget,
            },
        }
    else:
        # Legacy safety: refuse to delete imported rows without explicit merge.
        if row.col6_approved_budget is not None:
            return jsonify({
                "error": "Cannot delete a row imported from the approved 2026 budget. "
                         f"This row has historical data (col6 = ${row.col6_approved_budget:,.0f}). "
                         "Pass merge_into_label to consolidate values into another row, "
                         "or edit the row's values instead.",
            }), 403

    try:
        deleted_id = row.id
        db.session.delete(row)
        db.session.commit()
        resp = {"ok": True, "deleted": label, "id": deleted_id}
        if merge_summary:
            resp["merged_into"] = merge_summary
        return jsonify(resp)
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)[:300]}), 500


@app.route("/api/admin/summary-row-options", methods=["GET"])
@require_admin
def admin_summary_row_options():
    """Return canonical SUMMARY_ROW_MAP labels grouped by section, used to
    populate the Add Row dropdown on the FA dashboard. The FA picks from a
    canonical list so the new row inherits the right gl_prefix automatically.

    Returns:
      {
        "income": ["Maintenance", "Commercial Rent", "Storage Income", ...],
        "expenses": ["Payroll", "Insurance", "Repairs & Maintenance", ...],
        "non_operating_income": [...],
        "non_operating_expense": [...]
      }
    """
    try:
        from GL_TO_SUMMARY_MAP import SUMMARY_ROW_MAP, _CONDO_ROWS
    except ImportError:
        from budget_summary.GL_TO_SUMMARY_MAP import SUMMARY_ROW_MAP, _CONDO_ROWS

    grouped = {
        "income": [],
        "expenses": [],
        "non_operating_income": [],
        "non_operating_expense": [],
    }
    # Combine canonical map + condo-specific rows
    all_rows = {**SUMMARY_ROW_MAP, **_CONDO_ROWS}
    for label, cfg in all_rows.items():
        section = (cfg or {}).get("section") or "expenses"
        if section in grouped:
            grouped[section].append(label)
    # Sort each list alphabetically for FA scan-ability
    for k in grouped:
        grouped[k].sort()
    return jsonify(grouped)


@app.route("/api/admin/add-summary-row", methods=["POST"])
@require_admin
def admin_add_summary_row():
    """ADMIN: Insert a single new BudgetSummaryRow for one entity.

    Used when an FA wants to break out an audit line (via the Inspector's
    Move action) into its own summary row that wasn't in the original
    approved budget template. Without this, summary_overrides[<label>]
    becomes orphaned because no row matches it.

    Body: {
        "entity_code": "148",
        "label": "Supplies",
        "section": "Expenses"|"Income"|"Non-Operating Income"|"Non-Operating Expenses"|null,
        "after_label": "Repairs & Maintenance",  # optional — placement hint
        "gl_prefixes": ["4800"],  # optional — pre-attach GL prefix list to the new
                                  # row (used by the "Specific GL" Add Row mode so
                                  # the row pulls col3/4/5 data immediately without
                                  # needing a separate alias-resolve pass)
    }
    Returns the new row's display_order + label.
    """
    import json as _json
    BudgetSummaryRow = workflow_models["BudgetSummaryRow"]
    from workflow import BUDGET_YEAR as _BY

    data = request.get_json(silent=True) or {}
    entity_code = (data.get("entity_code") or "").strip()
    label = (data.get("label") or "").strip()
    section = (data.get("section") or "Expenses").strip() or None
    after_label = (data.get("after_label") or "").strip() or None
    raw_prefixes = data.get("gl_prefixes")
    gl_prefixes_json = None
    if isinstance(raw_prefixes, list):
        cleaned = [str(p).strip() for p in raw_prefixes if str(p or "").strip()]
        if cleaned:
            gl_prefixes_json = _json.dumps(cleaned)

    if not entity_code or not label:
        return jsonify({"error": "entity_code and label required"}), 400

    # Idempotent: if a row with this label already exists, return it.
    existing = BudgetSummaryRow.query.filter_by(
        entity_code=entity_code, budget_year=_BY, label=label
    ).first()
    if existing:
        return jsonify({
            "ok": True, "noop": "row already exists",
            "id": existing.id, "label": existing.label,
            "display_order": existing.display_order,
        })

    # Pick a display_order: just after `after_label` if provided + present,
    # else at the end of the section's existing range, else at the end.
    target_order = None
    if after_label:
        ref = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=_BY, label=after_label
        ).first()
        if ref and ref.display_order is not None:
            target_order = ref.display_order + 1
            # Shift everything at/after target_order by +1. The unique
            # constraint on (entity, year, display_order) means a single
            # UPDATE would violate uniqueness mid-statement, so use the
            # temporary-offset trick: push rows out of range first, then
            # pull them back to the desired positions.
            db.session.execute(db.text(
                "UPDATE budget_summary_rows SET display_order = display_order + 10000 "
                "WHERE entity_code = :ec AND budget_year = :by AND display_order >= :ord"
            ), {"ec": entity_code, "by": _BY, "ord": target_order})
            db.session.flush()
            db.session.execute(db.text(
                "UPDATE budget_summary_rows SET display_order = display_order - 9999 "
                "WHERE entity_code = :ec AND budget_year = :by AND display_order >= 10000"
            ), {"ec": entity_code, "by": _BY})
            db.session.flush()

    if target_order is None:
        # Fall back: max + 1 of all rows for this entity.
        max_row = db.session.execute(db.text(
            "SELECT COALESCE(MAX(display_order), 0) FROM budget_summary_rows "
            "WHERE entity_code = :ec AND budget_year = :by"
        ), {"ec": entity_code, "by": _BY}).scalar()
        target_order = int(max_row or 0) + 1

    try:
        row = BudgetSummaryRow(
            entity_code=entity_code,
            budget_year=_BY,
            display_order=target_order,
            label=label,
            section=section,
            row_type="data",
            gl_prefixes_json=gl_prefixes_json,
        )
        db.session.add(row)
        db.session.commit()
        return jsonify({
            "ok": True, "id": row.id, "label": row.label,
            "display_order": row.display_order, "section": row.section,
            "gl_prefixes": _json.loads(gl_prefixes_json) if gl_prefixes_json else [],
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)[:300]}), 500


@app.route("/api/admin/backfill-period/<entity_code>", methods=["POST"])
@require_admin
def admin_backfill_period(entity_code):
    """ADMIN: Set budget_period on an existing budget. Used to backfill
    entities whose budgets were generated before the period selector
    shipped (default ytd_months=2 was applied silently, producing wrong
    forecast formulas).

    Detection order:
    1. If body provides `{"month": N}` (1-12), use that explicitly.
    2. Else if a YSL file exists for this entity on SharePoint, download
       and parse to auto-detect the period from the header.
    3. Else return 400 with an error explaining what to do.

    On success, writes budget_period to assumptions_json AND triggers a
    full forecast/proposed_budget recompute for all lines (mirrors PUT
    /api/budget-assumptions behavior).

    Body:
      {"month": 4}            - explicit month override
      {"force": true}         - overwrite an existing period
      (no body)               - auto-detect from YSL

    Returns:
      {"ok": true, "entity_code": "...", "month": N, "source": "ysl"|"manual",
       "filename": "...", "lines_recomputed": N}
    """
    Budget = workflow_models["Budget"]
    BudgetLine = workflow_models["BudgetLine"]
    from workflow import BUDGET_YEAR as _BY

    body = request.get_json(silent=True) or {}
    explicit_month = body.get("month")
    force = bool(body.get("force"))

    budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {entity_code}"}), 404

    # Check if period is already set; bail unless force=true
    try:
        existing = json.loads(budget.assumptions_json or "{}")
    except Exception:
        existing = {}
    current_bp = existing.get("budget_period", "")
    if current_bp and not force and not explicit_month:
        return jsonify({
            "ok": False,
            "error": f"Period already set to {current_bp}. Pass {{\"force\": true}} to overwrite.",
            "current_period": current_bp,
        }), 400

    detected_month = None
    source = None
    filename = None

    # Branch 1: explicit override
    if explicit_month is not None:
        try:
            m = int(explicit_month)
        except Exception:
            return jsonify({"error": "month must be an integer 1-12"}), 400
        if not (1 <= m <= 12):
            return jsonify({"error": "month must be 1-12"}), 400
        detected_month = m
        source = "manual"
    else:
        # Branch 2: auto-detect from SharePoint YSL
        try:
            sp = _sharepoint_list_entity_sources(entity_code)
            ysl_files = sp.get("by_source_type", {}).get("ysl") or []
        except Exception as e:
            return jsonify({"error": f"SharePoint scan failed: {str(e)[:200]}"}), 500
        if not ysl_files:
            return jsonify({
                "ok": False,
                "error": f"No YSL file in SharePoint for entity {entity_code}. Provide explicit month: POST body {{\"month\": N}}",
            }), 400
        # Use the most recently modified YSL
        ysl_files_sorted = sorted(ysl_files, key=lambda x: x.get("last_modified", ""), reverse=True)
        ysl_sel = ysl_files_sorted[0]
        try:
            from budget_system.ysl_parser import parse_ysl_file
            tmp_path, filename = _wizard_download_sp_to_tmp(ysl_sel, ".xlsx")
            try:
                _gl_data, prop_info = parse_ysl_file(tmp_path)
                detected_month = prop_info.get("budget_period_month")
            finally:
                try:
                    import os as _os
                    _os.unlink(tmp_path)
                except Exception:
                    pass
        except Exception as e:
            return jsonify({"error": f"YSL parse failed: {str(e)[:200]}"}), 500
        if not detected_month or not (1 <= detected_month <= 12):
            return jsonify({
                "ok": False,
                "error": f"Could not auto-detect period from YSL header. Provide explicit month: POST body {{\"month\": N}}",
                "filename": filename,
            }), 400
        source = "ysl"

    # Apply: write to assumptions_json and recompute proposed_budget for all lines
    mm = str(int(detected_month)).zfill(2)
    new_period = f"{mm}/{_BY - 1}"
    existing["budget_period"] = new_period
    budget.assumptions_json = json.dumps(existing)

    # Recompute proposed_budget for all lines (matches workflow.py PUT logic)
    _ytd_months = detected_month
    _remaining = 12 - _ytd_months
    # Codex review (2026-05-03) caught two bugs in this block:
    #  1. ONE_TIME_FEE_GLS was used here without being imported into app.py
    #     → NameError on the first real call. Now imported from workflow.
    #  2. Capital lines were missing the FA #18 guard, so backfill-period
    #     would annualize and overwrite their proposed_budget — the very
    #     behavior tonight's #18 cap was added to prevent. Mirrors the
    #     guards in workflow.py PUT /api/budget-assumptions.
    from workflow import ONE_TIME_FEE_GLS as _OTFG
    lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
    recomputed = 0
    for line in lines:
        ytd = float(line.ytd_actual or 0)
        accrual = float(line.accrual_adj or 0)
        unpaid = float(line.unpaid_bills or 0)
        prior = float(line.prior_year or 0)
        base = ytd + accrual + unpaid
        _is_cap = (line.sheet_name == "Capital"
                   or (line.category or "").lower() == "capital")
        # FA one-time fee rule
        if (line.gl_code or "") in _OTFG and abs(base) > 0.01:
            estimate = 0
        # FA #18: Capital — never extrapolate, never auto-fill proposed
        elif _is_cap:
            estimate = 0
        # FA #7 anomaly cap
        elif base < 0 and prior >= 0:
            estimate = 0
        else:
            estimate = (base / _ytd_months) * _remaining if _ytd_months > 0 else 0
        forecast = base + estimate
        # FA #18: don't auto-fill proposed for Capital — leave whatever
        # the FA explicitly entered.
        if not _is_cap:
            line.proposed_budget = forecast * (1 + float(line.increase_pct or 0))
        recomputed += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.exception("backfill-period commit failed")
        return jsonify({"error": f"Commit failed: {str(e)[:200]}"}), 500

    logger.info(
        f"Period backfilled for {entity_code}: month={detected_month} source={source} "
        f"filename={filename} recomputed={recomputed}"
    )
    return jsonify({
        "ok": True,
        "entity_code": entity_code,
        "month": detected_month,
        "period": new_period,
        "source": source,
        "filename": filename,
        "lines_recomputed": recomputed,
    })


@app.route("/api/admin/backfill-period/all", methods=["POST"])
@require_admin
def admin_backfill_period_all():
    """ADMIN: Bulk backfill — try to auto-detect period for every entity
    whose budget_period is unset. Skips entities already set unless
    force=true. Returns per-entity status array.

    Body (optional):
      {"force": true}      - overwrite existing periods too
      {"only_missing": true} - default; skip entities with period set
    """
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY

    body = request.get_json(silent=True) or {}
    force = bool(body.get("force"))

    budgets = Budget.query.filter_by(year=_BY).all()
    results = []
    for b in budgets:
        try:
            existing = json.loads(b.assumptions_json or "{}")
        except Exception:
            existing = {}
        current_bp = existing.get("budget_period", "")
        if current_bp and not force:
            results.append({"entity_code": b.entity_code, "skipped": True, "reason": "period_already_set", "period": current_bp})
            continue
        # Inline call to the per-entity backfill logic. Use the test-client
        # to reuse the route handler and keep behavior consistent.
        try:
            with app.test_client() as tc:
                resp = tc.post(f"/api/admin/backfill-period/{b.entity_code}",
                               json={"force": force})
                data = resp.get_json() or {}
                data["entity_code"] = b.entity_code
                data["status_code"] = resp.status_code
                results.append(data)
        except Exception as e:
            results.append({"entity_code": b.entity_code, "error": str(e)[:200]})
    return jsonify({"ok": True, "count": len(results), "results": results})


@app.route("/api/admin/refresh-building-info/<entity_code>", methods=["POST"])
@require_admin
def admin_refresh_building_info(entity_code):
    """ADMIN: Re-pull the maintenance/common-charges history from the entity's
    2026 approved budget XLSX into BuildingInfo. Used to backfill entities that
    were imported before the auto-populate hook existed.

    Looks up the latest 2026-approved-budget item_id from the wizard selections
    record, downloads via SharePoint, parses the Income tab, persists.

    Body (optional): {"force": true} to overwrite even FA-edited histories.
    """
    Budget = workflow_models["Budget"]
    from workflow import BUDGET_YEAR as _BY

    data = request.get_json(silent=True) or {}
    force = bool(data.get("force"))

    budget = Budget.query.filter_by(entity_code=entity_code, year=_BY).first()
    if not budget:
        return jsonify({"error": f"No Budget row for entity {entity_code}"}), 404

    try:
        sels = json.loads(budget.wizard_selections_json or "{}")
    except Exception:
        sels = {}
    sel = sels.get("approved_2026") or {}
    item_id = (sel.get("item_id") or "").strip()
    if not item_id:
        return jsonify({"error": "No 2026 approved budget item_id in wizard_selections_json"}), 400

    try:
        import tempfile
        from pathlib import Path as _Path
        filename, file_bytes = _sharepoint_download_item(item_id)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        try:
            if force:
                # Wipe existing first so the populate helper rewrites it.
                BuildingInfo = workflow_models["BuildingInfo"]
                info = BuildingInfo.query.filter_by(entity_code=entity_code).first()
                if info:
                    info.maintenance_history_json = None
                    info.common_charges_history_json = None
                    db.session.commit()
            res = _populate_building_info_from_income(entity_code, tmp_path)
        finally:
            try:
                import os as _os
                _os.unlink(tmp_path)
            except Exception:
                pass
        if not res:
            return jsonify({"ok": False, "error": "parse_income_history found no data"}), 500
        return jsonify({"ok": True, "result": res, "filename": filename})
    except Exception as e:
        db.session.rollback()
        logger.exception("refresh-building-info failed")
        return jsonify({"ok": False, "error": str(e)[:300]}), 500


@app.route("/api/admin/wipe-entity-data", methods=["POST"])
@require_admin
def admin_wipe_entity_data():
    """ADMIN: Wipe budget_lines + budget_revisions + budget_summary_rows for the
    given entities (year=BUDGET_YEAR), and reset their Budget row state to Setup.

    Body: {"entity_codes": ["123","204","212","733"], "confirm": "WIPE"}

    Does NOT touch SharePoint files or Monday.com state. Idempotent — re-running
    on a clean entity is a no-op.
    """
    Budget = workflow_models["Budget"]
    BudgetLine = workflow_models["BudgetLine"]
    BudgetRevision = workflow_models["BudgetRevision"]
    BudgetSummaryRow = workflow_models["BudgetSummaryRow"]

    data = request.get_json() or {}
    entity_codes = data.get("entity_codes") or []
    confirm = data.get("confirm")

    if not isinstance(entity_codes, list) or not entity_codes:
        return jsonify({"error": "entity_codes (non-empty list) required"}), 400
    if confirm != "WIPE":
        return jsonify({"error": "confirm must equal \"WIPE\""}), 400

    from workflow import BUDGET_YEAR as _BY
    summary = {"entities": [], "lines_deleted": 0, "revisions_deleted": 0,
               "summary_rows_deleted": 0, "budgets_reset": 0}

    try:
        for ec in entity_codes:
            ec = str(ec).strip()
            if not ec:
                continue
            budget = Budget.query.filter_by(entity_code=ec, year=_BY).first()
            entity_record = {"entity_code": ec, "found": bool(budget)}
            if not budget:
                summary["entities"].append(entity_record)
                continue

            # Delete budget_revisions FIRST — they reference budget_lines via FK,
            # so removing lines before revisions would violate the constraint.
            rev_count = BudgetRevision.query.filter_by(budget_id=budget.id).count()
            BudgetRevision.query.filter_by(budget_id=budget.id).delete()
            db.session.flush()

            # Now safe to delete budget_lines
            line_count = BudgetLine.query.filter_by(budget_id=budget.id).count()
            BudgetLine.query.filter_by(budget_id=budget.id).delete()
            db.session.flush()

            # Count + delete budget_summary_rows (these can be re-imported from SharePoint)
            sum_count = BudgetSummaryRow.query.filter_by(entity_code=ec, budget_year=_BY).count()
            BudgetSummaryRow.query.filter_by(entity_code=ec, budget_year=_BY).delete()

            # Also clear expense_reports + audit_uploads + open_ap + maint_proof —
            # these power the FA Completion Checklist + Data Status indicators.
            # Child tables MUST be deleted first (FK constraints).
            expense_count = 0
            audit_count = 0
            open_ap_count = 0
            maint_count = 0
            data_sources_count = 0
            try:
                # Expense Distribution: invoices → reports
                db.session.execute(db.text(
                    "DELETE FROM expense_invoices WHERE report_id IN "
                    "(SELECT id FROM expense_reports WHERE entity_code = :ec)"
                ), {"ec": ec})
                row = db.session.execute(db.text(
                    "DELETE FROM expense_reports WHERE entity_code = :ec RETURNING id"
                ), {"ec": ec}).fetchall()
                expense_count = len(row)
            except Exception as _e:
                logger.warning(f"expense wipe failed for {ec}: {_e}")
                db.session.rollback()
            try:
                # Open AP: invoices → reports
                db.session.execute(db.text(
                    "DELETE FROM open_ap_invoices WHERE report_id IN "
                    "(SELECT id FROM open_ap_reports WHERE entity_code = :ec)"
                ), {"ec": ec})
                row = db.session.execute(db.text(
                    "DELETE FROM open_ap_reports WHERE entity_code = :ec RETURNING id"
                ), {"ec": ec}).fetchall()
                open_ap_count = len(row)
            except Exception as _e:
                logger.warning(f"open_ap wipe failed for {ec}: {_e}")
                db.session.rollback()
            try:
                # Maintenance Proof: units → reports
                db.session.execute(db.text(
                    "DELETE FROM maint_proof_units WHERE report_id IN "
                    "(SELECT id FROM maint_proof_reports WHERE entity_code = :ec)"
                ), {"ec": ec})
                row = db.session.execute(db.text(
                    "DELETE FROM maint_proof_reports WHERE entity_code = :ec RETURNING id"
                ), {"ec": ec}).fetchall()
                maint_count = len(row)
            except Exception as _e:
                logger.warning(f"maint_proof wipe failed for {ec}: {_e}")
                db.session.rollback()
            try:
                row = db.session.execute(db.text(
                    "DELETE FROM audit_uploads WHERE entity_code = :ec RETURNING id"
                ), {"ec": ec}).fetchall()
                audit_count = len(row)
            except Exception as _e:
                logger.warning(f"audit_uploads wipe failed for {ec}: {_e}")
                db.session.rollback()
            try:
                # data_sources references budgets.id
                row = db.session.execute(db.text(
                    "DELETE FROM data_sources WHERE budget_id = :bid RETURNING id"
                ), {"bid": budget.id}).fetchall()
                data_sources_count = len(row)
            except Exception as _e:
                logger.warning(f"data_sources wipe failed for {ec}: {_e}")
                db.session.rollback()

            # Reset Budget row state
            budget.status = "not_started"
            budget.wizard_step = 0
            budget.wizard_completed_at = None
            budget.assumptions_json = "{}"
            budget.assumptions_history_json = None
            budget.approved_by = None
            budget.approved_at = None
            budget.fa_notes = ""
            budget.ar_notes = ""
            budget.increase_pct = None
            budget.effective_date = None
            budget.initiated_by = None
            budget.initiated_at = None
            budget.return_to_status = None
            budget.pre_merge_snapshot_at = None if hasattr(budget, "pre_merge_snapshot_at") else None

            entity_record.update({
                "lines_deleted": line_count,
                "revisions_deleted": rev_count,
                "summary_rows_deleted": sum_count,
                "expense_reports_deleted": expense_count,
                "open_ap_reports_deleted": open_ap_count,
                "maint_proof_reports_deleted": maint_count,
                "audit_uploads_deleted": audit_count,
                "data_sources_deleted": data_sources_count,
                "budget_id": budget.id,
                "reset_to": "Setup",
            })
            summary["entities"].append(entity_record)
            summary["lines_deleted"] += line_count
            summary["revisions_deleted"] += rev_count
            summary["summary_rows_deleted"] += sum_count
            summary["budgets_reset"] += 1

        db.session.commit()
        return jsonify({"ok": True, **summary})
    except Exception as e:
        db.session.rollback()
        logger.exception("wipe-entity-data failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/upload-audit", methods=["GET"])
@require_admin
def admin_upload_audit():
    """Read-only audit summary for the 4/7-ish window: counts revisions and
    budget lines created by date so we can see when budgets were last touched.

    Query params:
      since=YYYY-MM-DD (default 2026-04-01)
      until=YYYY-MM-DD (default 2026-04-15)
    """
    from datetime import datetime as _dt
    since_str = request.args.get("since", "2026-04-01")
    until_str = request.args.get("until", "2026-04-15")
    try:
        since = _dt.strptime(since_str, "%Y-%m-%d")
        until = _dt.strptime(until_str, "%Y-%m-%d")
    except ValueError:
        return jsonify({"error": "since/until must be YYYY-MM-DD"}), 400

    BudgetLine = workflow_models["BudgetLine"]
    BudgetRevision = workflow_models["BudgetRevision"]
    Budget = workflow_models["Budget"]

    # BudgetLine has no created_at, only updated_at — using that as a proxy.
    # Initial inserts and later edits both update this; for the 4/7 audit window,
    # any line touched in that window indicates upload-or-edit activity.
    line_rows = db.session.execute(db.text("""
        SELECT b.entity_code,
               b.year,
               COUNT(bl.id) as line_count,
               MIN(bl.updated_at) as first_touched,
               MAX(bl.updated_at) as last_touched
          FROM budget_lines bl
          JOIN budgets b ON b.id = bl.budget_id
         WHERE bl.updated_at >= :since AND bl.updated_at < :until
         GROUP BY b.entity_code, b.year
         ORDER BY MIN(bl.updated_at) ASC
    """), {"since": since, "until": until}).fetchall()

    # Revisions in window
    rev_rows = db.session.execute(db.text("""
        SELECT br.action,
               br.source,
               COUNT(*) as cnt,
               MIN(br.created_at) as first_at,
               MAX(br.created_at) as last_at
          FROM budget_revisions br
         WHERE br.created_at >= :since AND br.created_at < :until
         GROUP BY br.action, br.source
         ORDER BY cnt DESC
    """), {"since": since, "until": until}).fetchall()

    return jsonify({
        "window": {"since": since_str, "until": until_str},
        "lines_touched_per_entity": [
            {
                "entity_code": r[0], "year": r[1], "line_count": r[2],
                "first_touched": r[3].isoformat() if r[3] else None,
                "last_touched": r[4].isoformat() if r[4] else None,
            }
            for r in line_rows
        ],
        "revisions_grouped": [
            {
                "action": r[0], "source": r[1], "count": r[2],
                "first_at": r[3].isoformat() if r[3] else None,
                "last_at": r[4].isoformat() if r[4] else None,
            }
            for r in rev_rows
        ],
        "total_lines_in_window": sum(r[2] for r in line_rows) if line_rows else 0,
        "total_revisions_in_window": sum(r[2] for r in rev_rows) if rev_rows else 0,
    })


@app.route("/api/sharepoint/_token-info", methods=["GET"])
def sharepoint_token_info():
    """DEBUG: decode the Graph access token and show its key claims (no secrets)."""
    try:
        import base64
        token = _get_graph_token()
        # JWT: header.payload.signature — decode payload (middle segment)
        parts = token.split(".")
        if len(parts) < 2:
            return jsonify({"error": "token is not a JWT"}), 500
        pad = "=" * (-len(parts[1]) % 4)
        payload = json.loads(base64.urlsafe_b64decode(parts[1] + pad).decode("utf-8"))
        return jsonify({
            "aud": payload.get("aud"),
            "iss": payload.get("iss"),
            "appid": payload.get("appid"),
            "tid": payload.get("tid"),
            "roles": payload.get("roles", []),
            "scp": payload.get("scp"),
            "idtyp": payload.get("idtyp"),
            "exp": payload.get("exp"),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/sharepoint/explore", methods=["GET"])
def sharepoint_explore():
    """Walk the 2027 Budget folder and return one level of structure.

    Query params:
      depth=1 (default) — top-level only
      depth=2           — also list each entity-folder's contents
      sub=<sub>         — explore a sub-path inside 2027 Budget instead of root
    """
    try:
        drive_id = _graph_get_drive_id()
        sub = (request.args.get("sub") or "").strip("/")
        base = SHAREPOINT_2027_FOLDER_PATH + ("/" + sub if sub else "")
        depth = int(request.args.get("depth", "1"))
        # URL-encode the path segment (Graph wants the full path after :/)
        import urllib.parse
        encoded = urllib.parse.quote(base, safe="/")
        listing = _graph_get(f"drives/{drive_id}/root:/{encoded}:/children")
        items = listing.get("value", [])
        out = {"path": base, "drive_id": drive_id, "count": len(items), "items": []}
        for it in items:
            entry = {
                "name": it.get("name"),
                "is_folder": "folder" in it,
                "size": it.get("size"),
                "last_modified": it.get("lastModifiedDateTime"),
                "web_url": it.get("webUrl"),
            }
            if depth >= 2 and "folder" in it:
                try:
                    sub_path = base + "/" + it["name"]
                    sub_encoded = urllib.parse.quote(sub_path, safe="/")
                    sub_listing = _graph_get(f"drives/{drive_id}/root:/{sub_encoded}:/children")
                    entry["children"] = [
                        {"name": c.get("name"), "is_folder": "folder" in c, "size": c.get("size")}
                        for c in sub_listing.get("value", [])
                    ]
                except Exception as e:
                    entry["children_error"] = str(e)
            out["items"].append(entry)
        return jsonify(out)
    except Exception as e:
        logger.error(f"sharepoint_explore failed: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    if IS_CLOUD:
        app.run(host="0.0.0.0", port=port)
    else:
        print("\n" + "=" * 50)
        print("  Century Budget & Assumptions System")
        print("  Open http://localhost:5000 in your browser")
        print("=" * 50 + "\n")
        app.run(debug=True, port=port)
