"""
Budget Summary Parser - Extracts yrlycomp tab data from approved budget Excel files.

Pulls every row from the yrlycomp tab exactly as structured in the Excel,
preserving sections, labels, footnote markers, and all year columns.
"""

import openpyxl
import os
import re
import json
from datetime import datetime


def find_yrlycomp_tab(workbook):
    """Find the yrlycomp tab (case-insensitive)."""
    for name in workbook.sheetnames:
        if 'yrlycomp' in name.lower().strip():
            return workbook[name]
    return None


def extract_entity_code(filename):
    """Extract entity code from filename (first numeric segment)."""
    match = re.match(r'(\d+)', os.path.basename(filename))
    return match.group(1) if match else None


def extract_building_name(sheet, filename=None):
    """Extract building name from header rows (typically row 1-3)."""
    for row_num in range(1, 6):
        for col_num in range(1, 20):
            val = sheet.cell(row=row_num, column=col_num).value
            if val and isinstance(val, str) and len(val) > 10 and 'yearly' not in val.lower() and 'year ending' not in val.lower() and 'update' not in val.lower():
                return val.strip()
    # Fallback: extract from filename
    if filename:
        name = os.path.basename(filename)
        # Remove entity code prefix and year/budget suffix
        match = re.match(r'\d+\s*-?\s*(.+?)\s*\d{4}', name)
        if match:
            return match.group(1).strip().rstrip('-').strip()
    return None


def parse_column_headers(sheet):
    """
    Parse rows 7-8 to identify year columns and their types.
    Row 7 = years (2021, 2022, etc.)
    Row 8 = sub-headers (Actual *, 8 Mo. Actual, 4 Mo. Est**, etc.)

    Returns dict: {col_num: {"year": int, "type": str, "header": str}}
    """
    columns = {}

    def _as_year(val):
        """Return year as int if val parses as a 2010-2030 year, else None.

        Handles:
          - ints/floats: 2024, 2024.0
          - plain strings: '2024', '2024.0', ' 2024 '
          - fiscal-year strings: '2024-25' → 2025, '2024-2025' → 2025
            (returns the LATTER year so condos with Jul-Jun fiscal years
            map to the calendar year the FY closes in; matches '06/30/2026'
            header convention in those files)

        FA dir 2026-05-20: 340 (Pierhouse Condo) revealed fiscal-year format
        in row 7 ('2016-17', '2017-18', ..., '2024-25'). 6 buildings hit
        'Could not parse column headers' for this reason."""
        if val is None:
            return None
        # int / float
        if isinstance(val, (int, float)):
            v = int(val)
            return v if 2010 <= v <= 2030 else None
        s = str(val).strip()
        if not s:
            return None
        # Plain numeric string
        try:
            v = int(float(s))
            if 2010 <= v <= 2030:
                return v
        except (TypeError, ValueError):
            pass
        # Fiscal-year string 'YYYY-YY', 'YYYY-YYYY', 'YYYY/YY', 'YYYY/YYYY'.
        # FA dir 2026-05-22: entity 358 (35 Prospect Park West) uses slash
        # separator ('2021/2022', '2024/2025', etc.) — original regex only
        # accepted hyphens, so all 8 year headers failed → "Could not parse
        # column headers." Widen to accept either separator.
        import re as _re_local
        m = _re_local.match(r'^(\d{4})\s*[-/]\s*(\d{2,4})$', s)
        if m:
            first = int(m.group(1))
            second_raw = m.group(2)
            if len(second_raw) == 2:
                # '2024-25' → 2025; carry the century from first year
                second = (first // 100) * 100 + int(second_raw)
                # Sanity: handle century rollover e.g. '1999-00' → 2000
                if second < first:
                    second += 100
            else:
                second = int(second_raw)
            if 2010 <= second <= 2030:
                return second
        return None

    # Find header row - look for a row with multiple year values.
    # FA dir 2026-05-20: widened range from (5-11) to (3-15) and added
    # stringy-year tolerance. Different building templates put the year
    # header at different rows; the narrow range failed 6 buildings.
    # Still requires >= 3 year columns so a stray year mention in a note
    # row can't accidentally match.
    header_row = None
    for r in range(3, 16):
        year_count = 0
        for c in range(1, 30):
            if _as_year(sheet.cell(row=r, column=c).value) is not None:
                year_count += 1
        if year_count >= 3:
            header_row = r
            break

    if header_row is None:
        return columns, None, None

    sub_header_row = header_row + 1

    for col_num in range(1, 30):
        year_val = sheet.cell(row=header_row, column=col_num).value
        sub_val = sheet.cell(row=sub_header_row, column=col_num).value

        year = _as_year(year_val)
        if year is not None:
            sub = str(sub_val).strip() if sub_val else ""

            # Classify the column type
            sub_lower = sub.lower()
            if 'actual *' == sub_lower or sub_lower == 'actual*':
                col_type = "audited_actual"
            elif 'mo. actual' in sub_lower or 'mo actual' in sub_lower:
                col_type = "partial_actual"
            elif 'est' in sub_lower:
                col_type = "estimate"
            elif 'forecast' in sub_lower:
                col_type = "forecast"
            elif 'budget' in sub_lower:
                col_type = "budget"
            elif sub_lower.startswith('v.') or sub_lower.startswith('vs'):
                col_type = "variance"
            else:
                col_type = "other"

            columns[col_num] = {
                "year": year,
                "type": col_type,
                "header": sub,
                "display": f"{year} {sub}" if sub else str(year)
            }

    return columns, header_row, sub_header_row


def classify_row(label):
    """Classify a row as header, data, subtotal, or skip."""
    if not label:
        return "empty"

    label_stripped = label.strip()
    label_lower = label_stripped.lower()

    # Section headers
    if label_lower in ('income', 'expenses', 'non-operating income', 'non-operating expense'):
        return "section_header"

    # Subtotals (indented with spaces or contain "total")
    if 'total' in label_lower:
        return "subtotal"

    # Net operating line / total-surplus / total-deficit lines (computed subtotals).
    # FA directive 2026-05-14 Phase 4.4.2: tightened from the old greedy
    # "surplus in label_lower" match, which mis-classified rows like
    # "Prior Year Surplus" as a subtotal. Prior Year Surplus is a data line
    # (last year's accumulated surplus carried into the budget), not a
    # computed total. Require the full subtotal phrase.
    if 'net operating' in label_lower:
        return "subtotal"
    if 'total surplus' in label_lower or 'total deficit' in label_lower:
        return "subtotal"

    # Footnote lines
    if label_stripped.startswith('*') or label_stripped.startswith('(') and len(label_stripped) < 5:
        return "footnote"

    # Data row
    return "data"


def parse_yrlycomp(filepath, target_year=2024):
    """
    Parse the yrlycomp tab from an approved budget Excel file.

    Returns a dict with:
    - entity_code: str
    - building_name: str
    - budget_year: int (the budget year, e.g. 2026)
    - columns: list of column metadata
    - rows: list of all data rows preserving exact Excel structure
    - footnotes: list of footnote text
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = find_yrlycomp_tab(wb)

    if sheet is None:
        return {"error": f"No yrlycomp tab found in {filepath}"}

    entity_code = extract_entity_code(filepath)
    building_name = extract_building_name(sheet, filepath)

    # Parse column structure
    columns, header_row, sub_header_row = parse_column_headers(sheet)

    if not columns:
        return {"error": f"Could not parse column headers in {filepath}"}

    # Determine budget year from the file header
    budget_year = None
    for r in range(1, 6):
        for c in range(1, 20):
            val = sheet.cell(row=r, column=c).value
            if val and isinstance(val, str) and 'year ending' in val.lower():
                match = re.search(r'(\d{4})', val)
                if match:
                    budget_year = int(match.group(1))

    data_start = sub_header_row + 1

    # FA dir 2026-05-22: Auto-detect label column — old logic only scanned
    # columns 1-2 looking for the literal word "Income". When the building's
    # Excel template puts labels in column C or D (entities 710, 847 — both
    # GREEN-tier with all files present, both silently produced 0 rows),
    # the parser found nothing and silently exited.
    #
    # New strategy (two-stage):
    #   1) Scan columns 1-5 for any common section/total label
    #      ("Income", "Revenue", "Expenses", "Total Revenue", "Maintenance",
    #      "Total Surplus"). First hit wins.
    #   2) If no keyword match, pick the column (1-5, excluding year columns
    #      detected by parse_column_headers) with the most text-string values
    #      in the first ~15 data rows — the label column is wherever the
    #      labels actually are.
    year_cols_set = set(columns.keys())
    KNOWN_LABELS = {
        "income", "revenue", "total income", "total revenue", "expenses",
        "operating expenses", "total expenses", "total operating expenses",
        "maintenance", "total surplus", "total surplus <deficit>",
    }
    label_col = None
    scan_end = min(data_start + 20, sheet.max_row + 1)
    # Stage 1 — keyword match
    for c in (1, 2, 3, 4, 5):
        if c in year_cols_set:
            continue
        for r in range(data_start, scan_end):
            val = sheet.cell(row=r, column=c).value
            if val and isinstance(val, str) and val.strip().lower() in KNOWN_LABELS:
                label_col = c
                break
        if label_col is not None:
            break
    # Stage 2 — fallback: column with most string content
    if label_col is None:
        best_count = 0
        for c in (1, 2, 3, 4, 5):
            if c in year_cols_set:
                continue
            count = 0
            for r in range(data_start, scan_end):
                val = sheet.cell(row=r, column=c).value
                if val and isinstance(val, str) and val.strip():
                    count += 1
            if count > best_count:
                best_count = count
                label_col = c
        if label_col is None:
            label_col = 2  # absolute fallback

    # Footnote marker is the column after labels
    marker_col = label_col + 1

    # Parse ALL data rows
    rows = []
    footnotes = []
    current_section = None
    display_order = 0
    found_total_surplus = False

    for row_num in range(data_start, sheet.max_row + 1):
        label_raw = sheet.cell(row=row_num, column=label_col).value
        footnote_marker = sheet.cell(row=row_num, column=marker_col).value

        # Skip completely empty rows
        has_any_data = False
        for c in range(2, max(columns.keys()) + 1 if columns else 20):
            if sheet.cell(row=row_num, column=c).value is not None:
                has_any_data = True
                break

        if not has_any_data:
            continue

        # Convert label
        label = str(label_raw).strip() if label_raw else None

        if not label:
            # Check if it's a note in a later column (skip these)
            continue

        # Stop parsing data after "Total Surplus <Deficit>" - everything after is footnotes
        if found_total_surplus:
            footnotes.append(label)
            continue

        # Check for footnote text (lines starting with *, (A), etc.)
        if label.startswith('* ') or label.startswith('** '):
            footnotes.append(label)
            continue

        # Classify the row
        row_type = classify_row(label)

        if row_type == "section_header":
            current_section = label.strip()

        if row_type in ("empty", "footnote"):
            continue

        # Extract values for each column
        values = {}
        for col_num, col_info in columns.items():
            cell_val = sheet.cell(row=row_num, column=col_num).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)):
                    values[col_num] = round(cell_val, 2)
                elif isinstance(cell_val, str):
                    # Handle text values like "#DIV/0!", "incl w/above"
                    values[col_num] = cell_val
                else:
                    values[col_num] = cell_val

        display_order += 1

        row_data = {
            "row_num": row_num,
            "display_order": display_order,
            "label": label.strip(),
            "footnote_marker": str(footnote_marker).strip() if footnote_marker else None,
            "section": current_section,
            "row_type": row_type,
            "values": values
        }

        rows.append(row_data)

        # Mark after adding - Total Surplus <Deficit> is the last real row
        if 'total surplus' in label.lower() or 'total deficit' in label.lower():
            found_total_surplus = True

    # Build column list sorted by position
    column_list = []
    for col_num in sorted(columns.keys()):
        col = columns[col_num]
        col["col_num"] = col_num
        column_list.append(col)

    result = {
        "entity_code": entity_code,
        "building_name": building_name,
        "budget_year": budget_year,
        "source_file": os.path.basename(filepath),
        "parsed_at": datetime.now().isoformat(),
        "columns": column_list,
        "rows": rows,
        "footnotes": footnotes,
        "stats": {
            "total_rows": len(rows),
            "data_rows": len([r for r in rows if r["row_type"] == "data"]),
            "subtotal_rows": len([r for r in rows if r["row_type"] == "subtotal"]),
            "section_headers": len([r for r in rows if r["row_type"] == "section_header"]),
            "year_columns": len(column_list),
        }
    }

    wb.close()
    return result


# ─── Income tab — Maintenance / Common Charges history ────────────────────
# Pulls the year-by-year history block from the "Income" sheet of an approved
# budget XLSX. Coops use GL 4010-0000 (Maintenance) with shares + $/share;
# condos use GL 4020-0000 (Common Charges) without shares. Auto-detects
# building type by which GL appears first in the tab.

_GL_RE = re.compile(r'^\d{4}-\d{4}$')

def _income_to_int_year(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        n = int(v)
        return n if 1900 <= n <= 2100 else None
    s = str(v).strip().replace(',', '')
    m = re.search(r'(?<!\d)(20[0-2]\d|19[5-9]\d)(?!\d)', s)
    return int(m.group(1)) if m else None


def _income_to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('$', '').replace('%', '')
    try:
        return float(s)
    except Exception:
        return None


def _income_gl_in_row(ws, row):
    for c in (1, 2):
        v = str(ws.cell(row=row, column=c).value or '').strip()
        if _GL_RE.match(v):
            return v
    return None


def parse_income_history(filepath):
    """Read the Income tab's year-by-year history for the first GL section
    (Maintenance for coops, Common Charges for condos). Stops at the next
    GL marker so we don't bleed into Storage/Garage/etc.

    Returns dict:
      {
        "building_type": "coop" | "condo" | "other",
        "gl_code": "4010-0000",
        "history": [
            {year, year_label?, monthly, annual, increase, shares?, perShare?},
            ...
        ]
      }
    On any structural problem, returns {"error": "..."}. Caller is expected
    to catch and treat as a non-fatal side-effect.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if 'Income' not in wb.sheetnames:
        return {"error": "No Income tab"}
    ws = wb['Income']

    # 1) Find the first GL marker — anchors our section.
    target_gl = None
    target_gl_row = None
    for r in range(1, min(ws.max_row + 1, 60)):
        gl = _income_gl_in_row(ws, r)
        if gl:
            target_gl = gl
            target_gl_row = r
            break
    if not target_gl:
        return {"error": "No GL code found in Income tab"}

    if target_gl.startswith('4010'):
        btype = 'coop'
    elif target_gl.startswith('4020'):
        btype = 'condo'
    else:
        btype = 'other'

    # 2) Locate header row (where 'Shares', 'Monthly', or 'Annual' appears).
    header_row = None
    for r in range(1, min(target_gl_row + 5, ws.max_row + 1)):
        for c in range(1, min(ws.max_column + 1, 20)):
            v = str(ws.cell(row=r, column=c).value or '').strip()
            if v in ('Shares', 'Monthly', 'Annual'):
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        return {"error": "No header row found"}

    # 3) Map columns by header text.
    col_map = {}
    for c in range(1, min(ws.max_column + 1, 20)):
        v = str(ws.cell(row=header_row, column=c).value or '').strip()
        if v == 'Shares':
            col_map['shares'] = c
        elif v in ('Mthly $/per sh', 'Mthly $/per Sh'):
            col_map['perShare'] = c
        elif v == 'Monthly':
            col_map['monthly'] = c
        elif v == 'Annual':
            col_map['annual'] = c
        elif v == 'Increase':
            col_map['increase'] = c

    # 4) Year column = the col immediately left of Shares (coop) or Monthly (condo).
    candidates = []
    if 'shares' in col_map:
        candidates.append(col_map['shares'] - 1)
    if 'monthly' in col_map:
        candidates.append(col_map['monthly'] - 1)
    year_col = None
    for cc in candidates:
        if cc < 1:
            continue
        for r in range(header_row + 1, min(header_row + 30, ws.max_row + 1)):
            v = ws.cell(row=r, column=cc).value
            if _income_to_int_year(v) is not None or (v and re.search(r'20\d{2}', str(v))):
                year_col = cc
                break
        if year_col:
            break
    if not year_col:
        return {"error": "No year column detected"}

    # 5) Walk rows, stop at next different GL.
    history = []
    for r in range(header_row + 1, ws.max_row + 1):
        gl_here = _income_gl_in_row(ws, r)
        if gl_here and gl_here != target_gl:
            break
        year_raw = ws.cell(row=r, column=year_col).value
        year_int = _income_to_int_year(year_raw)
        year_str = str(year_raw or '').strip() if year_raw else ''
        if year_int is None and not (year_str and re.search(r'20\d{2}', year_str)):
            continue

        shares = _income_to_float(ws.cell(row=r, column=col_map['shares']).value) if 'shares' in col_map else None
        per_sh = _income_to_float(ws.cell(row=r, column=col_map['perShare']).value) if 'perShare' in col_map else None
        monthly = _income_to_float(ws.cell(row=r, column=col_map['monthly']).value) if 'monthly' in col_map else None
        annual = _income_to_float(ws.cell(row=r, column=col_map['annual']).value) if 'annual' in col_map else None
        incr = _income_to_float(ws.cell(row=r, column=col_map['increase']).value) if 'increase' in col_map else None

        if not any(v is not None and v != 0 for v in (shares, per_sh, monthly, annual)):
            continue

        rec = {
            "year": year_int if year_int is not None else year_str,
            "year_label": year_str if year_str and year_str != str(year_int) else None,
            "monthly": monthly,
            "annual": annual,
            "increase": incr,
        }
        if 'shares' in col_map:
            rec["shares"] = shares
        if 'perShare' in col_map:
            rec["perShare"] = per_sh
        history.append(rec)

    wb.close()
    return {"building_type": btype, "gl_code": target_gl, "history": history}


def format_currency(val):
    """Format a number as currency string."""
    if val is None:
        return ""
    if isinstance(val, str):
        return val
    if val < 0:
        return f"({abs(val):,.0f})"
    return f"{val:,.0f}"


def print_summary(parsed):
    """Print a readable summary of parsed data."""
    if "error" in parsed:
        print(f"ERROR: {parsed['error']}")
        return

    print(f"\n{'='*80}")
    print(f"Entity: {parsed['entity_code']} - {parsed['building_name']}")
    print(f"Budget Year: {parsed['budget_year']}")
    print(f"Source: {parsed['source_file']}")
    print(f"Stats: {parsed['stats']}")
    print(f"{'='*80}")

    # Print columns
    print(f"\nColumns:")
    for col in parsed['columns']:
        print(f"  Col {col['col_num']}: {col['display']} ({col['type']})")

    # Print rows
    print(f"\nRows ({len(parsed['rows'])} total):")
    for row in parsed['rows']:
        indent = "  " if row['row_type'] == 'data' else ""
        marker = f" {row['footnote_marker']}" if row['footnote_marker'] else ""

        # Get the audited actual value for target year if available
        actual_val = ""
        for col in parsed['columns']:
            if col['type'] == 'audited_actual' and col['year'] == 2024:
                val = row['values'].get(col['col_num'])
                if val is not None:
                    actual_val = f"  →  2024 Actual: {format_currency(val)}"
                break

        if row['row_type'] == 'section_header':
            print(f"\n[{row['label']}]")
        elif row['row_type'] == 'subtotal':
            print(f"  {'─'*40}")
            print(f"  {row['label']}{marker}{actual_val}")
        else:
            print(f"  {indent}{row['label']}{marker}{actual_val}")


if __name__ == "__main__":
    # Test with 204 - 2026 budget
    test_file = "/sessions/ecstatic-sleepy-thompson/mnt/Budgets/budget_app/204 -  444 East 86th Street 2026 Operating Budget  - Approved.xlsx"

    if os.path.exists(test_file):
        parsed = parse_yrlycomp(test_file, target_year=2024)
        print_summary(parsed)

        # Save parsed JSON for inspection
        with open("/sessions/ecstatic-sleepy-thompson/mnt/Budgets/budget_summary/204_parsed.json", "w") as f:
            json.dump(parsed, f, indent=2, default=str)
        print(f"\nJSON saved to 204_parsed.json")
    else:
        print(f"File not found: {test_file}")
