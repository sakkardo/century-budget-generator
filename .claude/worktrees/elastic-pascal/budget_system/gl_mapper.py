"""
Map GL codes to template sheet and row locations.

Dynamically scans the template file to find ALL GL codes in column A,
regardless of whether they appear in any config list.
"""

import re
import logging
from pathlib import Path
from typing import Dict, Tuple, Optional, List, Any
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# GL code pattern: digits-digits (e.g., 4010-0000, 5105-0010, 6315-0035)
GL_CODE_PATTERN = re.compile(r"^\d+-\d+$")

# Sheets that contain GL code data rows (skip Setup, Budget Summary, RE Taxes, Insurance Schedule)
GL_DATA_SHEETS = [
    "Income",
    "Payroll",
    "Energy",
    "Water & Sewer",
    "Repairs & Supplies",
    "Gen & Admin",
]


class GLMapper:
    """Maps GL codes to their locations in the template workbook."""

    def __init__(self, template_path: Path):
        """
        Initialize mapper with template workbook.

        Args:
            template_path: Path to template.xlsx file
        """
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_path}")

        self.wb = load_workbook(self.template_path, data_only=False)
        self._mapping = {}

    def build_mapping(self) -> Dict[str, Tuple[str, int]]:
        """
        Scan template and build GL code to (sheet_name, row) mapping.

        Dynamically finds ALL GL codes in column A of each data sheet.
        No hardcoded list needed — any XXXX-XXXX pattern is captured.

        Returns:
            Dict: {gl_code: (sheet_name, row_number)}
            Row numbers are Excel row indices (1-based).
        """
        for sheet_name in GL_DATA_SHEETS:
            if sheet_name not in self.wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in template")
                continue

            ws = self.wb[sheet_name]
            self._scan_sheet_for_gl_codes(ws, sheet_name)

        logger.info(f"Built mapping for {len(self._mapping)} GL codes across {len(GL_DATA_SHEETS)} sheets")
        return self._mapping

    def _scan_sheet_for_gl_codes(
        self, ws: Any, sheet_name: str
    ) -> None:
        """
        Scan a sheet's column A to find ALL GL codes and record their row positions.

        Any cell matching the XXXX-XXXX pattern is treated as a GL code.

        Args:
            ws: Worksheet object
            sheet_name: Name of sheet
        """
        count = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws[f"A{row_idx}"].value
            if not cell_value:
                continue

            cell_str = str(cell_value).strip()

            # Match any GL code pattern (digits-digits)
            if GL_CODE_PATTERN.match(cell_str):
                self._mapping[cell_str] = (sheet_name, row_idx)
                count += 1
                logger.debug(f"Found GL {cell_str} at {sheet_name}!A{row_idx}")

        logger.info(f"  {sheet_name}: found {count} GL codes")

    def get_location(self, gl_code: str) -> Optional[Tuple[str, int]]:
        """
        Get the location (sheet, row) for a GL code.

        Args:
            gl_code: GL code to look up

        Returns:
            Tuple of (sheet_name, row_number) or None if not found
        """
        return self._mapping.get(gl_code)

    def get_sheet(self, sheet_name: str) -> Any:
        """
        Get a worksheet by name.

        Args:
            sheet_name: Name of sheet

        Returns:
            Worksheet object
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in template")
        return self.wb[sheet_name]

    def get_workbook(self) -> Any:
        """
        Get the workbook object.

        Returns:
            Workbook object
        """
        return self.wb

    def close(self) -> None:
        """Close the workbook."""
        self.wb.close()

    def get_mapping_summary(self) -> Dict[str, Any]:
        """
        Get a summary of the mapping.

        Returns:
            Dict with summary statistics
        """
        sheets_with_gl = {}
        for gl_code, (sheet_name, row) in self._mapping.items():
            if sheet_name not in sheets_with_gl:
                sheets_with_gl[sheet_name] = []
            sheets_with_gl[sheet_name].append((gl_code, row))

        return {
            "total_gl_codes_mapped": len(self._mapping),
            "sheets_with_gl": {
                sheet: len(codes) for sheet, codes in sheets_with_gl.items()
            },
        }


def build_gl_mapping(template_path: Path) -> Dict[str, Tuple[str, int]]:
    """
    Convenience function to build GL mapping from template.

    Args:
        template_path: Path to template.xlsx

    Returns:
        Dict: {gl_code: (sheet_name, row_number)}
    """
    mapper = GLMapper(template_path)
    return mapper.build_mapping()
