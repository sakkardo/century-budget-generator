"""
PM Budget Review Workflow Blueprint for Century Management.

Implements workflow for FA and PM budget review process with:
- User and building assignment management
- Budget and line item tracking
- PM data entry for R&M line items
- Status progression and approval workflow
"""

from flask import Blueprint, render_template_string, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from decimal import Decimal
import json
import logging

logger = logging.getLogger(__name__)

# Canonical GL-prefix logic — single source of truth (budget_app/gl_logic.py).
# Both this blueprint and app.py import from here so the matcher and the
# double-count guard can never drift. The browser mirror (_sumOrphanOverlap)
# is pinned to the same answers by budget_app/gl_test_vectors.json. Dual import
# mirrors the create_workflow_blueprint pattern (bare name, package fallback).
try:
    from gl_logic import (
        gl_matches_prefixes,
        gl_prefixes_overlap,
        gl_token_covered_by,
        gl_family,
    )
except ImportError:  # pragma: no cover - package-mode load
    from budget_app.gl_logic import (
        gl_matches_prefixes,
        gl_prefixes_overlap,
        gl_token_covered_by,
        gl_family,
    )

# R&M GL Code Mapping: gl_code|description|template_row|category
# Phase 3 step 3 (2026-06-08): GL classification maps moved to budget_constants.py.
# Re-imported so workflow.py refs + external imports of CAPITAL_GL_PREFIX /
# SUMMARY_PREFIX_OVERRIDES keep working.
try:
    from budget_constants import (RM_GL_MAP, CAPITAL_GL_PREFIX, ONE_TIME_FEE_GLS,
                                  SUMMARY_PREFIX_OVERRIDES)
except ImportError:
    from budget_app.budget_constants import (RM_GL_MAP, CAPITAL_GL_PREFIX, ONE_TIME_FEE_GLS,
                                             SUMMARY_PREFIX_OVERRIDES)

# Capital GL Prefix Map: first 4 digits → description
# 7xxx codes use entity-specific sub-accounts (e.g. 7110-1409) so we match on prefix

# One-time annual fee GLs — billed once per year, so once YTD > 0 the Mar-Dec
# estimate must be zeroed out (otherwise the forecast gets annualized as if
# it recurred monthly). Forecast then collapses to YTD + Accrual + Unpaid.
# FA can still override via estimate_override if a weird case ever comes up.

# Load GL_Mapping.csv (412+ entries) for routing and naming unmapped GLs.
# Indexed by 4-digit prefix so entity-specific sub-accounts (e.g. 4010-1409) match the
# base mapping entry (e.g. 4010-0000). Returns dict: prefix -> (description, sheet_name, category).
# Only codes with an explicit routing rule are included — balance sheet codes and codes
# not present in the mapping file stay Unmapped.
def _csv_row_to_sheet(cat, sub, code):
    """Return (sheet_name, category_key) or None if not explicitly routable.

    Note on R&S sub-categories: the UI groups the Repairs & Supplies tab into
    Supplies / Repairs / Maintenance Contracts buckets by matching BudgetLine.category
    against the strings 'supplies' / 'repairs' / 'maintenance'. Historically this
    function returned a lumped 'rm' bucket which made most R&S lines invisible in
    the Supplies group. We now preserve the sub-category from the CSV.
    """
    if cat == "Income":
        return ("Income", "income")
    if cat == "Gen & Admin Expenses":
        return ("Gen & Admin", "gen_admin")
    # Operating Expenses or blank category — use Sub-Category to pick the sheet
    if sub == "Payroll Expenses":
        return ("Payroll", "payroll")
    if sub == "Utility Expenses":
        # 63xx is Water/Sewer; all other utility codes are Energy
        if code.startswith("63"):
            return ("Water & Sewer", "water_sewer")
        return ("Energy", "energy")
    if sub == "Supplies":
        return ("Repairs & Supplies", "supplies")
    if sub == "Repairs":
        return ("Repairs & Supplies", "repairs")
    if sub == "Maintenance":
        return ("Repairs & Supplies", "maintenance")
    return None

def _load_gl_mapping_csv():
    import csv as _csv
    from pathlib import Path as _Path
    mapping = {}
    candidates = [
        _Path(__file__).parent.parent / "budget_system" / "GL_Mapping.csv",
        _Path(__file__).parent / "GL_Mapping.csv",
    ]
    for p in candidates:
        if p.exists():
            try:
                with open(p, newline="", encoding="utf-8-sig") as f:
                    for row in _csv.DictReader(f):
                        code = (row.get("GL Code") or "").strip()
                        desc = (row.get("Description") or "").strip()
                        cat = (row.get("Category Tab") or "").strip()
                        sub = (row.get("Sub-Category") or "").strip()
                        if not (code and desc):
                            continue
                        routing = _csv_row_to_sheet(cat, sub, code)
                        if routing is None:
                            continue  # Skip rows we can't confidently route
                        prefix = code[:4]
                        # First explicit routing wins (CSV is ordered by category)
                        if prefix not in mapping:
                            mapping[prefix] = (desc, routing[0], routing[1])
            except Exception:
                pass
            break
    return mapping

GL_MAPPING_CSV = _load_gl_mapping_csv()


# ─── SUMMARY ROW PREFIX OVERRIDES ────────────────────────────────────────
# Budget Summary tab rows carry gl_prefixes_json used to aggregate YTD/estimate/
# forecast from budget_lines. Historical push files (generated from the legacy
# GL_TO_SUMMARY_MAP.py) contain stale chart-of-accounts prefixes that predate
# the Yardi re-numbering. These overrides are the canonical Yardi prefixes
# keyed by canonical summary row label. Applied at BOTH import time (so future
# imports auto-correct) and via startup backfill (so existing DB rows are
# fixed). Source of truth: budget_system/GL_Mapping.csv (Utility Expenses +
# Supplies sub-categories). If a label is added to SUMMARY_PREFIX_OVERRIDES
# here, no per-building redeployment is needed.


def apply_summary_prefix_override(label, existing_prefixes):
    """Return corrected prefix list for a summary row label.

    Used at import time and in startup backfill. Only overrides labels
    explicitly listed in SUMMARY_PREFIX_OVERRIDES; all other rows pass
    through untouched.
    """
    if not label:
        return existing_prefixes
    override = SUMMARY_PREFIX_OVERRIDES.get(label.strip())
    if override:
        return list(override)
    return existing_prefixes


# ─── FIXED-FORECAST INCOME GLs ──────────────────────────────────────────
# Business rule (from Jacob, 2026-04-14): for Maintenance / Common Charges /
# Commercial Rent income rows on the Budget Summary tab, forecast (Col 5)
# must equal Approved Budget (Col 6) — these are predictable contractual
# amounts, not forecast-from-YTD. Col 4 (Estimate) is then set so the math
# ties out: Col 4 = Col 5 - Col 3. Matched by GL prefix (not label) since
# labels vary across buildings (co-op vs condo).
#
# Full GLs provided: 4010-0000, 4020-0000, 4020-0005, 4030-0000,
#                    4040-0000, 4040-0010
#
# FA dir 2026-06-02 (task #99): Operating Assessment (GL 4200) joins the set.
# Operating assessments are fully collectible (billed to all unit owners like
# maintenance), so forecast must equal the approved budget rather than an
# annualized-from-YTD figure. Matched by GL prefix 4200 — this deliberately
# EXCLUDES Capital Assessment and Tax-Abatement Assessment, which are
# unmapped manual Summary rows (no GL), not GL-4200 operating assessments.
# Stored as 4-digit bases to match _gl_matches_prefixes behavior.
# Phase 3 step 2 (2026-06-08): fixed-forecast GL constants + the budget-category
# map moved to budget_constants.py (pure data, internal-only). Imported here so
# every reference in this file resolves unchanged.
try:
    from budget_constants import (FIXED_FORECAST_GL_BASES, FIXED_FORECAST_GL_FULL,
                                  _row_has_fixed_forecast_gl, BUDGET_CAT_TO_CENTURY)
except ImportError:
    from budget_app.budget_constants import (FIXED_FORECAST_GL_BASES, FIXED_FORECAST_GL_FULL,
                                             _row_has_fixed_forecast_gl, BUDGET_CAT_TO_CENTURY)
try:
    from summary_engine import (compute_summary, _gl_matches_prefixes,
                                _section_key, _is_capital_line,
                                _aggregate_by_prefix)
except ImportError:
    from budget_app.summary_engine import (compute_summary, _gl_matches_prefixes,
                                           _section_key, _is_capital_line,
                                           _aggregate_by_prefix)

# Phase 3 step 1 (2026-06-08): the status state machine + lifecycle vocabulary
# now live in budget_status.py (one module). Imported here so every existing
# reference in this file resolves unchanged. Try/except covers both the bare
# (budget_app on sys.path) and package-qualified import contexts.
try:
    from budget_status import (BUDGET_STATUSES, USER_ROLES, VALID_TRANSITIONS,
                               LIFECYCLE_STAGES, derive_lifecycle_stage)
except ImportError:
    from budget_app.budget_status import (BUDGET_STATUSES, USER_ROLES, VALID_TRANSITIONS,
                                          LIFECYCLE_STAGES, derive_lifecycle_stage)

# Status UX Phase 1 (2026-06-09): shared per-source file-status model — the one
# brain the FA dashboard AND the wizard render tile colors from. Jacob's rule:
# green = the file is in a BUILT budget; amber = in SharePoint; red = missing/
# failed; gray = setup. See STATUS_UX_PLAN.md + source_status.py.
try:
    from source_status import compute_source_states
except ImportError:
    from budget_app.source_status import compute_source_states

# ─── Budget Year Config ─────────────────────────────────────────────────────
# Change this ONE value each cycle. All routes, queries, and column headers
# derive their years from this.  BY=2027 means:
#   Col 1 = 2024 Actual   (BY-3)
#   Col 2 = 2025 Actual   (BY-2)
#   Col 3 = 2026 YTD      (BY-1)
#   Col 4 = 2026 Est.     (BY-1)
#   Col 5 = 2026 Forecast (BY-1)
#   Col 6 = 2026 Budget   (BY-1)
#   Col 7 = 2027 Budget   (BY)
import os
# Phase 3 step 4 (2026-06-08): BUDGET_YEAR moved to budget_config.py (a leaf module
# that imports only stdlib). The models reference BUDGET_YEAR 13x; putting it below
# them in the import graph is the prerequisite for extracting them without a circular
# dependency. Re-imported here so every reference in this file — and app.py's
# `from workflow import BUDGET_YEAR` — resolves to the identical value, unchanged.
try:
    from budget_config import BUDGET_YEAR
except ImportError:
    from budget_app.budget_config import BUDGET_YEAR



# ─── Lifecycle stages + derive_lifecycle_stage moved to budget_status.py
#     (Phase 3 step 1) and imported above. No behavior change.


def create_workflow_blueprint(db):
    """
    Create and configure the workflow blueprint.

    Args:
        db: SQLAlchemy database instance from app.py

    Returns:
        tuple: (blueprint, models_dict, helpers_dict)
    """

    # ─── SQLAlchemy Models ────────────────────────────────────────────────────

    # Phase 3 step 7 (2026-06-08): the 18 SQLAlchemy models + _parse_backup_json
    # now live in models.py. register_models(db) builds them with the injected db
    # and returns the same name->class dict; we rebind each name locally so every
    # route + helper below (and the workflow_models dict returned at the end)
    # resolve unchanged. Verbatim move, no behavior change.
    try:
        from models import register_models
    except ImportError:
        from budget_app.models import register_models
    _wm_all = register_models(db)
    User = _wm_all["User"]
    BuildingAssignment = _wm_all["BuildingAssignment"]
    Budget = _wm_all["Budget"]
    BudgetLine = _wm_all["BudgetLine"]
    DataSource = _wm_all["DataSource"]
    AuditSyncRun = _wm_all["AuditSyncRun"]
    BudgetRevision = _wm_all["BudgetRevision"]
    BuildingVisit = _wm_all["BuildingVisit"]
    PresentationSession = _wm_all["PresentationSession"]
    PresentationEdit = _wm_all["PresentationEdit"]
    ARHandoff = _wm_all["ARHandoff"]
    PayrollPosition = _wm_all["PayrollPosition"]
    PayrollAssumption = _wm_all["PayrollAssumption"]
    CommercialTenant = _wm_all["CommercialTenant"]
    CommercialRentPeriod = _wm_all["CommercialRentPeriod"]
    CommercialTenantBillback = _wm_all["CommercialTenantBillback"]
    BudgetSummaryRow = _wm_all["BudgetSummaryRow"]
    BuildingInfo = _wm_all["BuildingInfo"]
    CamClass = _wm_all["CamClass"]
    CamAllocationOverride = _wm_all["CamAllocationOverride"]
    BudgetNarrative = _wm_all["BudgetNarrative"]









    # ─── New Pipeline Tables ─────────────────────────────────────────────────















    # ─── Payroll Models ───────────────────────────────────────────────────





    # ─── Helper Functions ────────────────────────────────────────────────────

    def store_rm_lines(entity_code, building_name, gl_data):
        """
        Store R&M lines from YSL data into the database.

        gl_data: dict of {gl_code: {period_2, period_3, period_4, period_5, ...}}

        - period_2 → prior_year
        - period_3 → ytd_actual
        - period_4 → ytd_budget
        - period_5 → current_budget

        Only stores lines where gl_code is in RM_GL_MAP.
        If budget exists with status='draft', updates lines.
        If status is anything else, only updates YSL columns (doesn't overwrite PM inputs).
        """
        try:
            # Get or create budget
            budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
            if not budget:
                budget = Budget(
                    entity_code=entity_code,
                    building_name=building_name,
                    year=BUDGET_YEAR,
                    status="draft"
                )
                db.session.add(budget)
                db.session.flush()

            is_draft = budget.status == "draft"

            # Process each GL code from YSL
            for gl_code, gl_values in gl_data.items():
                if gl_code not in RM_GL_MAP:
                    continue

                desc, row_num, category = RM_GL_MAP[gl_code]

                prior_year = float(gl_values.get("period_2", 0) or 0)
                ytd_actual = float(gl_values.get("period_3", 0) or 0)
                ytd_budget = float(gl_values.get("period_4", 0) or 0)
                current_budget = float(gl_values.get("period_5", 0) or 0)

                # Find existing line or create new
                line = BudgetLine.query.filter_by(budget_id=budget.id, gl_code=gl_code).first()

                if line:
                    # Always update YSL-sourced columns
                    line.prior_year = prior_year
                    line.ytd_actual = ytd_actual
                    line.ytd_budget = ytd_budget
                    line.current_budget = current_budget
                    # Only reset PM inputs if this is a draft
                    if is_draft:
                        line.accrual_adj = 0.0
                        line.unpaid_bills = 0.0
                        line.increase_pct = 0.0
                        line.notes = ""
                        line.reclass_to_gl = None
                        line.reclass_amount = 0.0
                        line.reclass_notes = ""
                        line.estimate_override = None
                        line.forecast_override = None
                        line.proposed_budget = 0.0
                else:
                    # Create new line
                    line = BudgetLine(
                        budget_id=budget.id,
                        gl_code=gl_code,
                        description=desc,
                        category=category,
                        row_num=row_num,
                        prior_year=prior_year,
                        ytd_actual=ytd_actual,
                        ytd_budget=ytd_budget,
                        current_budget=current_budget
                    )
                    db.session.add(line)

            db.session.commit()
            logger.info(f"Stored R&M lines for {entity_code}")
            return True
        except Exception as e:
            logger.error(f"Error storing R&M lines: {e}")
            db.session.rollback()
            return False


    SHEET_TO_CATEGORY = {
        "Income": "income",
        "Payroll": "payroll",
        "Energy": "energy",
        "Water & Sewer": "water_sewer",
        "Repairs & Supplies": "rm",
        "Gen & Admin": "gen_admin",
        "Capital": "capital",
    }

    def _delete_entity_data(entity_code):
        """Delete ALL entity-level supplementary data (expenses, open AP, etc.).
        Called by budget deletion to fully remove an entity's data.
        Each table is deleted in its own try/except with rollback to prevent
        one failure from poisoning the entire transaction."""
        ec = str(entity_code).strip()
        for sql in [
            "DELETE FROM expense_invoices WHERE report_id IN (SELECT id FROM expense_reports WHERE entity_code = :ec)",
            "DELETE FROM expense_reports WHERE entity_code = :ec",
            "DELETE FROM open_ap_invoices WHERE report_id IN (SELECT id FROM open_ap_reports WHERE entity_code = :ec)",
            "DELETE FROM open_ap_reports WHERE entity_code = :ec",
        ]:
            try:
                db.session.execute(db.text(sql), {"ec": ec})
            except Exception as e:
                db.session.rollback()
                logger.warning(f"_delete_entity_data skip: {e}")
        logger.info(f"Deleted entity-level data for {ec}")

    def store_all_lines(entity_code, building_name, gl_data, template_path, assumptions=None, fresh_start=False):
        """
        Store ALL GL codes from YSL data into budget_lines (not just R&M).
        Uses GLMapper to get sheet/row/description for every GL code.
        Optionally stores merged assumptions snapshot on the Budget record.
        If fresh_start=True, deletes all existing lines and resets to draft.
        """
        try:
            from gl_mapper import build_gl_mapping_with_descriptions
        except ImportError:
            from budget_system.gl_mapper import build_gl_mapping_with_descriptions

        try:
            gl_mapping = build_gl_mapping_with_descriptions(template_path)
        except Exception as e:
            logger.error(f"Failed to build GL mapping: {e}")
            gl_mapping = {}

        try:
            budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
            if not budget:
                budget = Budget(
                    entity_code=entity_code,
                    building_name=building_name,
                    year=BUDGET_YEAR,
                    status="draft"
                )
                db.session.add(budget)
                db.session.flush()

            # Fresh start: wipe all existing lines and reset to draft
            if fresh_start and budget.id:
                deleted = BudgetLine.query.filter_by(budget_id=budget.id).delete()
                budget.status = "draft"
                db.session.flush()
                logger.info(f"Fresh start: deleted {deleted} lines for {entity_code}")

            # Store assumptions snapshot if provided
            if assumptions:
                import json as _json_mod
                budget.assumptions_json = _json_mod.dumps(assumptions)

            is_draft = budget.status == "draft"
            stored = 0

            for gl_code, gl_values in gl_data.items():
                prior_year = float(gl_values.get("period_2", 0) or 0)
                ytd_actual = float(gl_values.get("period_3", 0) or 0)
                ytd_budget = float(gl_values.get("period_4", 0) or 0)
                current_budget = float(gl_values.get("period_5", 0) or 0)

                # Determine sheet, row, description, category
                if gl_code in RM_GL_MAP:
                    desc, row_num, category = RM_GL_MAP[gl_code]
                    sheet_name = "Repairs & Supplies"
                    pm_editable = True
                elif gl_code in gl_mapping:
                    sheet_name, row_num, desc = gl_mapping[gl_code]
                    category = SHEET_TO_CATEGORY.get(sheet_name, "other")
                    # Repairs & Supplies needs sub-category split (supplies/repairs/maintenance)
                    # for the UI grouping. SHEET_TO_CATEGORY returns the lumped "rm" bucket,
                    # so look up the actual sub-category in GL_Mapping.csv by 4-digit prefix.
                    if category == "rm":
                        _csv_hit = GL_MAPPING_CSV.get(gl_code[:4])
                        category = _csv_hit[2] if _csv_hit else "repairs"
                    pm_editable = False
                elif gl_code.startswith("7"):
                    prefix = gl_code[:4]
                    desc = CAPITAL_GL_PREFIX.get(prefix, f"Cap - {prefix}")
                    sheet_name = "Capital"
                    row_num = 0
                    category = "capital"
                    pm_editable = True
                else:
                    # Try GL_Mapping.csv for explicit routing to a real tab.
                    # Only codes present in the mapping file get routed; everything else
                    # (balance sheet codes, codes missing from mapping) stays Unmapped.
                    _csv_hit = GL_MAPPING_CSV.get(gl_code[:4])
                    if _csv_hit:
                        desc, sheet_name, category = _csv_hit
                        row_num = 0
                        pm_editable = True
                    else:
                        desc = gl_code
                        sheet_name = "Unmapped"
                        row_num = 0
                        category = "other"
                        pm_editable = False

                line = BudgetLine.query.filter_by(budget_id=budget.id, gl_code=gl_code).first()
                if line:
                    line.prior_year = prior_year
                    line.ytd_actual = ytd_actual
                    line.ytd_budget = ytd_budget
                    line.current_budget = current_budget
                    line.sheet_name = sheet_name
                    line.description = desc
                    line.category = category
                    line.pm_editable = pm_editable
                    if is_draft:
                        line.accrual_adj = 0.0
                        line.unpaid_bills = 0.0
                        line.increase_pct = 0.0
                        line.notes = ""
                        line.reclass_to_gl = None
                        line.reclass_amount = 0.0
                        line.reclass_notes = ""
                        line.estimate_override = None
                        line.forecast_override = None
                        line.proposed_budget = 0.0
                else:
                    line = BudgetLine(
                        budget_id=budget.id,
                        gl_code=gl_code,
                        description=desc,
                        category=category,
                        row_num=row_num,
                        sheet_name=sheet_name,
                        pm_editable=pm_editable,
                        prior_year=prior_year,
                        ytd_actual=ytd_actual,
                        ytd_budget=ytd_budget,
                        current_budget=current_budget
                    )
                    db.session.add(line)
                stored += 1

            db.session.commit()
            logger.info(f"Stored {stored} lines for {entity_code} (all GL codes)")
            return True
        except Exception as e:
            logger.error(f"Error storing lines: {e}")
            db.session.rollback()
            return False


    def get_pm_projections(entity_code):
        """
        Get PM-entered projections for a building.

        Returns: {gl_code: {accrual_adj, unpaid_bills, increase_pct, notes}}
        """
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return {}

        result = {}
        for line in budget.lines:
            result[line.gl_code] = {
                "accrual_adj": float(line.accrual_adj or 0),
                "unpaid_bills": float(line.unpaid_bills or 0),
                "increase_pct": float(line.increase_pct or 0),
                "notes": line.notes
            }

        return result


    def compute_forecast(ytd_actual, accrual_adj, unpaid_bills, prior_year, ytd_months=2):
        """
        Compute 12-month forecast.

        Formula: ytd_actual + accrual_adj + unpaid_bills + estimate
        where estimate = (ytd_total / ytd_months) * remaining_months

        FA item #7 anomaly guard: if YTD is negative but prior_year is
        zero/positive, treat as one-time refund/credit and skip extrapolation
        (estimate = 0, forecast = YTD only). Recurring negatives (tax
        credits where prior_year is also negative) keep extrapolating.
        """
        # Phase 1 (2026-06-08): delegate to the single source of truth. Behavior is
        # byte-identical to the prior inline version (anomaly cap preserved); the math
        # now lives in exactly one place. See budget_math.forecast + test_budget_math.py.
        import budget_math
        return budget_math.forecast(ytd_actual, accrual_adj, unpaid_bills, prior_year,
                                    ytd_months, anomaly_cap=True, payroll=False)


    def forecast_method(ytd_actual, accrual_adj, unpaid_bills, prior_year):
        """Return the forecast method label for display purposes."""
        return 'Annualized'


    def compute_proposed_budget(forecast, increase_pct, increase_dollar=None):
        """Compute proposed budget.

        FA directive 2026-05-11: PM can enter EITHER a % increase OR a $
        amount. $ wins when both are set (defensive — either-or is also
        enforced at save time). When both unset, proposed = forecast.

          $ set:      proposed = forecast + $
          % set:      proposed = forecast × (1 + %)
          both NULL:  proposed = forecast
        """
        # Phase 1 (2026-06-08): delegate to the single source of truth (byte-identical).
        import budget_math
        return budget_math.proposed(forecast, increase_pct, increase_dollar)


    # ─── Budget Summary Table ───────────────────────────────────────────────

    # ─── Commercial Rent (Phase 1) ──────────────────────────────────
    # Models for the new Commercial tab. Mirrors the Excel
    # "Comm Rent & Escalations" sheet structure: per-tenant lease
    # schedules + per-tenant escalation config + per-tenant utility/
    # insurance billback base-year amounts.
    # FA directive 2026-05-14 Phase 5 (commercial rent integration).











    # ─── Blueprint Creation ──────────────────────────────────────────────────

    bp = Blueprint("workflow", __name__)


    # ─── Admin Routes ────────────────────────────────────────────────────────

    @bp.route("/admin", methods=["GET"])
    def admin():
        """Admin dashboard for user and building management."""
        import json as json_mod
        users = User.query.all()
        assignments = BuildingAssignment.query.all()

        return render_template_string(
            ADMIN_TEMPLATE,
            users_json=json_mod.dumps([u.to_dict() for u in users]),
            assignments_json=json_mod.dumps([a.to_dict() for a in assignments]),
        )


    @bp.route("/dashboard", methods=["GET"])
    def dashboard():
        """FA Dashboard - view all buildings and budget status."""
        import json as json_mod
        budgets = Budget.query.all()

        status_counts = {}
        for status in BUDGET_STATUSES:
            status_counts[status] = len([b for b in budgets if b.status == status])

        return render_template_string(
            DASHBOARD_TEMPLATE,
            budgets_json=json_mod.dumps([b.to_dict() for b in budgets]),
            status_counts_json=json_mod.dumps(status_counts),
        )


    @bp.route("/dashboard/<entity_code>", methods=["GET"])
    def building_detail(entity_code):
        """FA Building Detail - combined view of budget, expenses, audit.

        FA directive 2026-05-14 (Dashboard Phase 3): the workbook view.
        Edit the budget numbers. Health/readiness/action items live on
        the Action Center at /action/<ec>. Both routes share the header
        strip, status pipeline, and KPI cards.
        """
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return "No budget found for this building", 404
        return render_template_string(BUILDING_DETAIL_TEMPLATE, entity_code=entity_code, budget_year=BUDGET_YEAR)


    @bp.route("/action/<entity_code>", methods=["GET"])
    def action_center(entity_code):
        """FA directive 2026-05-14 (Dashboard Phase 3): the Action Center
        landing. Consolidates 9 readiness gates + 2 wizard sidebar cards +
        2 below-workbook warnings + summary-tab warnings into ONE
        prioritized list grouped by Blockers / Warnings / Complete. Each
        item exists exactly once. Each action button goes somewhere useful
        (inline expand, tab switch + scroll, or dedicated review page —
        never a raw JSON endpoint).

        The workbook (existing /dashboard/<ec>) continues to handle the
        actual budget editing. Cross-links go both ways.
        """
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return "No budget found for this building", 404
        return render_template_string(
            ACTION_CENTER_TEMPLATE,
            entity_code=entity_code,
            budget_year=BUDGET_YEAR,
        )


    @bp.route("/pm", methods=["GET"])
    def pm_portal():
        """PM Portal - select building and edit R&M lines.

        FA dir 2026-05-24: lazy-sync Monday.com on page load so the PM
        dropdown reflects the current Active Buildings (non-Lemle) group
        and current pm8-column values. _ensure_monday_fresh is safe — it
        never raises, falls back to cached data if Monday is unreachable,
        and skips the network call entirely if the cache is <10 min old.
        """
        import json as json_mod

        # Lazy Monday sync (no-op if cache is fresh). Best-effort.
        try:
            try:
                from budget_app.app import _ensure_monday_fresh
            except ImportError:
                from app import _ensure_monday_fresh
            _ensure_monday_fresh()
        except Exception:
            # Never block portal load on a sync hiccup — pruning + stale data
            # is still a safer fallback than a 500.
            pass

        # Ensure at least one PM user exists for demo
        pm_users = User.query.filter_by(role="pm").all()
        if not pm_users:
            demo_pm = User(name="Test PM", email="testpm@centuryny.com", role="pm")
            db.session.add(demo_pm)
            db.session.commit()
            pm_users = [demo_pm]

        return render_template_string(
            PM_PORTAL_TEMPLATE,
            pm_users_json=json_mod.dumps([u.to_dict() for u in pm_users]),
        )


    @bp.route("/pm/<entity_code>", methods=["GET"])
    def pm_edit(entity_code):
        """PM Edit Page - spreadsheet-style R&M grid."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()

        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        # Check if PM can edit this budget
        # fa_review is included so the PM can re-enter and tweak a building
        # after it's been submitted for FA review.
        can_edit = budget.status in ["pm_pending", "pm_in_progress", "returned", "fa_review"]

        # PM sees Repairs & Supplies + Gen & Admin lines
        lines = BudgetLine.query.filter(
            BudgetLine.budget_id == budget.id,
            BudgetLine.sheet_name.in_(["Repairs & Supplies", "Gen & Admin"])
        ).order_by(BudgetLine.row_num).all()
        import json as json_mod

        lines_data = [l.to_dict() for l in lines]

        # Get ALL GL codes for reclass modal (not just pm_editable)
        all_gls = db.session.query(BudgetLine.gl_code, BudgetLine.description, BudgetLine.category).filter_by(budget_id=budget.id).order_by(BudgetLine.gl_code).all()
        all_gl_list = [{"gl_code": g.gl_code, "description": g.description, "category": g.category} for g in all_gls]

        # Derive dynamic YTD months from assumptions
        _ytd_months = 2
        try:
            assumptions = json_mod.loads(budget.assumptions_json) if budget.assumptions_json else {}
            bp = assumptions.get("budget_period", "")
            if "/" in str(bp):
                _ytd_months = int(str(bp).split("/")[0])
        except Exception:
            pass
        _remaining_months = 12 - _ytd_months

        return render_template_string(
            PM_EDIT_TEMPLATE,
            entity_code=entity_code,
            building_name=budget.building_name,
            status=budget.status,
            budget_status=budget.status,
            can_edit="true" if can_edit else "false",
            fa_notes=budget.fa_notes or "",
            lines_json=json_mod.dumps(lines_data),
            all_gl_json=json_mod.dumps(all_gl_list),
            ytd_months=_ytd_months,
            remaining_months=_remaining_months,
            estimate_label=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][_ytd_months] + '-Dec' if _ytd_months < 12 else 'Estimate',
        )


    # ─── API Routes: Users ───────────────────────────────────────────────────

    @bp.route("/api/users", methods=["GET"])
    def list_users():
        """List users.

        Query params:
          role=fa|pm|admin   — filter to one role
          active=1           — only users with at least one BuildingAssignment
                               on a current-year Budget. Excludes Lemle alumni
                               and other stale FAs from pre-filter Monday syncs.
                               Mirrors the wizard's FA-dropdown filter logic.
        """
        role = (request.args.get("role") or "").strip().lower()
        active_only = (request.args.get("active") or "").strip() in ("1", "true", "yes")

        q = User.query
        if role:
            q = q.filter_by(role=role)

        if active_only:
            # Resolve the set of currently-active entity_codes (those with a
            # Budget row for the current year — same population the wizard uses).
            active_entity_codes = {
                b.entity_code for b in
                Budget.query.with_entities(Budget.entity_code)
                            .filter_by(year=BUDGET_YEAR).all()
            }
            if not active_entity_codes:
                return jsonify([])
            # Find user_ids that have at least one assignment in those entities.
            assign_q = (
                BuildingAssignment.query
                .with_entities(BuildingAssignment.user_id)
                .filter(BuildingAssignment.entity_code.in_(active_entity_codes))
            )
            if role:
                assign_q = assign_q.filter_by(role=role)
            active_user_ids = {row[0] for row in assign_q.all()}
            if not active_user_ids:
                return jsonify([])
            q = q.filter(User.id.in_(active_user_ids))
            # FA directive 2026-05-10: Monday's people-column creates pseudo
            # users for joint assignments (e.g., "Jennifer Murman, Giovanni
            # Lizarazo"). Real FA names never contain ", ". Exclude these
            # pseudo-users from the picker.
            q = q.filter(~User.name.like('%, %'))

        users = q.order_by(User.name).all()
        return jsonify([u.to_dict() for u in users])


    @bp.route("/api/users", methods=["POST"])
    def create_user():
        """Create a new user."""
        data = request.get_json()

        if not data.get("name") or not data.get("email") or not data.get("role"):
            return jsonify({"error": "Missing required fields"}), 400

        if data["role"] not in USER_ROLES:
            return jsonify({"error": f"Invalid role. Must be one of {USER_ROLES}"}), 400

        # Check if email already exists
        if User.query.filter_by(email=data["email"]).first():
            return jsonify({"error": "Email already exists"}), 400

        user = User(
            name=data["name"],
            email=data["email"],
            role=data["role"]
        )

        db.session.add(user)
        db.session.commit()

        return jsonify(user.to_dict()), 201


    @bp.route("/api/users/<int:user_id>", methods=["PUT"])
    def update_user(user_id):
        """Update a user."""
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        data = request.get_json()

        if "name" in data:
            user.name = data["name"]
        if "email" in data:
            if data["email"] != user.email and User.query.filter_by(email=data["email"]).first():
                return jsonify({"error": "Email already exists"}), 400
            user.email = data["email"]
        if "role" in data:
            if data["role"] not in USER_ROLES:
                return jsonify({"error": f"Invalid role. Must be one of {USER_ROLES}"}), 400
            user.role = data["role"]

        db.session.commit()
        return jsonify(user.to_dict())


    @bp.route("/api/users/<int:user_id>", methods=["DELETE"])
    def delete_user(user_id):
        """Delete a user."""
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        db.session.delete(user)
        db.session.commit()

        return jsonify({"status": "deleted"}), 204


    # ─── API Routes: Assignments ─────────────────────────────────────────────

    @bp.route("/api/assignments", methods=["GET"])
    def list_assignments():
        """List all building assignments."""
        assignments = BuildingAssignment.query.all()
        return jsonify([a.to_dict() for a in assignments])


    @bp.route("/api/assignments", methods=["POST"])
    def create_assignment():
        """Create a new building assignment."""
        data = request.get_json()

        if not data.get("entity_code") or not data.get("user_id") or not data.get("role"):
            return jsonify({"error": "Missing required fields"}), 400

        if data["role"] not in ["fa", "pm"]:
            return jsonify({"error": "Role must be 'fa' or 'pm'"}), 400

        # Check if user exists
        user = User.query.get(data["user_id"])
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Check for duplicates
        existing = BuildingAssignment.query.filter_by(
            entity_code=data["entity_code"],
            user_id=data["user_id"],
            role=data["role"]
        ).first()

        if existing:
            return jsonify({"error": "Assignment already exists"}), 400

        assignment = BuildingAssignment(
            entity_code=data["entity_code"],
            user_id=data["user_id"],
            role=data["role"]
        )

        db.session.add(assignment)
        db.session.commit()

        return jsonify(assignment.to_dict()), 201


    @bp.route("/api/assignments/<int:assignment_id>", methods=["DELETE"])
    def delete_assignment(assignment_id):
        """Delete an assignment."""
        assignment = BuildingAssignment.query.get(assignment_id)
        if not assignment:
            return jsonify({"error": "Assignment not found"}), 404

        db.session.delete(assignment)
        db.session.commit()

        return jsonify({"status": "deleted"}), 204


    # ─── API Routes: Budgets ─────────────────────────────────────────────────

    @bp.route("/api/budgets", methods=["GET"])
    def list_budgets():
        """List all budgets with status and completeness data."""
        budgets = Budget.query.all()
        # Batch-fetch entities with expenses / audits in one query each (avoids session poisoning from per-row errors)
        try:
            expense_entities = {r[0] for r in db.session.execute(
                db.text("SELECT DISTINCT entity_code FROM expense_reports")
            ).fetchall()}
        except Exception:
            db.session.rollback()
            expense_entities = set()
        try:
            audit_entities = {r[0] for r in db.session.execute(
                db.text("SELECT DISTINCT entity_code FROM audit_uploads WHERE status = 'confirmed'")
            ).fetchall()}
        except Exception:
            db.session.rollback()
            audit_entities = set()

        # Batch-fetch data-loaded timestamps per entity
        # 1) Budget summary import timestamps (earliest imported_at per entity)
        summary_ts = {}
        try:
            # Phase 1 (2026-06-09): year-filtered — the canonical "B loaded" rule.
            # Was unfiltered, which disagreed with the foundation page (year-
            # filtered) and the wizard (row_type-filtered); prior-cycle rows no
            # longer count as this cycle's approved budget.
            rows = db.session.execute(
                db.text("SELECT entity_code, MIN(imported_at) FROM budget_summary_rows "
                        "WHERE budget_year = :by GROUP BY entity_code"),
                {"by": BUDGET_YEAR}
            ).fetchall()
            for r in rows:
                summary_ts[r[0]] = r[1].isoformat() if r[1] else None
        except Exception:
            db.session.rollback()

        # 2) YSL data timestamps — use earliest BudgetLine updated_at per entity as proxy
        #    (YSL import creates/updates budget_lines via store_all_lines)
        ysl_ts = {}
        try:
            rows = db.session.execute(
                db.text("""
                    SELECT b.entity_code, MIN(bl.updated_at)
                    FROM budget_lines bl
                    JOIN budgets b ON b.id = bl.budget_id
                    GROUP BY b.entity_code
                """)
            ).fetchall()
            for r in rows:
                ysl_ts[r[0]] = r[1].isoformat() if r[1] else None
        except Exception:
            db.session.rollback()

        # 3) Expense distribution upload timestamps
        expense_ts = {}
        try:
            rows = db.session.execute(
                db.text("SELECT entity_code, MAX(uploaded_at) FROM expense_reports GROUP BY entity_code")
            ).fetchall()
            for r in rows:
                expense_ts[r[0]] = r[1].isoformat() if r[1] else None
        except Exception:
            db.session.rollback()

        # 4) Open AP import timestamps
        open_ap_ts = {}
        try:
            rows = db.session.execute(
                db.text("SELECT entity_code, MAX(uploaded_at) FROM open_ap_reports GROUP BY entity_code")
            ).fetchall()
            for r in rows:
                open_ap_ts[r[0]] = r[1].isoformat() if r[1] else None
        except Exception:
            db.session.rollback()

        # 4b) Phase 1 (2026-06-09): Maintenance Proof staged timestamps — the M
        # tile's "loaded" signal. Tracked in data all along, invisible until now.
        maint_ts = {}
        try:
            rows = db.session.execute(
                db.text("SELECT entity_code, MAX(uploaded_at) FROM maint_proof_reports GROUP BY entity_code")
            ).fetchall()
            for r in rows:
                maint_ts[r[0]] = r[1].isoformat() if r[1] else None
        except Exception:
            db.session.rollback()

        # 4c) Phase 1: unresolved build failures -> the "failed" tile state
        # (build tried this source and it failed to parse; red with a fix link).
        failures_by_entity = {}
        try:
            rows = db.session.execute(db.text(
                "SELECT DISTINCT entity_code, source_type FROM build_failures "
                "WHERE resolved_at IS NULL"
            )).fetchall()
            for r in rows:
                failures_by_entity.setdefault(r[0], set()).add(r[1])
        except Exception:
            db.session.rollback()

        # 5) FA dir 2026-05-22: latest audit upload per entity for the new
        # "Au" tile on the FA dashboard. Returns id (for click-through to the
        # review page), status (drives green/amber/red), and timestamp.
        # DISTINCT ON keeps the most-recently-touched upload per entity.
        audit_info = {}
        try:
            rows = db.session.execute(db.text("""
                SELECT DISTINCT ON (entity_code)
                  entity_code, id, status, confirmed_at, created_at
                FROM audit_uploads
                ORDER BY entity_code,
                  COALESCE(confirmed_at, created_at) DESC NULLS LAST
            """)).fetchall()
            for r in rows:
                ts = r[3] or r[4]
                audit_info[r[0]] = {
                    "id": r[1],
                    "status": r[2] or "uploaded",
                    "ts": ts.isoformat() if ts else None,
                }
        except Exception:
            db.session.rollback()

        # 6) FA dir 2026-05-22 (Phase 2): SP inventory cache — populated by
        # /api/admin/sp-inventory/scan. Lets the dashboard show an "amber"
        # tile state when the SP file exists but hasn't been ingested yet
        # (vs red, which used to look identical regardless of SP state).
        # Returns {entity_code: {source_type: True/False}} for tile lookup.
        sp_inventory = {}
        sp_modified = {}  # parallel: {ec: {source_type: iso-date}} for tile labels
        try:
            rows = db.session.execute(db.text(
                "SELECT entity_code, source_type, found, file_modified, file_name "
                "FROM sp_inventory"
            )).fetchall()
            for r in rows:
                ec_r, st, found, fmod, fname = r[0], r[1], bool(r[2]), r[3], r[4]
                sp_inventory.setdefault(ec_r, {})[st] = found
                if found and fmod is not None:
                    # FA dir 2026-05-24: surface the SP file_modified timestamp
                    # so amber tiles can show "arrived 5/24" instead of a
                    # silent letter. Stored as ISO so JS Date() parses cleanly.
                    iso = fmod.isoformat() if hasattr(fmod, "isoformat") else str(fmod)
                    sp_modified.setdefault(ec_r, {})[st] = {
                        "modified": iso, "filename": fname or ""
                    }
        except Exception:
            db.session.rollback()

        result = []
        for b in budgets:
            d = b.to_dict()
            d["has_expenses"] = b.entity_code in expense_entities
            d["has_audit"] = b.entity_code in audit_entities
            ec = b.entity_code
            d["timestamps"] = {
                "budget_summary": summary_ts.get(ec),
                "ysl": ysl_ts.get(ec),
                "expense_dist": expense_ts.get(ec),
                "open_ap": open_ap_ts.get(ec),
                "audit": (audit_info.get(ec) or {}).get("ts"),
            }
            # Latest audit info for the Au tile renderer on the dashboard.
            d["audit"] = audit_info.get(ec) or None
            # SP inventory snapshot — drives amber state on dashboard tiles.
            # Keys map to source_type in sp_inventory (ysl, expense_distribution,
            # ap_aging, maint_proof, approved_2026, audit_2025). Renderer
            # checks: ingested → green; else if sp[…] → amber; else red.
            d["sp_inventory"] = sp_inventory.get(ec) or {}
            # FA dir 2026-05-24: parallel dict with file_modified + filename
            # per detected source so the wizard tiles can show arrival dates.
            d["sp_meta"] = sp_modified.get(ec) or {}
            # FA dir 2026-05-23: per-entity readiness tier — single source of
            # truth that both the FA Dashboard and the Wizard's Select Entity
            # page render from. Tiers stack ordered most-ready → least-ready,
            # so a desc sort by tier_order surfaces the FA's actionable queue.
            au_status = (audit_info.get(ec) or {}).get("status") if audit_info.get(ec) else None
            au_id = (audit_info.get(ec) or {}).get("id") if audit_info.get(ec) else None
            sp_e = sp_inventory.get(ec) or {}
            has_bud = bool(summary_ts.get(ec))
            has_exp = ec in expense_entities
            has_ysl = bool(ysl_ts.get(ec))
            has_ap  = bool(open_ap_ts.get(ec))
            audit_confirmed = (au_status == "confirmed")
            audit_extracted = (au_status in ("extracted", "mapped"))
            built = bool(b.wizard_completed_at)
            # Tier classification — checked in order, first match wins
            if built:
                tier = "BUILT"
                # Phase 1 fix (2026-06-09): was 0, colliding with NEEDS_FILES so
                # finished buildings interleaved with empty ones at the bottom of
                # the readiness sort. -1 = own rank: done = least prep-actionable.
                tier_order = -1
                next_action = "review_built"
                next_url = f"/dashboard/{ec}"
                tier_label = "Already built"
            elif audit_confirmed and has_bud and has_exp and has_ysl and has_ap:
                tier = "READY_TO_BUILD"
                tier_order = 5
                next_action = "build"
                next_url = f"/wizard/{ec}?step=2"
                tier_label = "Ready to build"
            elif audit_extracted:
                # Audit extracted but FA hasn't confirmed mapping yet
                tier = "IN_PROGRESS"
                tier_order = 4
                next_action = "audit_review"
                next_url = f"/audited-financials/review/{au_id}" if au_id else f"/wizard/{ec}?step=2"
                tier_label = "Audit review needed"
            elif audit_confirmed:
                # Audit confirmed but some Yardi sources missing
                tier = "IN_PROGRESS"
                tier_order = 3
                next_action = "build"
                next_url = f"/wizard/{ec}?step=2"
                tier_label = "Audit done, sources pending"
            elif (sp_e.get("audit_2025") and not au_status) or au_status in ("uploaded", "extracting"):
                # Label-truth fix (2026-06-10, found via 826/147 Waverly): an
                # audit row at status 'uploaded' (e.g. pulled in by the master-
                # folder scan) means the PDF is HERE and ready to extract — it
                # used to fall through to "Waiting for audit PDF", telling the
                # FA to wait for a file that already arrived.
                tier = "NEEDS_AUDIT_EXTRACT"
                tier_order = 2
                next_action = "audit_review"
                next_url = (f"/audited-financials/review/{au_id}" if au_id
                            else f"/wizard/{ec}?step=2&focus=audit_2025")
                tier_label = "Audit PDF ready to extract"
            elif (sp_e.get("approved_2026") or sp_e.get("ysl") or
                  sp_e.get("expense_distribution") or sp_e.get("ap_aging") or
                  has_bud or has_ysl or has_exp or has_ap):
                # Phase 1 fix (2026-06-09): also check DB-staged data, not just the
                # sp_inventory cache. A building with loaded sources but a stale SP
                # cache used to fall through to NEEDS_FILES ("Chase the FA") —
                # a false alarm manufactured by cache staleness.
                tier = "NEEDS_AUDIT"
                tier_order = 1
                next_action = "wait"
                next_url = f"/wizard/{ec}?step=2"
                tier_label = "Waiting for audit PDF"
            else:
                tier = "NEEDS_FILES"
                tier_order = 0
                next_action = "wait"
                next_url = f"/wizard/{ec}?step=2"
                tier_label = "Waiting for files"
            d["readiness"] = {
                "tier": tier,
                "tier_order": tier_order,
                "tier_label": tier_label,
                "next_action": next_action,
                "next_url": next_url,
            }
            # Phase 1 (2026-06-09): the shared per-source status — Jacob's rule
            # (green = in a BUILT budget; amber = in SharePoint; red = missing/
            # failed; gray = setup). Both the dashboard and the wizard will render
            # tiles from THIS dict, so the two pages can never disagree again.
            d["source_states"] = compute_source_states(
                built=built,
                is_setup=(d.get("lifecycle_stage") == "Setup"),
                staged={
                    "approved_2026": {"loaded": has_bud, "ts": summary_ts.get(ec)},
                    "expense_distribution": {"loaded": has_exp, "ts": expense_ts.get(ec)},
                    "ysl": {"loaded": has_ysl, "ts": ysl_ts.get(ec)},
                    "ap_aging": {"loaded": has_ap, "ts": open_ap_ts.get(ec)},
                    "maint_proof": {"loaded": bool(maint_ts.get(ec)), "ts": maint_ts.get(ec)},
                },
                sp_found=sp_e,
                sp_meta=sp_modified.get(ec) or {},
                audit=audit_info.get(ec),
                failures=failures_by_entity.get(ec) or set(),
            )
            result.append(d)
        return jsonify(result)


    def _re_tax_overrides_for(budget):
        """The ONE override-prep for every compute_re_taxes read site: load
        re_taxes_overrides and backfill missing/zero exemption keys from the
        G&A 6315 lines' current_budget (FA dir 2026-05-18: the RE-tax page
        and the G&A tab must agree on the same dollars). Extracted 2026-07-05
        so the dashboard bootstrap / summary pins / client doc / Excel export
        can't drift from GET /api/re-taxes again."""
        import json as _json
        overrides = None
        if budget and budget.assumptions_json:
            try:
                overrides = _json.loads(budget.assumptions_json).get("re_taxes_overrides")
            except Exception:
                overrides = None
        overrides = dict(overrides or {})
        if budget:
            _GL_TO_OVERRIDE = {
                "6315-0010": "abatement_current",  # Co-op Abatement
                "6315-0020": "star_current",        # STAR
                "6315-0025": "veteran_current",     # Veteran
                "6315-0035": "sche_current",        # SCHE
            }
            for _gl, _key in _GL_TO_OVERRIDE.items():
                _existing = overrides.get(_key)
                if _existing is None or float(_existing or 0) == 0:
                    _line = BudgetLine.query.filter_by(
                        budget_id=budget.id, gl_code=_gl).first()
                    if _line and _line.current_budget:
                        # Exemptions stored negative on G&A; the RE Tax calc
                        # expects positive amounts to subtract.
                        overrides[_key] = abs(float(_line.current_budget))
        return overrides

    def _orphan_gls_for_budget(budget):
        """Budget lines whose GL code is NOT claimed by any Summary row prefix
        but DO carry data (ytd or current_budget). Mirrors the orphan
        computation in /api/admin/summary-debug. Used to gate send-to-PM so a
        building can't ship with material GL data dropped from the budget."""
        import json as _json2
        if not budget:
            return []
        rows = BudgetSummaryRow.query.filter_by(
            entity_code=budget.entity_code, budget_year=BUDGET_YEAR).all()
        prefix_lists = []
        for r in rows:
            try:
                pl = _json2.loads(r.gl_prefixes_json) if r.gl_prefixes_json else []
            except Exception:
                pl = []
            if pl:
                prefix_lists.append(pl)
        orphans = []
        for l in BudgetLine.query.filter_by(budget_id=budget.id).all():
            gl = l.gl_code or ""
            if not gl:
                continue
            # Only P&L / operating GLs belong in the budget Summary (income 4xxx,
            # expenses 5xxx/6xxx, capital 7xxx). Balance-sheet & suspense accounts
            # (0xxx/1xxx/2xxx/3xxx, incl. 0000-0000) are correctly NOT mapped to
            # any operating row — they must NOT count as orphans, or the gate
            # would false-block every building (they all carry large BS balances).
            if gl[:1] not in ("4", "5", "6", "7"):
                continue
            if any(_gl_matches_prefixes(gl, pl) for pl in prefix_lists):
                continue
            ytd = float(l.ytd_actual or 0)
            cb = float(l.current_budget or 0)
            if abs(ytd) > 0.01 or abs(cb) > 0.01:
                orphans.append({"gl_code": gl, "description": l.description or "",
                                "ytd_actual": round(ytd, 2), "current_budget": round(cb, 2)})
        return orphans

    @bp.route("/api/budgets/<entity_code>/status", methods=["POST"])
    def change_budget_status(entity_code):
        """Change budget status with validation using VALID_TRANSITIONS."""
        data = request.get_json()
        new_status = data.get("status")

        if new_status not in BUDGET_STATUSES:
            return jsonify({"error": f"Invalid status. Must be one of {BUDGET_STATUSES}"}), 400

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        # Validate transition using VALID_TRANSITIONS
        allowed = VALID_TRANSITIONS.get(budget.status, [])
        if new_status not in allowed:
            return jsonify({"error": f"Cannot move from '{budget.status}' to '{new_status}'. Allowed: {allowed}"}), 400

        # FA directive 2026-05-11: PM R&M review gate.
        # When the PM tries to submit back to the FA (any → fa_review from
        # a PM-editing state), every "Repairs & Supplies" line on the
        # budget must have an explicit review action (pm_review_state set).
        # NULL = unreviewed = block submit. Returns 422 with the list of
        # unreviewed line IDs so the frontend can scroll/highlight them.
        # Only applies when the SOURCE state is one the PM owns.
        if (new_status == "fa_review"
                and budget.status in ("pm_pending", "pm_in_progress", "returned")):
            unreviewed = BudgetLine.query.filter_by(
                budget_id=budget.id, sheet_name="Repairs & Supplies",
            ).filter(BudgetLine.pm_review_state.is_(None)).all()
            if unreviewed:
                return jsonify({
                    "error": "rm_review_incomplete",
                    "message": (
                        f"{len(unreviewed)} Repairs & Maintenance line(s) "
                        "still need review. Enter a % or $ value, or click "
                        "'No change', on each highlighted row before submitting."
                    ),
                    "unreviewed_line_ids": [l.id for l in unreviewed],
                    "unreviewed_count": len(unreviewed),
                    "total_rm_lines": BudgetLine.query.filter_by(
                        budget_id=budget.id, sheet_name="Repairs & Supplies",
                    ).count(),
                }), 422

        # FA dir 2026-06-04: orphaned-data gate. Block FA → PM ("Send to PM")
        # when material GL data isn't mapped into the Summary — otherwise those
        # dollars are silently dropped from the budget (entity 500 shipped to
        # PM with $514K of orphaned expenses). The FA must map them via the
        # orphan review ("Add to existing row") before handoff. Sub-$500
        # stragglers don't block. Mirrors the R&M review gate above (422).
        if new_status == "pm_pending":
            try:
                _orphans = _orphan_gls_for_budget(budget)
            except Exception:
                _orphans = []
            _material = [o for o in _orphans
                         if abs(o["ytd_actual"]) >= 500 or abs(o["current_budget"]) >= 500]
            if _material:
                _tot = round(sum(o["ytd_actual"] for o in _material), 2)
                # FA dir 2026-06-08: allow an explicit override. The FA may know the
                # orphan GLs are intentional / handled outside the budget and choose
                # to send anyway. Default still BLOCKS (422) so nobody drops data by
                # accident; only an explicit override_orphans=true bypasses — and it
                # is logged (never silent) so the excluded $ is on the audit trail.
                if not data.get("override_orphans"):
                    return jsonify({
                        "error": "orphan_gls_unmapped",
                        "message": (
                            f"{len(_material)} GL account(s) holding ${abs(_tot):,.0f} "
                            "of data aren't mapped to any Summary row, so they'd be "
                            "dropped from the budget. Map them via the orphan review "
                            "(‘Add to existing row’) before sending to the PM."
                        ),
                        "orphan_count": len(_material),
                        "orphan_ytd_total": _tot,
                        "orphan_gls": [o["gl_code"] for o in _material][:50],
                        "can_override": True,
                    }), 422
                try:
                    db.session.add(BudgetRevision(
                        budget_id=budget.id, action="orphan_override",
                        field_name="status", old_value=budget.status, new_value="pm_pending",
                        notes=("OVERRIDE send-to-PM: %d unmapped GL(s) excluded ($%s ytd): %s" % (
                            len(_material), f"{abs(_tot):,.0f}",
                            ", ".join(o["gl_code"] for o in _material)))[:480],
                        source="web", user_id=_read_fa_id_from_cookie(),
                    ))
                except Exception:
                    pass

        if "notes" in data:
            budget.fa_notes = data["notes"]

        if new_status == "approved":
            budget.approved_by = data.get("approved_by", "system")
            budget.approved_at = datetime.utcnow()

        # Stamp pm_sent_at the first time the budget moves to PM. Surfaced on
        # the Building Detail page as "Sent on YYYY-MM-DD" so FAs can see at
        # a glance when handoff happened. Only set on first transition into
        # pm_pending — preserve the original send date if FA bounces it back
        # to PM later.
        if new_status == "pm_pending" and not budget.pm_sent_at:
            budget.pm_sent_at = datetime.utcnow()

        old_status = budget.status
        budget.status = new_status

        # Log status change
        db.session.add(BudgetRevision(
            budget_id=budget.id, action="status_change",
            field_name="status", old_value=old_status, new_value=new_status,
            notes=data.get("notes", ""), source="web",
            user_id=_read_fa_id_from_cookie(),
        ))
        db.session.commit()

        # FA directive 2026-05-10: Teams notification on PM↔FA handoffs.
        # Two trigger points:
        #   draft → pm_pending          (FA → PM)
        #   pm_in_progress → fa_review  (PM → FA)
        # Lazy import to avoid circular dependency (app.py imports workflow).
        # All notification logic lives in app.py and never raises.
        try:
            handoff_event = None
            sender_role, receiver_role = None, None
            if old_status != new_status:
                if new_status == "pm_pending":
                    handoff_event = "fa_to_pm"
                    sender_role, receiver_role = "fa", "pm"
                elif new_status == "fa_review":
                    handoff_event = "pm_to_fa"
                    sender_role, receiver_role = "pm", "fa"
            if handoff_event:
                # Determine the period label for the card if available.
                period_label = None
                try:
                    assum = json.loads(budget.assumptions_json or "{}")
                    bp = (assum.get("budget_period") or "").strip()
                    if bp and "/" in bp:
                        m_int = int(bp.split("/")[0])
                        names = ["Jan","Feb","Mar","Apr","May","Jun",
                                 "Jul","Aug","Sep","Oct","Nov","Dec"]
                        if 1 <= m_int <= 11:
                            period_label = (
                                f"Jan-{names[m_int-1]} actual / "
                                f"{names[m_int]}-Dec estimate"
                            )
                except Exception:
                    pass

                # Lazy import — app.py owns these helpers; we don't want a
                # circular import at module load time.
                try:
                    from app import _post_teams_handoff, _resolve_handoff_actors
                except ImportError:
                    from budget_app.app import _post_teams_handoff, _resolve_handoff_actors

                sender, receivers = _resolve_handoff_actors(
                    entity_code, sender_role, receiver_role
                )
                # Build dashboard URL from request host.
                try:
                    base = request.host_url.rstrip("/")
                except Exception:
                    base = ""
                dashboard_url = f"{base}/dashboard/{entity_code}" if base else None

                _post_teams_handoff(
                    event_type=handoff_event,
                    entity_code=entity_code,
                    sender_user=sender,
                    receiver_users=receivers,
                    building_name=budget.building_name,
                    dashboard_url=dashboard_url,
                    notes=data.get("notes"),
                    period_label=period_label,
                )
        except Exception as _hf_err:
            # Notifications must never break a status change.
            try:
                logger.warning("Teams handoff notify failed for %s: %s",
                               entity_code, _hf_err)
            except Exception:
                pass

        return jsonify(budget.to_dict())


    @bp.route("/api/budgets/<int:budget_id>", methods=["DELETE"])
    def delete_budget(budget_id):
        """Delete a non-approved budget and all its related records.
        Uses raw SQL to avoid ORM session poisoning issues."""
        # Always start with a clean session
        try:
            db.session.rollback()
        except Exception:
            pass

        # Look up budget via raw SQL — immune to session poisoning
        row = db.session.execute(
            db.text("SELECT id, entity_code, status, version FROM budgets WHERE id = :id"),
            {"id": budget_id}
        ).fetchone()
        if not row:
            return jsonify({"error": "Budget not found"}), 404

        bid, entity, status, ver = row[0], row[1], row[2], row[3] or 1
        if status == "approved":
            return jsonify({"error": "Cannot delete an approved budget."}), 400

        try:
            # Get line IDs for this budget
            line_rows = db.session.execute(
                db.text("SELECT id FROM budget_lines WHERE budget_id = :bid"), {"bid": bid}
            ).fetchall()
            line_ids = [r[0] for r in line_rows]

            # Delete in dependency order using raw SQL
            if line_ids:
                ids_str = ",".join(str(i) for i in line_ids)
                db.session.execute(db.text(f"DELETE FROM presentation_edits WHERE budget_line_id IN ({ids_str})"))
                db.session.execute(db.text(f"DELETE FROM budget_revisions WHERE budget_line_id IN ({ids_str})"))
            db.session.execute(db.text("DELETE FROM budget_revisions WHERE budget_id = :bid"), {"bid": bid})
            db.session.execute(db.text("DELETE FROM presentation_sessions WHERE budget_id = :bid"), {"bid": bid})
            db.session.execute(db.text("DELETE FROM ar_handoffs WHERE budget_id = :bid"), {"bid": bid})
            db.session.execute(db.text("DELETE FROM data_sources WHERE budget_id = :bid"), {"bid": bid})
            db.session.execute(db.text("DELETE FROM budget_lines WHERE budget_id = :bid"), {"bid": bid})
            # Wipe entity-level data
            _delete_entity_data(entity)
            # Delete the budget itself
            db.session.execute(db.text("DELETE FROM budgets WHERE id = :bid"), {"bid": bid})
            db.session.commit()
            logger.info(f"Deleted budget {bid} (entity {entity}, v{ver})")
            return jsonify({"message": f"Budget v{ver} for {entity} deleted", "id": bid})
        except Exception as e:
            db.session.rollback()
            logger.error(f"Failed to delete budget {bid}: {e}")
            return jsonify({"error": f"Failed to delete: {str(e)}"}), 500


    @bp.route("/api/dashboard/<entity_code>", methods=["GET"])
    def api_building_detail(entity_code):
        """Get combined budget data for building detail view."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        lines = BudgetLine.query.filter_by(budget_id=budget.id).order_by(BudgetLine.row_num).all()

        # Get assignments
        assignments = BuildingAssignment.query.filter_by(entity_code=entity_code).all()
        fa_name = next((a.user.name for a in assignments if a.role == "fa"), None)
        pm_name = next((a.user.name for a in assignments if a.role == "pm"), None)

        # Check expense report
        expense_data = {"exists": False}
        try:
            row = db.session.execute(
                db.text("SELECT id, period_from, period_to, total_amount FROM expense_reports WHERE entity_code = :ec ORDER BY uploaded_at DESC LIMIT 1"),
                {"ec": entity_code}
            ).fetchone()
            if row:
                invoice_count = db.session.execute(
                    db.text("SELECT COUNT(*) FROM expense_invoices WHERE report_id = :rid"),
                    {"rid": row[0]}
                ).fetchone()[0]
                expense_data = {
                    "exists": True,
                    "period_from": row[1],
                    "period_to": row[2],
                    "total_amount": float(row[3]) if row[3] else 0,
                    "invoice_count": invoice_count
                }
        except Exception:
            pass

        # Check audit data — fetch ALL confirmed uploads for multi-year comparison
        audit_data = {"exists": False, "years": {}, "summary_years": {}}
        try:
            import json as _json
            from budget_app.audited_financials import CENTURY_TO_SUMMARY

            audit_rows = db.session.execute(
                db.text("SELECT mapped_data, fiscal_year_end FROM audit_uploads WHERE entity_code = :ec AND status = 'confirmed' ORDER BY fiscal_year_end DESC"),
                {"ec": entity_code}
            ).fetchall()
            if audit_rows:
                years_data = {}
                summary_years_data = {}
                for row in audit_rows:
                    if not row[0]:
                        continue
                    # JSONB columns come back as dict already; plain JSON/TEXT come back as str
                    mapped = row[0] if isinstance(row[0], dict) else _json.loads(row[0])
                    fiscal_year = row[1] or "Unknown"
                    # Extract year_totals[0] for each category (the primary year)
                    year_cats = {}
                    summary_totals = {}
                    for cat, data in mapped.items():
                        if isinstance(data, dict):
                            totals = data.get("year_totals", data.get("years", []))
                            if totals and len(totals) > 0:
                                year_cats[cat] = totals[0]
                                # Also aggregate to summary row
                                summary_label = CENTURY_TO_SUMMARY.get(cat, cat)
                                summary_totals[summary_label] = summary_totals.get(summary_label, 0) + totals[0]
                            elif data.get("total"):
                                year_cats[cat] = data["total"]
                                summary_label = CENTURY_TO_SUMMARY.get(cat, cat)
                                summary_totals[summary_label] = summary_totals.get(summary_label, 0) + data["total"]
                    if year_cats:
                        years_data[fiscal_year] = year_cats
                    if summary_totals:
                        summary_years_data[fiscal_year] = summary_totals

                # Limit to 2 most recent fiscal years
                years_data_limited = dict(sorted(years_data.items(), reverse=True)[:2])
                summary_years_data_limited = dict(sorted(summary_years_data.items(), reverse=True)[:2])

                if years_data_limited:
                    audit_data = {
                        "exists": True,
                        "years": years_data_limited,
                        "summary_years": summary_years_data_limited,
                        "category_mapping": BUDGET_CAT_TO_CENTURY
                    }
        except Exception:
            pass

        # Group lines by sheet for tabbed view
        sheets = {}
        for l in lines:
            sn = l.sheet_name or "Unmapped"
            if sn not in sheets:
                sheets[sn] = []
            sheets[sn].append(l.to_dict())
        # FA directive 2026-05-05 (#10): sort each sheet's lines by GL code
        # ascending so the Income/Energy/etc tabs render in deterministic
        # GL order (was previously DB-insertion order, which looked random).
        # Sheets that group by sub-category (R&S, Gen&Admin) still respect
        # their grouping because the catConfig.groups iterator runs first
        # in the JS render — within each group, the lines are now sorted.
        for _sn in sheets:
            sheets[_sn].sort(key=lambda d: d.get("gl_code") or "")

        # ── Income forecast pin signal (task #99, FA dir 2026-06-02) ──────────
        # The Budget Summary pins forecast (Col5) to approved budget (Col6) for
        # fully-collectible income families (Maintenance 4010, Common Charges
        # 4020, Commercial CC 4030, Commercial Rent 4040, Operating Assessment
        # 4200) — but ONLY where that row actually has an approved budget. The
        # decision varies per row WITHIN a building (e.g. 500: maintenance has
        # no Col6 so it annualizes, while its operating-assessment row has Col6
        # so it pins). So we can't let the worksheet guess from current_budget;
        # we stamp each line with the Summary's own per-row decision so the
        # Income tab's forecast ties to the Summary instead of annualizing.
        try:
            import json as _json
            # budget_summary_rows is keyed by entity_code + budget_year (NOT
            # budget_id). The dashboard loads budget at year=BUDGET_YEAR, so the
            # covering summary rows live under the same year — mirror /api/summary.
            _summary_rows = BudgetSummaryRow.query.filter_by(
                entity_code=entity_code, budget_year=BUDGET_YEAR).all()
            # GL -> current_budget from the loaded worksheet lines. A row counts
            # as "has a budget" when its GL lines carry one even if the Summary's
            # imported col6 is missing (the two budget sources can disagree —
            # entity 500 maintenance: GL current_budget present, Summary col6
            # None). Without this the forecast pin wouldn't fire and the income
            # tab would annualize instead of matching the approved budget.
            # (Mirrors the /api/summary col6 fallback so tab == Summary.)
            _cb_by_gl = {}
            for _sn in sheets:
                for _ld in sheets[_sn]:
                    _g = _ld.get("gl_code") or ""
                    if _g:
                        _cb_by_gl[_g] = _cb_by_gl.get(_g, 0) + (_ld.get("current_budget") or 0)
            _pinned_prefixes = []
            for _sr in _summary_rows:
                if _sr.row_type != "data" or not _row_has_fixed_forecast_gl(_sr.gl_prefixes_json):
                    continue
                try:
                    _prefs = _json.loads(_sr.gl_prefixes_json) or []
                except Exception:
                    _prefs = []
                _has_budget = _sr.col6_approved_budget is not None
                if not _has_budget:
                    _row_cb = sum(cb for _g, cb in _cb_by_gl.items()
                                  if _g and gl_matches_prefixes(_g, _prefs))
                    _has_budget = abs(_row_cb) > 0.01
                if _has_budget:
                    for _p in _prefs:
                        _b = str(_p).split("-")[0].strip()
                        if _b in FIXED_FORECAST_GL_BASES:
                            _pinned_prefixes.append(_p)
            if _pinned_prefixes:
                for _sn in sheets:
                    for _ld in sheets[_sn]:
                        _glc = _ld.get("gl_code") or ""
                        if _glc and gl_matches_prefixes(_glc, _pinned_prefixes):
                            _ld["income_pinned"] = True
        except Exception:
            pass

        # Ordered sheet tab names
        sheet_order = ["Income", "Payroll", "Energy", "Water & Sewer", "Repairs & Supplies", "Gen & Admin", "RE Taxes", "Capital", "Unmapped"]

        # Parse stored assumptions
        import json as _json
        try:
            assumptions = _json.loads(budget.assumptions_json) if budget.assumptions_json else {}
        except Exception:
            assumptions = {}

        # Derive YTD months from assumptions or default to 2
        ytd_months = 2
        try:
            bp = assumptions.get("budget_period", "")
            if bp and "/" in bp:
                ytd_months = int(bp.split("/")[0])
        except Exception:
            pass
        remaining_months = 12 - ytd_months

        # Fetch RE Taxes data for co-ops
        re_taxes_data = None
        try:
            from dof_taxes import is_coop, compute_re_taxes
            if is_coop(entity_code):
                _rt_ovr = _re_tax_overrides_for(budget)
                re_taxes_data = compute_re_taxes(entity_code, _rt_ovr)
                if re_taxes_data:
                    # Parity with GET /api/re-taxes — the other producer of
                    # window._reTaxesData (saved cell edits + 10/31 toggle).
                    if _rt_ovr.get("cell_overrides"):
                        re_taxes_data["cell_overrides"] = _rt_ovr["cell_overrides"]
                    re_taxes_data["after_oct31"] = bool(_rt_ovr.get("after_oct31"))
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"RE Taxes fetch failed for {entity_code}: {e}")

        # FA #14 (2026-06-16): attach this building's custom RE-tax escalation /
        # adjustment lines (6315-xxxx beyond the 7 fixed GLs) so the RE Taxes tab
        # renders them as extra Section-3 rows on initial load. The page-load
        # bootstrap feeds window._reTaxesData directly, so the key must live here
        # too (not only on GET /api/re-taxes).
        if re_taxes_data:
            try:
                _RE_FIXED_GLS = {
                    "6315-0000", "6315-0010", "6315-0020", "6315-0025",
                    "6315-0030", "6315-0035", "6315-0040",
                }
                _custom = []
                for _cl in BudgetLine.query.filter(
                    BudgetLine.budget_id == budget.id,
                    BudgetLine.gl_code.like("6315-%"),
                ).order_by(BudgetLine.gl_code).all():
                    _gl = (_cl.gl_code or "").strip()
                    if _gl and _gl not in _RE_FIXED_GLS:
                        _custom.append({"gl": _gl, "label": _cl.description or _gl})
                re_taxes_data["custom_gl_rows"] = _custom
            except Exception:
                re_taxes_data.setdefault("custom_gl_rows", [])

        # FA directive 2026-05-05: audit-status chip on dashboard. Surface the
        # latest AuditUpload for this entity so the dashboard can render a
        # progression chip (Uploaded → Extracted → Mapped → Confirmed) without
        # an extra round-trip. Empty when no audit row exists for this entity.
        audit_summary = None
        try:
            row_au = db.session.execute(db.text(
                "SELECT id, fiscal_year_end, status, confirmed_at, confirmed_by, "
                "       updated_at, pdf_filename "
                "FROM audit_uploads "
                "WHERE entity_code = :ec "
                "ORDER BY (CASE status WHEN 'confirmed' THEN 4 WHEN 'mapped' THEN 3 "
                "                       WHEN 'extracted' THEN 2 WHEN 'uploaded' THEN 1 ELSE 0 END) DESC, "
                "         updated_at DESC NULLS LAST "
                "LIMIT 1"
            ), {"ec": entity_code}).fetchone()
            if row_au:
                audit_summary = {
                    "id": row_au[0],
                    "fiscal_year_end": row_au[1],
                    "status": row_au[2],
                    "confirmed_at": row_au[3].isoformat() if row_au[3] else None,
                    "confirmed_by": row_au[4] or "",
                    "updated_at": row_au[5].isoformat() if row_au[5] else None,
                    "pdf_filename": row_au[6] or "",
                    "review_url": f"/audited-financials/review/{row_au[0]}",
                }
        except Exception as _au_err:
            logger.warning(f"audit_summary lookup failed for {entity_code}: {_au_err}")

        return jsonify({
            "budget": budget.to_dict(),
            "building_type": (budget.building_type or ""),  # gates the CAM tab (condo/cond-op)
            "lines": [l.to_dict() for l in lines],
            "sheets": sheets,
            "sheet_order": [s for s in sheet_order if s in sheets or (s == "RE Taxes" and re_taxes_data)],
            "assignments": {"fa": fa_name, "pm": pm_name},
            "expenses": expense_data,
            "audit": audit_data,
            "audit_summary": audit_summary,
            "assumptions": assumptions,
            "ytd_months": ytd_months,
            "remaining_months": remaining_months,
            "re_taxes": re_taxes_data
        })


    # ─── API Routes: Budget Assumptions ──────────────────────────────────────

    @bp.route("/api/budget-assumptions/<entity_code>", methods=["GET"])
    def get_budget_assumptions(entity_code):
        """Get assumptions for a budget."""
        import json as _json
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        try:
            assumptions = _json.loads(budget.assumptions_json) if budget.assumptions_json else {}
        except Exception:
            assumptions = {}
        return jsonify(assumptions)

    @bp.route("/api/budget-assumptions/<entity_code>", methods=["PUT"])
    def update_budget_assumptions(entity_code):
        """Update assumptions for a budget and recalculate affected lines."""
        import json as _json
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        try:
            data = request.get_json(silent=True) or {}
        except Exception:
            return jsonify({"error": "Invalid JSON"}), 400
        if not data:
            return jsonify({"error": "No data provided"}), 400

        # Merge updates into existing assumptions
        try:
            current = _json.loads(budget.assumptions_json) if budget.assumptions_json else {}
        except Exception:
            current = {}

        # Deep merge: update each section
        for key, value in data.items():
            if isinstance(value, dict) and isinstance(current.get(key), dict):
                current[key].update(value)
            else:
                current[key] = value

        budget.assumptions_json = _json.dumps(current)

        # Log assumption changes
        for key, value in data.items():
            db.session.add(BudgetRevision(
                budget_id=budget.id, action="update",
                field_name="assumptions." + key,
                old_value="", new_value=_json.dumps(value) if isinstance(value, dict) else str(value),
                source="web",
                user_id=_read_fa_id_from_cookie(),
            ))

        # Recalculate affected BudgetLine increase_pct based on assumption changes
        changed_sections = list(data.keys())
        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
        recalc_count = 0

        # Map assumption sections to sheet names
        section_to_sheet = {
            "energy": "Energy",
            "water_sewer": "Water & Sewer",
            "insurance_renewal": "Insurance",  # Insurance lines are typically in Gen & Admin
        }

        for section in changed_sections:
            if section == "energy":
                # Apply energy rate increases to Energy sheet lines
                gas_inc = float(current.get("energy", {}).get("gas_rate_increase", 0) or 0)
                elec_inc = float(current.get("energy", {}).get("electric_rate_increase", 0) or 0)
                oil_inc = float(current.get("energy", {}).get("oil_rate_increase", 0) or 0)
                # Use the average of non-zero rates as default, or gas rate
                default_rate = gas_inc or elec_inc or oil_inc
                for line in lines:
                    if line.sheet_name == "Energy":
                        gl = line.gl_code or ""
                        # Gas GLs typically start with 5105, Electric with 5110, Oil with 5115
                        if "5105" in gl or "gas" in (line.description or "").lower():
                            line.increase_pct = gas_inc
                        elif "5110" in gl or "electric" in (line.description or "").lower():
                            line.increase_pct = elec_inc
                        elif "5115" in gl or "oil" in (line.description or "").lower() or "fuel" in (line.description or "").lower():
                            line.increase_pct = oil_inc
                        else:
                            line.increase_pct = default_rate
                        recalc_count += 1

            elif section == "water_sewer":
                # Apply water rate increase to Water & Sewer sheet lines
                water_inc = float(current.get("water_sewer", {}).get("rate_increase", 0) or 0)
                for line in lines:
                    if line.sheet_name == "Water & Sewer":
                        line.increase_pct = water_inc
                        recalc_count += 1

            elif section == "insurance_renewal":
                # Apply insurance renewal increase to insurance GL codes (6105-6195)
                ins_inc = float(current.get("insurance_renewal", {}).get("increase_percent", 0) or 0)
                for line in lines:
                    gl = line.gl_code or ""
                    if gl.startswith("61") and gl < "6200-0000":
                        line.increase_pct = ins_inc
                        recalc_count += 1

            elif section == "wage_increase":
                # Apply wage increase to all payroll lines
                wage_inc = float(current.get("wage_increase", {}).get("percent", 0) or 0)
                for line in lines:
                    if line.sheet_name == "Payroll":
                        line.increase_pct = wage_inc
                        recalc_count += 1

        # Derive YTD months from budget period assumption
        _ytd_months = 2
        try:
            bp = current.get("budget_period", "")
            if bp and "/" in bp:
                _ytd_months = int(bp.split("/")[0])
        except Exception:
            pass
        _remaining = 12 - _ytd_months

        # If budget_period changed, EVERY line's forecast/proposed needs to
        # recompute (estimate depends on _ytd_months). Otherwise only lines
        # touched by an assumption with increase_pct need to recompute.
        period_changed = "budget_period" in data
        for line in lines:
            if line.increase_pct or period_changed:
                ytd = float(line.ytd_actual or 0)
                accrual = float(line.accrual_adj or 0)
                unpaid = float(line.unpaid_bills or 0)
                prior = float(line.prior_year or 0)
                base = ytd + accrual + unpaid
                _is_cap = (line.sheet_name == "Capital"
                           or (line.category or "").lower() == "capital")
                # One-time fee rule: once YTD posted, no projection. Forecast = billed amount.
                if (line.gl_code or "") in ONE_TIME_FEE_GLS and abs(base) > 0.01:
                    estimate = 0
                # FA #18: Capital — never extrapolate, never auto-fill proposed
                elif _is_cap:
                    estimate = 0
                # 210 FA: RE-tax credit income (4105/4110/4115/4120/4125) — no
                # May-Dec estimate (posts at year-end, not monthly).
                elif (line.gl_code or "")[:4] in ("4105", "4110", "4115", "4120", "4125"):
                    estimate = 0
                # FA #7 anomaly cap: don't extrapolate one-time refund/credit.
                # Recurring negatives (tax abatements, where prior is also
                # negative) keep extrapolating normally.
                elif base < 0 and prior >= 0:
                    estimate = 0
                else:
                    estimate = (base / _ytd_months) * _remaining if _ytd_months > 0 else 0
                forecast = base + estimate
                # FA #18: don't auto-fill capital proposed_budget; leave it as-is
                # (FA can manually enter; auto-formula would predict spend that
                # shouldn't recur).
                if not _is_cap:
                    line.proposed_budget = forecast * (1 + float(line.increase_pct or 0))

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            logger.error(f"update_budget_assumptions failed for {entity_code}: {e}", exc_info=True)
            return jsonify({"error": "Failed to save assumptions"}), 500
        logger.info(f"Assumptions updated for {entity_code}, recalculated {recalc_count} lines")

        return jsonify({"status": "saved", "assumptions": current, "recalculated": recalc_count})


    # ─── API Routes: Source Uploads Status ──────────────────────────────────

    @bp.route("/api/entity/<entity_code>/sources", methods=["GET"])
    def get_entity_sources(entity_code):
        """Return last-uploaded timestamps + filenames for all 5 budget sources.

        Sources: YSL (Yardi), Expense Distribution (Yardi), AP Aging (Yardi),
        Maintenance Proof (Yardi), Audited Financials (PDF).
        """
        result = {
            "ysl": {"last_uploaded": None, "filename": None},
            "expense_distribution": {"last_uploaded": None, "filename": None},
            "ap_aging": {"last_uploaded": None, "filename": None},
            "maint_proof": {"last_uploaded": None, "filename": None},
            "audited_financials": {"last_uploaded": None, "filename": None},
        }

        # 1) YSL — proxy via earliest budget_lines.updated_at for this entity
        try:
            row = db.session.execute(
                db.text("""
                    SELECT MIN(bl.updated_at)
                    FROM budget_lines bl
                    JOIN budgets b ON b.id = bl.budget_id
                    WHERE b.entity_code = :ec
                """),
                {"ec": entity_code}
            ).fetchone()
            if row and row[0]:
                result["ysl"]["last_uploaded"] = row[0].isoformat()
        except Exception:
            db.session.rollback()

        # 2) Expense Distribution
        try:
            row = db.session.execute(
                db.text("SELECT file_name, uploaded_at FROM expense_reports WHERE entity_code = :ec ORDER BY uploaded_at DESC LIMIT 1"),
                {"ec": entity_code}
            ).fetchone()
            if row:
                result["expense_distribution"]["filename"] = row[0]
                result["expense_distribution"]["last_uploaded"] = row[1].isoformat() if row[1] else None
        except Exception:
            db.session.rollback()

        # 3) AP Aging
        try:
            row = db.session.execute(
                db.text("SELECT file_name, uploaded_at FROM open_ap_reports WHERE entity_code = :ec ORDER BY uploaded_at DESC LIMIT 1"),
                {"ec": entity_code}
            ).fetchone()
            if row:
                result["ap_aging"]["filename"] = row[0]
                result["ap_aging"]["last_uploaded"] = row[1].isoformat() if row[1] else None
        except Exception:
            db.session.rollback()

        # 4) Maintenance Proof
        try:
            row = db.session.execute(
                db.text("SELECT file_name, uploaded_at FROM maint_proof_reports WHERE entity_code = :ec ORDER BY uploaded_at DESC LIMIT 1"),
                {"ec": entity_code}
            ).fetchone()
            if row:
                result["maint_proof"]["filename"] = row[0]
                result["maint_proof"]["last_uploaded"] = row[1].isoformat() if row[1] else None
        except Exception:
            db.session.rollback()

        # 5) Audited Financials (confirmed only)
        try:
            row = db.session.execute(
                db.text("SELECT pdf_filename, created_at FROM audit_uploads WHERE entity_code = :ec AND status = 'confirmed' ORDER BY created_at DESC LIMIT 1"),
                {"ec": entity_code}
            ).fetchone()
            if row:
                result["audited_financials"]["filename"] = row[0]
                result["audited_financials"]["last_uploaded"] = row[1].isoformat() if row[1] else None
        except Exception:
            db.session.rollback()

        return jsonify(result)


    # ─── API Routes: YSL Merge Mode (BETA — 204/212 only) ───────────────────
    #
    # Partial re-upload of a YSL file that preserves all user edits (FA/PM).
    # Only refreshes Yardi-sourced columns: prior_year, ytd_actual, ytd_budget,
    # current_budget, plus the mapping metadata (sheet_name, description,
    # category, pm_editable). All user-edit columns are left untouched.
    #
    # Safety layers:
    #   1. Hardcoded allowlist — currently {"204", "212"}. All other entities
    #      return 403.
    #   2. Pre-merge snapshot — before any writes, serialize all BudgetLines
    #      for the entity to budget.pre_merge_snapshot (TEXT/JSON). Restore
    #      endpoint reads this to revert.
    #   3. Atomic transaction — single commit; any exception rolls back.
    #   4. Dry-run mode — ?mode=dry_run (or form mode=dry_run) returns the
    #      diff without writing. UI previews this before committing.
    #   5. Restore endpoint — /api/ysl/restore/<entity_code> rewinds the
    #      most recent merge from snapshot.
    # 724 added 2026-08-18 (Jacob: "go ahead with the ysl merge for 724"):
    # its FA re-uploaded a July YSL that the wizard never re-ingested (FA
    # note #5) - the merge is the sanctioned refresh path for a built
    # budget (preserves user edits, snapshots, writes BudgetRevisions).
    _YSL_MERGE_ALLOWLIST = {"204", "212", "724"}
    # Commits enabled 2026-08-18 with Jacob's explicit OK: dry-run stays the
    # default mode; a live commit still requires mode=commit explicitly.
    _YSL_MERGE_COMMIT_ENABLED = True
    _YSL_MERGE_REFRESH_FIELDS = (
        "prior_year", "ytd_actual", "ytd_budget", "current_budget",
    )
    _YSL_MERGE_META_FIELDS = (
        "sheet_name", "description", "category", "pm_editable",
    )


    def _ysl_merge_compute_diff(entity_code, gl_data, gl_mapping):
        """
        Given parsed YSL gl_data and the template mapping, compute the set of
        changes that would be applied. Returns a dict:
          {
            "updated": [ {gl_code, description, changes: [{field, old, new}]} ],
            "inserted": [ {gl_code, description} ],
            "orphaned": [ {gl_code, description} ],  # in DB but not in new YSL
            "totals": {"updated": n, "inserted": n, "orphaned": n, "total_gls_in_file": n}
          }
        Does NOT write anything.
        """
        try:
            from gl_mapper import build_gl_mapping_with_descriptions  # noqa: F401
        except ImportError:
            pass  # gl_mapping is already passed in, but import just in case

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return {
                "updated": [], "inserted": [], "orphaned": [],
                "totals": {"updated": 0, "inserted": 0, "orphaned": 0, "total_gls_in_file": len(gl_data)},
                "note": "No existing budget for this entity/year — would create from scratch."
            }

        # Index existing lines by gl_code
        existing = {ln.gl_code: ln for ln in BudgetLine.query.filter_by(budget_id=budget.id).all()}

        updated, inserted, orphaned = [], [], []
        new_gl_codes = set(gl_data.keys())

        for gl_code, gl_values in gl_data.items():
            prior_year = float(gl_values.get("period_2", 0) or 0)
            ytd_actual = float(gl_values.get("period_3", 0) or 0)
            ytd_budget = float(gl_values.get("period_4", 0) or 0)
            current_budget = float(gl_values.get("period_5", 0) or 0)

            if gl_code in existing:
                line = existing[gl_code]
                changes = []
                if round(float(line.prior_year or 0), 2) != round(prior_year, 2):
                    changes.append({"field": "prior_year", "old": float(line.prior_year or 0), "new": prior_year})
                if round(float(line.ytd_actual or 0), 2) != round(ytd_actual, 2):
                    changes.append({"field": "ytd_actual", "old": float(line.ytd_actual or 0), "new": ytd_actual})
                if round(float(line.ytd_budget or 0), 2) != round(ytd_budget, 2):
                    changes.append({"field": "ytd_budget", "old": float(line.ytd_budget or 0), "new": ytd_budget})
                if round(float(line.current_budget or 0), 2) != round(current_budget, 2):
                    changes.append({"field": "current_budget", "old": float(line.current_budget or 0), "new": current_budget})
                if changes:
                    updated.append({
                        "gl_code": gl_code,
                        "description": line.description,
                        "changes": changes,
                    })
            else:
                # New GL — would be inserted
                desc = gl_mapping.get(gl_code, (None, None, gl_code))[2] if gl_code in gl_mapping else gl_code
                inserted.append({"gl_code": gl_code, "description": desc})

        for gl_code, line in existing.items():
            if gl_code not in new_gl_codes:
                orphaned.append({"gl_code": gl_code, "description": line.description})

        return {
            "updated": updated,
            "inserted": inserted,
            "orphaned": orphaned,
            "totals": {
                "updated": len(updated),
                "inserted": len(inserted),
                "orphaned": len(orphaned),
                "total_gls_in_file": len(gl_data),
            },
        }


    def _ysl_merge_apply(entity_code, gl_data, gl_mapping):
        """
        Apply the merge. Snapshots existing lines to budget.pre_merge_snapshot,
        then refreshes Yardi fields + metadata, inserts new GLs. User-edit
        columns are never touched. Orphaned GLs (in DB, not in file) are
        left alone. Writes BudgetRevision entries for changed fields.
        Returns (success, summary_dict).
        """
        import json as _json_mod

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return False, {"error": "No existing budget. Use the standard upload flow first."}

        # 1) Snapshot every existing line (full to_dict) for restore
        snapshot = [ln.to_dict() for ln in BudgetLine.query.filter_by(budget_id=budget.id).all()]
        budget.pre_merge_snapshot = _json_mod.dumps(snapshot)
        budget.pre_merge_snapshot_at = datetime.utcnow()

        # 2) Index existing by gl_code
        existing = {ln.gl_code: ln for ln in BudgetLine.query.filter_by(budget_id=budget.id).all()}

        # 3) Resolve mapping helpers (reusing store_all_lines logic shape)
        try:
            from gl_mapper import build_gl_mapping_with_descriptions  # noqa: F401
        except ImportError:
            pass

        updated_ct, inserted_ct, revisions_ct = 0, 0, 0

        for gl_code, gl_values in gl_data.items():
            prior_year = float(gl_values.get("period_2", 0) or 0)
            ytd_actual = float(gl_values.get("period_3", 0) or 0)
            ytd_budget = float(gl_values.get("period_4", 0) or 0)
            current_budget = float(gl_values.get("period_5", 0) or 0)

            # Determine sheet/row/description/category/pm_editable using same
            # rules as store_all_lines (keep it consistent — Yardi metadata)
            if gl_code in RM_GL_MAP:
                desc, row_num, category = RM_GL_MAP[gl_code]
                sheet_name = "Repairs & Supplies"
                pm_editable = True
            elif gl_code in gl_mapping:
                sheet_name, row_num, desc = gl_mapping[gl_code]
                category = SHEET_TO_CATEGORY.get(sheet_name, "other")
                if category == "rm":
                    _csv_hit = GL_MAPPING_CSV.get(gl_code[:4])
                    category = _csv_hit[2] if _csv_hit else "repairs"
                pm_editable = False
            elif gl_code.startswith("7"):
                prefix = gl_code[:4]
                desc = CAPITAL_GL_PREFIX.get(prefix, f"Cap - {prefix}")
                sheet_name = "Capital"
                row_num = 0
                category = "capital"
                pm_editable = True
            else:
                _csv_hit = GL_MAPPING_CSV.get(gl_code[:4])
                if _csv_hit:
                    desc, sheet_name, category = _csv_hit
                    row_num = 0
                    pm_editable = True
                else:
                    desc = gl_code
                    sheet_name = "Unmapped"
                    row_num = 0
                    category = "other"
                    pm_editable = False

            if gl_code in existing:
                line = existing[gl_code]
                # Track changes to refresh fields for audit trail
                for fname, new_val in (
                    ("prior_year", prior_year),
                    ("ytd_actual", ytd_actual),
                    ("ytd_budget", ytd_budget),
                    ("current_budget", current_budget),
                ):
                    old_val = float(getattr(line, fname) or 0)
                    if round(old_val, 2) != round(new_val, 2):
                        db.session.add(BudgetRevision(
                            budget_id=budget.id,
                            budget_line_id=line.id,
                            action="update",
                            field_name=fname,
                            old_value=str(old_val),
                            new_value=str(new_val),
                            source="ysl_merge",
                            notes="YSL merge (beta)",
                        ))
                        revisions_ct += 1
                # Refresh Yardi-sourced fields only — preserve user edits
                line.prior_year = prior_year
                line.ytd_actual = ytd_actual
                line.ytd_budget = ytd_budget
                line.current_budget = current_budget
                # Refresh mapping metadata (Yardi-derived, not a user edit)
                line.sheet_name = sheet_name
                line.description = desc
                line.category = category
                line.pm_editable = pm_editable
                updated_ct += 1
            else:
                line = BudgetLine(
                    budget_id=budget.id,
                    gl_code=gl_code,
                    description=desc,
                    category=category,
                    row_num=row_num,
                    sheet_name=sheet_name,
                    pm_editable=pm_editable,
                    prior_year=prior_year,
                    ytd_actual=ytd_actual,
                    ytd_budget=ytd_budget,
                    current_budget=current_budget,
                )
                db.session.add(line)
                db.session.flush()
                db.session.add(BudgetRevision(
                    budget_id=budget.id,
                    budget_line_id=line.id,
                    action="create",
                    field_name="gl_code",
                    old_value="",
                    new_value=gl_code,
                    source="ysl_merge",
                    notes="YSL merge (beta): new GL inserted",
                ))
                inserted_ct += 1
                revisions_ct += 1

        # Top-level audit row
        db.session.add(BudgetRevision(
            budget_id=budget.id,
            budget_line_id=None,
            action="ysl_merge",
            field_name="__merge__",
            old_value="",
            new_value=f"updated={updated_ct} inserted={inserted_ct}",
            source="ysl_merge",
            notes="YSL merge (beta) applied; snapshot saved",
        ))

        db.session.commit()

        return True, {
            "updated": updated_ct,
            "inserted": inserted_ct,
            "revisions": revisions_ct,
            "snapshot_at": budget.pre_merge_snapshot_at.isoformat() if budget.pre_merge_snapshot_at else None,
        }


    @bp.route("/api/ysl/merge/<entity_code>", methods=["POST"])
    def ysl_merge(entity_code):
        """
        Merge a new YSL file into an existing budget, preserving user edits.

        Form fields:
          file: the YSL .xlsx (required)
          mode: 'dry_run' (default) | 'commit'

        Dry-run returns the diff only. Commit applies + snapshots + writes
        BudgetRevision entries.
        """
        if str(entity_code) not in _YSL_MERGE_ALLOWLIST:
            return jsonify({
                "error": f"YSL merge (beta) is only enabled for entities: {sorted(_YSL_MERGE_ALLOWLIST)}.",
                "entity_code": entity_code,
            }), 403

        mode = (request.form.get("mode") or request.args.get("mode") or "dry_run").lower()
        if mode not in ("dry_run", "commit"):
            return jsonify({"error": "mode must be 'dry_run' or 'commit'"}), 400

        f = request.files.get("file")
        if not f or not f.filename:
            return jsonify({"error": "Upload a .xlsx file in the 'file' field"}), 400
        if not f.filename.lower().endswith((".xlsx", ".xls")):
            return jsonify({"error": "File must be .xlsx or .xls"}), 400

        # Parse YSL
        import tempfile
        from pathlib import Path as _Path
        try:
            from ysl_parser import parse_ysl_file
        except ImportError:
            from budget_system.ysl_parser import parse_ysl_file

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = _Path(tmpdir) / f.filename
            f.save(str(tmp_path))
            try:
                gl_data, property_info = parse_ysl_file(tmp_path)
            except Exception as e:
                logger.exception("YSL merge: parse failed")
                return jsonify({"error": f"Failed to parse YSL file: {e}"}), 400

            # Sanity-check: file's property_code must match target entity
            file_entity = str(property_info.get("property_code") or "").strip()
            if file_entity and file_entity != str(entity_code):
                return jsonify({
                    "error": f"Entity mismatch: URL=/{entity_code}/ but file header says {file_entity}. Refusing to merge.",
                }), 400

            # Build gl_mapping for description resolution
            try:
                from gl_mapper import build_gl_mapping_with_descriptions
            except ImportError:
                from budget_system.gl_mapper import build_gl_mapping_with_descriptions
            template_path = _Path(__file__).parent.parent / "budget_system" / "Budget_Final_Template_v2.xlsx"
            try:
                gl_mapping = build_gl_mapping_with_descriptions(template_path)
            except Exception:
                gl_mapping = {}

            if mode == "dry_run":
                diff = _ysl_merge_compute_diff(entity_code, gl_data, gl_mapping)
                return jsonify({
                    "mode": "dry_run",
                    "entity_code": entity_code,
                    "filename": f.filename,
                    "property_info": property_info,
                    "commit_enabled": bool(_YSL_MERGE_COMMIT_ENABLED),
                    "diff": diff,
                })

            # Commit gate — beta preview-only until explicitly flipped on
            if not _YSL_MERGE_COMMIT_ENABLED:
                return jsonify({
                    "error": "YSL merge is in preview-only mode. Commit is disabled. Use mode=dry_run to preview the diff.",
                    "commit_enabled": False,
                }), 403

            # Commit path — atomic
            try:
                ok, summary = _ysl_merge_apply(entity_code, gl_data, gl_mapping)
                if not ok:
                    return jsonify({"error": summary.get("error", "Merge failed")}), 400
                return jsonify({
                    "mode": "commit",
                    "entity_code": entity_code,
                    "filename": f.filename,
                    "summary": summary,
                })
            except Exception as e:
                db.session.rollback()
                logger.exception("YSL merge commit failed")
                return jsonify({"error": f"Merge failed, rolled back: {e}"}), 500


    @bp.route("/api/ysl/restore/<entity_code>", methods=["POST"])
    def ysl_restore(entity_code):
        """
        Restore BudgetLines from the most recent pre_merge_snapshot.
        Only restores the Yardi-sourced fields (prior_year, ytd_actual,
        ytd_budget, current_budget) — user edits made after the merge are
        preserved. To fully rewind, restore all fields: use ?full=1.
        """
        if str(entity_code) not in _YSL_MERGE_ALLOWLIST:
            return jsonify({"error": "Restore only enabled for allowlisted entities."}), 403
        if not _YSL_MERGE_COMMIT_ENABLED:
            return jsonify({"error": "Restore is disabled in preview-only mode."}), 403

        import json as _json_mod
        full = (request.args.get("full") or "").strip() == "1"

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget or not budget.pre_merge_snapshot:
            return jsonify({"error": "No snapshot available for this entity."}), 404

        try:
            snapshot = _json_mod.loads(budget.pre_merge_snapshot)
        except Exception as e:
            return jsonify({"error": f"Snapshot unreadable: {e}"}), 500

        snap_by_gl = {row["gl_code"]: row for row in snapshot if row.get("gl_code")}
        existing = {ln.gl_code: ln for ln in BudgetLine.query.filter_by(budget_id=budget.id).all()}

        restored_ct = 0
        try:
            for gl_code, snap in snap_by_gl.items():
                line = existing.get(gl_code)
                if not line:
                    continue  # Line was inserted by merge but later deleted; skip
                # Always restore Yardi fields
                line.prior_year = float(snap.get("prior_year") or 0)
                line.ytd_actual = float(snap.get("ytd_actual") or 0)
                line.ytd_budget = float(snap.get("ytd_budget") or 0)
                line.current_budget = float(snap.get("current_budget") or 0)
                if full:
                    # Restore user-edit fields too (full rewind)
                    line.accrual_adj = float(snap.get("accrual_adj") or 0)
                    line.unpaid_bills = float(snap.get("unpaid_bills") or 0)
                    line.increase_pct = float(snap.get("increase_pct") or 0)
                    line.notes = snap.get("notes") or ""
                    line.reclass_to_gl = snap.get("reclass_to_gl")
                    line.reclass_amount = float(snap.get("reclass_amount") or 0)
                    line.reclass_notes = snap.get("reclass_notes") or ""
                    line.estimate_override = snap.get("estimate_override")
                    line.forecast_override = snap.get("forecast_override")
                    line.proposed_budget = float(snap.get("proposed_budget") or 0)
                    line.proposed_formula = snap.get("proposed_formula")
                    line.fa_proposed_status = snap.get("fa_proposed_status")
                    line.fa_proposed_note = snap.get("fa_proposed_note") or ""
                    line.fa_override_value = snap.get("fa_override_value")
                restored_ct += 1

            db.session.add(BudgetRevision(
                budget_id=budget.id,
                budget_line_id=None,
                action="ysl_restore",
                field_name="__restore__",
                old_value="",
                new_value=f"restored={restored_ct} full={full}",
                source="ysl_merge",
                notes="YSL restore from pre_merge_snapshot",
            ))
            db.session.commit()
            return jsonify({
                "entity_code": entity_code,
                "restored": restored_ct,
                "full": full,
                "snapshot_at": budget.pre_merge_snapshot_at.isoformat() if budget.pre_merge_snapshot_at else None,
            })
        except Exception as e:
            db.session.rollback()
            logger.exception("YSL restore failed")
            return jsonify({"error": f"Restore failed, rolled back: {e}"}), 500


    @bp.route("/api/ysl/snapshot/<entity_code>", methods=["GET"])
    def ysl_snapshot_info(entity_code):
        """Lightweight: does a snapshot exist? when was it taken?"""
        if str(entity_code) not in _YSL_MERGE_ALLOWLIST:
            return jsonify({"allowlisted": False})
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"allowlisted": True, "has_snapshot": False})
        return jsonify({
            "allowlisted": True,
            "has_snapshot": bool(budget.pre_merge_snapshot),
            "snapshot_at": budget.pre_merge_snapshot_at.isoformat() if budget.pre_merge_snapshot_at else None,
        })


    # ─── API Routes: Budget Wizard ────────────────────────────────────────────
    #
    # Guided 6-step budget creation flow:
    #   1. Select entity   2. Upload sources   3. Review portfolio assumptions
    #   4. Set building assumptions   5. Preview & generate   6. Open dashboard
    #
    # Gate mode: /wizard/<entity_code> (full page, first-time)
    # Sidebar mode: embedded in dashboard (re-entry after wizard_completed_at is set)

    @bp.route("/wizard", methods=["GET"])
    @bp.route("/wizard/<entity_code>", methods=["GET"])
    def wizard_page(entity_code=None):
        """Render the Budget Wizard gate page.

        Status UX Phase 4 (2026-06-09): the dashboard is the ONE portfolio home.
        Bare /wizard (no building) redirects there — the wizard's duplicate
        Select Entity grid is retired; the wizard is the per-building flow only
        (/wizard/<entity_code>, steps 2-5). Route stays registered so old
        bookmarks land somewhere sensible instead of 404ing.
        """
        if not entity_code:
            from flask import redirect as _redirect
            return _redirect("/dashboard")
        # FA dir 2026-05-23: defensive diagnostic — if anything in the
        # render fails, surface the real error instead of the generic
        # "Something went wrong" page. Remove once /wizard is stable.
        try:
            return _wizard_page_impl(entity_code)
        except Exception as _wiz_exc:
            import traceback as _tb
            tb_text = _tb.format_exc()
            print(f"[wizard_page] EXCEPTION: {_wiz_exc}\n{tb_text}", flush=True)
            from flask import make_response
            html = (
                "<h1>Wizard render error (diagnostic)</h1>"
                f"<p><b>{type(_wiz_exc).__name__}:</b> {str(_wiz_exc)[:500]}</p>"
                f"<pre style='background:#f4f1eb;padding:12px;overflow:auto;font-size:11px;'>{tb_text[:4000]}</pre>"
            )
            resp = make_response(html, 500)
            resp.headers["Content-Type"] = "text/html; charset=utf-8"
            return resp

    def _wizard_page_impl(entity_code=None):
        """Render the Budget Wizard gate page."""
        import json as json_mod
        from wizard_template import WIZARD_TEMPLATE
        # FA dir 2026-05-23: explicit package-qualified import. Bare
        # `from app import X` was ambiguous because budget_summary/app.py
        # exists and could shadow budget_app/app.py once budget_summary
        # got added to sys.path by a build-budget run.
        try:
            from budget_app.app import load_portfolio_defaults, load_building_assumptions, _get_monday_status
        except ImportError:
            # Fallback for legacy import path (running budget_app/ directly)
            from app import load_portfolio_defaults, load_building_assumptions, _get_monday_status

        # Bug #4: support querystring form ?entity=<code> in addition to
        # /wizard/<entity_code>. FAs commonly bookmark or paste links with
        # ?entity=168&step=3 — without this fallback the wizard lands on
        # Step 1 with "no entity selected" and the FA has to re-search.
        if not entity_code:
            entity_code = (request.args.get("entity") or "").strip() or None

        # Read-only snapshot of last Monday sync. Auto-sync (if stale) is
        # triggered by the client after the page renders — keeps the wizard
        # fast and avoids cross-blueprint app-context issues.
        monday_status = _get_monday_status()

        budgets_db = Budget.query.filter_by(year=BUDGET_YEAR).all()

        # Perf (2026-06-11 audit): one grouped query for "which budgets have
        # lines", instead of bool(b.lines) inside the loop below — that
        # lazy-loaded every one of ~150 buildings' full line sets (~30K rows
        # materialized) per /wizard view. budget_lines(budget_id) is the lead
        # column of ix_budget_lines_review_state, so this is index-only.
        _ids_with_lines = {
            r[0] for r in db.session.execute(db.text(
                "SELECT DISTINCT budget_id FROM budget_lines"
            )).fetchall()
        }

        # Build entity list from existing budgets in DB (not CSV)
        entity_list = []
        for b in budgets_db:
            entity_list.append({
                "entity_code": b.entity_code,
                "building_name": b.building_name or b.entity_code,
                "address": "",
                "wizard_step": b.wizard_step or 0,
                "wizard_completed_at": b.wizard_completed_at.isoformat() if b.wizard_completed_at else None,
                "status": b.status or "not_started",
                "has_lines": b.id in _ids_with_lines,
            })

        # Load FA users and building assignments for the selector.
        # Only include FAs who have at least one assignment in an entity that
        # has a Budget row for the current year (i.e., currently in the active
        # Monday list, post-auto-create). Stale FAs from pre-filter syncs are
        # excluded from the dropdown without deleting their data.
        fa_users = []
        assignments = []
        try:
            active_entities = {b["entity_code"] for b in entity_list}
            assign_rows = (
                BuildingAssignment.query
                .filter_by(role="fa")
                .filter(BuildingAssignment.entity_code.in_(active_entities))
                .all()
                if active_entities else []
            )
            assignments = [a.to_dict() for a in assign_rows]
            active_user_ids = {a["user_id"] for a in assignments}
            if active_user_ids:
                fa_rows = (
                    User.query
                    .filter_by(role="fa")
                    .filter(User.id.in_(active_user_ids))
                    .order_by(User.name)
                    .all()
                )
                fa_users = [u.to_dict() for u in fa_rows]
        except Exception:
            db.session.rollback()

        # Load assumptions for the selected entity
        portfolio = load_portfolio_defaults()
        building_overrides = {}
        sources = {}
        if entity_code:
            all_bldg = load_building_assumptions()
            building_overrides = all_bldg.get(entity_code, {})
            try:
                sources = _get_sources_dict(entity_code)
            except Exception:
                sources = {}

        return render_template_string(
            WIZARD_TEMPLATE,
            entity_code=entity_code or "",
            budget_year=BUDGET_YEAR,
            budgets_json=json_mod.dumps(entity_list),
            portfolio_json=json_mod.dumps(portfolio),
            building_json=json_mod.dumps(building_overrides),
            sources_json=json_mod.dumps(sources),
            fa_users_json=json_mod.dumps(fa_users),
            assignments_json=json_mod.dumps(assignments),
            monday_status_json=json_mod.dumps(monday_status),
        )


    def _get_sources_dict(entity_code):
        """Internal helper — returns source upload status dict for an entity."""
        result = {
            "ysl": {"uploaded": False, "last_uploaded": None},
            "expense_distribution": {"uploaded": False, "last_uploaded": None},
            "ap_aging": {"uploaded": False, "last_uploaded": None},
            "maint_proof": {"uploaded": False, "last_uploaded": None},
            "audited_financials": {"uploaded": False, "last_uploaded": None},
        }
        try:
            row = db.session.execute(db.text("""
                SELECT MIN(bl.updated_at) FROM budget_lines bl
                JOIN budgets b ON b.id = bl.budget_id
                WHERE b.entity_code = :ec
            """), {"ec": entity_code}).fetchone()
            if row and row[0]:
                result["ysl"]["uploaded"] = True
                result["ysl"]["last_uploaded"] = row[0].isoformat()
        except Exception:
            db.session.rollback()
        try:
            row = db.session.execute(db.text(
                "SELECT uploaded_at FROM expense_reports WHERE entity_code = :ec ORDER BY uploaded_at DESC LIMIT 1"
            ), {"ec": entity_code}).fetchone()
            if row and row[0]:
                result["expense_distribution"]["uploaded"] = True
                result["expense_distribution"]["last_uploaded"] = row[0].isoformat()
        except Exception:
            db.session.rollback()
        try:
            row = db.session.execute(db.text(
                "SELECT uploaded_at FROM open_ap_reports WHERE entity_code = :ec ORDER BY uploaded_at DESC LIMIT 1"
            ), {"ec": entity_code}).fetchone()
            if row and row[0]:
                result["ap_aging"]["uploaded"] = True
                result["ap_aging"]["last_uploaded"] = row[0].isoformat()
        except Exception:
            db.session.rollback()
        try:
            row = db.session.execute(db.text(
                "SELECT uploaded_at FROM maint_proof_reports WHERE entity_code = :ec ORDER BY uploaded_at DESC LIMIT 1"
            ), {"ec": entity_code}).fetchone()
            if row and row[0]:
                result["maint_proof"]["uploaded"] = True
                result["maint_proof"]["last_uploaded"] = row[0].isoformat()
        except Exception:
            db.session.rollback()
        try:
            row = db.session.execute(db.text(
                "SELECT status, created_at, COALESCE(confirmed_at, created_at) FROM audit_uploads WHERE entity_code = :ec ORDER BY created_at DESC LIMIT 1"
            ), {"ec": entity_code}).fetchone()
            if row and row[0]:
                result["audited_financials"]["uploaded"] = True
                result["audited_financials"]["audit_status"] = row[0]
                result["audited_financials"]["last_uploaded"] = row[1].isoformat() if row[1] else None
                if row[0] == "confirmed" and row[2]:
                    result["audited_financials"]["confirmed_at"] = row[2].isoformat()
        except Exception:
            db.session.rollback()
        return result


    @bp.route("/api/wizard/<entity_code>/status", methods=["GET"])
    def wizard_status(entity_code):
        """Return current wizard state for an entity."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        sources = _get_sources_dict(entity_code)

        # Parse assumption history
        history = []
        if budget and budget.assumptions_history_json:
            try:
                history = json.loads(budget.assumptions_history_json)
            except Exception:
                history = []

        # FA directive 2026-05-05: surface notable orphan-GL situations as
        # wizard sidebar notes. Currently scoped to Interest Income (GL 4800)
        # per FA spec — when 4800 has actual data on the building but no
        # summary row aggregates it, the FA needs to know to either add a
        # row or override a related row's COL 3.
        notes = _wizard_notes_for_entity(entity_code, budget)

        return jsonify({
            "entity_code": entity_code,
            "wizard_step": budget.wizard_step if budget else 0,
            "wizard_completed_at": budget.wizard_completed_at.isoformat() if budget and budget.wizard_completed_at else None,
            "status": budget.status if budget else "not_started",
            "has_lines": bool(budget and budget.lines) if budget else False,
            "sources": sources,
            "assumptions_version": len(history),
            "assumptions_history": history,
            "notes": notes,
        })


    def _wizard_notes_for_entity(entity_code, budget):
        """Detect notable budget-state issues that the FA should see in the
        wizard sidebar (right-side panel on the dashboard). Returns a list of
        note dicts {type, severity, title, message, ...}.

        Currently checks:
          - **orphan_interest_income**: GL 4800-XXXX has non-zero ytd or
            current_budget on this building, but no summary row's
            gl_prefixes_json aggregates the 4800 family. Without this note,
            interest income is silently missing from COL 3 (YTD).

        Designed to be additive — drop in more checks here as portfolio-wide
        invariants emerge. Falls back to empty list on any error so the
        wizard sidebar always renders.
        """
        notes = []
        if not budget:
            return notes
        try:
            lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
            summary_rows = BudgetSummaryRow.query.filter_by(
                entity_code=entity_code, budget_year=BUDGET_YEAR
            ).all()

            # Build set of prefixes covered by any summary row
            covered_prefixes = set()
            for row in summary_rows:
                if not row.gl_prefixes_json:
                    continue
                try:
                    for p in (json.loads(row.gl_prefixes_json) or []):
                        s = str(p).strip()
                        if s:
                            covered_prefixes.add(s)
                except Exception:
                    pass

            # Check Interest Income (4800-XXXX)
            interest_lines = [l for l in lines if (l.gl_code or "").startswith("4800")]
            interest_with_data = [
                l for l in interest_lines
                if (l.ytd_actual or 0) != 0 or (l.current_budget or 0) != 0
            ]
            if interest_with_data:
                # 4800 is covered if any summary prefix matches: bare "4800",
                # the full sub-account "4800-0000", etc.
                is_covered = (
                    "4800" in covered_prefixes
                    or any(p == "4800" or p.startswith("4800-") for p in covered_prefixes)
                )
                if not is_covered:
                    total_ytd = sum((l.ytd_actual or 0) for l in interest_lines)
                    total_cb = sum((l.current_budget or 0) for l in interest_lines)
                    gl_codes = sorted({l.gl_code for l in interest_with_data})
                    notes.append({
                        "type": "orphan_interest_income",
                        "severity": "high",
                        "title": "Interest Income not in summary",
                        "message": (
                            "GL 4800 has data (YTD ${ytd:,.0f}, "
                            "Current Budget ${cb:,.0f}) but no summary row "
                            "aggregates it. Add a row via the summary tab "
                            "(+ Add Row → Specific GL → search 4800) or "
                            "override a related row's COL 3."
                        ).format(ytd=total_ytd, cb=total_cb),
                        "gl_codes": gl_codes,
                        "totals": {
                            "ytd": round(total_ytd, 2),
                            "current_budget": round(total_cb, 2),
                        },
                    })

            # FA directive 2026-05-14: surface portfolio-scan findings for
            # this building. Pulls the latest building_scan_findings row
            # written by /api/wizard/<ec>/scan-findings or the nightly
            # scanner. Tells the FA which labels in her approved 2026 file
            # won't aggregate to any canonical row, BEFORE she imports.
            try:
                scan_row = db.session.execute(db.text(
                    "SELECT scanned_at, has_file, labels_unmapped, parse_error, "
                    "unmapped_labels_json "
                    "FROM building_scan_findings "
                    "WHERE entity_code = :ec "
                    "ORDER BY scanned_at DESC LIMIT 1"
                ), {"ec": entity_code}).fetchone()
                if scan_row:
                    has_file = bool(scan_row[1])
                    parse_err = scan_row[3]
                    # One brain with the Pre-import card (2026-06-10, Jacob —
                    # 224's note flagged 5 labels that all had their own rows
                    # with values on the summary): filter the stored unmapped
                    # list through the building's rows + FA overrides before
                    # deciding whether to warn at all.
                    try:
                        ul_j = json.loads(scan_row[4] or '{"unmapped":[]}')
                        raw_unmapped = ul_j.get("unmapped", []) or []
                    except Exception:
                        raw_unmapped = []
                    try:
                        from app import _effective_unmapped_labels
                        still_unmapped, _res, _ov, _bl = _effective_unmapped_labels(
                            entity_code, raw_unmapped)
                    except Exception:
                        still_unmapped = raw_unmapped
                    unmapped_n = len(still_unmapped)
                    if parse_err and not has_file:
                        notes.append({
                            "type": "scan_no_approved_file",
                            "severity": "critical",
                            "title": "No approved 2026 budget file in SharePoint",
                            "message": (
                                "The wizard's import step won't find a file. "
                                "Either upload one to the entity's 2027 Budget "
                                f"folder, or open /api/wizard/{entity_code}/"
                                "scan-findings?refresh=1 after uploading."
                            ),
                        })
                    elif unmapped_n > 0:
                        sample = [
                            (u.get("label") if isinstance(u, dict) else u)
                            for u in still_unmapped
                        ][:6]
                        sev = "high" if unmapped_n > 2 else "medium"
                        notes.append({
                            "type": "scan_unmapped_labels",
                            "severity": sev,
                            "title": (
                                f"{unmapped_n} approved-budget label"
                                f"{'s' if unmapped_n != 1 else ''} won't aggregate"
                            ),
                            "message": (
                                "These labels in the approved 2026 file have "
                                "no canonical row and no row of their own on "
                                "this building's summary — their amounts won't "
                                "land anywhere. Use the Pre-import label check "
                                "card to map, add a row, or override. "
                                "Sample: " + ", ".join(sample[:5]) +
                                ("..." if len(sample) > 5 else "")
                            ),
                            "unmapped_count": unmapped_n,
                            "sample_labels": sample,
                            "review_url": f"/api/wizard/{entity_code}/scan-findings",
                        })
            except Exception as _scan_err:
                logger.warning(f"_wizard_notes_for_entity scan-findings lookup "
                               f"failed for {entity_code}: {_scan_err}")
        except Exception as _err:
            logger.warning(f"_wizard_notes_for_entity failed for {entity_code}: {_err}")
        return notes


    @bp.route("/api/wizard/<entity_code>/step", methods=["POST"])
    def wizard_update_step(entity_code):
        """Update the wizard step for an entity."""
        data = request.get_json(force=True)
        step = data.get("step", 0)

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"success": False, "error": "No budget found for this entity"}), 404

        budget.wizard_step = step
        if step >= 6 and not budget.wizard_completed_at:
            budget.wizard_completed_at = datetime.utcnow()
            # Mirror wizard_complete_build: marking a budget built must also
            # advance status past "not_started", or the dashboard's Send-to-PM
            # transition (draft -> pm_pending) is permanently blocked —
            # VALID_TRANSITIONS only lets not_started step to data_collection.
            if budget.status in (None, "", "not_started", "data_collection", "data_ready"):
                budget.status = "draft"
        db.session.commit()

        return jsonify({"success": True, "wizard_step": step})


    @bp.route("/api/wizard/<entity_code>/assumptions", methods=["POST"])
    def wizard_save_assumptions(entity_code):
        """Save building assumption overrides and bump version history."""
        from app import load_portfolio_defaults, load_building_assumptions, save_building_assumptions, merge_assumptions

        data = request.get_json(force=True)
        overrides = data.get("overrides", {})

        # Save to building_assumptions.json
        all_bldg = load_building_assumptions()
        if entity_code not in all_bldg:
            all_bldg[entity_code] = {}

        # Merge incoming overrides into existing building assumptions
        for key, val in overrides.items():
            if isinstance(val, dict):
                if key not in all_bldg[entity_code]:
                    all_bldg[entity_code][key] = {}
                all_bldg[entity_code][key].update(val)
            else:
                all_bldg[entity_code][key] = val

        save_building_assumptions(all_bldg)

        # Build merged assumptions and update version history
        merged = merge_assumptions(entity_code)
        portfolio = load_portfolio_defaults()

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if budget:
            # Load existing history
            history = []
            if budget.assumptions_history_json:
                try:
                    history = json.loads(budget.assumptions_history_json)
                except Exception:
                    history = []

            # Determine changed fields
            changed = list(overrides.keys()) if overrides else []

            # Append new version
            history.append({
                "version": len(history) + 1,
                "timestamp": datetime.utcnow().isoformat(),
                "portfolio": portfolio,
                "building_overrides": all_bldg.get(entity_code, {}),
                "merged": merged,
                "changed_fields": changed,
            })

            budget.assumptions_history_json = json.dumps(history)
            budget.assumptions_json = json.dumps(merged)
            if budget.wizard_step < 4:
                budget.wizard_step = 4
            db.session.commit()

        return jsonify({
            "success": True,
            "version": len(history) if budget else 1,
            "merged": merged,
        })


    @bp.route("/api/wizard/<entity_code>/flag", methods=["POST"])
    def wizard_flag_assumption(entity_code):
        """Add a flag/note about portfolio assumptions to the budget record."""
        data = request.get_json(force=True)
        note = data.get("note", "")

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"success": False, "error": "No budget found"}), 404

        # Append flag to fa_notes
        flag_entry = f"[ASSUMPTION FLAG {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}] {note}"
        if budget.fa_notes:
            budget.fa_notes = budget.fa_notes + "\n" + flag_entry
        else:
            budget.fa_notes = flag_entry

        db.session.commit()
        return jsonify({"success": True})


    @bp.route("/api/wizard/<entity_code>/preview", methods=["GET"])
    def wizard_preview(entity_code):
        """Return category-level preview showing Yardi raw vs assumptions-applied."""
        from app import merge_assumptions

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"success": False, "error": "No budget found"}), 404

        merged = merge_assumptions(entity_code)
        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()

        # Group lines by category and sum current_budget (Yardi raw)
        categories = {}
        for line in lines:
            cat = _wizard_categorize_gl(line.gl_code, line.sheet_name or line.category)
            if cat not in categories:
                categories[cat] = {"raw": 0.0, "adjusted": 0.0}
            raw_val = line.current_budget or 0.0
            categories[cat]["raw"] += raw_val

            # Apply assumptions to compute adjusted value
            adjusted = _wizard_apply_assumption_to_line(line, merged)
            categories[cat]["adjusted"] += adjusted

        # Build preview rows in P&L order: Income → operating expenses →
        # Capital → catch-alls. Empty buckets are skipped so the FA sees a
        # focused view; "Other" only appears if the categorizer truly didn't
        # know where to put something (should be rare).
        # NB: "Unmapped" is intentionally excluded — the underlying sheet
        # absorbs subtotal rows ("TOTAL OPERATING EXPENSES" etc.) and balance-
        # sheet codes that double-count or don't belong in the operating
        # budget. The template's hide-zero-rows mechanism already keeps these
        # out of the rendered output, so surfacing the bucket in the Preview
        # produces a misleading "$X unmapped" alarm. Lines still land on the
        # Unmapped sheet in budget_lines if anyone needs to inspect them.
        preview = []
        category_order = [
            "Income",
            "Payroll",
            "Insurance",
            "Energy",
            "Water / Sewer",
            "Repairs & Supplies",
            "Gen & Admin",
            "Capital",
            "Other",
        ]
        for cat_name in category_order:
            data = categories.get(cat_name)
            if not data:
                continue
            raw = data["raw"]
            adjusted = data["adjusted"]
            # Skip buckets where both columns round to zero — keeps the table
            # focused on what's actually populated for this entity.
            if abs(raw) < 0.005 and abs(adjusted) < 0.005:
                continue
            delta = adjusted - raw
            pct = (delta / raw * 100) if raw != 0 else 0
            preview.append({
                "category": cat_name,
                "raw": round(raw, 2),
                "adjusted": round(adjusted, 2),
                "delta": round(delta, 2),
                "delta_pct": round(pct, 1),
            })

        return jsonify({"success": True, "preview": preview, "assumptions": merged})


    def _wizard_categorize_gl(gl_code, sheet_or_category):
        """Map a GL code/sheet to a wizard preview category. Categories mirror
        SHEET_TO_CATEGORY (Income, Payroll, Energy, Water & Sewer, Repairs &
        Supplies, Gen & Admin, Capital) plus Insurance (carved out of Gen &
        Admin via GL prefix 61), plus Unmapped (lines whose GL isn't in the
        template — surfaced so FAs see their own data, not "R&M / Other"
        absorbing it). Order matters: Insurance before Gen & Admin (61xx
        lives on the G&A sheet); specific buckets before generic ones.
        """
        sheet = (sheet_or_category or "").lower()
        gl_prefix = (gl_code or "")[:2]

        # Insurance Schedule = GL 6105–6195 (Century KB). Insurance lines live
        # on the Gen & Admin sheet — check this first so they don't fall into
        # the broader Gen & Admin bucket below.
        if "insurance" in sheet or gl_prefix == "61":
            return "Insurance"
        if "payroll" in sheet or gl_prefix in ("50", "51"):
            return "Payroll"
        if "energy" in sheet or gl_prefix == "64":
            return "Energy"
        if "water" in sheet or "sewer" in sheet or gl_prefix == "65":
            return "Water / Sewer"
        if "income" in sheet:
            return "Income"
        if "capital" in sheet:
            return "Capital"
        if "repairs" in sheet or "supplies" in sheet:
            return "Repairs & Supplies"
        if "gen" in sheet or "admin" in sheet:
            return "Gen & Admin"
        if "unmapped" in sheet:
            return "Unmapped"
        return "Other"


    def _wizard_apply_assumption_to_line(line, merged):
        """Apply the relevant assumption to a single budget line, returning adjusted value.

        If the line has an FA override (estimate_override or forecast_override),
        the override is preserved — we return the override value, not the
        assumption-adjusted value.

        IMPORTANT: All percent values in `merged` are stored as decimal
        fractions (0.05 = 5%, 0.15 = 15%) — the form serializes them divided
        by 100 already. Multiply by raw directly; do NOT divide by 100 again.
        """
        # If FA already overrode this line, preserve that
        if line.estimate_override is not None:
            return line.estimate_override
        if line.forecast_override is not None:
            return line.forecast_override

        raw = line.current_budget or 0.0
        gl_prefix = (line.gl_code or "")[:2]
        sheet = (line.sheet_name or "").lower()

        # Payroll — apply wage_increase weighted by pre/post weeks.
        # The form stores: percent (decimal fraction), pre_increase_weeks,
        # post_increase_weeks. Effective annual increase =
        # percent * (post_weeks / 52). Old code averaged ALL numeric fields
        # in wage_increase, including the week counts (15 + 37 + 0.03)/3 = 17.34
        # → 17.3% inflation instead of ~3% (or ~2.1% weighted).
        if "payroll" in sheet or gl_prefix in ("50", "51"):
            wage_inc = merged.get("wage_increase", {}) or {}
            pct = wage_inc.get("percent", 0) or 0
            pre_weeks = wage_inc.get("pre_increase_weeks", 0) or 0
            post_weeks = wage_inc.get("post_increase_weeks", 0) or 0
            total_weeks = pre_weeks + post_weeks
            if pct and total_weeks > 0:
                # Weighted: pre-weeks at 0%, post-weeks at pct, divided by 52
                effective = pct * (post_weeks / 52.0)
                return raw * (1 + effective)
            if pct:
                # Fallback: full-year application if weeks not set
                return raw * (1 + pct)
            return raw

        # Insurance — apply insurance_renewal increase_percent (decimal fraction).
        # GL prefix 61 (Century KB: Insurance Schedule = 6105–6195). Pre/post
        # months weight the renewal so a Mar 2027 renewal at 15% only applies
        # to 9 of 12 months → ~11.25% effective.
        if "insurance" in sheet or gl_prefix == "61":
            ins = merged.get("insurance_renewal", {}) or {}
            pct = ins.get("increase_percent", 0) or 0
            pre_months = ins.get("pre_renewal_months", 0) or 0
            post_months = ins.get("post_renewal_months", 0) or 0
            total = pre_months + post_months
            if pct and total > 0:
                effective = pct * (post_months / 12.0)
                return raw * (1 + effective)
            if pct:
                return raw * (1 + pct)
            return raw

        # Energy — keys depend on GL: gas (640x), electric (641x), oil (645x).
        # Form stores electric_rate_increase / gas_rate_increase /
        # oil_rate_increase as decimal fractions. Old code looked for
        # escalation_pct / increase_pct which don't exist → always $0 delta.
        if "energy" in sheet or gl_prefix == "64":
            energy = merged.get("energy", {}) or {}
            gl = (line.gl_code or "")
            desc = (line.description or "").lower()
            pct = 0
            # GL-prefix routing: 6400-series gas, 6410-series electric, 6450 oil.
            # Fall back to description keywords if GL doesn't match cleanly.
            if gl.startswith("641") or "electric" in desc:
                pct = energy.get("electric_rate_increase", 0) or 0
            elif gl.startswith("645") or "oil" in desc:
                pct = energy.get("oil_rate_increase", 0) or 0
            elif gl.startswith("640") or "gas" in desc:
                pct = energy.get("gas_rate_increase", 0) or 0
            else:
                # Unknown energy line — average the three rates as a safe default
                rates = [
                    energy.get("electric_rate_increase", 0) or 0,
                    energy.get("gas_rate_increase", 0) or 0,
                    energy.get("oil_rate_increase", 0) or 0,
                ]
                non_zero = [r for r in rates if r]
                pct = (sum(non_zero) / len(non_zero)) if non_zero else 0
            if pct:
                return raw * (1 + pct)
            return raw

        # Water/Sewer — form stores rate_increase (decimal fraction). Old code
        # used escalation_pct / increase_pct (don't exist).
        if "water" in sheet or "sewer" in sheet or gl_prefix == "65":
            ws = merged.get("water_sewer", {}) or {}
            pct = ws.get("rate_increase", 0) or 0
            if pct:
                return raw * (1 + pct)
            return raw

        # R&M and other — no assumption, pass through
        return raw


    @bp.route("/api/wizard/<entity_code>/generate", methods=["POST"])
    def wizard_generate(entity_code):
        """Apply assumptions to budget lines and mark wizard complete.

        This is an ADDITIVE operation:
        - Only recalculates lines in assumption-affected categories
        - Preserves per-line FA overrides (estimate_override, forecast_override)
        - Snapshots assumptions to history
        - Propagates Step 3 staged edits (incl. budget_period from YSL auto-
          detect) into Budget.assumptions_json so the dashboard sees them.
        """
        from app import merge_assumptions

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"success": False, "error": "No budget found"}), 404

        # ── FA #3 gate (148 working session 2026-05-13): budget_period MUST be
        # set before Generate. Without it, dashboard falls back to ytd_months=2
        # and every forecast inflates 6× ("Mar-Dec estimate" wrong-period bug).
        # We check the staged value (Step 3) — same source the merge below uses.
        try:
            _staged_for_gate = json.loads(budget.wizard_selections_json or "{}")
            _staged_period = (_staged_for_gate.get("assumptions") or {}).get("budget_period")
        except Exception:
            _staged_period = None
        # Accept "MM/YYYY" only — explicit format so we don't accept "" or junk.
        _valid_period = False
        if isinstance(_staged_period, str) and "/" in _staged_period:
            _mm, _, _yyyy = _staged_period.partition("/")
            try:
                _mm_i = int(_mm); _yyyy_i = int(_yyyy)
                _valid_period = 1 <= _mm_i <= 12 and 2000 <= _yyyy_i <= 2100
            except Exception:
                _valid_period = False
        if not _valid_period:
            return jsonify({
                "success": False,
                "error": "missing_budget_period",
                "message": (
                    "Set the budget period (Step 3) before generating. "
                    "Without it, all forecasts default to a 2-month YTD window "
                    "and inflate ~6×. Pick the month YTD actuals run through."
                ),
            }), 400

        # CFO defaults + per-building overrides (file-based)
        merged = merge_assumptions(entity_code)

        # Deep-merge Step 3 staged edits (FA overrides + auto-detected budget_period)
        # on top so Step 3 wins. Mirrors the build-budget endpoint's deep-merge
        # (app.py ~6955). Without this, budget_period gets dropped and the
        # dashboard falls back to the 2-month YTD default → forecasts inflate 6×.
        # NB: forecast/estimate are computed on-the-fly at render time from
        # assumptions_json["budget_period"], so we don't need to write them to
        # BudgetLine here — saving the merged dict is sufficient.
        try:
            _staged = json.loads(budget.wizard_selections_json or "{}")
            staged_assumptions = _staged.get("assumptions") or {}
        except Exception:
            staged_assumptions = {}
        # Scrub literal "undefined"/"null" keys that can leak in from
        # JS event handlers reading missing data-* attributes. Mirror the
        # guard in /api/wizard/<id>/selections/assumptions so the same
        # defense applies at both write-time and read-time.
        _BAD_KEYS = {"undefined", "null", "None", ""}
        def _scrub_assumptions(d):
            if not isinstance(d, dict):
                return d
            return {k: _scrub_assumptions(v) for k, v in d.items() if k not in _BAD_KEYS}
        if isinstance(staged_assumptions, dict):
            staged_assumptions = _scrub_assumptions(staged_assumptions)
            for key, value in staged_assumptions.items():
                if isinstance(value, dict) and isinstance(merged.get(key), dict):
                    merged[key].update(value)
                else:
                    merged[key] = value
        # Also scrub merged itself in case CFO defaults or per-building overrides
        # got polluted historically.
        merged = _scrub_assumptions(merged)

        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()

        updated_count = 0
        for line in lines:
            # Skip lines with FA overrides — those are preserved
            if line.estimate_override is not None or line.forecast_override is not None:
                continue

            adjusted = _wizard_apply_assumption_to_line(line, merged)
            raw = line.current_budget or 0.0

            # Only update if assumption actually changed the value
            if abs(adjusted - raw) > 0.005:
                line.proposed_budget = adjusted
                updated_count += 1

        # Save assumptions snapshot (now includes staged budget_period + FA edits)
        budget.assumptions_json = json.dumps(merged)

        # Mark wizard complete
        budget.wizard_step = 6
        budget.wizard_completed_at = datetime.utcnow()
        # Advance status to "draft" so the dashboard's Send-to-PM transition
        # (draft -> pm_pending in VALID_TRANSITIONS) is unblocked. Without
        # this, status stays at "not_started" and Send to PM silently fails
        # the transition check.
        if budget.status in (None, "", "not_started", "data_collection", "data_ready"):
            budget.status = "draft"

        db.session.commit()

        return jsonify({
            "success": True,
            "lines_updated": updated_count,
            "total_lines": len(lines),
            "assumptions_version": len(json.loads(budget.assumptions_history_json or "[]")),
        })


    # ─── API Routes: Sheet Subtotal / Sheet-Total Overrides ───────────────
    # FA dir 2026-05-19: FA can override Income/Payroll/etc. tab subtotal +
    # Sheet Total cells via the formula bar (same UX as line cells). Stored
    # per-(entity, row_id, col) in budget.assumptions_json under
    # "sheet_subtotal_overrides".

    @bp.route("/api/sheet-subtotal-override/<entity_code>", methods=["PUT"])
    def put_sheet_subtotal_override(entity_code):
        """Set or clear a subtotal-cell override on the FA dashboard sheet tabs.

        Body: {row_id: str, col: str, value: float|null, formula: str|null}
        Pass value=null to clear the override (reverts to computed sum).
        """
        import json as _json
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        data = request.get_json() or {}
        row_id = (data.get("row_id") or "").strip()
        col = (data.get("col") or "").strip()
        if not row_id or not col:
            return jsonify({"error": "row_id and col required"}), 400
        value = data.get("value")
        formula = data.get("formula")
        try:
            current = _json.loads(budget.assumptions_json) if budget.assumptions_json else {}
        except Exception:
            current = {}
        overrides = current.setdefault("sheet_subtotal_overrides", {})
        row_overrides = overrides.setdefault(row_id, {})
        if value is None and not formula:
            # Clear
            row_overrides.pop(col, None)
            row_overrides.pop(col + "__formula", None)
            if not row_overrides:
                overrides.pop(row_id, None)
        else:
            if value is not None:
                row_overrides[col] = float(value)
            if formula:
                row_overrides[col + "__formula"] = formula
            else:
                row_overrides.pop(col + "__formula", None)
        budget.assumptions_json = _json.dumps(current)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)[:200]}), 500
        return jsonify({"ok": True, "row_id": row_id, "col": col, "value": value, "formula": formula})

    @bp.route("/api/sheet-subtotal-override/<entity_code>", methods=["GET"])
    def get_sheet_subtotal_overrides(entity_code):
        """Return all subtotal overrides for this entity, keyed by row_id → col → value."""
        import json as _json
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"overrides": {}})
        try:
            current = _json.loads(budget.assumptions_json) if budget.assumptions_json else {}
        except Exception:
            current = {}
        return jsonify({"overrides": current.get("sheet_subtotal_overrides", {})})


    # ─── API Routes: Per-Building Data Quality Health Check ─────────────────
    # FA dir 2026-05-19: data-quality checks per building, exposed to the
    # existing Health drawer as a second tab ("Data Quality" alongside
    # "Readiness"). Each check encodes a bug class we hit on 148/168 so we
    # catch the same pattern on every future entity automatically.

    @bp.route("/api/health-check/<entity_code>", methods=["GET"])
    def get_health_check(entity_code):
        """Per-building data-quality health check.

        Returns a list of checks with status (pass/warn/fail), human-readable
        detail, suggested fix (when applicable), and supporting data. Powers
        the "Data Quality" tab in the Health drawer.

        Designed to catch every bug class from the 148/168 cycle:
          - GL routing (Cable TV in own row, not Other Income)
          - Maintenance prefix completeness (4010 + 4060 + 4070)
          - Cable TV breakout
          - Flip Tax routes below the line
          - Working Capital Contribution row exists when GL has data
          - RE Tax exemptions reconcile with G&A budget
          - Capital lines proposed = $0
          - Summary math ties (Total Income = sum of income rows)
          - Audited financials confirmed for BUDGET_YEAR-2
          - Commercial RE Tax row in income section, not at the bottom
          - Cross-building shape diff against a known-good baseline
        """
        import json as _json
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found", "entity_code": entity_code}), 404

        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
        rows = (
            BudgetSummaryRow.query
            .filter_by(entity_code=entity_code, budget_year=BUDGET_YEAR)
            .order_by(BudgetSummaryRow.display_order)
            .all()
        )
        # Index by GL prefix → row label for the "what bucket does X live in" question
        rows_by_label = {r.label: r for r in rows}

        checks = []
        def _check(name, status, detail, fix=None, data=None):
            checks.append({
                "name": name,
                "status": status,
                "detail": detail,
                "fix": fix,
                "data": data or [],
            })

        # Calibration (2026-06-08): an unbuilt budget (no lines) has nothing to
        # data-quality-check. Short-circuit so the preflight doesn't cry wolf about
        # "missing" rows on a building that simply hasn't been built yet.
        if not lines:
            _check("Budget built", "warn",
                   "No budget lines yet. This building hasn't been built; run the wizard before data-quality checks apply.")
            return jsonify({"entity_code": entity_code,
                            "scanned_at": datetime.utcnow().isoformat() + "Z",
                            "summary": {"pass": 0, "warn": 1, "fail": 0},
                            "checks": checks})

        # ── Check 1: GL coverage — no orphan lines on Unmapped sheet ─────
        orphans = [l for l in lines if (l.sheet_name or "") == "Unmapped"
                   and (abs(float(l.ytd_actual or 0)) > 0.01 or abs(float(l.current_budget or 0)) > 0.01)]
        if not orphans:
            _check("GL coverage", "pass",
                   f"All {len(lines)} account numbers route to a sheet. Zero orphans with real data.")
        else:
            _check("GL coverage", "fail",
                   f"{len(orphans)} account number{'s' if len(orphans) != 1 else ''} on the Unmapped sheet have YTD or budget data with nowhere to go.",
                   fix={"label": "Review unmapped lines", "url": f"/dashboard/{entity_code}#unmapped"},
                   data=[{"gl": l.gl_code, "desc": l.description or "",
                          "ytd": float(l.ytd_actual or 0)} for l in orphans[:8]])

        # ── Check 2: All standard summary rows present ────────────────────
        # Calibration (2026-06-08): only the rows EVERY building must have. Conditional
        # income rows (Commercial Rent, Cable TV, Flip Tax, Tax Benefit Credits, RE
        # Taxes) have their own data-conditional checks; listing them here warned
        # 147/148 buildings for line items they legitimately don't carry.
        STRUCTURAL_LABELS = [
            "Maintenance", "Total Income",
            "Payroll", "Electric", "Water & Sewer", "Supplies",
            "Repairs & Maintenance", "Insurance",
            "Professional Fees", "Administrative & Other",
            "Total Expenses", "Net Operating Surplus <Deficit>",
            "Total Surplus <Deficit>",
        ]
        missing = [lbl for lbl in STRUCTURAL_LABELS if lbl not in rows_by_label]
        if not missing:
            _check("All standard summary rows present", "pass",
                   f"{len(rows)} summary rows. All structural rows (income total, expense categories, totals) present.")
        else:
            _check("All standard summary rows present", "warn",
                   f"{len(missing)} structural summary row{'s' if len(missing) != 1 else ''} missing on this building's summary.",
                   data=[{"gl": "", "desc": lbl, "ytd": 0} for lbl in missing[:8]])

        # ── Check 3: Maintenance row includes 4010 + 4060 + 4070 ──────────
        m_row = rows_by_label.get("Maintenance")
        try:
            m_prefixes = _json.loads(m_row.gl_prefixes_json) if m_row and m_row.gl_prefixes_json else []
        except Exception:
            m_prefixes = []
        # Calibration (2026-06-08): only require a maintenance-income prefix if the
        # building actually HAS data on it. 4060 (arrears)/4070 (prepaid) are often $0,
        # which previously warned 92/148 buildings for income families they don't have.
        def _has_prefix_data(pfx):
            return any((l.gl_code or "").startswith(pfx) and
                       (abs(float(l.ytd_actual or 0)) > 0.01 or abs(float(l.current_budget or 0)) > 0.01)
                       for l in lines)
        m_expected = {p for p in ("4010", "4060", "4070") if _has_prefix_data(p)}
        m_missing = m_expected - set(p for p in m_prefixes)
        if not m_row:
            _check("Maintenance income aggregation", "warn",
                   "Maintenance row not present on this building.")
        elif not m_missing:
            _check("Maintenance income aggregation", "pass",
                   f"Maintenance row pulls every maintenance-income family that has data ({', '.join(sorted(m_expected)) or 'none'}).")
        else:
            _check("Maintenance income aggregation", "fail",
                   f"Maintenance row is missing account prefixes that HAVE data: {', '.join(sorted(m_missing))}. That income won't roll up here.",
                   fix={"label": "Run resolve-summary-aliases", "endpoint": f"/api/admin/resolve-summary-aliases/{entity_code}"},
                   data=[{"gl": p, "desc": "Has data but missing from Maintenance prefix list", "ytd": 0} for p in sorted(m_missing)])

        # ── Check 4: Cable TV breakout (own row, not in Other Income) ─────
        cable_row = rows_by_label.get("Cable TV")
        try:
            cable_prefixes = _json.loads(cable_row.gl_prefixes_json) if cable_row and cable_row.gl_prefixes_json else []
        except Exception:
            cable_prefixes = []
        # Calibration (2026-06-08): only flag cable routing if the building HAS cable
        # income (4250). Warning a building with $0 cable about a "missing cable row"
        # fired on 143/148 buildings and buried the real problems.
        cable_ytd = sum(float(l.ytd_actual or 0) for l in lines if (l.gl_code or "").startswith("4250"))
        if cable_row and "4250" in cable_prefixes:
            _check("Cable TV breakout", "pass",
                   f"Cable TV income (account 4250) routes to its own summary row. ${cable_ytd:,.0f} YTD on this building.")
        elif abs(cable_ytd) < 0.01:
            _check("Cable TV breakout", "pass",
                   "No cable income (account 4250) on this building. Not applicable.")
        else:
            _check("Cable TV breakout", "warn",
                   f"${cable_ytd:,.0f} of cable income (4250) exists but isn't broken out into its own row; it's likely lumped into Other Income.",
                   fix={"label": "Run resolve-summary-aliases", "endpoint": f"/api/admin/resolve-summary-aliases/{entity_code}"})

        # ── Check 5: Flip Tax routes below the line ───────────────────────
        flip_row = rows_by_label.get("Flip Tax/Transfer Fees") or rows_by_label.get("Flip Tax")
        try:
            flip_prefixes = _json.loads(flip_row.gl_prefixes_json) if flip_row and flip_row.gl_prefixes_json else []
        except Exception:
            flip_prefixes = []
        flip_has_7025 = any(p == "7025" for p in flip_prefixes)
        # Calibration (2026-06-08): only flag flip-tax routing if the building HAS flip
        # tax income (7025). Previously fired on 131/148 buildings with no flip tax.
        flip_ytd = sum(float(l.ytd_actual or 0) for l in lines if (l.gl_code or "").startswith("7025"))
        if flip_row and flip_has_7025:
            _check("Flip Tax routes below the line", "pass",
                   "Account 7025 (Flip Tax - Capital) routes to non-operating income, not Capital Expenses.")
        elif abs(flip_ytd) < 0.01:
            _check("Flip Tax routes below the line", "pass",
                   "No flip tax income (account 7025) on this building. Not applicable.")
        elif flip_row:
            _check("Flip Tax routes below the line", "warn",
                   f"${flip_ytd:,.0f} of flip tax exists but the Flip Tax row doesn't include account 7025, so it may be stuck on the Capital sheet.",
                   fix={"label": "Run resolve-summary-aliases", "endpoint": f"/api/admin/resolve-summary-aliases/{entity_code}"})
        else:
            _check("Flip Tax routes below the line", "warn",
                   f"${flip_ytd:,.0f} of flip tax income (7025) exists but there's no 'Flip Tax/Transfer Fees' row to hold it.")

        # ── Check 6: Working Capital Contribution row exists when GL has data ─
        wc_lines = [l for l in lines if (l.gl_code or "") == "4725-0040"
                    and (abs(float(l.ytd_actual or 0)) > 0.01 or abs(float(l.current_budget or 0)) > 0.01)]
        wc_row = rows_by_label.get("Working Capital Contribution")
        if not wc_lines:
            _check("Working Capital Contribution", "pass",
                   "No Working Capital Contribution data on this building (account 4725-0040 is empty). Skip.")
        elif wc_row:
            _check("Working Capital Contribution", "pass",
                   f"Working Capital row exists. Account 4725-0040 contributes ${float(wc_lines[0].ytd_actual or 0):,.0f} YTD here.")
        else:
            _check("Working Capital Contribution", "fail",
                   f"Account 4725-0040 has ${float(wc_lines[0].ytd_actual or 0):,.0f} YTD but no destination row on this summary.",
                   fix={"label": "Add 'Working Capital Contribution' row", "endpoint": "/api/admin/add-summary-row",
                        "body": {"entity_code": entity_code, "label": "Working Capital Contribution",
                                  "section": "Non-Operating Income", "display_order": 33}},
                   data=[{"gl": l.gl_code, "desc": l.description or "",
                          "ytd": float(l.ytd_actual or 0)} for l in wc_lines])

        # ── Check 7: Capital lines proposed = $0 ───────────────────────────
        capital_with_proposed = [l for l in lines
                                  if ((l.sheet_name or "") == "Capital" or (l.category or "").lower() == "capital")
                                  and l.proposed_budget is not None and abs(float(l.proposed_budget or 0)) > 0.01]
        capital_count = sum(1 for l in lines if (l.sheet_name or "") == "Capital" or (l.category or "").lower() == "capital")
        if capital_count == 0:
            _check("Capital lines have no proposed budget", "pass",
                   "No capital lines on this building.")
        elif not capital_with_proposed:
            _check("Capital lines have no proposed budget", "pass",
                   f"All {capital_count} capital lines display $0 proposed (correct — capital projects aren't proposed via the operating budget).")
        else:
            _check("Capital lines have no proposed budget", "warn",
                   f"{len(capital_with_proposed)} of {capital_count} capital lines have a proposed_budget value set. Capital should always be $0.",
                   data=[{"gl": l.gl_code, "desc": l.description or "",
                          "ytd": float(l.proposed_budget or 0)} for l in capital_with_proposed[:8]])

        # ── Check 8: RE Tax exemptions reconcile with G&A budget ───────────
        is_coop_b = (budget.building_type or "").lower() in ("coop", "co-op")
        if is_coop_b:
            try:
                from dof_taxes import compute_re_taxes
                overrides = None
                if budget.assumptions_json:
                    try:
                        overrides = _json.loads(budget.assumptions_json).get("re_taxes_overrides")
                    except Exception:
                        overrides = None
                # Apply the same GL-line fallback the live endpoint uses
                _GL_TO_OVR = {
                    "6315-0010": "abatement_current", "6315-0020": "star_current",
                    "6315-0025": "veteran_current",   "6315-0035": "sche_current",
                }
                ovr = dict(overrides or {})
                for _gl, _key in _GL_TO_OVR.items():
                    if not ovr.get(_key):
                        _l = next((l for l in lines if l.gl_code == _gl), None)
                        if _l and _l.current_budget:
                            ovr[_key] = abs(float(_l.current_budget))
                re_taxes = compute_re_taxes(entity_code, ovr)
                rt_total = float(re_taxes.get("total_exemptions_budget") or 0)
                ga_total = sum(abs(float(l.current_budget or 0)) for l in lines
                                if (l.gl_code or "") in _GL_TO_OVR)
                if rt_total > 0 and ga_total > 0:
                    delta = abs(rt_total - ga_total)
                    if delta < 1.0:  # to the dollar
                        _check("RE Tax exemptions reconcile", "pass",
                               f"RE Tax page total credits (${rt_total:,.0f}) matches sum of G&A tax credit budget lines (${ga_total:,.0f}).")
                    else:
                        _check("RE Tax exemptions reconcile", "warn",
                               f"RE Tax page shows ${rt_total:,.0f} in credits; G&A budget lines sum to ${ga_total:,.0f}. Off by ${delta:,.0f}.")
                else:
                    _check("RE Tax exemptions reconcile", "warn",
                           "RE Tax exemption totals are zero. FA may need to enter values on the RE Tax tab or set the 6315-* budget lines.")
            except Exception as e:
                _check("RE Tax exemptions reconcile", "warn",
                       f"Could not compute RE Tax check: {str(e)[:80]}")
        else:
            _check("RE Tax exemptions reconcile", "pass",
                   "Not applicable for condos (no building-level RE tax).")

        # ── Check 9: Math ties — Total Income = sum of income rows ─────────
        # Best-effort: only checks col6 (current budget) which is most stable.
        income_rows = [r for r in rows if (r.section or "").lower() == "income" and r.row_type == "data"]
        total_income_row = rows_by_label.get("Total Income")
        if income_rows and total_income_row and total_income_row.col6_approved_budget is not None:
            sum_income = sum(float(r.col6_approved_budget or 0) for r in income_rows)
            total = float(total_income_row.col6_approved_budget)
            delta = abs(sum_income - total)
            if delta < 1.0:
                _check("Summary math ties (Total Income)", "pass",
                       f"Total Income (${total:,.0f}) equals sum of {len(income_rows)} income rows.")
            else:
                _check("Summary math ties (Total Income)", "warn",
                       f"Total Income (${total:,.0f}) differs from sum of income rows (${sum_income:,.0f}) by ${delta:,.0f}.")
        else:
            _check("Summary math ties (Total Income)", "pass",
                   "Skip — totals not yet populated on this building.")

        # ── Check 10: Audited financials confirmed for BUDGET_YEAR - 2 ─────
        try:
            target_year = str(BUDGET_YEAR - 2)
            audits = db.session.execute(
                db.text("SELECT id, status, fiscal_year_end, pdf_filename FROM audit_uploads "
                         "WHERE entity_code = :ec AND fiscal_year_end = :fy"),
                {"ec": entity_code, "fy": target_year}
            ).fetchall()
            confirmed = [a for a in audits if (a[1] or "") == "confirmed"]
            if confirmed:
                _check(f"Audited financials confirmed ({target_year})", "pass",
                       f"FY{target_year} audit upload confirmed. Column 2 (BY-2 Actual) populated on summary.")
            elif audits:
                _check(f"Audited financials confirmed ({target_year})", "fail",
                       f"FY{target_year} audit uploaded but not confirmed yet. Column 2 will be blank on summary.",
                       fix={"label": "Open Audited Financials", "url": "/audited-financials"})
            elif is_coop_b:
                _check(f"Audited financials confirmed ({target_year})", "warn",
                       f"No FY{target_year} audit upload found for this co-op. Column 2 will be blank.")
            else:
                _check(f"Audited financials confirmed ({target_year})", "pass",
                       "Not applicable for non-coop.")
        except Exception as e:
            _check(f"Audited financials confirmed", "warn",
                   f"Could not check audit status: {str(e)[:80]}")

        # ── Check 11: Commercial RE Tax row position ───────────────────────
        cret_row = rows_by_label.get("Commercial Real Estate Tax")
        total_income_idx = (total_income_row.display_order if total_income_row else None)
        if not cret_row:
            _check("Commercial RE Tax row position", "pass",
                   "No Commercial Real Estate Tax row on this building.")
        elif total_income_idx is not None and cret_row.display_order > total_income_idx + 5:
            _check("Commercial RE Tax row position", "warn",
                   f"Commercial RE Tax row is at display position {cret_row.display_order}, well below Total Income (position {total_income_idx}). Belongs in the income section.",
                   fix={"label": "Move row up", "endpoint": "/api/admin/move-summary-row",
                        "body": {"entity_code": entity_code, "label": "Commercial Real Estate Tax", "to_position": total_income_idx - 2}})
        else:
            _check("Commercial RE Tax row position", "pass",
                   f"Commercial RE Tax row is in the income section (position {cret_row.display_order}).")

        # ── Summary counts ─────────────────────────────────────────────────
        summary = {
            "pass": sum(1 for c in checks if c["status"] == "pass"),
            "warn": sum(1 for c in checks if c["status"] == "warn"),
            "fail": sum(1 for c in checks if c["status"] == "fail"),
        }
        return jsonify({
            "entity_code": entity_code,
            "scanned_at": datetime.utcnow().isoformat() + "Z",
            "summary": summary,
            "checks": checks,
        })


    # ─── API Routes: RE Taxes (NYC DOF) ─────────────────────────────────────

    @bp.route("/api/re-taxes/<entity_code>", methods=["GET"])
    def get_re_taxes(entity_code):
        """Get RE Taxes calculation for a co-op property, pulling from NYC DOF."""
        try:
            from dof_taxes import is_coop, compute_re_taxes
            if not is_coop(entity_code):
                return jsonify({"error": "Not a co-op — condos do not have building-level RE taxes", "is_coop": False}), 200
            import json as _json
            budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
            # Override prep (load + FA-dir-2026-05-18 G&A 6315 backfill) lives
            # in _re_tax_overrides_for — shared with every other producer.
            overrides = _re_tax_overrides_for(budget)

            result = compute_re_taxes(entity_code, overrides)
            # Pass through saved per-cell overrides (numeric + formula sources)
            # so the frontend can restore user edits on reload.
            if isinstance(overrides, dict) and overrides.get("cell_overrides"):
                result["cell_overrides"] = overrides["cell_overrides"]
            # FA dir 2026-05-19: After-10/31 toggle round-trip.
            if isinstance(overrides, dict):
                result["after_oct31"] = bool(overrides.get("after_oct31"))
            # FA #14 (2026-06-16): surface any custom RE-tax escalation /
            # adjustment lines this building has — 6315-xxxx budget lines beyond
            # the 7 fixed GLs — so the page renders them as extra Section-3 rows.
            _RE_FIXED_GLS = {
                "6315-0000", "6315-0010", "6315-0020", "6315-0025",
                "6315-0030", "6315-0035", "6315-0040",
            }
            custom_rows = []
            if budget:
                _clines = BudgetLine.query.filter(
                    BudgetLine.budget_id == budget.id,
                    BudgetLine.gl_code.like("6315-%"),
                ).order_by(BudgetLine.gl_code).all()
                for _cl in _clines:
                    _gl = (_cl.gl_code or "").strip()
                    if _gl and _gl not in _RE_FIXED_GLS:
                        custom_rows.append({"gl": _gl, "label": _cl.description or _gl})
            result["custom_gl_rows"] = custom_rows
            return jsonify({"is_coop": True, "re_taxes": result})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @bp.route("/api/re-taxes/<entity_code>/add-line", methods=["POST"])
    def add_re_tax_line(entity_code):
        """FA #14: add a custom RE-tax escalation / adjustment line.

        Creates a new 6315-xxxx BudgetLine on this co-op's budget. The new line
        then renders as an extra Section-3 row on the RE Taxes page (via the GET
        endpoint's custom_gl_rows) AND rolls into Gen & Admin + the Summary
        through the 6315 prefix aggregation. The 7 fixed GLs use suffixes
        0000-0040; the next free 6315-00NN suffix is allocated for the new line.
        """
        try:
            from dof_taxes import is_coop
            import re as _re
            if not is_coop(entity_code):
                return jsonify({"error": "Not a co-op — RE taxes only apply to co-ops"}), 400
            data = request.get_json(silent=True) or {}
            label = str(data.get("label") or "").strip()
            if not label:
                return jsonify({"error": "A label is required"}), 400
            try:
                amount = float(data.get("amount") or 0)
            except (TypeError, ValueError):
                amount = 0.0

            budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
            if not budget:
                return jsonify({"error": "No %s budget for %s — build it first" % (BUDGET_YEAR, entity_code)}), 404

            # Allocate the next free 6315-00NN suffix. Fixed GLs occupy 0000-0040;
            # scan existing 6315-#### lines and take max(suffix, 40) + 5. Also grab
            # a template line to copy sheet/category so the new line lands in the
            # same tab + summary bucket as the building's other RE-tax lines.
            existing = BudgetLine.query.filter(
                BudgetLine.budget_id == budget.id,
                BudgetLine.gl_code.like("6315-%"),
            ).all()
            max_suffix = 40
            template = None
            for _ln in existing:
                m = _re.match(r"^6315-(\d{4})$", (_ln.gl_code or "").strip())
                if m:
                    max_suffix = max(max_suffix, int(m.group(1)))
                if template is None:
                    template = _ln
            new_gl = "6315-%04d" % (max_suffix + 5)

            sheet_name = (template.sheet_name if template and template.sheet_name else "Gen & Admin")
            category = (template.category if template and template.category else "Real Estate Taxes")

            max_row = db.session.query(db.func.max(BudgetLine.row_num)).filter(
                BudgetLine.budget_id == budget.id
            ).scalar() or 0

            line = BudgetLine(
                budget_id=budget.id,
                gl_code=new_gl,
                description=label,
                category=category,
                row_num=int(max_row) + 1,
                sheet_name=sheet_name,
                pm_editable=False,
                prior_year=0,
                ytd_actual=amount,
                ytd_budget=0,
                current_budget=amount,
            )
            db.session.add(line)
            db.session.flush()
            db.session.add(BudgetRevision(
                budget_id=budget.id,
                budget_line_id=line.id,
                action="create",
                field_name="gl_code",
                old_value="",
                new_value=new_gl,
                source="re_tax_add_row",
                notes="RE Taxes #14: custom line '%s' = %s" % (label, amount),
            ))
            db.session.commit()
            return jsonify({"ok": True, "gl": new_gl, "label": label, "amount": amount})
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

    @bp.route("/api/re-taxes/<entity_code>", methods=["PUT"])
    def update_re_taxes(entity_code):
        """Save RE Taxes overrides (exemptions, transitional increase, etc.)."""
        import json as _json
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400
        try:
            current = _json.loads(budget.assumptions_json) if budget.assumptions_json else {}
        except Exception:
            current = {}
        current["re_taxes_overrides"] = data
        budget.assumptions_json = _json.dumps(current)
        # Also update the Gen & Admin GL lines that reference RE Taxes
        try:
            from dof_taxes import compute_re_taxes
            result = compute_re_taxes(entity_code, data)
            # Update GL 6315-0000 (Gross Tax) proposed budget
            _update_gl_line(budget.id, "6315-0000", result["gross_tax"])
            # Update exemption lines (negated — they reduce tax)
            _update_gl_line(budget.id, "6315-0010", -result["exemptions"]["coop_abatement"]["budget_year"])
            _update_gl_line(budget.id, "6315-0020", -result["exemptions"]["star"]["budget_year"])
            _update_gl_line(budget.id, "6315-0025", -result["exemptions"]["veteran"]["budget_year"])
            _update_gl_line(budget.id, "6315-0035", -result["exemptions"]["sche"]["budget_year"])
            # FA #29 (2026-06-15): pin the G&A 6315 lines' Estimate (col4, May-Dec)
            # and Forecast (col5, 12-mo) to the RE-tax page values so Gen&Admin
            # matches the RE-tax page (the authoritative source) instead of
            # annualizing YTD. Mirrors the RE-tax Section-3 per-GL formulas:
            # E = first_half_tax/2 (tax line), -(exemption base)/4 (exemptions),
            # 0 (SCRIE/J-51), all 0 after 10/31; F = YTD + E. The summary
            # aggregator + the G&A faComputeForecast both honor these overrides.
            _after = bool(data.get("after_oct31") or data.get("afterOct31"))
            _fh = float(result.get("first_half_tax") or 0)
            _ex = result.get("exemptions") or {}
            def _exbase(k):
                return float((_ex.get(k) or {}).get("current_year") or 0)
            _gl_est = {
                "6315-0000": (0.0 if _after else _fh / 2),
                "6315-0010": (0.0 if _after else -_exbase("coop_abatement") / 4),
                "6315-0020": (0.0 if _after else -_exbase("star") / 4),
                "6315-0025": (0.0 if _after else -_exbase("veteran") / 4),
                "6315-0030": 0.0,
                "6315-0035": (0.0 if _after else -_exbase("sche") / 4),
                "6315-0040": 0.0,
            }
            for _gl, _est in _gl_est.items():
                _ln = BudgetLine.query.filter_by(budget_id=budget.id, gl_code=_gl).first()
                if _ln:
                    _ln.estimate_override = round(_est, 2)
                    _ln.forecast_override = round(float(_ln.ytd_actual or 0) + _est, 2)
        except Exception as e:
            logger.warning(f"Failed to update Gen & Admin tax lines: {e}")
        db.session.commit()
        # FA #14: keep custom escalation/adjustment rows attached on the save
        # response so they don't vanish if the tab is re-rendered after an
        # autosave (window._reTaxesData is refreshed from this payload).
        try:
            _RE_FIXED_GLS = {
                "6315-0000", "6315-0010", "6315-0020", "6315-0025",
                "6315-0030", "6315-0035", "6315-0040",
            }
            _custom = []
            for _cl in BudgetLine.query.filter(
                BudgetLine.budget_id == budget.id,
                BudgetLine.gl_code.like("6315-%"),
            ).order_by(BudgetLine.gl_code).all():
                _gl = (_cl.gl_code or "").strip()
                if _gl and _gl not in _RE_FIXED_GLS:
                    _custom.append({"gl": _gl, "label": _cl.description or _gl})
            if isinstance(result, dict):
                result["custom_gl_rows"] = _custom
        except Exception:
            pass
        return jsonify({"status": "saved", "re_taxes": result})

    def _update_gl_line(budget_id, gl_code, value):
        """Update the proposed_budget for a specific GL line."""
        line = BudgetLine.query.filter_by(budget_id=budget_id, gl_code=gl_code).first()
        if line:
            line.proposed_budget = round(value, 2)

    # ─── API Routes: Lines ───────────────────────────────────────────────────

    @bp.route("/api/lines/<entity_code>", methods=["GET"])
    def get_lines(entity_code):
        """Get all R&M lines for a building."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        lines = BudgetLine.query.filter_by(budget_id=budget.id).order_by(BudgetLine.row_num).all()
        return jsonify([l.to_dict() for l in lines])


    @bp.route("/api/lines/<entity_code>", methods=["PUT"])
    def update_lines(entity_code):
        """Update R&M lines for a building (PM data entry)."""
        data = request.get_json()

        # FA directive 2026-05-10: structured logging instead of print().
        try:
            _incoming_lines = (data or {}).get("lines", []) or []
            _notes_in = [(l.get("gl_code"), l.get("notes")) for l in _incoming_lines if (l.get("notes") or "").strip()]
            logger.info(
                "[update_lines] entity=%s total_lines=%d with_notes=%d sample=%s",
                entity_code, len(_incoming_lines), len(_notes_in), _notes_in[:5],
            )
        except Exception as _diag_err:
            logger.warning("[update_lines] diag err: %s", _diag_err)

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            logger.warning("[update_lines] entity=%s NOT FOUND for year %s", entity_code, BUDGET_YEAR)
            return jsonify({"error": "Budget not found"}), 404

        logger.info("[update_lines] entity=%s budget.id=%s status=%s",
                    entity_code, budget.id, budget.status)

        # Check if PM can edit
        # fa_review is allowed so the PM can re-enter and save edits after submit.
        if budget.status not in ["pm_pending", "pm_in_progress", "returned", "fa_review"]:
            logger.info("[update_lines] REJECTED — status %s not editable", budget.status)
            return jsonify({"error": "Budget is not in editable status"}), 400

        # Mark as in progress
        if budget.status == "pm_pending":
            budget.status = "pm_in_progress"

        # Suggestion 5 (204 dry run): count what actually happened so a
        # payload that matches nothing can't read as success.
        _pm_applied = 0
        _pm_skipped = []
        # Update each line — match by gl_code (or id if provided)
        for line_data in data.get("lines", []):
            line = None
            if line_data.get("id"):
                line = BudgetLine.query.get(line_data["id"])
                if line and line.budget_id != budget.id:
                    line = None
            elif line_data.get("gl_code"):
                line = BudgetLine.query.filter_by(
                    budget_id=budget.id, gl_code=line_data["gl_code"]
                ).first()
            if not line:
                _pm_skipped.append(str(line_data.get("gl_code") or line_data.get("id")))
                continue
            _pm_applied += 1

            # Track changes for PM audit trail
            changes = []

            # Float fields that are always present in PM payload
            for fname in ("accrual_adj", "unpaid_bills", "increase_pct"):
                if fname in line_data:
                    new_val = float(line_data.get(fname, 0) or 0)
                    old_val = getattr(line, fname, None) or 0
                    if old_val != new_val:
                        changes.append((fname, str(old_val), str(new_val)))
                    setattr(line, fname, new_val)

            # FA directive 2026-05-11: PM can now enter EITHER an % increase
            # OR a $ amount (`increase_dollar`). Either-or enforced HERE at
            # save time — setting one clears the other. NULL = unset.
            # The frontend's pmCellBlur also clears the sibling in-memory
            # so the UI stays consistent without a round-trip.
            if "increase_dollar" in line_data:
                raw = line_data.get("increase_dollar")
                if raw is None or raw == "":
                    new_val = None
                else:
                    try:
                        new_val = float(raw)
                    except (TypeError, ValueError):
                        new_val = None
                old_val = line.increase_dollar
                if old_val != new_val:
                    changes.append(("increase_dollar", str(old_val), str(new_val)))
                line.increase_dollar = new_val
                # When the PM enters a $ amount, clear the % so either-or
                # is enforced even if the FE didn't (defense in depth).
                if new_val is not None and (line.increase_pct or 0) != 0:
                    changes.append(("increase_pct", str(line.increase_pct), "0 (cleared by $-entry)"))
                    line.increase_pct = 0.0
            # Symmetric: if FE explicitly sent a non-zero % AND the row
            # currently has a $ amount set, clear the $.
            if ("increase_pct" in line_data
                    and float(line_data.get("increase_pct") or 0) != 0
                    and line.increase_dollar is not None):
                changes.append(("increase_dollar", str(line.increase_dollar), "NULL (cleared by %-entry)"))
                line.increase_dollar = None

            # FA directive 2026-05-11: PM R&M review-gate state stamping.
            # Only stamped for sheet_name == "Repairs & Supplies" and only
            # when the FE sends an explicit `pm_action` signal — saveAll
            # sends increase_pct for every line on every save, so we cannot
            # use key-presence as the signal (would create false positives).
            #
            # Wire protocol (pm_action values):
            #   "review_pct"     PM typed in the % field (incl. 0%)
            #   "review_dollar"  PM typed in the $ field (incl. 0)
            #   "no_change"      PM clicked the row's "No change" button —
            #                    forces increase_pct = 0, increase_dollar = NULL
            #
            # Absence of pm_action leaves pm_review_state untouched (so
            # saving notes or accrual_adj on an unreviewed line keeps it
            # unreviewed; saving the same on an already-reviewed line keeps
            # its prior state).
            try:
                _pm_action = (line_data.get("pm_action") or "").strip().lower()
            except Exception:
                _pm_action = ""
            if line.sheet_name == "Repairs & Supplies" and _pm_action:
                _new_state = None
                if _pm_action == "no_change":
                    # Explicit "No change" click overrides everything else.
                    # FA directive 2026-05-18: under single-entry, "No change"
                    # means proposed_budget = current_budget (flat-line vs
                    # current). Clear the legacy increase fields too.
                    line.proposed_budget = float(line.current_budget or 0)
                    line.increase_pct = 0.0
                    line.increase_dollar = None
                    _new_state = "no_change"
                elif _pm_action == "review_dollar":
                    _new_state = "typed_dollar"
                elif _pm_action == "review_pct":
                    _new_state = "typed_pct"
                elif _pm_action == "review_proposed":
                    # FA directive 2026-05-18: PM typed a proposed_budget value
                    # directly. Single-entry stamping signal.
                    _new_state = "typed_proposed"
                if _new_state:
                    _old_state = line.pm_review_state or "unreviewed"
                    if _old_state != _new_state:
                        line.pm_review_state = _new_state
                        line.pm_reviewed_at = datetime.utcnow()
                        line.pm_reviewed_by = _read_fa_id_from_cookie()
                        changes.append(("pm_review_state", _old_state, _new_state))

            # Notes
            if "notes" in line_data:
                new_val = line_data.get("notes", "")
                if (line.notes or "") != new_val:
                    changes.append(("notes", line.notes or "", new_val))
                    logger.info("[update_lines] notes change gl=%s '%s' -> '%s'",
                                line.gl_code, line.notes or '', new_val)
                line.notes = new_val

            # Category
            if "category" in line_data and line_data["category"]:
                old_val = line.category or ""
                new_val = line_data["category"]
                if old_val != new_val:
                    changes.append(("category", old_val, new_val))
                line.category = new_val

            # Nullable override fields
            for ofield in ("estimate_override", "forecast_override"):
                if ofield in line_data:
                    raw = line_data[ofield]
                    new_val = float(raw) if raw is not None else None
                    old_val = getattr(line, ofield, None)
                    if old_val != new_val:
                        changes.append((ofield, str(old_val), str(new_val)))
                    setattr(line, ofield, new_val)

            # Proposed budget and formula
            # FA directive 2026-05-18: accept null/empty to mean "PM has not
            # entered a value" — distinguish wiped/unset from explicit zero.
            if "proposed_budget" in line_data:
                raw = line_data["proposed_budget"]
                if raw is None or raw == "":
                    new_val = None
                else:
                    try:
                        new_val = float(raw)
                    except (TypeError, ValueError):
                        new_val = None
                old_val = line.proposed_budget
                if (old_val or 0) != (new_val or 0):
                    changes.append(("proposed_budget", str(old_val), str(new_val)))
                line.proposed_budget = new_val
            if "proposed_formula" in line_data:
                new_val = line_data["proposed_formula"] or None
                old_val = line.proposed_formula or ""
                if old_val != (new_val or ""):
                    changes.append(("proposed_formula", old_val, new_val or ""))
                line.proposed_formula = new_val

            # FA dir 2026-05-17: typed-formula strings for estimate / forecast
            # parallel to proposed_formula. Persist so re-clicks restore the
            # expression (e.g. edit "*4" → "*3" without retyping).
            for ffield in ("estimate_formula", "forecast_formula"):
                if ffield in line_data:
                    new_val = line_data[ffield] or None
                    old_val = getattr(line, ffield, None) or ""
                    if old_val != (new_val or ""):
                        changes.append((ffield, old_val, new_val or ""))
                    setattr(line, ffield, new_val)

            # Other numeric fields
            for fname in ("prior_year", "ytd_actual", "ytd_budget", "current_budget"):
                if fname in line_data:
                    new_val = float(line_data[fname] or 0)
                    old_val = getattr(line, fname, None) or 0
                    if old_val != new_val:
                        changes.append((fname, str(old_val), str(new_val)))
                    setattr(line, fname, new_val)

            # Write audit trail entries
            for field, old_v, new_v in changes:
                db.session.add(BudgetRevision(
                    budget_id=budget.id, budget_line_id=line.id,
                    action="update", field_name=field,
                    old_value=old_v, new_value=new_v, source="pm",
                    user_id=_read_fa_id_from_cookie(),
                ))

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            import logging
            logging.getLogger(__name__).error(f'PM lines save failed: {e}')
            return jsonify({"error": "Failed to save changes"}), 500

        _resp = budget.to_dict()
        _resp["applied"] = _pm_applied
        _resp["skipped"] = _pm_skipped[:50]
        _resp["skipped_count"] = len(_pm_skipped)
        if _pm_skipped:
            logger.warning("[update_lines] entity=%s applied=%d skipped=%d (%s)",
                           entity_code, _pm_applied, len(_pm_skipped), _pm_skipped[:5])
        return jsonify(_resp)


    # ─── PM R&M review gate: bulk "no change" endpoint ────────────────────
    # FA directive 2026-05-11. The PM can click "Mark all unreviewed R&M
    # as 0% (no change)" in the section header to sweep remaining
    # unreviewed lines after the friction-modal confirms count + $ exposure.
    # Distinct audit trail (`bulk_no_change`) so the FA can see in the diff
    # which lines got individual attention vs. were swept in a bulk pass.

    @bp.route("/api/pm/<entity_code>/rm-bulk-no-change", methods=["POST"])
    def pm_rm_bulk_no_change(entity_code):
        """Mark all unreviewed R&M lines on this budget as no_change.

        Body (optional): {"confirm": true}  — frontend should always send
        this after the user clicks through the confirm modal. Without it,
        we still proceed (the friction is on the FE), but it's logged.

        Response:
          200 {"ok": true, "marked": N, "lines": [<ids>], "skipped": M}
              where `marked` = previously-unreviewed lines now stamped
              `bulk_no_change`, and `skipped` = already-reviewed lines
              (their state is left untouched).
        """
        budget = Budget.query.filter_by(
            entity_code=entity_code, year=BUDGET_YEAR
        ).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        # Editable status check — same gate as update_lines.
        if budget.status not in ("pm_pending", "pm_in_progress", "returned", "fa_review"):
            return jsonify({"error": "Budget is not in editable status"}), 400

        # Pull every R&M line on this budget.
        all_rm = BudgetLine.query.filter_by(
            budget_id=budget.id, sheet_name="Repairs & Supplies"
        ).all()
        unreviewed = [l for l in all_rm if l.pm_review_state is None]
        skipped = len(all_rm) - len(unreviewed)

        if not unreviewed:
            return jsonify({
                "ok": True, "marked": 0, "lines": [],
                "skipped": skipped, "total_rm": len(all_rm),
            })

        # Promote pm_pending → pm_in_progress (same convention as
        # update_lines so the section header counter reflects PM activity).
        if budget.status == "pm_pending":
            budget.status = "pm_in_progress"

        user_id = _read_fa_id_from_cookie()
        now = datetime.utcnow()
        marked_ids = []
        for line in unreviewed:
            old_pct = line.increase_pct
            old_dollar = line.increase_dollar
            line.increase_pct = 0.0
            line.increase_dollar = None
            line.pm_review_state = "bulk_no_change"
            line.pm_reviewed_at = now
            line.pm_reviewed_by = user_id
            marked_ids.append(line.id)
            # One audit row per line so the FA can see exactly which lines
            # the bulk action touched.
            db.session.add(BudgetRevision(
                budget_id=budget.id, budget_line_id=line.id,
                action="pm_bulk_no_change", field_name="pm_review_state",
                old_value="unreviewed", new_value="bulk_no_change",
                source="pm", user_id=user_id,
            ))

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            logger.error("PM bulk no-change save failed: %s", e)
            return jsonify({"error": "Failed to save changes"}), 500

        # Also log a single wizard_events row summarizing the sweep, for
        # the "audit at a glance" view.
        try:
            from app import _log_wizard_event
            _log_wizard_event(
                entity_code, step="pm_review", action="bulk_no_change",
                ok=True,
                payload={
                    "marked": len(marked_ids),
                    "skipped": skipped,
                    "total_rm": len(all_rm),
                    "line_ids": marked_ids,
                },
            )
        except Exception:
            pass

        return jsonify({
            "ok": True,
            "marked": len(marked_ids),
            "lines": marked_ids,
            "skipped": skipped,
            "total_rm": len(all_rm),
        })


    # ─── FA Line Edit & Reclass Endpoints ────────────────────────────────────

    @bp.route("/api/fa-lines/<entity_code>", methods=["PUT"])
    def update_fa_lines(entity_code):
        """FA edits to any budget line (all sheets, not just R&M)."""
        data = request.get_json()
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        # QA fix 2 (2026-07-03): stamp every revision from this request with one
        # batch id so "Undo last" reverts the whole edit, never half of it.
        import uuid as _uuid
        _batch_id = _uuid.uuid4().hex

        # 204 dry run 2026-07-06 (F4): count what actually happened — this
        # endpoint used to 200 "ok" while matching 0 of 69 lines.
        _applied = 0
        _skipped = []
        for line_data in data.get("lines", []):
            line = None
            if line_data.get("gl_code"):
                line = BudgetLine.query.filter_by(
                    budget_id=budget.id, gl_code=line_data.get("gl_code")
                ).first()
            # Suggestion 6 (204 dry run): accept id too — the PM endpoint
            # already does; the identifier asymmetry is exactly how 69 edits
            # once no-op'd in testing.
            if line is None and line_data.get("id"):
                line = BudgetLine.query.get(line_data["id"])
                if line is not None and line.budget_id != budget.id:
                    line = None
            if not line:
                _skipped.append(str(line_data.get("gl_code") or line_data.get("id")))
                continue
            _applied += 1
            # Track changes for audit trail — all editable fields
            changes = []
            editable_float_fields = [
                "increase_pct", "proposed_budget", "accrual_adj", "unpaid_bills",
                "prior_year", "ytd_actual", "ytd_budget", "current_budget"
            ]
            for fname in editable_float_fields:
                if fname in line_data:
                    new_val = float(line_data[fname] or 0)
                    old_val = getattr(line, fname, None) or 0
                    if old_val != new_val:
                        changes.append((fname, str(old_val), str(new_val)))
                    setattr(line, fname, new_val)
            if "notes" in line_data:
                new_val = line_data["notes"]
                if (line.notes or "") != new_val:
                    changes.append(("notes", line.notes or "", new_val))
                line.notes = new_val
            if "proposed_formula" in line_data:
                new_val = line_data["proposed_formula"]  # string or None
                old_val = line.proposed_formula or ""
                if old_val != (new_val or ""):
                    changes.append(("proposed_formula", old_val, new_val or ""))
                line.proposed_formula = new_val or None
            # FA dir 2026-05-17: estimate/forecast formula companions.
            for ffield in ("estimate_formula", "forecast_formula"):
                if ffield in line_data:
                    new_val = line_data[ffield] or None
                    old_val = getattr(line, ffield, None) or ""
                    if old_val != (new_val or ""):
                        changes.append((ffield, old_val, new_val or ""))
                    setattr(line, ffield, new_val)

            # Nullable override fields (null = use formula, number = manual override)
            for ofield in ("estimate_override", "forecast_override"):
                if ofield in line_data:
                    raw = line_data[ofield]
                    new_val = float(raw) if raw is not None else None
                    old_val = getattr(line, ofield, None)
                    if old_val != new_val:
                        changes.append((ofield, str(old_val), str(new_val)))
                    setattr(line, ofield, new_val)

            # Ancillary backup worksheet (JSON list of line items)
            if "backup_json" in line_data:
                raw_bj = line_data["backup_json"]
                if raw_bj is None or raw_bj == "":
                    new_json = None
                elif isinstance(raw_bj, (list, dict)):
                    new_json = json.dumps(raw_bj)
                else:
                    # string — validate it parses
                    try:
                        json.loads(raw_bj)
                        new_json = raw_bj
                    except Exception:
                        new_json = None
                old_json = line.backup_json or ""
                if old_json != (new_json or ""):
                    changes.append(("backup_json", old_json[:200], (new_json or "")[:200]))
                line.backup_json = new_json

            for field, old_v, new_v in changes:
                db.session.add(BudgetRevision(
                    budget_id=budget.id, budget_line_id=line.id,
                    action="update", field_name=field,
                    old_value=old_v, new_value=new_v, source="web",
                    user_id=_read_fa_id_from_cookie(),
                    batch_id=_batch_id,
                ))

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            import logging
            logging.getLogger(__name__).error(f'FA lines save failed: {e}')
            return jsonify({"error": "Failed to save changes"}), 500
        if _skipped:
            logger.warning("[fa-lines] entity=%s applied=%d skipped=%d (%s)",
                           entity_code, _applied, len(_skipped), _skipped[:5])
        return jsonify({"status": "ok", "applied": _applied,
                        "skipped": _skipped[:50], "skipped_count": len(_skipped)})


    @bp.route("/api/lines/<entity_code>/reclass", methods=["PUT"])
    def update_reclass(entity_code):
        """PM suggests reclassifying a GL line (FA acts on it)."""
        data = request.get_json()
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        gl_code = data.get("gl_code")
        line = BudgetLine.query.filter_by(budget_id=budget.id, gl_code=gl_code).first()
        if not line:
            return jsonify({"error": "Line not found"}), 404

        line.reclass_to_gl = data.get("reclass_to_gl") or None
        line.reclass_amount = float(data.get("reclass_amount", 0) or 0)
        line.reclass_notes = data.get("reclass_notes", "")
        db.session.commit()
        return jsonify(line.to_dict())


    # ── B7 (210 FA Notes, 2026-06-17): add an Unmapped GL line to a budget ──
    # The FA's "6744 Parking shows nowhere; reclass doesn't add it to G&A"
    # gap. Reclass only moves $ between two existing GLs; it never put an
    # orphan onto a tab + into the budget. These two endpoints do: the picker
    # lists the building's Summary lines, and map-to-summary appends the GL to
    # the chosen row's prefixes (→ rolls into the budget total) AND moves the
    # line onto that row's detail tab (source_tab) so it shows where expected.
    @bp.route("/api/budget-summary-rows/<entity_code>", methods=["GET"])
    def api_budget_summary_rows(entity_code):
        """Minimal list of DATA Summary rows (mapping targets) for the B7
        'add to budget' picker on the Unmapped tab."""
        rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).order_by(BudgetSummaryRow.display_order).all()
        out = [{"id": r.id, "label": r.label, "section": r.section,
                "source_tab": r.source_tab}
               for r in rows if (r.row_type == "data") and r.label]
        return jsonify({"rows": out})

    @bp.route("/api/summary/<entity_code>/organize-sections", methods=["POST"])
    def api_summary_organize_sections(entity_code):
        """FA 724 #4 (Jennifer 2026-08-18): organize a data-only summary into
        the canonical Income / Expenses / Non-Operating framework with live
        totals. Idempotent — a summary that already has any section_header
        or subtotal row (sectioned or flat-with-totals) is a noop."""
        try:
            from summary_engine import plan_section_organization, apply_section_organization
        except ImportError:
            from budget_app.summary_engine import plan_section_organization, apply_section_organization
        rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).order_by(BudgetSummaryRow.display_order).all()
        if not rows:
            return jsonify({"error": "No summary rows for %s" % entity_code}), 404
        plan = plan_section_organization(rows)
        if plan is None:
            return jsonify({"status": "noop",
                            "reason": "summary already has section/subtotal structure"})
        try:
            result = apply_section_organization(
                db, BudgetSummaryRow, entity_code, BUDGET_YEAR, rows, plan)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": "organize failed: %s" % str(e)[:200]}), 500
        logger.info("[organize-sections] %s: moved=%d inserted=%d",
                    entity_code, result["moved"], result["inserted"])
        return jsonify({"status": "ok", "entity_code": entity_code,
                        "moved": result["moved"], "inserted": result["inserted"]})

    @bp.route("/api/lines/<entity_code>/map-to-summary", methods=["PUT"])
    def map_line_to_summary(entity_code):
        """Map an Unmapped budget line into a Summary row: append the GL to that
        row's gl_prefixes (so it rolls into the budget) and move the line onto
        the row's detail tab. Idempotent on the prefix (dedupes)."""
        import json as _json
        data = request.get_json() or {}
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        gl_code = (data.get("gl_code") or "").strip()
        row_id = data.get("summary_row_id")
        line = BudgetLine.query.filter_by(budget_id=budget.id, gl_code=gl_code).first()
        if not line:
            return jsonify({"error": "Line not found"}), 404
        srow = BudgetSummaryRow.query.filter_by(
            id=row_id, entity_code=entity_code, budget_year=BUDGET_YEAR).first()
        if not srow:
            return jsonify({"error": "Summary row not found"}), 404
        try:
            try:
                prefixes = _json.loads(srow.gl_prefixes_json) if srow.gl_prefixes_json else []
                if not isinstance(prefixes, list):
                    prefixes = []
            except Exception:
                prefixes = []
            prefix_added = False
            if gl_code not in prefixes:
                prefixes.append(gl_code)
                srow.gl_prefixes_json = _json.dumps(prefixes)
                prefix_added = True
            old_sheet = line.sheet_name
            target_sheet = srow.source_tab or line.sheet_name
            line.sheet_name = target_sheet
            line.category = SHEET_TO_CATEGORY.get(target_sheet, line.category or "other")
            db.session.add(BudgetRevision(
                budget_id=budget.id, budget_line_id=line.id,
                action="map_to_summary", field_name="sheet_name",
                old_value=str(old_sheet), new_value=str(target_sheet),
                notes="Mapped %s into summary line '%s' (tab %s)" % (
                    gl_code, srow.label, target_sheet),
                source="web", user_id=_read_fa_id_from_cookie(),
            ))
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": "map failed: %s" % str(e)}), 500
        return jsonify({"status": "ok", "gl_code": gl_code, "summary_row": srow.label,
                        "target_sheet": target_sheet, "prefix_added": prefix_added,
                        "line": line.to_dict()})

    # ── B3 (210 FA Notes, 2026-06-17): interest income above/below the line ──
    # The FA wants interest income (4800 family) positioned per building —
    # operating (above the line) or non-operating (below). On the Summary that
    # placement is the "Interest Income" row's `section`. This flips it.
    @bp.route("/api/budget/<entity_code>/interest-placement", methods=["PUT"])
    def set_interest_placement(entity_code):
        data = request.get_json() or {}
        placement = (data.get("placement") or "").strip().lower()
        section = ("Non-Operating Income"
                   if placement in ("non_operating", "non-operating", "below", "nonop")
                   else "Income")
        rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR).all()
        target = None
        for r in rows:
            if r.row_type != "data":
                continue
            lab = (r.label or "").lower()
            pf = r.gl_prefixes_json or ""
            if "interest" in lab or "4800" in pf:
                target = r
                break
        if not target:
            return jsonify({"error": "Interest Income summary row not found"}), 404
        try:
            old = target.section
            target.section = section
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": "update failed: %s" % str(e)}), 500
        return jsonify({"status": "ok", "label": target.label,
                        "section": section, "old_section": old})


    @bp.route("/api/reclass/accept", methods=["POST"])
    def accept_pm_reclass():
        """FA accepts PM's invoice reclass — moves ytd_actual between GL lines."""
        data = request.get_json()
        entity_code = data.get("entity_code")
        from_gl = data.get("from_gl")
        to_gl = data.get("to_gl")
        amount = float(data.get("amount", 0))

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        from_line = BudgetLine.query.filter_by(budget_id=budget.id, gl_code=from_gl).first()
        to_line = BudgetLine.query.filter_by(budget_id=budget.id, gl_code=to_gl).first()
        if not from_line or not to_line:
            return jsonify({"error": "GL line not found"}), 404

        old_from_ytd = float(from_line.ytd_actual or 0)
        old_to_ytd = float(to_line.ytd_actual or 0)

        from_line.ytd_actual = old_from_ytd - amount
        to_line.ytd_actual = old_to_ytd + amount

        # FA directive 2026-06-03: a reclass must carry the source line's
        # accrual adjustment with it. Forecast = ytd + accrual + unpaid +
        # estimate, so if the actuals move to a new GL but the accrual stays
        # behind, the source forecast is overstated and the target understated.
        # Per FA: move the FULL accrual on any reclass (whole-line semantics).
        # Idempotent against repeat/multi-target reclasses: once the source
        # accrual is drained to 0, later reclasses move nothing.
        old_from_accrual = float(from_line.accrual_adj or 0)
        old_to_accrual = float(to_line.accrual_adj or 0)
        moved_accrual = old_from_accrual
        if moved_accrual:
            from_line.accrual_adj = old_from_accrual - moved_accrual  # → 0.0
            to_line.accrual_adj = old_to_accrual + moved_accrual

        # Audit trail
        _fa_uid = _read_fa_id_from_cookie()
        db.session.add(BudgetRevision(
            budget_id=budget.id, budget_line_id=from_line.id,
            action="reclass_accept", field_name="ytd_actual",
            old_value=str(old_from_ytd), new_value=str(from_line.ytd_actual),
            notes=f"FA accepted reclass of ${amount:,.0f} to {to_gl}", source="web",
            user_id=_fa_uid,
        ))
        db.session.add(BudgetRevision(
            budget_id=budget.id, budget_line_id=to_line.id,
            action="reclass_accept", field_name="ytd_actual",
            old_value=str(old_to_ytd), new_value=str(to_line.ytd_actual),
            notes=f"FA accepted reclass of ${amount:,.0f} from {from_gl}", source="web",
            user_id=_fa_uid,
        ))
        if moved_accrual:
            db.session.add(BudgetRevision(
                budget_id=budget.id, budget_line_id=from_line.id,
                action="reclass_accept", field_name="accrual_adj",
                old_value=str(old_from_accrual), new_value=str(from_line.accrual_adj),
                notes=f"Accrual ${moved_accrual:,.0f} moved with reclass to {to_gl}", source="web",
                user_id=_fa_uid,
            ))
            db.session.add(BudgetRevision(
                budget_id=budget.id, budget_line_id=to_line.id,
                action="reclass_accept", field_name="accrual_adj",
                old_value=str(old_to_accrual), new_value=str(to_line.accrual_adj),
                notes=f"Accrual ${moved_accrual:,.0f} moved with reclass from {from_gl}", source="web",
                user_id=_fa_uid,
            ))

        db.session.commit()

        return jsonify({
            "status": "ok",
            "from_line": from_line.to_dict(),
            "to_line": to_line.to_dict()
        })


    @bp.route("/api/budget-proposal/review", methods=["POST"])
    def review_budget_proposal():
        """FA accepts, rejects, or comments on a PM budget proposal for a GL line."""
        data = request.get_json()
        entity_code = data.get("entity_code")
        gl_code = data.get("gl_code")
        action = data.get("action")          # "accepted", "rejected", "commented"
        note = data.get("note", "")
        override_value = data.get("override_value")  # only for reject

        if action not in ("accepted", "rejected", "commented"):
            return jsonify({"error": "Invalid action"}), 400

        budget = Budget.query.filter_by(entity_code=entity_code).order_by(Budget.id.desc()).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        line = BudgetLine.query.filter_by(budget_id=budget.id, gl_code=gl_code).first()
        if not line:
            return jsonify({"error": "GL line not found"}), 404

        old_status = line.fa_proposed_status or "pending"
        line.fa_proposed_status = action
        line.fa_proposed_note = note

        if action == "rejected" and override_value is not None:
            try:
                override_value = float(override_value)
            except (ValueError, TypeError):
                return jsonify({"error": "Invalid override value"}), 400
            line.fa_override_value = override_value
            # When FA rejects with a custom value, write it as proposed_budget
            line.proposed_budget = override_value
        elif action == "accepted":
            line.fa_override_value = None  # clear any prior override

        # Append to notes for visibility in both dashboards
        timestamp = datetime.utcnow().strftime("%m/%d %H:%M")
        if action == "rejected":
            ov_str = f" | FA override: ${override_value:,.0f}" if override_value is not None else ""
            note_entry = f"[FA REJECTED {timestamp}] {note}{ov_str}"
        elif action == "commented":
            note_entry = f"[FA COMMENT {timestamp}] {note}"
        else:
            note_entry = f"[FA ACCEPTED {timestamp}]"

        existing_notes = line.notes or ""
        line.notes = f"{existing_notes}\n{note_entry}".strip() if existing_notes else note_entry

        # Audit trail
        rev = BudgetRevision(
            budget_id=budget.id,
            budget_line_id=line.id,
            action="fa_proposal_review",
            field_name="fa_proposed_status",
            old_value=old_status,
            new_value=action,
            notes=note or "",
            source="web",
            user_id=_read_fa_id_from_cookie()
        )
        db.session.add(rev)
        db.session.commit()

        return jsonify({"status": "ok", "line": line.to_dict()})


    # ─── Payroll Roster & Assumptions API ────────────────────────────────────

    @bp.route("/api/payroll/positions/<entity_code>", methods=["GET"])
    def get_payroll_positions(entity_code):
        """Get all payroll positions for an entity."""
        positions = PayrollPosition.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).order_by(PayrollPosition.sort_order).all()
        return jsonify([p.to_dict() for p in positions])

    @bp.route("/api/payroll/positions/<entity_code>", methods=["POST"])
    def save_payroll_positions(entity_code):
        """Save/update all payroll positions for an entity (full replace)."""
        try:
            data = request.get_json(silent=True) or {}
        except Exception:
            return jsonify({"error": "Invalid JSON"}), 400
        positions_data = data.get("positions", [])
        if not isinstance(positions_data, list):
            return jsonify({"error": "positions must be a list"}), 400
        # Delete existing and re-insert
        PayrollPosition.query.filter_by(entity_code=entity_code, budget_year=BUDGET_YEAR).delete()
        for i, p in enumerate(positions_data):
            wi_mode_raw = p.get("wage_increase_mode")
            wi_mode = wi_mode_raw if wi_mode_raw in ("pct", "dollar") else None
            wi_val_raw = p.get("wage_increase_value")
            try:
                wi_val = float(wi_val_raw) if wi_val_raw not in (None, "") else None
            except (TypeError, ValueError):
                wi_val = None
            # Extra bonus lines — normalize, validate, json-serialize. Empty list => NULL.
            raw_extras = p.get("extra_bonuses") or []
            clean_extras = []
            if isinstance(raw_extras, list):
                for e in raw_extras:
                    if not isinstance(e, dict):
                        continue
                    basis = e.get("basis")
                    if basis not in ("per_emp", "lump", "pct_wages"):
                        continue
                    try:
                        amt = float(e.get("amount") or 0)
                    except (TypeError, ValueError):
                        amt = 0.0
                    clean_extras.append({
                        "label": str(e.get("label") or "").strip()[:80],
                        "amount": amt,
                        "basis": basis,
                    })
            extras_json = json.dumps(clean_extras) if clean_extras else None
            # FA directive 2026-05-05: benefit_adjustments — validate + serialize.
            # Shape: {adjusted_count, label?, benefits: {<key>:{rate,periods,label?}|null}}
            # Allowed benefit keys (must match the 6 the FA can adjust on the building).
            ALLOWED_BENEFIT_KEYS = {
                "welfare", "pension", "supp_retirement",
                "legal", "training", "profit_sharing",
            }
            raw_adj = p.get("benefit_adjustments")
            adj_json = None
            if isinstance(raw_adj, dict):
                emp_count = int(p.get("employee_count", 0) or 0)
                try:
                    adj_count = int(raw_adj.get("adjusted_count") or 0)
                except (TypeError, ValueError):
                    adj_count = 0
                # Clamp adjusted_count to [0, employee_count]; 0 disables the adjustment.
                if adj_count < 0:
                    adj_count = 0
                if adj_count > emp_count:
                    adj_count = emp_count
                clean_benefits = {}
                raw_benefits = raw_adj.get("benefits") or {}
                if isinstance(raw_benefits, dict):
                    for key, val in raw_benefits.items():
                        if key not in ALLOWED_BENEFIT_KEYS or not isinstance(val, dict):
                            continue
                        try:
                            r = float(val.get("rate") or 0)
                            pp = float(val.get("periods") or 0)
                        except (TypeError, ValueError):
                            continue
                        # Treat zero rate/periods as "no adjustment" — drop the entry.
                        if abs(r) < 1e-9 or abs(pp) < 1e-9:
                            continue
                        clean_benefits[key] = {
                            "rate": r,
                            "periods": pp,
                            "label": str(val.get("label") or "").strip()[:80],
                        }
                # Persist only when the adjustment actually does something.
                if adj_count > 0 and clean_benefits:
                    adj_json = json.dumps({
                        "adjusted_count": adj_count,
                        "label": str(raw_adj.get("label") or "").strip()[:120],
                        "benefits": clean_benefits,
                    })
            pos = PayrollPosition(
                entity_code=entity_code,
                budget_year=BUDGET_YEAR,
                position_name=p.get("position_name", "").strip(),
                employee_count=int(p.get("employee_count", 0) or 0),
                hourly_rate=float(p.get("hourly_rate", 0) or 0),
                bonus_per_employee=float(p.get("bonus_per_employee", 0) or 0),
                additional_weekly=(float(p.get("additional_weekly") or 0) or None),
                effective_week_override=(float(p["effective_week_override"]) if p.get("effective_week_override") not in (None, "", 0) else None),
                wage_increase_mode=wi_mode,
                wage_increase_value=wi_val,
                extra_bonuses_json=extras_json,
                benefit_adjustments_json=adj_json,
                sort_order=i
            )
            db.session.add(pos)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            logger.error(f"save_payroll_positions failed for {entity_code}: {e}", exc_info=True)
            return jsonify({"error": "Failed to save positions"}), 500
        positions = PayrollPosition.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).order_by(PayrollPosition.sort_order).all()
        return jsonify({"status": "ok", "positions": [p.to_dict() for p in positions]})

    @bp.route("/api/payroll/compute/<entity_code>", methods=["GET"])
    def api_payroll_compute(entity_code):
        """Server-side mirror of the payroll tab math (suggestion 3,
        2026-07-06). READ-ONLY parity check: returns the engine components,
        totals, and the GL pushes the roster WOULD make next to what is
        stored — so tab-vs-server drift is visible on any building. The
        browser stays the editor; nothing is written here."""
        try:
            from payroll_engine import compute_payroll, roster_gl_values
        except ImportError:
            from budget_app.payroll_engine import compute_payroll, roster_gl_values
        positions = [p.to_dict() for p in PayrollPosition.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).order_by(PayrollPosition.sort_order).all()]
        pa = PayrollAssumption.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR).first()
        assumptions, overrides = {}, {}
        if pa:
            d = pa.to_dict()
            assumptions = d.get("assumptions") or {}
            overrides = d.get("overrides") or {}
        result = compute_payroll(positions, assumptions, overrides)
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        gl_lines = []
        if budget:
            gl_lines = [l.to_dict() for l in BudgetLine.query.filter_by(
                budget_id=budget.id).filter(BudgetLine.sheet_name == "Payroll").all()]
        result["gl_push_preview"] = roster_gl_values(result["components"], gl_lines)
        result["roster_empty"] = len(positions) == 0
        return jsonify(result)

    @bp.route("/api/payroll/assumptions/<entity_code>", methods=["GET"])
    def get_payroll_assumptions(entity_code):
        """Get payroll-tab-specific assumptions. Falls back to main assumptions if none saved."""
        pa = PayrollAssumption.query.filter_by(entity_code=entity_code, budget_year=BUDGET_YEAR).first()
        if pa:
            return jsonify(pa.to_dict())
        # Fall back: seed from main assumptions tab
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"assumptions": {}})
        import json as _json
        main_a = _json.loads(budget.assumptions_json or "{}")
        # Build payroll-specific structure from main assumptions
        pt = main_a.get("payroll_tax", {})
        ub = main_a.get("union_benefits", {})
        wc = main_a.get("workers_comp", {})
        wi = main_a.get("wage_increase", {})
        _seed_pct = float(wi.get("percent", 0) or 0)
        seeded = {
            "wage_increase_pct": _seed_pct,
            "wage_increase_mode": "pct",
            "wage_increase_value": _seed_pct,
            "effective_week": wi.get("effective_week", "16"),
            "pre_increase_weeks": int(wi.get("pre_increase_weeks", 15) or 15),
            "post_increase_weeks": int(wi.get("post_increase_weeks", 37) or 37),
            "ot_factor": 0.002,
            "vac_sick_hol_factor": 0.10,
            "fica": float(pt.get("FICA", 0) or 0),
            "sui": float(pt.get("SUI", 0) or 0),
            "fui": float(pt.get("FUI", 0) or 0),
            "mta": float(pt.get("MTA", 0) or 0),
            "nys_disability": float(pt.get("NYS_Disability", 0) or 0),
            "pfl": float(pt.get("PFL", 0) or 0),
            "workers_comp": float(wc.get("percent", 0) or 0),
            "welfare_monthly": float(ub.get("welfare_monthly", 0) or 0),
            "pension_weekly": float(ub.get("pension_weekly", 0) or 0),
            "supp_retirement_weekly": float(ub.get("supp_retirement_weekly", 0) or 0),
            "legal_monthly": float(ub.get("legal_monthly", 0) or 0),
            "training_monthly": float(ub.get("training_monthly", 0) or 0),
            "profit_sharing_quarterly": float(ub.get("profit_sharing_quarterly", 0) or 0),
        }
        # No PayrollAssumption row yet — return seeded + empty overrides.
        return jsonify({"assumptions": seeded, "overrides": {}, "source": "main_assumptions"})

    @bp.route("/api/payroll/assumptions/<entity_code>", methods=["POST"])
    def save_payroll_assumptions(entity_code):
        """Save payroll-tab-specific assumptions and/or per-cell overrides.

        Body can contain either or both:
          {"assumptions": {...}}        — full assumption replace
          {"overrides": {...}}          — full overrides replace; pass {} to clear all
          {"override": {key, value}}    — single-cell update; value=null clears that key

        FA directive 2026-05-17: per-cell overrides on green tax/benefit totals.
        """
        try:
            data = request.get_json(silent=True) or {}
        except Exception:
            return jsonify({"error": "Invalid JSON"}), 400
        # 204 dry run 2026-07-06 (F4): an unrecognized payload used to 200 and
        # persist NOTHING. Fail loud instead.
        if not any(k in data for k in ("assumptions", "overrides", "override")):
            return jsonify({"error": "payload must contain 'assumptions', "
                                     "'overrides', or 'override'"}), 400
        import json as _json
        pa = PayrollAssumption.query.filter_by(entity_code=entity_code, budget_year=BUDGET_YEAR).first()
        if not pa:
            pa = PayrollAssumption(entity_code=entity_code, budget_year=BUDGET_YEAR)
            db.session.add(pa)
        # 1. Full assumption replace (legacy path).
        if "assumptions" in data:
            pa.assumptions_json = _json.dumps(data.get("assumptions") or {})
        # 2. Full overrides replace.
        if "overrides" in data:
            ov = data.get("overrides") or {}
            if not isinstance(ov, dict):
                return jsonify({"error": "overrides must be a JSON object"}), 400
            pa.overrides_json = _json.dumps(ov)
        # 3. Single-cell upsert.
        if "override" in data:
            single = data.get("override") or {}
            key = (single.get("key") or "").strip()
            if not key:
                return jsonify({"error": "override.key required"}), 400
            try:
                current_ov = _json.loads(pa.overrides_json or "{}")
            except Exception:
                current_ov = {}
            val = single.get("value")
            if val is None:
                current_ov.pop(key, None)
            else:
                try:
                    current_ov[key] = float(val)
                except (TypeError, ValueError):
                    return jsonify({"error": f"override.value must be numeric (got {val!r})"}), 400
            pa.overrides_json = _json.dumps(current_ov)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            logger.error(f"save_payroll_assumptions failed for {entity_code}: {e}", exc_info=True)
            return jsonify({"error": "Failed to save assumptions"}), 500
        return jsonify({
            "status": "ok",
            "assumptions": _json.loads(pa.assumptions_json or "{}"),
            "overrides": _json.loads(pa.overrides_json or "{}"),
        })


    @bp.route("/api/budget-history/<entity_code>", methods=["GET"])
    def get_budget_history(entity_code):
        """Get change history (revisions) for a budget."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        revisions = BudgetRevision.query.filter_by(budget_id=budget.id)\
            .order_by(BudgetRevision.created_at.desc()).limit(200).all()

        # Enrich with GL code info where applicable
        result = []
        for r in revisions:
            entry = r.to_dict()
            if r.budget_line_id:
                line = BudgetLine.query.get(r.budget_line_id)
                if line:
                    entry["gl_code"] = line.gl_code
                    entry["description"] = line.description
            result.append(entry)

        return jsonify({"revisions": result})


    @bp.route("/api/building-info/<entity_code>", methods=["GET"])
    def get_building_info(entity_code):
        """Fetch reference/illustrative building data (maint history, amort schedule params).

        Also surfaces Budget.building_type so the FA can edit the coop/condo
        flag from the Building Info card (replaces the legacy is_coop heuristic).
        """
        info = BuildingInfo.query.filter_by(entity_code=entity_code).first()
        # Pull the building_type from the Budget record (per-year SoT for is_coop()).
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        building_type = (budget.building_type or "") if budget else ""
        if not info:
            return jsonify({
                "entity_code": entity_code,
                "maintenance_history": None,
                "common_charges_history": None,
                "amort_config": None,
                "building_type": building_type,
                "building_name": (budget.building_name or "") if budget else "",
                "updated_at": None,
                "updated_by": None,
            })
        d = info.to_dict()
        d["building_type"] = building_type
        d["building_name"] = (budget.building_name or "") if budget else ""
        # FA directive 2026-05-10: dedup maintenance/common-charges history
        # by year on read. Auto-heals data damaged by old append paths
        # (168 had 24 maintenance rows for 14 unique years). Helper inlined
        # here because the named helper lives inside the PUT closure.
        def _dedup_inline(rows):
            if not isinstance(rows, list):
                return rows
            seen = {}
            for r in rows:
                if not isinstance(r, dict):
                    continue
                y = r.get("year")
                if y is None:
                    continue
                cur = seen.get(y)
                if cur is None:
                    seen[y] = r
                    continue
                cur_a = float(cur.get("annual") or 0)
                new_a = float(r.get("annual") or 0)
                if abs(new_a) > abs(cur_a):
                    seen[y] = r
            return [seen[y] for y in sorted(seen.keys())]

        if d.get("maintenance_history"):
            cleaned_mh = _dedup_inline(d["maintenance_history"])
            if len(cleaned_mh) != len(d["maintenance_history"]):
                # Persist the cleaned version so it stays clean on next save.
                info.maintenance_history_json = json.dumps(cleaned_mh)
                try:
                    db.session.commit()
                except Exception:
                    db.session.rollback()
            d["maintenance_history"] = cleaned_mh
        if d.get("common_charges_history"):
            cleaned_cc = _dedup_inline(d["common_charges_history"])
            if len(cleaned_cc) != len(d["common_charges_history"]):
                info.common_charges_history_json = json.dumps(cleaned_cc)
                try:
                    db.session.commit()
                except Exception:
                    db.session.rollback()
            d["common_charges_history"] = cleaned_cc
        return jsonify(d)


    @bp.route("/api/building-info/<entity_code>", methods=["PUT"])
    def update_building_info(entity_code):
        """Upsert reference/illustrative building data. Body: {section: value}.

        Also accepts `building_type` ("coop" | "condo" | "" / null) which writes
        to Budget.building_type for the current BUDGET_YEAR — the source of
        truth used by is_coop() to decide whether to render the RE Tax tab.
        """
        try:
            body = request.get_json(force=True) or {}
        except Exception:
            return jsonify({"error": "Invalid JSON"}), 400

        info = BuildingInfo.query.filter_by(entity_code=entity_code).first()
        if not info:
            info = BuildingInfo(entity_code=entity_code)
            db.session.add(info)

        # FA dir 2026-05-19: snapshot the CURRENT state before applying writes.
        # Keeps last 20 versions so the FA can undo accidental wipes (e.g.
        # someone clears maintenance history by mistake). Restore endpoint
        # reads from this list. Skip snapshot on first save (no prior state).
        try:
            had_any_state = bool(
                info.maintenance_history_json
                or info.common_charges_history_json
                or info.amort_config_json
            )
            if had_any_state:
                try:
                    _snaps = json.loads(info.snapshots_json) if info.snapshots_json else []
                    if not isinstance(_snaps, list):
                        _snaps = []
                except Exception:
                    _snaps = []
                # Capture pre-write state
                _snaps.append({
                    "ts": datetime.utcnow().isoformat() + "Z",
                    "by": _read_fa_id_from_cookie() if "_read_fa_id_from_cookie" in dir() else None,
                    "maintenance_history_json": info.maintenance_history_json,
                    "common_charges_history_json": info.common_charges_history_json,
                    "amort_config_json": info.amort_config_json,
                })
                # Cap at 20
                if len(_snaps) > 20:
                    _snaps = _snaps[-20:]
                info.snapshots_json = json.dumps(_snaps)
        except Exception as _e:
            # Never fail the save because of snapshot bookkeeping
            pass

        # FA directive 2026-05-10: dedup maintenance_history and
        # common_charges_history by year. Past versions of the parser
        # / append paths produced duplicate-by-year rows (168 had 24
        # rows for 14 years). When we see two rows for the same year,
        # prefer the one with non-zero `annual` (the computed/illustrative
        # value), falling back to non-zero `monthly` * 12, then to whichever
        # comes first.
        def _dedup_by_year(rows):
            if not isinstance(rows, list):
                return rows
            seen = {}
            for r in rows:
                if not isinstance(r, dict):
                    continue
                y = r.get("year")
                if y is None:
                    continue
                cur = seen.get(y)
                if cur is None:
                    seen[y] = r
                    continue
                # Prefer the one with the more informative `annual`
                cur_annual = float(cur.get("annual") or 0)
                new_annual = float(r.get("annual") or 0)
                if abs(new_annual) > abs(cur_annual):
                    seen[y] = r
                # Else keep cur (already more informative or equally so)
            # Return sorted by year
            return [seen[y] for y in sorted(seen.keys())]

        # Known section fields — silently ignore unknown keys so future sections
        # can be added without breaking existing clients.
        if "maintenance_history" in body:
            mh = body.get("maintenance_history")
            if mh is not None:
                mh = _dedup_by_year(mh)
            info.maintenance_history_json = json.dumps(mh) if mh is not None else None
        if "common_charges_history" in body:
            cc = body.get("common_charges_history")
            if cc is not None:
                cc = _dedup_by_year(cc)
            info.common_charges_history_json = json.dumps(cc) if cc is not None else None
        if "amort_config" in body:
            ac = body.get("amort_config")
            info.amort_config_json = json.dumps(ac) if ac is not None else None

        # FA directive 2026-05-05: building_type editable on Building Info card.
        # Validates to known values; empty string clears it.
        if "building_type" in body:
            raw_bt = body.get("building_type")
            bt = (str(raw_bt).strip().lower() if raw_bt is not None else "")
            if bt and bt not in ("coop", "co-op", "condo", "rental", "mixed", "other"):
                return jsonify({"error": f"Invalid building_type: {bt!r}"}), 400
            # Normalize "co-op" → "coop" for consistency with is_coop().
            if bt == "co-op":
                bt = "coop"
            budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
            if budget:
                budget.building_type = bt or ""

        try:
            user = current_user if hasattr(current_user, "is_authenticated") and current_user.is_authenticated else None
            info.updated_by = (user.username if user else None)
        except Exception:
            pass

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500

        return jsonify({"ok": True, **info.to_dict()})


    # ─── Recent Changes Feed + Undo (FA dir 2026-05-19 Phase 2) ──────────
    # The BudgetRevision table already logs every field-level change (old +
    # new values). These endpoints surface that as a feed in the Health
    # drawer, with a one-click Undo for any entry.

    # FA dir 2026-05-19: Summary tab uses BudgetSummaryRow.col*_override (and
    # col7_proposed_budget). Map the field-name prefix written by
    # api_summary_edit ("col3_override:Maintenance") to the actual DB attr
    # so undo can write the old_value back.
    _SUMMARY_UNDOABLE_FIELDS = {
        "col7": "col7_proposed_budget",
        "col1_override": "col1_override",
        "col2_override": "col2_override",
        "col3_override": "col3_override",
        "col4_override": "col4_override",
        "col5_override": "col5_override",
        "col6_override": "col6_override",
    }

    # Fields safe to undo on BudgetLine. Each entry: {parser_for_old_value, sensitive_flag}
    # Fields not in this list reject the undo (don't pretend to revert
    # something we don't know how to write back).
    _UNDOABLE_FIELDS = {
        "proposed_budget": "float",
        "increase_pct": "float",
        "increase_dollar": "float_nullable",
        "estimate_override": "float_nullable",
        "forecast_override": "float_nullable",
        "estimate_formula": "text_nullable",
        "forecast_formula": "text_nullable",
        "proposed_formula": "text_nullable",
        "accrual_adj": "float",
        "unpaid_bills": "float",
        "current_budget": "float",
        "prior_year": "float",
        "ytd_actual": "float",
        "notes": "text",
        "category": "text",
        "pm_review_state": "text_nullable",
        "fa_proposed_status": "text_nullable",
        "fa_proposed_note": "text",
        "fa_override_value": "float_nullable",
    }

    def _parse_undo_value(raw, kind):
        """Parse a stored old_value string into the right Python type for the
        field. Handles None / empty for nullable fields."""
        if raw is None:
            return None
        s = str(raw).strip()
        if kind == "float":
            try:
                return float(s) if s else 0.0
            except Exception:
                return 0.0
        if kind == "float_nullable":
            if s in ("", "None", "null"):
                return None
            try:
                return float(s)
            except Exception:
                return None
        if kind == "text":
            return "" if s in ("None", "null") else s
        if kind == "text_nullable":
            return None if s in ("", "None", "null") else s
        return s

    @bp.route("/api/recent-changes/<entity_code>", methods=["GET"])
    def get_recent_changes(entity_code):
        """Return the most recent BudgetRevision entries for this building.

        Up to 50 entries by default. Each is augmented with the GL code +
        description from BudgetLine so the UI can show "Maintenance Fees /
        proposed_budget changed from $145,738 → $153,025 by FA at 4:23 PM".

        Query params:
          limit    — max entries returned (default 50)
          sheet    — if set, return only changes whose line is on that sheet
                     (e.g. ?sheet=Income, ?sheet=Payroll). Used by the per-tab
                     Undo/History controls.
        """
        limit = int(request.args.get("limit", "50"))
        sheet_filter = (request.args.get("sheet") or "").strip()
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"changes": []})

        # Three filter modes:
        #   - sheet="Summary" → revisions with action="summary_edit" (no line_id)
        #   - sheet="<other>" → revisions on lines whose sheet_name matches
        #   - no sheet → all revisions for this budget
        q = BudgetRevision.query.filter_by(budget_id=budget.id)
        if sheet_filter == "Summary":
            q = q.filter(BudgetRevision.action == "summary_edit")
        elif sheet_filter:
            line_ids_on_sheet = [
                l.id for l in BudgetLine.query
                    .filter_by(budget_id=budget.id, sheet_name=sheet_filter)
                    .all()
            ]
            if not line_ids_on_sheet:
                return jsonify({"changes": [], "entity_code": entity_code, "sheet": sheet_filter})
            q = q.filter(BudgetRevision.budget_line_id.in_(line_ids_on_sheet))
        revisions = q.order_by(BudgetRevision.id.desc()).limit(limit).all()
        # QA fix 2: batch sizes so the UI can say "this edit changed N fields"
        _batch_counts = {}
        _bids = {r.batch_id for r in revisions if getattr(r, "batch_id", None)}
        if _bids:
            for _bid, _cnt in (db.session.query(BudgetRevision.batch_id, db.func.count(BudgetRevision.id))
                               .filter(BudgetRevision.batch_id.in_(_bids))
                               .group_by(BudgetRevision.batch_id).all()):
                _batch_counts[_bid] = _cnt
        # Build GL lookup for any line_ids referenced
        line_ids = {r.budget_line_id for r in revisions if r.budget_line_id}
        gl_lookup = {}
        if line_ids:
            for l in BudgetLine.query.filter(BudgetLine.id.in_(line_ids)).all():
                gl_lookup[l.id] = {"gl_code": l.gl_code, "description": l.description}
        out = []
        for r in revisions:
            gl_info = gl_lookup.get(r.budget_line_id, {})
            # Detect summary edits — field_name shape "col3_override:Maintenance".
            # Parse so the UI can show "Maintenance / Proposed → $X".
            is_summary = (r.action == "summary_edit") and (":" in (r.field_name or ""))
            sum_col = ""
            sum_label = ""
            if is_summary:
                try:
                    parts = (r.field_name or "").split(":", 1)
                    sum_col = parts[0]   # e.g. "col3_override" or "col7"
                    sum_label = parts[1] # e.g. "Maintenance"
                except Exception:
                    pass
            undoable = (
                (r.budget_line_id is not None
                 and r.action in ("update", "fa_proposal_review")
                 and r.field_name in _UNDOABLE_FIELDS)
                or
                (is_summary and sum_col in _SUMMARY_UNDOABLE_FIELDS)
            )
            out.append({
                "id": r.id,
                "ts": r.created_at.isoformat() + "Z" if r.created_at else None,
                "user_id": r.user_id,
                "action": r.action,
                "field": sum_col if is_summary else (r.field_name or ""),
                "old_value": r.old_value or "",
                "new_value": r.new_value or "",
                "source": r.source or "",
                "notes": r.notes or "",
                "gl_code": gl_info.get("gl_code", "") if not is_summary else "",
                "description": gl_info.get("description", "") if not is_summary else sum_label,
                "summary_label": sum_label,  # for UI to show even when desc is generic
                "is_summary": is_summary,
                "undoable": undoable,
                "batch_id": getattr(r, "batch_id", None),
                "batch_size": _batch_counts.get(getattr(r, "batch_id", None), 1),
            })
        return jsonify({"changes": out, "entity_code": entity_code})

    @bp.route("/api/recent-changes/<entity_code>/undo", methods=["POST"])
    def undo_recent_change(entity_code):
        """Undo one BudgetRevision: write the old_value back to the field.

        Body: {revision_id: N}. Creates a new revision tagged action="undo"
        so the undo itself is auditable + redoable.
        """
        body = request.get_json(silent=True) or {}
        rev_id = body.get("revision_id")
        if not rev_id:
            return jsonify({"error": "revision_id required"}), 400

        rev = BudgetRevision.query.get(rev_id)
        if not rev:
            return jsonify({"error": "Revision not found"}), 404
        budget = Budget.query.get(rev.budget_id)
        if not budget or budget.entity_code != entity_code:
            return jsonify({"error": "Revision does not belong to this entity"}), 400

        # Batch undo (QA fix 2, 2026-07-03): one FA edit can write several
        # revisions in a single request (e.g. an INC% edit writes increase_pct
        # AND proposed_budget). Reverting only the newest field left rows
        # half-undone, so undo now reverts every revision sharing batch_id.
        # Old revisions (batch_id NULL) keep single-revision behavior.
        import uuid as _uuid
        if getattr(rev, "batch_id", None):
            batch_revs = (BudgetRevision.query
                          .filter_by(budget_id=budget.id, batch_id=rev.batch_id)
                          .order_by(BudgetRevision.id.desc()).all())
        else:
            batch_revs = [rev]
        undo_batch_id = _uuid.uuid4().hex

        reverted = []
        skipped = []
        for r in batch_revs:
            if r.action == "summary_edit":
                sum_col, _, sum_label = (r.field_name or "").partition(":")
                if sum_col not in _SUMMARY_UNDOABLE_FIELDS:
                    skipped.append({"revision_id": r.id, "reason": f"Summary column '{sum_col}' not undoable."})
                    continue
                db_attr = _SUMMARY_UNDOABLE_FIELDS[sum_col]
                srow = (
                    BudgetSummaryRow.query
                    .filter_by(entity_code=entity_code, budget_year=budget.year, label=sum_label)
                    .first()
                )
                if not srow:
                    skipped.append({"revision_id": r.id, "reason": f"Summary row '{sum_label}' not found."})
                    continue
                raw_old = r.old_value
                if raw_old in (None, "", "None", "null"):
                    new_value = None
                else:
                    try:
                        new_value = float(raw_old)
                    except Exception:
                        skipped.append({"revision_id": r.id, "reason": f"Cannot parse old value '{raw_old}' as number."})
                        continue
                previous_current = getattr(srow, db_attr, None)
                try:
                    setattr(srow, db_attr, new_value)
                except Exception as e:
                    skipped.append({"revision_id": r.id, "reason": f"Cannot set {db_attr}: {e}"})
                    continue
                db.session.add(BudgetRevision(
                    budget_id=budget.id,
                    action="summary_edit",
                    field_name=r.field_name,  # keep "col3_override:Maintenance" shape
                    old_value=str(previous_current) if previous_current is not None else "",
                    new_value=str(new_value) if new_value is not None else "",
                    notes=f"Undo of revision #{r.id}",
                    source="web",
                    batch_id=undo_batch_id,
                    user_id=_read_fa_id_from_cookie() if "_read_fa_id_from_cookie" in dir() else None,
                ))
                reverted.append({"revision_id": r.id, "field": sum_col,
                                 "summary_label": sum_label, "new_value": new_value})
                continue

            # Line-level revision
            if r.field_name not in _UNDOABLE_FIELDS:
                skipped.append({"revision_id": r.id, "reason": f"Field '{r.field_name}' is not undoable from this UI."})
                continue
            if not r.budget_line_id:
                skipped.append({"revision_id": r.id, "reason": "Revision has no line attached"})
                continue
            line = BudgetLine.query.get(r.budget_line_id)
            if not line:
                skipped.append({"revision_id": r.id, "reason": "Budget line missing"})
                continue
            kind = _UNDOABLE_FIELDS[r.field_name]
            new_value_for_line = _parse_undo_value(r.old_value, kind)
            previous_current = getattr(line, r.field_name, None)
            try:
                setattr(line, r.field_name, new_value_for_line)
            except Exception as e:
                skipped.append({"revision_id": r.id, "reason": f"Cannot set {r.field_name}: {e}"})
                continue
            db.session.add(BudgetRevision(
                budget_id=budget.id,
                budget_line_id=line.id,
                action="undo",
                field_name=r.field_name,
                old_value=str(previous_current) if previous_current is not None else "",
                new_value=str(new_value_for_line) if new_value_for_line is not None else "",
                notes=f"Undo of revision #{r.id}",
                source="web",
                batch_id=undo_batch_id,
                user_id=_read_fa_id_from_cookie() if "_read_fa_id_from_cookie" in dir() else None,
            ))
            reverted.append({"revision_id": r.id, "field": r.field_name,
                             "gl_code": line.gl_code,
                             "description": line.description or "",
                             "new_value": new_value_for_line})

        if not reverted:
            db.session.rollback()
            first_reason = skipped[0]["reason"] if skipped else "Nothing undoable in this batch."
            return jsonify({"error": first_reason, "skipped": skipped}), 400
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)[:200]}), 500
        first = reverted[0]
        return jsonify({
            "ok": True,
            "reverted_count": len(reverted),
            "reverted": reverted,
            "skipped": skipped,
            # Back-compat keys read by existing callers
            "reverted_field": first.get("field"),
            "new_value": first.get("new_value"),
            "gl_code": first.get("gl_code", ""),
            "description": first.get("description", ""),
            "summary_label": first.get("summary_label", ""),
            "undo_batch_id": undo_batch_id,
        })


    # ─── Building Info — Undo / History (FA dir 2026-05-19) ─────────────
    # Snapshot-on-save powers per-building undo. Every PUT pushes the prior
    # state into BuildingInfo.snapshots_json (capped at 20). These endpoints
    # let the FA list snapshots and restore one.

    @bp.route("/api/building-info/<entity_code>/history", methods=["GET"])
    def get_building_info_history(entity_code):
        """List the last N building-info snapshots for this entity.

        Returns a brief summary per snapshot: timestamp, who, what kind of
        data was present (maint history row count, amort config present, etc.).
        The frontend uses this to render the "History" modal with restore
        buttons.
        """
        info = BuildingInfo.query.filter_by(entity_code=entity_code).first()
        if not info or not info.snapshots_json:
            return jsonify({"snapshots": []})
        try:
            snaps = json.loads(info.snapshots_json)
            if not isinstance(snaps, list):
                snaps = []
        except Exception:
            snaps = []

        # Build summaries (oldest → newest, so frontend can show newest first)
        out = []
        for idx, s in enumerate(snaps):
            mh = s.get("maintenance_history_json")
            cc = s.get("common_charges_history_json")
            ac = s.get("amort_config_json")
            def _safe_count(raw):
                try:
                    arr = json.loads(raw) if raw else []
                    return len(arr) if isinstance(arr, list) else 0
                except Exception:
                    return 0
            out.append({
                "index": idx,
                "ts": s.get("ts"),
                "by": s.get("by"),
                "maintenance_rows": _safe_count(mh),
                "common_charges_rows": _safe_count(cc),
                "has_amort_config": bool(ac),
            })
        # Newest first for the UI
        out.reverse()
        return jsonify({"snapshots": out})

    @bp.route("/api/building-info/<entity_code>/restore", methods=["POST"])
    def restore_building_info(entity_code):
        """Restore a snapshot. Body: {snapshot_index: N} where N is the index
        returned by /history. Pushes the CURRENT state into snapshots first
        (so a restore is itself undoable). Then writes the snapshot's JSON
        blobs back to the live columns.
        """
        body = request.get_json(silent=True) or {}
        idx = body.get("snapshot_index")
        if idx is None:
            return jsonify({"error": "snapshot_index required"}), 400
        try:
            idx = int(idx)
        except Exception:
            return jsonify({"error": "snapshot_index must be integer"}), 400

        info = BuildingInfo.query.filter_by(entity_code=entity_code).first()
        if not info:
            return jsonify({"error": "Building info not found"}), 404
        try:
            snaps = json.loads(info.snapshots_json) if info.snapshots_json else []
            if not isinstance(snaps, list):
                snaps = []
        except Exception:
            snaps = []
        if not snaps or idx < 0 or idx >= len(snaps):
            return jsonify({"error": "Snapshot index out of range"}), 400

        target = snaps[idx]

        # Push the CURRENT state as a new snapshot before restoring (so the
        # FA can undo the restore itself if they realize the restore was wrong)
        try:
            snaps.append({
                "ts": datetime.utcnow().isoformat() + "Z",
                "by": _read_fa_id_from_cookie() if "_read_fa_id_from_cookie" in dir() else None,
                "maintenance_history_json": info.maintenance_history_json,
                "common_charges_history_json": info.common_charges_history_json,
                "amort_config_json": info.amort_config_json,
                "note": f"pre-restore snapshot (about to restore snapshot from {target.get('ts')})",
            })
            if len(snaps) > 20:
                snaps = snaps[-20:]
            info.snapshots_json = json.dumps(snaps)
        except Exception:
            pass

        # Apply the target snapshot's blobs
        info.maintenance_history_json = target.get("maintenance_history_json")
        info.common_charges_history_json = target.get("common_charges_history_json")
        info.amort_config_json = target.get("amort_config_json")

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)[:200]}), 500

        return jsonify({"ok": True, "restored_from": target.get("ts"), **info.to_dict()})


    # ─── Readiness Inspector ─────────────────────────────────────────────
    # FA directive 2026-05-09: a single consolidated "what's left on this
    # building" panel at the top of the dashboard. Each gate already has a
    # signal somewhere in the app (audit chip, period banner, orphan
    # warnings, etc.); this endpoint just unifies them into one structured
    # response so the FA can scan in 5 seconds instead of clicking through
    # 8 places.
    @bp.route("/api/readiness/<entity_code>", methods=["GET"])
    def api_readiness(entity_code):
        """Per-building readiness inspector. Returns 9 gates with status +
        click-through actions. Read-only. Safe to call on every dashboard
        load. Gate 8 ("Approved-file labels") added 2026-05-14 — reflects
        portfolio scan findings for this building.

        Response shape:
          {
            "entity_code": "168",
            "summary": {"ok": 6, "warn": 1, "fail": 1, "total": 8, "ready": false},
            "gates": [
              {key, label, status, detail, action_url, action_label}, ...
            ]
          }
        Status values: "ok" (green), "warn" (amber, soft issue), "fail"
        (red, hard blocker), "skip" (n/a for this building).
        """
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        # Pre-fetch everything we need in one go to avoid N+1 queries.
        try:
            assum = json.loads(budget.assumptions_json or "{}")
        except Exception:
            assum = {}
        rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).all()
        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()

        # Audit uploads — query via raw SQL (same pattern as the dashboard
        # endpoint above). Avoids cross-module model-registry coupling.
        audit_rows = []
        try:
            rows_au = db.session.execute(db.text(
                "SELECT id, fiscal_year_end, status, confirmed_at, confirmed_by, "
                "       updated_at, pdf_filename "
                "FROM audit_uploads "
                "WHERE entity_code = :ec "
                "ORDER BY (CASE status WHEN 'confirmed' THEN 4 WHEN 'mapped' THEN 3 "
                "                       WHEN 'extracted' THEN 2 WHEN 'uploaded' THEN 1 ELSE 0 END) DESC, "
                "         updated_at DESC NULLS LAST"
            ), {"ec": entity_code}).fetchall()
            # Wrap rows into a small object exposing the fields we use below
            class _AuRow:
                def __init__(self, r):
                    self.id = r[0]
                    self.fiscal_year_end = r[1]
                    self.status = r[2]
                    self.confirmed_at = r[3]
                    self.confirmed_by = r[4]
                    self.updated_at = r[5]
                    self.pdf_filename = r[6]
            audit_rows = [_AuRow(r) for r in rows_au]
        except Exception as _au_err:
            logger.warning(f"readiness audit lookup failed for {entity_code}: {_au_err}")
            audit_rows = []

        # Payroll positions.
        try:
            PayrollPosition = workflow_models.get("PayrollPosition")
            positions = (
                PayrollPosition.query.filter_by(entity_code=entity_code).all()
                if PayrollPosition else []
            )
        except Exception:
            positions = []

        gates = []

        # ── Gate 1: Source files / Audit upload exists ──
        if audit_rows:
            au = audit_rows[0]
            fy = au.fiscal_year_end or "?"
            gates.append({
                "key": "source_files",
                "label": "Source files found",
                "status": "ok",
                "detail": f"FY{fy} audit uploaded",
                "action_url": "/audited-financials",
                "action_label": "View",
            })
        else:
            gates.append({
                "key": "source_files",
                "label": "Source files found",
                "status": "fail",
                "detail": "No audit uploaded for this entity",
                "action_url": f"/audited-financials/bulk-upload?entity={entity_code}",
                "action_label": "Upload",
            })

        # ── Gate 2: Audit confirmed ──
        confirmed = [au for au in audit_rows if (au.status or "") == "confirmed"]
        latest = audit_rows[0] if audit_rows else None
        if confirmed:
            au = confirmed[0]
            fy = au.fiscal_year_end or "?"
            gates.append({
                "key": "audit_confirmed",
                "label": "Audit confirmed",
                "status": "ok",
                "detail": f"FY{fy} confirmed",
                "action_url": f"/audited-financials/review/{au.id}",
                "action_label": "View",
            })
        elif latest:
            status_label = (latest.status or "uploaded").lower()
            cta = "Run extraction"
            if status_label == "extracted": cta = "Map + Confirm"
            elif status_label == "mapped": cta = "Confirm"
            gates.append({
                "key": "audit_confirmed",
                "label": "Audit confirmed",
                "status": "warn",
                "detail": f"Audit at status '{status_label}' — needs FA action",
                "action_url": f"/audited-financials/review/{latest.id}",
                "action_label": cta,
            })
        else:
            gates.append({
                "key": "audit_confirmed",
                "label": "Audit confirmed",
                "status": "fail",
                "detail": "No audit to confirm",
                "action_url": "/audited-financials/bulk-upload",
                "action_label": "Upload",
            })

        # ── Gate 3: Period set ──
        bp_str = (assum.get("budget_period") or "").strip()
        if bp_str and "/" in bp_str:
            try:
                m_int = int(bp_str.split("/")[0])
                month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                               "Jul","Aug","Sep","Oct","Nov","Dec"]
                actuals_end = month_names[m_int - 1] if 1 <= m_int <= 12 else "?"
                est_start   = month_names[m_int]      if 1 <= m_int <= 11 else "?"
                gates.append({
                    "key": "period_set",
                    "label": "Period set",
                    "status": "ok",
                    "detail": f"Actuals: Jan–{actuals_end} · Estimate: {est_start}–Dec",
                    "action_url": None,
                    "action_label": None,
                })
            except Exception:
                gates.append({
                    "key": "period_set",
                    "label": "Period set",
                    "status": "warn",
                    "detail": f"Period stored as '{bp_str}' (could not parse)",
                    "action_url": None,
                    "action_label": "Fix",
                })
        else:
            gates.append({
                "key": "period_set",
                "label": "Period set",
                "status": "fail",
                "detail": "Forecast will use default 2-month YTD until set",
                "action_url": None,
                "action_label": "Set period",
            })

        # ── Gate 4: Building type set ──
        bt = (budget.building_type or "").strip().lower()
        if bt:
            type_label = {"coop": "Co-op", "condo": "Condo", "rental": "Rental",
                          "mixed": "Mixed-use", "other": "Other"}.get(bt, bt.title())
            gates.append({
                "key": "building_type_set",
                "label": "Building type set",
                "status": "ok",
                "detail": type_label,
                "action_url": None,
                "action_label": None,
            })
        else:
            gates.append({
                "key": "building_type_set",
                "label": "Building type set",
                "status": "warn",
                "detail": "Affects coop/condo math + RE Tax tab visibility",
                "action_url": "#building-info",
                "action_label": "Set type",
            })

        # ── Gate 5: Orphan GLs (cap at high-severity bar) ──
        # Reuse the orphan detection logic from /api/summary — count GLs
        # with non-trivial data (|ytd|>=100 or |current_budget|>=100)
        # not claimed by any summary row.
        matched_gls = set()
        for r in rows:
            if r.row_type != "data" or not r.gl_prefixes_json:
                continue
            try:
                pfx = json.loads(r.gl_prefixes_json)
            except Exception:
                continue
            for ln in lines:
                gl_str = str(ln.gl_code or "").strip()
                gl_base = gl_str.split("-")[0]
                for p in pfx:
                    if "-" in str(p):
                        if gl_str.startswith(str(p)):
                            matched_gls.add(ln.gl_code); break
                    else:
                        if gl_base.startswith(str(p)):
                            matched_gls.add(ln.gl_code); break
        orphan_count = 0
        for ln in lines:
            if ln.gl_code in matched_gls:
                continue
            ytd = float(ln.ytd_actual or 0)
            cb = float(ln.current_budget or 0)
            if abs(ytd) >= 100 or abs(cb) >= 100:
                # Skip placeholder GLs where description == gl_code (stub rows)
                if (ln.description or "").strip() == (ln.gl_code or "").strip():
                    continue
                orphan_count += 1
        if orphan_count == 0:
            gates.append({
                "key": "no_orphans",
                "label": "No orphan GLs",
                "status": "ok",
                "detail": "All GLs aggregated into a summary row",
                "action_url": None,
                "action_label": None,
            })
        else:
            gates.append({
                "key": "no_orphans",
                "label": "No orphan GLs",
                "status": "warn",
                "detail": f"{orphan_count} GL{'s' if orphan_count != 1 else ''} with data not aggregated",
                # FA directive 2026-05-17: deep-link to the orphan banner on
                # the Summary tab so the FA lands on the exact widget, not
                # just on the tab. readinessAction handles #sumOrphans.
                "action_url": "#sumOrphans",
                "action_label": "Review",
            })

        # ── Gate 6: No duplicate-row alerts ──
        # Two summary rows pulling from the same gl_prefixes_json is a
        # guaranteed duplicate. Same as /api/summary warnings detection.
        prefix_to_labels = {}
        for r in rows:
            if r.row_type != "data" or not r.gl_prefixes_json:
                continue
            key = r.gl_prefixes_json.strip()
            if key in ("[]", "null", ""):
                continue
            prefix_to_labels.setdefault(key, []).append(r.label or "?")
        dup_groups = [labels for labels in prefix_to_labels.values() if len(labels) > 1]
        if not dup_groups:
            gates.append({
                "key": "no_duplicates",
                "label": "No duplicate rows",
                "status": "ok",
                "detail": "Each GL prefix maps to one row",
                "action_url": None,
                "action_label": None,
            })
        else:
            n = sum(len(g) for g in dup_groups)
            gates.append({
                "key": "no_duplicates",
                "label": "No duplicate rows",
                "status": "warn",
                "detail": f"{len(dup_groups)} duplicate set{'s' if len(dup_groups) != 1 else ''} ({n} rows)",
                # FA directive 2026-05-17: deep-link to the actual duplicate
                # rows. readinessAction handles #sumDuplicateRows: switches to
                # Summary tab, scrolls to first duplicate, flashes all rows.
                "action_url": "#sumDuplicateRows",
                "action_label": "Review",
            })

        # ── Gate 7: Payroll reviewed ──
        if positions:
            # Count positions with non-zero employee_count and rate.
            valid = sum(
                1 for p in positions
                if (p.employee_count or 0) > 0 and (p.hourly_rate or 0) > 0
            )
            if valid >= 1:
                gates.append({
                    "key": "payroll_reviewed",
                    "label": "Payroll reviewed",
                    "status": "ok",
                    "detail": f"{valid} position{'s' if valid != 1 else ''} configured",
                    "action_url": "#tab=Payroll",
                    "action_label": "View",
                })
            else:
                gates.append({
                    "key": "payroll_reviewed",
                    "label": "Payroll reviewed",
                    "status": "warn",
                    "detail": f"{len(positions)} position{'s' if len(positions) != 1 else ''} present but missing rate/count",
                    "action_url": "#tab=Payroll",
                    "action_label": "Fix",
                })
        else:
            # No payroll at all — could be a building with no staff (small
            # condos sometimes), so flag amber not red.
            gates.append({
                "key": "payroll_reviewed",
                "label": "Payroll reviewed",
                "status": "warn",
                "detail": "No payroll positions configured",
                "action_url": "#tab=Payroll",
                "action_label": "Add",
            })

        # ── Gate 8: Approved-file labels (FA directive 2026-05-14) ──
        # Reads the latest building_scan_findings row for this entity.
        # If no scan has run yet, status="skip" so this gate doesn't block
        # a building that's never been scanned. Once the wizard's
        # scan-findings endpoint runs, this gate reflects the result.
        try:
            label_row = db.session.execute(db.text(
                "SELECT scanned_at, has_file, labels_unmapped, parse_error "
                "FROM building_scan_findings "
                "WHERE entity_code = :ec "
                "ORDER BY scanned_at DESC LIMIT 1"
            ), {"ec": entity_code}).fetchone()
        except Exception:
            label_row = None

        if not label_row:
            gates.append({
                "key": "approved_file_labels",
                "label": "Approved-file labels",
                "status": "skip",
                "detail": "Not yet scanned — open the wizard or run /api/admin/scan-building to check",
                "action_url": None,
                "action_label": None,
            })
        elif label_row[3] and not label_row[1]:  # parse_error + no file
            gates.append({
                "key": "approved_file_labels",
                "label": "Approved-file labels",
                "status": "fail",
                "detail": "No approved 2026 budget file found in SharePoint",
                "action_url": None,
                "action_label": None,
            })
        elif label_row[3]:  # parse_error
            gates.append({
                "key": "approved_file_labels",
                "label": "Approved-file labels",
                "status": "fail",
                "detail": f"Parse error: {str(label_row[3])[:60]}",
                "action_url": None,
                "action_label": None,
            })
        else:
            # FA directive 2026-05-14 Phase 4.5: severity is now computed from
            # the "truly missing" count, not raw unmapped count. Reason: the
            # Excel import auto-creates a summary row for every label it sees.
            # If an unmapped Excel label already has a matching row on the
            # building's summary, that data WILL aggregate on next import —
            # the gate shouldn't scream about it. The "needs new standard row
            # in the master list" cleanup is a separate (admin-level) concern,
            # not a per-building blocker.
            unmapped_n = int(label_row[2] or 0)
            truly_missing_n = unmapped_n
            unmapped_already_on_summary = 0
            if unmapped_n > 0:
                try:
                    import json as _json
                    # Pull the unmapped labels from the scan's stored JSON.
                    ul_row = db.session.execute(db.text(
                        "SELECT unmapped_labels_json FROM building_scan_findings "
                        "WHERE entity_code = :ec ORDER BY scanned_at DESC LIMIT 1"
                    ), {"ec": entity_code}).fetchone()
                    ul_blob = _json.loads((ul_row[0] if ul_row else None) or '{"unmapped":[]}')
                    unmapped_label_set = {
                        str((u.get("label") if isinstance(u, dict) else u) or "").strip().lower()
                        for u in (ul_blob.get("unmapped") or [])
                    }
                    # Pull every label currently on this building's summary
                    # (any row_type — the FA sees the row regardless).
                    rows_now = BudgetSummaryRow.query.filter_by(
                        entity_code=entity_code, budget_year=BUDGET_YEAR
                    ).all()
                    current_label_set = {
                        (r.label or "").strip().lower() for r in rows_now if r.label
                    }
                    unmapped_already_on_summary = sum(
                        1 for l in unmapped_label_set if l in current_label_set
                    )
                    truly_missing_n = max(0, unmapped_n - unmapped_already_on_summary)
                except Exception as _e_truly:
                    logger.warning(
                        f"approved_file_labels: truly-missing calc failed for {entity_code}: {_e_truly}"
                    )

            if truly_missing_n == 0:
                # Two sub-cases:
                #   - unmapped_n == 0: clean import, nothing to fix
                #   - unmapped_n > 0 but all already on summary: data will land,
                #     the only "fix" is adding labels to the master list which
                #     is an admin concern, not an FA blocker.
                if unmapped_n == 0:
                    detail_text = "All labels match standard summary rows — import will be clean"
                else:
                    detail_text = (
                        f"{unmapped_n} label(s) in the approved file aren't in the master list, "
                        f"but every one already has a matching row on this building's summary. "
                        f"Data will aggregate cleanly on import."
                    )
                gates.append({
                    "key": "approved_file_labels",
                    "label": "Approved-file labels",
                    "status": "ok",
                    "detail": detail_text,
                    "action_url": f"/action/{entity_code}#expand-labels" if unmapped_n > 0 else None,
                    "action_label": "View" if unmapped_n > 0 else None,
                })
            elif truly_missing_n <= 2:
                gates.append({
                    "key": "approved_file_labels",
                    "label": "Approved-file labels",
                    "status": "warn",
                    "detail": (
                        f"{truly_missing_n} label(s) in the approved file have no matching summary row yet — "
                        f"add a row or rename in the Excel before import"
                    ),
                    "action_url": f"/action/{entity_code}#expand-labels",
                    "action_label": "Review",
                })
            else:
                gates.append({
                    "key": "approved_file_labels",
                    "label": "Approved-file labels",
                    "status": "fail",
                    "detail": (
                        f"{truly_missing_n} labels have no matching summary row — multiple "
                        f"summary rows would be $0 on import"
                    ),
                    "action_url": f"/action/{entity_code}#expand-labels",
                    "action_label": "Review",
                })

        # ── Gate 9: Generated ──
        bstatus = (budget.status or "").lower()
        if bstatus == "generated":
            gates.append({
                "key": "generated",
                "label": "Budget generated",
                "status": "ok",
                "detail": "Ready to send to FA / export",
                "action_url": None,
                "action_label": None,
            })
        else:
            # Only block (fail) if all upstream gates are green; otherwise warn.
            upstream_ok = all(
                g["status"] == "ok"
                for g in gates
                if g["key"] in ("source_files", "audit_confirmed", "period_set")
            )
            gates.append({
                "key": "generated",
                "label": "Budget generated",
                "status": "warn",
                "detail": (
                    "Ready to generate" if upstream_ok
                    else "Generate when upstream gates are green"
                ),
                "action_url": "#generate",
                "action_label": "Generate" if upstream_ok else None,
            })

        # ── Summary tally ──
        ok_n   = sum(1 for g in gates if g["status"] == "ok")
        warn_n = sum(1 for g in gates if g["status"] == "warn")
        fail_n = sum(1 for g in gates if g["status"] == "fail")

        return jsonify({
            "entity_code": entity_code,
            "budget_year": BUDGET_YEAR,
            "summary": {
                "ok": ok_n,
                "warn": warn_n,
                "fail": fail_n,
                "total": len(gates),
                "ready": (fail_n == 0 and warn_n == 0),
            },
            "gates": gates,
        })


    # ─── Diff-strip + FA identity ────────────────────────────────────────
    # FA directive 2026-05-10: when an FA reopens a building they last
    # visited >24h ago, surface what's changed since (orphan/duplicate
    # deltas, audit status flips, edits by other FAs).
    #
    # Identity model: a `century_fa_id` cookie identifies the FA. Set via
    # POST /api/whoami (writes cookie + validates user exists in `users`
    # with role='fa'). Read via GET /api/whoami. No Flask-Login involved
    # — this is for personalization, not security.
    #
    # Diff endpoint: GET /api/diff/<entity_code> compares current state
    # to the latest building_visits snapshot for this FA. Returns pills
    # (auto/system/fa) for what changed. Inserts a fresh snapshot row
    # ONLY when >24h has elapsed since last visit AND there's something
    # to show (avoids refresh-race).
    #
    # Dismiss endpoint: POST /api/diff/<entity_code>/dismiss flags the
    # latest visit row as dismissed so subsequent calls suppress the
    # strip until something new changes. Wired to beforeunload via
    # navigator.sendBeacon.

    def _fa_signer():
        """Build the itsdangerous signer for the century_fa_id cookie.
        Salts the SECRET_KEY so this signer's tokens can't be confused
        with any other token type (e.g. session, CSRF) signed by Flask."""
        from itsdangerous import URLSafeSerializer
        from flask import current_app
        return URLSafeSerializer(
            current_app.config.get("SECRET_KEY", "century-budget-dev-key"),
            salt="century-fa-id",
        )

    def _read_fa_id_from_cookie():
        """Resolve the current FA's user_id from the century_fa_id cookie.
        Returns None when missing, invalid signature, or unknown user.

        FA directive 2026-05-10: cookie is signed via itsdangerous to
        prevent trivial impersonation. Pre-signing legacy cookies (raw
        integer values) are silently rejected — the user just re-picks
        from the FA modal and gets a freshly-signed cookie.
        """
        from itsdangerous import BadSignature
        try:
            raw = request.cookies.get("century_fa_id")
            if not raw:
                return None
            try:
                uid = int(_fa_signer().loads(raw))
            except (BadSignature, ValueError, TypeError):
                return None
            user = User.query.filter_by(id=uid).first()
            if not user:
                return None
            return uid
        except Exception:
            return None

    def _compute_diff_state(entity_code, budget):
        """Compute the small JSON snapshot we diff against. Cheap: ~3-4
        DB calls. Returns a dict with the same shape we persist to
        building_visits.snapshot_json.
        """
        # Counts of orphan GLs (data not aggregated) and duplicate row
        # groups (two summary rows w/ identical gl_prefixes_json).
        rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).all()
        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()

        matched_gls = set()
        for r in rows:
            if r.row_type != "data" or not r.gl_prefixes_json:
                continue
            try:
                pfx = json.loads(r.gl_prefixes_json)
            except Exception:
                continue
            for ln in lines:
                gl_str = str(ln.gl_code or "").strip()
                gl_base = gl_str.split("-")[0]
                for p in pfx:
                    if "-" in str(p):
                        if gl_str.startswith(str(p)):
                            matched_gls.add(ln.gl_code); break
                    else:
                        if gl_base.startswith(str(p)):
                            matched_gls.add(ln.gl_code); break
        orphan_count = 0
        for ln in lines:
            if ln.gl_code in matched_gls:
                continue
            ytd = float(ln.ytd_actual or 0)
            cb = float(ln.current_budget or 0)
            if abs(ytd) >= 100 or abs(cb) >= 100:
                if (ln.description or "").strip() == (ln.gl_code or "").strip():
                    continue
                orphan_count += 1

        prefix_to_labels = {}
        for r in rows:
            if r.row_type != "data" or not r.gl_prefixes_json:
                continue
            key = r.gl_prefixes_json.strip()
            if key in ("[]", "null", ""):
                continue
            prefix_to_labels.setdefault(key, []).append(r.label or "?")
        duplicate_groups = sum(1 for labels in prefix_to_labels.values() if len(labels) > 1)

        # Audit status (latest by status priority then updated_at)
        audit_status = None
        audit_id = None
        last_audit_confirmed_at = None
        try:
            row_au = db.session.execute(db.text(
                "SELECT id, status, confirmed_at "
                "FROM audit_uploads "
                "WHERE entity_code = :ec "
                "ORDER BY (CASE status WHEN 'confirmed' THEN 4 WHEN 'mapped' THEN 3 "
                "                       WHEN 'extracted' THEN 2 WHEN 'uploaded' THEN 1 ELSE 0 END) DESC, "
                "         updated_at DESC NULLS LAST "
                "LIMIT 1"
            ), {"ec": entity_code}).fetchone()
            if row_au:
                audit_id = row_au[0]
                audit_status = row_au[1]
                last_audit_confirmed_at = (
                    row_au[2].isoformat() if row_au[2] else None
                )
        except Exception:
            pass

        # Max revision id for this budget — lets us cheaply find revisions
        # newer than last visit by id range.
        last_revision_id = 0
        try:
            row_rev = db.session.execute(db.text(
                "SELECT COALESCE(MAX(id), 0) FROM budget_revisions WHERE budget_id = :bid"
            ), {"bid": budget.id}).fetchone()
            if row_rev:
                last_revision_id = int(row_rev[0] or 0)
        except Exception:
            pass

        return {
            "v": 1,
            "orphan_count": int(orphan_count),
            "duplicate_groups": int(duplicate_groups),
            "audit_status": audit_status,
            "audit_id": audit_id,
            "last_revision_id": last_revision_id,
            "last_audit_confirmed_at": last_audit_confirmed_at,
        }

    @bp.route("/api/whoami", methods=["GET"])
    def api_whoami():
        """Return the current FA identity from the century_fa_id cookie.
        No identity → {user_id: null}. Used by the dashboard to decide
        whether to render the FA-name chip + enable edits."""
        uid = _read_fa_id_from_cookie()
        if not uid:
            return jsonify({"user_id": None})
        user = User.query.filter_by(id=uid).first()
        if not user:
            return jsonify({"user_id": None})
        return jsonify({
            "user_id": user.id,
            "name": user.name,
            "email": getattr(user, "email", None) or "",
            "role": getattr(user, "role", None) or "",
        })

    @bp.route("/api/whoami", methods=["POST"])
    def api_set_whoami():
        """Set the century_fa_id cookie. Body: {user_id: int}.
        Validates the user exists. Writes a 90-day max-age cookie.
        """
        try:
            body = request.get_json(force=True) or {}
        except Exception:
            return jsonify({"error": "Invalid JSON"}), 400
        try:
            uid = int(body.get("user_id"))
        except Exception:
            return jsonify({"error": "user_id required"}), 400
        user = User.query.filter_by(id=uid).first()
        if not user:
            return jsonify({"error": "Unknown user_id"}), 404

        resp = jsonify({
            "ok": True,
            "user_id": user.id,
            "name": user.name,
            "role": getattr(user, "role", None) or "",
        })
        # 90-day cookie, SIGNED via itsdangerous (FA directive 2026-05-10).
        # samesite=Lax to allow normal navigation. JS-readable so the chip
        # can refresh live, but the value is opaque (no longer the raw uid).
        signed = _fa_signer().dumps(user.id)
        resp.set_cookie(
            "century_fa_id", signed,
            max_age=90 * 24 * 3600,
            samesite="Lax",
            httponly=False,
            secure=False,
        )
        return resp

    @bp.route("/api/whoami", methods=["DELETE"])
    def api_clear_whoami():
        """Clear the century_fa_id cookie (logout for this FA)."""
        resp = jsonify({"ok": True})
        resp.set_cookie("century_fa_id", "", max_age=0, samesite="Lax")
        return resp

    @bp.route("/api/diff/<entity_code>", methods=["GET"])
    def api_diff(entity_code):
        """Return the diff between current building state and the FA's
        last-visited snapshot. Records a fresh snapshot when >24h elapsed
        AND a strip was actually shown.

        Query params:
          include_dismissed=1  — include rows where diff_dismissed_at is
                                 set (used by "View recent changes" link)

        Response:
          {
            "show": bool,
            "reason": str,           # always returned (debug aid)
            "since": iso str | null, # last visit timestamp
            "has_prev_row": bool,    # whether there's a prior visit
            "pills": [
              {kind, title, body, ts?}, ...
            ]
          }
        """
        uid = _read_fa_id_from_cookie()
        if not uid:
            return jsonify({
                "show": False, "reason": "no_identity",
                "has_prev_row": False, "pills": [],
            })

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        include_dismissed = (request.args.get("include_dismissed") or "") in ("1", "true", "yes")

        # Lazy GC — drop rows older than 90 days for this (user, entity).
        try:
            db.session.execute(db.text(
                "DELETE FROM building_visits "
                "WHERE user_id = :uid AND entity_code = :ec "
                "AND visited_at < (CURRENT_TIMESTAMP - INTERVAL '90 days')"
            ), {"uid": uid, "ec": entity_code})
            db.session.commit()
        except Exception:
            db.session.rollback()

        # Find the latest visit (optionally including dismissed).
        q = BuildingVisit.query.filter_by(user_id=uid, entity_code=entity_code)
        if not include_dismissed:
            q = q.filter(BuildingVisit.diff_dismissed_at.is_(None))
        prev = q.order_by(BuildingVisit.visited_at.desc()).first()

        has_prev_row = (BuildingVisit.query
                        .filter_by(user_id=uid, entity_code=entity_code)
                        .first() is not None)

        # Compute current snapshot (cheap; ~3-4 queries).
        current = _compute_diff_state(entity_code, budget)

        # First-ever visit for this (user, entity) — record + suppress.
        if prev is None:
            try:
                v = BuildingVisit(
                    user_id=uid,
                    entity_code=entity_code,
                    snapshot_json=json.dumps(current),
                )
                db.session.add(v)
                db.session.commit()
            except Exception:
                db.session.rollback()
            return jsonify({
                "show": False, "reason": "first_visit",
                "has_prev_row": False, "pills": [],
            })

        # Recent visit — same FA loaded the page within 24h. Don't bump
        # the timestamp (refresh-race avoidance) and suppress the strip.
        elapsed = (datetime.utcnow() - prev.visited_at).total_seconds()
        if elapsed < (24 * 3600) and not include_dismissed:
            return jsonify({
                "show": False, "reason": "recent_visit",
                "since": prev.visited_at.isoformat() if prev.visited_at else None,
                "has_prev_row": True, "pills": [],
            })

        # Compute pills from the prev → current delta.
        try:
            prev_snap = json.loads(prev.snapshot_json or "{}")
        except Exception:
            prev_snap = {}

        pills = []

        # AUTO pill — audit auto-confirmed by the system since last visit.
        prev_status = (prev_snap.get("audit_status") or "").lower()
        cur_status = (current.get("audit_status") or "").lower()
        if cur_status == "confirmed" and prev_status in ("uploaded", "extracted", "mapped", ""):
            try:
                row_au = db.session.execute(db.text(
                    "SELECT confirmed_by, confirmed_at, fiscal_year_end "
                    "FROM audit_uploads WHERE id = :id"
                ), {"id": current.get("audit_id")}).fetchone()
                if row_au:
                    confirmed_by = (row_au[0] or "").strip().lower()
                    fy = row_au[2] or ""
                    if confirmed_by in ("system", "auto", ""):
                        pills.append({
                            "kind": "auto",
                            "title": "Audit auto-confirmed",
                            "body": (f"FY{fy} audit was auto-confirmed by the system since you were last here."
                                     if fy else "Audit was auto-confirmed by the system since you were last here."),
                            "ts": row_au[1].isoformat() if row_au[1] else None,
                        })
                    else:
                        pills.append({
                            "kind": "fa",
                            "title": "Audit confirmed",
                            "body": (f"FY{fy} audit was confirmed by {row_au[0]} since you were last here."
                                     if fy else f"Audit was confirmed by {row_au[0]} since you were last here."),
                            "ts": row_au[1].isoformat() if row_au[1] else None,
                        })
            except Exception:
                pass

        # SYSTEM pill — orphan-count delta.
        prev_orphans = int(prev_snap.get("orphan_count") or 0)
        cur_orphans = int(current.get("orphan_count") or 0)
        if cur_orphans != prev_orphans:
            direction = "dropped" if cur_orphans < prev_orphans else "increased"
            pills.append({
                "kind": "system",
                "title": "Orphan GL warnings",
                "body": f"Orphan GL warnings {direction}: {prev_orphans} → {cur_orphans}.",
            })

        # SYSTEM pill — duplicate-group delta.
        prev_dups = int(prev_snap.get("duplicate_groups") or 0)
        cur_dups = int(current.get("duplicate_groups") or 0)
        if cur_dups != prev_dups:
            direction = "dropped" if cur_dups < prev_dups else "increased"
            pills.append({
                "kind": "system",
                "title": "Duplicate row groups",
                "body": f"Duplicate row groups {direction}: {prev_dups} → {cur_dups}.",
            })

        # FA pill — edits by another FA since last visit.
        prev_max_rev = int(prev_snap.get("last_revision_id") or 0)
        if prev_max_rev > 0 or current.get("last_revision_id", 0) > prev_max_rev:
            try:
                rows_rev = db.session.execute(db.text(
                    "SELECT br.id, br.user_id, br.field_name, br.old_value, br.new_value, "
                    "       br.action, br.created_at, COALESCE(u.name, '') "
                    "FROM budget_revisions br "
                    "LEFT JOIN users u ON u.id = br.user_id "
                    "WHERE br.budget_id = :bid AND br.id > :prev "
                    "  AND br.user_id IS NOT NULL AND br.user_id != :uid "
                    "  AND COALESCE(br.source, 'web') = 'web' "
                    "ORDER BY br.id ASC LIMIT 5"
                ), {"bid": budget.id, "prev": prev_max_rev, "uid": uid}).fetchall()
                for r in rows_rev:
                    actor = r[7] or "Another FA"
                    field = r[2] or ""
                    old_v = r[3] or ""
                    new_v = r[4] or ""
                    action = r[5] or "update"
                    if action == "update" and field:
                        body = f"{actor} edited {field}"
                        if old_v or new_v:
                            body += f": {old_v or '(empty)'} → {new_v or '(empty)'}"
                        body += "."
                    elif action == "create":
                        body = f"{actor} created a new entry ({field or 'row'})."
                    else:
                        body = f"{actor} {action} on {field or 'this budget'}."
                    pills.append({
                        "kind": "fa",
                        "title": f"Edited by {actor}",
                        "body": body,
                        "ts": r[6].isoformat() if r[6] else None,
                    })
            except Exception:
                pass

        if not pills:
            # Nothing changed since last visit — bump the prev row's
            # visited_at (so we don't keep re-evaluating the same delta
            # forever) and suppress the strip.
            try:
                prev.visited_at = datetime.utcnow()
                prev.snapshot_json = json.dumps(current)
                db.session.commit()
            except Exception:
                db.session.rollback()
            return jsonify({
                "show": False, "reason": "no_changes",
                "since": prev.visited_at.isoformat() if prev.visited_at else None,
                "has_prev_row": True, "pills": [],
            })

        # We have something to show — record a fresh visit row with the
        # current snapshot. The prev row stays as the historical pointer
        # the FA can re-open via "View recent changes".
        since_iso = prev.visited_at.isoformat() if prev.visited_at else None
        try:
            v = BuildingVisit(
                user_id=uid,
                entity_code=entity_code,
                snapshot_json=json.dumps(current),
            )
            db.session.add(v)
            db.session.commit()
        except Exception:
            db.session.rollback()

        return jsonify({
            "show": True, "reason": "diff",
            "since": since_iso,
            "has_prev_row": True,
            "pills": pills,
        })

    @bp.route("/api/diff/<entity_code>/dismiss", methods=["POST"])
    def api_diff_dismiss(entity_code):
        """Mark the latest visit row as dismissed. Called via
        navigator.sendBeacon on beforeunload — fire-and-forget."""
        uid = _read_fa_id_from_cookie()
        if not uid:
            return jsonify({"ok": False, "reason": "no_identity"}), 200
        try:
            db.session.execute(db.text(
                "UPDATE building_visits "
                "SET diff_dismissed_at = CURRENT_TIMESTAMP "
                "WHERE id = ("
                "  SELECT id FROM building_visits "
                "  WHERE user_id = :uid AND entity_code = :ec "
                "  ORDER BY visited_at DESC LIMIT 1"
                ")"
            ), {"uid": uid, "ec": entity_code})
            db.session.commit()
            return jsonify({"ok": True})
        except Exception as e:
            db.session.rollback()
            return jsonify({"ok": False, "error": str(e)}), 500


    @bp.route("/api/download-budget/<entity_code>", methods=["GET"])
    def download_budget(entity_code):
        """Regenerate and download budget Excel from DB data."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
        if not lines:
            return jsonify({"error": "No budget lines found"}), 404

        # Rebuild gl_data dict from budget_lines
        gl_data = {}
        for l in lines:
            gl_data[l.gl_code] = {
                "period_2": l.prior_year or 0,
                "period_3": l.ytd_actual or 0,
                "period_4": l.ytd_budget or 0,
                "period_5": l.current_budget or 0,
            }

        try:
            from template_populator import populate_template
        except ImportError:
            from budget_system.template_populator import populate_template

        import tempfile
        from pathlib import Path as _Path
        from flask import send_file as _send_file

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = _Path(tmpdir) / f"{entity_code}_{budget.building_name}_{BUDGET_YEAR}_Budget.xlsx"
            template_path = _Path(__file__).parent.parent / "budget_system" / "Budget_Final_Template_v2.xlsx"

            property_info = {
                "property_code": entity_code,
                "property_name": budget.building_name,
            }

            # Dynamic YTD months for Excel export
            import json as _json_mod
            _exp_ytd = 2
            try:
                _exp_assumptions = _json_mod.loads(budget.assumptions_json) if budget.assumptions_json else {}
                _exp_bp = _exp_assumptions.get("budget_period", "")
                if "/" in str(_exp_bp):
                    _exp_ytd = int(str(_exp_bp).split("/")[0])
            except Exception:
                pass

            success = populate_template(
                template_path=template_path,
                gl_data=gl_data,
                property_info=property_info,
                output_path=output_path,
                ytd_months=_exp_ytd,
                remaining_months=12 - _exp_ytd,
            )

            if not success or not output_path.exists():
                return jsonify({"error": "Failed to generate Excel"}), 500

            return _send_file(
                output_path,
                as_attachment=True,
                download_name=output_path.name,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


    # ─── New Excel Export (Phase 6, FA directive 2026-05-15) ────────────────
    # Replaces the old /api/download-budget. Starts from the building's OWN
    # approved Excel (from SharePoint), so structure/formulas/cross-sheet
    # refs all carry through. We selectively rewrite the sheets we own
    # in the product. Pass 1a: Commercial Rent & Escalations + DRAFT
    # watermark + FA-edit cell markers. Pass 1b: yrlycomp summary cols.
    # Future passes: Payroll, Energy, yardi_data refresh, maint proof, etc.

    @bp.route("/api/export-excel/<entity_code>", methods=["GET"])
    def api_export_excel(entity_code):
        """Generate a building's Excel budget by starting from its approved
        2026 file and overlaying product data. Streams as download.

        Add ?debug=timing to get a JSON timing report instead of the file.
        """
        import tempfile, traceback, time
        from pathlib import Path as _Path
        from flask import send_file as _send_file
        t0 = time.time()
        timings = {}
        def lap(label):
            timings[label] = round(time.time() - t0, 3)
        debug_timing = request.args.get("debug") == "timing"

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.comments import Comment
        except Exception as e:
            return jsonify({"error": f"openpyxl unavailable: {e!r}"}), 500
        lap("imports")

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        lap("budget_lookup")

        # YTD months actual — drives the Remaining-Projection annualization in
        # the detail tabs (and the yardi input cell). From the budget assumptions
        # (budget_period like "2/10"); defaults to 2.
        export_ytd_months = 2
        try:
            import json as _jm
            _asum = _jm.loads(budget.assumptions_json) if (budget and budget.assumptions_json) else {}
            _bp = _asum.get("budget_period", "")
            if "/" in str(_bp):
                export_ytd_months = int(str(_bp).split("/")[0])
        except Exception:
            pass

        # ── Find template source ─────────────────────────────────────
        # FA directive 2026-05-17: default back to generic template.
        # SharePoint overlay was attempted but openpyxl load+save on the
        # building's full Excel (~10MB, 26 sheets, thousands of formulas)
        # exceeds even a 300s timeout. The full-fidelity SharePoint path
        # needs async job processing OR zipfile/XML surgery (deferred).
        # Use ?source=sharepoint to opt in (still likely to time out).
        file_bytes = None
        template_source = None
        use_sharepoint = request.args.get("source") == "sharepoint"
        if use_sharepoint:
            try:
                import app as _app_mod  # type: ignore
                files = _app_mod._sharepoint_list_approved_budgets(entity_code)
                if files:
                    files.sort(key=lambda f: f.get("last_modified", ""), reverse=True)
                    target = files[0]
                    if target.get("item_id"):
                        _name, file_bytes = _app_mod._sharepoint_download_item(target["item_id"])
                        template_source = f"sharepoint:{target.get('name', '?')}"
            except Exception as _e:
                logger.warning(f"export-excel sharepoint fetch failed for {entity_code}: {_e}")

        if not file_bytes:
            # Default: generic master template (fast, proven by old endpoint)
            try:
                gen = _Path(__file__).parent.parent / "budget_system" / "Budget_Final_Template_v2.xlsx"
                if gen.exists():
                    with open(gen, "rb") as f:
                        file_bytes = f.read()
                    template_source = "generic_master_template"
            except Exception as _e:
                logger.warning(f"export-excel generic template load failed: {_e}")

        if not file_bytes:
            return jsonify({"error": "No template available"}), 500
        lap("template_fetch")

        # ── Pre-populate via template_populator (Pass 2) ─────────────
        # populate_template handles Income/Payroll/Energy/Water/R&S/Gen&Admin
        # by mapping budget_lines GL codes to template cells. This is the
        # existing battle-tested path used by the old endpoint. We layer
        # our own rewrites (Budget Summary + Comm Rent) on top, which
        # OVERWRITE the populator's simplified Budget Summary work.
        in_path = tempfile.mktemp(suffix=".xlsx")
        out_path = tempfile.mktemp(suffix=".xlsx")
        edit_log = []
        try:
            with open(in_path, "wb") as f:
                f.write(file_bytes)
            lap("write_template_tmp")

            # Pre-populate (Pass 2) — only when using the generic template
            # (the populator was built for that file layout).
            if template_source == "generic_master_template":
                try:
                    try:
                        from template_populator import TemplatePopulator
                    except ImportError:
                        from budget_system.template_populator import TemplatePopulator
                    # Build gl_data smartly: query template's GL map FIRST, then
                    # for each line decide: keep as-is if template knows the
                    # exact GL, else aggregate into canonical -0000 form. This
                    # preserves direct matches AND captures sub-account data
                    # the template's specific row mapping would otherwise drop.
                    # FA directive 2026-05-15 Pass 2 (Option A — smart variant).
                    lines = BudgetLine.query.filter_by(budget_id=budget.id).all() if budget else []
                    # Peek at template's known GLs (one-time read; cheap)
                    try:
                        peek = TemplatePopulator(_Path(in_path), _Path(in_path + ".peek"))
                        known_gls = set(peek.gl_mapping.keys())
                        peek.close()
                    except Exception:
                        known_gls = set()
                    gl_data = {}
                    agg_stats = {"raw_lines": 0, "exact_match_kept": 0,
                                  "aggregated_to_canonical": 0,
                                  "template_known_gls": len(known_gls)}
                    for l in lines:
                        full = (l.gl_code or "").strip()
                        if not full:
                            continue
                        agg_stats["raw_lines"] += 1
                        if full in known_gls:
                            target = full
                            agg_stats["exact_match_kept"] += 1
                        else:
                            prefix = full.split("-")[0]
                            if len(prefix) == 4 and prefix.isdigit():
                                target = f"{prefix}-0000"
                            else:
                                target = full
                            agg_stats["aggregated_to_canonical"] += 1
                        if target not in gl_data:
                            gl_data[target] = {"period_2": 0, "period_3": 0, "period_4": 0, "period_5": 0}
                        gl_data[target]["period_2"] += float(l.prior_year or 0)
                        gl_data[target]["period_3"] += float(l.ytd_actual or 0)
                        gl_data[target]["period_4"] += float(l.ytd_budget or 0)
                        gl_data[target]["period_5"] += float(l.current_budget or 0)
                    edit_log.append({"gl_aggregation": agg_stats})
                    property_info = {
                        "property_code": entity_code,
                        "property_name": (budget.building_name if budget else "") or "",
                    }
                    # YTD months from assumptions
                    import json as _json_mod
                    ytd_months = 2
                    try:
                        assumptions = _json_mod.loads(budget.assumptions_json) if (budget and budget.assumptions_json) else {}
                        bp = assumptions.get("budget_period", "")
                        if "/" in str(bp):
                            ytd_months = int(str(bp).split("/")[0])
                    except Exception:
                        pass
                    # populate_template writes to a separate output path
                    populated_path = tempfile.mktemp(suffix=".xlsx")
                    populator = TemplatePopulator(_Path(in_path), _Path(populated_path))
                    success = populator.populate(gl_data, property_info, ytd_months, 12 - ytd_months)
                    if success:
                        populator.save()
                    pop_stats = populator.get_stats()
                    populator.close()
                    edit_log.append({
                        "populate_template": {
                            "success": bool(success),
                            "gl_count_sent": len(gl_data),
                            "matched": pop_stats.get("gl_codes_matched"),
                            "unmatched": pop_stats.get("gl_codes_unmatched"),
                            "cells_filled": pop_stats.get("cells_filled"),
                            "errors": (pop_stats.get("errors") or [])[:5],
                        }
                    })
                    # Reroute: use the populated file as our new working file
                    if _Path(populated_path).exists():
                        in_path = populated_path
                except Exception as e:
                    logger.warning(f"export-excel populate_template failed: {str(e)[:300]} — continuing without it")
            lap("populate_template")

            try:
                wb = openpyxl.load_workbook(in_path, data_only=False)
            except Exception as e:
                return jsonify({"error": f"workbook load failed: {str(e)[:300]}"}), 500
            lap("openpyxl_load")

            # ── Apply rewrites ────────────────────────────────────
            # Foundational source layer FIRST — every other tab references it.
            try:
                _export_write_yardi_data(wb, entity_code, budget, edit_log, ytd_months=export_ytd_months)
            except Exception as e:
                edit_log.append({"sheet": "yardi_data (2)", "error": str(e)[:200]})
                logger.warning(f"export-excel yardi_data write failed: {traceback.format_exc()[-500:]}")
            lap("write_yardi_data")

            # Cover sheet rebuilt next (so it's the title page)
            try:
                _export_rewrite_cover_sheet(wb, entity_code, budget, edit_log)
            except Exception as e:
                edit_log.append({"sheet": "Cover Sheet", "error": str(e)[:200]})
                logger.warning(f"export-excel cover rebuild failed: {traceback.format_exc()[-500:]}")
            lap("rewrite_cover")

            # When using the building's own SharePoint Excel, the original
            # yrlycomp already has 40+ rows of detailed line items with
            # cross-sheet formulas. Replacing it with my product-derived
            # version would LOSE that detail. Instead, only update col7
            # cells in the existing yrlycomp via overlay.
            # When using the generic template, the Budget Summary is
            # simplified — my rewrite produces a richer version that
            # mirrors product detail.
            is_sharepoint = template_source and template_source.startswith("sharepoint:")
            if not is_sharepoint:
                try:
                    _export_rewrite_budget_summary(wb, entity_code, edit_log)
                except Exception as e:
                    edit_log.append({"sheet": "Budget Summary", "error": str(e)[:200]})
                    logger.warning(f"export-excel budget summary rewrite failed: {traceback.format_exc()[-500:]}")
            else:
                # SharePoint path: overlay product col7 onto existing yrlycomp
                try:
                    _export_overlay_yrlycomp_col7(wb, entity_code, edit_log)
                except Exception as e:
                    edit_log.append({"sheet": "yrlycomp overlay", "error": str(e)[:200]})
                    logger.warning(f"export-excel yrlycomp overlay failed: {traceback.format_exc()[-500:]}")
            lap("rewrite_budget_summary")
            # Live RE Taxes tab (co-ops) — defines cbg_re_tax_net used by the
            # Commercial escalation formula below.
            try:
                _export_rewrite_re_taxes(wb, entity_code, budget, edit_log)
            except Exception as e:
                edit_log.append({"sheet": "RE Taxes", "error": str(e)[:200]})
                logger.warning(f"export-excel re-taxes rewrite failed: {traceback.format_exc()[-500:]}")
            lap("rewrite_re_taxes")
            try:
                _export_rewrite_comm_rent(wb, entity_code, edit_log)
            except Exception as e:
                edit_log.append({"sheet": "Comm Rent & Escalations", "error": str(e)[:200]})
                logger.warning(f"export-excel comm rent rewrite failed: {traceback.format_exc()[-500:]}")
            lap("rewrite_comm_rent")
            # Condo CAM Allocation (Schedule A-1) — no-op unless cam_enabled, so
            # non-CAM buildings' exports are unchanged. Ships computed VALUES.
            try:
                _export_rewrite_cam_allocation(wb, entity_code, edit_log)
            except Exception as e:
                edit_log.append({"sheet": "CAM Allocation", "error": str(e)[:200]})
                logger.warning(f"export-excel cam allocation rewrite failed: {traceback.format_exc()[-500:]}")
            lap("rewrite_cam_allocation")

            # Pass 3: rewrite detail tabs from BudgetLine. Each tab gets every
            # line matching its sheet_name filter. Drops dependence on the
            # template's known-GL list — 100% of product GL data flows.
            if budget and not is_sharepoint:
                detail_tabs = [
                    ("Income",             "Income"),
                    ("Payroll",            "Payroll"),
                    ("Energy",             "Energy"),
                    ("Water & Sewer",      "Water & Sewer"),
                    ("Repairs & Supplies", "Repairs & Supplies"),
                    ("Gen & Admin",        "Gen & Admin"),
                    ("Capital",            "Capital"),
                    ("Unmapped",           "Unmapped"),
                ]
                for tab_name, sheet_filter in detail_tabs:
                    try:
                        _export_rewrite_detail_tab(
                            wb, entity_code, budget,
                            tab_name=tab_name,
                            sheet_filter=sheet_filter,
                            edit_log=edit_log,
                            ytd_months=export_ytd_months,
                        )
                    except Exception as e:
                        edit_log.append({"sheet": tab_name, "error": str(e)[:200]})
                        logger.warning(f"export-excel {tab_name} rewrite failed: {traceback.format_exc()[-500:]}")
            lap("rewrite_detail_tabs")

            # DRAFT watermark on yrlycomp if budget isn't approved.
            # Generic path: the rebuilt Budget Summary carries its own DRAFT
            # banner (row 4, no row insert), so we must NOT insert_rows() here —
            # that would shift every live SUM formula out of alignment. Only the
            # SharePoint path (untouched existing yrlycomp) still uses it.
            if is_sharepoint and (budget.status or "").lower() not in ("approved",):
                try:
                    _export_apply_draft_watermark(wb, edit_log)
                except Exception as e:
                    edit_log.append({"sheet": "yrlycomp watermark", "error": str(e)[:200]})
            lap("watermark")

            # Apply Century branding (tab colors, freeze panes, page setup)
            try:
                _export_apply_branding(wb, edit_log)
            except Exception as e:
                edit_log.append({"action": "branding", "error": str(e)[:200]})
            lap("branding")

            # Pass 4 polish — hide empty sheets, turn off gridlines, etc.
            # Run LAST so it can detect which sheets have real data.
            try:
                _export_apply_polish(wb, edit_log)
            except Exception as e:
                edit_log.append({"action": "polish", "error": str(e)[:200]})
            lap("polish")

            if debug_timing:
                return jsonify({
                    "timings_s": timings,
                    "total_s": round(time.time() - t0, 3),
                    "template_source": template_source,
                    "edit_log": edit_log,
                    "skipped_save": True,
                })

            # Force Excel to recalculate every formula on open. openpyxl writes
            # formula strings with NO cached value, and the Railway server has no
            # LibreOffice to pre-compute — without this a client opening the file
            # would see blanks/zeros until pressing F9. fullCalcOnLoad makes
            # desktop Excel + Google Sheets recompute the whole workbook on open.
            # (A browser / Quick Look preview may still show blanks until the
            # file is opened in a real calc engine — noted on the cover.)
            try:
                wb.calculation.fullCalcOnLoad = True
                wb.calculation.calcMode = "auto"
            except Exception as _e:
                edit_log.append({"action": "fullCalcOnLoad", "error": str(_e)[:120]})

            # ── Save + stream ────────────────────────────────────
            wb.save(out_path)
            lap("save")

            # Filename: "148 - 130 East 18 Owners Corp - 2027 Operating Budget.xlsx"
            building_name = (budget.building_name or "Building").strip()
            # Strip filesystem-unsafe chars
            safe_name = "".join(c for c in building_name if c.isalnum() or c in " -_.,")[:80]
            filename = f"{entity_code} - {safe_name} - {BUDGET_YEAR} Operating Budget.xlsx"

            return _send_file(
                out_path,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        finally:
            try:
                import os as _os
                _os.unlink(in_path)
            except Exception:
                pass

    # ─── Excel Export helpers (Pass 1a) ──────────────────────────────────
    # These are module-level (not nested in routes) so future passes can
    # extend them. Kept in this file for proximity to the models.

    def _export_rewrite_detail_tab(wb, entity_code, budget, tab_name,
                                    sheet_filter, edit_log=None,
                                    extra_filter_fn=None, ytd_months=2):
        """Pass 3+4: rewrite a detail tab from BudgetLine data, presentation
        quality. Tight Calibri typography, alt-row cream banding, thin gray
        borders, real number formats with red parens for negatives.
        Yellow fill ONLY on FA-editable inputs (Increase %, Proposed).
        Other cells are display-only with no fill.
        Subtotal row in green band, bold. Single comment on the title.
        FA directive 2026-05-17 (presentation polish).
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.comments import Comment

        # Replace existing
        old_index = None
        for i, name in enumerate(wb.sheetnames):
            if name == tab_name:
                old_index = i
                del wb[name]
                break
        ws = wb.create_sheet(tab_name, index=old_index if old_index is not None else None)

        sheet_filters = sheet_filter if isinstance(sheet_filter, list) else [sheet_filter]
        q = BudgetLine.query.filter_by(budget_id=budget.id).filter(
            BudgetLine.sheet_name.in_(sheet_filters)
        )
        lines = q.order_by(BudgetLine.row_num, BudgetLine.gl_code).all()
        if extra_filter_fn:
            lines = [l for l in lines if extra_filter_fn(l)]

        # ── Style tokens (presentation quality) ──────────────────
        FONT_BODY = Font(name="Calibri", size=10, color="1A1714")
        FONT_BODY_MUTED = Font(name="Calibri", size=10, color="8A7E72")
        FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="001721")
        FONT_SUBTITLE = Font(name="Calibri", size=10, italic=True, color="8A7E72")
        FONT_SUBTOTAL = Font(name="Calibri", size=10, bold=True, color="001721")
        FONT_INPUT = Font(name="Calibri", size=10, color="065F46")
        FONT_FORMULA = Font(name="Calibri", size=10, color="334155")

        FILL_HEADER = PatternFill(start_color="001721", end_color="001721", fill_type="solid")
        FILL_ALT_ROW = PatternFill(start_color="FAFAF7", end_color="FAFAF7", fill_type="solid")
        FILL_INPUT = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        FILL_SUBTOTAL = PatternFill(start_color="E3EAEE", end_color="E3EAEE", fill_type="solid")

        thin_gray = Side(border_style="thin", color="E5E0D5")
        medium_brown = Side(border_style="medium", color="001721")
        ROW_BORDER = Border(bottom=thin_gray)
        HEADER_BORDER = Border(bottom=medium_brown)
        SUBTOTAL_BORDER = Border(top=medium_brown, bottom=medium_brown)

        ALIGN_LEFT = Alignment(horizontal="left", vertical="center", indent=0)
        ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
        ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

        FMT_CURRENCY = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
        FMT_PERCENT = "0.0%;[Red]-0.0%"

        gen_stamp = datetime.utcnow().strftime("%b %d, %Y")
        attribution_comment = Comment(
            f"Generated by Century Budget on {gen_stamp} from product database.",
            "Century",
        )

        # ── Title block (rows 1-3) ───────────────────────────────
        building_name = (budget.building_name if budget else entity_code) or entity_code
        ws.cell(row=1, column=1, value=tab_name).font = FONT_TITLE
        ws.cell(row=2, column=1,
                value=f"{building_name}  ·  Entity {entity_code}  ·  Fiscal Year {BUDGET_YEAR}").font = FONT_SUBTITLE
        ws.cell(row=3, column=1,
                value=f"{len(lines)} general ledger line(s)").font = FONT_SUBTITLE
        # Optional title comment with build stamp
        ws.cell(row=1, column=1).comment = attribution_comment

        # ── Column headers (row 5) ───────────────────────────────
        headers = [
            ("GL Code", "left"),
            ("Description", "left"),
            ("Notes", "left"),
            (f"{BUDGET_YEAR-2}\nActual", "right"),  # prior_year = YSL Col E (period_2) = the prior full year = BUDGET_YEAR-2 (2025), the column just left of the BY-1 YTD column (Col F)
            (f"{BUDGET_YEAR-1}\nYTD Actual", "right"),
            ("Accrual\nAdj", "right"),
            ("Unpaid\nBills", "right"),
            (f"{BUDGET_YEAR-1}\nYTD Budget", "right"),
            ("Remaining\nProjection", "right"),
            (f"{BUDGET_YEAR-1}\n12-Mo Forecast", "right"),
            (f"{BUDGET_YEAR-1}\nCurrent Budget", "right"),
            ("Increase\n%", "right"),
            # FA 2026-06-17: the builder writes an editable Proposed-$ Override
            # column (M) between Increase% and the computed Proposed. The header
            # list had been missing it, so every column from Proposed rightward
            # rendered one cell left of its label (proposed values showed under
            # "$ Var"). Add the label so all 16 columns line up.
            ("Proposed $\nOverride", "right"),
            (f"{BUDGET_YEAR}\nProposed", "right"),
            ("$ Var\nvs Prior", "right"),
            ("% Change", "right"),
        ]
        for col_i, (h, align) in enumerate(headers, start=1):
            c = ws.cell(row=5, column=col_i, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            c.border = HEADER_BORDER
        ws.row_dimensions[5].height = 36

        # ── Data rows ────────────────────────────────────────────
        # Filter out lines with no values at all — skip writing them.
        # The FA can still find them in the (hidden) yardi_data tab.
        # Cleaner than hiding because Excel sometimes ignores openpyxl's
        # hidden flag.
        def line_has_data(l):
            return any([l.prior_year, l.ytd_actual, l.accrual_adj,
                        l.unpaid_bills, l.ytd_budget, l.current_budget,
                        l.proposed_budget])
        visible_lines = [l for l in lines if line_has_data(l)]
        skipped_count = len(lines) - len(visible_lines)
        # QA fix 12 (2026-07-03): the row-3 subtitle counted pre-filter lines
        # ("7 general ledger line(s)" over a 3-row table). Count what ships.
        _subtitle = f"{len(visible_lines)} general ledger line(s)"
        if skipped_count:
            _subtitle += f" ({skipped_count} zero-value line(s) omitted)"
        ws.cell(row=3, column=1, value=_subtitle).font = FONT_SUBTITLE

        # Per-tab calculation flavor: Capital does NOT extrapolate or propose;
        # Payroll annualizes on YTD only (accrual/unpaid excluded). These mirror
        # budget_math exactly so the workbook's live formulas match the app.
        is_capital = (tab_name or "").strip().lower() == "capital"
        is_payroll = (tab_name or "").strip().lower() == "payroll"
        _ym = int(ytd_months or 0)
        # Live reference to the YTD-months input cell on the yardi sheet, so a
        # client editing that one cell re-annualizes every detail line.
        ym_ref = "'yardi_data (2)'!$M$2"
        import budget_math
        # FA 2026-06-17: mirror the dashboard's no-budget rule (B1 prepaid /
        # B4 dividend+messenger) in the export so proposed ships $0 too.
        try:
            from models import gl_is_non_budgeted as _gl_is_non_budgeted
        except ImportError:
            from budget_app.models import gl_is_non_budgeted as _gl_is_non_budgeted

        # FA 2026-06-16 (values-snapshot): the detail tabs now write computed
        # VALUES, not =SUMIF/formula strings. The old formula-only export shipped
        # un-evaluated (no cached result), so any viewer that didn't recalc on
        # open showed blank columns. Values mirror budget_math exactly, so the
        # file opens fully populated everywhere and ties to the app. _subtot
        # accumulates each money column for the TOTAL row.
        _subtot = {}

        r = 6
        for i, l in enumerate(visible_lines):
            alt = (i % 2 == 1)
            row_fill = FILL_ALT_ROW if alt else None

            def set_cell(col, value, *, font=None, align=None, fmt=None, fill=None, border=None):
                cell = ws.cell(row=r, column=col, value=value)
                cell.font = font or FONT_BODY
                cell.alignment = align or ALIGN_RIGHT
                if fmt: cell.number_format = fmt
                if fill: cell.fill = fill
                elif alt: cell.fill = FILL_ALT_ROW
                cell.border = border or ROW_BORDER
                return cell

            # A: GL Code
            set_cell(1, l.gl_code or "", align=ALIGN_LEFT, font=FONT_BODY_MUTED)
            # B: Description
            set_cell(2, l.description or "", align=ALIGN_LEFT)
            # C: Notes
            if l.notes:
                set_cell(3, l.notes, align=ALIGN_LEFT, font=FONT_BODY_MUTED)
            else:
                set_cell(3, None, align=ALIGN_LEFT)
            # D-P: computed VALUES (snapshot), mirroring budget_math + the
            # dashboard rules exactly. (Was =SUMIF/=IF formula strings that
            # shipped un-evaluated → blank in non-recalc viewers.)
            _prior  = float(l.prior_year or 0)
            _ytd    = float(l.ytd_actual or 0)
            _accr   = float(l.accrual_adj or 0)
            _unpaid = float(l.unpaid_bills or 0)
            _ytdbud = float(l.ytd_budget or 0)
            _curbud = float(l.current_budget or 0)
            # D: prior_year
            set_cell(4, _prior, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # E: ytd_actual
            set_cell(5, _ytd, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # F: accrual_adj
            set_cell(6, _accr, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # G: unpaid_bills
            set_cell(7, _unpaid, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # H: ytd_budget
            set_cell(8, _ytdbud, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # Income fixed-forecast pin (Maintenance / Common Charges / Commercial
            # Rent — GL bases 4010/4020/4030/4040): forecast locks to the approved
            # budget, estimate = budget − YTD, proposed = budget. Mirrors the
            # dashboard's isFixedToBudgetLine rule so the workbook ties out.
            _glbase = (l.gl_code or "").split("-")[0].strip()
            _glfull = (l.gl_code or "").strip()
            is_fixed = (not is_capital) and (
                _glbase in FIXED_FORECAST_GL_BASES or _glfull in FIXED_FORECAST_GL_FULL)
            # FA pins so the workbook ties to the product: manual estimate/
            # forecast overrides win (the FA's "or manual" path), RE-tax credit
            # income (4105-4125) posts at year-end → no May-Dec estimate (#B2),
            # and prepaid/dividend/messenger are never budgeted → proposed 0
            # (#B1/#B4). Without these the export re-extrapolated credits and
            # budgeted prepaid, diverging from what the FA sees on the tabs.
            _est_ovr = l.estimate_override
            _fcst_ovr = l.forecast_override
            _is_credit = (l.gl_code or "")[:4] in ("4105", "4110", "4115", "4120", "4125")
            _no_budget = _gl_is_non_budgeted(l.gl_code, entity_code)
            # I: Remaining Projection
            if is_capital:
                _remaining = 0.0
            elif _est_ovr is not None:
                _remaining = float(_est_ovr)
            elif is_fixed:
                _remaining = _curbud - _ytd
            elif _is_credit:
                _remaining = 0.0
            else:
                _remaining = budget_math.estimate(_ytd, _accr, _unpaid, _prior, _ym, payroll=is_payroll)
            set_cell(9, _remaining, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # J: 12-Month Forecast
            if _fcst_ovr is not None:
                _fcst = float(_fcst_ovr)
            elif is_fixed:
                _fcst = _curbud
            elif is_capital:
                _fcst = _ytd + _accr + _unpaid + _remaining
            elif _is_credit:
                _fcst = _ytd + _accr + _unpaid + _remaining
            else:
                _fcst = budget_math.forecast(_ytd, _accr, _unpaid, _prior, _ym, payroll=is_payroll)
            set_cell(10, _fcst, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # K: current_budget
            set_cell(11, _curbud, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # L: Increase % — back-solved from the stored proposal off its forecast
            #    basis (so the displayed % ties to the app). M (override) only when
            #    there is no forecast basis to drive off.
            incr_input = float(l.increase_pct or 0)
            override_val = None
            if is_fixed:
                incr_input = 0.0
            elif not is_capital:
                try:
                    _stored = float(l.proposed_budget) if l.proposed_budget is not None else None
                    if _stored is not None and abs(_stored) > 0.005:
                        if abs(_fcst) > 0.005:
                            incr_input = _stored / _fcst - 1.0
                        else:
                            override_val = _stored
                except Exception:
                    pass
            set_cell(12, incr_input, fmt=FMT_PERCENT, fill=FILL_INPUT, font=FONT_INPUT)
            # M: Proposed $ Override
            set_cell(13, override_val, fmt=FMT_CURRENCY, fill=FILL_INPUT, font=FONT_INPUT)
            # N: Proposed (capital + never-budgeted = 0; fixed=budget×(1+incr);
            #    else override or forecast×(1+incr))
            if is_capital or _no_budget:
                _proposed = 0.0
            elif is_fixed:
                _proposed = _curbud * (1 + incr_input)
            elif override_val is not None:
                _proposed = float(override_val)
            else:
                _proposed = _fcst * (1 + incr_input)
            set_cell(14, _proposed, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # O: $ Var vs prior
            _ovar = _proposed - _prior
            set_cell(15, _ovar, fmt=FMT_CURRENCY, font=FONT_FORMULA)
            # P: % Change
            set_cell(16, ((_proposed / _prior - 1) if abs(_prior) > 0.005 else 0),
                     fmt=FMT_PERCENT, font=FONT_FORMULA)
            # Accumulate money-column subtotals (written as values on TOTAL row)
            for _cc, _vv in ((4, _prior), (5, _ytd), (6, _accr), (7, _unpaid),
                             (8, _ytdbud), (9, _remaining), (10, _fcst),
                             (11, _curbud), (14, _proposed), (15, _ovar)):
                _subtot[_cc] = _subtot.get(_cc, 0.0) + _vv

            r += 1

        # ── Subtotal row ─────────────────────────────────────────
        if r > 6:
            sub_start, sub_end = 6, r - 1
            sub_label = ws.cell(row=r, column=2, value=f"TOTAL {tab_name.upper()}")
            sub_label.font = FONT_SUBTOTAL
            sub_label.alignment = ALIGN_LEFT
            sub_label.fill = FILL_SUBTOTAL
            sub_label.border = SUBTOTAL_BORDER
            ws.cell(row=r, column=1).fill = FILL_SUBTOTAL
            ws.cell(row=r, column=1).border = SUBTOTAL_BORDER
            ws.cell(row=r, column=3).fill = FILL_SUBTOTAL
            ws.cell(row=r, column=3).border = SUBTOTAL_BORDER
            for col_i in [4, 5, 6, 7, 8, 9, 10, 11, 14, 15]:
                sc = ws.cell(row=r, column=col_i,
                              value=round(_subtot.get(col_i, 0.0), 2))
                sc.font = FONT_SUBTOTAL
                sc.fill = FILL_SUBTOTAL
                sc.alignment = ALIGN_RIGHT
                sc.number_format = FMT_CURRENCY
                sc.border = SUBTOTAL_BORDER
            # L (Increase %) + M (Proposed override) — blank in subtotal
            for _blank_col in (12, 13):
                bc = ws.cell(row=r, column=_blank_col)
                bc.fill = FILL_SUBTOTAL
                bc.border = SUBTOTAL_BORDER
            # P = Proposed/Prior - 1 for the subtotal (value)
            _tot_prop = _subtot.get(14, 0.0)
            _tot_prior = _subtot.get(4, 0.0)
            o_cell = ws.cell(row=r, column=16,
                              value=((_tot_prop / _tot_prior - 1) if abs(_tot_prior) > 0.005 else 0))
            o_cell.font = FONT_SUBTOTAL
            o_cell.fill = FILL_SUBTOTAL
            o_cell.alignment = ALIGN_RIGHT
            o_cell.number_format = FMT_PERCENT
            o_cell.border = SUBTOTAL_BORDER

        # ── Column widths + row heights ──────────────────────────
        widths = {"A": 12, "B": 38, "C": 20,
                  "D": 14, "E": 14, "F": 11, "G": 11, "H": 14,
                  "I": 14, "J": 16, "K": 14, "L": 10,
                  "M": 14, "N": 15, "O": 14, "P": 11}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # Freeze
        ws.freeze_panes = "D6"

        # Variance data-bars on $ Var (O) and % Change (P): bar length shows
        # magnitude, direction shows sign (positive right / negative left).
        if r > 6:
            try:
                from openpyxl.formatting.rule import DataBarRule
                for _col in ("O", "P"):
                    ws.conditional_formatting.add(
                        f"{_col}6:{_col}{r-1}",
                        DataBarRule(start_type="min", end_type="max", color="001721"))
            except Exception:
                pass

        if edit_log is not None:
            edit_log.append({
                "sheet": tab_name,
                "filter": sheet_filters,
                "lines_written": len(visible_lines),
                "lines_skipped_empty": skipped_count,
                "lines_total_in_db": len(lines),
            })


    def _export_overlay_yrlycomp_col7(wb, entity_code, edit_log=None):
        """SharePoint path: take the building's existing yrlycomp tab AS-IS
        but overlay product col7_proposed_budget values into the right cells.
        Match rows by label (case-insensitive, whitespace-normalized).

        Detects the col7 column by reading the header row (looks for a year
        match against BUDGET_YEAR or a "Budget" heading). Yellow-fill +
        comment marker on every overlaid cell.
        FA directive 2026-05-17 Phase 2 (SharePoint overlay).
        """
        from openpyxl.styles import PatternFill, Font
        from openpyxl.comments import Comment

        target = None
        for name in wb.sheetnames:
            n = name.lower().strip()
            if n == "yrlycomp" or n.startswith("yrlycomp"):
                target = wb[name]
                break
        if not target:
            return
        EDIT_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        gen_stamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        edit_comment = Comment(
            f"Updated by Century Budget product on {gen_stamp} from budget_summary_rows.col7_proposed_budget.",
            "Century Product",
        )

        # ── Find header row + col7 column ────────────────────────
        # Look for a row whose cells contain the BUDGET_YEAR string or
        # "Budget" heading. Typical yrlycomp header is around row 5-8.
        header_row = None
        col7_col = None
        for r in range(1, min(target.max_row or 0, 20) + 1):
            for c in range(1, min(target.max_column or 0, 16) + 1):
                v = target.cell(row=r, column=c).value
                if isinstance(v, str) and str(BUDGET_YEAR) in v and "budget" in v.lower():
                    header_row = r
                    col7_col = c
                    break
            if header_row is not None:
                break
        if header_row is None:
            edit_log and edit_log.append({"yrlycomp_overlay": "header not detected"})
            return

        # ── Build product col7 lookup by normalized label ────────
        rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).all()
        def norm(s):
            return "".join((s or "").lower().split())
        col7_by_label = {norm(r.label): r.col7_proposed_budget for r in rows
                          if r.col7_proposed_budget is not None}

        # ── Walk yrlycomp rows below the header; match labels ───
        updated = 0
        unmatched_labels = []
        # Labels in yrlycomp are usually in column A or B. Try both.
        for r in range(header_row + 1, (target.max_row or 0) + 1):
            label_val = None
            for c in (1, 2):
                v = target.cell(row=r, column=c).value
                if isinstance(v, str) and len(v.strip()) > 2:
                    label_val = v.strip()
                    break
            if not label_val:
                continue
            key = norm(label_val)
            if key in col7_by_label:
                cell = target.cell(row=r, column=col7_col)
                cell.value = float(col7_by_label[key])
                cell.number_format = "$#,##0"
                cell.fill = EDIT_FILL
                cell.comment = edit_comment
                updated += 1
            else:
                unmatched_labels.append(label_val[:50])

        if edit_log is not None:
            edit_log.append({
                "yrlycomp_overlay": {
                    "header_row": header_row,
                    "col7_column": col7_col,
                    "rows_updated": updated,
                    "unmatched_label_sample": unmatched_labels[:10],
                    "product_labels_total": len(col7_by_label),
                }
            })


    def _export_write_yardi_data(wb, entity_code, budget, edit_log=None, ytd_months=2):
        """Foundational source-layer tab. One row per BudgetLine, all values
        hardcoded. Every other tab in the workbook references this tab via
        SUMIF/SUMIFS — so when an FA opens the Excel and edits a cell here,
        the whole workbook recomputes.
        FA directive 2026-05-17 (Pass 5 — dynamic formulas).
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.comments import Comment

        sheet_name = "yardi_data (2)"
        # Replace existing
        old_index = None
        for i, name in enumerate(wb.sheetnames):
            if name.lower().strip() in ("yardi_data (2)", "yardi_data", "yardi import"):
                old_index = i
                del wb[name]
                break
        ws = wb.create_sheet(sheet_name, index=old_index if old_index is not None else None)

        lines = BudgetLine.query.filter_by(budget_id=budget.id).order_by(
            BudgetLine.sheet_name, BudgetLine.row_num, BudgetLine.gl_code
        ).all() if budget else []

        # Style
        FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        FONT_BODY = Font(name="Calibri", size=10, color="1A1714")
        FONT_BODY_MUTED = Font(name="Calibri", size=9, color="8A7E72")
        FILL_HEADER = PatternFill(start_color="001721", end_color="001721", fill_type="solid")
        FILL_ALT = PatternFill(start_color="FAFAF7", end_color="FAFAF7", fill_type="solid")
        thin = Side(border_style="thin", color="E5E0D5")
        ROW_BORDER = Border(bottom=thin)
        FMT_CURRENCY = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

        # Title
        ws.cell(row=1, column=1, value="Yardi Data — Source Layer").font = Font(
            name="Calibri", size=14, bold=True, color="001721"
        )
        ws.cell(row=2, column=1,
                value=f"Entity {entity_code} · {len(lines)} GL line(s) · Every other tab references this data via SUMIF formulas"
                ).font = Font(name="Calibri", size=9, italic=True, color="8A7E72")

        # Headers row 4
        headers = ["GL Code", "Description", "Sheet", "Category",
                   f"{BUDGET_YEAR-2}\nActual",
                   f"{BUDGET_YEAR-1}\nYTD Actual",
                   "Accrual\nAdj",
                   "Unpaid\nBills",
                   f"{BUDGET_YEAR-1}\nYTD Budget",
                   f"{BUDGET_YEAR-1}\nCurrent Budget",
                   "Notes"]
        for col_i, h in enumerate(headers, start=1):
            c = ws.cell(row=4, column=col_i, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = Alignment(
                horizontal="right" if 5 <= col_i <= 10 else "left",
                vertical="center", wrap_text=True,
            )
            c.border = ROW_BORDER
        ws.row_dimensions[4].height = 32

        # Data rows
        r = 5
        for i, l in enumerate(lines):
            alt = (i % 2 == 1)
            row_fill = FILL_ALT if alt else None

            def cell(col, value, *, font=None, align=None, fmt=None):
                c = ws.cell(row=r, column=col, value=value)
                c.font = font or FONT_BODY
                c.alignment = align or Alignment(horizontal="right", vertical="center")
                if fmt: c.number_format = fmt
                if row_fill: c.fill = row_fill
                c.border = ROW_BORDER
                return c

            cell(1, l.gl_code or "", font=FONT_BODY_MUTED,
                 align=Alignment(horizontal="left", vertical="center"))
            cell(2, l.description or "",
                 align=Alignment(horizontal="left", vertical="center"))
            cell(3, l.sheet_name or "",
                 font=FONT_BODY_MUTED,
                 align=Alignment(horizontal="left", vertical="center"))
            cell(4, l.category or "",
                 font=FONT_BODY_MUTED,
                 align=Alignment(horizontal="left", vertical="center"))
            cell(5, float(l.prior_year) if l.prior_year else None, fmt=FMT_CURRENCY)
            cell(6, float(l.ytd_actual) if l.ytd_actual else None, fmt=FMT_CURRENCY)
            cell(7, float(l.accrual_adj) if l.accrual_adj else None, fmt=FMT_CURRENCY)
            cell(8, float(l.unpaid_bills) if l.unpaid_bills else None, fmt=FMT_CURRENCY)
            cell(9, float(l.ytd_budget) if l.ytd_budget else None, fmt=FMT_CURRENCY)
            cell(10, float(l.current_budget) if l.current_budget else None, fmt=FMT_CURRENCY)
            # Increase% and Proposed are intentionally NOT stored here anymore.
            # They live on the detail tabs (Increase% as an editable input,
            # Proposed as a formula) so there is a single source of truth and
            # every computed value flows from the raw leaf inputs via formulas.
            cell(11, l.notes or "",
                 font=FONT_BODY_MUTED,
                 align=Alignment(horizontal="left", vertical="center", wrap_text=True))
            r += 1

        # Column widths
        widths = {"A": 14, "B": 36, "C": 16, "D": 14,
                  "E": 14, "F": 14, "G": 12, "H": 12, "I": 14, "J": 14,
                  "K": 24}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # Workbook assumption input: YTD months actual. The detail tabs'
        # Remaining-Projection formulas reference this exact cell
        # ('yardi_data (2)'!$M$2), so editing it re-annualizes every line.
        ws.cell(row=1, column=13, value="YTD months actual (input):").font = Font(
            name="Calibri", size=10, bold=True, color="001721")
        ym_cell = ws.cell(row=2, column=13, value=int(ytd_months or 0))
        ym_cell.fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        ym_cell.font = Font(name="Calibri", size=12, bold=True, color="065F46")
        ym_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["M"].width = 18

        ws.freeze_panes = "C5"

        # Hide this tab by default — FAs don't usually need to see raw data.
        # Showing it makes the workbook feel cluttered. It's there for power
        # users who want to inspect or edit the source values.
        ws.sheet_state = "hidden"

        if edit_log is not None:
            edit_log.append({"sheet": sheet_name, "lines_written": len(lines), "hidden": True})


    def _export_rewrite_cover_sheet(wb, entity_code, budget, edit_log=None):
        """Rebuild Cover Sheet as a real presentation cover. Replaces the
        template's generic cover with a clean, branded title page that the
        FA can show first at a board meeting.
        FA directive 2026-05-17 (presentation quality).
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.drawing.image import Image as XLImage
        from pathlib import Path as _P

        # Century brand (sourced from centuryny.com): navy wordmark + red mark.
        NAVY = "001721"; RED = "DE1C23"; INK = "1A1714"; MUTED = "7A8791"
        TINT = "F2F5F6"; RULE = "D7DEE2"
        MONEY = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

        # Find + replace existing cover
        old_index = None
        old_name = None
        for i, name in enumerate(wb.sheetnames):
            if name.lower().strip() in ("cover sheet", "cov", "cover"):
                old_index = i
                old_name = name
                del wb[name]
                break
        ws = wb.create_sheet(old_name or "Cover Sheet", index=0)

        building_name = (budget.building_name if budget else entity_code) or entity_code
        status = (budget.status if budget else None) or "draft"
        approved = status.lower() in ("approved", "fa_approved")
        gen_stamp = datetime.utcnow().strftime("%B %d, %Y")

        for col, w in {"A": 3, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20, "G": 3}.items():
            ws.column_dimensions[col].width = w
        ws.sheet_view.showGridLines = False

        # ── Logo (dark navy wordmark on the white cover) ─────────
        try:
            lp = _P(__file__).parent / "brand" / "century_logo_dark.png"
            if lp.exists():
                img = XLImage(str(lp))
                try:
                    from PIL import Image as _PILImg
                    with _PILImg.open(str(lp)) as _pi:
                        _iw, _ih = _pi.size
                    tw = 280
                    img.width = tw
                    img.height = max(40, int(_ih * tw / _iw))
                except Exception:
                    img.width = 280; img.height = 60
                ws.add_image(img, "B2")
        except Exception:
            pass
        for rr0 in range(1, 6):
            ws.row_dimensions[rr0].height = 16

        # Tagline + red brand rule
        tg = ws.cell(row=6, column=2, value="A New York Property Management Company")
        tg.font = Font(name="Calibri", size=10, color=MUTED)
        ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=6)
        for c in range(2, 7):
            ws.cell(row=7, column=c).border = Border(bottom=Side(border_style="thick", color=RED))
        ws.row_dimensions[7].height = 6

        # Title block
        ws.row_dimensions[8].height = 10
        t = ws.cell(row=9, column=2, value="Operating Budget")
        t.font = Font(name="Calibri", size=34, bold=True, color=NAVY)
        ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=6)
        ws.row_dimensions[9].height = 44
        bn = ws.cell(row=10, column=2, value=building_name)
        bn.font = Font(name="Calibri", size=18, color=NAVY)
        ws.merge_cells(start_row=10, start_column=2, end_row=10, end_column=6)
        ws.row_dimensions[10].height = 26
        fy = ws.cell(row=11, column=2, value=f"Fiscal Year {BUDGET_YEAR}  ·  Entity {entity_code}")
        fy.font = Font(name="Calibri", size=12, color=MUTED)
        ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=6)
        ws.row_dimensions[11].height = 22

        if not approved:
            d = ws.cell(row=12, column=2, value="DRAFT — not yet board-approved")
            d.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            d.fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
            d.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=12, start_column=2, end_row=12, end_column=3)
            ws.row_dimensions[12].height = 22

        # ── KPI cards (live: reference the summary's total cells) ─
        kpis = [("Total revenue", '=IFERROR(cbg_income,"")'),
                ("Total expenses", '=IFERROR(cbg_expenses,"")'),
                ("Net operating", '=IFERROR(cbg_netop,"")'),
                ("Total surplus", '=IFERROR(cbg_surplus,"")')]
        kcol = 2
        for klabel, kformula in kpis:
            lc = ws.cell(row=14, column=kcol, value=klabel.upper())
            lc.font = Font(name="Calibri", size=9, bold=True, color=MUTED)
            lc.fill = PatternFill(start_color=TINT, end_color=TINT, fill_type="solid")
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            vc = ws.cell(row=15, column=kcol, value=kformula)
            vc.font = Font(name="Calibri", size=15, bold=True, color=NAVY)
            vc.number_format = MONEY
            vc.fill = PatternFill(start_color=TINT, end_color=TINT, fill_type="solid")
            vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            kcol += 1
        ws.row_dimensions[14].height = 18
        ws.row_dimensions[15].height = 30

        # ── Building details ─────────────────────────────────────
        info_pairs = [("Entity code", entity_code)]
        try:
            bi = BuildingInfo.query.filter_by(entity_code=entity_code).first()
            if bi:
                parts = []
                if getattr(bi, "address", None): parts.append(bi.address)
                if getattr(bi, "city", None): parts.append(bi.city)
                if parts: info_pairs.append(("Address", ", ".join(parts)))
                if getattr(bi, "building_type", None): info_pairs.append(("Type", bi.building_type))
        except Exception:
            pass
        info_pairs.append(("Status", status.replace("_", " ").title()))
        info_pairs.append(("Prepared", gen_stamp))

        ws.cell(row=17, column=2, value="BUILDING").font = Font(name="Calibri", size=9, bold=True, color=RED)
        rr = 18
        for label, value in info_pairs:
            lc = ws.cell(row=rr, column=2, value=label.upper())
            lc.font = Font(name="Calibri", size=9, bold=True, color=MUTED)
            vc = ws.cell(row=rr, column=3, value=str(value))
            vc.font = Font(name="Calibri", size=11, color=INK)
            ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=6)
            for c in range(2, 7):
                ws.cell(row=rr, column=c).border = Border(bottom=Side(border_style="thin", color=RULE))
            ws.row_dimensions[rr].height = 20
            rr += 1

        # ── Contents (hyperlinks to the always-present sheets) ───
        rr += 1
        ws.cell(row=rr, column=2, value="CONTENTS").font = Font(name="Calibri", size=9, bold=True, color=RED)
        rr += 1
        for nm in ["Budget Summary", "Income", "Payroll", "Energy", "Water & Sewer",
                   "Repairs & Supplies", "Gen & Admin", "Capital"]:
            # E3 values-snapshot: visible text + a real internal hyperlink, not a
            # =HYPERLINK() formula (which showed blank in non-recalc previews).
            hc = ws.cell(row=rr, column=2, value=nm)
            hc.hyperlink = f"#'{nm}'!A1"
            hc.font = Font(name="Calibri", size=11, color="185FA5", underline="single")
            ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
            ws.row_dimensions[rr].height = 18
            rr += 1

        # ── How to read it (legend) ──────────────────────────────
        rr += 1
        ws.cell(row=rr, column=2, value="HOW TO READ THIS WORKBOOK").font = Font(name="Calibri", size=9, bold=True, color=RED)
        rr += 1
        legend = [("Editable input", "FFF7DA", "9A3412", "Type over these (Increase %, assumptions) and everything recalculates."),
                  ("Calculated", "FFFFFF", "334155", "A live formula — recalculates from the inputs and raw data."),
                  ("Subtotal / total", "E3EAEE", NAVY, "SUM and net formulas; no need to edit these directly."),
                  ("Imported actuals", "FFFFFF", INK, "Prior-year, audited 2025, and approved budget — source figures.")]
        for nm, fill, fontc, desc in legend:
            sw = ws.cell(row=rr, column=2, value=nm)
            sw.font = Font(name="Calibri", size=10, bold=True, color=fontc)
            sw.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            sw.alignment = Alignment(horizontal="center", vertical="center")
            _b = Side(border_style="thin", color=RULE)
            sw.border = Border(top=_b, bottom=_b, left=_b, right=_b)
            dc = ws.cell(row=rr, column=3, value=desc)
            dc.font = Font(name="Calibri", size=10, color=INK)
            ws.merge_cells(start_row=rr, start_column=3, end_row=rr, end_column=6)
            ws.row_dimensions[rr].height = 18
            rr += 1

        # ── Footnote ─────────────────────────────────────────────
        rr += 1
        fn = ws.cell(row=rr, column=2,
                     value="Figures populate when opened in Excel or Google Sheets; a quick browser preview may show blanks until the file is opened. Generated by Century Budget Manager from the live database.")
        fn.font = Font(name="Calibri", size=9, italic=True, color=MUTED)
        fn.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=rr, start_column=2, end_row=rr + 2, end_column=6)

        ws.print_options.horizontalCentered = True

        if edit_log is not None:
            edit_log.append({"sheet": ws.title, "action": "cover_rebuilt_branded"})


    def _export_apply_polish(wb, edit_log=None):
        """Pass 4 polish: hide empty sheets, turn off gridlines workbook-wide,
        consistent print setup. Run after all rewrites so we can detect
        which sheets have real data.
        FA directive 2026-05-17 (presentation quality).
        """
        # Sheets we explicitly keep visible even when light on data
        ALWAYS_VISIBLE = {"cover sheet", "budget summary", "yrlycomp",
                          "comm rent & escalations"}
        # Sheets we explicitly hide (template defaults the product doesn't
        # fill from product data). RE Taxes + Insurance Schedule have
        # template structure but no product-populated values today; once
        # those data sources are wired (Phase 5) we'll remove from this list.
        # "re taxes" is no longer force-hidden — _export_rewrite_re_taxes makes
        # it visible (co-ops, live formulas) or hides it itself (non-co-ops).
        ALWAYS_HIDE = {"setup", "contents", "exp-pie",
                       "insurance schedule"}

        hidden = []
        kept = []
        for name in list(wb.sheetnames):
            low = name.lower().strip()
            ws = wb[name]
            # Always-hide list
            if low in ALWAYS_HIDE:
                ws.sheet_state = "hidden"
                hidden.append(name)
                continue
            # Always-visible list
            if low in ALWAYS_VISIBLE:
                ws.sheet_state = "visible"
                kept.append(name)
                continue
            # Empty-sheet detection: count cells with non-trivial data
            # (skip title rows 1-5; look for numeric values or formulas)
            has_data = False
            try:
                for r in range(6, min((ws.max_row or 0) + 1, 200)):
                    for c in range(1, min((ws.max_column or 0) + 1, 16)):
                        v = ws.cell(row=r, column=c).value
                        if isinstance(v, (int, float)) and v != 0:
                            has_data = True; break
                        if isinstance(v, str) and v.startswith("="):
                            has_data = True; break
                        if isinstance(v, str) and len(v.strip()) > 2 and not v.startswith("Total") and not v.startswith("TOTAL"):
                            # Label rows count as data too (lists of GL codes etc.)
                            has_data = True; break
                    if has_data:
                        break
            except Exception:
                has_data = True  # err on visibility
            if has_data:
                kept.append(name)
            else:
                ws.sheet_state = "hidden"
                hidden.append(name)

        # Workbook-wide visual settings
        for name in wb.sheetnames:
            try:
                ws = wb[name]
                ws.sheet_view.showGridLines = False  # cleaner — no grid
                ws.sheet_view.showRowColHeaders = True
                # Print: landscape, fit-to-page width
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.fitToWidth = 1
                # Presentation pages (cover + dashboard) fit on ONE page so charts
                # never split or clip; data sheets fit width and flow tall.
                ws.page_setup.fitToHeight = 1 if name.lower().strip() in ("at a glance", "dashboard", "cover sheet") else 0
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_margins.left = 0.4
                ws.page_margins.right = 0.4
                ws.page_margins.top = 0.6
                ws.page_margins.bottom = 0.5
                ws.page_margins.header = 0.3
                ws.page_margins.footer = 0.3
                # Center horizontally
                ws.print_options.horizontalCentered = True
                # Footer + repeating header rows on print (skip the cover page)
                low = name.lower().strip()
                if low not in ("cover sheet", "cov", "cover"):
                    ws.oddFooter.left.text = "&D"
                    ws.oddFooter.center.text = "CONFIDENTIAL  —  Prepared for the Board"
                    ws.oddFooter.right.text = "Page &P of &N"
                    ws.print_title_rows = "1:5"
            except Exception:
                continue

        # First visible sheet becomes the default active sheet
        for name in wb.sheetnames:
            if wb[name].sheet_state == "visible":
                wb.active = wb.sheetnames.index(name)
                break

        if edit_log is not None:
            edit_log.append({
                "polish": {
                    "kept_visible": kept,
                    "hidden": hidden,
                    "gridlines_off": True,
                }
            })


    def _export_apply_branding(wb, edit_log=None):
        """Apply Century brand styling across the whole workbook.
        - Sheet tab colors (income green, expenses orange, summary brown)
        - Frozen panes on detail sheets (header + GL Code column)
        - Page setup (landscape, fit-to-page width)
        - Default font tweak where not already styled
        FA directive 2026-05-15 Phase 2 polish.
        """
        TAB_COLORS = [
            ("budget summary", "001721"),    # brown — primary
            ("yrlycomp",       "001721"),
            ("summary",        "001721"),
            ("cover sheet",    "001721"),
            ("cov",            "001721"),
            ("setup",          "8A7E72"),    # neutral
            ("contents",       "8A7E72"),
            ("income",         "16A34A"),    # green — income
            ("comm rent",      "16A34A"),
            ("bud",            "001721"),
            ("payroll",        "9A3412"),    # orange — expense detail
            ("pyrl",           "9A3412"),
            ("energy",         "9A3412"),
            ("water",          "9A3412"),
            ("wtrswr",         "9A3412"),
            ("gas",            "9A3412"),
            ("electric",       "9A3412"),
            ("elect",          "9A3412"),
            ("steam",          "9A3412"),
            ("repairs",        "9A3412"),
            ("gen & admin",    "9A3412"),
            ("re taxes",       "9A3412"),
            ("insurance",      "9A3412"),
            ("mortgage",       "9A3412"),
            ("cap",            "1E40AF"),    # blue — capital
            ("exp-pie",        "001721"),
            ("vlookup",        "8A7E72"),
            ("yardi",          "8A7E72"),
            ("maint proof",    "8A7E72"),
            ("unmapped",       "8A7E72"),    # neutral gray — orphan GLs
        ]
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                low = sheet_name.lower().strip()
                for key, color in TAB_COLORS:
                    if key in low:
                        ws.sheet_properties.tabColor = color
                        break
                # Page setup: landscape + fit to 1 wide, no height limit
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                # Margins (narrow)
                ws.page_margins.left = 0.4
                ws.page_margins.right = 0.4
                ws.page_margins.top = 0.5
                ws.page_margins.bottom = 0.5
            except Exception:
                continue

        # Freeze panes on detail sheets (header row + GL column)
        for sheet_name in wb.sheetnames:
            low = sheet_name.lower().strip()
            if any(k in low for k in ("income", "payroll", "energy", "water",
                                       "repairs", "gen & admin", "insurance schedule",
                                       "comm rent", "budget summary", "yrlycomp",
                                       "gas", "electric", "steam")):
                try:
                    ws = wb[sheet_name]
                    if not ws.freeze_panes:
                        ws.freeze_panes = "C6"
                except Exception:
                    pass

        if edit_log is not None:
            edit_log.append({"action": "branding_applied",
                             "tabs_colored": len(wb.sheetnames)})


    def _export_apply_draft_watermark(wb, edit_log=None):
        """Add a DRAFT marker to the summary sheet's top row when the budget
        isn't yet approved. Looks for "yrlycomp" (building-specific Excels)
        OR "Budget Summary" (generic master template) OR the first sheet
        named like a summary. Yellow fill + bold red text.
        """
        from openpyxl.styles import PatternFill, Font
        target = None
        candidates = ("yrlycomp", "budget summary", "summary", "cover sheet")
        for cand in candidates:
            for name in wb.sheetnames:
                if name.lower().strip() == cand:
                    target = wb[name]
                    break
            if target:
                break
        if not target:
            # Last resort: first sheet in the workbook
            if wb.sheetnames:
                target = wb[wb.sheetnames[0]]
        if not target:
            return
        # Insert a row at the top with DRAFT marker
        target.insert_rows(1)
        cell = target.cell(row=1, column=1)
        cell.value = f"DRAFT — exported from product, not yet approved. Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}."
        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        cell.font = Font(bold=True, color="9A3412", size=11)
        # Merge across first ~10 cols if possible (safe even if sheet is narrow)
        try:
            target.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        except Exception:
            pass
        if edit_log is not None:
            edit_log.append({"sheet": "yrlycomp", "action": "draft_watermark"})

    def _export_rewrite_budget_summary(wb, entity_code, edit_log=None):
        """Pass 1b + Pass 4 polish: rewrite the Budget Summary tab with
        full-detail rows from BudgetSummaryRow. Presentation quality:
        clean Calibri typography, alt-row banding, thin gray row borders,
        thick brown header underline, real number formats, restrained yellow
        fill (only on col7 — the FA's editable Proposed Budget column).
        Section headers in brown band, subtotals in green, grand total in
        navy.
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.comments import Comment
        import json as _json

        # Find + replace existing summary sheet
        candidates = ("yrlycomp", "budget summary", "summary")
        old_name = None
        old_index = None
        for i, name in enumerate(wb.sheetnames):
            if name.lower().strip() in candidates:
                old_name = name
                old_index = i
                break
        if old_name is not None:
            del wb[old_name]
        new_title = old_name or "Budget Summary"
        ws = wb.create_sheet(new_title, index=old_index if old_index is not None else 0)

        rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).order_by(BudgetSummaryRow.display_order).all()
        if not rows:
            ws.cell(row=1, column=1, value=f"No summary data for {entity_code} / FY{BUDGET_YEAR}.")
            if edit_log is not None:
                edit_log.append({"sheet": new_title, "rows": 0})
            return

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()

        # ── Style tokens (presentation quality) ──────────────────
        FONT_TITLE = Font(name="Calibri", size=18, bold=True, color="001721")
        FONT_SUBTITLE = Font(name="Calibri", size=10, italic=True, color="8A7E72")
        FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        FONT_BODY = Font(name="Calibri", size=10, color="1A1714")
        FONT_BODY_MUTED = Font(name="Calibri", size=9, color="8A7E72")
        FONT_SECTION = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        FONT_SUBTOTAL = Font(name="Calibri", size=10, bold=True, color="001721")
        FONT_GRAND = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        FONT_INPUT = Font(name="Calibri", size=10, color="065F46", bold=True)
        FONT_FORMULA = Font(name="Calibri", size=10, color="334155")

        FILL_HEADER = PatternFill(start_color="001721", end_color="001721", fill_type="solid")
        FILL_SECTION = PatternFill(start_color="8A7E72", end_color="8A7E72", fill_type="solid")
        FILL_ALT_ROW = PatternFill(start_color="FAFAF7", end_color="FAFAF7", fill_type="solid")
        FILL_INPUT = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        FILL_SUBTOTAL = PatternFill(start_color="E3EAEE", end_color="E3EAEE", fill_type="solid")
        FILL_GRAND = PatternFill(start_color="001721", end_color="001721", fill_type="solid")

        thin_gray = Side(border_style="thin", color="E5E0D5")
        med_brown = Side(border_style="medium", color="001721")
        thick_navy = Side(border_style="medium", color="001721")
        ROW_BORDER = Border(bottom=thin_gray)
        HEADER_BORDER = Border(bottom=med_brown)
        SUBTOTAL_BORDER = Border(top=med_brown, bottom=med_brown)
        GRAND_BORDER = Border(top=thick_navy, bottom=thick_navy)

        ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
        ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
        ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

        FMT_CURRENCY = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
        FMT_PERCENT = "0.0%;[Red]-0.0%"

        gen_stamp = datetime.utcnow().strftime("%b %d, %Y")
        attribution = Comment(
            f"Generated by Century Budget on {gen_stamp} from budget_summary_rows.",
            "Century",
        )

        # ── Title block (rows 1-3) ───────────────────────────────
        building_name = (budget.building_name if budget else entity_code) or entity_code
        title_cell = ws.cell(row=1, column=1, value="Budget Summary")
        title_cell.font = FONT_TITLE
        title_cell.comment = attribution
        ws.cell(row=2, column=1,
                value=f"{building_name}  ·  Entity {entity_code}  ·  Fiscal Year {BUDGET_YEAR}").font = FONT_SUBTITLE
        if budget and budget.status:
            ws.cell(row=3, column=1, value=f"Status: {budget.status}").font = FONT_SUBTITLE

        # ── Column headers (row 5) ───────────────────────────────
        headers = [
            ("GL", "left"),
            ("Line Item", "left"),
            (f"{BUDGET_YEAR-3}\nActual", "right"),
            (f"{BUDGET_YEAR-2}\nActual", "right"),
            (f"{BUDGET_YEAR-1}\nYTD", "right"),
            (f"{BUDGET_YEAR-1}\nEstimate", "right"),
            (f"{BUDGET_YEAR-1}\nForecast", "right"),
            (f"{BUDGET_YEAR-1}\nBudget", "right"),
            (f"{BUDGET_YEAR}\nProposed", "right"),
            ("% Var\nvs Budget", "right"),
        ]
        for col_i, (h, align) in enumerate(headers, start=1):
            c = ws.cell(row=5, column=col_i, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = Alignment(horizontal=align if align != "right" else "right",
                                     vertical="center", wrap_text=True)
            c.border = HEADER_BORDER
        ws.row_dimensions[5].height = 36

        # ── DRAFT banner (row 4, no row insert) when not yet board-approved.
        #    Writing it here instead of insert_rows() keeps the live SUM
        #    formulas below from being shifted out of alignment.
        if not (budget and (budget.status or "").lower() == "approved"):
            _d = ws.cell(row=4, column=1,
                         value=f"DRAFT — not yet board-approved. Generated {datetime.utcnow().strftime('%b %d, %Y')}.")
            _d.font = Font(name="Calibri", size=10, bold=True, color="9A3412")

        # ── Live-formula scaffolding ──────────────────────────────
        # Walk rows in order: accumulate the data rows since the last subtotal
        # (_block) and remember each subtotal's Excel row by category (_subcell).
        # Section totals = SUM of their block; Net Operating and Total Surplus
        # reference the recorded subtotal cells. No static aggregate is written.
        # Works for both the sectioned and flat summary layouts (no dependence
        # on the `section` field).
        _block = []
        _subcell = {}
        # FA 2026-06-16 (values-snapshot): write computed VALUES, not formulas.
        # _blockcols accumulates data-row column sums for the current section
        # (→ section subtotals as values, footing in both layouts). _subcat_vals
        # records each subtotal's per-column value so Net Operating / Total
        # Surplus are arithmetic on them. Data-row values come from the product's
        # own /api/summary (_pl_by_order) so the workbook == the app exactly.
        _blockcols = {}
        _subcat_vals = {}
        _exp_first = None   # first/last Excel row of the contiguous expense
        _exp_last = None    # data block — used to anchor the charts

        def _subtotal_cat(lbl):
            s = (lbl or "").lower()
            if "net operating" in s: return "netop"
            if "surplus" in s or "deficit" in s: return "grand"
            if "non" in s and "income" in s: return "noi"
            if "non" in s and "expense" in s: return "noe"
            if "income" in s: return "income"
            if "expense" in s: return "expenses"
            return None

        def _subtotal_formula(col_letter, cat):
            if cat == "netop":
                inc, exp = _subcell.get("income"), _subcell.get("expenses")
                return f"={col_letter}{inc}-{col_letter}{exp}" if (inc and exp) else None
            if cat == "grand":
                netop = _subcell.get("netop")
                if netop:
                    base = f"={col_letter}{netop}"
                else:
                    inc, exp = _subcell.get("income"), _subcell.get("expenses")
                    base = f"={col_letter}{inc}-{col_letter}{exp}" if (inc and exp) else None
                if base is None:
                    return None
                if _subcell.get("noi"): base += f"+{col_letter}{_subcell['noi']}"
                if _subcell.get("noe"): base += f"-{col_letter}{_subcell['noe']}"
                return base
            if _block:
                return "=SUM(" + ",".join(f"{col_letter}{rr}" for rr in _block) + ")"
            return None

        # GL head (4-digit) -> set of detail tab names holding it, so Summary
        # rows SUMIFS over the DETAIL TABS (live, client-editable) by GL prefix.
        _detail_tab_names = {"Income", "Payroll", "Energy", "Water & Sewer",
                             "Repairs & Supplies", "Gen & Admin", "Capital", "Unmapped"}
        _prefix_to_sheets = {}
        try:
            for _bl in (BudgetLine.query.filter_by(budget_id=budget.id).all() if budget else []):
                _gc = (_bl.gl_code or "").strip()
                _sn = (_bl.sheet_name or "").strip()
                if not _gc or _sn not in _detail_tab_names:
                    continue
                _hd = _gc.split("-")[0][:4]
                if _hd:
                    _prefix_to_sheets.setdefault(_hd, set()).add(_sn)
        except Exception:
            _prefix_to_sheets = {}

        # 2025 Actual (col2) is audit-sourced (not GL-summable, not stored on the
        # row). Reuse the dashboard's exact computation (api_summary_get) so the
        # workbook's 2025 column ties to the app. Also keep col3/4/5 as a static
        # fallback for rows that have no detail-tab home. Heavy-ish but one call
        # per export; failure degrades gracefully (col2 blank, col3-5 still live).
        _pl_by_order = {}
        try:
            _resp = api_summary_get(entity_code)
            _pl = _resp.get_json() if hasattr(_resp, "get_json") else None
            if isinstance(_pl, dict):
                for _pr in (_pl.get("rows") or []):
                    _do = _pr.get("display_order")
                    if _do is not None:
                        _pl_by_order[_do] = _pr
        except Exception:
            _pl_by_order = {}

        # ── Data rows ────────────────────────────────────────────
        r = 6
        for i, row in enumerate(rows):
            row_label = row.label or ""

            # Determine row classification
            is_section = (row.row_type == "section_header")
            is_subtotal = (row.row_type == "subtotal")
            is_grand = is_subtotal and ("surplus" in row_label.lower()
                                         or "deficit" in row_label.lower()) and "non" not in row_label.lower()
            alt = (i % 2 == 1) and not is_section and not is_subtotal

            if is_section:
                for col_i in range(1, 11):
                    sc = ws.cell(row=r, column=col_i)
                    sc.fill = FILL_SECTION
                    sc.border = ROW_BORDER
                sc = ws.cell(row=r, column=2, value=row_label.upper())
                sc.font = FONT_SECTION
                sc.alignment = ALIGN_LEFT
                ws.row_dimensions[r].height = 22
                r += 1
                continue

            row_fill = FILL_GRAND if is_grand else (FILL_SUBTOTAL if is_subtotal else (FILL_ALT_ROW if alt else None))
            row_border = GRAND_BORDER if is_grand else (SUBTOTAL_BORDER if is_subtotal else ROW_BORDER)

            # GL prefixes (col A) — only for data rows
            if not is_subtotal and row.gl_prefixes_json:
                try:
                    prefixes = _json.loads(row.gl_prefixes_json)
                    if isinstance(prefixes, list) and prefixes:
                        gl_cell = ws.cell(row=r, column=1, value=", ".join(str(p) for p in prefixes[:3]))
                        gl_cell.font = FONT_BODY_MUTED
                        gl_cell.alignment = ALIGN_LEFT
                        if row_fill: gl_cell.fill = row_fill
                        gl_cell.border = row_border
                except Exception:
                    pass
            elif row_fill:
                a = ws.cell(row=r, column=1)
                a.fill = row_fill
                a.border = row_border

            # Label (col B)
            label = ws.cell(row=r, column=2, value=row_label + (f" {row.footnote_marker}" if (row.footnote_marker and not is_section) else ""))
            label.font = FONT_GRAND if is_grand else (FONT_SUBTOTAL if is_subtotal else FONT_BODY)
            label.alignment = ALIGN_LEFT
            if row_fill: label.fill = row_fill
            label.border = row_border

            # Numeric cells (C-J)
            def write_num(col, value, *, formula=False, is_input=False):
                cell = ws.cell(row=r, column=col, value=value)
                cell.number_format = FMT_CURRENCY
                cell.alignment = ALIGN_RIGHT
                if is_grand:
                    cell.font = FONT_GRAND
                elif is_subtotal:
                    cell.font = FONT_SUBTOTAL
                elif formula:
                    cell.font = FONT_FORMULA
                elif is_input:
                    cell.font = FONT_INPUT
                else:
                    cell.font = FONT_BODY
                if is_input and not is_grand and not is_subtotal:
                    cell.fill = FILL_INPUT
                elif row_fill:
                    cell.fill = row_fill
                cell.border = row_border

            # ── Numbers ──────────────────────────────────────────
            # Data rows SUMIFS over the DETAIL TABS by GL prefix (their Proposed
            # cells are the live, client-editable values, so a detail-tab edit
            # flows up here). Subtotals SUM the section's data rows; Net Op /
            # Total Surplus reference the recorded subtotal cells.
            prefixes = []
            if not is_subtotal and row.gl_prefixes_json:
                try:
                    parsed = _json.loads(row.gl_prefixes_json)
                    if isinstance(parsed, list):
                        prefixes = [str(p).strip() for p in parsed if p]
                except Exception:
                    pass
            # Label-only fallback for rows with a clean 1:1 GL but no stored
            # prefixes. Rows legitimately not GL-driven (Prior Year Surplus,
            # Flip Tax, Commercial Escalations) are intentionally omitted and
            # keep their stored value.
            LABEL_PREFIX_FALLBACK = {
                "cable tv": ["4250-0010"],
            }
            if not is_subtotal and not prefixes:
                _k = (row_label or "").strip().lower()
                if _k in LABEL_PREFIX_FALLBACK:
                    prefixes = LABEL_PREFIX_FALLBACK[_k]

            def detail_sumifs(detail_col):
                """SUM of SUMIFS over each (detail tab, prefix). Returns None if
                any prefix has no detail-tab home, so the caller keeps the stored
                value (correct, just not a formula)."""
                if not prefixes:
                    return None
                terms = []
                for p in prefixes:
                    hd = str(p).split("-")[0][:4]
                    sheets = [s for s in (_prefix_to_sheets.get(hd) or set())
                              if s in _detail_tab_names]
                    if not sheets:
                        # This one prefix has no detail-tab line (no data) — it
                        # contributes 0, so SKIP it rather than dropping the
                        # whole (often multi-prefix) row to blank.
                        continue
                    for s in sheets:
                        terms.append(f"SUMIFS('{s}'!${detail_col}:${detail_col},'{s}'!$A:$A,\"{p}*\")")
                return ("=" + "+".join(terms)) if terms else None

            def prefix_sumifs_yardi(yardi_col):
                """Historical columns (Prior, Approved budget) come from the raw
                yardi sheet by prefix. These aren't client-editable drivers, and
                yardi holds EVERY line, so this stays live + complete even for
                GLs that have no detail-tab home."""
                if not prefixes:
                    return None
                yd = "'yardi_data (2)'"
                return "=" + "+".join(
                    f'SUMIFS({yd}!${yardi_col}:${yardi_col},{yd}!$A:$A,"{p}*")'
                    for p in prefixes)

            if is_subtotal:
                _cat = _subtotal_cat(row_label)
                # Subtotals as VALUES: section totals = sum of the section's data
                # rows (foots in both sectioned + flat layouts, avoiding the
                # flat-format API subtotal quirk); Net Operating / Total Surplus =
                # arithmetic on the recorded subtotal values.
                _vals = {}
                if _cat == "netop":
                    inc = _subcat_vals.get("income", {}); exp = _subcat_vals.get("expenses", {})
                    for _ci in (3, 4, 5, 6, 7, 8, 9):
                        _vals[_ci] = round(inc.get(_ci, 0.0) - exp.get(_ci, 0.0), 2)
                elif _cat == "grand":
                    netop = _subcat_vals.get("netop")
                    inc = _subcat_vals.get("income", {}); exp = _subcat_vals.get("expenses", {})
                    noi = _subcat_vals.get("noi", {}); noe = _subcat_vals.get("noe", {})
                    for _ci in (3, 4, 5, 6, 7, 8, 9):
                        _b = netop.get(_ci, 0.0) if netop else (inc.get(_ci, 0.0) - exp.get(_ci, 0.0))
                        _vals[_ci] = round(_b + noi.get(_ci, 0.0) - noe.get(_ci, 0.0), 2)
                else:
                    for _ci in (3, 4, 5, 6, 7, 8, 9):
                        _vals[_ci] = round(_blockcols.get(_ci, 0.0), 2)
                for _ci in (3, 4, 5, 6, 7, 8, 9):
                    write_num(_ci, _vals[_ci])
                _row_h, _row_i = _vals.get(8), _vals.get(9)
                if _cat:
                    _subcell[_cat] = r
                    _subcat_vals[_cat] = _vals
            else:
                # Data row: write the product's own computed columns as VALUES so
                # the workbook == the app (includes the row-level pins #18/#19/#26
                # and op-assessment). Fall back to the stored row value per column.
                _plr = _pl_by_order.get(row.display_order, {})
                def _pv(_k, _fallback=None):
                    _v = _plr.get(_k)
                    if _v is None:
                        _v = _fallback
                    try:
                        return float(_v) if _v is not None else None
                    except Exception:
                        return None
                _c1 = _pv("col1", row.col1_prior_actual)
                _c2 = _pv("col2", row.col2_override)
                _c3 = _pv("col3")
                _c4 = _pv("col4")
                _c5 = _pv("col5")
                _c6 = _pv("col6", row.col6_approved_budget)
                _c7 = _pv("col7", row.col7_proposed_budget)
                write_num(3, _c1)
                write_num(4, _c2)
                write_num(5, _c3)
                write_num(6, _c4)
                write_num(7, _c5)
                write_num(8, _c6)
                write_num(9, _c7, is_input=True)
                _row_h, _row_i = _c6, _c7
                for _ci, _vv in ((3, _c1), (4, _c2), (5, _c3), (6, _c4),
                                 (7, _c5), (8, _c6), (9, _c7)):
                    if _vv is not None:
                        _blockcols[_ci] = _blockcols.get(_ci, 0.0) + _vv
                _block.append(r)
                _sec = (row.section or "").lower()
                if "expense" in _sec and "non" not in _sec:
                    if _exp_first is None:
                        _exp_first = r
                    _exp_last = r

            # J: % variance (Proposed vs Approved Budget) — value
            _jh = _row_h if (_row_h is not None) else 0.0
            _ji = _row_i if (_row_i is not None) else 0.0
            _jvar = ((_ji - _jh) / _jh) if abs(_jh) > 0.005 else 0
            vc = ws.cell(row=r, column=10, value=_jvar)
            vc.number_format = FMT_PERCENT
            vc.alignment = ALIGN_RIGHT
            vc.font = FONT_GRAND if is_grand else (FONT_SUBTOTAL if is_subtotal else FONT_FORMULA)
            if row_fill: vc.fill = row_fill
            vc.border = row_border

            if is_subtotal:
                ws.row_dimensions[r].height = 22
            r += 1
            if is_subtotal:
                _block = []
                _blockcols = {}

        # ── Column widths ────────────────────────────────────────
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 38
        for col_letter in ["C","D","E","F","G","H","I"]:
            ws.column_dimensions[col_letter].width = 14
        ws.column_dimensions["J"].width = 11

        ws.freeze_panes = "C6"

        # Defined names for the cover's live KPI cards. Workbook-scoped, so they
        # resolve at open time regardless of sheet build order. Point at the
        # grand-total + section-total Proposed cells (and Approved for variance).
        try:
            from openpyxl.workbook.defined_name import DefinedName
            def _dn(nm, cell):
                try:
                    if nm in wb.defined_names:
                        del wb.defined_names[nm]
                except Exception:
                    pass
                wb.defined_names[nm] = DefinedName(nm, attr_text=f"'{new_title}'!{cell}")
            if _subcell.get("income"):
                _dn("cbg_income",      f"$I${_subcell['income']}")
                _dn("cbg_income_appr", f"$H${_subcell['income']}")
            if _subcell.get("expenses"):
                _dn("cbg_expenses",      f"$I${_subcell['expenses']}")
                _dn("cbg_expenses_appr", f"$H${_subcell['expenses']}")
            if _subcell.get("netop"):
                _dn("cbg_netop",      f"$I${_subcell['netop']}")
                _dn("cbg_netop_appr", f"$H${_subcell['netop']}")
            if _subcell.get("grand"):
                _dn("cbg_surplus",      f"$I${_subcell['grand']}")
                _dn("cbg_surplus_appr", f"$H${_subcell['grand']}")
        except Exception:
            pass

        # ── Native charts (live: reference the summary cells, redraw on edit) ──
        # Anchored to the right of the table over the contiguous expense block:
        # a doughnut (proposed expense mix) + a prior-vs-proposed column chart.
        try:
            if _exp_first and _exp_last and _exp_last >= _exp_first + 1:
                from openpyxl.chart import DoughnutChart, BarChart, Reference
                cats = Reference(ws, min_col=2, min_row=_exp_first, max_row=_exp_last)
                dough = DoughnutChart()
                dough.title = f"{BUDGET_YEAR} proposed expense mix"
                dough.add_data(Reference(ws, min_col=9, min_row=_exp_first, max_row=_exp_last),
                               titles_from_data=False)
                dough.set_categories(cats)
                dough.height = 7.2
                dough.width = 12.5
                ws.add_chart(dough, "L2")
                bar = BarChart()
                bar.type = "col"
                bar.title = "Prior actual vs proposed"
                bar.add_data(Reference(ws, min_col=3, min_row=_exp_first, max_row=_exp_last),
                             titles_from_data=False)
                bar.add_data(Reference(ws, min_col=9, min_row=_exp_first, max_row=_exp_last),
                             titles_from_data=False)
                bar.set_categories(cats)
                try:
                    bar.series[0].graphicalProperties.solidFill = "9AA7AF"
                    bar.series[1].graphicalProperties.solidFill = "001721"
                    bar.x_axis.delete = False; bar.y_axis.delete = False
                    bar.y_axis.majorGridlines = None
                    bar.y_axis.numFmt = '"$"#,##0'
                    bar.legend.position = "b"
                except Exception:
                    pass
                bar.height = 7.2
                bar.width = 12.5
                ws.add_chart(bar, "L18")
        except Exception as _ce:
            if edit_log is not None:
                edit_log.append({"sheet": new_title, "chart_error": str(_ce)[:120]})

        # ── Variance data-bars on the Summary %Var column ────────
        try:
            from openpyxl.formatting.rule import DataBarRule
            if r > 6:
                ws.conditional_formatting.add(
                    f"J6:J{r-1}",
                    DataBarRule(start_type="min", end_type="max", color="001721"))
        except Exception:
            pass

        # ── Four-year trend (helper table + column chart) on the Summary ─────
        inc = _subcell.get("income"); exp = _subcell.get("expenses")
        net = _subcell.get("netop") or _subcell.get("grand")
        try:
            from openpyxl.chart import BarChart, Reference
            if inc and exp:
                ht = 50
                for j, yl in enumerate([BUDGET_YEAR-3, BUDGET_YEAR-2, BUDGET_YEAR-1, BUDGET_YEAR]):
                    ws.cell(row=ht, column=13+j, value=str(yl)).font = Font(name="Calibri", size=9, color="7A8791")
                # E3 values-snapshot: write each trend cell as the VALUE pulled
                # from the summary's own subtotal rows (was a live =$C$srow ref
                # that left the chart blank in non-recalc previews). Net falls
                # back to Revenue-Expenses if its source cell is a formula; if a
                # source value can't be resolved, keep the live ref.
                _COLN = {"C": 3, "D": 4, "H": 8, "I": 9}
                def _srcnum(_srow, _coln):
                    _v = ws.cell(row=_srow, column=_coln).value
                    return _v if isinstance(_v, (int, float)) else None
                for i, (lab, srow) in enumerate([("Revenue", inc), ("Expenses", exp), ("Net", net)]):
                    if not srow:
                        continue
                    ws.cell(row=ht+1+i, column=12, value=lab).font = Font(name="Calibri", size=9, color="1A1714")
                    for j, colL in enumerate(["C", "D", "H", "I"]):
                        _cn = _COLN[colL]
                        _val = _srcnum(srow, _cn)
                        if _val is None and lab == "Net":
                            _iv = _srcnum(inc, _cn) if inc else None
                            _ev = _srcnum(exp, _cn) if exp else None
                            _val = (_iv - _ev) if (_iv is not None and _ev is not None) else None
                        tc2 = (ws.cell(row=ht+1+i, column=13+j, value=round(_val, 2))
                               if _val is not None
                               else ws.cell(row=ht+1+i, column=13+j, value=f"=${colL}${srow}"))
                        tc2.number_format = FMT_CURRENCY
                tr = BarChart(); tr.type = "col"
                tr.title = f"Revenue · Expenses · Net  ({BUDGET_YEAR-3} → {BUDGET_YEAR})"
                tr.add_data(Reference(ws, min_col=12, max_col=16, min_row=ht+1, max_row=ht+3),
                            titles_from_data=True, from_rows=True)
                tr.set_categories(Reference(ws, min_col=13, max_col=16, min_row=ht))
                try:
                    tr.series[0].graphicalProperties.solidFill = "001721"
                    tr.series[1].graphicalProperties.solidFill = "9AA7AF"
                    tr.series[2].graphicalProperties.solidFill = "DE1C23"
                    tr.x_axis.delete = False; tr.y_axis.delete = False
                    tr.y_axis.majorGridlines = None
                    tr.y_axis.numFmt = '"$"#,##0'
                    tr.legend.position = "b"
                except Exception:
                    pass
                tr.height = 7.2; tr.width = 15
                ws.add_chart(tr, "L34")
        except Exception:
            pass

        # ── "At a Glance" executive tab (inserted right after the cover) ─────
        try:
            from openpyxl.chart import DoughnutChart, BarChart, Reference
            from openpyxl.chart.label import DataLabelList
            from pathlib import Path as _P2
            NAVY = "001721"; RED = "DE1C23"; MUTED = "7A8791"; INKc = "1A1714"; CARDB = "DCE3E7"
            MONEY = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
            for _n in list(wb.sheetnames):
                if _n.lower().strip() in ("at a glance", "dashboard"):
                    del wb[_n]
            dws = wb.create_sheet("At a Glance", index=1)
            dws.sheet_view.showGridLines = False
            # Card columns B/D/F/H with thin spacers C/E/G
            for col, w in {"A": 2, "B": 22, "C": 2, "D": 22, "E": 2, "F": 22, "G": 2, "H": 22, "I": 2}.items():
                dws.column_dimensions[col].width = w

            # ── Header band (rows 1-3) navy + light logo + title; red rule row 4
            for rr0 in range(1, 4):
                for c in range(1, 10):
                    dws.cell(row=rr0, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
            dws.row_dimensions[1].height = 8; dws.row_dimensions[2].height = 26; dws.row_dimensions[3].height = 14
            try:
                from openpyxl.drawing.image import Image as _XLImg
                _lp = _P2(__file__).parent / "brand" / "century_logo_light.png"
                if _lp.exists():
                    _im = _XLImg(str(_lp))
                    try:
                        from PIL import Image as _PI
                        with _PI.open(str(_lp)) as _p:
                            _iw, _ih = _p.size
                        _im.width = 150; _im.height = max(22, int(_ih * 150 / _iw))
                    except Exception:
                        _im.width = 150; _im.height = 30
                    dws.add_image(_im, "B2")
            except Exception:
                pass
            tt = dws.cell(row=2, column=8, value=building_name)
            tt.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
            tt.alignment = Alignment(horizontal="right", vertical="center")
            dws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=8)
            tt2 = dws.cell(row=3, column=8, value=f"{BUDGET_YEAR} Operating Budget · At a Glance")
            tt2.font = Font(name="Calibri", size=10, color="C4CCD1")
            tt2.alignment = Alignment(horizontal="right", vertical="top")
            dws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=8)
            for c in range(1, 10):
                dws.cell(row=4, column=c).fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
            dws.row_dimensions[4].height = 3

            # ── KPI cards (accent bar row 6, label 7, big number 8, delta 9) ──
            up_good = '[Green]"▲ "0.0%;[Red]"▼ "0.0%'   # increase favorable
            up_bad  = '[Red]"▲ "0.0%;[Green]"▼ "0.0%'   # increase unfavorable (expenses)
            def _delta(cur, appr):
                # % vs approved budget; blank when approved is ~0, and "n/m" when
                # the swing is off-the-charts (a draft proposal not yet refined
                # toward budget — avoids nonsense like ▼45057%).
                return (f'=IF(ABS({appr})<100,"",IF(ABS(({cur}-{appr})/{appr})>5,'
                        f'"n/m",({cur}-{appr})/{appr}))')
            cards = [("TOTAL REVENUE",  '=IFERROR(cbg_income,"")',   _delta("cbg_income", "cbg_income_appr"),     NAVY, up_good),
                     ("TOTAL EXPENSES", '=IFERROR(cbg_expenses,"")', _delta("cbg_expenses", "cbg_expenses_appr"), NAVY, up_bad),
                     ("NET OPERATING",  '=IFERROR(cbg_netop,"")',    _delta("cbg_netop", "cbg_netop_appr"),       RED, up_good),
                     ("TOTAL SURPLUS",  '=IFERROR(cbg_surplus,"")',  _delta("cbg_surplus", "cbg_surplus_appr"),   NAVY, up_good)]
            dws.row_dimensions[6].height = 3
            dws.row_dimensions[7].height = 15
            dws.row_dimensions[8].height = 30
            dws.row_dimensions[9].height = 16
            for idx, (lab, vf, df, accent, dfmt) in enumerate(cards):
                cc = 2 + idx * 2  # B, D, F, H
                ab = dws.cell(row=6, column=cc)
                ab.fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
                lc = dws.cell(row=7, column=cc, value=lab)
                lc.font = Font(name="Calibri", size=9, bold=True, color=MUTED)
                lc.alignment = Alignment(indent=1, vertical="center")
                vc = dws.cell(row=8, column=cc, value=vf)
                vc.font = Font(name="Calibri", size=16, bold=True, color=NAVY)
                vc.number_format = MONEY
                vc.alignment = Alignment(indent=1, vertical="center")
                dc = dws.cell(row=9, column=cc, value=df)
                dc.font = Font(name="Calibri", size=10, color=MUTED)
                dc.number_format = dfmt
                dc.alignment = Alignment(indent=1, vertical="center")
                # card border (rows 6-9)
                for rr1 in range(6, 10):
                    cell = dws.cell(row=rr1, column=cc)
                    cell.border = Border(
                        left=Side(style="thin", color=CARDB),
                        right=Side(style="thin", color=CARDB),
                        top=(Side(style="thin", color=CARDB) if rr1 == 6 else None),
                        bottom=(Side(style="thin", color=CARDB) if rr1 == 9 else None))

            def _section(rrow, text):
                s = dws.cell(row=rrow, column=2, value=text)
                s.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
                for c in range(2, 9):
                    dws.cell(row=rrow, column=c).border = Border(bottom=Side(style="medium", color=RED))

            # ── Four-year trend (full-width column chart) ────────────
            _section(11, "FOUR-YEAR TREND")
            ht = 62  # helper table, parked low/right and out of the way
            for j, yl in enumerate([BUDGET_YEAR-3, BUDGET_YEAR-2, BUDGET_YEAR-1, BUDGET_YEAR]):
                dws.cell(row=ht, column=13+j, value=str(yl))
            # E3 values-snapshot: pull VALUES from the main summary's subtotal
            # rows (was a live ='Summary'!$C$srow ref → blank chart in previews).
            _COLN2 = {"C": 3, "D": 4, "H": 8, "I": 9}
            _mainws = wb[new_title]
            def _srcnum2(_srow, _coln):
                _v = _mainws.cell(row=_srow, column=_coln).value
                return _v if isinstance(_v, (int, float)) else None
            for i, (lab, srow) in enumerate([("Revenue", inc), ("Expenses", exp), ("Net", net)]):
                if not srow:
                    continue
                dws.cell(row=ht+1+i, column=12, value=lab)
                for j, colL in enumerate(["C", "D", "H", "I"]):
                    _cn = _COLN2[colL]
                    _val = _srcnum2(srow, _cn)
                    if _val is None and lab == "Net":
                        _iv = _srcnum2(inc, _cn) if inc else None
                        _ev = _srcnum2(exp, _cn) if exp else None
                        _val = (_iv - _ev) if (_iv is not None and _ev is not None) else None
                    dc2 = (dws.cell(row=ht+1+i, column=13+j, value=round(_val, 2))
                           if _val is not None
                           else dws.cell(row=ht+1+i, column=13+j, value=f"='{new_title}'!${colL}${srow}"))
                    dc2.number_format = '"$"#,##0'
            if inc and exp:
                tb = BarChart(); tb.type = "col"; tb.title = None
                tb.add_data(Reference(dws, min_col=12, max_col=16, min_row=ht+1, max_row=ht+3),
                            titles_from_data=True, from_rows=True)
                tb.set_categories(Reference(dws, min_col=13, max_col=16, min_row=ht))
                try:
                    tb.series[0].graphicalProperties.solidFill = NAVY
                    tb.series[1].graphicalProperties.solidFill = "9AA7AF"
                    tb.series[2].graphicalProperties.solidFill = RED
                    tb.x_axis.delete = False; tb.y_axis.delete = False
                    tb.y_axis.majorGridlines = None
                    tb.y_axis.numFmt = '"$"#,##0'
                    tb.legend.position = "b"
                except Exception:
                    pass
                tb.height = 7.5; tb.width = 19
                dws.add_chart(tb, "B12")

            # ── Expense mix doughnut (references the summary expense block) ──
            _section(30, f"{BUDGET_YEAR} EXPENSE MIX")
            if _exp_first and _exp_last and _exp_last >= _exp_first + 1:
                dg = DoughnutChart(); dg.title = None
                dg.add_data(Reference(ws, min_col=9, min_row=_exp_first, max_row=_exp_last), titles_from_data=False)
                dg.set_categories(Reference(ws, min_col=2, min_row=_exp_first, max_row=_exp_last))
                try:
                    dg.holeSize = 62
                    dg.legend.position = "r"
                except Exception:
                    pass
                dg.height = 7.5; dg.width = 18
                dws.add_chart(dg, "B31")

            # ── Biggest changes vs approved (data-bar table) ─────────
            movers = []
            for _pr in _pl_by_order.values():
                if _pr.get("row_type") != "data":
                    continue
                ch = (_pr.get("col7") or 0) - (_pr.get("col6") or 0)
                if abs(ch) < 0.5:
                    continue
                _isinc = "income" in (_pr.get("section") or "").lower()
                fav = (ch > 0) if _isinc else (ch < 0)
                movers.append((_pr.get("label") or "", ch, fav))
            movers.sort(key=lambda t: abs(t[1]), reverse=True)
            _section(49, "BIGGEST CHANGES VS APPROVED BUDGET")
            mr = 50
            for lab, ch, fav in movers[:6]:
                nm = dws.cell(row=mr, column=2, value=lab)
                nm.font = Font(name="Calibri", size=11, color=INKc)
                dws.merge_cells(start_row=mr, start_column=2, end_row=mr, end_column=3)
                # magnitude column (F) drives the data bar; number itself hidden
                mag = dws.cell(row=mr, column=6, value=float(abs(ch)))
                mag.number_format = ';;;'
                ac = dws.cell(row=mr, column=8, value=float(ch))
                ac.number_format = MONEY
                ac.font = Font(name="Calibri", size=11, bold=True, color=("15803D" if fav else "991B1B"))
                ac.alignment = Alignment(horizontal="right")
                for c in range(2, 9):
                    dws.cell(row=mr, column=c).border = Border(bottom=Side(style="thin", color="EEF1F3"))
                mr += 1
            if mr > 50:
                try:
                    from openpyxl.formatting.rule import DataBarRule
                    dws.conditional_formatting.add(
                        f"F50:F{mr-1}",
                        DataBarRule(start_type="min", end_type="max", color=RED))
                except Exception:
                    pass

            ft = dws.cell(row=mr + 2, column=2,
                          value="CONFIDENTIAL — Prepared for the Board · Generated by Century Budget Manager · Figures recalculate live in Excel.")
            ft.font = Font(name="Calibri", size=9, italic=True, color=MUTED)
            dws.merge_cells(start_row=mr + 2, start_column=2, end_row=mr + 2, end_column=8)
            dws.print_options.horizontalCentered = True
            # Print only the visible board area (cols A-I); the trend helper
            # table lives out in cols L-P and must not print.
            dws.print_area = "A1:I60"
            if edit_log is not None:
                edit_log.append({"sheet": "At a Glance", "movers": len(movers[:6])})
        except Exception as _de:
            if edit_log is not None:
                edit_log.append({"sheet": "At a Glance", "error": str(_de)[:160]})

        if edit_log is not None:
            data_rows = sum(1 for x in rows if x.row_type == "data")
            edit_log.append({
                "sheet": new_title, "total_rows": len(rows),
                "data_rows": data_rows,
            })


    def _export_rewrite_comm_rent(wb, entity_code, edit_log=None):
        """Pass 1a: rewrite the Comm Rent & Escalations sheet from product
        tenant data. Drops the existing sheet contents and writes a fresh
        layout from CommercialTenant + CommercialRentPeriod. Includes
        Excel formulas for annual totals and summary feed rows.

        Structure written (per tenant block):
          - Tenant name header
          - Year / Period / $/mo / Months / Annualized columns
          - Sum row per tenant per year
        Followed by a summary section that totals across all tenants for
        each year. Numbers match what the Commercial tab shows.

        FA-edit marker: every cell we write gets a yellow fill + comment.
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.comments import Comment

        # Find any existing Comm Rent sheet to replace. openpyxl's delete_rows
        # is O(N×cells) and was hanging on heavyweight workbooks — instead
        # we delete the old sheet entirely and create a fresh one at the
        # same index in the tab order.
        old_index = None
        old_name = None
        for i, name in enumerate(wb.sheetnames):
            n = name.lower().strip()
            if "comm" in n and ("rent" in n or "escal" in n):
                old_name = name
                old_index = i
                break
        if old_name is not None:
            del wb[old_name]
        # Create new sheet at the same position
        new_title = old_name or "Comm Rent & Escalations"
        ws = wb.create_sheet(new_title, index=old_index if old_index is not None else None)
        target_name = ws.title

        tenants = CommercialTenant.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).order_by(CommercialTenant.sort_order, CommercialTenant.id).all()

        # Styling tokens
        TITLE_FONT = Font(name="Plus Jakarta Sans", size=14, bold=True, color="001721")
        SECTION_FONT = Font(name="Plus Jakarta Sans", size=11, bold=True, color="9A3412")
        TENANT_FONT = Font(name="Plus Jakarta Sans", size=12, bold=True)
        HEADER_FONT = Font(name="Plus Jakarta Sans", size=10, bold=True, color="8A7E72")
        EDIT_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        TOTAL_FILL = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        thin = Side(border_style="thin", color="E5E0D5")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        gen_stamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        edit_comment = Comment(
            f"Generated by Century Budget product on {gen_stamp}. Source: commercial_tenants table.",
            "Century Product",
        )

        def stamp(cell, fill=True):
            if fill:
                cell.fill = EDIT_FILL
            cell.border = BORDER

        # ── Title rows ────────────────────────────────────────────
        r = 1
        ws.cell(row=r, column=2, value=f"Commercial Rent & Escalations").font = TITLE_FONT
        r += 1
        ws.cell(row=r, column=2, value="Schedule A-1").font = Font(italic=True, color="8A7E72")
        ws.cell(row=r, column=4, value=f"For Year Ending 12/31/{BUDGET_YEAR}")
        r += 2

        if not tenants:
            ws.cell(row=r, column=2, value="No commercial tenants configured for this building.")
            ws.cell(row=r, column=2).font = Font(italic=True, color="8A7E72")
            if edit_log is not None:
                edit_log.append({"sheet": target_name, "tenants": 0})
            return

        # ── COMMERCIAL RENT section ──────────────────────────────
        ws.cell(row=r, column=3, value="COMMERCIAL RENT").font = SECTION_FONT
        r += 1
        # Column headers
        headers = ["", "Tenant", "Period", "$/Month", "Months", "Annualized"]
        for col_i, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=col_i, value=h)
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="right" if col_i >= 4 else "left")
        r += 1

        # Per-tenant rent table
        from sqlalchemy import asc as _asc
        for t in tenants:
            # Tenant header row
            name_cell = ws.cell(row=r, column=2, value=t.tenant_name + (f"  ({t.unit_label})" if t.unit_label else ""))
            name_cell.font = TENANT_FONT
            stamp(name_cell)
            name_cell.comment = edit_comment
            if t.lease_end:
                ws.cell(row=r, column=9, value=f"Lease ends {t.lease_end.isoformat()}").font = Font(italic=True, size=10, color="8A7E72")
            r += 1

            # Rent periods — grouped by year
            periods = CommercialRentPeriod.query.filter_by(tenant_id=t.id).order_by(
                _asc(CommercialRentPeriod.year), _asc(CommercialRentPeriod.sort_order)
            ).all()
            if not periods:
                ws.cell(row=r, column=3, value="(no rent periods)").font = Font(italic=True, color="8A7E72")
                r += 1
                continue
            current_year = None
            year_start_row = None
            year_sum = 0.0   # E3 values-snapshot: subtotal as a VALUE (no =SUM)
            for p in periods:
                if p.year != current_year:
                    # Close previous year subtotal if any
                    if year_start_row is not None and year_start_row < r:
                        sub_cell = ws.cell(row=r, column=6, value=round(year_sum, 2))
                        sub_cell.number_format = "$#,##0"
                        sub_cell.font = Font(bold=True)
                        sub_cell.fill = TOTAL_FILL
                        ws.cell(row=r, column=2, value=f"  {current_year} total").font = Font(italic=True, color="8A7E72")
                        r += 1
                    current_year = p.year
                    year_sum = 0.0
                    ws.cell(row=r, column=3, value=str(current_year)).font = Font(bold=True, color="8A7E72")
                    r += 1
                    year_start_row = r
                # Period row
                ws.cell(row=r, column=3, value=p.period_label)
                # Monthly rent (FA-editable input)
                rent_cell = ws.cell(row=r, column=4, value=float(p.monthly_rent or 0))
                rent_cell.number_format = "$#,##0.00"
                stamp(rent_cell)
                rent_cell.comment = edit_comment
                # Months count (FA-editable input)
                months_cell = ws.cell(row=r, column=5, value=int(p.months_count or 0))
                stamp(months_cell)
                # Annualized — computed VALUE (E3 values-snapshot; was =D*E, which
                # opened blank in non-recalc previews like Quick Look / Drive).
                _ann = round(float(p.monthly_rent or 0) * int(p.months_count or 0), 2)
                year_sum += _ann
                ann_cell = ws.cell(row=r, column=6, value=_ann)
                ann_cell.number_format = "$#,##0"
                r += 1
            # Close final year for this tenant
            if year_start_row is not None and year_start_row < r:
                sub_cell = ws.cell(row=r, column=6, value=round(year_sum, 2))
                sub_cell.number_format = "$#,##0"
                sub_cell.font = Font(bold=True)
                sub_cell.fill = TOTAL_FILL
                ws.cell(row=r, column=2, value=f"  {current_year} total").font = Font(italic=True, color="8A7E72")
                r += 1
            r += 1  # blank row between tenants

        # ── ESCALATIONS section (if any tenant has a real model) ─
        active_escalation_tenants = [t for t in tenants if (t.escalation_model or "none") != "none"
                                     and t.tenant_share_pct]
        if active_escalation_tenants:
            r += 1
            ws.cell(row=r, column=3, value="ESCALATIONS").font = SECTION_FONT
            r += 1
            # Headers
            esc_headers = ["", "Tenant", "Model", "Share %", "Base year", "Escalation Due"]
            for col_i, h in enumerate(esc_headers, start=1):
                c = ws.cell(row=r, column=col_i, value=h)
                c.font = HEADER_FONT
                c.alignment = Alignment(horizontal="right" if col_i >= 4 else "left")
            r += 1
            for t in active_escalation_tenants:
                ws.cell(row=r, column=2, value=t.tenant_name)
                ws.cell(row=r, column=3, value=t.escalation_model)
                share_cell = ws.cell(row=r, column=4,
                                     value=float(t.tenant_share_pct or 0))
                share_cell.number_format = "0.0000%"
                stamp(share_cell)
                base = (t.base_year_re_tax if t.escalation_model == "re_tax"
                        else t.base_year_opex if t.escalation_model == "opex"
                        else None)
                if base is not None:
                    bc = ws.cell(row=r, column=5, value=float(base))
                    bc.number_format = "$#,##0"
                    stamp(bc)
                # Escalation Due — LIVE formula: share × max(0, current basis −
                # base year). re_tax model → current RE tax (cbg_re_tax_net from
                # the RE Taxes tab); opex model → total operating expenses
                # (cbg_expenses from the Summary). Server snapshot only as a
                # fallback when neither basis name resolves.
                _emodel = (t.escalation_model or "").lower()
                # E3 values-snapshot: ship the computed escalation VALUE. Was a
                # live =MAX(0,IFERROR(cbg_re_tax_net/cbg_expenses,0)-base)*share
                # formula that opened BLANK in non-recalc previews. Same math the
                # product + Summary escalation feed (4520) use.
                _esc = _commercial_compute_escalations(entity_code)
                amt = next((e["amount"] for e in _esc if e["tenant_id"] == t.id), 0)
                ac = ws.cell(row=r, column=6, value=round(float(amt or 0), 2))
                ac.number_format = "$#,##0"
                ac.fill = TOTAL_FILL
                ac.font = Font(bold=True)
                ac.comment = Comment(
                    f"Escalation = share × max(0, current {(_emodel or 'basis')} − base year). "
                    f"Snapshot of the product value at export time.",
                    "Century Product",
                )
                r += 1
            r += 1

        # ── Lease notes appendix (collapsed at the bottom) ────────
        tenants_with_notes = [t for t in tenants if t.lease_notes]
        if tenants_with_notes:
            ws.cell(row=r, column=3, value="LEASE NOTES").font = SECTION_FONT
            r += 1
            for t in tenants_with_notes:
                ws.cell(row=r, column=2, value=t.tenant_name).font = Font(bold=True)
                r += 1
                for line in (t.lease_notes or "").split("\n"):
                    if line.strip():
                        ws.cell(row=r, column=3, value=line.strip()).font = Font(italic=True, color="8A7E72")
                        r += 1
                r += 1

        # Column widths
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 14
        for col_letter in ["G", "H", "I", "J", "K"]:
            ws.column_dimensions[col_letter].width = 14

        if edit_log is not None:
            edit_log.append({
                "sheet": target_name,
                "tenants": len(tenants),
                "with_escalation": len(active_escalation_tenants),
            })


    def _export_rewrite_cam_allocation(wb, entity_code, edit_log=None):
        """Pass 1a: write the condo CAM Allocation (Schedule A-1) sheet from
        product data (_cam_compute). ONLY for CAM-enabled buildings with classes
        — otherwise no sheet is touched (so every non-CAM building's export is
        unchanged). Ships computed VALUES (matrix cells, line totals, per-class
        column totals, grand total) — no formula strings — per the values-snapshot
        rule, so it opens correctly in any viewer (Quick Look / Drive preview)."""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter

        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget or not getattr(budget, "cam_enabled", False):
            return
        comp = _cam_compute(entity_code)
        classes = comp.get("classes") or []
        lines = comp.get("lines") or []
        if not classes or not lines:
            return

        # Find/replace an existing CAM/Schedule-A-1 sheet, else create one.
        old_index = None
        old_name = None
        for i, name in enumerate(wb.sheetnames):
            n = name.lower().strip()
            if "cam" in n or "schedule a-1" in n:
                old_name = name
                old_index = i
                break
        if old_name is not None:
            del wb[old_name]
        ws = wb.create_sheet(old_name or "CAM Allocation",
                             index=old_index if old_index is not None else None)
        target = ws.title

        TITLE_FONT = Font(name="Plus Jakarta Sans", size=14, bold=True, color="001721")
        SUB_FONT = Font(italic=True, color="8A7E72")
        HEADER_FONT = Font(name="Plus Jakarta Sans", size=10, bold=True, color="3730A3")
        SECTION_FONT = Font(name="Plus Jakarta Sans", size=10, bold=True, color="8A7E72")
        TOTAL_FONT = Font(bold=True, color="3730A3")
        TOTAL_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
        thin = Side(border_style="thin", color="E5E0D5")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        col_gl = 2
        col_first = 3
        ncls = len(classes)
        col_total = col_first + ncls

        r = 1
        ws.cell(row=r, column=col_gl, value="CAM Allocation").font = TITLE_FONT
        r += 1
        ws.cell(row=r, column=col_gl, value="Schedule A-1").font = SUB_FONT
        ws.cell(row=r, column=4, value=f"For Year Ending 12/31/{BUDGET_YEAR}").font = SUB_FONT
        r += 1
        ws.cell(row=r, column=col_gl,
                value="Operating expenses allocated across unit classes by common-interest share.").font = SUB_FONT
        r += 2

        # Class/share legend
        ws.cell(row=r, column=col_gl, value="Unit classes:").font = SECTION_FONT
        for j, cl in enumerate(classes):
            ws.cell(row=r, column=col_first + j,
                    value=f"{cl['name']} ({round((cl.get('share_pct') or 0) * 100, 4)}%)").font = HEADER_FONT
        r += 2

        # Matrix header
        ws.cell(row=r, column=col_gl, value="Expense Line").font = HEADER_FONT
        for j, cl in enumerate(classes):
            hc = ws.cell(row=r, column=col_first + j, value=cl["name"])
            hc.font = HEADER_FONT
            hc.alignment = Alignment(horizontal="right")
        tc = ws.cell(row=r, column=col_total, value="Line Total")
        tc.font = HEADER_FONT
        tc.alignment = Alignment(horizontal="right")
        r += 1

        by_sheet = {}
        for l in lines:
            by_sheet.setdefault(l["sheet_name"], []).append(l)
        for sn in ["Payroll", "Energy", "Water & Sewer", "Repairs & Supplies", "Gen & Admin"]:
            rows = by_sheet.get(sn)
            if not rows:
                continue
            ws.cell(row=r, column=col_gl, value=sn).font = SECTION_FONT
            r += 1
            for l in rows:
                ws.cell(row=r, column=col_gl,
                        value=f"{l['gl_code']}  {l.get('description') or ''}")
                cells = l.get("cells") or {}
                for j, cl in enumerate(classes):
                    v = round(float(cells.get(cl["id"], 0) or 0), 2)
                    cell = ws.cell(row=r, column=col_first + j, value=v)  # VALUE, not formula
                    cell.number_format = "$#,##0"
                    cell.border = BORDER
                tcell = ws.cell(row=r, column=col_total, value=round(float(l.get("total") or 0), 2))
                tcell.number_format = "$#,##0"
                r += 1

        # Column totals (per-class allocated expense) — VALUES
        ct = comp.get("column_totals") or {}
        lc = ws.cell(row=r, column=col_gl, value="Allocated expense")
        lc.font = TOTAL_FONT
        lc.fill = TOTAL_FILL
        for j, cl in enumerate(classes):
            v = round(float(ct.get(cl["id"], 0) or 0), 2)
            cell = ws.cell(row=r, column=col_first + j, value=v)
            cell.number_format = "$#,##0"
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
        gc = ws.cell(row=r, column=col_total, value=round(float(comp.get("grand_total") or 0), 2))
        gc.number_format = "$#,##0"
        gc.font = TOTAL_FONT
        gc.fill = TOTAL_FILL
        r += 2

        # Per-class common charges (each class funds its allocated expense)
        ws.cell(row=r, column=col_gl,
                value="Common charges by class (= allocated expense):").font = SECTION_FONT
        r += 1
        for cl in classes:
            v = round(float(ct.get(cl["id"], 0) or 0), 2)
            ws.cell(row=r, column=col_gl, value=f"  {cl['name']}")
            vc = ws.cell(row=r, column=col_first, value=v)
            vc.number_format = "$#,##0"
            r += 1
        r += 1

        # Required increase (plan: "CAM Allocation -- FA feedback on 343",
        # Cluster C): back out other income, compare to current common
        # charges. Values, not formulas -- same discipline as the rest of
        # this sheet.
        req = _cam_compute_required_increase(entity_code)
        if not req.get("error"):
            ws.cell(row=r, column=col_gl, value="Required increase:").font = SECTION_FONT
            r += 1
            oi = ws.cell(row=r, column=col_gl, value="  Less: other income")
            oc = ws.cell(row=r, column=col_first, value=-round(float(req.get("other_income") or 0), 2))
            oc.number_format = "$#,##0"
            r += 1
            cc = ws.cell(row=r, column=col_gl, value="  Amount to be covered by common charges")
            ccv = ws.cell(row=r, column=col_first, value=round(float(req.get("amount_to_be_covered") or 0), 2))
            ccv.number_format = "$#,##0"
            ccv.font = TOTAL_FONT
            r += 1
            hdr_cur = ws.cell(row=r, column=col_gl, value="  Class")
            hdr_cur.font = HEADER_FONT
            ws.cell(row=r, column=col_first, value="Current").font = HEADER_FONT
            ws.cell(row=r, column=col_first + 1, value="Required").font = HEADER_FONT
            ws.cell(row=r, column=col_first + 2, value="Increase $").font = HEADER_FONT
            ws.cell(row=r, column=col_first + 3, value="Increase %").font = HEADER_FONT
            r += 1
            for rc in req.get("classes") or []:
                ws.cell(row=r, column=col_gl, value=f"  {rc.get('class_name')}")
                c1 = ws.cell(row=r, column=col_first, value=round(float(rc.get("current_common_charges") or 0), 2))
                c1.number_format = "$#,##0"
                c2 = ws.cell(row=r, column=col_first + 1, value=round(float(rc.get("required_common_charges") or 0), 2))
                c2.number_format = "$#,##0"
                c3 = ws.cell(row=r, column=col_first + 2, value=round(float(rc.get("increase_dollar") or 0), 2))
                c3.number_format = "$#,##0"
                c4 = ws.cell(row=r, column=col_first + 3, value=round(float(rc.get("increase_pct") or 0), 2) / 100)
                c4.number_format = "0.00%"
                r += 1

        ws.column_dimensions["B"].width = 36
        for j in range(ncls + 1):
            ws.column_dimensions[get_column_letter(col_first + j)].width = 15

        if edit_log is not None:
            edit_log.append({"sheet": target, "cam_classes": ncls,
                             "cam_lines": len(lines), "grand_total": comp.get("grand_total")})

    def _export_rewrite_re_taxes(wb, entity_code, budget, edit_log=None):
        """Build a LIVE Real Estate Taxes tab from compute_re_taxes (co-ops with
        DOF data). Blue input cells (AVs, rates, exemptions, op-assessment %)
        drive the formula cells (half-taxes, gross, net), so flexing an
        assumption recomputes the net tax. Defines `cbg_re_tax_net` for the
        Commercial escalation formula. Non-co-ops / no DOF: leave the tab hidden.
        """
        from openpyxl.styles import PatternFill, Font, Alignment
        import json as _json
        ws = None; idx = None
        for i, nm in enumerate(wb.sheetnames):
            if nm.lower().strip() in ("re taxes", "re tax", "real estate taxes"):
                idx = i; ws = wb[nm]; break
        try:
            from dof_taxes import is_coop, compute_re_taxes
        except ImportError:
            from budget_app.dof_taxes import is_coop, compute_re_taxes
        try:
            _coop = is_coop(entity_code)
        except Exception:
            _coop = False
        if not _coop:
            if ws is not None:
                ws.sheet_state = "hidden"
            return
        try:
            rt = compute_re_taxes(entity_code, _re_tax_overrides_for(budget))
        except Exception as _e:
            if edit_log is not None:
                edit_log.append({"sheet": "RE Taxes", "error": str(_e)[:120]})
            if ws is not None:
                ws.sheet_state = "hidden"
            return
        if not rt or abs(float(rt.get("gross_tax") or 0)) < 1:
            if ws is not None:
                ws.sheet_state = "hidden"
            return

        name = ws.title if ws is not None else "RE Taxes"
        if ws is not None:
            del wb[name]
        ws = wb.create_sheet(name, index=idx if idx is not None else None)
        ws.sheet_state = "visible"
        ws.sheet_view.showGridLines = False
        NAVY = "001721"; RED = "DE1C23"; MUTED = "7A8791"
        INPUT_FILL = "FFF7DA"; INPUT_FONT = "065F46"
        MONEY = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
        PCT = "0.000%"
        for col, w in {"A": 2, "B": 34, "C": 16, "D": 13, "E": 16}.items():
            ws.column_dimensions[col].width = w

        def lbl(row, text, col=2, size=10, bold=False, color="1A1714"):
            c = ws.cell(row=row, column=col, value=text)
            c.font = Font(name="Calibri", size=size, bold=bold, color=color)
            return c

        def inp(row, val, col=3, fmt=MONEY):
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = fmt
            c.font = Font(name="Calibri", size=11, bold=True, color=INPUT_FONT)
            c.fill = PatternFill(start_color=INPUT_FILL, end_color=INPUT_FILL, fill_type="solid")
            c.alignment = Alignment(horizontal="right")
            return c

        def fml(row, formula, col=3, fmt=MONEY, bold=False):
            c = ws.cell(row=row, column=col, value=formula)
            c.number_format = fmt
            c.font = Font(name="Calibri", size=(12 if bold else 11), bold=bold,
                          color=(NAVY if bold else "334155"))
            c.alignment = Alignment(horizontal="right")
            return c

        ws.cell(row=1, column=2, value="Real Estate Taxes").font = Font(name="Calibri", size=16, bold=True, color=NAVY)
        ws.cell(row=2, column=2,
                value=f"{(budget.building_name if budget else entity_code) or entity_code} · Entity {entity_code} · FY {BUDGET_YEAR}"
                ).font = Font(name="Calibri", size=10, italic=True, color=MUTED)
        lbl(4, "FIRST HALF (Jul–Dec)", bold=True, size=9, color=RED)
        lbl(5, "Assessed value (transitional)"); inp(5, round(float(rt.get("assessed_value") or 0)))
        lbl(6, "Tax rate"); inp(6, float(rt.get("tax_rate") or 0), fmt=PCT)
        lbl(7, "First-half tax", bold=True, size=11, color=NAVY); fml(7, "=C5*C6/2", bold=True)
        lbl(9, "SECOND HALF (Jan–Jun)", bold=True, size=9, color=RED)
        lbl(10, "Est. assessed value"); inp(10, round(float(rt.get("est_assessed_value") or 0)))
        lbl(11, "Est. tax rate"); inp(11, float(rt.get("est_tax_rate") or 0), fmt=PCT)
        lbl(12, "Second-half tax", bold=True, size=11, color=NAVY); fml(12, "=C10*C11/2", bold=True)
        lbl(14, "Gross tax", bold=True, size=12, color=NAVY); fml(14, "=C7+C12", bold=True)
        lbl(16, "EXEMPTIONS", bold=True, size=9, color=RED)
        for _c, _t in ((3, "Current"), (4, "Growth"), (5, "Budget")):
            ws.cell(row=16, column=_c, value=_t).font = Font(name="Calibri", size=9, bold=True, color=MUTED)
        exns = rt.get("exemptions") or {}
        ex_rows = []
        er = 17
        for key, disp in (("veteran", "Veteran"), ("sche", "SCHE"), ("star", "STAR"), ("coop_abatement", "Co-op abatement")):
            ex = exns.get(key) or {}
            lbl(er, disp)
            inp(er, round(float(ex.get("current_year") or 0)), col=3)
            inp(er, float(ex.get("growth_pct") or 0), col=4, fmt="0.0%")
            fml(er, f"=C{er}*(1+D{er})", col=5)
            ex_rows.append(er); er += 1
        lbl(er, "Total exemptions", bold=True, size=11, color=NAVY)
        fml(er, f"=SUM(E{ex_rows[0]}:E{ex_rows[-1]})", col=5, bold=True)
        tot_ex_row = er; er += 2
        lbl(er, "NET REAL ESTATE TAX", bold=True, size=12, color=NAVY)
        fml(er, f"=C14-E{tot_ex_row}", bold=True)
        net_row = er; er += 2
        lbl(er, "Operating assessment %"); inp(er, float(rt.get("operating_assessment_pct") or 0.175), fmt="0.0%")
        opct_row = er; er += 1
        lbl(er, "Operating assessment"); fml(er, f"=C7*2*C{opct_row}")
        try:
            from openpyxl.workbook.defined_name import DefinedName
            if "cbg_re_tax_net" in wb.defined_names:
                del wb.defined_names["cbg_re_tax_net"]
            wb.defined_names["cbg_re_tax_net"] = DefinedName("cbg_re_tax_net", attr_text=f"'{name}'!$C${net_row}")
        except Exception:
            pass
        if edit_log is not None:
            edit_log.append({"sheet": name, "action": "re_taxes_live", "net_row": net_row})


    # ─── Budget Summary API ──────────────────────────────────────────────

    # Summary helpers (_gl_matches_prefixes / _section_key / _is_capital_line /
    # _aggregate_by_prefix) moved to summary_engine.py (tranche 2a, 2026-07-05).
    # They resolve via the module-level import — call sites unchanged.

    @bp.route("/api/budget/ensure", methods=["POST"])
    def api_budget_ensure():
        """Create a Budget record if one doesn't already exist for this entity/year.

        Used by bulk onboarding to seed buildings before importing summary rows.
        """
        data = request.get_json()
        if not data or "entity_code" not in data:
            return jsonify({"error": "entity_code required"}), 400

        entity_code = data["entity_code"]
        building_name = data.get("building_name", "Unknown")
        year = data.get("year", BUDGET_YEAR)

        existing = Budget.query.filter_by(entity_code=entity_code, year=year).first()
        if existing:
            return jsonify({"status": "exists", "entity_code": entity_code, "budget_id": existing.id})

        budget = Budget(
            entity_code=entity_code,
            building_name=building_name,
            year=year,
            status="not_started",
        )
        db.session.add(budget)
        db.session.commit()
        return jsonify({"status": "created", "entity_code": entity_code, "budget_id": budget.id})


    @bp.route("/api/summary/import/<entity_code>", methods=["POST"])
    def api_summary_import(entity_code):
        """Import budget summary row framework + Col 1 / Col 6 from parsed Excel.

        Accepts JSON matching batch_import.extract_importable_data() output.
        Upserts rows by entity_code + budget_year + display_order.
        """
        import json as _json

        data = request.get_json()
        if not data or "rows" not in data:
            return jsonify({"error": "Missing rows data"}), 400

        budget_year = BUDGET_YEAR  # Current cycle year — all imports target BUDGET_YEAR
        source_file = data.get("source_file", "")

        # Auto-create Budget record if missing (belt + suspenders for bulk onboard)
        building_name = data.get("building_name", "Unknown")
        if not Budget.query.filter_by(entity_code=entity_code, year=budget_year).first():
            db.session.add(Budget(
                entity_code=entity_code,
                building_name=building_name,
                year=budget_year,
                status="not_started",
            ))
            db.session.flush()

        imported = 0
        updated = 0

        for i, row in enumerate(data["rows"]):
            display_order = row.get("display_order") or (i + 1)

            existing = BudgetSummaryRow.query.filter_by(
                entity_code=entity_code,
                budget_year=budget_year,
                display_order=display_order,
            ).first()

            # Apply canonical Yardi prefix overrides for known-stale labels.
            # This catches push files generated from the legacy chart-of-accounts
            # (Electric/Steam/Gas/Water & Sewer/Supplies) and auto-corrects them
            # on the way in, so no future per-building redeployment is needed.
            incoming_prefixes = row.get("gl_prefixes") or []
            corrected_prefixes = apply_summary_prefix_override(row.get("label"), incoming_prefixes)
            gl_pj = None
            if corrected_prefixes:
                gl_pj = _json.dumps(corrected_prefixes)

            if existing:
                existing.label = row["label"]
                existing.section = row.get("section")
                existing.row_type = row.get("row_type", "data")
                existing.footnote_marker = row.get("footnote_marker")
                existing.col1_prior_actual = row.get("col1_prior_actual")
                existing.col6_approved_budget = row.get("col6_approved_budget")
                existing.source_tab = row.get("source_tab") or existing.source_tab
                existing.gl_prefixes_json = gl_pj or existing.gl_prefixes_json
                existing.source_file = source_file or existing.source_file
                existing.updated_at = datetime.utcnow()
                updated += 1
            else:
                db.session.add(BudgetSummaryRow(
                    entity_code=entity_code,
                    budget_year=budget_year,
                    display_order=display_order,
                    label=row["label"],
                    section=row.get("section"),
                    row_type=row.get("row_type", "data"),
                    footnote_marker=row.get("footnote_marker"),
                    col1_prior_actual=row.get("col1_prior_actual"),
                    col6_approved_budget=row.get("col6_approved_budget"),
                    col7_proposed_budget=None,
                    source_tab=row.get("source_tab"),
                    gl_prefixes_json=gl_pj,
                    source_file=source_file,
                ))
                imported += 1

        db.session.commit()

        # FA 724 #4 (Jennifer 2026-08-18): an import that lands only data
        # rows (no Total Income / Total Expenses framework) renders as one
        # flat list and nothing totals. Organize it right here so the
        # Summary is presentation-ready from the first paint.
        _organized = None
        try:
            try:
                from summary_engine import plan_section_organization, apply_section_organization
            except ImportError:
                from budget_app.summary_engine import plan_section_organization, apply_section_organization
            _rows_now = BudgetSummaryRow.query.filter_by(
                entity_code=entity_code, budget_year=budget_year
            ).order_by(BudgetSummaryRow.display_order).all()
            _plan = plan_section_organization(_rows_now)
            if _plan is not None:
                _organized = apply_section_organization(
                    db, BudgetSummaryRow, entity_code, budget_year, _rows_now, _plan)
                db.session.commit()
        except Exception as _oe:
            db.session.rollback()
            logger.warning("summary import organize skipped for %s: %s", entity_code, _oe)

        return jsonify({
            "status": "ok",
            "entity_code": entity_code,
            "budget_year": budget_year,
            "imported": imported,
            "updated": updated,
            "total_rows": len(data["rows"]),
            "organized": _organized,
        })


    @bp.route("/api/summary/<entity_code>", methods=["GET"])
    def api_summary_get(entity_code):
        """Return full 8-column budget summary for a building.

        Stored columns: col1 (2024 Actual), col6 (2026 Budget), col7 (2027 Budget).
        Computed columns: col2 (2025 Actual — TBD), col3-col5 from budget_lines
        via GL prefix aggregation, col8 = % variance.
        """
        import json as _json

        budget_year = request.args.get("year", BUDGET_YEAR, type=int)

        summary_rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=budget_year
        ).order_by(BudgetSummaryRow.display_order).all()

        # Fallback: if no rows for requested year, use latest available year
        if not summary_rows:
            latest = db.session.query(db.func.max(BudgetSummaryRow.budget_year)).filter_by(
                entity_code=entity_code
            ).scalar()
            if latest:
                budget_year = latest
                summary_rows = BudgetSummaryRow.query.filter_by(
                    entity_code=entity_code, budget_year=budget_year
                ).order_by(BudgetSummaryRow.display_order).all()

        if not summary_rows:
            return jsonify({"error": "No summary data found", "entity_code": entity_code}), 404

        # Fetch budget_lines for GL aggregation (cols 3-5)
        budget = Budget.query.filter_by(entity_code=entity_code, year=budget_year).first()
        bl_dicts = []
        ytd_months = 2

        if budget:
            lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
            bl_dicts = [l.to_dict() for l in lines]
            try:
                assumptions = _json.loads(budget.assumptions_json) if budget.assumptions_json else {}
                bp_val = assumptions.get("budget_period", "")
                if "/" in str(bp_val):
                    ytd_months = int(str(bp_val).split("/")[0])
            except Exception:
                pass

        # FA 724 (Jennifer 2026-08-20): insurance col7 = forecast x
        # (1 + renewal increase) — thread the assumption into the engine.
        _ins_renewal_pct = 0.0
        try:
            _ins = (_json.loads(budget.assumptions_json)
                    if budget and budget.assumptions_json else {}).get("insurance_renewal") or {}
            _ins_renewal_pct = float(_ins.get("increase_percent") or 0)
        except Exception:
            _ins_renewal_pct = 0.0

        # ── FA dir 2026-06-03 (#6): operating-assessment proposed driver ──
        # The operating-assessment (GL 4200) row's proposed budget (Col 7) =
        # first-half RE tax × 2 × pct (default 17.5%, editable per-property on
        # the RE Tax page → re_taxes_overrides). Co-ops only; condos have no
        # building-level RE tax. Computed ONCE here (DOF data is cached) and
        # applied to the 4200 row below when no explicit Summary override.
        _op_assess_proposed = None
        _re_tax_exemptions_budget = None  # FA #18: RE-tax 2026-27 exemptions total (negated → income Tax Benefit Credits col7)
        try:
            from dof_taxes import is_coop as _is_coop, compute_re_taxes as _compute_re_taxes
            if _is_coop(entity_code):
                _rt = _compute_re_taxes(entity_code, _re_tax_overrides_for(budget))
                _val = _rt.get("operating_assessment_proposed")
                if _val is not None and abs(float(_val)) > 0.005:
                    _op_assess_proposed = round(float(_val), 2)
                # FA #18 (2026-06-16): capture the RE-tax 2026-27 exemptions
                # total so the income "Tax Benefit Credits" row can pin its
                # proposed (col7) to the NEGATIVE of it (see col7 cascade).
                _exb = _rt.get("total_exemptions_budget")
                if _exb is not None and abs(float(_exb)) > 0.005:
                    _re_tax_exemptions_budget = round(float(_exb), 2)
        except Exception as _e:
            logger.warning(f"operating-assessment proposed compute failed for {entity_code}: {_e}")
            _op_assess_proposed = None
            _re_tax_exemptions_budget = None

        # ── Audit row for Col 2 (prefetched here; engine is DB-free) ──────
        fy = str(budget_year - 2)  # Col 2 = BY-2 actual
        row_au = None
        col2_sql_error = None
        try:
            # Fiscal-year match is preferred but NOT required: uploads whose
            # filename had no parseable year carry fiscal_year_end='' and used
            # to vanish from Col 2 silently (826, 2026-06-10 — "none of the
            # mapping got into the budget"). A confirmed audit with an unknown
            # year is still this entity's audit; exact-year rows win the sort.
            row_au = db.session.execute(db.text(
                "SELECT id, mapped_data, fiscal_year_end, confirmed_at, confirmed_by, pdf_filename, raw_extraction, summary_overrides FROM audit_uploads "
                "WHERE entity_code = :ec AND status = 'confirmed' "
                "AND (fiscal_year_end = :fy OR fiscal_year_end IS NULL OR fiscal_year_end = '') "
                # NULLS LAST (2026-06-10 audit): Postgres DESC defaults to
                # NULLS FIRST, so a NULL-fiscal-year confirmed upload outranked
                # the exact-year row — Col 2 from the wrong audit. Force the
                # exact-year match to win.
                "ORDER BY (fiscal_year_end = :fy) DESC NULLS LAST, confirmed_at DESC LIMIT 1"
            ), {"ec": entity_code, "fy": fy}).fetchone()
        except Exception as _sql_err:
            col2_sql_error = str(_sql_err)

        # Pure computation lives in summary_engine.compute_summary
        # (tranche 2a) — deterministic over these inputs, test-vectored.
        return jsonify(compute_summary(
            entity_code, budget_year, summary_rows, bl_dicts, ytd_months,
            row_au=row_au, col2_sql_error=col2_sql_error,
            op_assess_proposed=_op_assess_proposed,
            re_tax_exemptions_budget=_re_tax_exemptions_budget,
            insurance_renewal_pct=_ins_renewal_pct,
        ))


    @bp.route("/api/admin/soft-reset/<entity_code>", methods=["POST"])
    def api_soft_reset_entity(entity_code):
        """ADMIN: Soft-reset an entity for a fresh wizard run, KEEPING source uploads.

        Clears the computed budget (lines, summary rows, commercial tenants,
        revisions) and resets wizard state to step 0 / status draft. Does NOT
        touch source uploads (yardi YSL, expense_reports, open_ap_reports,
        maint_proof_reports), audited financials (AuditUpload rows),
        BuildingInfo, payroll_positions, or building assignments. Use this
        when you want the FA to re-run the wizard from scratch but skip
        re-uploading source files.

        Returns counts of what was deleted so the caller can verify.
        Idempotent: safe to call on an already-reset entity (counts return 0).
        """
        import json as _json
        ec = str(entity_code).strip()
        budget = Budget.query.filter_by(entity_code=ec, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": f"No budget found for {ec} / FY{BUDGET_YEAR}"}), 404

        counts = {
            "budget_lines": 0,
            "summary_rows": 0,
            "commercial_tenants": 0,
            "commercial_periods": 0,
            "commercial_billbacks": 0,
            "budget_revisions": 0,
            "presentation_edits": 0,
        }
        warnings = []

        def _try_delete(sql, params=None, label=""):
            """Run one DELETE in its own txn. Returns rowcount or 0; never
            poisons the session. The whole point of the soft-reset is to be
            tolerant of schema drift across environments."""
            try:
                r = db.session.execute(db.text(sql), params or {})
                db.session.commit()
                return r.rowcount or 0
            except Exception as _e:
                db.session.rollback()
                warnings.append(f"{label}: {str(_e)[:160]}")
                logger.warning(f"soft-reset skip {label}: {_e}")
                return 0

        # 1) Line-scoped dependents
        line_rows = db.session.execute(
            db.text("SELECT id FROM budget_lines WHERE budget_id = :bid"),
            {"bid": budget.id}
        ).fetchall()
        line_ids = [r[0] for r in line_rows]
        db.session.commit()  # close the SELECT txn cleanly

        if line_ids:
            ids_str = ",".join(str(i) for i in line_ids)
            counts["presentation_edits"] = _try_delete(
                f"DELETE FROM presentation_edits WHERE budget_line_id IN ({ids_str})",
                label="presentation_edits"
            )
            counts["budget_revisions"] += _try_delete(
                f"DELETE FROM budget_revisions WHERE budget_line_id IN ({ids_str})",
                label="budget_revisions (line-scoped)"
            )

        # 2) Budget-scoped revisions
        counts["budget_revisions"] += _try_delete(
            "DELETE FROM budget_revisions WHERE budget_id = :bid",
            {"bid": budget.id},
            label="budget_revisions (budget-scoped)"
        )

        # 3) Budget lines
        counts["budget_lines"] = _try_delete(
            "DELETE FROM budget_lines WHERE budget_id = :bid",
            {"bid": budget.id},
            label="budget_lines"
        )

        # 4) Summary rows
        counts["summary_rows"] = _try_delete(
            "DELETE FROM budget_summary_rows WHERE entity_code = :ec AND budget_year = :y",
            {"ec": ec, "y": BUDGET_YEAR},
            label="budget_summary_rows"
        )

        # 5) Commercial rent data — three sibling tables, any could differ across
        # environments. Each gets its own committed delete.
        try:
            tenant_rows = db.session.execute(db.text(
                "SELECT id FROM commercial_tenants WHERE entity_code = :ec AND budget_year = :y"
            ), {"ec": ec, "y": BUDGET_YEAR}).fetchall()
            db.session.commit()
            t_ids = [r[0] for r in tenant_rows]
        except Exception as _e:
            db.session.rollback()
            warnings.append(f"commercial_tenants SELECT: {str(_e)[:160]}")
            t_ids = []

        if t_ids:
            ids_str = ",".join(str(i) for i in t_ids)
            counts["commercial_periods"] = _try_delete(
                f"DELETE FROM commercial_rent_periods WHERE tenant_id IN ({ids_str})",
                label="commercial_rent_periods"
            )
            counts["commercial_billbacks"] = _try_delete(
                f"DELETE FROM commercial_tenant_billbacks WHERE tenant_id IN ({ids_str})",
                label="commercial_tenant_billbacks"
            )
            counts["commercial_tenants"] = _try_delete(
                f"DELETE FROM commercial_tenants WHERE id IN ({ids_str})",
                label="commercial_tenants"
            )

        # 6) Reset wizard / budget header — its own commit
        try:
            budget = Budget.query.get(budget.id)  # re-fetch in case session was cleared
            budget.wizard_step = 0
            budget.wizard_completed_at = None
            budget.status = "draft"
            budget.assumptions_json = None
            budget.lock_state = None
            budget.locked_at = None
            budget.locked_by = None
            budget.updated_at = datetime.utcnow()
            db.session.commit()
        except Exception as _e:
            db.session.rollback()
            warnings.append(f"budget header reset: {str(_e)[:160]}")

        # What we deliberately KEPT — for FA reassurance
        kept = {
            "budget_id": budget.id,
            "audit_uploads": db.session.execute(db.text(
                "SELECT COUNT(*) FROM audit_uploads WHERE entity_code = :ec"
            ), {"ec": ec}).scalar() or 0,
            "expense_reports": db.session.execute(db.text(
                "SELECT COUNT(*) FROM expense_reports WHERE entity_code = :ec"
            ), {"ec": ec}).scalar() or 0,
            "open_ap_reports": db.session.execute(db.text(
                "SELECT COUNT(*) FROM open_ap_reports WHERE entity_code = :ec"
            ), {"ec": ec}).scalar() or 0,
        }
        try:
            _log_wizard_event(ec, step="admin", action="soft_reset", payload={
                "deleted": counts, "kept": kept,
            })
        except Exception:
            pass
        logger.info(f"Soft-reset {ec}: deleted={counts} kept={kept} warnings={warnings}")
        return jsonify({
            "status": "ok",
            "entity_code": ec,
            "deleted": counts,
            "kept": kept,
            "warnings": warnings,
        })


    @bp.route("/api/admin/resolve-summary-aliases/<entity_code>", methods=["POST"])
    def api_resolve_summary_aliases(entity_code):
        """ADMIN: Re-resolve label aliases for an existing entity's summary rows.

        Use case: when we add a new entry to LABEL_ALIASES (e.g., Storage ->
        Storage Income), existing BudgetSummaryRow records still have their
        original empty gl_prefixes_json because aliases resolve only at
        import time. This endpoint walks the entity's rows, looks up each
        label via LABEL_ALIASES + SUMMARY_ROW_MAP, and updates gl_prefixes_json
        when it finds a match the previous import missed.

        Body (optional): {"all_entities": true} to run for every entity.

        Idempotent. Safe to run repeatedly. Only updates rows where the new
        prefix list differs from what's already stored.
        """
        import json as _json
        # GL_TO_SUMMARY_MAP lives in budget_summary; both flat and packaged
        # forms appear elsewhere — try them in order.
        SUMMARY_ROW_MAP = LABEL_ALIASES = _CONDO_ROWS = None
        try:
            from GL_TO_SUMMARY_MAP import SUMMARY_ROW_MAP, LABEL_ALIASES, _CONDO_ROWS
        except ImportError:
            try:
                from budget_summary.GL_TO_SUMMARY_MAP import SUMMARY_ROW_MAP, LABEL_ALIASES, _CONDO_ROWS
            except ImportError as e:
                return jsonify({"error": f"GL_TO_SUMMARY_MAP import failed: {e}"}), 500

        body = request.get_json(silent=True) or {}
        all_entities = bool(body.get("all_entities"))

        if all_entities:
            entities = [b.entity_code for b in Budget.query.filter_by(year=BUDGET_YEAR).all()]
        else:
            entities = [entity_code]

        # FA dir 2026-05-19 (148 RE Tax redesign): the new "Real Estate Tax
        # Benefit Credits" row didn't exist on most entities — credits used to
        # roll up under the "Real Estate Taxes" row (prefix "6315"). Now that
        # "Real Estate Taxes" is narrowed to just "6315-0000", the credit GLs
        # (6315-0010..0040) need a home or they'll orphan. Auto-insert the
        # row right after "Real Estate Taxes" with the same section.
        _AUTO_INSERT_ROWS = {
            "Real Estate Tax Benefit Credits": {
                "after_label": "Real Estate Taxes",
                "section": "expenses",
            },
        }

        results = []
        total_updated = 0
        total_inserted = 0
        for ec in entities:
            rows = BudgetSummaryRow.query.filter_by(entity_code=ec, budget_year=BUDGET_YEAR).all()
            existing_labels = {r.label for r in rows}
            updated = 0
            inserted = 0
            updated_labels = []
            inserted_labels = []
            for row in rows:
                if row.row_type != "data":
                    continue
                label = row.label
                # Resolution chain: direct -> alias -> condo-specific
                cfg = SUMMARY_ROW_MAP.get(label)
                if not cfg:
                    canonical = LABEL_ALIASES.get(label)
                    if canonical and canonical != label:
                        cfg = SUMMARY_ROW_MAP.get(canonical)
                if not cfg:
                    cfg = _CONDO_ROWS.get(label)
                if not cfg:
                    continue
                new_prefixes = cfg.get("gl_prefix", [])
                if not new_prefixes:
                    continue
                try:
                    current = _json.loads(row.gl_prefixes_json) if row.gl_prefixes_json else []
                except Exception:
                    current = []
                if current != new_prefixes:
                    row.gl_prefixes_json = _json.dumps(new_prefixes)
                    updated += 1
                    updated_labels.append({"label": label, "old": current, "new": new_prefixes})

            # Auto-insert missing summary rows defined above.
            auto_insert_error = None
            for new_label, meta in _AUTO_INSERT_ROWS.items():
                if new_label in existing_labels:
                    continue
                after_label = meta["after_label"]
                anchor = next((r for r in rows if r.label == after_label), None)
                if not anchor:
                    # No anchor — skip. Don't make up an arbitrary position.
                    continue
                new_cfg = SUMMARY_ROW_MAP.get(new_label)
                if not new_cfg:
                    continue
                # Shift every row after the anchor down by +1. The uq constraint
                # on (entity_code, budget_year, display_order) is NOT deferrable,
                # and Postgres checks UNIQUE per-row inside an UPDATE statement,
                # so the naive `+1` can hit transient collisions. Two-step with
                # a large offset avoids any conflict:
                #   1) shift to display_order + 100000 (no positive collisions)
                #   2) shift back by 99999 → effectively +1, room for new row
                try:
                    db.session.flush()  # send any pending ORM updates first
                    db.session.execute(
                        db.text(
                            "UPDATE budget_summary_rows SET display_order = display_order + 100000 "
                            "WHERE entity_code = :ec AND budget_year = :year AND display_order > :anchor"
                        ),
                        {"ec": ec, "year": BUDGET_YEAR, "anchor": anchor.display_order},
                    )
                    db.session.execute(
                        db.text(
                            "UPDATE budget_summary_rows SET display_order = display_order - 99999 "
                            "WHERE entity_code = :ec AND budget_year = :year AND display_order > 100000"
                        ),
                        {"ec": ec, "year": BUDGET_YEAR},
                    )
                    # Keep in-memory `rows` consistent with DB so any subsequent
                    # _AUTO_INSERT_ROWS iteration in this loop sees the new state.
                    for r in rows:
                        if r.display_order > anchor.display_order:
                            r.display_order = r.display_order + 1
                    new_row = BudgetSummaryRow(
                        entity_code=ec,
                        budget_year=BUDGET_YEAR,
                        display_order=anchor.display_order + 1,
                        label=new_label,
                        section=meta.get("section") or anchor.section,
                        row_type="data",
                        gl_prefixes_json=_json.dumps(new_cfg.get("gl_prefix", [])),
                    )
                    db.session.add(new_row)
                    db.session.flush()
                    rows.append(new_row)
                    existing_labels.add(new_label)
                    inserted += 1
                    inserted_labels.append({"label": new_label, "after": after_label, "order": anchor.display_order + 1})
                except Exception as e:
                    db.session.rollback()
                    auto_insert_error = f"{new_label}: {type(e).__name__}: {str(e)[:200]}"
                    # Stop trying to insert further rows on this entity
                    break

            if updated or inserted:
                try:
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    results.append({"entity_code": ec, "error": str(e)[:200]})
                    continue
            total_updated += updated
            total_inserted += inserted
            row_result = {
                "entity_code": ec,
                "rows_examined": len(rows),
                "rows_updated": updated,
                "rows_inserted": inserted,
                "updates": updated_labels,
                "inserts": inserted_labels,
            }
            if auto_insert_error:
                row_result["auto_insert_error"] = auto_insert_error
            results.append(row_result)
        return jsonify({
            "ok": True,
            "total_entities": len(entities),
            "total_rows_updated": total_updated,
            "total_rows_inserted": total_inserted,
            "results": results,
        })


    @bp.route("/api/admin/summary-debug/<entity_code>", methods=["GET"])
    def api_summary_debug(entity_code):
        try:
            return _api_summary_debug_impl(entity_code)
        except Exception as e:
            import traceback
            return jsonify({
                "error": str(e),
                "traceback": traceback.format_exc().split("\n")[-15:],
            }), 500

    def _api_summary_debug_impl(entity_code):
        """ADMIN: Diagnostic view of an entity's Budget Summary aggregation.

        Lists every BudgetSummaryRow + its matched GL prefixes + GL line
        match counts, so we can see at a glance:
          - Which rows have no GL prefix configured (label-alias miss)
          - Which rows have GL prefixes but zero matching budget_lines
          - Which budget_lines aren't picked up by ANY summary row
            (orphan GL codes — the FA's "missing data" complaints)

        This is the diagnostic tool for FA #9 / #12 / #16 (and any future
        "X isn't pulling into the summary"). Run it for any entity, look
        for empty/orphan rows, then add the appropriate LABEL_ALIASES
        entry or extend the SUMMARY_ROW_MAP gl_prefix list.

        Returns:
          {
            "entity_code": "...",
            "summary_rows": [...],
            "orphan_gl_codes": [...],
            "stats": {...}
          }
        """
        import json as _json
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404

        summary_rows = BudgetSummaryRow.query.filter_by(entity_code=entity_code, budget_year=BUDGET_YEAR).all()
        budget_lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
        bl_dicts = [l.to_dict() for l in budget_lines]

        # Track which GL codes are matched by SOME summary row's prefix
        matched_gls = set()
        out_rows = []

        for row in summary_rows:
            label = row.label
            try:
                prefixes = _json.loads(row.gl_prefixes_json) if row.gl_prefixes_json else []
            except Exception:
                prefixes = []

            matches = []
            ytd_total = 0.0
            for line in bl_dicts:
                gl = line.get("gl_code", "")
                if not _gl_matches_prefixes(gl, prefixes):
                    continue
                matches.append({
                    "gl_code": gl,
                    "description": line.get("description") or "",
                    "ytd_actual": round(float(line.get("ytd_actual", 0) or 0), 2),
                    "category": line.get("category"),
                })
                matched_gls.add(gl)
                ytd_total += float(line.get("ytd_actual", 0) or 0)

            out_rows.append({
                "label": label,
                "row_type": row.row_type,
                "section": row.section,
                "display_order": row.display_order,
                "gl_prefixes": prefixes,
                "match_count": len(matches),
                "ytd_total": round(ytd_total, 2),
                "matched_lines": matches[:10],  # cap for readability; full count in match_count
                "has_data": ytd_total != 0,
                "is_orphan": row.row_type == "data" and not prefixes and not (row.special if hasattr(row, "special") else None),
            })

        # Orphan GL codes (have data but no summary row claims them)
        orphans = []
        for line in bl_dicts:
            gl = line.get("gl_code", "")
            if gl and gl not in matched_gls:
                ytd = float(line.get("ytd_actual", 0) or 0)
                cb = float(line.get("current_budget", 0) or 0)
                if abs(ytd) > 0.01 or abs(cb) > 0.01:  # only show non-zero
                    orphans.append({
                        "gl_code": gl,
                        "description": line.get("description") or "",
                        "sheet_name": line.get("sheet_name"),
                        "category": line.get("category"),
                        "ytd_actual": round(ytd, 2),
                        "current_budget": round(cb, 2),
                    })

        stats = {
            "total_rows": len(out_rows),
            "rows_with_no_prefix": sum(1 for r in out_rows if r["row_type"] == "data" and not r["gl_prefixes"]),
            "rows_with_zero_data": sum(1 for r in out_rows if r["row_type"] == "data" and r["match_count"] == 0),
            "orphan_gl_count": len(orphans),
            "orphan_ytd_total": round(sum(o["ytd_actual"] for o in orphans), 2),
        }

        return jsonify({
            "entity_code": entity_code,
            "summary_rows": out_rows,
            "orphan_gl_codes": orphans,
            "stats": stats,
        })


    @bp.route("/api/summary/<entity_code>/export.xlsx", methods=["GET"])
    def api_summary_export_xlsx(entity_code):
        """Download the summary as a DYNAMIC .xlsx — computed cells are live Excel
        formulas (Forecast =D+E, subtotals =SUM(...), Net Op / Total Surplus cross-row,
        %Var), so the workbook recalculates in Excel. Reuses api_summary_get so the file
        is always byte-identical in meaning to the on-screen summary (single source)."""
        res = api_summary_get(entity_code)
        if isinstance(res, tuple):   # (error_response, status)
            return res
        data = res.get_json() if hasattr(res, "get_json") else None
        if not data or not data.get("rows"):
            return jsonify({"error": "No summary data for %s" % entity_code}), 404
        try:
            try:
                from excel_export import build_summary_workbook
            except ImportError:
                from budget_app.excel_export import build_summary_workbook
            xb = build_summary_workbook(data)
        except Exception as e:
            logger.error("xlsx export failed for %s: %s", entity_code, e)
            return jsonify({"error": "Export failed: %s" % str(e)[:200]}), 500
        from flask import Response
        fname = "Budget_Summary_%s_%s.xlsx" % (entity_code, data.get("budget_year") or "")
        return Response(
            xb,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="%s"' % fname},
        )

    @bp.route("/api/summary/<entity_code>", methods=["PUT"])
    def api_summary_edit(entity_code):
        """FA edits cells on summary rows.

        Accepts JSON: {"edits": [
            {"display_order": N, "col7": value},          // legacy proposed-budget edit
            {"display_order": N, "col3_override": value}, // FA-set computed-cell override
            {"display_order": N, "col4_override": value},
            {"display_order": N, "col5_override": value},
            ...
        ]}

        Each edit object can carry one or more of the editable fields. Pass
        null to clear an override (revert to computed). Logs each change
        to budget_revisions for audit trail.
        """
        data = request.get_json()
        if not data or "edits" not in data:
            return jsonify({"error": "Missing edits"}), 400

        budget_year = data.get("budget_year", BUDGET_YEAR)
        user_id = data.get("user_id")

        # Need a budget record for revision logging
        budget = Budget.query.filter_by(entity_code=entity_code, year=budget_year).first()

        # Editable fields: legacy col7 + all column overrides (FA dir 2026-05-17).
        # Adding col1/col2/col6 overrides made every numeric Summary cell editable.
        EDITABLE_FIELDS = {
            "col7": "col7_proposed_budget",
            "col1_override": "col1_override",
            "col2_override": "col2_override",
            "col3_override": "col3_override",
            "col4_override": "col4_override",
            "col5_override": "col5_override",
            "col6_override": "col6_override",
        }
        # FA dir 2026-05-17: typed-formula strings persist in cell_formulas_json.
        # Edit shape: {col1_formula: "=300*12*4"} or {col1_formula: null} to clear.
        # Stored as a JSON blob keyed by col so we can add/clear per-col without
        # touching siblings.
        FORMULA_FIELDS = ("col1_formula", "col2_formula", "col3_formula",
                          "col4_formula", "col5_formula", "col6_formula", "col7_formula")

        updated = 0
        import uuid as _uuid
        _batch_id = _uuid.uuid4().hex  # QA fix 2: batch undo
        for edit in data["edits"]:
            display_order = edit.get("display_order")
            if display_order is None:
                continue

            row = BudgetSummaryRow.query.filter_by(
                entity_code=entity_code,
                budget_year=budget_year,
                display_order=display_order,
            ).first()
            if not row:
                continue

            row_changed = False
            for api_field, db_attr in EDITABLE_FIELDS.items():
                if api_field not in edit:
                    continue
                new_val = edit.get(api_field)
                old_val = getattr(row, db_attr, None)
                # Coerce: None passes through (clears override / proposed);
                # non-None coerces to float.
                coerced = float(new_val) if new_val is not None else None
                # Skip no-op writes (avoids spurious revision rows)
                if (old_val is None and coerced is None) or (old_val == coerced):
                    continue
                setattr(row, db_attr, coerced)
                row_changed = True
                # Log to budget_revisions if budget exists
                if budget:
                    db.session.add(BudgetRevision(
                        budget_id=budget.id,
                        user_id=user_id,
                        action="summary_edit",
                        field_name=f"{api_field}:{row.label}",
                        old_value=str(old_val) if old_val is not None else "",
                        new_value=str(coerced) if coerced is not None else "",
                        source="web",
                        batch_id=_batch_id,
                    ))
            # FA dir 2026-05-17: persist typed-formula strings in cell_formulas_json.
            # Edit shape:
            #   {"col1_formula": "=300*12*4"} → set / replace col1 formula
            #   {"col1_formula": null}        → clear col1 formula (drop key)
            # Stored as a single JSON keyed by col so revoke / add is surgical.
            import json as _json
            try:
                row_formulas = _json.loads(row.cell_formulas_json or "{}")
                if not isinstance(row_formulas, dict):
                    row_formulas = {}
            except Exception:
                row_formulas = {}
            for f_field in FORMULA_FIELDS:
                if f_field not in edit:
                    continue
                col_key = f_field.split("_")[0]  # "col1_formula" → "col1"
                new_formula = edit.get(f_field)
                old_formula = row_formulas.get(col_key)
                if new_formula is None or new_formula == "":
                    # Clear the formula for this col.
                    if col_key in row_formulas:
                        row_formulas.pop(col_key, None)
                        row_changed = True
                else:
                    # Store the formula. Trim + coerce to string.
                    new_formula_s = str(new_formula).strip()
                    if old_formula != new_formula_s:
                        row_formulas[col_key] = new_formula_s
                        row_changed = True
                        if budget:
                            db.session.add(BudgetRevision(
                                budget_id=budget.id,
                                user_id=user_id,
                                action="summary_formula",
                                field_name=f"{f_field}:{row.label}",
                                old_value=str(old_formula or ""),
                                new_value=new_formula_s,
                                source="web",
                            ))
            # Persist the formulas blob (or NULL when empty so we don't carry
            # a stale "{}" string forever).
            row.cell_formulas_json = _json.dumps(row_formulas) if row_formulas else None

            if row_changed:
                row.updated_at = datetime.utcnow()
                updated += 1

        db.session.commit()
        return jsonify({"status": "ok", "updated": updated})


    # ─── Commercial Rent Routes (FA directive 2026-05-14 Phase 5) ──────────
    # Read API + one-time importer for the new Commercial tab. The importer
    # parses the Excel "Comm Rent & Escalations" sheet on first request and
    # populates CommercialTenant + CommercialRentPeriod tables. Subsequent
    # requests just read from the DB. Phase 2 will add edit endpoints.

    def _comm_rent_parse_excel(workbook):
        """Parse the Comm Rent & Escalations sheet into structured tenant data.
        Best-effort — handles the 148 / 212 / 829 patterns observed in research.
        Returns: [{tenant_name, unit_label, rent_periods: [...], lease_notes,
                   escalation_model, ...}]
        """
        sheet = None
        for name in workbook.sheetnames:
            n = name.lower().strip()
            if "comm" in n and ("rent" in n or "escal" in n):
                sheet = workbook[name]
                break
        if not sheet:
            return []

        months_short = ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]
        month_idx = {m: i + 1 for i, m in enumerate(months_short)}

        def infer_months(period_label):
            s = (period_label or "").lower().replace("—", "-").replace("–", "-").replace(" ", "")
            if not s:
                return 0
            if "-" not in s:
                for m in months_short:
                    if s.startswith(m):
                        return 1
                return 0
            parts = s.split("-")
            if len(parts) != 2:
                return 0
            def find_idx(p):
                for m in months_short:
                    if p.startswith(m):
                        return month_idx[m]
                return 0
            a, b = find_idx(parts[0]), find_idx(parts[1])
            if a == 0 or b == 0:
                return 0
            return (b - a + 1) if b >= a else 0

        HEADER_TOKENS = {
            "tenant", "tenant ", "category", "annual expense", "base year",
            "difference", "tenants share", "balance due from tenant",
            "schedule a-1", "commercial rent", "commercial(garage rent)",
            "real estate taxes", "operating expenses", "escalatable portion",
            "less:  base year", "less: base year", "tenant's proportionate share",
            "real estate tax escalations", "operating escalations",
            "tenant escalations", "less abatements and credits", "base amount",
            "escalated portion", "commercial rent escalations",
        }

        # Strategy: do a two-pass parse so we know all tenant rows BEFORE
        # detecting unit codes (fixes the "1AB before Cobblestone" issue on
        # building 829). Also extract years from inside strings like
        # "2025 Total" / "2026 Budget" (fixes 212 which has no pure year rows).

        import re as _re
        YEAR_RE = _re.compile(r"\b(20\d{2})\b")  # any 4-digit year 2000-2099

        # ── Pass 1: collect rows ─────────────────────────────────────
        rows = []
        for r in range(1, min((sheet.max_row or 0) + 1, 120)):
            def cell(c, _row=r):
                try:
                    return sheet.cell(row=_row, column=c).value
                except Exception:
                    return None
            cb, cc, cd, ce = cell(2), cell(3), cell(4), cell(5)
            ci, ck = cell(9), cell(11)
            b_str = str(cb).strip() if cb is not None else ""
            c_str = str(cc).strip() if cc is not None else ""
            rows.append({
                "r": r, "cb": cb, "cc": cc, "cd": cd, "ce": ce, "ci": ci, "ck": ck,
                "b_str": b_str, "c_str": c_str,
                "b_low": b_str.lower(), "c_low": c_str.lower(),
            })

        def is_real_tenant_name(rd):
            """True if column B looks like a tenant name (not a unit code,
            not a header). A unit code is short + has no whitespace +
            mostly digits (e.g. '1AB', '1C', '2A')."""
            s = rd["b_str"]
            if not s or s.startswith("="):
                return False
            if rd["b_low"] in HEADER_TOKENS:
                return False
            # Unit code: short, no spaces, starts with digit
            if len(s) <= 5 and not any(ch.isspace() for ch in s) and s[0].isdigit():
                return False
            return True

        # ── Pass 2: walk rows in order with state ────────────────────
        tenants = []
        current = None
        current_year = None
        section = "rent"  # rent | re_tax | utility_billback | opex | summary
        pending_unit_label = None  # buffered unit code waiting for a tenant
        # Periods seen before any year was detected for the current tenant.
        # When the first year is discovered, these are assigned to that year.
        # Fixes building 212 where rent row precedes year-total row.
        pending_yearless_periods = []

        def assign_pending(tenant, year):
            for p in pending_yearless_periods:
                p["year"] = year
                tenant["rent_periods"].append(p)
            pending_yearless_periods.clear()

        for idx, rd in enumerate(rows):
            b_str, c_str = rd["b_str"], rd["c_str"]
            b_low, c_low = rd["b_low"], rd["c_low"]
            cb, cd, ci, ck = rd["cb"], rd["cd"], rd["ci"], rd["ck"]

            # Section transitions
            if "real estate tax" in c_low and "escal" in c_low:
                section = "re_tax"
                continue
            if "operating escal" in c_low or "operating expense" in c_low:
                section = "opex"
                continue
            if "tenant escal" in c_low:
                section = "utility_billback"
                continue

            if section != "rent":
                # Stop attaching to tenants once we exit the rent section.
                continue

            # Year detection (handles "2025", "2025 Total", "2026 Budget").
            # Important: do this BEFORE tenant creation so a tenant row that
            # also carries period data sees the year set earlier.
            ymatch = YEAR_RE.search(c_str)
            if ymatch:
                try:
                    new_year = int(ymatch.group(1))
                    # If we have pending yearless periods buffered, assign them
                    # to this year now. Fixes 212 (rent row precedes year row).
                    if current is not None and pending_yearless_periods:
                        assign_pending(current, new_year)
                    current_year = new_year
                except Exception:
                    pass

            # Tenant / unit code detection
            if b_str and not b_str.startswith("=") and b_low not in HEADER_TOKENS:
                is_unit_code = (
                    len(b_str) <= 5
                    and not any(ch.isspace() for ch in b_str)
                    and b_str[0].isdigit()
                )
                if is_unit_code:
                    # Look ahead: does the next row have a real tenant name?
                    # If yes, this unit code is a label for that tenant.
                    next_rd = rows[idx + 1] if idx + 1 < len(rows) else None
                    if next_rd and is_real_tenant_name(next_rd):
                        pending_unit_label = b_str
                    elif current is not None and not current.get("unit_label"):
                        current["unit_label"] = b_str
                else:
                    # Real tenant name. Create new tenant.
                    # CRITICAL: don't reset current_year — same-row period data
                    # depends on the year set earlier (Mack r10 case).
                    new_tenant = {
                        "tenant_name": b_str,
                        "unit_label": pending_unit_label,  # consume any pending
                        "rent_periods": [],
                        "notes_lines": [],
                        "tenant_share_pct": None,
                        "base_year_re_tax": None,
                        "base_year_opex": None,
                        "escalation_model": "none",
                    }
                    tenants.append(new_tenant)
                    current = new_tenant
                    pending_unit_label = None

            # Rent period detection — three sub-cases:
            #
            # (a) Year-only header row (e.g., c="2025"). Skip — period rows
            #     come on subsequent rows.
            # (b) Year + label + rent value on same row (e.g., c="2026 Budget"
            #     d=$16,666). Create a 12-month period for that year.
            # (c) Normal period row (e.g., c="Jan-Feb" d=$7,593). Period
            #     label is c_str; months_count from label parsing.
            looks_like_pure_year = bool(ymatch and YEAR_RE.fullmatch(c_str))
            has_rent_value = isinstance(cd, (int, float)) and cd and cd > 0

            if current is not None and has_rent_value and c_str and not looks_like_pure_year:
                if ymatch:
                    # Case (b): year+label+rent in same row → 12-month period.
                    # Use the year from the match (which may not be current_year
                    # if multiple years appear — prefer the one in this row's c_str).
                    try:
                        row_year = int(ymatch.group(1))
                    except Exception:
                        row_year = current_year
                    label_clean = YEAR_RE.sub("", c_str).strip(" -:") or "Annual"
                    current["rent_periods"].append({
                        "year": row_year,
                        "period_label": label_clean[:50],
                        "monthly_rent": float(cd),
                        "months_count": 12,
                    })
                else:
                    # Case (c): normal period row.
                    months = infer_months(c_str)
                    if months > 0:
                        period = {
                            "year": current_year if current_year is not None else 0,
                            "period_label": c_str,
                            "monthly_rent": float(cd),
                            "months_count": months,
                        }
                        if current_year is None:
                            # Buffer — assign when first year seen.
                            pending_yearless_periods.append(period)
                        else:
                            current["rent_periods"].append(period)

            # Capture lease notes from columns I / K (commentary)
            if current is not None:
                for note_cell in (ci, ck):
                    if note_cell and isinstance(note_cell, str) and len(note_cell) > 5:
                        if note_cell not in current["notes_lines"]:
                            current["notes_lines"].append(note_cell)

        # Determine escalation model from sheet content (single-mode for now)
        sheet_text_lower = ""
        for r in range(1, min((sheet.max_row or 0) + 1, 100)):
            for c in range(1, 16):
                v = None
                try:
                    v = sheet.cell(row=r, column=c).value
                except Exception:
                    pass
                if isinstance(v, str):
                    sheet_text_lower += v.lower() + " "
        if "real estate tax escal" in sheet_text_lower:
            esc_model = "re_tax"
        elif "operating escal" in sheet_text_lower:
            esc_model = "opex"
        elif "tenant escal" in sheet_text_lower and ("gas" in sheet_text_lower or "steam" in sheet_text_lower):
            esc_model = "utility_billback"
        else:
            esc_model = "none"

        # Trim periods to a sensible year window. The Excel may include deep
        # history (829 has 2020-2026). For budget work, only the prior +
        # current + near-future commitments matter. Also helps mask the
        # cross-tenant-block pollution that affects buildings like 829
        # where tenant names appear mid-block (years before the name belong
        # to a previous tenant that's already finished its block).
        year_lo = BUDGET_YEAR - 2  # 2025 if BUDGET_YEAR=2027
        year_hi = BUDGET_YEAR + 3
        for t in tenants:
            t["rent_periods"] = [
                p for p in t.get("rent_periods", [])
                if year_lo <= p.get("year", 0) <= year_hi
            ]

        # Flatten notes into a single text field, fill escalation model.
        for t in tenants:
            t["escalation_model"] = esc_model
            t["lease_notes"] = "\n".join(t.get("notes_lines", []))[:2000] or None
            t.pop("notes_lines", None)

        # Drop tenants with no usable rent periods AFTER trimming. Better to
        # show a clean empty state than misleading entries.
        tenants = [t for t in tenants if t.get("rent_periods")]

        return tenants


    def _comm_rent_run_import(entity_code):
        """One-time importer: parse the building's approved Excel and create
        CommercialTenant + CommercialRentPeriod rows. Idempotent: returns
        existing data if any tenants already exist for this building/year.
        """
        existing = CommercialTenant.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).count()
        if existing > 0:
            return {"status": "exists", "imported": 0, "tenant_count": existing}

        # Use the app.py SharePoint helpers via the registered models — we can't
        # import them directly from here without a circular dep, so go through
        # current_app's view of the existing scan endpoint plumbing.
        import sys, tempfile
        from pathlib import Path
        try:
            import openpyxl
        except Exception as e:
            return {"status": "error", "error": f"openpyxl not available: {e!r}"}

        # The SharePoint helpers live in app.py — pull them via current_app.
        from flask import current_app
        sp_list = current_app.view_functions.get("admin_research_comm_rent")
        # We need _sharepoint_list_approved_budgets + _sharepoint_download_item.
        # They're not bound to the route directly; they exist in app.py globals.
        try:
            import app as _app_mod  # type: ignore
            files = _app_mod._sharepoint_list_approved_budgets(entity_code)
        except Exception as e:
            return {"status": "error", "error": f"sharepoint list failed: {str(e)[:200]}"}

        if not files:
            return {"status": "no_file", "error": "no approved 2026 budget file"}

        files.sort(key=lambda f: f.get("last_modified", ""), reverse=True)
        target = files[0]
        item_id = target.get("item_id")
        if not item_id:
            return {"status": "error", "error": "no item_id on file"}

        try:
            _name, file_bytes = _app_mod._sharepoint_download_item(item_id)
        except Exception as e:
            return {"status": "error", "error": f"download failed: {str(e)[:200]}"}

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = tmp.name
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            parsed = _comm_rent_parse_excel(wb)
        except Exception as e:
            return {"status": "error", "error": f"parse failed: {str(e)[:200]}"}
        finally:
            if tmp_path:
                try:
                    import os as _os
                    _os.unlink(tmp_path)
                except Exception:
                    pass

        # Persist
        imported = 0
        for i, t in enumerate(parsed):
            tenant = CommercialTenant(
                entity_code=entity_code,
                budget_year=BUDGET_YEAR,
                tenant_name=t["tenant_name"][:200],
                unit_label=(t.get("unit_label") or None) and t["unit_label"][:100],
                lease_notes=t.get("lease_notes"),
                escalation_model=t.get("escalation_model") or "none",
                sort_order=i,
                imported_from_excel=True,
            )
            db.session.add(tenant)
            db.session.flush()  # need tenant.id for periods
            for j, p in enumerate(t.get("rent_periods") or []):
                db.session.add(CommercialRentPeriod(
                    tenant_id=tenant.id,
                    year=p["year"],
                    period_label=p["period_label"][:50],
                    monthly_rent=p["monthly_rent"],
                    months_count=p["months_count"],
                    sort_order=j,
                ))
            imported += 1
        db.session.commit()
        return {
            "status": "imported",
            "imported": imported,
            "file_name": target.get("name"),
        }


    @bp.route("/api/commercial/<entity_code>", methods=["GET"])
    def api_commercial_get(entity_code):
        """List commercial tenants for a building with their rent periods +
        billbacks. Auto-imports from Excel on first call (idempotent).
        Query params:
          - skip_import=1: don't auto-import (just return whatever's in DB)
        """
        skip_import = request.args.get("skip_import", "0") == "1"
        if not skip_import:
            import_result = _comm_rent_run_import(entity_code)
        else:
            import_result = {"status": "skipped"}

        tenants = CommercialTenant.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).order_by(CommercialTenant.sort_order, CommercialTenant.id).all()

        # Ensure the Summary row is in sync — cheap idempotent op, makes the
        # auto-feed visible to FAs reading building dashboards that haven't
        # had any commercial edits since this code shipped.
        try:
            sync = _commercial_recompute_summary(entity_code)
        except Exception as _se:
            sync = {"error": str(_se)[:200]}

        return jsonify({
            "entity_code": entity_code,
            "budget_year": BUDGET_YEAR,
            "tenant_count": len(tenants),
            "tenants": [t.to_dict() for t in tenants],
            "import_result": import_result,
            "summary_sync": sync,
        })


    # ─── Summary auto-feed helper (Phase 3a) ─────────────────────────
    # After any commercial rent change, recompute the BudgetSummaryRow
    # for Commercial Rent (GL 4040) and write the budget-year total into
    # col7_proposed_budget. Called from every CRUD endpoint below.

    def _commercial_find_summary_row(entity_code, label_candidates=None, gl_prefix_candidates=None):
        """Find the BudgetSummaryRow that represents Commercial Rent for a
        building. Match by label first (alias map flattens to "Commercial"
        canonical, but some imports leave "Commercial Rent" intact), then
        by GL prefix in the row's stored prefix list. Returns the row or None.
        """
        labels = {(l or "").strip().lower() for l in (label_candidates or ["Commercial Rent", "Commercial"])}
        prefixes_want = gl_prefix_candidates or ["4040"]
        rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).all()
        # Pass 1: label match
        for r in rows:
            if (r.label or "").strip().lower() in labels:
                return r
        # Pass 2: GL prefix match
        import json as _j
        for r in rows:
            try:
                stored = _j.loads(r.gl_prefixes_json or "[]")
            except Exception:
                continue
            if not isinstance(stored, list):
                continue
            for p in stored:
                base = str(p).split("-")[0].strip()
                for cand in prefixes_want:
                    if base == cand:
                        return r
        return None

    def _commercial_compute_escalations(entity_code, year=None):
        """Compute every commercial tenant's escalation amount for `year`.
        Returns: [{tenant_id, tenant_name, model, amount, breakdown}].
        breakdown includes the inputs so the UI can show the math.

        FA dir 2026-06-03 (#1): parameterized by year. The escalation math is
        identical across years (share × max(0, currentYearTax − fixed base));
        only the "current-year" RE-tax / opex BASIS changes:
          - year == BUDGET_YEAR (2027, proposed): RE Taxes col7 (proposed),
            expenses col7. This is the original behavior.
          - year  < BUDGET_YEAR (2026, BY-1): RE Taxes col6 (2026 Budget),
            expenses col6. Per FA: the 2026 escalation is driven off the 2026
            Budget. The tenant's base_year_* and share stay fixed (standard
            lease escalation — only the current-year figure moves).
        """
        year = year or BUDGET_YEAR
        is_budget_year = (year >= BUDGET_YEAR)
        tenants = CommercialTenant.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).all()
        if not tenants:
            return []

        # Pull "Real Estate Taxes" row once (used by re_tax model). Basis col
        # depends on the year: col7 (proposed) for BUDGET_YEAR, col6 (2026
        # Budget) for the prior year.
        re_tax_row = _commercial_find_summary_row(
            entity_code,
            label_candidates=["Real Estate Taxes", "Real Estate Tax"],
            gl_prefix_candidates=["6310", "6311", "6315", "6320"],
        )
        if re_tax_row:
            re_tax_current = ((re_tax_row.col7_proposed_budget
                               or re_tax_row.col6_approved_budget or 0)
                              if is_budget_year
                              else (re_tax_row.col6_approved_budget or 0))
        else:
            re_tax_current = 0

        # Sum all expense-section rows (used by opex model) — same basis col.
        expense_rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR
        ).all()
        import json as _cj
        total_expense = 0
        for r in expense_rows:
            if r.row_type != "data":
                continue
            _sect = (r.section or "").lower().strip()
            _is_expense = (_sect == "expenses")
            # FA #20 (2026-06-16): flat-format summaries (e.g. 829) leave
            # `section` NULL, so the section-only filter found nothing and the
            # opex basis came out $0 — the escalation silently never pulled to
            # the Summary. Fall back to GL-prefix classification (5xxx / 6xxx =
            # operating expense) when there's no section bucket. Sectioned
            # buildings are unchanged.
            if not _sect:
                try:
                    _prefs = _cj.loads(r.gl_prefixes_json or "[]")
                except Exception:
                    _prefs = []
                _ebases = {str(p).split("-")[0].strip() for p in _prefs if p}
                _is_expense = bool(_ebases) and all(
                    b[:1] in ("5", "6") for b in _ebases)
            if _is_expense:
                if is_budget_year:
                    total_expense += (r.col7_proposed_budget
                                      or r.col6_approved_budget or 0)
                else:
                    total_expense += (r.col6_approved_budget or 0)

        results = []
        for t in tenants:
            model = t.escalation_model or "none"
            share = t.tenant_share_pct or 0
            amount = 0.0
            breakdown = {"model": model, "share_pct": share}
            if model == "re_tax" and share and t.base_year_re_tax:
                escalatable = max(0, re_tax_current - t.base_year_re_tax)
                amount = escalatable * share
                breakdown.update({
                    "current_re_tax": re_tax_current,
                    "base_year": t.base_year_re_tax,
                    "escalatable": escalatable,
                })
            elif model == "opex" and share and t.base_year_opex:
                escalatable = max(0, total_expense - t.base_year_opex)
                amount = escalatable * share
                breakdown.update({
                    "current_opex": total_expense,
                    "base_year": t.base_year_opex,
                    "escalatable": escalatable,
                })
            # utility_billback handled in Phase 3b.3 (per-category)
            results.append({
                "tenant_id": t.id,
                "tenant_name": t.tenant_name,
                "model": model,
                "amount": round(amount, 2),
                "breakdown": breakdown,
            })
        return results


    def _commercial_recompute_summary(entity_code):
        """Recompute the Commercial Rent (4040) AND Commercial Escalations
        (4520) summary rows from current commercial tenant data.

        Sums:
          - 4040 col7 = total of BUDGET_YEAR rent periods across all tenants
          - 4520 col7 = total of per-tenant escalation amounts (re_tax + opex)

        No-op (returns sync stub) if the row isn't on Summary. Idempotent.
        """
        # ── Row 4040: Commercial Rent ─────────────────────────────────
        rent_row = _commercial_find_summary_row(entity_code)
        rent_sync = None
        if rent_row:
            total_rent = (db.session.query(
                db.func.coalesce(
                    db.func.sum(CommercialRentPeriod.monthly_rent
                                * CommercialRentPeriod.months_count),
                    0,
                )
            )
            .join(CommercialTenant,
                  CommercialTenant.id == CommercialRentPeriod.tenant_id)
            .filter(
                CommercialTenant.entity_code == entity_code,
                CommercialTenant.budget_year == BUDGET_YEAR,
                CommercialRentPeriod.year == BUDGET_YEAR,
            ).scalar()) or 0
            old_rent = rent_row.col7_proposed_budget
            rent_row.col7_proposed_budget = float(total_rent) if total_rent else None
            rent_row.updated_at = datetime.utcnow()
            rent_sync = {
                "row_id": rent_row.id, "label": rent_row.label,
                "old_col7": old_rent, "new_col7": rent_row.col7_proposed_budget,
                "total_periods_summed": total_rent,
            }

        # ── Row 4520: Commercial Escalations ─────────────────────────
        # Match on label OR GL prefix. Different buildings use different
        # labels ("Commercial Rent Escalations", "Commercial RE Tax
        # Escalation", "Commercial RE Tax", etc.) — the GL prefix 4520
        # is the consistent identifier.
        esc_row = _commercial_find_summary_row(
            entity_code,
            label_candidates=[
                "Commercial Rent Escalations",
                "Commercial RE Tax Escalation",
                "Commercial R/E Tax Escalation",
                "Commercial R/E Tax Escalation  (A)",
                "Comercial RE Tax",
                "Commercial Operating Tax Escalation",
                "Commercial Tenant Escalation",
                "Commercial Escalations",
            ],
            gl_prefix_candidates=["4520"],
        )
        esc_sync = None
        if esc_row:
            # 2027 (BUDGET_YEAR, proposed) → col7, the original behavior.
            escalations = _commercial_compute_escalations(entity_code)
            esc_total = sum(e["amount"] for e in escalations)
            old_esc = esc_row.col7_proposed_budget
            esc_row.col7_proposed_budget = float(esc_total) if esc_total else None
            # FA dir 2026-06-03 (#1): 2026 (BY-1) → col5 (2026 Forecast) via
            # col5_override, so it surfaces on the Summary's 2026 Forecast
            # column mirroring how the 2027 total flows to col7. Same
            # methodology (re_tax + opex), driven off the 2026 Budget basis.
            # Leaves the imported 2026 Budget (col6) untouched for reference.
            escalations_2026 = _commercial_compute_escalations(entity_code, year=BUDGET_YEAR - 1)
            esc_total_2026 = sum(e["amount"] for e in escalations_2026)
            old_esc_2026 = esc_row.col5_override
            esc_row.col5_override = float(esc_total_2026) if esc_total_2026 else None
            esc_row.updated_at = datetime.utcnow()
            esc_sync = {
                "row_id": esc_row.id, "label": esc_row.label,
                "old_col7": old_esc, "new_col7": esc_row.col7_proposed_budget,
                "total_escalations": esc_total,
                "per_tenant": escalations,
                # 2026 mirror (#1)
                "old_col5": old_esc_2026, "new_col5": esc_row.col5_override,
                "total_escalations_2026": esc_total_2026,
                "per_tenant_2026": escalations_2026,
            }

        db.session.commit()
        return {"rent": rent_sync, "escalations": esc_sync}


    # ─── Tenant CRUD (Phase 2a) ──────────────────────────────────────
    # Add / edit / delete commercial tenants. Endpoints follow the existing
    # patterns from other CRUD endpoints in this file.

    @bp.route("/api/commercial/<entity_code>/tenant", methods=["POST"])
    def api_commercial_tenant_create(entity_code):
        data = request.get_json(silent=True) or {}
        name = (data.get("tenant_name") or "").strip()
        if not name:
            return jsonify({"error": "tenant_name required"}), 400
        max_sort = db.session.query(db.func.max(CommercialTenant.sort_order)).filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR).scalar() or 0
        t = CommercialTenant(
            entity_code=entity_code,
            budget_year=BUDGET_YEAR,
            tenant_name=name[:200],
            unit_label=(data.get("unit_label") or None) and str(data["unit_label"])[:100],
            lease_notes=data.get("lease_notes"),
            escalation_model=data.get("escalation_model") or "none",
            tenant_share_pct=data.get("tenant_share_pct"),
            base_year_re_tax=data.get("base_year_re_tax"),
            base_year_opex=data.get("base_year_opex"),
            sort_order=max_sort + 1,
            imported_from_excel=False,
        )
        db.session.add(t)
        db.session.commit()
        return jsonify(t.to_dict())

    @bp.route("/api/commercial/<entity_code>/tenant/<int:tenant_id>", methods=["PUT"])
    def api_commercial_tenant_update(entity_code, tenant_id):
        t = CommercialTenant.query.filter_by(id=tenant_id, entity_code=entity_code).first()
        if not t:
            return jsonify({"error": "tenant not found"}), 404
        data = request.get_json(silent=True) or {}
        # Whitelisted fields the FA can edit. Skip None values so callers can
        # send partial updates without clearing fields.
        UPDATABLE = {
            "tenant_name", "unit_label", "lease_notes",
            "escalation_model", "tenant_share_pct",
            "base_year_re_tax", "base_year_opex",
            "lease_start", "lease_end",
        }
        for k, v in data.items():
            if k not in UPDATABLE:
                continue
            if k in ("lease_start", "lease_end") and v:
                try:
                    from datetime import date as _date
                    setattr(t, k, _date.fromisoformat(str(v)[:10]))
                except Exception:
                    pass
            else:
                setattr(t, k, v)
        t.updated_at = datetime.utcnow()
        db.session.commit()
        # Recompute Summary in case escalation config or share % changed.
        sync = _commercial_recompute_summary(entity_code)
        result = t.to_dict()
        result["summary_sync"] = sync
        return jsonify(result)

    @bp.route("/api/commercial/<entity_code>/tenant/<int:tenant_id>", methods=["DELETE"])
    def api_commercial_tenant_delete(entity_code, tenant_id):
        t = CommercialTenant.query.filter_by(id=tenant_id, entity_code=entity_code).first()
        if not t:
            return jsonify({"error": "tenant not found"}), 404
        # Explicit cascade — don't rely on FK ondelete (SQLite may not honor it)
        CommercialRentPeriod.query.filter_by(tenant_id=t.id).delete()
        CommercialTenantBillback.query.filter_by(tenant_id=t.id).delete()
        db.session.delete(t)
        db.session.commit()
        sync = _commercial_recompute_summary(entity_code)
        return jsonify({"status": "deleted", "id": tenant_id, "summary_sync": sync})

    @bp.route("/api/commercial/<entity_code>/tenant/<int:tenant_id>/period", methods=["POST"])
    def api_commercial_period_create(entity_code, tenant_id):
        t = CommercialTenant.query.filter_by(id=tenant_id, entity_code=entity_code).first()
        if not t:
            return jsonify({"error": "tenant not found"}), 404
        data = request.get_json(silent=True) or {}
        year = int(data.get("year") or 0)
        if not year:
            return jsonify({"error": "year required"}), 400
        max_sort = db.session.query(db.func.max(CommercialRentPeriod.sort_order)).filter_by(
            tenant_id=t.id).scalar() or 0
        p = CommercialRentPeriod(
            tenant_id=t.id,
            year=year,
            period_label=(data.get("period_label") or "Jan-Dec")[:50],
            monthly_rent=float(data.get("monthly_rent") or 0),
            months_count=int(data.get("months_count") or 12),
            sort_order=max_sort + 1,
        )
        db.session.add(p)
        db.session.commit()
        sync = _commercial_recompute_summary(entity_code)
        return jsonify({"period": p.to_dict(), "summary_sync": sync})

    @bp.route("/api/commercial/<entity_code>/tenant/<int:tenant_id>/period/<int:period_id>",
             methods=["PUT"])
    def api_commercial_period_update(entity_code, tenant_id, period_id):
        p = CommercialRentPeriod.query.filter_by(id=period_id, tenant_id=tenant_id).first()
        if not p:
            return jsonify({"error": "period not found"}), 404
        # Verify tenant belongs to entity
        t = CommercialTenant.query.filter_by(id=tenant_id, entity_code=entity_code).first()
        if not t:
            return jsonify({"error": "tenant not found"}), 404
        data = request.get_json(silent=True) or {}
        if "year" in data and data["year"]:
            p.year = int(data["year"])
        if "period_label" in data and data["period_label"] is not None:
            p.period_label = str(data["period_label"])[:50]
        if "monthly_rent" in data:
            p.monthly_rent = float(data["monthly_rent"] or 0)
        if "months_count" in data:
            p.months_count = int(data["months_count"] or 12)
        db.session.commit()
        sync = _commercial_recompute_summary(entity_code)
        # to_dict() already includes annualized for client-side update.
        result = p.to_dict()
        result["summary_sync"] = sync
        return jsonify(result)

    @bp.route("/api/commercial/<entity_code>/tenant/<int:tenant_id>/period/<int:period_id>",
             methods=["DELETE"])
    def api_commercial_period_delete(entity_code, tenant_id, period_id):
        p = CommercialRentPeriod.query.filter_by(id=period_id, tenant_id=tenant_id).first()
        if not p:
            return jsonify({"error": "period not found"}), 404
        db.session.delete(p)
        db.session.commit()
        sync = _commercial_recompute_summary(entity_code)
        return jsonify({"status": "deleted", "id": period_id, "summary_sync": sync})

    @bp.route("/api/commercial/<entity_code>/project-year", methods=["POST"])
    def api_commercial_project_year(entity_code):
        """Project rent periods from a source year to a target year by cloning
        each tenant's source-year periods, optionally multiplied by an
        increase percentage. If target-year periods already exist for a
        tenant, they are skipped (use ?overwrite=1 to replace).

        Body: {
            "from_year": 2026,
            "to_year": 2027,
            "increase_pct": 0.03,        # decimal (3% = 0.03), optional, default 0
            "tenant_ids": [1,2,3],        # optional; default all tenants for building
            "overwrite": false            # optional; default false
        }
        """
        data = request.get_json(silent=True) or {}
        from_year = int(data.get("from_year") or (BUDGET_YEAR - 1))
        to_year = int(data.get("to_year") or BUDGET_YEAR)
        increase_pct = float(data.get("increase_pct") or 0)
        overwrite = bool(data.get("overwrite", False))
        tenant_ids = data.get("tenant_ids")

        q = CommercialTenant.query.filter_by(entity_code=entity_code, budget_year=BUDGET_YEAR)
        if tenant_ids:
            q = q.filter(CommercialTenant.id.in_(tenant_ids))
        tenants = q.all()

        created_total = 0
        skipped_tenants = []
        for t in tenants:
            existing_to = CommercialRentPeriod.query.filter_by(
                tenant_id=t.id, year=to_year).all()
            if existing_to and not overwrite:
                skipped_tenants.append(t.id)
                continue
            if existing_to and overwrite:
                for p in existing_to:
                    db.session.delete(p)
                db.session.flush()
            source_periods = CommercialRentPeriod.query.filter_by(
                tenant_id=t.id, year=from_year).order_by(
                CommercialRentPeriod.sort_order).all()
            if not source_periods:
                continue
            for sp in source_periods:
                new_rent = (sp.monthly_rent or 0) * (1 + increase_pct)
                db.session.add(CommercialRentPeriod(
                    tenant_id=t.id,
                    year=to_year,
                    period_label=sp.period_label,
                    monthly_rent=round(new_rent, 2),
                    months_count=sp.months_count,
                    sort_order=sp.sort_order,
                ))
                created_total += 1
        db.session.commit()
        sync = _commercial_recompute_summary(entity_code)
        return jsonify({
            "status": "ok",
            "from_year": from_year,
            "to_year": to_year,
            "increase_pct": increase_pct,
            "periods_created": created_total,
            "skipped_tenants": skipped_tenants,
            "summary_sync": sync,
        })


    @bp.route("/api/commercial/<entity_code>/import", methods=["POST"])
    def api_commercial_import(entity_code):
        """Force a re-import from Excel. WARNING: skips if tenants already
        exist — use ?force=1 to wipe + reimport (admin action)."""
        force = request.args.get("force", "0") == "1"
        if force:
            tenants = CommercialTenant.query.filter_by(
                entity_code=entity_code, budget_year=BUDGET_YEAR
            ).all()
            for t in tenants:
                # Cascading delete handles periods + billbacks via ondelete=CASCADE
                # but Flask-SQLAlchemy may not pass through to SQLite-style FKs
                # in all setups. Explicitly delete children first to be safe.
                CommercialRentPeriod.query.filter_by(tenant_id=t.id).delete()
                CommercialTenantBillback.query.filter_by(tenant_id=t.id).delete()
                db.session.delete(t)
            db.session.commit()

        result = _comm_rent_run_import(entity_code)
        return jsonify(result)


    # ─── CAM Allocation (condos) — Schedule A-1 ─────────────────────────────
    # Split operating-expense GLs across unit classes by proportionate share to
    # drive per-class common charges. Mirrors the commercial-escalation cluster.
    # Design + coverage: CAM_ALLOCATION_DESIGN_2026-06-17.md, CAM_COVERAGE_2026-06-17.md.
    CAM_EXPENSE_SHEETS = ("Payroll", "Energy", "Water & Sewer",
                          "Repairs & Supplies", "Gen & Admin")

    def _cam_line_amount(l):
        """The budgeted expense a CAM line allocates: proposed budget, falling
        back to the approved (current) budget when proposed isn't set."""
        try:
            p = float(l.proposed_budget) if l.proposed_budget is not None else 0.0
        except (TypeError, ValueError):
            p = 0.0
        if abs(p) > 0.005:
            return p
        try:
            return float(l.current_budget or 0)
        except (TypeError, ValueError):
            return 0.0

    def _cam_parse_excel(workbook):
        """Parse a condo's CAM Allocation / Schedule A-1 sheet into unit
        classes + their proportionate common-interest shares.

        Grounded in the confirmed 347 layout (a HEADER ROW of class names —
        "Total Expenses | Residential | Retail | Garage | Total" — with the
        matching % shares in the row directly below — "2026 Budget | 76.5953%
        | 16.6464% | 6.7583% | 100.0000%" — repeated as the column header
        down the whole GL matrix). Falls back to a VERTICAL layout (class name
        in one column, its share in the next, one row per class) for
        buildings that lay it out that way (see coverage review,
        CAM_COVERAGE_2026-06-17.md — cond-ops and a few non-standard condos).

        Never guesses: only returns a block whose shares reconcile to ~100%
        (0.90-1.10 tolerance for OCR/rounding slop, then re-normalized to
        exactly 1.0). Returns [{"name", "share_pct"}] or None if nothing
        reconciling was found.
        """
        EXCLUDE = {
            "total", "total expenses", "total operating expenses", "totals",
            "g/l code", "gl code", "g/l", "code", "description", "amount",
            "for year ending", "cam allocation", "schedule a-1", "a-1",
            "proportionate share", "% common interest", "common interest",
            "unit type", "category", "r/b", "r/b flag", "flag", "notes", "note",
            "2026 budget", "2027 budget", "budget", "current budget", "approved",
            "prior year", "increase", "variance", "%", "share", "class",
        }

        def find_sheet():
            for name in workbook.sheetnames:
                n = name.lower().strip()
                if "cam" in n or "schedule a-1" in n or n == "a-1":
                    ws = workbook[name]
                    if getattr(ws, "max_row", None) is None:
                        continue  # chart-only sheet, no cells
                    return ws
            for name in workbook.sheetnames:
                ws = workbook[name]
                if getattr(ws, "max_row", None) is None:
                    continue  # chart-only sheet, no cells
                for r in range(1, min((ws.max_row or 0) + 1, 8)):
                    for c in range(1, min((ws.max_column or 0) + 1, 12)):
                        v = ws.cell(row=r, column=c).value
                        if isinstance(v, str) and (
                                "cam allocation" in v.lower() or "schedule a-1" in v.lower()):
                            return ws
            return None

        sheet = find_sheet()
        if not sheet:
            return None

        def norm_share(v):
            if v is None:
                return None
            if isinstance(v, (int, float)):
                f = float(v)
            elif isinstance(v, str):
                s = v.strip().rstrip("%").replace(",", "")
                if not s:
                    return None
                try:
                    f = float(s)
                except ValueError:
                    return None
                if "%" in v:
                    f = f / 100.0
            else:
                return None
            if f > 1.5:      # a raw 76.59 is a percent number, not a fraction
                f = f / 100.0
            return f

        def is_class_label(v):
            if not isinstance(v, str):
                return False
            s = v.strip()
            if not s or len(s) > 40:
                return False
            if s.lower() in EXCLUDE:
                return False
            if any(ch.isdigit() for ch in s) and len(s) <= 3:
                return False   # bare GL-code-looking fragments
            return True

        max_r = min((sheet.max_row or 0), 80)
        max_c = min((sheet.max_column or 0), 30)

        def cell(r, c):
            try:
                return sheet.cell(row=r, column=c).value
            except Exception:
                return None

        # ── Horizontal scan: a row of class-name labels + an adjacent row
        # (below OR above, same columns) of matching % shares. ────────────
        best = None
        for r in range(1, max_r + 1):
            label_cols = [c for c in range(1, max_c + 1) if is_class_label(cell(r, c))]
            if len(label_cols) < 2:
                continue
            for other_r in (r + 1, r - 1):
                if other_r < 1 or other_r > max_r:
                    continue
                pairs = []
                for c in label_cols:
                    share = norm_share(cell(other_r, c))
                    if share is not None and 0 < share <= 1.0001:
                        pairs.append((str(cell(r, c)).strip(), share))
                if len(pairs) < 2:
                    continue
                total = sum(s for _, s in pairs)
                if 0.90 <= total <= 1.10:
                    score = (abs(total - 1.0), -len(pairs))
                    if best is None or score < best[0]:
                        best = (score, pairs)
        if best:
            pairs = best[1]
            total = sum(s for _, s in pairs)
            return [{"name": n, "share_pct": round(s / total, 6)} for n, s in pairs]

        # ── Vertical fallback: label in one column, share in an adjacent
        # column, one row per class, a tight contiguous run. ──────────────
        for label_col in range(1, max_c + 1):
            run = []
            for r in range(1, max_r + 1):
                v = cell(r, label_col)
                if not is_class_label(v):
                    continue
                share = None
                for share_col in (label_col + 1, label_col + 2, label_col - 1):
                    if 1 <= share_col <= max_c:
                        s = norm_share(cell(r, share_col))
                        if s is not None and 0 < s <= 1.0001:
                            share = s
                            break
                if share is not None:
                    run.append((str(v).strip(), share, r))
            if len(run) < 2:
                continue
            run.sort(key=lambda x: x[2])
            groups, cur = [], [run[0]]
            for item in run[1:]:
                if item[2] - cur[-1][2] <= 3:
                    cur.append(item)
                else:
                    groups.append(cur)
                    cur = [item]
            groups.append(cur)
            for g in groups:
                if len(g) < 2:
                    continue
                total = sum(s for _, s, _ in g)
                if 0.90 <= total <= 1.10:
                    return [{"name": n, "share_pct": round(s / total, 6)} for n, s, _ in g]

        return None

    def _cam_run_import(entity_code, force=False):
        """One-time importer: parse the building's approved Excel Schedule A-1
        and create CamClass rows (name + share_pct). Idempotent unless
        force=True. force=True replaces the FULL class set for the year (and,
        since overrides key on cam_class_id, drops any per-cell overrides tied
        to the replaced classes — same trade-off as Commercial Rent's force
        re-import). Mirrors _comm_rent_run_import's file-discovery path.
        """
        existing = CamClass.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR).count()
        if existing > 0 and not force:
            return {"status": "exists", "imported": 0, "class_count": existing}

        import tempfile
        try:
            import openpyxl
        except Exception as e:
            return {"status": "error", "error": f"openpyxl not available: {e!r}"}

        try:
            import app as _app_mod  # type: ignore
            files = _app_mod._sharepoint_list_approved_budgets(entity_code)
        except Exception as e:
            return {"status": "error", "error": f"sharepoint list failed: {str(e)[:200]}"}

        if not files:
            return {"status": "no_file", "error": "no approved budget file"}

        files.sort(key=lambda f: f.get("last_modified", ""), reverse=True)
        target = files[0]
        item_id = target.get("item_id")
        if not item_id:
            return {"status": "error", "error": "no item_id on file"}

        try:
            _name, file_bytes = _app_mod._sharepoint_download_item(item_id)
        except Exception as e:
            return {"status": "error", "error": f"download failed: {str(e)[:200]}"}

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = tmp.name
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            parsed = _cam_parse_excel(wb)
        except Exception as e:
            return {"status": "error", "error": f"parse failed: {str(e)[:200]}"}
        finally:
            if tmp_path:
                try:
                    import os as _os
                    _os.unlink(tmp_path)
                except Exception:
                    pass

        if not parsed:
            return {"status": "not_found",
                    "error": "no reconciling CAM class/share block found in the Excel",
                    "file_name": target.get("name")}

        if force and existing > 0:
            CamClass.query.filter_by(
                entity_code=entity_code, budget_year=BUDGET_YEAR).delete()
            db.session.flush()

        for i, c in enumerate(parsed):
            db.session.add(CamClass(
                entity_code=entity_code, budget_year=BUDGET_YEAR,
                name=c["name"][:80], share_pct=c["share_pct"], sort_order=i,
            ))
        db.session.commit()
        return {"status": "imported", "imported": len(parsed), "file_name": target.get("name")}

    def _cam_compute(entity_code, year=None):
        """Compute the CAM allocation matrix (Schedule A-1) for a building.

        Per-(line,class) override wins; else the line's cam_code drives the
        default split — 'B'/'S' = building-wide (x share), 'R'/<class-name> =
        100% to that class, 'SUBSET:a|b' = split across the named classes by
        their re-normalized shares. Rounding residual goes to the largest cell
        so every row reconciles to its line total. Returns the matrix, per-class
        column totals (allocated expense), the grand total, and reconciliation
        flags. (Per-line % is entered client-side and persisted as $ cells.)
        """
        yr = year or BUDGET_YEAR
        budget = Budget.query.filter_by(entity_code=entity_code, year=yr).first()
        classes = (CamClass.query
                   .filter_by(entity_code=entity_code, budget_year=yr)
                   .order_by(CamClass.sort_order, CamClass.id).all())
        result = {
            "classes": [c.to_dict() for c in classes],
            "lines": [],
            "column_totals": {c.id: 0.0 for c in classes},
            "grand_total": 0.0,
            "reconciles": True,
            "shares_ok": False,
            "share_sum": 0.0,
        }
        share_sum = round(sum(float(c.share_pct or 0) for c in classes), 6)
        result["share_sum"] = share_sum
        result["shares_ok"] = bool(classes) and abs(share_sum - 1.0) < 0.0001
        if not budget or not classes:
            return result

        cls_by_name = {(c.name or "").strip().lower(): c for c in classes}
        ovr = {}
        for o in CamAllocationOverride.query.filter_by(budget_id=budget.id).all():
            if o.amount is not None:
                ovr[(o.gl_code, o.cam_class_id)] = float(o.amount)

        lines = (BudgetLine.query
                 .filter_by(budget_id=budget.id)
                 .filter(BudgetLine.sheet_name.in_(CAM_EXPENSE_SHEETS))
                 .order_by(BudgetLine.sheet_name, BudgetLine.row_num).all())

        for l in lines:
            total = round(_cam_line_amount(l), 2)
            if abs(total) < 0.005:
                continue
            code = (l.cam_code or "").strip()
            # The code decides which classes RECEIVE the split (B/S = all;
            # R = residential; <class-name> = that class; SUBSET:a|b = those).
            up = code.upper()
            if up in ("B", "S", ""):
                target_ids = [c.id for c in classes]
            elif up == "R":
                res = cls_by_name.get("residential")
                target_ids = [res.id] if res else [classes[0].id]
            elif up.startswith("SUBSET:"):
                names = [n.strip().lower() for n in code.split(":", 1)[1].split("|")]
                target_ids = [c.id for c in classes
                              if (c.name or "").strip().lower() in names] \
                             or [c.id for c in classes]
            elif code.strip().lower() in cls_by_name:
                target_ids = [cls_by_name[code.strip().lower()].id]
            else:
                target_ids = [c.id for c in classes]

            # Per-(line,class) overrides are FIXED; the code then splits the
            # REMAINING amount across the non-overridden target classes by their
            # relative share. So a single-cell override (e.g. a small Garage
            # portion on an R line) leaves the rest on Residential — matching the
            # Excel's mixed lines (HVAC 3,730 Res / 270 Garage). A full set of
            # overrides just reconciles to the cent. (P6 fix 2026-07-01: a partial
            # override used to zero the other classes + dump the residual.)
            line_ovr = {c.id: round(ovr[(l.gl_code, c.id)], 2) for c in classes
                        if (l.gl_code, c.id) in ovr}
            cells = {c.id: line_ovr.get(c.id, 0.0) for c in classes}
            recv = [c for c in classes if c.id in target_ids and c.id not in line_ovr]
            remainder = round(total - sum(line_ovr.values()), 2)
            if recv and abs(remainder) > 0.005:
                ssum = sum(float(c.share_pct or 0) for c in recv) or 1.0
                for c in recv:
                    cells[c.id] = round(remainder * (float(c.share_pct or 0) / ssum), 2)
            # Reconcile the row to its total (rounding residual → largest receiver,
            # else the largest touched cell).
            diff = round(total - sum(cells.values()), 2)
            if abs(diff) >= 0.01:
                pool = recv or [c for c in classes if abs(cells.get(c.id, 0.0)) > 0.005] or classes
                rc = max(pool, key=lambda c: abs(cells.get(c.id, 0.0)) or float(c.share_pct or 0))
                cells[rc.id] = round(cells.get(rc.id, 0.0) + diff, 2)
            for cid, amt in cells.items():
                result["column_totals"][cid] = round(result["column_totals"][cid] + amt, 2)
            result["grand_total"] = round(result["grand_total"] + total, 2)
            result["lines"].append({
                "gl_code": l.gl_code, "description": l.description,
                "sheet_name": l.sheet_name, "total": total,
                "cam_code": l.cam_code, "cells": cells,
            })
        col_sum = round(sum(result["column_totals"].values()), 2)
        result["reconciles"] = abs(col_sum - result["grand_total"]) < 1.0
        return result

    def _cam_find_summary_row(entity_code, label, class_name=None):
        """Find the Summary income row a CAM class funds. Prefer an exact match
        on the class's summary_row_label; else a Common-Charges/Maintenance row
        whose label contains the class name. Returns the row or None."""
        want = (label or "").strip().lower()
        cname = (class_name or "").strip().lower()
        rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR).all()
        if want:
            for r in rows:
                if (r.label or "").strip().lower() == want:
                    return r
        if cname:
            for r in rows:
                rl = (r.label or "").strip().lower()
                if cname in rl and ("common charge" in rl or "maintenance" in rl):
                    return r
        return None

    def _cam_recompute_summary(entity_code, write=True):
        """Feed each class's allocated-expense column total into its Summary
        common-charge income row (col7 = the common charges that class must
        cover — like _commercial_recompute_summary feeds 4040/4520). A CAM-set
        col7_proposed_budget is read first in the Summary cascade and survives
        the income/fixed-forecast pins (both only fire when col7 is None), so it
        displays exactly like commercial rent. NO-OP (no write) when CAM is
        disabled, so non-CAM / disabled buildings are never touched (silo-safe).
        write=False returns the same mapping WITHOUT committing (GET preview)."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget or not getattr(budget, "cam_enabled", False):
            return {"enabled": False, "classes": []}
        comp = _cam_compute(entity_code)
        ct = comp["column_totals"]
        out = []
        touched = False
        for c in comp["classes"]:
            total = round(float(ct.get(c["id"], 0.0) or 0.0), 2)
            new_col7 = total if abs(total) > 0.005 else None
            row = _cam_find_summary_row(entity_code, c.get("summary_row_label"), c.get("name"))
            out.append({
                "class_id": c["id"], "class_name": c["name"],
                "allocated_expense": total,
                "row_id": row.id if row else None,
                "label": (row.label if row else (c.get("summary_row_label") or c.get("name"))),
                "old_col7": (row.col7_proposed_budget if row else None),
                "new_col7": new_col7,
                "matched": bool(row),
            })
            if row and write:
                row.col7_proposed_budget = new_col7
                row.updated_at = datetime.utcnow()
                touched = True
        if write and touched:
            db.session.commit()
        return {"enabled": True, "grand_total": comp["grand_total"],
                "reconciles": comp["reconciles"], "shares_ok": comp["shares_ok"],
                "classes": out}

    def _cam_compute_required_increase(entity_code):
        """The FA's actual worksheet math (plan: "CAM Allocation -- FA
        feedback on 343", Cluster C): back out non-common-charge income from
        total allocated expense, split what's left by class share, and
        compare to each class's CURRENT common charges to get the $ / %
        increase needed to balance the budget. Read-only -- never written
        anywhere automatically; an FA reviews these numbers like every other
        CAM figure before they mean anything downstream."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return {"error": "Budget not found"}
        comp = _cam_compute(entity_code)
        classes = comp.get("classes") or []
        grand_total = float(comp.get("grand_total") or 0)

        # Match each class to the Summary row carrying its CURRENT common
        # charges -- the same lookup _cam_recompute_summary already uses to
        # feed col7, so "current" and "proposed" always agree on which row
        # is whose.
        matched_rows = {}
        for c in classes:
            matched_rows[c["id"]] = _cam_find_summary_row(entity_code, c.get("summary_row_label"), c.get("name"))

        if classes and not any(matched_rows.values()):
            # Without a matched row for AT LEAST one class, there's no
            # reliable "current common charges" figure to anchor "other
            # income" against -- every Income line would get counted as
            # "other," producing confidently-wrong negative numbers instead
            # of an honest "can't compute this yet."
            return {
                "grand_total_expense": round(grand_total, 2),
                "other_income": None,
                "amount_to_be_covered": None,
                "warning": ("No CAM class matched a Summary common-charge row, so \"other income\" "
                           "can't be reliably identified. Link at least one class to its Summary row "
                           "(the link icon in the class editor) before these figures mean anything."),
                "classes": [{"class_id": c["id"], "class_name": c["name"],
                            "current_common_charges": 0.0, "required_common_charges": None,
                            "increase_dollar": None, "increase_pct": None,
                            "matched_row_label": None} for c in classes],
            }

        # Other income = every Income-sheet line's CURRENT budget, minus the
        # CURRENT common charges recorded directly on the matched Summary
        # row(s) -- anchored on col6 (this year's approved figure, never
        # touched by CAM) rather than trying to trace individual GL lines
        # back to a row via gl_prefixes_json, which is frequently empty even
        # on rows that DID match by name/label (confirmed on 347: both
        # matched "Common Charges" rows carry gl_prefixes_json == [], which
        # silently left every income line -- including the real common
        # charges -- counted as "other," producing a negative
        # amount-to-be-covered). col7 is deliberately NOT used here: CAM
        # itself writes col7 with the allocated expense, so reading it back
        # would make "current vs. required" circular.
        total_current_cc = sum(float(r.col6_approved_budget or 0) for r in matched_rows.values() if r)
        income_lines = BudgetLine.query.filter_by(budget_id=budget.id, sheet_name="Income").all()
        total_income_current = sum(float(l.current_budget or 0) for l in income_lines)
        other_income = round(total_income_current - total_current_cc, 2)

        to_be_covered = round(grand_total - other_income, 2)

        out_classes = []
        for c in classes:
            row = matched_rows.get(c["id"])
            current_cc = round(float(row.col6_approved_budget or 0), 2) if row else 0.0
            required_cc = round(to_be_covered * float(c.get("share_pct") or 0), 2)
            increase_dollar = round(required_cc - current_cc, 2)
            increase_pct = round((increase_dollar / abs(current_cc) * 100) if abs(current_cc) > 0.01 else 0.0, 2)
            out_classes.append({
                "class_id": c["id"], "class_name": c["name"],
                "current_common_charges": current_cc,
                "required_common_charges": required_cc,
                "increase_dollar": increase_dollar,
                "increase_pct": increase_pct,
                "matched_row_label": row.label if row else None,
            })

        return {
            "grand_total_expense": round(grand_total, 2),
            "other_income": other_income,
            "amount_to_be_covered": to_be_covered,
            "classes": out_classes,
        }

    @bp.route("/api/cam/<entity_code>/required-increase", methods=["GET"])
    def api_cam_required_increase(entity_code):
        return jsonify(_cam_compute_required_increase(entity_code))

    @bp.route("/api/admin/cam-audit", methods=["GET"])
    def api_cam_audit():
        """Portfolio-wide CAM data-integrity check: flags any building whose
        CamClass shares don't sum to ~100% -- the pattern behind the 343
        %-parsing incident (2026-07-01). Read-only; added specifically to
        give a definitive answer on whether that bug affected buildings
        beyond the one an FA happened to report, not just the handful
        touched during that session's verification work."""
        from collections import defaultdict
        all_classes = CamClass.query.filter_by(budget_year=BUDGET_YEAR).all()
        by_entity = defaultdict(list)
        for c in all_classes:
            by_entity[c.entity_code].append(c)
        results = []
        for ec, classes in by_entity.items():
            total = sum(float(c.share_pct or 0) for c in classes)
            results.append({
                "entity_code": ec,
                "class_count": len(classes),
                "share_sum_pct": round(total * 100, 4),
                "ok": abs(total - 1.0) < 0.0005,
                "classes": [{"id": c.id, "name": c.name, "share_pct": c.share_pct} for c in classes],
            })
        results.sort(key=lambda r: r["ok"])  # flagged (not ok) ones first
        return jsonify({
            "total_entities_with_cam_classes": len(results),
            "flagged_count": sum(1 for r in results if not r["ok"]),
            "results": results,
        })

    @bp.route("/api/cam/<entity_code>", methods=["GET"])
    def api_cam_get(entity_code):
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        data = _cam_compute(entity_code)
        data["entity_code"] = entity_code
        data["cam_enabled"] = bool(budget and getattr(budget, "cam_enabled", False))
        data["building_type"] = (budget.building_type if budget else "") or ""
        data["summary_sync"] = _cam_recompute_summary(entity_code, write=False)  # read-only preview
        return jsonify(data)

    @bp.route("/api/cam/<entity_code>/import", methods=["POST"])
    def api_cam_import(entity_code):
        """Manual-trigger import of unit classes + shares from the building's
        approved SharePoint Excel (Schedule A-1). Unlike Commercial Rent's
        import (which auto-fires on GET), this is button-only — CAM writes
        stay explicit until the parser is validated across the portfolio.
        ?force=1 wipes + re-imports (like /api/commercial/<ec>/import)."""
        force = request.args.get("force", "0") == "1"
        result = _cam_run_import(entity_code, force=force)
        if result.get("status") == "imported":
            result["summary_sync"] = _cam_recompute_summary(entity_code)
        return jsonify(result)

    @bp.route("/api/cam/<entity_code>/enable", methods=["PUT"])
    def api_cam_enable(entity_code):
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        data = request.get_json() or {}
        budget.cam_enabled = bool(data.get("enabled", True))
        db.session.commit()
        sync = _cam_recompute_summary(entity_code)
        return jsonify({"status": "ok", "cam_enabled": bool(budget.cam_enabled),
                        "summary_sync": sync})

    def _cam_shares_total_with(entity_code, exclude_class_id, candidate_share):
        """What the Σ of all classes' share_pct would be if `exclude_class_id`
        (or no class, for a new one) were set to `candidate_share`. Used to
        hard-block a save that would push the total over 100% -- see the 343
        incident (a 100x parsing bug let shares reach 112.91% live, unblocked).
        Under-100% is allowed through (an FA building up classes one at a time
        is a normal, expected intermediate state)."""
        others = CamClass.query.filter_by(entity_code=entity_code, budget_year=BUDGET_YEAR).all()
        total = sum(float(c.share_pct or 0) for c in others if c.id != exclude_class_id)
        return total + candidate_share

    @bp.route("/api/cam/<entity_code>/class", methods=["POST"])
    def api_cam_class_create(entity_code):
        data = request.get_json() or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "name required"}), 400
        try:
            share_pct = float(data.get("share_pct") or 0)
        except (TypeError, ValueError):
            share_pct = 0.0
        new_total = _cam_shares_total_with(entity_code, None, share_pct)
        if new_total > 1.0 + 0.0001:
            return jsonify({"error": "Classes would total %.4f%% -- cannot exceed 100%%."
                                     % round(new_total * 100, 4)}), 409
        n = CamClass.query.filter_by(entity_code=entity_code, budget_year=BUDGET_YEAR).count()
        c = CamClass(entity_code=entity_code, budget_year=BUDGET_YEAR, name=name,
                     share_pct=share_pct,
                     summary_row_label=data.get("summary_row_label"),
                     notes=data.get("notes"),
                     sort_order=int(data.get("sort_order") or n))
        db.session.add(c)
        db.session.commit()
        result = c.to_dict()
        result["summary_sync"] = _cam_recompute_summary(entity_code)
        return jsonify(result)

    @bp.route("/api/cam/<entity_code>/class/<int:class_id>", methods=["PUT"])
    def api_cam_class_update(entity_code, class_id):
        c = CamClass.query.filter_by(id=class_id, entity_code=entity_code,
                                     budget_year=BUDGET_YEAR).first()
        if not c:
            return jsonify({"error": "class not found"}), 404
        data = request.get_json() or {}
        if "name" in data:
            c.name = (data["name"] or "").strip() or c.name
        if "summary_row_label" in data:
            c.summary_row_label = data["summary_row_label"]
        if "notes" in data:
            c.notes = data["notes"]
        if "share_pct" in data:
            try:
                new_share = float(data["share_pct"] or 0)
            except (TypeError, ValueError):
                new_share = None
            if new_share is not None:
                new_total = _cam_shares_total_with(entity_code, c.id, new_share)
                if new_total > 1.0 + 0.0001:
                    return jsonify({"error": "Classes would total %.4f%% -- cannot exceed 100%%."
                                             % round(new_total * 100, 4)}), 409
                c.share_pct = new_share
        if "sort_order" in data:
            try:
                c.sort_order = int(data["sort_order"])
            except (TypeError, ValueError):
                pass
        db.session.commit()
        result = c.to_dict()
        result["summary_sync"] = _cam_recompute_summary(entity_code)
        return jsonify(result)

    @bp.route("/api/cam/<entity_code>/class/<int:class_id>", methods=["DELETE"])
    def api_cam_class_delete(entity_code, class_id):
        c = CamClass.query.filter_by(id=class_id, entity_code=entity_code,
                                     budget_year=BUDGET_YEAR).first()
        if not c:
            return jsonify({"error": "class not found"}), 404
        CamAllocationOverride.query.filter_by(cam_class_id=c.id).delete()
        db.session.delete(c)
        db.session.commit()
        sync = _cam_recompute_summary(entity_code)
        return jsonify({"status": "deleted", "id": class_id, "summary_sync": sync})

    @bp.route("/api/cam/<entity_code>/line-code", methods=["PUT"])
    def api_cam_line_code(entity_code):
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        data = request.get_json() or {}
        gl = (data.get("gl_code") or "").strip()
        line = BudgetLine.query.filter_by(budget_id=budget.id, gl_code=gl).first()
        if not line:
            return jsonify({"error": "line not found"}), 404
        line.cam_code = (data.get("cam_code") or None)
        db.session.commit()
        sync = _cam_recompute_summary(entity_code)
        return jsonify({"status": "ok", "gl_code": gl, "cam_code": line.cam_code,
                        "summary_sync": sync})

    @bp.route("/api/cam/<entity_code>/cell", methods=["PUT"])
    def api_cam_cell(entity_code):
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        data = request.get_json() or {}
        gl = (data.get("gl_code") or "").strip()
        cid = data.get("cam_class_id")
        amt = data.get("amount")
        o = CamAllocationOverride.query.filter_by(
            budget_id=budget.id, gl_code=gl, cam_class_id=cid).first()
        if amt is None or amt == "":
            if o:
                db.session.delete(o)   # clear → revert to code default
        else:
            try:
                amt = float(amt)
            except (TypeError, ValueError):
                return jsonify({"error": "bad amount"}), 400
            if o:
                o.amount = amt
            else:
                db.session.add(CamAllocationOverride(
                    budget_id=budget.id, gl_code=gl, cam_class_id=cid, amount=amt))
        db.session.commit()
        sync = _cam_recompute_summary(entity_code)
        return jsonify({"status": "ok", "summary_sync": sync})


    # ─── Board Presentation (client narrative) ─────────────────────────────
    # Board Presentation redesign (2026-07-01, plan: "Client Board Presentation").
    # Replaces two dead surfaces (openBoardPresentation() overlay + the orphaned
    # /presentation/<token> route) with a system-drafted, FA-reviewed narrative
    # that only becomes client-visible after explicit FA sign-off. Mirrors
    # AuditUpload's draft -> reviewed -> confirmed review-gate pattern.

    NARRATIVE_EXPENSE_SHEETS = ("Payroll", "Energy", "Water & Sewer",
                                "Repairs & Supplies", "Gen & Admin",
                                # 2026-07-05: RE Taxes sheet lines belong in the
                                # client story — without them, buildings with a
                                # dedicated RE-tax sheet (829/210 class) understate
                                # the headline and the detail by the tax amount.
                                "RE Taxes", "Capital")

    def _narrative_line_amounts(l):
        """(current, proposed) for one BudgetLine, mirroring _cam_line_amount's
        proposed-falls-back-to-current rule so a not-yet-budgeted line doesn't
        read as a $0 cliff in the narrative."""
        try:
            cur = float(l.current_budget or 0)
        except (TypeError, ValueError):
            cur = 0.0
        try:
            p = float(l.proposed_budget) if l.proposed_budget is not None else 0.0
        except (TypeError, ValueError):
            p = 0.0
        return cur, (p if abs(p) > 0.005 else cur)

    def _generate_client_narrative(budget):
        """Draft the client-facing Board Presentation narrative from real
        budget data. Every dollar figure is computed directly from BudgetLine;
        every "why" is a neutral, factual statement of WHAT changed, never an
        invented external cause (no fabricated "due to a new contract" claims)
        — an FA adds that color explicitly during review (BudgetNarrative
        .reviewed_narrative), it is never auto-asserted as fact.
        """
        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
        bt = (budget.building_type or "").strip().lower()
        is_condo = bt in ("condo", "condominium", "cond-op")
        charge_word = "common charge" if is_condo else "maintenance"

        by_sheet = {}
        for l in lines:
            sn = l.sheet_name or "Other"
            if sn != "Income" and sn not in NARRATIVE_EXPENSE_SHEETS:
                continue
            cur, prop = _narrative_line_amounts(l)
            d = by_sheet.setdefault(sn, {"current": 0.0, "proposed": 0.0})
            d["current"] += cur
            d["proposed"] += prop

        income = by_sheet.get("Income", {"current": 0.0, "proposed": 0.0})
        exp_current = sum(d["current"] for sn, d in by_sheet.items() if sn in NARRATIVE_EXPENSE_SHEETS)
        exp_proposed = sum(d["proposed"] for sn, d in by_sheet.items() if sn in NARRATIVE_EXPENSE_SHEETS)
        net_change = round(exp_proposed - exp_current, 2)
        pct_change = round((net_change / abs(exp_current) * 100) if abs(exp_current) > 0.01 else 0.0, 1)

        movers = []
        for sn in NARRATIVE_EXPENSE_SHEETS:
            d = by_sheet.get(sn)
            if not d:
                continue
            chg = round(d["proposed"] - d["current"], 2)
            if abs(chg) < 0.5:
                continue
            movers.append({
                "category": sn, "current": round(d["current"], 2),
                "proposed": round(d["proposed"], 2), "change": chg,
                "pct": round((chg / abs(d["current"]) * 100) if abs(d["current"]) > 0.01 else 0.0, 1),
            })
        movers.sort(key=lambda m: -m["change"])
        increases = [m for m in movers if m["change"] > 0]
        decreases = [m for m in movers if m["change"] < 0]
        savings_total = round(-sum(m["change"] for m in decreases), 2)
        drivers_text = " and ".join(m["category"] for m in increases[:2]) or "operating cost changes"

        owner_word = "unit owner" if is_condo else "shareholder"
        gov_doc = "bylaws" if is_condo else "proprietary lease"
        entity_word = "condominium association" if is_condo else "cooperative"

        faq = [
            {"q": f"Why can't these increases be absorbed without raising {charge_word}s?",
             "a": (f"As a not-for-profit {entity_word}, {budget.building_name} has no operating "
                   f"margin to absorb rising costs. {charge_word.capitalize()}s are the sole funding "
                   f"source for building operations, and expenses are passed through to "
                   f"{owner_word}s directly under the building's {gov_doc}.")},
            {"q": "Is this increase specific to this building, or happening everywhere?",
             "a": ("[FA: note here whether specific drivers reflect portfolio-wide trends "
                   "(e.g. insurance market conditions) versus something specific to this building.]")},
            {"q": "What happens if the Board does not approve this budget?",
             "a": (f"The building would continue operating under the {budget.year - 1} budget, "
                   f"which may not fund the cost changes described above. The Board would need to "
                   f"either draw down reserves to cover any shortfall or work with Century on a "
                   f"revised proposal.")},
            {"q": f"When would the new {charge_word}s take effect?",
             "a": (f"If approved as proposed, the new {charge_word} schedule takes effect "
                   f"{budget.effective_date}." if budget.effective_date else
                   f"[FA: enter the proposed effective date once set.]")},
        ]

        return {
            "building_type_words": {"charge_word": charge_word, "owner_word": owner_word,
                                    "entity_word": entity_word, "gov_doc": gov_doc, "is_condo": is_condo},
            "opening": (f"Century Management has completed its analysis of the proposed "
                       f"{budget.year} operating budget for {budget.building_name}, submitted "
                       f"herewith for the Board's review ahead of your scheduled budget call."),
            "headline": {"pct_change": pct_change, "net_change": net_change,
                        "exp_current": round(exp_current, 2), "exp_proposed": round(exp_proposed, 2)},
            "driver_summary": (
                f"The proposed budget reflects a {pct_change:+.1f}% change in operating expense, "
                f"driven primarily by {drivers_text}." if increases else
                f"The proposed budget reflects a {pct_change:+.1f}% change in operating expense."),
            "drivers": increases[:4],
            "savings": decreases,
            "savings_total": savings_total,
            "income": {"current": round(income["current"], 2), "proposed": round(income["proposed"], 2)},
            "categories": [{"sheet": sn, **d} for sn, d in by_sheet.items() if sn in NARRATIVE_EXPENSE_SHEETS],
            "faq": faq,
            "timeline": {"effective_date": budget.effective_date or "",
                        "board_review_through": "", "board_vote_by": ""},  # FA fills in during review
            "additional_notes": "",  # e.g. reserve-fund status — FA-authored, not derivable from BudgetLine
            "generated_at": datetime.utcnow().isoformat(),
        }

    # ── Full Budget Detail (plan: "Board Presentation — Full Budget Detail",
    # 2026-07-01) — every tab an FA sees, read-only, numbers only, behind the
    # narrative memo above. CAM excluded until it clears its own review gate.

    CLIENT_DETAIL_SHEETS = ("Income",) + NARRATIVE_EXPENSE_SHEETS

    def _client_safe_line(l, ytd_months=2):
        """Allowlist a BudgetLine down to what a client may see: the full
        numeric trail (prior-year actual, YTD actual, current budget,
        forecast, proposed) plus the variance derived from current→proposed.
        Never GL codes, formulas, notes, or FA/PM review-state fields.
        Jacob directive 2026-07-03 (Concept C): every budget number belongs
        in the presentation's backup detail, not just the two-year compare."""
        cur, prop = _narrative_line_amounts(l)
        variance = round(prop - cur, 2)
        try:
            fcst = compute_forecast(l.ytd_actual or 0, l.accrual_adj or 0,
                                    l.unpaid_bills or 0, l.prior_year or 0, ytd_months)
        except Exception:
            fcst = 0.0
        return {
            "description": l.description or l.category or "",
            "category": l.category or "",
            "prior_actual": round(float(l.prior_year or 0), 2),
            "ytd_actual": round(float(l.ytd_actual or 0), 2),
            "current": round(cur, 2),
            "forecast": round(float(fcst or 0), 2),
            "proposed": round(prop, 2),
            "variance": variance,
            "variance_pct": round((variance / abs(cur) * 100) if abs(cur) > 0.01 else 0.0, 1),
        }

    def _add_line_bars(rows, is_income):
        """Mutates each line dict in place: favorable (income up / expense
        down = good) and bar_pct (magnitude relative to this tab's own
        biggest variance, 0 when every line is unchanged) -- an HTML port
        of the Excel At-a-Glance DataBarRule pattern (~8625-8628:
        start_type=min, end_type=max) rather than a new visual language."""
        tab_max = max((abs(r["variance"]) for r in rows), default=0)
        for r in rows:
            r["favorable"] = (r["variance"] >= 0) if is_income else (r["variance"] <= 0)
            r["bar_pct"] = round(abs(r["variance"]) / tab_max * 100, 1) if tab_max else 0.0
        return rows

    def _generate_client_detail_tabs(budget):
        """Full-tab, read-only, numbers-only detail behind the narrative memo.
        Snapshot-isolation invariant: computed ONCE at publish time and frozen
        into PresentationSession.snapshot_data — the client route never calls
        this (or any compute it wires in) live. See check_client_narrative_publish.py."""
        entity_code = budget.entity_code
        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()

        # How many months of actuals are in YTD -- same "budget_period"
        # assumption ("8/12" style) the rest of the app reads; the client
        # footnote and forecast math both depend on it.
        ytd_months = 2
        try:
            if budget.assumptions_json:
                bp_val = json.loads(budget.assumptions_json).get("budget_period", "")
                if "/" in str(bp_val):
                    ytd_months = int(str(bp_val).split("/")[0])
        except Exception:
            pass

        by_sheet = {}
        for l in lines:
            sn = l.sheet_name or "Other"
            if sn not in CLIENT_DETAIL_SHEETS:
                continue
            row = _client_safe_line(l, ytd_months)
            # Jacob 2026-07-05: lines with zeros across the board are noise in
            # a client document — drop them (mirrors the Excel export's
            # line_has_data rule; totals unchanged since they sum to zero).
            if not any(abs(row.get(_f) or 0) > 0.005 for _f in
                       ("prior_actual", "ytd_actual", "current", "forecast", "proposed")):
                continue
            by_sheet.setdefault(sn, []).append(row)

        tabs = []
        for sn in CLIENT_DETAIL_SHEETS:
            rows = by_sheet.get(sn)
            if rows:
                _add_line_bars(rows, is_income=(sn == "Income"))
                tabs.append({"name": sn, "lines": rows})

        # Summary — QA fix 11 (2026-07-03): subtotal/total rows persist col7=0
        # and are computed at render time, so raw to_dict() rows shipped $0
        # totals into the client document. Use the SAME computed view the FA
        # sees (api_summary_get runs the col7 cascade + total-row math; it is
        # a pure read, no commits), mapped onto the template's stored-row
        # keys. Falls back to raw rows if the computed view errors.
        summary_tab_rows = []
        try:
            _resp = api_summary_get(entity_code)
            _payload = _resp[0] if isinstance(_resp, tuple) else _resp
            _sdata = _payload.get_json(silent=True) or {}
            for _r in (_sdata.get("rows") or []):
                if (_r.get("row_type") == "data"
                        and not any(abs(_r.get(_c) or 0) > 0.005 for _c in ("col1", "col6", "col7"))):
                    continue  # Jacob 2026-07-05: all-zero data rows stay out of the client doc
                summary_tab_rows.append({
                    "label": _r.get("label"),
                    "row_type": _r.get("row_type"),
                    "col1_prior_actual": _r.get("col1"),
                    "col6_approved_budget": _r.get("col6"),
                    "col7_proposed_budget": _r.get("col7"),
                })
        except Exception:
            summary_tab_rows = []
        if not summary_tab_rows:
            summary_rows = BudgetSummaryRow.query.filter_by(
                entity_code=entity_code, budget_year=BUDGET_YEAR
            ).order_by(BudgetSummaryRow.display_order).all()
            summary_tab_rows = [r.to_dict() for r in summary_rows]
        if summary_tab_rows:
            tabs.append({"name": "Summary", "rows": summary_tab_rows})

        # ── Reconciliation disclosure (2026-07-05) ──────────────────────────
        # The detail tabs value lines with the flat-fallback rule (proposed =
        # stored else current) while the Summary tab carries Century's computed
        # totals (pins + cascade). When the two part ways by more than 1% or
        # $1,000, disclose the difference under the Summary table instead of
        # letting a board discover it with a calculator.
        recon = {}
        try:
            sum_exp = sum_inc = None
            for _r in summary_tab_rows or []:
                _lbl = (_r.get("label") or "").strip().lower()
                _v = _r.get("col7_proposed_budget")
                if _lbl == "total income" and isinstance(_v, (int, float)):
                    sum_inc = _v
                elif _lbl == "total expenses" and isinstance(_v, (int, float)):
                    sum_exp = _v
            detail_exp = 0.0
            for _sn in NARRATIVE_EXPENSE_SHEETS:
                if _sn == "Capital":
                    continue  # capital is outside the operating Total Expenses
                for _row in by_sheet.get(_sn) or []:
                    detail_exp += _row.get("proposed") or 0
            detail_inc = sum((_row.get("proposed") or 0) for _row in (by_sheet.get("Income") or []))
            if sum_exp and abs(detail_exp - sum_exp) > max(abs(sum_exp) * 0.01, 1000):
                recon["detail_exp"] = round(detail_exp, 2)
                recon["summary_exp"] = round(sum_exp, 2)
                recon["diff_exp"] = round(abs(detail_exp - sum_exp), 2)
            if sum_inc and abs(detail_inc - sum_inc) > max(abs(sum_inc) * 0.01, 1000):
                recon["detail_inc"] = round(detail_inc, 2)
                recon["summary_inc"] = round(sum_inc, 2)
                recon["diff_inc"] = round(abs(detail_inc - sum_inc), 2)
        except Exception:
            recon = {}

        # RE Taxes — coop-only, numbers only (no exemption-formula mechanics).
        # Mirrors the overrides-extraction + safe-fallback pattern used by
        # _export_rewrite_re_taxes (~9026) — a failed/slow DOF lookup skips
        # this tab rather than blocking the whole publish.
        try:
            try:
                from dof_taxes import is_coop, compute_re_taxes
            except ImportError:
                from budget_app.dof_taxes import is_coop, compute_re_taxes
            if is_coop(entity_code):
                rt = compute_re_taxes(entity_code, _re_tax_overrides_for(budget))
                if rt and abs(float(rt.get("gross_tax") or 0)) > 0.5:
                    tabs.append({"name": "RE Tax Bill", "re_taxes": {
                        "gross_tax": round(float(rt.get("gross_tax") or 0), 2),
                        "net_tax": round(float(rt.get("net_tax") or 0), 2),
                        "first_half_tax": round(float(rt.get("first_half_tax") or 0), 2),
                        "second_half_tax": round(float(rt.get("second_half_tax") or 0), 2),
                    }})
        except Exception:
            pass  # RE-tax detail is supplementary; never block publish on it.

        # Commercial — tenant-based (not GL-based), so "current/proposed" is
        # keyed by tenant, not by line description. Read-only: calls
        # _commercial_compute_escalations directly, never
        # _commercial_recompute_summary (that one writes to the DB).
        try:
            tenants = CommercialTenant.query.filter_by(
                entity_code=entity_code, budget_year=BUDGET_YEAR
            ).order_by(CommercialTenant.sort_order, CommercialTenant.id).all()
            if tenants:
                escalations = {e["tenant_id"]: e for e in
                               _commercial_compute_escalations(entity_code, BUDGET_YEAR)}
                comm_rows = []
                for t in tenants:
                    periods = CommercialRentPeriod.query.filter_by(tenant_id=t.id).all()
                    cur_rent = sum(p.annualized() for p in periods if p.year == BUDGET_YEAR - 1)
                    next_periods = [p for p in periods if p.year == BUDGET_YEAR]
                    if next_periods:
                        prop_rent = sum(p.annualized() for p in next_periods)
                    else:
                        esc = escalations.get(t.id, {}).get("amount") or 0.0
                        prop_rent = cur_rent + float(esc)
                    if abs(cur_rent) <= 0.005 and abs(prop_rent) <= 0.005:
                        continue  # Jacob 2026-07-05: all-zero rows stay out of the client doc
                    variance = round(prop_rent - cur_rent, 2)
                    comm_rows.append({
                        "description": t.tenant_name + (f" ({t.unit_label})" if t.unit_label else ""),
                        "category": "Commercial",
                        # Tenant-based rows have no GL prior/YTD/forecast trail --
                        # None renders as an em-dash client-side, never a fake 0.
                        "prior_actual": None, "ytd_actual": None, "forecast": None,
                        "current": round(cur_rent, 2), "proposed": round(prop_rent, 2),
                        "variance": variance,
                        "variance_pct": round((variance / abs(cur_rent) * 100) if abs(cur_rent) > 0.01 else 0.0, 1),
                    })
                if comm_rows:
                    _add_line_bars(comm_rows, is_income=True)
                    tabs.append({"name": "Commercial", "lines": comm_rows})
        except Exception:
            pass  # Commercial detail is supplementary; never block publish on it.

        # Overview chart data, pre-computed server-side (simpler/more robust
        # than doing this math in Jinja or JS). Donut = expense mix by the 6
        # narrative categories; bars = current vs proposed per category.
        cat_totals = []
        for sn in NARRATIVE_EXPENSE_SHEETS:
            rows = by_sheet.get(sn)
            if not rows:
                continue
            cat_totals.append({"name": sn, "current": sum(r["current"] for r in rows),
                               "proposed": sum(r["proposed"] for r in rows)})
        total_proposed = sum(c["proposed"] for c in cat_totals) or 1.0
        max_amt = max([c["current"] for c in cat_totals] + [c["proposed"] for c in cat_totals] + [1.0])
        DONUT_COLORS = ["#001721", "#DE1C23", "#8a7e72", "#5b7a8c", "#a8763e", "#4d5d53"]
        cum = 0.0
        donut_slices = []
        for i, c in enumerate(cat_totals):
            pct = round(c["proposed"] / total_proposed * 100, 1) if total_proposed else 0.0
            donut_slices.append({"name": c["name"], "pct": pct, "start": round(cum, 1),
                                 "color": DONUT_COLORS[i % len(DONUT_COLORS)]})
            cum += pct
        bars = [{"name": c["name"],
                "current_pct": round(c["current"] / max_amt * 100, 1),
                "proposed_pct": round(c["proposed"] / max_amt * 100, 1),
                "current": round(c["current"], 2), "proposed": round(c["proposed"], 2)}
               for c in cat_totals]

        # "Biggest changes" -- top individual EXPENSE-line movers across
        # every sheet (Income/Commercial excluded: the hero stat + driver
        # summary already tell the category-level income story; this is
        # the line-level "what's actually costing more" drill-down).
        # Mirrors the Excel "Biggest Changes vs Approved Budget" list
        # (~8607-8630) rather than a new concept.
        expense_lines = []
        for sn in NARRATIVE_EXPENSE_SHEETS:
            for r in (by_sheet.get(sn) or []):
                if abs(r["variance"]) > 0.5:
                    expense_lines.append(r)
        expense_lines.sort(key=lambda r: abs(r["variance"]), reverse=True)
        top_movers = expense_lines[:6]
        movers_max = max((abs(r["variance"]) for r in top_movers), default=0)
        movers = [{"label": r["description"], "change": r["variance"], "favorable": r["favorable"],
                  "bar_pct": round(abs(r["variance"]) / movers_max * 100, 1) if movers_max else 0.0}
                 for r in top_movers]

        return {"tabs": tabs, "chart_data": {"donut": donut_slices, "bars": bars, "movers": movers},
                "meta": {"ytd_months": ytd_months, "recon": recon}}

    @bp.route("/api/board-notice/<entity_code>", methods=["GET"])
    def api_board_notice_get(entity_code):
        """Fetch (auto-drafting on first call) the Board Presentation narrative
        for FA review. Never regenerates over an existing reviewed/published
        narrative — regeneration is an explicit action (?regenerate=1)."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        narrative = BudgetNarrative.query.filter_by(budget_id=budget.id).first()
        regenerate = request.args.get("regenerate", "0") == "1"
        if not narrative:
            narrative = BudgetNarrative(budget_id=budget.id)
            db.session.add(narrative)
        if not narrative.raw_narrative or regenerate:
            narrative.raw_narrative = json.dumps(_generate_client_narrative(budget))
            if regenerate and narrative.status != "draft":
                # Regenerating after review/publish starts a fresh review cycle —
                # never let a stale reviewed_narrative silently persist alongside
                # newly regenerated numbers.
                narrative.reviewed_narrative = None
                narrative.status = "draft"
        db.session.commit()
        result = narrative.to_dict()
        result["active"] = json.loads(narrative.reviewed_narrative) if narrative.reviewed_narrative else json.loads(narrative.raw_narrative)
        return jsonify(result)

    @bp.route("/api/board-notice/<entity_code>", methods=["PUT"])
    def api_board_notice_save(entity_code):
        """FA saves edits to the narrative. mark_reviewed=true additionally
        stamps reviewed_by/reviewed_at and advances status to 'reviewed' — the
        gate that must pass before /publish will create a client link."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        narrative = BudgetNarrative.query.filter_by(budget_id=budget.id).first()
        if not narrative:
            return jsonify({"error": "No draft narrative — GET /api/board-notice first"}), 404
        data = request.get_json(silent=True) or {}
        if "narrative" in data:
            narrative.reviewed_narrative = json.dumps(data["narrative"])
        if data.get("mark_reviewed"):
            narrative.status = "reviewed"
            narrative.reviewed_by = (data.get("reviewed_by") or "")[:120]
            narrative.reviewed_at = datetime.utcnow()
        db.session.commit()
        return jsonify(narrative.to_dict())

    @bp.route("/api/board-notice/<entity_code>/publish", methods=["POST"])
    def api_board_notice_publish(entity_code):
        """Create the client-facing link. Requires the narrative to already be
        'reviewed' (an explicit FA sign-off, per plan decision #1/#4) — a
        'draft' narrative can never be published directly. Snapshots the
        current BudgetLine/BudgetSummaryRow + the reviewed narrative into
        PresentationSession.snapshot_data so a later budget edit can't
        silently change what the client already received."""
        import secrets
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return jsonify({"error": "Budget not found"}), 404
        narrative = BudgetNarrative.query.filter_by(budget_id=budget.id).first()
        if not narrative or narrative.status not in ("reviewed", "published") or not narrative.reviewed_narrative:
            return jsonify({"error": "Narrative must be reviewed and saved before it can be published. "
                                     "Save your edits with mark_reviewed=true first."}), 409

        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
        summary_rows = BudgetSummaryRow.query.filter_by(
            entity_code=entity_code, budget_year=BUDGET_YEAR).order_by(BudgetSummaryRow.display_order).all()
        snapshot = {
            "budget": {"building_name": budget.building_name, "entity_code": budget.entity_code,
                      "year": budget.year, "building_type": budget.building_type},
            "narrative": json.loads(narrative.reviewed_narrative),
            "lines": [l.to_dict() for l in lines],
            "summary_rows": [r.to_dict() for r in summary_rows],
            "detail_tabs": _generate_client_detail_tabs(budget),
            "snapshot_at": datetime.utcnow().isoformat(),
        }

        session_row = PresentationSession.query.filter_by(budget_id=budget.id, is_active=True).first()
        if not session_row:
            session_row = PresentationSession(
                budget_id=budget.id, token=secrets.token_urlsafe(32), created_by=None)
            db.session.add(session_row)
        session_row.snapshot_data = json.dumps(snapshot)
        session_row.is_active = True
        narrative.status = "published"
        db.session.commit()

        url = request.host_url.rstrip("/") + "/board-notice/" + session_row.token
        return jsonify({"status": "ok", "token": session_row.token, "url": url})

    @bp.route("/api/board-notice/<entity_code>/preview", methods=["GET"])
    def api_board_notice_preview(entity_code):
        """FA-side live preview of the client document — renders the exact
        BOARD_NOTICE_TEMPLATE from CURRENT budget data + the narrative as it
        stands (reviewed edits if any, else the draft), WITHOUT creating a
        PresentationSession or touching narrative.status. The 'see it before
        you publish' view; the client-facing board_notice_view stays
        snapshot-isolated and is unaffected by this route."""
        budget = Budget.query.filter_by(entity_code=entity_code, year=BUDGET_YEAR).first()
        if not budget:
            return "No budget found for this building", 404
        narrative = BudgetNarrative.query.filter_by(budget_id=budget.id).first()
        if narrative and narrative.reviewed_narrative:
            narr = json.loads(narrative.reviewed_narrative)
        elif narrative and narrative.raw_narrative:
            narr = json.loads(narrative.raw_narrative)
        else:
            narr = _generate_client_narrative(budget)  # in-memory only; nothing persisted

        snapshot = {
            "budget": {"building_name": budget.building_name, "entity_code": budget.entity_code,
                      "year": budget.year, "building_type": budget.building_type},
            "narrative": narr,
            "detail_tabs": _generate_client_detail_tabs(budget),
        }

        def _fmt_money(n):
            try:
                n = float(n)
            except (TypeError, ValueError):
                return "$0"
            return ("-$" if n < 0 else "$") + f"{abs(n):,.0f}"

        html = render_template_string(BOARD_NOTICE_TEMPLATE, snapshot=snapshot, fmt=_fmt_money)
        # Static (non-sticky) banner so it scrolls away and the document's own
        # sticky sub-nav takes over cleanly.
        banner = ('<div style="background:#92400e;color:#fff;font-family:-apple-system,sans-serif;'
                  'font-size:13px;font-weight:700;text-align:center;padding:9px;letter-spacing:0.03em;">'
                  'PREVIEW &mdash; not published. This is what the board will see once you publish; '
                  'numbers reflect the budget as of right now.</div>')
        return html.replace("<body>", "<body>" + banner, 1)

    @bp.route("/board-notice/<token>", methods=["GET"])
    def board_notice_view(token):
        """Client-facing Board Presentation. Renders ONLY from the frozen
        snapshot captured at publish time — never a live BudgetLine/
        BudgetSummaryRow query — so this page cannot change underneath a
        board member after they've been sent the link."""
        session_row = PresentationSession.query.filter_by(token=token, is_active=True).first()
        if not session_row or not session_row.snapshot_data:
            return "<h1>Presentation not found</h1><p>This link may have expired or is invalid.</p>", 404
        if session_row.expires_at and session_row.expires_at < datetime.utcnow():
            return "<h1>This link has expired</h1><p>Please contact your Century Financial Analyst for an updated link.</p>", 410
        snapshot = json.loads(session_row.snapshot_data)

        def _fmt_money(n):
            try:
                n = float(n)
            except (TypeError, ValueError):
                return "$0"
            return ("-$" if n < 0 else "$") + f"{abs(n):,.0f}"

        return render_template_string(BOARD_NOTICE_TEMPLATE, snapshot=snapshot, fmt=_fmt_money)


    # ─── HTML Templates ─────────────────────────────────────────────────────

    return (bp, {"User": User, "BuildingAssignment": BuildingAssignment, "Budget": Budget, "BudgetLine": BudgetLine, "BudgetRevision": BudgetRevision, "BuildingVisit": BuildingVisit, "PayrollPosition": PayrollPosition, "PayrollAssumption": PayrollAssumption, "BudgetSummaryRow": BudgetSummaryRow, "BuildingInfo": BuildingInfo, "AuditSyncRun": AuditSyncRun, "DataSource": DataSource, "CommercialTenant": CommercialTenant, "CommercialRentPeriod": CommercialRentPeriod, "CommercialTenantBillback": CommercialTenantBillback},
            {"store_rm_lines": store_rm_lines, "store_all_lines": store_all_lines,
             "get_pm_projections": get_pm_projections,
             "compute_forecast": compute_forecast, "compute_proposed_budget": compute_proposed_budget})


# ─── HTML Template Strings ───────────────────────────────────────────────────

# ── Page templates ──────────────────────────────────────────────────────
# Extracted to page_templates/ 2026-07-05 (clean-architecture tranche 1):
# 7 byte-identical constants, one module per page. Template edits happen
# in that package now; the JS/math/invariant deploy gates scan it too.
try:
    from page_templates import (
        ADMIN_TEMPLATE, DASHBOARD_TEMPLATE, ACTION_CENTER_TEMPLATE,
        BUILDING_DETAIL_TEMPLATE, PM_PORTAL_TEMPLATE, PM_EDIT_TEMPLATE,
        BOARD_NOTICE_TEMPLATE,
    )
except ImportError:
    from budget_app.page_templates import (
        ADMIN_TEMPLATE, DASHBOARD_TEMPLATE, ACTION_CENTER_TEMPLATE,
        BUILDING_DETAIL_TEMPLATE, PM_PORTAL_TEMPLATE, PM_EDIT_TEMPLATE,
        BOARD_NOTICE_TEMPLATE,
    )
