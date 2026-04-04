"""
Excel Export V2 — Full Budget Download with FA/PM Edits (TEST / REFERENCE)

STATUS: Test script only. NOT wired into the app. Does not modify any existing files.
Run: py excel_export_v2.py  →  generates test_budget_export_v4.xlsx

What this does:
1. Pulls real data from production API for entity 204
2. Copies Budget_Final_Template_v2.xlsx as a base (preserves formatting)
3. Fills ALL fields including FA/PM edits (accrual, unpaid, inc%, notes, overrides)
4. Uses Excel formulas for calculated fields (Estimate, Forecast, Proposed, Variance, %Chg)
5. Fixes Payroll template's broken L-column formula references (labels were in L, values
   in M; formulas rewritten to point at M)

PATH TO PRODUCTION (when ready to make this a real feature):
- Extract fill_type1, fill_type2, fill_re_taxes, fill_payroll, build_all_lines_sheet
  into budget_app/excel_exporter.py
- Replace body of /api/download-budget/<entity_code> endpoint in workflow.py (~line 1950)
  to call new exporter instead of template_populator.populate_template()

KNOWN ISSUES (to fix before production):
- Budget Summary's RE Taxes total pulls from Gen & Admin 6315 GL lines, not from the
  RE Taxes calculation sheet. Need to wire B25 back into the 6315 GL values.
- Payroll GL rows Inc % (column L) now reflects per-line FA edits, overwriting the
  template's =$L$8 global wage-increase link. May need to reconsider which source
  wins when both are set.
- Payroll template had a bug: formulas referenced $L$7/$L$8/$L$10/$L$11 (text labels)
  instead of $M$7/$M$8/$M$10/$M$11 (values). This script patches those refs on copy.

See: plans/iridescent-booping-dongarra.md for original design doc.
"""

import requests
import shutil
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ── Config ──────────────────────────────────────────────────────────────────

API_BASE = "https://century-budget-generator-production.up.railway.app"
ENTITY = "204"
TEMPLATE_PATH = Path(__file__).parent / "budget_system" / "Budget_Final_Template_v2.xlsx"
OUTPUT_PATH = Path(__file__).parent / "test_budget_export_v4.xlsx"

GL_CODE_RE = re.compile(r"^\d{4}-\d{4}$")

# Type 1 sheets: full 15-column layout (A-O) with GL codes, accrual, etc.
TYPE1_SHEETS = ["Repairs & Supplies", "Gen & Admin"]

# Type 2 sheets: custom layouts with specialized calculation sections
# (Payroll handled separately by fill_payroll)
TYPE2_SHEETS = ["Income", "Energy", "Water & Sewer"]

ALL_GL_SHEETS = TYPE1_SHEETS + TYPE2_SHEETS + ["Payroll"]


# ── Step 1: Fetch data ──────────────────────────────────────────────────────

def fetch_data():
    """Pull dashboard data for entity 204 from production."""
    print("Fetching data from production API...")
    resp = requests.get(f"{API_BASE}/api/dashboard/{ENTITY}")
    resp.raise_for_status()
    data = resp.json()
    budget = data.get("budget", {})
    print(f"  Building: {budget.get('building_name', '?')}")
    print(f"  Lines: {len(data.get('lines', []))}")
    print(f"  YTD Months: {data.get('ytd_months', '?')}")
    return data


# ── Step 2: Build GL mapping from template ──────────────────────────────────

def build_gl_mapping(wb):
    """Scan template sheets for GL codes in column A. Returns {gl_code: (sheet_name, row)}."""
    mapping = {}
    for sn in ALL_GL_SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            val = str(ws.cell(r, 1).value or "").strip()
            if GL_CODE_RE.match(val):
                mapping[val] = (sn, r)
    return mapping


# ── Step 3: Fill Setup sheet ────────────────────────────────────────────────

def fill_setup(wb, data):
    """Fill Setup sheet with entity info and time periods."""
    if "Setup" not in wb.sheetnames:
        print("  WARNING: No Setup sheet found")
        return
    ws = wb["Setup"]
    budget = data.get("budget", {})
    ws["B4"] = budget.get("entity_code", ENTITY)
    ws["B5"] = budget.get("building_name", "")
    ws["B10"] = data.get("ytd_months", 2)
    ws["B11"] = data.get("remaining_months", 10)
    print(f"  Setup: B4={ENTITY}, B5={budget.get('building_name','')}, B10={data.get('ytd_months',2)}, B11={data.get('remaining_months',10)}")


# ── Step 4: Fill Type 1 sheets (R&S, Gen & Admin) ──────────────────────────

def fill_type1(wb, data, gl_mapping):
    """
    Fill standard GL sheets that have the full 15-column layout:
    A=GL  B=Desc  C=Notes  D=Prior  E=YTD  F=Accrual  G=Unpaid
    H=YTD Budget  I=Estimate(fx)  J=Forecast(fx)  K=Curr Budget
    L=Inc%  M=Proposed(fx)  N=$Var(fx)  O=%Chg(fx)
    """
    ytd = data.get("ytd_months", 2)
    rem = data.get("remaining_months", 10)
    line_lookup = {l["gl_code"]: l for l in data.get("lines", [])}
    filled = 0

    for sn in TYPE1_SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            gl = str(ws.cell(r, 1).value or "").strip()
            if not GL_CODE_RE.match(gl):
                continue
            line = line_lookup.get(gl)
            if not line:
                continue

            # Raw data columns
            ws.cell(r, 3, value=line.get("notes", "") or "")           # C: Notes
            ws.cell(r, 4, value=line.get("prior_year", 0) or 0)       # D: Prior Year
            ws.cell(r, 5, value=line.get("ytd_actual", 0) or 0)       # E: YTD Actual
            ws.cell(r, 6, value=line.get("accrual_adj", 0) or 0)      # F: Accrual Adj
            ws.cell(r, 7, value=line.get("unpaid_bills", 0) or 0)     # G: Unpaid Bills
            ws.cell(r, 8, value=line.get("ytd_budget", 0) or 0)       # H: YTD Budget
            ws.cell(r, 11, value=line.get("current_budget", 0) or 0)  # K: Current Budget
            ws.cell(r, 12, value=line.get("increase_pct", 0) or 0)    # L: Increase %

            # I: Mar-Dec Estimate — override or formula
            est_override = line.get("estimate_override")
            if est_override is not None:
                ws.cell(r, 9, value=float(est_override))
            else:
                ws.cell(r, 9).value = (
                    f"=IFERROR(IF(AND((E{r}+F{r}+G{r})>=D{r},D{r}>0),"
                    f"(E{r}+F{r}+G{r})/{ytd}*{rem},"
                    f"MAX(D{r}-(E{r}+F{r}+G{r}),0)),0)"
                )

            # J: 12 Mo Forecast — override or formula
            fcst_override = line.get("forecast_override")
            if fcst_override is not None:
                ws.cell(r, 10, value=float(fcst_override))
            else:
                ws.cell(r, 10).value = f"=E{r}+F{r}+G{r}+I{r}"

            # M: Proposed Budget — formula
            ws.cell(r, 13).value = f"=ROUND(J{r}*(1+L{r}),0)"

            # N: $ Variance = Proposed - Prior Year
            ws.cell(r, 14).value = f"=M{r}-D{r}"

            # O: % Change = (Proposed / Prior) - 1
            ws.cell(r, 15).value = f"=IFERROR(M{r}/D{r}-1,0)"

            filled += 1

    print(f"  Type 1 sheets: filled {filled} GL rows")
    return filled


# ── Step 5: Fill Type 2 sheets (Income, Payroll, Energy, Water & Sewer) ────

def fill_type2(wb, data, gl_mapping):
    """
    Fill custom-layout sheets. Only inject raw data into known columns.
    Leave specialized formulas intact.
    """
    line_lookup = {l["gl_code"]: l for l in data.get("lines", [])}
    filled = 0

    for sn in TYPE2_SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in range(1, ws.max_row + 1):
            gl = str(ws.cell(r, 1).value or "").strip()
            if not GL_CODE_RE.match(gl):
                continue
            line = line_lookup.get(gl)
            if not line:
                continue

            # Always fill the 4 Yardi columns (D, E, H, K)
            ws.cell(r, 4, value=line.get("prior_year", 0) or 0)       # D
            ws.cell(r, 5, value=line.get("ytd_actual", 0) or 0)       # E
            ws.cell(r, 8, value=line.get("ytd_budget", 0) or 0)       # H
            ws.cell(r, 11, value=line.get("current_budget", 0) or 0)  # K

            # Try to fill accrual/unpaid/inc% if columns F, G, L have data or headers nearby
            # Check if column F has a value or formula already (indicates it's an input column)
            f_val = ws.cell(r, 6).value
            if f_val is not None or line.get("accrual_adj", 0):
                ws.cell(r, 6, value=line.get("accrual_adj", 0) or 0)
            g_val = ws.cell(r, 7).value
            if g_val is not None or line.get("unpaid_bills", 0):
                ws.cell(r, 7, value=line.get("unpaid_bills", 0) or 0)
            l_val = ws.cell(r, 12).value
            if l_val is not None or line.get("increase_pct", 0):
                ws.cell(r, 12, value=line.get("increase_pct", 0) or 0)

            # Notes in column C if it exists as input
            if line.get("notes"):
                ws.cell(r, 3, value=line.get("notes", ""))

            filled += 1

    print(f"  Type 2 sheets: filled {filled} GL rows")
    return filled


# ── Step 5b: Fill RE Taxes sheet ────────────────────────────────────────────

def fill_re_taxes(wb, data):
    """
    Fill the RE Taxes sheet with DOF tax calculation data.
    Template layout:
      Row 7: B=Assessed Valuation (1st half AV)
      Row 8: B=Tax Rate, D=1st Half Tax (formula)
      Row 11: B=Transitional AV Increase %
      Row 12: B=Estimated AV (formula), D=formula
      Row 13: B=Estimated Tax Rate, D=2nd Half Tax (formula)
      Row 15: B=Gross Tax (formula)
      Rows 19-22: Exemptions (Veteran, SCHE, STAR, Co-op)
      Row 23: Total Exemptions (formula)
      Row 25: Net Tax (formula)
    """
    if "RE Taxes" not in wb.sheetnames:
        print("  WARNING: No RE Taxes sheet found")
        return 0

    re_taxes = data.get("re_taxes")
    if not re_taxes:
        print("  No RE Taxes data from API (not a co-op or data unavailable)")
        return 0

    ws = wb["RE Taxes"]

    # 1st Half: Assessed Valuation & Tax Rate
    av1 = re_taxes.get("assessed_value", 0) or re_taxes.get("prior_trans_av", 0)
    rate1 = re_taxes.get("tax_rate", 0)
    ws.cell(7, 2, value=av1)                    # B7: Assessed Valuation
    ws.cell(8, 2, value=rate1)                   # B8: Tax Rate
    ws.cell(8, 4).value = "=B7*B8/2"            # D8: 1st Half Tax (formula)

    # 2nd Half: Transitional increase, Estimated AV, Estimated Rate
    trans_pct = re_taxes.get("transitional_av_increase", 0)
    est_rate = re_taxes.get("est_tax_rate", 0)
    ws.cell(11, 2, value=trans_pct)              # B11: Transitional AV Increase %
    ws.cell(12, 2).value = "=B7*(1+B11)"        # B12: Estimated AV (formula)
    ws.cell(13, 2, value=est_rate)               # B13: Estimated Tax Rate
    ws.cell(13, 4).value = "=B12*B13/2"         # D13: 2nd Half Tax (formula)

    # Gross Tax
    ws.cell(15, 2).value = "=D8+D13"            # B15: Gross Tax (formula)

    # Exemptions — each has: growth %, current year, budget year (formula)
    exemptions = re_taxes.get("exemptions", {})

    # Row 19: Veteran
    vet = exemptions.get("veteran", {})
    ws.cell(19, 2, value=vet.get("growth_pct", 0))      # B19: Growth %
    ws.cell(19, 3, value=vet.get("current_year", 0))     # C19: Current Year
    ws.cell(19, 4).value = "=C19*(1+B19)"                # D19: Budget Year (formula)

    # Row 20: SCHE (Senior Citizen)
    sche = exemptions.get("sche", {})
    ws.cell(20, 2, value=sche.get("growth_pct", 0))
    ws.cell(20, 3, value=sche.get("current_year", 0))
    ws.cell(20, 4).value = "=C20*(1+B20)"

    # Row 21: STAR
    star = exemptions.get("star", {})
    ws.cell(21, 2, value=star.get("growth_pct", 0))
    ws.cell(21, 3, value=star.get("current_year", 0))
    ws.cell(21, 4).value = "=C21*(1+B21)"

    # Row 22: Co-op Abatement
    coop = exemptions.get("coop_abatement", {})
    ws.cell(22, 2, value=coop.get("growth_pct", 0))
    ws.cell(22, 3, value=coop.get("current_year", 0))
    ws.cell(22, 4).value = "=C22*(1+B22)"

    # Row 23: Total Exemptions
    ws.cell(23, 3).value = "=SUM(C19:C22)"      # C23: Total Current
    ws.cell(23, 4).value = "=SUM(D19:D22)"      # D23: Total Budget (formula)

    # Row 25: Net Tax = Gross - Total Exemptions
    ws.cell(25, 2).value = "=B15-D23"            # B25: Net Tax (formula)

    # Metadata
    ws.cell(27, 1, value="Source")
    ws.cell(27, 2, value=re_taxes.get("source", ""))
    ws.cell(28, 1, value="BBL")
    ws.cell(28, 2, value=re_taxes.get("bbl", ""))
    ws.cell(29, 1, value="Tax Class")
    ws.cell(29, 2, value=re_taxes.get("tax_class", ""))

    print(f"  RE Taxes: AV={av1:,.0f}, Rate={rate1:.6f}, Gross={re_taxes.get('gross_tax',0):,.0f}, Net={re_taxes.get('net_tax',0):,.0f}")
    return 1


# ── Helper: fix Payroll template formulas that reference wrong column ──────

def _fix_payroll_formula_refs(ws):
    """
    The template's Payroll sheet has formulas that reference $L$7, $L$8, $L$10, $L$11
    but those cells contain text labels, not values. The actual values live in M7,
    M8, M10, M11. Rewrite those formula references from L to M.
    Also fix: H37's =L7*... and L55-L99 GL rows where =$L$8.
    """
    import re as _re
    # Map: (col from L -> col M) for specific rows only
    targets = {
        r"\$L\$7": "$M$7",
        r"\$L\$8": "$M$8",
        r"\$L\$10": "$M$10",
        r"\$L\$11": "$M$11",
        # Unqualified L7/L8/L10/L11 at formula start/boundary only
    }
    # H37 has =L7*... — that L7 is unqualified. Also need to handle that.
    # Use a single regex: match L with optional $ prefixes, and absolute/relative rows 7,8,10,11

    fixes = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            original = v
            # Replace $L$7 -> $M$7 etc.
            v = v.replace("$L$7", "$M$7")
            v = v.replace("$L$8", "$M$8")
            v = v.replace("$L$10", "$M$10")
            v = v.replace("$L$11", "$M$11")
            # Replace unqualified L7/L8/L10/L11 (but NOT L55, L56 etc. that are roster refs)
            # Use word boundary to only match exactly L7, L8, L10, L11
            v = _re.sub(r"(?<![A-Z$])L7(?!\d)", "M7", v)
            v = _re.sub(r"(?<![A-Z$])L8(?!\d)", "M8", v)
            v = _re.sub(r"(?<![A-Z$])L10(?!\d)", "M10", v)
            v = _re.sub(r"(?<![A-Z$])L11(?!\d)", "M11", v)
            if v != original:
                cell.value = v
                fixes += 1
    return fixes


# ── Step 5c: Fill Payroll sheet ─────────────────────────────────────────────

def fill_payroll(wb, data):
    """
    Fill the Payroll sheet with:
    - Section 1 (rows 7-12): Payroll assumptions (tax rates, union benefits)
    - Section 2 (rows 17-24): Employee roster (positions, counts, rates)
    - Section 3 (rows 55-105): GL detail lines (same as Type 1 fill)
    Template formulas in sections 2-3 handle wage calcs, taxes, benefits.
    """
    if "Payroll" not in wb.sheetnames:
        print("  WARNING: No Payroll sheet found")
        return 0

    ws = wb["Payroll"]

    # Fetch payroll-specific data from API
    try:
        pos_resp = requests.get(f"{API_BASE}/api/payroll/positions/{ENTITY}")
        positions = pos_resp.json() if pos_resp.ok else []
        assum_resp = requests.get(f"{API_BASE}/api/payroll/assumptions/{ENTITY}")
        assum_data = assum_resp.json() if assum_resp.ok else {}
        assumptions = assum_data.get("assumptions", assum_data)
    except Exception as e:
        print(f"  WARNING: Could not fetch payroll data: {e}")
        positions = []
        assumptions = {}

    filled = 0

    # ── Section 1: Payroll Assumptions (tax rates + union benefits) ──
    if assumptions:
        # Payroll tax rates (column B, rows 7-12)
        tax_map = {
            7: "fica",           # B7: FICA
            8: "sui",            # B8: SUI
            9: "fui",            # B9: FUI
            10: "mta",           # B10: MTA
            11: "nys_disability", # B11: NYS Disability
            12: "pfl",           # B12: Paid Family Leave
        }
        for row, key in tax_map.items():
            val = assumptions.get(key)
            if val is not None:
                ws.cell(row, 2, value=float(val))

        # Union benefit rates (column G, rows 7-12)
        benefit_map = {
            7: "welfare_monthly",
            8: "pension_weekly",
            9: "supp_retirement_weekly",
            10: "legal_monthly",
            11: "training_monthly",
            12: "profit_sharing_quarterly",
        }
        for row, key in benefit_map.items():
            val = assumptions.get(key)
            if val is not None:
                ws.cell(row, 7, value=float(val))

        # Workers comp, wage increase, and week counts — values go in column M (display)
        # The template's labels stay in column L.
        # NOTE: Template formulas originally referenced $L$7, $L$8, $L$10, $L$11
        # (which contain text labels, not numbers — a template bug). We rewrite those
        # formulas below to reference column M instead, where the values actually live.
        if assumptions.get("workers_comp") is not None:
            ws.cell(7, 13, value=float(assumptions["workers_comp"]))   # M7
        if assumptions.get("wage_increase_pct") is not None:
            ws.cell(8, 13, value=float(assumptions["wage_increase_pct"]))  # M8
        if assumptions.get("effective_week"):
            ws.cell(9, 13, value=assumptions["effective_week"])        # M9
        if assumptions.get("pre_increase_weeks") is not None:
            ws.cell(10, 13, value=float(assumptions["pre_increase_weeks"]))  # M10
        if assumptions.get("post_increase_weeks") is not None:
            ws.cell(11, 13, value=float(assumptions["post_increase_weeks"]))  # M11

        # Rewrite template formulas that wrongly reference column L (text labels)
        # to reference column M (actual values)
        _fix_payroll_formula_refs(ws)

        print(f"  Payroll assumptions: {len(assumptions)} fields loaded (values in M, formulas rewritten)")

    # ── Section 2: Employee Roster (rows 17-24) ──
    if isinstance(positions, list) and positions:
        for i, pos in enumerate(positions[:8]):  # Max 8 positions in template
            r = 17 + i
            ws.cell(r, 1, value=pos.get("position_name", ""))
            ws.cell(r, 2, value=pos.get("employee_count", 0))
            ws.cell(r, 3, value=pos.get("hourly_rate", 0))
            # Leave columns D-M as template formulas (weekly pay, pre/post wages, etc.)

        print(f"  Payroll roster: {len(positions)} positions loaded")

    # ── Section 3: GL Detail (rows 55-105) — same as Type 1 fill ──
    ytd = data.get("ytd_months", 2)
    rem = data.get("remaining_months", 10)
    line_lookup = {l["gl_code"]: l for l in data.get("lines", [])}

    for r in range(1, ws.max_row + 1):
        gl = str(ws.cell(r, 1).value or "").strip()
        if not GL_CODE_RE.match(gl):
            continue
        line = line_lookup.get(gl)
        if not line:
            continue

        # Raw data columns
        ws.cell(r, 3, value=line.get("notes", "") or "")           # C
        ws.cell(r, 4, value=line.get("prior_year", 0) or 0)       # D
        ws.cell(r, 5, value=line.get("ytd_actual", 0) or 0)       # E
        ws.cell(r, 6, value=line.get("accrual_adj", 0) or 0)      # F
        ws.cell(r, 7, value=line.get("unpaid_bills", 0) or 0)     # G
        ws.cell(r, 8, value=line.get("ytd_budget", 0) or 0)       # H
        ws.cell(r, 11, value=line.get("current_budget", 0) or 0)  # K
        ws.cell(r, 12, value=line.get("increase_pct", 0) or 0)    # L

        # I: Estimate
        est_override = line.get("estimate_override")
        if est_override is not None:
            ws.cell(r, 9, value=float(est_override))
        else:
            ws.cell(r, 9).value = (
                f"=IFERROR(IF(AND((E{r}+F{r}+G{r})>=D{r},D{r}>0),"
                f"(E{r}+F{r}+G{r})/{ytd}*{rem},"
                f"MAX(D{r}-(E{r}+F{r}+G{r}),0)),0)"
            )

        # J: Forecast
        fcst_override = line.get("forecast_override")
        if fcst_override is not None:
            ws.cell(r, 10, value=float(fcst_override))
        else:
            ws.cell(r, 10).value = f"=E{r}+F{r}+G{r}+I{r}"

        # M: Proposed
        ws.cell(r, 13).value = f"=ROUND(J{r}*(1+L{r}),0)"

        # N: $ Variance
        ws.cell(r, 14).value = f"=M{r}-D{r}"

        # O: % Change
        ws.cell(r, 15).value = f"=IFERROR(M{r}/D{r}-1,0)"

        filled += 1

    print(f"  Payroll GL detail: {filled} rows filled")
    return filled


# ── Step 6: Build "All Lines" summary sheet ─────────────────────────────────

def build_all_lines_sheet(wb, data):
    """
    New sheet with every line from the API in flat table format with formulas.
    Grouped by sheet_name. This ensures nothing is lost — even unmapped lines.
    """
    lines = data.get("lines", [])
    ytd = data.get("ytd_months", 2)
    rem = data.get("remaining_months", 10)

    ws = wb.create_sheet(title="All Lines")

    headers = [
        "Sheet", "GL Code", "Description", "Notes",
        "Prior Year", "YTD Actual", "Accrual Adj", "Unpaid Bills",
        "YTD Budget", "Mar-Dec Est", "12 Mo Forecast", "Curr Budget",
        "Inc %", "Proposed Budget", "$ Variance", "% Change"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Sort by sheet_name then row_num
    sorted_lines = sorted(lines, key=lambda x: (x.get("sheet_name", "ZZZ"), x.get("row_num", 0)))

    for i, line in enumerate(sorted_lines):
        r = i + 2

        ws.cell(r, 1, value=line.get("sheet_name", ""))
        ws.cell(r, 2, value=line.get("gl_code", ""))
        ws.cell(r, 3, value=line.get("description", ""))
        ws.cell(r, 4, value=line.get("notes", "") or "")
        ws.cell(r, 5, value=line.get("prior_year", 0) or 0)        # E
        ws.cell(r, 6, value=line.get("ytd_actual", 0) or 0)        # F
        ws.cell(r, 7, value=line.get("accrual_adj", 0) or 0)       # G
        ws.cell(r, 8, value=line.get("unpaid_bills", 0) or 0)      # H
        ws.cell(r, 9, value=line.get("ytd_budget", 0) or 0)        # I

        # J: Mar-Dec Estimate (formula)
        est = line.get("estimate_override")
        if est is not None:
            ws.cell(r, 10, value=float(est))
        else:
            ws.cell(r, 10).value = (
                f"=IFERROR(IF(AND((F{r}+G{r}+H{r})>=E{r},E{r}>0),"
                f"(F{r}+G{r}+H{r})/{ytd}*{rem},"
                f"MAX(E{r}-(F{r}+G{r}+H{r}),0)),0)"
            )

        # K: 12 Mo Forecast (formula)
        fcst = line.get("forecast_override")
        if fcst is not None:
            ws.cell(r, 11, value=float(fcst))
        else:
            ws.cell(r, 11).value = f"=F{r}+G{r}+H{r}+J{r}"

        ws.cell(r, 12, value=line.get("current_budget", 0) or 0)   # L
        ws.cell(r, 13, value=line.get("increase_pct", 0) or 0)     # M

        # N: Proposed Budget (formula)
        ws.cell(r, 14).value = f"=ROUND(K{r}*(1+M{r}),0)"

        # O: $ Variance = Curr Budget - Forecast
        ws.cell(r, 15).value = f"=L{r}-K{r}"

        # P: % Change
        ws.cell(r, 16).value = f"=IFERROR((L{r}-K{r})/K{r},0)"

    # Grand total row
    total_r = len(sorted_lines) + 2
    ws.cell(total_r, 1, value="GRAND TOTAL")
    for col in [5, 6, 7, 8, 9, 10, 11, 12, 14, 15]:
        cl = chr(64 + col)
        ws.cell(total_r, col).value = f"=SUM({cl}2:{cl}{total_r - 1})"
    # % Change for total
    ws.cell(total_r, 16).value = f"=IFERROR((L{total_r}-K{total_r})/K{total_r},0)"

    print(f"  All Lines sheet: {len(sorted_lines)} rows + grand total")
    return len(sorted_lines)


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    # Verify template exists
    if not TEMPLATE_PATH.exists():
        print(f"ERROR: Template not found at {TEMPLATE_PATH}")
        return

    # Step 1: Fetch data
    data = fetch_data()

    # Copy template
    print(f"\nCopying template: {TEMPLATE_PATH.name} -> {OUTPUT_PATH.name}")
    shutil.copy2(TEMPLATE_PATH, OUTPUT_PATH)

    # Open the copy
    wb = openpyxl.load_workbook(OUTPUT_PATH)
    print(f"  Template sheets: {wb.sheetnames}")

    # Step 2: Build GL mapping
    gl_mapping = build_gl_mapping(wb)
    print(f"  GL mapping: {len(gl_mapping)} codes found in template")

    # Step 3: Fill Setup
    print("\nFilling Setup sheet...")
    fill_setup(wb, data)

    # Step 4: Fill Type 1 sheets
    print("\nFilling Type 1 sheets (R&S, Gen & Admin)...")
    fill_type1(wb, data, gl_mapping)

    # Step 5: Fill Type 2 sheets
    print("\nFilling Type 2 sheets (Income, Energy, Water & Sewer)...")
    fill_type2(wb, data, gl_mapping)

    # Step 5b: Fill RE Taxes
    print("\nFilling RE Taxes sheet...")
    fill_re_taxes(wb, data)

    # Step 5c: Fill Payroll sheet
    print("\nFilling Payroll sheet...")
    fill_payroll(wb, data)

    # Step 6: Build All Lines sheet
    print("\nBuilding All Lines sheet...")
    build_all_lines_sheet(wb, data)

    # Save
    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")

    # Verification stats
    print("\nVerification:")
    for ws in wb.worksheets:
        total = 0
        formulas = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    total += 1
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas += 1
        if total > 0:
            print(f"  {ws.title}: {total} cells, {formulas} formulas")

    # Count how many API lines were mapped vs unmapped
    mapped_gls = set(gl_mapping.keys())
    api_gls = {l["gl_code"] for l in data.get("lines", [])}
    mapped = api_gls & mapped_gls
    unmapped = api_gls - mapped_gls
    print(f"\n  Mapped to template: {len(mapped)} GL codes")
    print(f"  Not in template: {len(unmapped)} GL codes (in All Lines sheet)")
    print(f"\nDone. Open {OUTPUT_PATH.name} in Excel to review.")


if __name__ == "__main__":
    main()
