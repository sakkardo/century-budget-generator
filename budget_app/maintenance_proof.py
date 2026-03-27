"""
Maintenance Proof / Common Charges Report Parser for Century Management.

Parses Yardi "Maint Proof" / "Common Charges" .xlsx exports and provides:
- Unit-level detail: unit code, shares/% common, charge code, monthly amount
- Summary totals: total shares, total monthly, annual projection
- Auto-population of Income budget lines from maintenance proof data
- Building type awareness: Coops run "Maintenance", Condos run "Common Charges"
"""

from flask import Blueprint, request, jsonify
from datetime import datetime
import logging
import re
import csv
import os

logger = logging.getLogger(__name__)


def create_maintenance_proof_blueprint(db, workflow_models):
    """Factory function to create the maintenance proof blueprint with DB access."""

    bp = Blueprint("maintenance_proof", __name__)
    Budget = workflow_models["Budget"]
    BudgetLine = workflow_models["BudgetLine"]
    BudgetRevision = workflow_models["BudgetRevision"]

    # ─── DB Models ────────────────────────────────────────────────────────────

    class MaintenanceProofReport(db.Model):
        """Stores metadata about an uploaded maintenance proof report."""
        __tablename__ = "maint_proof_reports"

        id = db.Column(db.Integer, primary_key=True)
        entity_code = db.Column(db.String(20), nullable=False, index=True)
        building_type = db.Column(db.String(30))  # Coop, Condo, Cond-Op
        charge_label = db.Column(db.String(50))    # "Maintenance" or "Common Charges"
        file_name = db.Column(db.String(255))
        total_units = db.Column(db.Integer, default=0)
        total_shares = db.Column(db.Float, default=0)
        total_monthly = db.Column(db.Float, default=0)
        total_annual = db.Column(db.Float, default=0)
        uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

        units = db.relationship("MaintenanceProofUnit", back_populates="report",
                                cascade="all, delete-orphan", lazy="dynamic")

        def to_dict(self):
            return {
                "id": self.id,
                "entity_code": self.entity_code,
                "building_type": self.building_type,
                "charge_label": self.charge_label,
                "file_name": self.file_name,
                "total_units": self.total_units,
                "total_shares": self.total_shares,
                "total_monthly": round(self.total_monthly, 2),
                "total_annual": round(self.total_annual, 2),
                "uploaded_at": self.uploaded_at.isoformat() if self.uploaded_at else None,
            }

    class MaintenanceProofUnit(db.Model):
        """Individual unit record from the maintenance proof."""
        __tablename__ = "maint_proof_units"

        id = db.Column(db.Integer, primary_key=True)
        report_id = db.Column(db.Integer, db.ForeignKey("maint_proof_reports.id"), nullable=False)
        unit_code = db.Column(db.String(30), nullable=False)
        shares = db.Column(db.Float, default=0)       # Shares (coop) or % common (condo)
        status = db.Column(db.String(30))              # Current, Vacant, etc.
        charge_code = db.Column(db.String(50))         # maint, common, storage, etc.
        monthly_amount = db.Column(db.Float, default=0)

        report = db.relationship("MaintenanceProofReport", back_populates="units")

        def to_dict(self):
            return {
                "id": self.id,
                "unit_code": self.unit_code.strip() if self.unit_code else "",
                "shares": self.shares,
                "status": self.status,
                "charge_code": self.charge_code.strip() if self.charge_code else "",
                "monthly_amount": round(self.monthly_amount, 2),
                "annual_amount": round(self.monthly_amount * 12, 2),
            }

    # ─── Parser ───────────────────────────────────────────────────────────────

    def parse_maintenance_proof(file_path):
        """
        Parse a Yardi Maintenance Proof / Common Charges .xlsx export.

        Expected format:
        - Row 2: Report title (e.g., "Maint Proof" or "Common Charges")
        - Row 4-5: Property/charge code filter info
        - Row 7: Headers (Property, Unit Code, %common/Share, Status, Charge Code, Amount, Unit)
        - Row 8+: Data rows (unit_code in col B, shares in col C, status in col D,
                   charge_code in col E, amount in col F)
        - Last row: Total shares in col C

        Returns:
            tuple: (report_title, units_list, total_shares_from_footer)
        """
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Row 2: Report title
        report_title = str(ws.cell(row=2, column=1).value or "").strip()

        # Parse data rows (start from row 8 based on format)
        units = []
        total_shares_footer = 0

        for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=8, values_only=False):
            unit_code = str(row[1].value or "").strip()  # Col B
            shares = row[2].value                         # Col C
            status = str(row[3].value or "").strip()      # Col D
            charge_code = str(row[4].value or "").strip() # Col E
            amount = row[5].value                         # Col F

            # Footer row: only shares column has a value (total shares)
            if shares and not unit_code and not charge_code:
                total_shares_footer = float(shares)
                continue

            # Skip empty rows
            if not unit_code or amount is None:
                continue

            units.append({
                "unit_code": unit_code,
                "shares": float(shares or 0),
                "status": status,
                "charge_code": charge_code,
                "monthly_amount": float(amount or 0),
            })

        wb.close()
        return report_title, units, total_shares_footer

    # ─── Building Type Lookup ─────────────────────────────────────────────────

    def get_building_type(entity_code):
        """Look up building type from the buildings CSV."""
        csv_paths = [
            os.path.join(os.path.dirname(__file__), "budget_system", "buildings.csv"),
            os.path.join(os.path.dirname(__file__), "..", "budget_system", "buildings.csv"),
        ]
        for csv_path in csv_paths:
            if os.path.exists(csv_path):
                with open(csv_path) as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if str(row.get("entity_code", "")).strip() == str(entity_code).strip():
                            return row.get("type", "").strip()
        return ""

    def get_charge_label(building_type):
        """Determine the charge label based on building type."""
        btype = (building_type or "").lower()
        if btype in ["condo"]:
            return "Common Charges"
        elif btype in ["coop", "cond-op"]:
            return "Maintenance"
        return "Maintenance"  # Default

    # ─── Store Report ─────────────────────────────────────────────────────────

    def store_maintenance_proof(entity_code, report_title, units, total_shares_footer, file_name=""):
        """Store parsed maintenance proof data. Replaces existing report for same entity."""
        building_type = get_building_type(entity_code)
        charge_label = get_charge_label(building_type)

        # Remove existing report for this entity
        existing = MaintenanceProofReport.query.filter_by(entity_code=entity_code).first()
        if existing:
            db.session.delete(existing)
            db.session.flush()

        total_shares = total_shares_footer or sum(u["shares"] for u in units)
        total_monthly = sum(u["monthly_amount"] for u in units)

        report = MaintenanceProofReport(
            entity_code=entity_code,
            building_type=building_type,
            charge_label=charge_label,
            file_name=file_name,
            total_units=len(units),
            total_shares=total_shares,
            total_monthly=total_monthly,
            total_annual=total_monthly * 12,
        )
        db.session.add(report)
        db.session.flush()

        for u in units:
            db.session.add(MaintenanceProofUnit(
                report_id=report.id,
                unit_code=u["unit_code"],
                shares=u["shares"],
                status=u.get("status", ""),
                charge_code=u.get("charge_code", ""),
                monthly_amount=u["monthly_amount"],
            ))

        db.session.commit()
        logger.info(f"Stored maintenance proof for entity {entity_code}: {len(units)} units, "
                     f"${total_monthly:,.2f}/mo, ${total_monthly*12:,.2f}/yr ({charge_label})")
        return report

    # ─── Auto-populate Income Lines ──────────────────────────────────────────

    def apply_income_from_proof(entity_code, report):
        """
        Auto-populate income budget lines from the maintenance proof data.

        Maps charge codes to income GL codes:
        - maint/maintenance → 4010-0000 (Maintenance Income)
        - common → 4010-0000 (Common Charges Income, same GL)
        - storage → 4030-0000 (Storage Income)
        - parking/garage → 4025-0000 (Garage Income)
        - assessment → 4200-0000 (Assessment Income)

        The annual total for each charge code is written to the matching
        BudgetLine's ytd_actual or current_budget as appropriate.
        """
        CHARGE_TO_GL = {
            "maint": "4010-0000",
            "maintenance": "4010-0000",
            "common": "4010-0000",
            "common charges": "4010-0000",
            "storage": "4030-0000",
            "stor": "4030-0000",
            "parking": "4025-0000",
            "garage": "4025-0000",
            "assessment": "4200-0000",
            "assess": "4200-0000",
        }

        budget = Budget.query.filter_by(entity_code=entity_code, year=2027).first()
        if not budget:
            logger.warning(f"No budget found for entity {entity_code}, cannot apply income from proof")
            return {"applied": 0, "details": {}}

        # Group proof units by charge code
        charge_totals = {}
        for unit in report.units.all():
            code = (unit.charge_code or "").strip().lower()
            if code not in charge_totals:
                charge_totals[code] = 0
            charge_totals[code] += unit.monthly_amount

        # Map to GL codes and apply
        lines = BudgetLine.query.filter_by(budget_id=budget.id).all()
        gl_to_line = {line.gl_code: line for line in lines}

        applied = 0
        details = {}
        for charge_code, monthly_total in charge_totals.items():
            gl = CHARGE_TO_GL.get(charge_code)
            if not gl:
                logger.info(f"No GL mapping for charge code '{charge_code}', skipping")
                continue

            annual = round(monthly_total * 12, 2)

            if gl in gl_to_line:
                line = gl_to_line[gl]
                old_val = float(line.current_budget or 0)
                line.current_budget = annual

                # Log the change
                db.session.add(BudgetRevision(
                    budget_id=budget.id, action="update",
                    field_name="current_budget",
                    old_value=str(old_val),
                    new_value=str(annual),
                    notes=f"GL {gl}: auto-populated from {report.charge_label} proof "
                          f"({charge_code}, {report.total_units} units, ${monthly_total:,.2f}/mo)",
                    source="maint_proof"
                ))
                applied += 1
                details[gl] = {"charge_code": charge_code, "monthly": monthly_total, "annual": annual}
                logger.info(f"Applied {charge_code} → {gl}: ${annual:,.2f}/yr")

        db.session.commit()
        return {"applied": applied, "details": details}

    # ─── API Routes ───────────────────────────────────────────────────────────

    @bp.route("/api/maint-proof/upload", methods=["POST"])
    def upload_maint_proof():
        """Upload and parse a Maintenance Proof / Common Charges .xlsx file."""
        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files["file"]
        if not file.filename.endswith(".xlsx"):
            return jsonify({"error": "File must be .xlsx"}), 400

        override_entity = request.form.get("entity_code", "").strip()

        import tempfile
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        file.save(tmp.name)
        tmp.close()

        try:
            report_title, units, total_shares = parse_maintenance_proof(tmp.name)

            entity_code = override_entity
            if not entity_code:
                return jsonify({"error": "entity_code is required"}), 400

            report = store_maintenance_proof(
                entity_code, report_title, units, total_shares, file.filename
            )

            # Auto-populate income lines
            income_result = {"applied": 0, "details": {}}
            try:
                income_result = apply_income_from_proof(entity_code, report)
            except Exception as e:
                logger.error(f"Failed to apply income from proof: {e}")

            return jsonify({
                "status": "ok",
                "report": report.to_dict(),
                "units_parsed": len(units),
                "income_applied": income_result["applied"],
                "income_details": income_result.get("details", {}),
            })
        except Exception as e:
            logger.error(f"Failed to parse maintenance proof: {e}")
            return jsonify({"error": str(e)}), 500
        finally:
            os.unlink(tmp.name)

    @bp.route("/api/maint-proof/<entity_code>", methods=["GET"])
    def get_maint_proof(entity_code):
        """Get the latest maintenance proof report for an entity."""
        report = MaintenanceProofReport.query.filter_by(entity_code=entity_code)\
            .order_by(MaintenanceProofReport.uploaded_at.desc()).first()
        if not report:
            return jsonify({"exists": False})

        # Get charge code breakdown
        charge_summary = {}
        for unit in report.units.all():
            code = (unit.charge_code or "").strip()
            if code not in charge_summary:
                charge_summary[code] = {"count": 0, "shares": 0, "monthly": 0}
            charge_summary[code]["count"] += 1
            charge_summary[code]["shares"] += unit.shares
            charge_summary[code]["monthly"] += unit.monthly_amount

        # Annualize
        for code in charge_summary:
            charge_summary[code]["annual"] = round(charge_summary[code]["monthly"] * 12, 2)
            charge_summary[code]["monthly"] = round(charge_summary[code]["monthly"], 2)

        return jsonify({
            "exists": True,
            "report": report.to_dict(),
            "charge_summary": charge_summary,
        })

    @bp.route("/api/maint-proof/<entity_code>/units", methods=["GET"])
    def get_maint_proof_units(entity_code):
        """Get unit-level detail for an entity's maintenance proof."""
        report = MaintenanceProofReport.query.filter_by(entity_code=entity_code)\
            .order_by(MaintenanceProofReport.uploaded_at.desc()).first()
        if not report:
            return jsonify({"units": []})

        charge_filter = request.args.get("charge_code", "").strip()
        query = MaintenanceProofUnit.query.filter_by(report_id=report.id)
        if charge_filter:
            query = query.filter(MaintenanceProofUnit.charge_code.ilike(f"%{charge_filter}%"))

        units = query.order_by(MaintenanceProofUnit.unit_code).all()
        return jsonify({
            "report": report.to_dict(),
            "units": [u.to_dict() for u in units],
        })

    @bp.route("/api/building-type/<entity_code>", methods=["GET"])
    def get_building_type_api(entity_code):
        """Get the building type for an entity."""
        btype = get_building_type(entity_code)
        return jsonify({
            "entity_code": entity_code,
            "building_type": btype,
            "charge_label": get_charge_label(btype),
            "needs_maint_proof": btype.lower() in ["coop", "condo", "cond-op"],
        })

    # ─── Return blueprint and models ──────────────────────────────────────────

    models = {
        "MaintenanceProofReport": MaintenanceProofReport,
        "MaintenanceProofUnit": MaintenanceProofUnit,
    }
    helpers = {
        "parse_maintenance_proof": parse_maintenance_proof,
        "store_maintenance_proof": store_maintenance_proof,
        "get_building_type": get_building_type,
        "get_charge_label": get_charge_label,
        "apply_income_from_proof": apply_income_from_proof,
    }

    return bp, models, helpers
