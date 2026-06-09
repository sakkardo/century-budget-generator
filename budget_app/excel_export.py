"""Dynamic .xlsx export of a building's Summary (Excel Phase 2/3, 2026-06-09).

PURE + additive: takes the exact payload /api/summary returns and produces an .xlsx
whose computed cells are LIVE Excel formulas referencing other cells — change an input
in Excel and Forecast, %Var, every subtotal, Net Operating, and Total Surplus
recalculate. Nothing here is imported by the running app until a route wires it, so it
cannot affect existing behavior.

Formula model (one source of truth = the summary payload; mirrors the backend grouping
in workflow.py: data rows bucket by _section_key(section); subtotals match by label):
  Excel layout: A=label, B..I = col1..col8 (1 header row, data from row 2).
  data row:  value cells for col1-4,6,7; Forecast (F) = '=D{r}+E{r}' (col5==col3+col4
             holds exactly, even for fixed-forecast where col4=col5-col3); %Var (I) =
             '=(H{r}-F{r})/ABS(F{r})'.
  subtotal:  Total Income/Expenses/Non-Op = SUM of that section's data-row cells (by
             section_key, so the trailing Interest-Income row still rolls into Non-Op
             Income); Net Operating = Total Income - Total Expenses; Total Surplus =
             Net Op + Total Non-Op Income - Total Non-Op Expenses.
"""
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    _HAVE_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAVE_OPENPYXL = False

# Summary col key -> 1-based Excel column index (A=1 label, B=2 .. I=9).
_COLIDX = {"col1": 2, "col2": 3, "col3": 4, "col4": 5, "col5": 6, "col6": 7, "col7": 8}
_COLLET = {"col1": "B", "col2": "C", "col3": "D", "col4": "E", "col5": "F", "col6": "G", "col7": "H"}
_MONEY = "#,##0;(#,##0)"
_PCT = "0.0%"


def _section_key(section):
    """Mirror workflow.py _section_key: map a row's section label to its bucket."""
    s = (section or "").strip().lower()
    if "non" in s and "income" in s:
        return "non_operating_income"
    if "non" in s and "expense" in s:
        return "non_operating_expense"
    if s == "income":
        return "income"
    if "expense" in s:
        return "expenses"
    return ""


def _num(v):
    try:
        return None if v is None else round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _sum_formula(col_letter, excel_rows):
    """=SUM(F2,F3,F38) over (possibly non-contiguous) data-row cells; '' if none."""
    if not excel_rows:
        return None
    refs = ",".join("%s%d" % (col_letter, r) for r in excel_rows)
    return "=SUM(%s)" % refs


def build_summary_workbook(summary_data):
    """Return .xlsx bytes for one building's summary, with live formulas."""
    if not _HAVE_OPENPYXL:
        raise RuntimeError("openpyxl not available")
    rows = summary_data.get("rows", []) or []
    by = int(summary_data.get("budget_year") or 0)
    entity = str(summary_data.get("entity_code") or "")

    # Excel row per summary row: header is row 1, data starts at row 2.
    xlrow = {i: i + 2 for i in range(len(rows))}

    # Authoritative grouping: bucket data-row Excel rows by section_key.
    buckets = {"income": [], "expenses": [], "non_operating_income": [], "non_operating_expense": []}
    for i, r in enumerate(rows):
        if r.get("row_type") == "data":
            sk = _section_key(r.get("section"))
            if sk in buckets:
                buckets[sk].append(xlrow[i])

    # Locate the subtotal Excel rows we cross-reference (Net Op / Total Surplus).
    sub_row = {}
    for i, r in enumerate(rows):
        if r.get("row_type") != "subtotal":
            continue
        lbl = (r.get("label") or "").lower()
        if "total income" in lbl:
            sub_row["income"] = xlrow[i]
        elif "total expense" in lbl and "non" not in lbl:
            sub_row["expenses"] = xlrow[i]
        elif "net operating" in lbl:
            sub_row["netop"] = xlrow[i]
        elif "total non" in lbl and "income" in lbl:
            sub_row["noi"] = xlrow[i]
        elif "total non" in lbl and "expense" in lbl:
            sub_row["noe"] = xlrow[i]
        elif "total surplus" in lbl or "total deficit" in lbl:
            sub_row["grand"] = xlrow[i]

    def subtotal_cell_formula(label_lower, col_letter):
        if "total income" in label_lower:
            return _sum_formula(col_letter, buckets["income"])
        if "total expense" in label_lower and "non" not in label_lower:
            return _sum_formula(col_letter, buckets["expenses"])
        if "total non" in label_lower and "income" in label_lower:
            return _sum_formula(col_letter, buckets["non_operating_income"])
        if "total non" in label_lower and "expense" in label_lower:
            return _sum_formula(col_letter, buckets["non_operating_expense"])
        if "net operating" in label_lower:
            inc, exp = sub_row.get("income"), sub_row.get("expenses")
            if inc and exp:
                return "=%s%d-%s%d" % (col_letter, inc, col_letter, exp)
            return None
        if "total surplus" in label_lower or "total deficit" in label_lower:
            netop, noi, noe = sub_row.get("netop"), sub_row.get("noi"), sub_row.get("noe")
            if netop:
                f = "=%s%d" % (col_letter, netop)
                if noi:
                    f += "+%s%d" % (col_letter, noi)
                if noe:
                    f += "-%s%d" % (col_letter, noe)
                return f
            return None
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = ("Summary %s" % entity)[:31]

    headers = ["Line Item", "%d Actual" % (by - 3), "%d Actual" % (by - 2),
               "%d YTD" % (by - 1), "%d Est." % (by - 1), "%d Forecast" % (by - 1),
               "%d Budget" % (by - 1), "%d Budget" % by, "% Var"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 34

    for i, r in enumerate(rows):
        er = xlrow[i]
        rt = r.get("row_type")
        ws.cell(row=er, column=1, value=r.get("label"))
        if rt == "section_header":
            ws.cell(row=er, column=1).font = Font(bold=True)
            continue
        label_lower = (r.get("label") or "").lower()

        for ckey, colidx in _COLIDX.items():
            cl = _COLLET[ckey]
            if rt == "data":
                if ckey == "col5":
                    c3, c4, c5 = _num(r.get("col3")), _num(r.get("col4")), _num(r.get("col5"))
                    if c3 is not None and c4 is not None and c5 is not None and abs((c3 + c4) - c5) < 0.5:
                        cell = ws.cell(row=er, column=colidx, value="=D%d+E%d" % (er, er))
                        cell.number_format = _MONEY
                    elif c5 is not None:
                        cell = ws.cell(row=er, column=colidx, value=c5)
                        cell.number_format = _MONEY
                else:
                    v = _num(r.get(ckey))
                    if v is not None:
                        cell = ws.cell(row=er, column=colidx, value=v)
                        cell.number_format = _MONEY
            elif rt == "subtotal":
                f = subtotal_cell_formula(label_lower, cl)
                if f:
                    cell = ws.cell(row=er, column=colidx, value=f)
                    cell.number_format = _MONEY

        if rt in ("data", "subtotal"):
            # %Var = (Proposed - Forecast) / |Forecast|; guard div/0.
            cell = ws.cell(row=er, column=9,
                           value='=IF(ABS(F%d)>0,(H%d-F%d)/ABS(F%d),"")' % (er, er, er, er))
            cell.number_format = _PCT

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
